#!/usr/bin/env python3
"""
Create Excel file with 对比上月表 sheet based on the screenshot structure.
This creates a properly formatted comparison sheet matching the provided image.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

def create_comparison_excel():
    """Create Excel file with comparison sheet matching the screenshot"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "对比上月表"
    
    # Title
    title = "加拿大-各门店2025年6月9日环比数据-星期一"
    ws.merge_cells('A1:J1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    
    # Headers
    headers = ["项目", "内容", "加拿大一店", "加拿大二店", "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大七店", "加拿大片区"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data from the screenshot
    data_rows = [
        # 桌数(考核) section - Light yellow background
        ("桌数\n(考核)", "今日总客数", 158.1, 81.7, 255.6, 197.3, 194.5, 117.3, 143.1, 1147.6, "FFFF99"),
        ("", "今日外卖客数", 0.9, 10, 2.4, 0, 1, 0, 0, 14.3, "FFFF99"),
        ("", "今日未计入考核客数", 9.9, 4.3, 9.4, 6.7, 10.5, 11.7, 5.9, 58.4, "FFFF99"),
        ("", "6月总客数", 1821.1, 956.7, 2148.2, 2195.5, 2176.1, 1531.4, 1738.6, 12567.6, "FFFF99"),
        ("", "上月同期总客数", 1949.9, 1061.3, 2158, 2424, 2410.7, 1758.8, 1911.5, 13674.2, "FFFF99"),
        ("", "对比上月同期总客数", "下降128.8卓", "下降104.6卓", "下降9.8卓", "下降228.5卓", "下降234.6卓", "下降227.4卓", "下降172.9卓", "下降1106.6卓", "FFFF00"),
        
        # 收入 section - Light blue background
        ("收入\n(不含税-万加元)", "今日营业收入", 2.39, 1.03, 2.75, 2.11, 2.64, 1.72, 1.85, 14.42, "E6F3FF"),
        ("", "本月截止目前营业收入", 28.59, 13.93, 27.70, 28.70, 30.18, 20.03, 25.17, 174.30, "E6F3FF"),
        ("", "上月截止目前营业收入", 29.48, 14.87, 27.33, 32.86, 35.01, 22.69, 28.20, 190.45, "E6F3FF"),
        ("", "环比营业收入变化", -0.89, -0.94, 0.37, -4.16, -4.83, -2.66, -3.04, -16.15, "E6F3FF"),
        ("", "本月营业收入目标", 101.49, 53.00, 92.85, 110.00, 114.40, 71.00, 98.00, 640.74, "E6F3FF"),
        ("", "本月截止目标完成率", "28.2%", "26.3%", "29.8%", "26.1%", "26.4%", "28.2%", "25.7%", "27.2%", "FFFF00"),
        ("", "标准时间进度", "30.0%", "30.0%", "30.0%", "30.0%", "30.0%", "30.0%", "30.0%", "30.0%", "E6F3FF"),
        ("", "优惠总金额", 2.22, 1.34, 1.88, 1.60, 1.91, 0.99, 1.96, 11.90, "E6F3FF"),
        ("", "优惠占比", "7.76%", "9.61%", "6.80%", "5.56%", "6.34%", "4.96%", "7.77%", "6.83%", "E6F3FF"),
        ("", "今日人均消费", 46.04, 45.96, 32.27, 35.57, 48.77, 44.01, 44.00, 40.90, "E6F3FF"),
        ("", "今日消费客数", 505, 224, 851, 594, 541, 390, 421, 3526, "E6F3FF"),
        
        # 单桌消费 section - Light yellow background
        ("单桌消费\n(不含税)", "今日单桌消费", 138.40, 119.72, 103.62, 103.56, 128.69, 133.04, 124.32, 119.57, "FFFF99"),
        ("", "截止今日单桌消费", 149.95, 136.15, 124.73, 126.70, 133.24, 121.49, 138.42, 132.57, "FFFF00"),
        ("", "上月单桌消费", 147.36, 140.39, 126.07, 134.18, 137.79, 126.16, 144.23, 136.39, "FFFF99"),
        ("", "环比上月变化", 2.58, -4.24, -1.34, -7.48, -4.55, -4.67, -5.81, -3.82, "FFFF99"),
        ("", "名次", "第1名", "第2名", "第3名", "第4名", "第5名", "第6名", "第7名", "当月累计平均翻台率", "FFFF00"),
        
        # 翻台率 section - Light blue background
        ("翻台率", "6月9日翻台率排名", 5.33, 3.54, 2.98, 2.82, 2.51, 2.27, 2.09, 3.72, "E6F3FF"),
        ("", "6月平均翻台率排名", 4.97, 4.4, 3.82, 3.48, 3.39, 3.04, 2.95, "", "E6F3FF"),
    ]
    
    # Add data to worksheet
    current_row = 3
    category_start_rows = {}
    
    for i, row_data in enumerate(data_rows):
        category, content = row_data[0], row_data[1]
        values = row_data[2:-1]  # All values except the color
        color = row_data[-1]
        
        # Track category start rows for merging
        if category and category not in category_start_rows:
            category_start_rows[category] = current_row
        
        # Add category (column A)
        if category:
            ws.cell(row=current_row, column=1, value=category)
        
        # Add content (column B)
        ws.cell(row=current_row, column=2, value=content)
        
        # Add values (columns C-J)
        for col, value in enumerate(values, 3):
            cell = ws.cell(row=current_row, column=col, value=value)
            
            # Apply background color
            if color:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        # Apply background color to category and content cells
        if color:
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=current_row, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
        current_row += 1
    
    # Merge category cells
    merge_ranges = [
        ("桌数\n(考核)", 3, 8),  # 6 rows
        ("收入\n(不含税-万加元)", 9, 18),  # 10 rows
        ("单桌消费\n(不含税)", 19, 22),  # 4 rows
        ("翻台率", 23, 24)  # 2 rows
    ]
    
    for category, start_row, end_row in merge_ranges:
        if start_row < end_row:
            ws.merge_cells(f'A{start_row}:A{end_row}')
            cell = ws[f'A{start_row}']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
    
    # Special formatting for the final average (3.72)
    ws.cell(row=23, column=10).font = Font(bold=True, color="FF0000", size=14)
    
    # Apply borders to all cells with data
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(1, current_row):
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = thin_border
    
    # Set column widths
    column_widths = [12, 20, 12, 12, 12, 12, 12, 12, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Set row height for title
    ws.row_dimensions[1].height = 25
    
    return wb

def main():
    """Main function"""
    print("🍲 Creating Haidilao Comparison Excel File")
    print("=" * 50)
    
    # Create the Excel file
    wb = create_comparison_excel()
    
    # Save the file
    output_file = "data/haidilao_comparison_2025_6_9.xlsx"
    Path("data").mkdir(exist_ok=True)
    
    wb.save(output_file)
    
    print(f"✅ Excel file created successfully: {output_file}")
    print("\n📋 File contains:")
    print("  • Sheet name: 对比上月表")
    print("  • Title: 加拿大-各门店2025年6月9日环比数据-星期一")
    print("  • All 7 stores + 片区 data")
    print("  • Color-coded sections:")
    print("    - 桌数(考核): Light yellow")
    print("    - 收入(不含税-万加元): Light blue")
    print("    - 单桌消费(不含税): Light yellow")
    print("    - 翻台率: Light blue")
    print("  • Highlighted comparison rows in bright yellow")
    print("  • Proper Chinese formatting and merged cells")
    
    print(f"\n🎯 You can now use this file with the automation system:")
    print(f"   npm run open-automation-menu")
    print(f"   Then select option 5 and use: {output_file}")

if __name__ == "__main__":
    main() 