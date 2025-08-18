#!/usr/bin/env python3
"""
Test script to identify RBC duplicate transaction issues
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

def analyze_rbc_transactions():
    """Analyze RBC transaction files for potential duplicate issues"""
    
    # Check RBC files
    input_dir = Path("Input/daily_report/bank_transactions_reports")
    rbc_files = list(input_dir.glob("RBC Business Bank Account (5401)*.xlsx"))
    
    if rbc_files:
        file_path = rbc_files[0]
        print(f"\nAnalyzing RBC file: {file_path.name}")
        
        # Read the file
        df = pd.read_excel(file_path, engine='openpyxl')
        print(f"Total rows in source file: {len(df)}")
        print(f"Columns: {list(df.columns)}")
        
        # Check for duplicates in source data
        print("\nChecking for duplicates in source data...")
        
        # Define key columns for duplicate detection
        key_columns = ['Date', 'Description 1', 'Withdrawals', 'Deposits']
        
        # Check if all key columns exist
        missing_cols = [col for col in key_columns if col not in df.columns]
        if missing_cols:
            print(f"Missing columns: {missing_cols}")
            print(f"Available columns: {list(df.columns)}")
            return
        
        # Find duplicates
        duplicates = df[df.duplicated(subset=key_columns, keep=False)]
        if not duplicates.empty:
            print(f"\nFound {len(duplicates)} duplicate rows in source file!")
            print("\nFirst 10 duplicate entries:")
            print(duplicates.head(10)[key_columns])
        else:
            print("No duplicates found in source data")
        
        # Check date range
        print(f"\nDate range in file: {df['Date'].min()} to {df['Date'].max()}")
        
        # Check for July 2025 transactions
        july_2025 = df[pd.to_datetime(df['Date']).dt.to_period('M') == '2025-07']
        print(f"\nJuly 2025 transactions: {len(july_2025)}")
        
        # Check template file
        template_file = input_dir / "CA全部7家店明细.xlsx"
        if template_file.exists():
            print(f"\nAnalyzing template file...")
            wb = load_workbook(template_file)
            
            # Check RBC 5401 sheet
            sheet_name = "RBC 5401"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Find last row with data
                last_row = 2
                for row in range(3, ws.max_row + 1):
                    if any(ws.cell(row=row, column=col).value for col in range(1, 12)):
                        last_row = row
                
                print(f"Last row with data in {sheet_name}: {last_row}")
                
                # Check last few transactions
                print(f"\nLast 5 transactions in template:")
                for row in range(max(3, last_row - 4), last_row + 1):
                    date_val = ws.cell(row=row, column=2).value  # RBC date is in column 2
                    desc_val = ws.cell(row=row, column=1).value  # Description in column 1
                    debit_val = ws.cell(row=row, column=4).value  # Debit in column 4
                    credit_val = ws.cell(row=row, column=5).value  # Credit in column 5
                    
                    if date_val:
                        print(f"Row {row}: Date={date_val}, Desc={desc_val}, Debit={debit_val}, Credit={credit_val}")
            
            wb.close()

def check_duplicate_detection_logic():
    """Test the duplicate detection logic"""
    print("\n\nTesting duplicate detection logic...")
    
    # Test date normalization
    test_dates = [
        "Jul 31, 2025",
        "2025-07-31",
        datetime(2025, 7, 31),
        "31/07/2025"
    ]
    
    print("\nDate normalization test:")
    for date_val in test_dates:
        try:
            normalized = pd.to_datetime(date_val).strftime('%Y-%m-%d')
            print(f"{date_val} -> {normalized}")
        except Exception as e:
            print(f"{date_val} -> ERROR: {e}")
    
    # Test description comparison
    test_descriptions = [
        ("INTERAC PURCHASE", "INTERAC PURCHASE"),
        ("INTERAC PURCHASE ", "INTERAC PURCHASE"),  # Trailing space
        ("", None),  # Empty vs None
        ("Payment", "payment"),  # Case difference
    ]
    
    print("\nDescription comparison test:")
    for desc1, desc2 in test_descriptions:
        str1 = str(desc1) if desc1 else ""
        str2 = str(desc2) if desc2 else ""
        match = str1 == str2
        print(f"'{desc1}' vs '{desc2}' -> {match}")

if __name__ == "__main__":
    analyze_rbc_transactions()
    check_duplicate_detection_logic()