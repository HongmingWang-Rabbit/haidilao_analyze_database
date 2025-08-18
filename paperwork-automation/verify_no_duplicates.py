#!/usr/bin/env python3
"""
Verify no duplicates in the latest output
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import sys
import io
from collections import defaultdict

# Set UTF-8 encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def verify_no_duplicates():
    """Quick check for duplicates in latest output"""
    
    # Find the latest output file
    output_dir = Path("output")
    output_files = list(output_dir.glob("Bank_Transactions_Report_*.xlsx"))
    
    if not output_files:
        print("No output files found")
        return
    
    latest_file = max(output_files, key=lambda p: p.stat().st_mtime)
    print(f"Checking file: {latest_file.name}\n")
    
    # Load the workbook
    wb = load_workbook(latest_file)
    
    total_duplicates = 0
    
    # Check RBC sheets specifically
    rbc_sheets = ['RBC 5401', 'RBC 5419', 'RBC0517-Hi Bowl', 'RBC 0922（USD）', 'RBC3088（USD）-Hi Bowl']
    
    for sheet_name in rbc_sheets:
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        
        # Count transactions
        transactions = []
        for row in range(3, ws.max_row + 1):
            date_val = ws.cell(row=row, column=2).value  # Date
            if not date_val:
                break
                
            desc_val = ws.cell(row=row, column=1).value  # Description
            debit_val = ws.cell(row=row, column=4).value  # Debit
            credit_val = ws.cell(row=row, column=5).value  # Credit
            
            key = (
                str(date_val).strip(),
                str(desc_val).strip() if desc_val else '',
                str(debit_val).strip() if debit_val else '',
                str(credit_val).strip() if credit_val else ''
            )
            transactions.append(key)
        
        # Check for duplicates
        unique_count = len(set(transactions))
        duplicate_count = len(transactions) - unique_count
        
        print(f"{sheet_name}: {len(transactions)} transactions, {duplicate_count} duplicates")
        total_duplicates += duplicate_count
    
    wb.close()
    
    print(f"\nTOTAL DUPLICATES: {total_duplicates}")
    
    if total_duplicates > 0:
        print("\nWARNING: Duplicates still exist!")
    else:
        print("\nSUCCESS: No duplicates found!")

if __name__ == "__main__":
    verify_no_duplicates()