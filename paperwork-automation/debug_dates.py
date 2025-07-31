#!/usr/bin/env python3
"""Debug date detection in templates"""

from openpyxl import load_workbook
import pandas as pd

def debug_get_last_existing_date(ws, sheet_name):
    """Debug version of get_last_existing_date"""
    print(f"\n=== Debugging {sheet_name} ===")
    print(f"Max row: {ws.max_row}")
    
    last_date = None
    start_row = max(3, ws.max_row - 10)  # Check last 10 rows for debugging
    
    print(f"Checking rows {start_row} to {ws.max_row}")
    
    for row in range(start_row, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        print(f"Row {row}, Col 1: {repr(date_val)} (type: {type(date_val)})")
        
        if date_val and str(date_val) not in ['Date', '']:
            try:
                parsed_date = pd.to_datetime(date_val)
                print(f"  -> Parsed as: {parsed_date}")
                if last_date is None or parsed_date > last_date:
                    last_date = parsed_date
                    print(f"  -> New last date: {last_date}")
            except Exception as e:
                print(f"  -> Parse failed: {e}")
                continue
    
    result = last_date.strftime('%Y-%m-%d') if last_date else ""
    print(f"Final result: {repr(result)}")
    return result

# Test with template
wb = load_workbook('Input/daily_report/bank_transactions_reports/CA全部7家店明细.xlsx')

# Test CIBC
ws_cibc = wb['CA7D-CIBC 0401']
cibc_result = debug_get_last_existing_date(ws_cibc, 'CA7D-CIBC 0401')

# Test RBC  
ws_rbc = wb['RBC 5401']
rbc_result = debug_get_last_existing_date(ws_rbc, 'RBC 5401')

print(f"\nSUMMARY:")
print(f"CIBC last date: {cibc_result}")
print(f"RBC last date: {rbc_result}")