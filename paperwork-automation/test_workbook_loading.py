#!/usr/bin/env python3
"""Test different workbook loading methods to preserve images"""

from openpyxl import load_workbook
import sys
import os

# Fix encoding issue
sys.stdout.reconfigure(encoding='utf-8')

template_path = 'Input/daily_report/bank_transactions_reports/CA全部7家店明细.xlsx'

def test_loading_method(method_name, **kwargs):
    """Test a specific loading method"""
    print(f"\n=== Testing {method_name} ===")
    print(f"Parameters: {kwargs}")
    
    try:
        wb = load_workbook(template_path, **kwargs)
        
        # Check a few sheets for images
        sheets_to_check = ['CA1D-3817', 'CA7D-CIBC 0401', 'RBC 5401']
        
        for sheet_name in sheets_to_check:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Check for images
                image_count = 0
                if hasattr(ws, '_images') and ws._images:
                    image_count = len(ws._images)
                
                # Check for drawing objects
                drawing_exists = hasattr(ws, '_drawing') and ws._drawing is not None
                
                print(f"  {sheet_name}: {image_count} images, drawing={drawing_exists}")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"  ERROR: {e}")
        return False

# Test different loading approaches
methods = [
    ("Default", {}),
    ("Keep VBA", {"keep_vba": True}),
    ("Data Only False", {"data_only": False}),
    ("Keep Links", {"keep_links": True}),
    ("Rich Text", {"rich_text": True}),
    ("All Options", {"keep_vba": True, "data_only": False, "keep_links": True, "rich_text": True})
]

for method_name, kwargs in methods:
    test_loading_method(method_name, **kwargs)