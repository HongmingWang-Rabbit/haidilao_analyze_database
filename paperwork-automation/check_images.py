#!/usr/bin/env python3
"""Check for images and other objects in the template file"""

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import sys

# Fix encoding issue
sys.stdout.reconfigure(encoding='utf-8')

def check_worksheet_objects(ws, sheet_name):
    """Check what objects exist in a worksheet"""
    print(f"\n=== Checking {sheet_name} ===")
    
    # Check for images
    if hasattr(ws, '_images') and ws._images:
        print(f"Images found: {len(ws._images)}")
        for i, img in enumerate(ws._images):
            print(f"  Image {i+1}: {img}")
            if hasattr(img, 'anchor'):
                print(f"    Anchor: {img.anchor}")
    else:
        print("No images found")
    
    # Check for charts
    if hasattr(ws, '_charts') and ws._charts:
        print(f"Charts found: {len(ws._charts)}")
        for i, chart in enumerate(ws._charts):
            print(f"  Chart {i+1}: {chart}")
    else:
        print("No charts found")
    
    # Check for drawings
    if hasattr(ws, '_drawing') and ws._drawing:
        print(f"Drawing objects found")
        print(f"  Drawing: {ws._drawing}")
    else:
        print("No drawing objects found")
    
    # Check for shapes
    if hasattr(ws, '_shapes') and ws._shapes:
        print(f"Shapes found: {len(ws._shapes)}")
        for i, shape in enumerate(ws._shapes):
            print(f"  Shape {i+1}: {shape}")
    else:
        print("No shapes found")

# Load the template file
try:
    wb = load_workbook('Input/daily_report/bank_transactions_reports/CA全部7家店明细.xlsx')
    print(f"Template file loaded successfully")
    print(f"Sheets: {wb.sheetnames}")
    
    # Check key sheets for objects
    sheets_to_check = ['CA7D-CIBC 0401', 'RBC 5401', 'CA1D-3817']
    
    for sheet_name in sheets_to_check:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            check_worksheet_objects(ws, sheet_name)
    
    wb.close()
    
except Exception as e:
    print(f"Error loading template file: {e}")
    sys.exit(1)