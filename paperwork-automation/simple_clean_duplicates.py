#!/usr/bin/env python3
"""
Simple script to clean duplicate transactions
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import sys
import io
from datetime import datetime

# Set UTF-8 encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def simple_clean():
    """Simple duplicate removal"""
    
    # Find the latest output file
    output_dir = Path("output")
    output_files = list(output_dir.glob("Bank_Transactions_Report_*.xlsx"))
    
    if not output_files:
        print("No output files found")
        return
    
    latest_file = max(output_files, key=lambda p: p.stat().st_mtime)
    print(f"Processing file: {latest_file.name}")
    
    # Create a cleaned output file
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    cleaned_file = output_dir / f"Bank_Transactions_Report_CLEANED_{timestamp}.xlsx"
    
    # Copy the file first
    import shutil
    shutil.copy2(latest_file, cleaned_file)
    
    # Load the cleaned copy
    wb = load_workbook(cleaned_file)
    
    sheets_processed = 0
    total_original = 0
    total_after_cleaning = 0
    
    for sheet_name in wb.sheetnames:
        # Skip non-transaction sheets
        if 'Summary' in sheet_name or 'Cover' in sheet_name:
            continue
            
        ws = wb[sheet_name]
        
        # Determine column positions
        if 'RBC' in sheet_name:
            date_col = 2
            desc_col = 1
            debit_col = 4
            credit_col = 5
        elif 'CIBC' in sheet_name:
            date_col = 1
            desc_col = 2
            debit_col = 3
            credit_col = 4
        else:  # BMO format
            date_col = 1
            desc_col = 3
            debit_col = 6
            credit_col = 7
        
        # Read all data into memory
        all_rows = []
        for row in range(3, ws.max_row + 1):
            date_val = ws.cell(row=row, column=date_col).value
            if not date_val or str(date_val).lower() in ['date', '']:
                break
                
            # Read entire row
            row_data = []
            for col in range(1, ws.max_column + 1):
                row_data.append(ws.cell(row=row, column=col).value)
            
            # Add key fields for duplicate detection
            all_rows.append({
                'row_data': row_data,
                'key': (
                    str(date_val).strip(),
                    str(ws.cell(row=row, column=desc_col).value or '').strip(),
                    str(ws.cell(row=row, column=debit_col).value or '').strip(),
                    str(ws.cell(row=row, column=credit_col).value or '').strip()
                ),
                'date': date_val
            })
        
        original_count = len(all_rows)
        total_original += original_count
        
        # Remove duplicates
        seen_keys = set()
        unique_rows = []
        for row_info in all_rows:
            if row_info['key'] not in seen_keys:
                seen_keys.add(row_info['key'])
                unique_rows.append(row_info)
        
        # Sort by date
        try:
            unique_rows.sort(key=lambda x: pd.to_datetime(x['date']))
        except:
            unique_rows.sort(key=lambda x: str(x['date']))
        
        final_count = len(unique_rows)
        total_after_cleaning += final_count
        
        print(f"\n{sheet_name}:")
        print(f"  Original: {original_count} transactions")
        print(f"  After cleaning: {final_count} transactions")
        print(f"  Removed: {original_count - final_count} duplicates")
        
        # Clear and rewrite the sheet
        # First, clear all transaction rows
        for row in range(3, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
        
        # Write back unique rows
        for idx, row_info in enumerate(unique_rows):
            new_row = idx + 3
            for col_idx, value in enumerate(row_info['row_data'], 1):
                ws.cell(row=new_row, column=col_idx).value = value
        
        sheets_processed += 1
    
    # Save the cleaned file
    wb.save(cleaned_file)
    wb.close()
    
    print(f"\n{'='*60}")
    print(f"SUMMARY:")
    print(f"  Sheets processed: {sheets_processed}")
    print(f"  Total original transactions: {total_original}")
    print(f"  Total after cleaning: {total_after_cleaning}")
    print(f"  Total duplicates removed: {total_original - total_after_cleaning}")
    print(f"\nCleaned file saved as: {cleaned_file.name}")

if __name__ == "__main__":
    simple_clean()