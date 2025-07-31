#!/usr/bin/env python3
"""Check for visible duplicates in the bank transaction sheets"""

from openpyxl import load_workbook

wb = load_workbook('output/Bank_Transactions_Report_2025-07-29_155059.xlsx')
ws = wb['CA7D-CIBC 0401']

transactions = {}
duplicates = []

for r in range(3, ws.max_row+1):
    date = ws.cell(row=r, column=1).value
    desc = str(ws.cell(row=r, column=2).value) if ws.cell(row=r, column=2).value else ''
    debit = ws.cell(row=r, column=3).value  
    credit = ws.cell(row=r, column=4).value
    
    if date or desc:
        key = (str(date), desc[:50], str(debit), str(credit))
        if key in transactions:
            duplicates.append((r, transactions[key], key))
        else:
            transactions[key] = r

print(f'Found {len(duplicates)} visible duplicates in CA7D-CIBC 0401:')
for dup in duplicates[:10]:
    print(f'Row {dup[0]} duplicates Row {dup[1]}: {dup[2][0]} | {dup[2][1][:30]}... | {dup[2][2]} | {dup[2][3]}')

# Check RBC sheet too
ws_rbc = wb['RBC 5401']
transactions_rbc = {}
duplicates_rbc = []

for r in range(3, ws_rbc.max_row+1):
    date = ws_rbc.cell(row=r, column=2).value  # RBC date is in column 2
    desc = str(ws_rbc.cell(row=r, column=1).value) if ws_rbc.cell(row=r, column=1).value else ''  # RBC desc is in column 1
    debit = ws_rbc.cell(row=r, column=4).value  # RBC debit is in column 4
    credit = ws_rbc.cell(row=r, column=5).value  # RBC credit is in column 5
    
    if date or desc:
        key = (str(date), desc[:50], str(debit), str(credit))
        if key in transactions_rbc:
            duplicates_rbc.append((r, transactions_rbc[key], key))
        else:
            transactions_rbc[key] = r

print(f'\nFound {len(duplicates_rbc)} visible duplicates in RBC 5401:')
for dup in duplicates_rbc[:10]:
    print(f'Row {dup[0]} duplicates Row {dup[1]}: {dup[2][0]} | {dup[2][1][:30]}... | {dup[2][2]} | {dup[2][3]}')