#!/usr/bin/env python3
"""
Process Hi-Bowl Report
Copy template and fill in blue-marked sections with data from daily report
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
from pathlib import Path

# Add encoding support
import os
os.environ['PYTHONIOENCODING'] = 'utf-8'

def identify_blue_cells(template_path):
    """Identify cells with blue background in the template"""
    wb = load_workbook(template_path, data_only=False)
    blue_cells = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        blue_cells[sheet_name] = []
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill and cell.fill.patternType:
                    # Check if cell has blue fill
                    if cell.fill.fgColor and cell.fill.fgColor.rgb:
                        # Common blue colors in Excel
                        blue_colors = ['FF0070C0', 'FF00B0F0', 'FF5B9BD5', 'FF4472C4', 'FFDCE6F1']
                        if any(cell.fill.fgColor.rgb.startswith(color) for color in blue_colors):
                            blue_cells[sheet_name].append({
                                'cell': cell.coordinate,
                                'row': cell.row,
                                'column': cell.column,
                                'value': cell.value,
                                'color': cell.fill.fgColor.rgb
                            })
    
    return blue_cells

def analyze_files():
    """Analyze both files to understand structure"""
    
    # File paths
    daily_report = "Input/daily_report/hi-bowl-report/daily-data/HaiDiLao-report-8003-2025-7 (1).xlsx"
    template_file = "Input/daily_report/hi-bowl-report/output-template/海外新业态管报数据-本位币-新模板.xlsx"
    
    print("=== ANALYZING HI-BOWL REPORT FILES ===\n")
    
    # 1. Read daily report
    print(f"1. Reading daily report: {daily_report}")
    try:
        # Read all sheets
        xl_file = pd.ExcelFile(daily_report)
        print(f"   Sheets found: {xl_file.sheet_names}")
        
        # Read first few rows of each sheet
        for sheet_name in xl_file.sheet_names:
            print(f"\n   Sheet: {sheet_name}")
            df = pd.read_excel(daily_report, sheet_name=sheet_name, nrows=10)
            print(f"   Shape: {df.shape}")
            print(f"   Columns: {list(df.columns)[:5]}..." if len(df.columns) > 5 else f"   Columns: {list(df.columns)}")
            
    except Exception as e:
        print(f"   Error reading daily report: {e}")
    
    # 2. Read template file
    print(f"\n2. Reading template file: {template_file}")
    try:
        # Read all sheets
        xl_file = pd.ExcelFile(template_file)
        print(f"   Sheets found: {xl_file.sheet_names}")
        
        # Identify blue cells
        print("\n3. Identifying blue cells in template...")
        blue_cells = identify_blue_cells(template_file)
        
        for sheet_name, cells in blue_cells.items():
            if cells:
                print(f"\n   Sheet '{sheet_name}' has {len(cells)} blue cells:")
                for cell_info in cells[:5]:  # Show first 5
                    print(f"     Cell {cell_info['cell']}: {cell_info['value']}")
                if len(cells) > 5:
                    print(f"     ... and {len(cells) - 5} more")
                    
    except Exception as e:
        print(f"   Error reading template: {e}")

def fill_template_with_data():
    """Fill the template with data from daily report"""
    
    # File paths
    daily_report = "Input/daily_report/hi-bowl-report/daily-data/HaiDiLao-report-8003-2025-7 (1).xlsx"
    template_file = "Input/daily_report/hi-bowl-report/output-template/海外新业态管报数据-本位币-新模板.xlsx"
    output_file = "Input/daily_report/hi-bowl-report/output-template/海外新业态管报数据-本位币-filled.xlsx"
    
    print("\n\n=== FILLING TEMPLATE WITH DATA ===\n")
    
    # Load workbook for editing
    wb = load_workbook(template_file)
    
    # Read daily report data
    daily_data = pd.read_excel(daily_report, sheet_name=None)  # Read all sheets
    
    # Identify blue cells
    blue_cells = identify_blue_cells(template_file)
    
    # Process each sheet
    for sheet_name in wb.sheetnames:
        if sheet_name not in blue_cells or not blue_cells[sheet_name]:
            continue
            
        ws = wb[sheet_name]
        print(f"\nProcessing sheet: {sheet_name}")
        
        # For each blue cell, try to find corresponding data
        for cell_info in blue_cells[sheet_name]:
            cell_coord = cell_info['cell']
            cell_value = cell_info['value']
            
            print(f"  Blue cell {cell_coord}: {cell_value}")
            
            # TODO: Add logic to map template fields to daily report data
            # This will depend on the specific structure of both files
            
    # Save the filled template
    wb.save(output_file)
    print(f"\n\nTemplate filled and saved to: {output_file}")

if __name__ == "__main__":
    # First analyze the files
    analyze_files()
    
    # Then fill the template
    # fill_template_with_data()  # Uncomment after understanding the structure