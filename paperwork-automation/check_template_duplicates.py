#!/usr/bin/env python3
"""
Check if duplicates exist in the template file
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import sys
import io

# Set UTF-8 encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def check_template():
    """Check template file for existing duplicates"""
    
    template_file = Path("Input/daily_report/bank_transactions_reports/CA全部7家店明细.xlsx")
    print(f"Checking template file: {template_file.name}")
    
    # Load the workbook
    wb = load_workbook(template_file)
    
    # Check RBC 5401 sheet specifically
    sheet_name = "RBC 5401"
    if sheet_name in wb.sheetnames:
        print(f"\nChecking {sheet_name}...")
        ws = wb[sheet_name]
        
        # Extract all transactions
        transactions = []
        for row in range(3, ws.max_row + 1):
            # Check if row has data
            date_val = ws.cell(row=row, column=2).value  # Date column
            if not date_val:
                continue
                
            desc_val = ws.cell(row=row, column=1).value  # Description
            debit_val = ws.cell(row=row, column=4).value  # Debit
            credit_val = ws.cell(row=row, column=5).value  # Credit
            
            transactions.append({
                'row': row,
                'date': str(date_val),
                'desc': str(desc_val) if desc_val else '',
                'debit': str(debit_val) if debit_val else '',
                'credit': str(credit_val) if credit_val else ''
            })
        
        print(f"  Total transactions in template: {len(transactions)}")
        
        # Show all transactions
        print("\n  All transactions:")
        for trans in transactions:
            print(f"    Row {trans['row']}: {trans['date']} | {trans['desc'][:30]}... | Debit: {trans['debit']} | Credit: {trans['credit']}")
        
        # Check for duplicates
        seen = {}
        duplicates = []
        
        for trans in transactions:
            key = (trans['date'], trans['desc'], trans['debit'], trans['credit'])
            if key in seen:
                duplicates.append((trans['row'], seen[key]))
                print(f"\n  DUPLICATE FOUND: Row {trans['row']} is duplicate of row {seen[key]}")
            else:
                seen[key] = trans['row']
        
        if duplicates:
            print(f"\n  Total duplicates in template: {len(duplicates)}")
        else:
            print("\n  No duplicates found in template")
    
    wb.close()

if __name__ == "__main__":
    check_template()