#!/usr/bin/env python3
"""
Examine Sales File Structure

This script examines the monthly dish sales Excel file to understand:
1. How many sheets it contains
2. What the sheet names are
3. Whether sales modes are in separate sheets or columns
4. Sample data structure
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
import logging
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def examine_sales_file_structure():
    """Examine the structure of the monthly dish sales file"""
    
    sales_file = Path("Input/monthly_report/monthly_dish_sale/海外菜品销售报表_20250706_0957.xlsx")
    
    if not sales_file.exists():
        logger.error(f"Sales file not found: {sales_file}")
        return False
    
    try:
        # Load workbook to see sheet structure
        logger.info(f"Examining file: {sales_file}")
        workbook = load_workbook(sales_file, read_only=True)
        
        logger.info(f"Found {len(workbook.sheetnames)} sheets:")
        for i, sheet_name in enumerate(workbook.sheetnames):
            logger.info(f"  Sheet {i+1}: {sheet_name}")
        
        # Examine each sheet
        for sheet_name in workbook.sheetnames:
            logger.info(f"\n--- Examining Sheet: {sheet_name} ---")
            
            try:
                # Read the sheet with pandas
                df = pd.read_excel(sales_file, sheet_name=sheet_name, nrows=10)
                
                logger.info(f"Sheet '{sheet_name}' has {len(df.columns)} columns:")
                for col in df.columns:
                    logger.info(f"  - {col}")
                
                # Check if there's a sales mode column
                mode_columns = [col for col in df.columns if '模式' in str(col) or '方式' in str(col) or 'mode' in str(col).lower()]
                if mode_columns:
                    logger.info(f"Found potential sales mode columns: {mode_columns}")
                    
                    # Show unique values in mode columns
                    for mode_col in mode_columns:
                        unique_values = df[mode_col].dropna().unique()
                        logger.info(f"  Unique values in {mode_col}: {list(unique_values)}")
                
                # Check for sales type indicators
                type_columns = [col for col in df.columns if '类型' in str(col) or '外卖' in str(col) or '堂食' in str(col)]
                if type_columns:
                    logger.info(f"Found potential sales type columns: {type_columns}")
                    
                    for type_col in type_columns:
                        unique_values = df[type_col].dropna().unique()
                        logger.info(f"  Unique values in {type_col}: {list(unique_values)}")
                
                # Look for the specific dish mentioned by user
                dish_columns = [col for col in df.columns if '菜品' in str(col) or '名称' in str(col) or 'name' in str(col).lower()]
                if dish_columns:
                    logger.info(f"Found dish name columns: {dish_columns}")
                    
                    for dish_col in dish_columns:
                        # Look for beef dishes
                        beef_dishes = df[df[dish_col].str.contains('牛仔骨', na=False)]
                        if len(beef_dishes) > 0:
                            logger.info(f"Found {len(beef_dishes)} beef rib dishes in column {dish_col}")
                            for _, row in beef_dishes.iterrows():
                                logger.info(f"  Beef dish: {row[dish_col]}")
                        
                        # Look for the specific dish code mentioned
                        code_matches = df[df[dish_col].str.contains('4510316', na=False)]
                        if len(code_matches) > 0:
                            logger.info(f"Found {len(code_matches)} matches for code 4510316 in column {dish_col}")
                
                # Show sample data (first few rows)
                logger.info(f"Sample data from sheet '{sheet_name}':")
                for i, row in df.head(3).iterrows():
                    logger.info(f"  Row {i+1}: {dict(row)}")
                
            except Exception as e:
                logger.warning(f"Could not read sheet '{sheet_name}': {e}")
        
        workbook.close()
        
    except Exception as e:
        logger.error(f"Error examining sales file: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return True


if __name__ == "__main__":
    success = examine_sales_file_structure()
    if success:
        print("File examination completed!")
    else:
        print("File examination failed!")