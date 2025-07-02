#!/usr/bin/env python3
"""
Tests for MonthlyDishesWorksheetGenerator
"""

from openpyxl import Workbook
from lib.monthly_dishes_worksheet import MonthlyDishesWorksheetGenerator
import unittest
import sys
from pathlib import Path
from datetime import datetime
from unittest.mock import Mock, patch, MagicMock

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))


class TestMonthlyDishesWorksheetGenerator(unittest.TestCase):
    """Comprehensive tests for MonthlyDishesWorksheetGenerator"""

    def setUp(self):
        """Set up test fixtures"""
        self.store_names = {
            1: "加拿大一店", 2: "加拿大二店", 3: "加拿大三店", 4: "加拿大四店",
            5: "加拿大五店", 6: "加拿大六店", 7: "加拿大七店"
        }
        self.target_date = "2025-06-10"
        self.generator = MonthlyDishesWorksheetGenerator(
            self.store_names, self.target_date)

    def test_generator_initialization(self):
        """Test generator initialization"""
        self.assertEqual(self.generator.store_names, self.store_names)
        self.assertEqual(self.generator.target_date, self.target_date)

    def test_generator_initialization_with_different_stores(self):
        """Test generator initialization with different store configuration"""
        different_stores = {1: "Test Store 1", 2: "Test Store 2"}
        generator = MonthlyDishesWorksheetGenerator(
            different_stores, "2025-01-01")
        self.assertEqual(generator.store_names, different_stores)
        self.assertEqual(generator.target_date, "2025-01-01")

    def create_mock_dish_data(self, include_materials=True):
        """Create mock dish data for testing"""
        mock_data = []

        if include_materials:
            # Dish with materials
            mock_data.extend([
                {
                    'dish_type_name': '锅底类',
                    'dish_child_type_name': '锅底类',
                    'dish_id': 1,
                    'dish_name': '清油麻辣火锅',
                    'dish_system_name': '清油麻辣火锅',
                    'dish_code': '1060061',
                    'dish_short_code': None,
                    'dish_size': '单锅',
                    'dish_unit': '锅',
                    'serving_size_kg': 1.2000,
                    'material_id': 1,
                    'material_name': '清油底料',
                    'material_number': '3000759',
                    'material_unit': '公斤',
                    'package_spec': '300G*40包/件',
                    'standard_quantity': 396.000,
                    'current_price': 28.50,
                    'currency': 'CAD',
                    'price_date': datetime(2025, 6, 1),
                    'material_count_for_dish': 2
                },
                {
                    'dish_type_name': '锅底类',
                    'dish_child_type_name': '锅底类',
                    'dish_id': 1,
                    'dish_name': '清油麻辣火锅',
                    'dish_system_name': '清油麻辣火锅',
                    'dish_code': '1060061',
                    'dish_short_code': None,
                    'dish_size': '单锅',
                    'dish_unit': '锅',
                    'serving_size_kg': 1.2000,
                    'material_id': 2,
                    'material_name': '三花淡奶',
                    'material_number': '1500882',
                    'material_unit': '瓶',
                    'package_spec': '354ML*48瓶/箱',
                    'standard_quantity': 101.952,
                    'current_price': 12.30,
                    'currency': 'CAD',
                    'price_date': datetime(2025, 6, 1),
                    'material_count_for_dish': 2
                }
            ])

        # Dish without materials
        mock_data.append({
            'dish_type_name': '素菜类',
            'dish_child_type_name': '素菜类',
            'dish_id': 2,
            'dish_name': '测试菜品',
            'dish_system_name': '测试菜品',
            'dish_code': '2000001',
            'dish_short_code': 'TEST',
            'dish_size': '标准',
            'dish_unit': '份',
            'serving_size_kg': 0.5000,
            'material_id': None,
            'material_name': None,
            'material_number': None,
            'material_unit': None,
            'package_spec': None,
            'standard_quantity': None,
            'current_price': None,
            'currency': None,
            'price_date': None,
            'material_count_for_dish': 0
        })

        return mock_data

    def test_worksheet_creation_with_data(self):
        """Test worksheet creation with mock data"""
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        # Mock data provider
        mock_data_provider = Mock()
        mock_dish_data = self.create_mock_dish_data(include_materials=True)

        with patch.object(self.generator, 'get_dish_material_data', return_value=mock_dish_data):
            ws = self.generator.generate_worksheet(wb, mock_data_provider)

        # Verify worksheet was created
        self.assertIsNotNone(ws)
        self.assertEqual(ws.title, "菜品用料月报")

        # Verify basic structure
        self.assertTrue(ws.max_row > 1)  # Should have content
        self.assertTrue(ws.max_column >= 20)  # Should have all columns

    def test_worksheet_creation_with_no_data(self):
        """Test worksheet creation when no data is available"""
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        # Mock data provider
        mock_data_provider = Mock()

        with patch.object(self.generator, 'get_dish_material_data', return_value=[]):
            ws = self.generator.generate_worksheet(wb, mock_data_provider)

        # Verify worksheet was created with no data message
        self.assertIsNotNone(ws)
        self.assertEqual(ws.title, "菜品用料月报")
        self.assertEqual(ws['A1'].value, "无可用数据 - No Data Available")

    def test_summary_section_calculation(self):
        """Test summary statistics calculation"""
        wb = Workbook()
        ws = wb.active
        mock_dish_data = self.create_mock_dish_data(include_materials=True)

        # Test summary section
        self.generator.add_summary_section(ws, 1, mock_dish_data)

        # Check if summary data was added
        found_summary = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "数据统计概览":
                    found_summary = True
                    break

        self.assertTrue(found_summary, "Summary section should be present")

    def test_dish_data_section_creation(self):
        """Test dish data section creation"""
        wb = Workbook()
        ws = wb.active
        mock_dish_data = self.create_mock_dish_data(include_materials=True)

        # Test dish data section
        self.generator.add_dish_data_section(ws, 1, mock_dish_data)

        # Check if headers were added
        found_headers = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "序号":
                    found_headers = True
                    break

        self.assertTrue(found_headers, "Data headers should be present")

    def test_dish_row_addition_with_material(self):
        """Test adding dish row with material data"""
        wb = Workbook()
        ws = wb.active
        mock_dish_data = self.create_mock_dish_data(include_materials=True)[0]

        # Test adding dish row
        self.generator.add_dish_row(
            ws, 1, 1, mock_dish_data, has_material=True)

        # Verify data was added
        self.assertEqual(ws.cell(1, 1).value, 1)  # Row number
        self.assertEqual(ws.cell(1, 2).value, '锅底类')  # Dish type
        self.assertEqual(ws.cell(1, 4).value, '1060061')  # Dish code
        self.assertEqual(ws.cell(1, 6).value, '清油麻辣火锅')  # Dish name

    def test_dish_row_addition_without_material(self):
        """Test adding dish row without material data"""
        wb = Workbook()
        ws = wb.active
        mock_dish_data = self.create_mock_dish_data(include_materials=False)[0]

        # Test adding dish row without material
        self.generator.add_dish_row(
            ws, 1, 1, mock_dish_data, has_material=False)

        # Verify data was added
        self.assertEqual(ws.cell(1, 1).value, 1)  # Row number
        self.assertEqual(ws.cell(1, 6).value, '测试菜品')  # Dish name
        # Remarks for no material
        self.assertEqual(ws.cell(1, 20).value, '暂无配方')

    def test_formatting_application(self):
        """Test formatting application"""
        wb = Workbook()
        ws = wb.active

        # Add some test data
        ws['A1'] = "序号"
        ws['A2'] = 1
        ws['A3'] = 2

        # Apply formatting
        self.generator.apply_common_formatting(ws, 3)

        # Verify borders were applied (checking if border is not None)
        self.assertIsNotNone(ws.cell(1, 1).border)

    def test_database_query_error_handling(self):
        """Test error handling in database queries"""
        # Mock data provider with database error
        mock_data_provider = Mock()
        mock_data_provider.db_manager.get_connection.side_effect = Exception(
            "Database connection failed")

        # Test that error is handled gracefully
        result = self.generator.get_dish_material_data(mock_data_provider)
        self.assertEqual(result, [])

    def test_date_formatting_in_title(self):
        """Test date formatting in worksheet title"""
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        mock_data_provider = Mock()
        mock_dish_data = self.create_mock_dish_data(include_materials=True)

        with patch.object(self.generator, 'get_dish_material_data', return_value=mock_dish_data):
            ws = self.generator.generate_worksheet(wb, mock_data_provider)

        # Check title contains formatted date
        title_cell = ws['A1']
        self.assertIn("2025年06月", title_cell.value)

    def test_column_width_settings(self):
        """Test column width settings"""
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        mock_data_provider = Mock()
        mock_dish_data = self.create_mock_dish_data(include_materials=True)

        with patch.object(self.generator, 'get_dish_material_data', return_value=mock_dish_data):
            ws = self.generator.generate_worksheet(wb, mock_data_provider)

        # Check that column widths were set
        for col_num in range(1, 11):  # Check first 10 columns
            col_letter = chr(64 + col_num)  # A, B, C, etc.
            self.assertIsNotNone(ws.column_dimensions[col_letter].width)

    def test_mixed_dish_data_processing(self):
        """Test processing dishes with and without materials"""
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        mock_data_provider = Mock()
        # Mix of dishes with and without materials
        mock_dish_data = self.create_mock_dish_data(include_materials=True)

        with patch.object(self.generator, 'get_dish_material_data', return_value=mock_dish_data):
            ws = self.generator.generate_worksheet(wb, mock_data_provider)

        # Verify both types of dishes were processed
        self.assertIsNotNone(ws)
        self.assertTrue(ws.max_row > 5)  # Should have multiple rows

    def test_error_recovery(self):
        """Test error recovery during worksheet generation"""
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        # Mock data provider that raises exception during data fetch
        mock_data_provider = Mock()

        # Test that worksheet is still created even with data fetch error
        with patch.object(self.generator, 'get_dish_material_data', side_effect=Exception("Test error")):
            # Should not raise exception
            try:
                ws = self.generator.generate_worksheet(wb, mock_data_provider)
                self.assertIsNotNone(ws)
            except Exception as e:
                self.fail(
                    f"Worksheet generation should handle errors gracefully: {e}")


if __name__ == '__main__':
    unittest.main()
