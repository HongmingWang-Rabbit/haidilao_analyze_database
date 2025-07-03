#!/usr/bin/env python3
"""
Tests for Detailed Material Spending Worksheet Generator
"""

from lib.detailed_material_spending_worksheet import DetailedMaterialSpendingGenerator
import unittest
from unittest.mock import Mock, patch, MagicMock
import sys
from pathlib import Path
from openpyxl import Workbook
from datetime import datetime

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))


class TestDetailedMaterialSpendingGenerator(unittest.TestCase):
    """Comprehensive tests for DetailedMaterialSpendingGenerator."""

    def setUp(self):
        """Set up test fixtures."""
        self.mock_data_provider = Mock()
        self.mock_db_manager = Mock()
        self.mock_data_provider.db_manager = self.mock_db_manager
        self.generator = DetailedMaterialSpendingGenerator(
            self.mock_data_provider)

    def test_generator_initialization(self):
        """Test generator initialization."""
        self.assertIsNotNone(self.generator)
        self.assertEqual(self.generator.data_provider, self.mock_data_provider)
        self.assertIsNotNone(self.generator.logger)

    def test_get_store_data_success(self):
        """Test successful store data retrieval."""
        # Mock store data
        mock_stores = [
            {'id': 1, 'name': '加拿大一店', 'district_id': 1, 'is_active': True},
            {'id': 2, 'name': '加拿大二店', 'district_id': 1, 'is_active': True}
        ]
        self.mock_db_manager.fetch_all.return_value = mock_stores

        result = self.generator._get_store_data()

        # Verify query execution
        self.mock_db_manager.fetch_all.assert_called_once()
        call_args = self.mock_db_manager.fetch_all.call_args[0][0]
        self.assertIn("SELECT id, name, district_id, is_active", call_args)
        self.assertIn("FROM store", call_args)
        self.assertIn("WHERE is_active = TRUE", call_args)

        # Verify result structure
        self.assertEqual(len(result), 2)
        self.assertEqual(result[1]['name'], '加拿大一店')
        self.assertEqual(result[2]['name'], '加拿大二店')

    def test_get_store_data_empty(self):
        """Test store data retrieval with empty result."""
        self.mock_db_manager.fetch_all.return_value = []

        result = self.generator._get_store_data()

        self.assertEqual(result, {})

    def test_get_store_data_error(self):
        """Test store data retrieval with database error."""
        self.mock_db_manager.fetch_all.side_effect = Exception("DB Error")

        result = self.generator._get_store_data()

        self.assertEqual(result, {})

    def test_get_detailed_material_spending_success(self):
        """Test successful detailed material spending data retrieval."""
        # Mock spending data
        mock_data = [
            {
                'material_number': '1001',
                'material_name': '牛肉片',
                'material_type_name': '成本-荤菜类',
                'material_child_type_name': None,
                'material_used': 100.0,
                'unit_price': 25.50,
                'total_amount': 2550.0,
                'unit': 'KG',
                'package_spec': '10KG/箱',
                'type_sort_order': 1,
                'child_type_sort_order': 999
            },
            {
                'material_number': '2001',
                'material_name': '白菜',
                'material_type_name': '成本-素菜类',
                'material_child_type_name': None,
                'material_used': 50.0,
                'unit_price': 3.20,
                'total_amount': 160.0,
                'unit': 'KG',
                'package_spec': '20KG/箱',
                'type_sort_order': 2,
                'child_type_sort_order': 999
            }
        ]
        self.mock_db_manager.fetch_all.return_value = mock_data

        result = self.generator._get_detailed_material_spending(1, 2025, 6)

        # Verify query execution
        self.mock_db_manager.fetch_all.assert_called_once()
        call_args = self.mock_db_manager.fetch_all.call_args
        query = call_args[0][0]
        params = call_args[0][1]

        # Verify query structure
        self.assertIn("SELECT", query)
        self.assertIn("material_number", query)
        self.assertIn("material_name", query)
        self.assertIn("material_type_name", query)
        self.assertIn("FROM material_monthly_usage mmu", query)
        self.assertIn("JOIN material m ON mmu.material_id = m.id", query)
        self.assertIn("WHERE mmu.store_id = %s", query)
        self.assertIn("AND mmu.year = %s", query)
        self.assertIn("AND mmu.month = %s", query)

        # Verify parameters
        self.assertEqual(params, (1, 1, 2025, 6))

        # Verify result
        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]['material_name'], '牛肉片')
        self.assertEqual(result[1]['material_name'], '白菜')

    def test_get_detailed_material_spending_empty(self):
        """Test detailed material spending retrieval with empty result."""
        self.mock_db_manager.fetch_all.return_value = []

        result = self.generator._get_detailed_material_spending(1, 2025, 6)

        self.assertEqual(result, [])

    def test_get_detailed_material_spending_error(self):
        """Test detailed material spending retrieval with database error."""
        self.mock_db_manager.fetch_all.side_effect = Exception("DB Error")

        result = self.generator._get_detailed_material_spending(1, 2025, 6)

        self.assertEqual(result, [])

    def test_create_detailed_spending_worksheet_structure(self):
        """Test detailed spending worksheet structure creation."""
        # Create test workbook and worksheet
        wb = Workbook()
        ws = wb.create_sheet("Test")

        # Mock store info
        store_info = {
            'id': 1,
            'name': '加拿大一店',
            'district_id': 1
        }

        # Mock spending data
        spending_data = [
            {
                'material_number': '1001',
                'material_name': '牛肉片',
                'material_type_name': '成本-荤菜类',
                'material_child_type_name': None,
                'material_used': 100.0,
                'unit_price': 25.50,
                'total_amount': 2550.0,
                'unit': 'KG',
                'package_spec': '10KG/箱',
                'type_sort_order': 1,
                'child_type_sort_order': 999
            }
        ]

        # Generate worksheet
        self.generator._create_detailed_spending_worksheet(
            ws, store_info, spending_data, "2025-06")

        # Verify header structure
        self.assertIn("海底捞物料消耗明细报表", ws['A1'].value)
        self.assertIn("加拿大一店", ws['A1'].value)

        # Verify column headers
        self.assertEqual(ws['A4'].value, "物料编号 - 物料名称")
        self.assertEqual(ws['B4'].value, "本月使用金额")
        self.assertEqual(ws['C4'].value, "分类")
        self.assertEqual(ws['D4'].value, "使用数量")
        self.assertEqual(ws['E4'].value, "单价")
        self.assertEqual(ws['F4'].value, "单位")

        # Verify data exists - check multiple rows to find material type header
        found_type_header = False
        found_material_item = False

        for row_num in range(5, 15):  # Check rows 5-14
            cell_value = str(ws.cell(row=row_num, column=1).value or "")
            if "成本-荤菜类" in cell_value:
                found_type_header = True
            if "1001 牛肉片" in cell_value:
                found_material_item = True

        self.assertTrue(found_type_header, "Material type header not found")
        self.assertTrue(found_material_item, "Material item not found")

    def test_create_detailed_spending_worksheet_calculations(self):
        """Test worksheet calculations and totals."""
        wb = Workbook()
        ws = wb.create_sheet("Test")

        store_info = {'id': 1, 'name': '测试店', 'district_id': 1}

        # Mock multiple materials with different types
        spending_data = [
            {
                'material_number': '1001',
                'material_name': '牛肉片',
                'material_type_name': '成本-荤菜类',
                'material_child_type_name': None,
                'material_used': 100.0,
                'unit_price': 25.50,
                'total_amount': 2550.0,
                'unit': 'KG',
                'package_spec': '10KG/箱',
                'type_sort_order': 1,
                'child_type_sort_order': 999
            },
            {
                'material_number': '1002',
                'material_name': '猪肉片',
                'material_type_name': '成本-荤菜类',
                'material_child_type_name': None,
                'material_used': 80.0,
                'unit_price': 22.00,
                'total_amount': 1760.0,
                'unit': 'KG',
                'package_spec': '10KG/箱',
                'type_sort_order': 1,
                'child_type_sort_order': 999
            },
            {
                'material_number': '2001',
                'material_name': '白菜',
                'material_type_name': '成本-素菜类',
                'material_child_type_name': None,
                'material_used': 50.0,
                'unit_price': 3.20,
                'total_amount': 160.0,
                'unit': 'KG',
                'package_spec': '20KG/箱',
                'type_sort_order': 2,
                'child_type_sort_order': 999
            }
        ]

        self.generator._create_detailed_spending_worksheet(
            ws, store_info, spending_data, "2025-06")

        # Find and verify subtotals
        found_meat_subtotal = False
        found_vegetable_subtotal = False
        found_grand_total = False

        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=6):
            cell_value = str(row[0].value) if row[0].value else ""
            amount_value = row[1].value

            if "小计 - 成本-荤菜类" in cell_value:
                found_meat_subtotal = True
                self.assertEqual(amount_value, 4310.0)  # 2550 + 1760
            elif "小计 - 成本-素菜类" in cell_value:
                found_vegetable_subtotal = True
                self.assertEqual(amount_value, 160.0)
            elif cell_value == "总计":
                found_grand_total = True
                self.assertEqual(amount_value, 4470.0)  # 4310 + 160

        self.assertTrue(found_meat_subtotal,
                        "Meat category subtotal not found")
        self.assertTrue(found_vegetable_subtotal,
                        "Vegetable category subtotal not found")
        self.assertTrue(found_grand_total, "Grand total not found")

    def test_generate_worksheets_success(self):
        """Test successful worksheet generation for multiple stores."""
        # Mock store data
        mock_stores = [
            {'id': 1, 'name': '加拿大一店', 'district_id': 1, 'is_active': True},
            {'id': 2, 'name': '加拿大二店', 'district_id': 1, 'is_active': True}
        ]

        # Mock spending data
        mock_spending_data = [
            {
                'material_number': '1001',
                'material_name': '牛肉片',
                'material_type_name': '成本-荤菜类',
                'material_child_type_name': None,
                'material_used': 100.0,
                'unit_price': 25.50,
                'total_amount': 2550.0,
                'unit': 'KG',
                'package_spec': '10KG/箱',
                'type_sort_order': 1,
                'child_type_sort_order': 999
            }
        ]

        # Setup mocks
        self.mock_db_manager.fetch_all.side_effect = [
            mock_stores,  # First call for stores
            mock_spending_data,  # Second call for store 1 spending
            mock_spending_data   # Third call for store 2 spending
        ]

        # Create test workbook
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        # Generate worksheets
        self.generator.generate_worksheets(wb, "2025-06-30")

        # Verify worksheets were created
        self.assertEqual(len(wb.worksheets), 2)
        worksheet_names = [ws.title for ws in wb.worksheets]
        self.assertIn("物料明细-加拿大一店", worksheet_names)
        self.assertIn("物料明细-加拿大二店", worksheet_names)

    def test_generate_worksheets_no_stores(self):
        """Test worksheet generation with no stores."""
        # Mock empty store data
        self.mock_db_manager.fetch_all.return_value = []

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        # Should not raise exception
        self.generator.generate_worksheets(wb, "2025-06-30")

        # No worksheets should be created
        self.assertEqual(len(wb.worksheets), 0)

    def test_generate_worksheets_no_spending_data(self):
        """Test worksheet generation with stores but no spending data."""
        # Mock store data but no spending data
        mock_stores = [
            {'id': 1, 'name': '加拿大一店', 'district_id': 1, 'is_active': True}
        ]

        self.mock_db_manager.fetch_all.side_effect = [
            mock_stores,  # First call for stores
            []            # Second call for spending data (empty)
        ]

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        # Should not raise exception
        self.generator.generate_worksheets(wb, "2025-06-30")

        # No worksheets should be created due to no spending data
        self.assertEqual(len(wb.worksheets), 0)

    def test_generate_worksheets_error_handling(self):
        """Test worksheet generation error handling."""
        # Mock database error
        self.mock_db_manager.fetch_all.side_effect = Exception("DB Error")

        wb = Workbook()

        # Should handle error gracefully and not crash
        try:
            self.generator.generate_worksheets(wb, "2025-06-30")
            # If we get here, error was handled gracefully
            self.assertTrue(True)
        except Exception:
            self.fail("Exception should be handled gracefully")

    def test_apply_formatting(self):
        """Test worksheet formatting application."""
        wb = Workbook()
        ws = wb.create_sheet("Test")

        # Add some test data
        ws['A1'] = "Test"
        ws['B1'] = "Data"

        # Apply formatting
        self.generator._apply_formatting(ws, 1, 5)

        # Verify column widths are set
        self.assertEqual(ws.column_dimensions['A'].width, 35)
        self.assertEqual(ws.column_dimensions['B'].width, 15)
        self.assertEqual(ws.column_dimensions['C'].width, 20)

        # Verify borders are applied
        for row in range(1, 6):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                self.assertIsNotNone(cell.border)

    def test_worksheet_name_length_handling(self):
        """Test handling of long store names in worksheet names."""
        # Mock store with very long name
        mock_stores = [
            {'id': 999, 'name': '非常长的店名超过三十一个字符限制测试用店铺名称',
                'district_id': 1, 'is_active': True}
        ]

        mock_spending_data = [
            {
                'material_number': '1001',
                'material_name': '牛肉片',
                'material_type_name': '成本-荤菜类',
                'material_child_type_name': None,
                'material_used': 100.0,
                'unit_price': 25.50,
                'total_amount': 2550.0,
                'unit': 'KG',
                'package_spec': '10KG/箱',
                'type_sort_order': 1,
                'child_type_sort_order': 999
            }
        ]

        self.mock_db_manager.fetch_all.side_effect = [
            mock_stores,
            mock_spending_data
        ]

        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        self.generator.generate_worksheets(wb, "2025-06-30")

        # Verify worksheet name falls back to store ID
        self.assertEqual(len(wb.worksheets), 1)
        worksheet_title = wb.worksheets[0].title
        # Should either be the truncated version or fallback to ID
        self.assertTrue(
            worksheet_title == "物料明细-999" or "999" in worksheet_title,
            f"Unexpected worksheet title: {worksheet_title}"
        )
        self.assertLessEqual(len(worksheet_title), 31)


if __name__ == '__main__':
    unittest.main()
