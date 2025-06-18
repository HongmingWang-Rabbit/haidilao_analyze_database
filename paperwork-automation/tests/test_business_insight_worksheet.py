#!/usr/bin/env python3
"""
Tests for business insight worksheet generator.
"""

import unittest
from datetime import datetime
from openpyxl import Workbook
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent))
from lib.business_insight_worksheet import BusinessInsightWorksheetGenerator

class TestBusinessInsightWorksheet(unittest.TestCase):
    """Test business insight worksheet generation"""
    
    def setUp(self):
        """Set up test data"""
        self.store_names = {
            1: "加拿大一店", 2: "加拿大二店", 3: "加拿大三店", 4: "加拿大四店",
            5: "加拿大五店", 6: "加拿大六店", 7: "加拿大七店"
        }
        self.target_date = "2025-06-10"
        self.generator = BusinessInsightWorksheetGenerator(self.store_names, self.target_date)
        
        # Test data
        self.daily_data = {
            1: {
                'revenue_tax_not_included': 22194.3,
                'tables_served': 188.8,
                'tables_served_validated': 188.8,
                'customers': 4229.3,
                'takeout_tables': 10.2,
                'turnover_rate': 1.5
            },
            2: {
                'revenue_tax_not_included': 12139.0,
                'tables_served': 76.0,
                'tables_served_validated': 76.0,
                'customers': 2246.6,
                'takeout_tables': 8.0,
                'turnover_rate': 1.2
            }
        }
        
        self.monthly_data = {
            1: {
                'monthly_revenue': 221943.0,
                'monthly_tables': 1888.0,
                'monthly_tables_validated': 1888.0,
                'customers': 42293.0,
                'takeout_tables': 102.0,
                'tables_served': 1900.0,
                'avg_turnover_rate': 1.8,
                'target_revenue': 250000.0,
                'prev_monthly_revenue': 199748.7,
                'prev_monthly_tables_validated': 1699.2,
                'prev_month_avg_per_table': 117.5
            },
            2: {
                'monthly_revenue': 121390.0,
                'monthly_tables': 760.0,
                'monthly_tables_validated': 760.0,
                'customers': 22466.0,
                'takeout_tables': 80.0,
                'tables_served': 770.0,
                'avg_turnover_rate': 1.6,
                'target_revenue': 150000.0,
                'prev_monthly_revenue': 109251.0,
                'prev_monthly_tables_validated': 684.0,
                'prev_month_avg_per_table': 159.7
            }
        }
        
        self.previous_month_data = {
            1: {
                'prev_monthly_revenue': 199748.7,
                'prev_monthly_tables': 1699.2,
                'prev_monthly_tables_validated': 1699.2,
                'prev_monthly_customers': 38063.7
            },
            2: {
                'prev_monthly_revenue': 109251.0,
                'prev_monthly_tables': 684.0,
                'prev_monthly_tables_validated': 684.0,
                'prev_monthly_customers': 20229.7
            }
        }
        
        self.current_mtd = {
            1: {
                'mtd_revenue': 221943.0,
                'mtd_tables': 1888.0,
                'mtd_tables_served': 1888.0,
                'mtd_customers': 42293.0,
                'mtd_discount_total': 2000.0
            },
            2: {
                'mtd_revenue': 121390.0,
                'mtd_tables': 760.0,
                'mtd_tables_served': 760.0,
                'mtd_customers': 22466.0,
                'mtd_discount_total': 1500.0
            }
        }
        
        self.prev_mtd = {
            1: {
                'prev_mtd_revenue': 199748.7,
                'prev_mtd_tables': 1699.2,
                'customers': 38063.7
            },
            2: {
                'prev_mtd_revenue': 109251.0,
                'prev_mtd_tables': 684.0,
                'customers': 20229.7
            }
        }
        
        self.daily_ranking = {1: 1, 2: 2}
        self.monthly_ranking = {1: 1, 2: 2}
        self.daily_ranking_values = [22194.3, 12139.0]
        self.monthly_ranking_values = [221943.0, 121390.0]
    
    def test_generator_initialization(self):
        """Test generator initialization"""
        self.assertEqual(self.generator.store_names, self.store_names)
        self.assertEqual(self.generator.target_date, self.target_date)
    
    def test_worksheet_creation(self):
        """Test worksheet creation"""
        wb = Workbook()
        ws = self.generator.generate_worksheet(
            wb, self.daily_data, self.monthly_data, self.previous_month_data,
            self.monthly_data, self.current_mtd, self.prev_mtd,
            self.daily_ranking, self.monthly_ranking, 
            self.daily_ranking_values, self.monthly_ranking_values
        )
        
        # Check worksheet exists
        self.assertIsNotNone(ws)
        self.assertEqual(ws.title, "营业透视")
        
        # Check worksheet is in workbook
        self.assertIn("营业透视", [sheet.title for sheet in wb.worksheets])
    
    def test_date_title(self):
        """Test date title formatting"""
        wb = Workbook()
        ws = self.generator.generate_worksheet(
            wb, self.daily_data, self.monthly_data, self.previous_month_data,
            self.monthly_data, self.current_mtd, self.prev_mtd,
            self.daily_ranking, self.monthly_ranking,
            self.daily_ranking_values, self.monthly_ranking_values
        )
        
        # Check date title
        title_cell = ws['A1']
        self.assertIn("2025-06-10", title_cell.value)
        self.assertTrue(title_cell.font.bold)
    
    def test_store_basic_info_section(self):
        """Test store basic information section"""
        wb = Workbook()
        ws = self.generator.generate_worksheet(
            wb, self.daily_data, self.monthly_data, self.previous_month_data,
            self.monthly_data, self.current_mtd, self.prev_mtd,
            self.daily_ranking, self.monthly_ranking,
            self.daily_ranking_values, self.monthly_ranking_values
        )
        
        # Find headers row (should be row 2)
        headers_found = False
        for row in ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                if cell.value == "门店":
                    headers_found = True
                    break
            if headers_found:
                break
        
        self.assertTrue(headers_found, "Store basic info headers not found")
        
        # Check data rows contain store names
        store_names_found = []
        for row in ws.iter_rows(min_row=3, max_row=10):
            if row[0].value in self.store_names.values():
                store_names_found.append(row[0].value)
        
        self.assertGreater(len(store_names_found), 0, "No store data found")
    
    def test_business_analysis_section(self):
        """Test business data analysis section"""
        wb = Workbook()
        ws = self.generator.generate_worksheet(
            wb, self.daily_data, self.monthly_data, self.previous_month_data,
            self.monthly_data, self.current_mtd, self.prev_mtd,
            self.daily_ranking, self.monthly_ranking,
            self.daily_ranking_values, self.monthly_ranking_values
        )
        
        # Find business analysis section
        analysis_section_found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "营业数据分析":
                    analysis_section_found = True
                    break
            if analysis_section_found:
                break
        
        self.assertTrue(analysis_section_found, "Business analysis section not found")
    
    def test_turnover_analysis_section(self):
        """Test turnover rate analysis section"""
        wb = Workbook()
        ws = self.generator.generate_worksheet(
            wb, self.daily_data, self.monthly_data, self.previous_month_data,
            self.monthly_data, self.current_mtd, self.prev_mtd,
            self.daily_ranking, self.monthly_ranking,
            self.daily_ranking_values, self.monthly_ranking_values
        )
        
        # Find turnover analysis section
        turnover_section_found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "翻台率分析":
                    turnover_section_found = True
                    break
            if turnover_section_found:
                break
        
        self.assertTrue(turnover_section_found, "Turnover analysis section not found")
    
    def test_data_calculations(self):
        """Test data calculations are correct"""
        wb = Workbook()
        ws = self.generator.generate_worksheet(
            wb, self.daily_data, self.monthly_data, self.previous_month_data,
            self.monthly_data, self.current_mtd, self.prev_mtd,
            self.daily_ranking, self.monthly_ranking,
            self.daily_ranking_values, self.monthly_ranking_values
        )
        
        # Find data rows and verify calculations
        data_found = False
        for row in ws.iter_rows(min_row=3, max_row=15):
            if row[0].value == "加拿大一店":
                data_found = True
                # Check daily revenue calculation (22194.3 / 10000 = 2.22)
                daily_revenue = row[1].value
                self.assertAlmostEqual(daily_revenue, 2.22, places=1)
                break
        
        self.assertTrue(data_found, "Store data calculations not found")
    
    def test_formatting_applied(self):
        """Test formatting is applied correctly"""
        wb = Workbook()
        ws = self.generator.generate_worksheet(
            wb, self.daily_data, self.monthly_data, self.previous_month_data,
            self.monthly_data, self.current_mtd, self.prev_mtd,
            self.daily_ranking, self.monthly_ranking,
            self.daily_ranking_values, self.monthly_ranking_values
        )
        
        # Check that cells have borders and alignment
        cells_with_formatting = 0
        for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=5):
            for cell in row:
                if cell.value is not None:
                    if cell.border.left.style or cell.alignment.horizontal:
                        cells_with_formatting += 1
        
        self.assertGreater(cells_with_formatting, 0, "No formatting applied")
    
    def test_column_widths_set(self):
        """Test column widths are set properly"""
        wb = Workbook()
        ws = self.generator.generate_worksheet(
            wb, self.daily_data, self.monthly_data, self.previous_month_data,
            self.monthly_data, self.current_mtd, self.prev_mtd,
            self.daily_ranking, self.monthly_ranking,
            self.daily_ranking_values, self.monthly_ranking_values
        )
        
        # Check that column widths are set
        column_widths_set = 0
        for col in ['A', 'B', 'C', 'D', 'E']:
            if ws.column_dimensions[col].width > 0:
                column_widths_set += 1
        
        self.assertGreater(column_widths_set, 0, "Column widths not set")

if __name__ == '__main__':
    unittest.main() 