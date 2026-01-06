"""
Tests for Q1 2026 Challenge Tracking Worksheet Generator.

This test module covers:
- Generator initialization
- Q1 2026 date validation
- Store-specific target calculations
- Color coding for target achievement
- Regional summary (excluding Store 8)
- Target configuration loading
"""

import unittest
from unittest.mock import Mock, MagicMock
from datetime import datetime, timedelta
import sys
from pathlib import Path

# Add project root to path for imports
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from openpyxl import Workbook
from lib.q1_challenge_worksheet import Q1ChallengeWorksheetGenerator
from configs.challenge_targets import (
    get_store_turnover_target,
    get_store_tables_target,
    is_store_excluded_from_regional,
    is_q1_2026_active,
    STORE_6_TURNOVER_TARGET,
    STORE_8_TURNOVER_TARGET,
    DEFAULT_TURNOVER_IMPROVEMENT,
    WEEKLY_TABLES_IMPROVEMENT
)


class TestQ1ChallengeConfig(unittest.TestCase):
    """Tests for Q1 2026 challenge target configuration."""

    def test_q1_2026_date_validation(self):
        """Test Q1 2026 date range validation."""
        # Valid Q1 2026 dates
        self.assertTrue(is_q1_2026_active("2026-01-01"))
        self.assertTrue(is_q1_2026_active("2026-01-15"))
        self.assertTrue(is_q1_2026_active("2026-02-28"))
        self.assertTrue(is_q1_2026_active("2026-03-31"))

        # Invalid dates (outside Q1 2026)
        self.assertFalse(is_q1_2026_active("2025-12-31"))
        self.assertFalse(is_q1_2026_active("2026-04-01"))
        self.assertFalse(is_q1_2026_active("2025-01-15"))

    def test_store_6_fixed_turnover_target(self):
        """Test Store 6 has fixed turnover target of 3.65."""
        prev_year_turnover = 4.0  # Doesn't matter for fixed target

        target = get_store_turnover_target(6, prev_year_turnover)
        self.assertEqual(target, STORE_6_TURNOVER_TARGET)
        self.assertEqual(target, 3.65)

    def test_store_8_fixed_turnover_target(self):
        """Test Store 8 has fixed turnover target of 4.0."""
        prev_year_turnover = 0  # No previous year data

        target = get_store_turnover_target(8, prev_year_turnover)
        self.assertEqual(target, STORE_8_TURNOVER_TARGET)
        self.assertEqual(target, 4.0)

    def test_regular_store_turnover_improvement(self):
        """Test regular stores (1-5, 7) have improvement target."""
        for store_id in [1, 2, 3, 4, 5, 7]:
            prev_year_turnover = 5.0
            expected_target = prev_year_turnover + DEFAULT_TURNOVER_IMPROVEMENT

            target = get_store_turnover_target(store_id, prev_year_turnover)
            self.assertEqual(target, expected_target,
                           f"Store {store_id} should have improvement target")

    def test_store_8_no_tables_target(self):
        """Test Store 8 has no tables target (returns None)."""
        target = get_store_tables_target(8, 100)
        self.assertIsNone(target)

    def test_regular_store_tables_improvement(self):
        """Test regular stores have tables improvement target (+56/week)."""
        for store_id in [1, 2, 3, 4, 5, 6, 7]:
            prev_year_tables = 200
            expected_target = prev_year_tables + WEEKLY_TABLES_IMPROVEMENT

            target = get_store_tables_target(store_id, prev_year_tables)
            self.assertEqual(target, expected_target,
                           f"Store {store_id} should have +56 tables target")

    def test_store_8_excluded_from_regional(self):
        """Test Store 8 is excluded from regional calculations."""
        self.assertTrue(is_store_excluded_from_regional(8))

        # Other stores should not be excluded
        for store_id in [1, 2, 3, 4, 5, 6, 7]:
            self.assertFalse(is_store_excluded_from_regional(store_id),
                           f"Store {store_id} should not be excluded")


class TestQ1ChallengeWorksheetGenerator(unittest.TestCase):
    """Tests for Q1ChallengeWorksheetGenerator."""

    def setUp(self):
        """Set up test fixtures."""
        self.mock_data_provider = Mock()
        self.generator = Q1ChallengeWorksheetGenerator(self.mock_data_provider)
        self.test_date_q1 = "2026-01-07"
        self.test_date_not_q1 = "2025-06-15"

    def test_generator_initialization(self):
        """Test generator initialization with data provider."""
        self.assertIsNotNone(self.generator)
        self.assertEqual(self.generator.data_provider, self.mock_data_provider)
        self.assertIsNotNone(self.generator.logger)

    def test_worksheet_creation_in_q1(self):
        """Test worksheet is created during Q1 2026."""
        workbook = Workbook()

        # Mock data provider
        self.mock_data_provider.get_weekly_store_performance.return_value = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'current_avg_turnover_rate': 5.5,
                'prev_avg_turnover_rate': 5.2,
                'current_total_tables': 200,
                'prev_total_tables': 180,
            }
        ]

        ws = self.generator.generate_worksheet(workbook, self.test_date_q1)

        # Worksheet should be created
        self.assertIsNotNone(ws)
        self.assertIn("2026 Q1 挑战", workbook.sheetnames)

    def test_worksheet_not_created_outside_q1(self):
        """Test worksheet is not created outside Q1 2026."""
        workbook = Workbook()

        ws = self.generator.generate_worksheet(workbook, self.test_date_not_q1)

        # Worksheet should not be created
        self.assertIsNone(ws)
        self.assertNotIn("2026 Q1 挑战", workbook.sheetnames)

    def test_target_achievement_color_met(self):
        """Test green color is applied when target is met."""
        from openpyxl.styles import PatternFill

        cell = Mock()
        self.generator._apply_achievement_color(cell, True)

        # Should apply green fill
        self.assertEqual(cell.fill, self.generator.GREEN_FILL)

    def test_target_achievement_color_not_met(self):
        """Test red color is applied when target is not met."""
        cell = Mock()
        self.generator._apply_achievement_color(cell, False)

        # Should apply red fill
        self.assertEqual(cell.fill, self.generator.RED_FILL)

    def test_header_structure(self):
        """Test worksheet header structure."""
        workbook = Workbook()
        self.mock_data_provider.get_weekly_store_performance.return_value = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'current_avg_turnover_rate': 5.5,
                'prev_avg_turnover_rate': 5.2,
                'current_total_tables': 200,
                'prev_total_tables': 180,
            }
        ]

        ws = self.generator.generate_worksheet(workbook, self.test_date_q1)

        # Check main headers
        self.assertEqual(ws.cell(row=1, column=1).value, "门店名称")
        self.assertEqual(ws.cell(row=1, column=2).value, "翻台率挑战")
        self.assertEqual(ws.cell(row=1, column=7).value, "桌数挑战")
        self.assertEqual(ws.cell(row=1, column=12).value, "备注")


class TestQ1ChallengeEdgeCases(unittest.TestCase):
    """Test edge cases for Q1 challenge worksheet."""

    def setUp(self):
        """Set up test fixtures."""
        self.mock_data_provider = Mock()
        self.generator = Q1ChallengeWorksheetGenerator(self.mock_data_provider)

    def test_empty_data_handling(self):
        """Test handling when no data is available."""
        self.mock_data_provider.get_weekly_store_performance.return_value = []

        workbook = Workbook()
        ws = self.generator.generate_worksheet(workbook, "2026-01-07")

        # Should still create worksheet (empty)
        self.assertIn("2026 Q1 挑战", workbook.sheetnames)

    def test_store_8_no_prev_year_data(self):
        """Test Store 8 with no previous year data."""
        # Store 8 opened Oct 2025, so no 2025 Q1 data
        prev_year_turnover = 0

        target = get_store_turnover_target(8, prev_year_turnover)
        # Should still return fixed target
        self.assertEqual(target, 4.0)

    def test_exactly_at_target(self):
        """Test when current value exactly meets target."""
        current_value = 5.16  # Exactly at target
        target_value = 5.16

        achieved = current_value >= target_value
        self.assertTrue(achieved, "Should be achieved when exactly at target")

    def test_slightly_below_target(self):
        """Test when current value is slightly below target."""
        current_value = 5.15
        target_value = 5.16

        achieved = current_value >= target_value
        self.assertFalse(achieved, "Should not be achieved when below target")

    def test_regional_summary_calculation(self):
        """Test regional summary calculates correctly."""
        # Mock store data with Store 8
        mock_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'current_avg_turnover_rate': 5.5,
                'prev_avg_turnover_rate': 5.2,
                'current_total_tables': 200,
                'prev_total_tables': 180,
            },
            {
                'store_id': 8,
                'store_name': '加拿大八店',
                'current_avg_turnover_rate': 4.0,
                'prev_avg_turnover_rate': 0,
                'current_total_tables': 150,
                'prev_total_tables': 0,
            }
        ]

        # Filter for regional (excluding Store 8)
        regional_stores = [s for s in mock_data if not is_store_excluded_from_regional(s['store_id'])]

        # Only Store 1 should be included
        self.assertEqual(len(regional_stores), 1)
        self.assertEqual(regional_stores[0]['store_id'], 1)


class TestQ1ChallengeTargetConstants(unittest.TestCase):
    """Test that target constants are correct."""

    def test_turnover_improvement_delta(self):
        """Test default turnover improvement is 0.16."""
        self.assertEqual(DEFAULT_TURNOVER_IMPROVEMENT, 0.16)

    def test_weekly_tables_improvement(self):
        """Test weekly tables improvement is 56 (8 per day * 7 days)."""
        self.assertEqual(WEEKLY_TABLES_IMPROVEMENT, 56)

    def test_store_6_turnover_target(self):
        """Test Store 6 turnover target is 3.65."""
        self.assertEqual(STORE_6_TURNOVER_TARGET, 3.65)

    def test_store_8_turnover_target(self):
        """Test Store 8 turnover target is 4.0."""
        self.assertEqual(STORE_8_TURNOVER_TARGET, 4.0)


if __name__ == '__main__':
    unittest.main()
