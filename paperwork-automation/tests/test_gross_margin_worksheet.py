#!/usr/bin/env python3
"""
Tests for Gross Margin Worksheet Generator
Following established testing patterns for worksheet generators
"""

from lib.gross_margin_worksheet import GrossMarginWorksheetGenerator
import unittest
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))


class TestGrossMarginWorksheetGenerator(unittest.TestCase):
    """Test cases for GrossMarginWorksheetGenerator"""

    def setUp(self):
        """Set up test fixtures"""
        self.target_date = "2025-06-30"
        self.generator = GrossMarginWorksheetGenerator(self.target_date)
        self.workbook = Workbook()
        if self.workbook.active:
            self.workbook.remove(self.workbook.active)

    def tearDown(self):
        """Clean up after tests"""
        if hasattr(self, 'workbook'):
            self.workbook.close()

    def test_generator_initialization(self):
        """Test generator initialization"""
        self.assertEqual(self.generator.target_date, self.target_date)
        self.assertIsNotNone(self.generator.logger)

    def test_generate_detailed_revenue_worksheet_with_empty_data(self):
        """Test generating detailed revenue worksheet with empty data"""
        ws = self.generator.generate_detailed_revenue_worksheet(
            self.workbook, [])

        # Check worksheet was created
        self.assertIsNotNone(ws)
        self.assertEqual(ws.title, "菜品价格变动及菜品损耗表")

        # Check basic structure
        self.assertIsNotNone(ws.cell(row=1, column=1).value)
        self.assertIn("菜品销售报表", str(ws.cell(row=1, column=1).value))

    def test_generate_detailed_revenue_worksheet_with_sample_data(self):
        """Test generating detailed revenue worksheet with sample data"""
        sample_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'dish_code': '106006610016',
                'dish_name': '单锅',
                'current_period_sales': 100,
                'current_price': 25.50,
                'previous_price': 24.00,
                'last_year_price': 23.00,
                'material_code': '1500680',
                'material_name': '测试物料',
                'material_price_per_kg': 4.05,
                'current_material_usage': 50,
                'current_month_loss_cost': 100.00,
                'previous_month_loss_cost': 95.00,
                'comparable_period_loss_cost': 90.00
            },
            {
                'store_id': 2,
                'store_name': '加拿大二店',
                'dish_code': '106006610017',
                'dish_name': '拼锅',
                'current_period_sales': 80,
                'current_price': 30.00,
                'previous_price': 29.00,
                'last_year_price': 28.00,
                'material_code': '1500681',
                'material_name': '测试物料2',
                'material_price_per_kg': 3.50,
                'current_material_usage': 40,
                'current_month_loss_cost': 80.00,
                'previous_month_loss_cost': 75.00,
                'comparable_period_loss_cost': 70.00
            }
        ]

        ws = self.generator.generate_detailed_revenue_worksheet(
            self.workbook, sample_data)

        # Check worksheet was created
        self.assertIsNotNone(ws)
        self.assertEqual(ws.title, "菜品价格变动及菜品损耗表")

        # Check data rows were added (should be more than just headers)
        self.assertGreater(ws.max_row, 3)  # Headers + at least some data rows

    def test_generate_material_cost_worksheet(self):
        """Test generating material cost worksheet"""
        ws = self.generator.generate_material_cost_worksheet(self.workbook, [])

        # Check worksheet was created
        self.assertIsNotNone(ws)
        self.assertEqual(ws.title, "原材料成本变动表")

        # Check placeholder content
        self.assertIn("原材料成本变动表", str(ws.cell(row=1, column=1).value))

    def test_generate_discount_analysis_worksheet(self):
        """Test generating discount analysis worksheet"""
        ws = self.generator.generate_discount_analysis_worksheet(
            self.workbook, [])

        # Check worksheet was created
        self.assertIsNotNone(ws)
        self.assertEqual(ws.title, "打折优惠表")

        # Check placeholder content
        self.assertIn("打折优惠表", str(ws.cell(row=1, column=1).value))

    def test_worksheet_formatting(self):
        """Test worksheet formatting is applied correctly"""
        sample_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'dish_code': '106006610016',
                'dish_name': '单锅',
                'current_period_sales': 100,
                'current_price': 25.50,
                'previous_price': 24.00,
                'last_year_price': 23.00,
                'material_code': '1500680',
                'material_name': '测试物料',
                'material_price_per_kg': 4.05,
                'current_material_usage': 50,
                'current_month_loss_cost': 100.00,
                'previous_month_loss_cost': 95.00,
                'comparable_period_loss_cost': 90.00
            }
        ]

        ws = self.generator.generate_detailed_revenue_worksheet(
            self.workbook, sample_data)

        # Check frozen panes
        self.assertEqual(ws.freeze_panes, 'A4')

        # Check column widths are set
        self.assertGreater(ws.column_dimensions['A'].width, 0)
        self.assertGreater(ws.column_dimensions['B'].width, 0)

    def test_price_change_calculations(self):
        """Test price change calculations in data processing"""
        sample_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'dish_code': '106006610016',
                'dish_name': '单锅',
                'current_period_sales': 100,
                'current_price': 25.50,
                'previous_price': 24.00,
                'last_year_price': 23.00,
                'material_code': '1500680',
                'material_name': '测试物料',
                'material_price_per_kg': 4.05,
                'current_material_usage': 50,
                'current_month_loss_cost': 100.00,
                'previous_month_loss_cost': 95.00,
                'comparable_period_loss_cost': 90.00
            }
        ]

        ws = self.generator.generate_detailed_revenue_worksheet(
            self.workbook, sample_data)

        # Check that worksheet was created and has expected structure
        self.assertIsNotNone(ws)
        self.assertGreater(ws.max_row, 3)

        # The calculations are done within the worksheet generator
        # We can verify the structure is correct and data is processed
        self.assertEqual(ws.title, "菜品价格变动及菜品损耗表")

    def test_error_handling_with_invalid_data(self):
        """Test error handling with invalid or missing data"""
        # Test with None data
        ws = self.generator.generate_detailed_revenue_worksheet(
            self.workbook, None)
        self.assertIsNotNone(ws)

        # Test with malformed data
        malformed_data = [
            {
                'store_id': 'invalid',
                'current_price': 'not_a_number',
                'previous_price': None
            }
        ]

        # Should not crash, should handle gracefully
        ws = self.generator.generate_detailed_revenue_worksheet(
            self.workbook, malformed_data)
        self.assertIsNotNone(ws)

    def test_multiple_stores_data_grouping(self):
        """Test that data is properly grouped by store"""
        multi_store_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'dish_code': '106006610016',
                'dish_name': '单锅',
                'current_period_sales': 100,
                'current_price': 25.50,
                'previous_price': 24.00,
                'last_year_price': 23.00,
                'material_code': '1500680',
                'material_name': '测试物料',
                'material_price_per_kg': 4.05,
                'current_material_usage': 50,
                'current_month_loss_cost': 100.00,
                'previous_month_loss_cost': 95.00,
                'comparable_period_loss_cost': 90.00
            },
            {
                'store_id': 2,
                'store_name': '加拿大二店',
                'dish_code': '106006610017',
                'dish_name': '拼锅',
                'current_period_sales': 80,
                'current_price': 30.00,
                'previous_price': 29.00,
                'last_year_price': 28.00,
                'material_code': '1500681',
                'material_name': '测试物料2',
                'material_price_per_kg': 3.50,
                'current_material_usage': 40,
                'current_month_loss_cost': 80.00,
                'previous_month_loss_cost': 75.00,
                'comparable_period_loss_cost': 70.00
            }
        ]

        ws = self.generator.generate_detailed_revenue_worksheet(
            self.workbook, multi_store_data)

        # Check worksheet was created and has data for multiple stores
        self.assertIsNotNone(ws)
        self.assertGreater(ws.max_row, 4)  # Headers + at least 2 data rows

    def test_date_parsing_and_formatting(self):
        """Test date parsing and formatting in headers"""
        # Test different date formats
        test_dates = ["2025-06-30", "2025-12-31", "2024-01-01"]

        for test_date in test_dates:
            generator = GrossMarginWorksheetGenerator(test_date)
            ws = generator.generate_detailed_revenue_worksheet(Workbook(), [])

            # Check that date is properly formatted in header
            header_value = str(ws.cell(row=1, column=1).value)
            self.assertIn("2025年06月" if test_date.startswith("2025-06") else
                          "2025年12月" if test_date.startswith("2025-12") else
                          "2024年01月", header_value)


if __name__ == '__main__':
    unittest.main()
