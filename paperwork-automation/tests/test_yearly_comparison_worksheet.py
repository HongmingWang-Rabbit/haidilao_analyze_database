#!/usr/bin/env python3
"""
Unit tests for yearly_comparison_worksheet.py
Tests the YearlyComparisonWorksheetGenerator class functionality.
"""

from lib.yearly_comparison_worksheet import YearlyComparisonWorksheetGenerator
import unittest
import sys
import os
from unittest.mock import MagicMock, patch
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Add parent directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))


class TestYearlyComparisonWorksheetGenerator(unittest.TestCase):
    """Test the yearly comparison worksheet generator"""

    def setUp(self):
        """Set up test fixtures"""
        self.store_names = {
            1: "加拿大一店", 2: "加拿大二店", 3: "加拿大三店", 4: "加拿大四店",
            5: "加拿大五店", 6: "加拿大六店", 7: "加拿大七店"
        }
        self.target_date = "2025-06-10"
        self.generator = YearlyComparisonWorksheetGenerator(
            self.store_names, self.target_date)

        # Sample data based on actual database structure
        self.sample_yearly_current = [
            {
                'store_id': 1,
                # Non-validated (for consumption)
                'total_tables': 150.0,
                # Validated (for table display)
                'total_tables_validated': 150.0,
                'total_revenue': 450000.0,  # In cents
                'avg_turnover_rate': 2.5,
                'avg_per_table': 3000.0
            },
            {
                'store_id': 2,
                # Non-validated (for consumption)
                'total_tables': 180.0,
                # Validated (for table display)
                'total_tables_validated': 180.0,
                'total_revenue': 540000.0,
                'avg_turnover_rate': 2.8,
                'avg_per_table': 3000.0
            },
            {
                'store_id': 3,
                # Non-validated (for consumption)
                'total_tables': 120.0,
                # Validated (for table display)
                'total_tables_validated': 120.0,
                'total_revenue': 360000.0,
                'avg_turnover_rate': 2.2,
                'avg_per_table': 3000.0
            }
        ]

        self.sample_yearly_previous = [
            {
                'store_id': 1,
                # Non-validated (for consumption)
                'total_tables': 140.0,
                # Validated (for table display)
                'total_tables_validated': 140.0,
                'total_revenue': 420000.0,
                'avg_turnover_rate': 2.3,
                'avg_per_table': 3000.0
            },
            {
                'store_id': 2,
                # Non-validated (for consumption)
                'total_tables': 160.0,
                # Validated (for table display)
                'total_tables_validated': 160.0,
                'total_revenue': 480000.0,
                'avg_turnover_rate': 2.6,
                'avg_per_table': 3000.0
            },
            {
                'store_id': 3,
                # Non-validated (for consumption)
                'total_tables': 110.0,
                # Validated (for table display)
                'total_tables_validated': 110.0,
                'total_revenue': 330000.0,
                'avg_turnover_rate': 2.0,
                'avg_per_table': 3000.0
            }
        ]

    def test_generator_initialization(self):
        """Test that the generator initializes correctly"""
        self.assertEqual(self.generator.store_names, self.store_names)
        self.assertEqual(self.generator.target_date, self.target_date)
        self.assertEqual(self.generator.current_year, 2025)
        self.assertEqual(self.generator.previous_year, 2024)
        self.assertEqual(self.generator.month, 6)
        self.assertEqual(self.generator.day, 10)

    def test_calculate_percentage_change_positive(self):
        """Test percentage change calculation for positive growth"""
        result = self.generator.calculate_percentage_change(120, 100)
        self.assertEqual(result, 20.0)

    def test_calculate_percentage_change_negative(self):
        """Test percentage change calculation for negative growth"""
        result = self.generator.calculate_percentage_change(80, 100)
        self.assertEqual(result, -20.0)

    def test_calculate_percentage_change_zero_previous(self):
        """Test percentage change when previous value is zero"""
        result = self.generator.calculate_percentage_change(100, 0)
        self.assertEqual(result, 0.0)

    def test_calculate_percentage_change_none_previous(self):
        """Test percentage change when previous value is None"""
        result = self.generator.calculate_percentage_change(100, None)
        self.assertEqual(result, 0.0)

    def test_format_percentage_change_positive(self):
        """Test formatting of positive percentage change"""
        result = self.generator.format_percentage_change(15.7)
        self.assertEqual(result, "15.7%")

    def test_format_percentage_change_negative(self):
        """Test formatting of negative percentage change"""
        result = self.generator.format_percentage_change(-8.3)
        self.assertEqual(result, "-8.3%")

    def test_format_percentage_change_zero(self):
        """Test formatting of zero percentage change"""
        result = self.generator.format_percentage_change(0.0)
        self.assertEqual(result, "0.0%")

    def test_generate_worksheet_with_valid_data(self):
        """Test worksheet generation with valid current and previous year data"""
        wb = Workbook()

        ws = self.generator.generate_worksheet(
            wb, self.sample_yearly_current, self.sample_yearly_previous)

        # Check that worksheet was created
        self.assertIsNotNone(ws)
        self.assertEqual(ws.title, "同比数据")

        # Check title
        title_cell = ws['A1']
        self.assertIn("加拿大-各门店2025年6月10日同比数据", title_cell.value)

        # Check headers structure
        self.assertEqual(ws['A2'].value, "分类")
        self.assertEqual(ws['C2'].value, "西部")
        self.assertEqual(ws['F2'].value, "东部")
        self.assertEqual(ws['J2'].value, "加拿大片区")

        # Check store name headers in row 3
        expected_headers = ["项目", "内容", "加拿大一店", "加拿大二店", "加拿大七店",
                            "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大片区"]
        for col, expected_header in enumerate(expected_headers, 1):
            actual_value = ws.cell(row=3, column=col).value
            self.assertEqual(actual_value, expected_header)

    def test_generate_worksheet_with_empty_data(self):
        """Test worksheet generation with empty data"""
        wb = Workbook()

        ws = self.generator.generate_worksheet(wb, [], [])

        # Should create worksheet with warning message
        self.assertIsNotNone(ws)
        self.assertEqual(ws.title, "同比数据")
        self.assertIn("年度对比数据不足", ws['A1'].value)

    def test_generate_worksheet_with_none_data(self):
        """Test worksheet generation with None data"""
        wb = Workbook()

        ws = self.generator.generate_worksheet(wb, None, None)

        # Should create worksheet with warning message
        self.assertIsNotNone(ws)
        self.assertEqual(ws.title, "同比数据")
        self.assertIn("年度对比数据不足", ws['A1'].value)

    def test_data_calculations(self):
        """Test that data calculations are performed correctly"""
        wb = Workbook()

        # Create worksheet to trigger data processing
        ws = self.generator.generate_worksheet(
            wb, self.sample_yearly_current, self.sample_yearly_previous)

        # Check that data rows are populated (桌数对比同期数据 section starts at row 4)
        # Row 4: "本月截止目前" - should have current year data
        store1_current_tables = ws.cell(row=4, column=3).value  # 加拿大一店
        self.assertEqual(store1_current_tables, 150.0)

        # Row 5: "去年截止同期" - should have previous year data
        store1_previous_tables = ws.cell(row=5, column=3).value
        self.assertEqual(store1_previous_tables, 140.0)

        # Row 6: "对比去年同期" - should have difference
        store1_difference = ws.cell(row=6, column=3).value
        self.assertEqual(store1_difference, 10.0)  # 150 - 140

        # Row 7: "桌数增长率" - should have percentage
        store1_growth_rate = ws.cell(row=7, column=3).value
        # (150-140)/140 * 100 ≈ 7.1%
        self.assertEqual(store1_growth_rate, "7.1%")

    def test_total_calculations(self):
        """Test that total calculations for 加拿大片区 are correct"""
        wb = Workbook()

        ws = self.generator.generate_worksheet(
            wb, self.sample_yearly_current, self.sample_yearly_previous)

        # Check total current tables (column J, row 4)
        total_current = ws.cell(row=4, column=10).value  # 加拿大片区
        expected_total_current = 150.0 + 180.0 + 120.0  # Sum of current tables
        self.assertEqual(total_current, expected_total_current)

        # Check total previous tables (row 5)
        total_previous = ws.cell(row=5, column=10).value
        expected_total_previous = 140.0 + 160.0 + 110.0  # Sum of previous tables
        self.assertEqual(total_previous, expected_total_previous)

        # Check total difference (row 6)
        total_difference = ws.cell(row=6, column=10).value
        expected_difference = expected_total_current - expected_total_previous
        self.assertEqual(total_difference, expected_difference)

    def test_revenue_calculations(self):
        """Test revenue calculations and conversions to 万加元"""
        wb = Workbook()

        ws = self.generator.generate_worksheet(
            wb, self.sample_yearly_current, self.sample_yearly_previous)

        # Revenue section starts at row 12
        # Row 12: "本月截止目前" revenue for store 1
        store1_current_revenue = ws.cell(row=12, column=3).value
        expected_revenue = 450000.0 / 10000  # Convert to 万加元
        self.assertEqual(store1_current_revenue, expected_revenue)

        # Check total revenue for 加拿大片区
        total_current_revenue = ws.cell(row=12, column=10).value
        expected_total_revenue = (450000.0 + 540000.0 + 360000.0) / 10000
        self.assertEqual(total_current_revenue, expected_total_revenue)

    def test_worksheet_formatting(self):
        """Test that Excel formatting is applied correctly"""
        wb = Workbook()

        ws = self.generator.generate_worksheet(
            wb, self.sample_yearly_current, self.sample_yearly_previous)

        # Check title formatting
        title_cell = ws['A1']
        self.assertTrue(title_cell.font.bold)
        self.assertEqual(title_cell.font.size, 12)
        self.assertEqual(title_cell.alignment.horizontal, 'center')
        self.assertIn('FFD700', title_cell.fill.start_color.rgb)

        # Check header formatting (row 2 and 3)
        header_cell = ws['A2']
        self.assertTrue(header_cell.font.bold)
        self.assertIn('FFD700', header_cell.fill.start_color.rgb)

        # Check data cell formatting - growth rate cells should be highlighted
        growth_rate_cell = ws.cell(row=7, column=3)  # 桌数增长率
        # Highlighted
        self.assertIn('FFFF00', growth_rate_cell.fill.start_color.rgb)

        # Check that borders are applied
        data_cell = ws.cell(row=4, column=3)
        self.assertIsNotNone(data_cell.border)
        self.assertEqual(data_cell.border.left.style, 'thin')

    def test_column_widths_and_row_heights(self):
        """Test that column widths and row heights are set correctly"""
        wb = Workbook()

        ws = self.generator.generate_worksheet(
            wb, self.sample_yearly_current, self.sample_yearly_previous)

        # Check column widths
        expected_widths = [15, 20, 12, 12, 12, 12, 12, 12, 12, 15]
        for i, expected_width in enumerate(expected_widths, 1):
            column_letter = chr(64 + i)  # A=65, B=66, etc.
            actual_width = ws.column_dimensions[column_letter].width
            self.assertEqual(actual_width, expected_width)

        # Check row heights
        self.assertEqual(ws.row_dimensions[1].height, 25)  # Title row
        self.assertEqual(ws.row_dimensions[2].height, 20)  # Header row
        self.assertEqual(ws.row_dimensions[3].height, 20)  # Store names row

    def test_merged_cells(self):
        """Test that cells are merged correctly"""
        wb = Workbook()

        ws = self.generator.generate_worksheet(
            wb, self.sample_yearly_current, self.sample_yearly_previous)

        # Check title merge (A1:J1)
        merged_ranges = [str(range_obj)
                         for range_obj in ws.merged_cells.ranges]
        self.assertIn('A1:J1', merged_ranges)

        # Check header merges
        self.assertIn('A2:B2', merged_ranges)  # 分类
        self.assertIn('C2:E2', merged_ranges)  # 西部
        self.assertIn('F2:I2', merged_ranges)  # 东部

        # Check category merges
        self.assertIn('A4:A7', merged_ranges)   # 桌数对比同期数据
        self.assertIn('A8:A11', merged_ranges)  # 翻台率对比同期数据
        self.assertIn('A12:A15', merged_ranges)  # 营业收入
        self.assertIn('A16:A19', merged_ranges)  # 单桌消费对比同期数据

    def test_negative_percentage_formatting(self):
        """Test that negative percentages are formatted in red"""
        # Create data with negative growth
        negative_data_current = [
            {
                'store_id': 1,
                'total_tables': 100.0,
                'total_tables_validated': 100.0,
                'total_revenue': 300000.0,
                'avg_turnover_rate': 2.0,
                'avg_per_table': 3000.0
            }
        ]

        negative_data_previous = [
            {
                'store_id': 1,
                'total_tables': 120.0,
                'total_tables_validated': 120.0,
                'total_revenue': 360000.0,
                'avg_turnover_rate': 2.4,
                'avg_per_table': 3000.0
            }
        ]

        wb = Workbook()
        ws = self.generator.generate_worksheet(
            wb, negative_data_current, negative_data_previous)

        # Check that negative percentage has red font
        growth_rate_cell = ws.cell(row=7, column=3)  # 桌数增长率
        self.assertTrue(growth_rate_cell.value.startswith('-'))
        self.assertIn('FF0000', growth_rate_cell.font.color.rgb)  # Red color

    def test_store_order_consistency(self):
        """Test that stores are ordered correctly (西部: 一店,二店,七店; 东部: 三店,四店,五店,六店)"""
        wb = Workbook()

        ws = self.generator.generate_worksheet(
            wb, self.sample_yearly_current, self.sample_yearly_previous)

        # Check store order in row 3
        expected_order = ["项目", "内容", "加拿大一店", "加拿大二店", "加拿大七店",
                          "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大片区"]

        for col, expected_store in enumerate(expected_order, 1):
            actual_store = ws.cell(row=3, column=col).value
            self.assertEqual(actual_store, expected_store)

    def test_data_sections_structure(self):
        """Test that all required data sections are present"""
        wb = Workbook()

        ws = self.generator.generate_worksheet(
            wb, self.sample_yearly_current, self.sample_yearly_previous)

        # Check section headers in column A
        expected_sections = [
            (4, "桌数\n对比同期数据"),
            (8, "翻台率\n对比同期数据"),
            (12, "营业收入\n(不含税-万加元)"),
            (16, "单桌消费\n对比同期数据")
        ]

        for row, expected_section in expected_sections:
            actual_section = ws.cell(row=row, column=1).value
            self.assertEqual(actual_section, expected_section)

        # Check content labels in column B for each section
        expected_contents = [
            (4, "本月截止目前"), (5, "去年截止同期"), (6, "对比去年同期"), (7, "桌数增长率"),
            (8, "本月截止目前"), (9, "去年截止同期"), (10, "对比去年同期"), (11, "翻台率增长率"),
            (12, "本月截止目前"), (13, "去年截止同期"), (14, "对比去年同期"), (15, "收入增长率"),
            (16, "本月截止目前"), (17, "去年截止同期"), (18, "对比去年同期"), (19, "单桌消费增长率")
        ]

        for row, expected_content in expected_contents:
            actual_content = ws.cell(row=row, column=2).value
            self.assertEqual(actual_content, expected_content)


class TestYearlyComparisonWorksheetIntegration(unittest.TestCase):
    """Integration tests for yearly comparison worksheet with realistic data"""

    def setUp(self):
        """Set up test fixtures with realistic data volumes"""
        self.store_names = {
            1: "加拿大一店", 2: "加拿大二店", 3: "加拿大三店", 4: "加拿大四店",
            5: "加拿大五店", 6: "加拿大六店", 7: "加拿大七店"
        }
        self.target_date = "2025-06-10"
        self.generator = YearlyComparisonWorksheetGenerator(
            self.store_names, self.target_date)

    def test_realistic_data_volumes(self):
        """Test with realistic data volumes matching actual restaurant data"""
        # Realistic current year data (month-to-date for June)
        realistic_current = [
            {'store_id': 1, 'total_tables': 280.5, 'total_tables_validated': 280.5, 'total_revenue': 842000.0,
                'avg_turnover_rate': 2.8, 'avg_per_table': 3007.0},
            {'store_id': 2, 'total_tables': 310.0, 'total_tables_validated': 310.0, 'total_revenue': 925000.0,
                'avg_turnover_rate': 3.1, 'avg_per_table': 2984.0},
            {'store_id': 3, 'total_tables': 245.5, 'total_tables_validated': 245.5, 'total_revenue': 736500.0,
                'avg_turnover_rate': 2.5, 'avg_per_table': 3000.0},
            {'store_id': 4, 'total_tables': 290.0, 'total_tables_validated': 290.0, 'total_revenue': 870000.0,
                'avg_turnover_rate': 2.9, 'avg_per_table': 3000.0},
            {'store_id': 5, 'total_tables': 265.0, 'total_tables_validated': 265.0, 'total_revenue': 795000.0,
                'avg_turnover_rate': 2.7, 'avg_per_table': 3000.0},
            {'store_id': 6, 'total_tables': 300.5, 'total_tables_validated': 300.5, 'total_revenue': 901500.0,
                'avg_turnover_rate': 3.0, 'avg_per_table': 3000.0},
            {'store_id': 7, 'total_tables': 220.0, 'total_tables_validated': 220.0, 'total_revenue': 660000.0,
                'avg_turnover_rate': 2.2, 'avg_per_table': 3000.0}
        ]

        # Previous year data (slightly lower due to business growth)
        realistic_previous = [
            {'store_id': 1, 'total_tables': 260.0, 'total_tables_validated': 260.0, 'total_revenue': 780000.0,
                'avg_turnover_rate': 2.6, 'avg_per_table': 3000.0},
            {'store_id': 2, 'total_tables': 290.0, 'total_tables_validated': 290.0, 'total_revenue': 870000.0,
                'avg_turnover_rate': 2.9, 'avg_per_table': 3000.0},
            {'store_id': 3, 'total_tables': 230.0, 'total_tables_validated': 230.0, 'total_revenue': 690000.0,
                'avg_turnover_rate': 2.3, 'avg_per_table': 3000.0},
            {'store_id': 4, 'total_tables': 270.0, 'total_tables_validated': 270.0, 'total_revenue': 810000.0,
                'avg_turnover_rate': 2.7, 'avg_per_table': 3000.0},
            {'store_id': 5, 'total_tables': 250.0, 'total_tables_validated': 250.0, 'total_revenue': 750000.0,
                'avg_turnover_rate': 2.5, 'avg_per_table': 3000.0},
            {'store_id': 6, 'total_tables': 280.0, 'total_tables_validated': 280.0, 'total_revenue': 840000.0,
                'avg_turnover_rate': 2.8, 'avg_per_table': 3000.0},
            {'store_id': 7, 'total_tables': 200.0, 'total_tables_validated': 200.0, 'total_revenue': 600000.0,
                'avg_turnover_rate': 2.0, 'avg_per_table': 3000.0}
        ]

        wb = Workbook()
        ws = self.generator.generate_worksheet(
            wb, realistic_current, realistic_previous)

        # Verify the worksheet was created successfully
        self.assertIsNotNone(ws)
        self.assertEqual(ws.title, "同比数据")

        # Check that all stores have data populated
        for store_col in range(3, 10):  # Columns C through I (stores 1-7)
            current_tables_cell = ws.cell(row=4, column=store_col)
            self.assertIsNotNone(current_tables_cell.value)
            self.assertGreater(current_tables_cell.value, 0)

        # Check total calculations are reasonable
        total_current_tables = ws.cell(row=4, column=10).value  # 加拿大片区
        expected_total = sum(store['total_tables_validated']
                             for store in realistic_current)
        self.assertEqual(total_current_tables, expected_total)

        # Check revenue conversion to 万加元
        store1_revenue = ws.cell(row=12, column=3).value
        expected_revenue = 842000.0 / 10000  # Convert to 万加元
        self.assertEqual(store1_revenue, expected_revenue)


if __name__ == '__main__':
    unittest.main(verbosity=2)
