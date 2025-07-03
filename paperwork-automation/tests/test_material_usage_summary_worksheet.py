#!/usr/bin/env python3
"""
Comprehensive tests for MaterialUsageSummaryGenerator
Tests material usage summary worksheet generation functionality.
"""

from lib.material_usage_summary_worksheet import MaterialUsageSummaryGenerator
import unittest
from datetime import datetime
from unittest.mock import Mock, MagicMock, patch
from openpyxl import Workbook
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))


class TestMaterialUsageSummaryGenerator(unittest.TestCase):
    """Comprehensive tests for MaterialUsageSummaryGenerator."""

    def setUp(self):
        """Set up test fixtures."""
        self.mock_data_provider = Mock()
        self.mock_db_manager = Mock()
        self.mock_data_provider.db_manager = self.mock_db_manager

        self.generator = MaterialUsageSummaryGenerator(self.mock_data_provider)
        self.target_date = "2025-06-01"

    def test_generator_initialization(self):
        """Test generator initialization."""
        self.assertEqual(self.generator.data_provider, self.mock_data_provider)
        self.assertIsNotNone(self.generator.logger)

    def test_worksheet_creation_success(self):
        """Test successful worksheet creation."""
        # Mock data
        mock_usage_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'material_type_id': 1,
                'material_type_name': '成本-锅底类',
                'usage_amount': 25016.36
            },
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'material_type_id': 2,
                'material_type_name': '成本-荤菜类',
                'usage_amount': 65793.20
            },
            {
                'store_id': 2,
                'store_name': '加拿大二店',
                'material_type_id': 1,
                'material_type_name': '成本-锅底类',
                'usage_amount': 15000.00
            }
        ]

        mock_store_data = {
            1: {'id': 1, 'name': '加拿大一店', 'district_id': 1, 'is_active': True},
            2: {'id': 2, 'name': '加拿大二店', 'district_id': 1, 'is_active': True}
        }

        mock_material_type_data = {
            1: {'id': 1, 'name': '成本-锅底类', 'sort_order': 1},
            2: {'id': 2, 'name': '成本-荤菜类', 'sort_order': 2}
        }

        # Mock the database calls
        self.mock_db_manager.execute_query.side_effect = [
            mock_usage_data,  # First call for usage data
            mock_store_data.values(),  # Second call for store data
            mock_material_type_data.values()  # Third call for material type data
        ]

        # Create workbook and run generator
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        self.generator.generate_worksheet(wb, self.target_date)

        # Verify worksheet was created
        self.assertEqual(len(wb.worksheets), 1)
        ws = wb.worksheets[0]
        self.assertEqual(ws.title, "物料使用汇总")

        # Verify header content
        self.assertEqual(ws['A1'].value, "海底捞物料使用汇总报表 - 2025-06")
        self.assertTrue(ws['A2'].value.startswith("生成时间:"))

    def test_worksheet_creation_no_data(self):
        """Test worksheet creation with no usage data."""
        # Mock empty data
        self.mock_db_manager.execute_query.side_effect = [
            [],  # Empty usage data
            [],  # Empty store data
            []   # Empty material type data
        ]

        wb = Workbook()
        wb.remove(wb.active)

        self.generator.generate_worksheet(wb, self.target_date)

        # Verify worksheet was created with header only
        self.assertEqual(len(wb.worksheets), 1)
        ws = wb.worksheets[0]
        self.assertEqual(ws.title, "物料使用汇总")
        self.assertEqual(ws['A1'].value, "海底捞物料使用汇总报表 - 2025-06")

    def test_get_material_usage_data_success(self):
        """Test successful material usage data retrieval."""
        expected_data = [
            {'store_id': 1, 'store_name': '加拿大一店', 'material_type_id': 1,
             'material_type_name': '成本-锅底类', 'usage_amount': 25016.36}
        ]

        self.mock_db_manager.execute_query.return_value = expected_data

        result = self.generator._get_material_usage_data(2025, 6)

        self.assertEqual(result, expected_data)
        self.mock_db_manager.execute_query.assert_called_once()

        # Verify the query parameters
        call_args = self.mock_db_manager.execute_query.call_args
        self.assertEqual(call_args[0][1], (2025, 6))

    def test_get_material_usage_data_error(self):
        """Test material usage data retrieval with database error."""
        self.mock_db_manager.execute_query.side_effect = Exception(
            "Database error")

        result = self.generator._get_material_usage_data(2025, 6)

        self.assertEqual(result, [])

    def test_get_store_data_success(self):
        """Test successful store data retrieval."""
        expected_stores = [
            {'id': 1, 'name': '加拿大一店', 'district_id': 1, 'is_active': True},
            {'id': 2, 'name': '加拿大二店', 'district_id': 1, 'is_active': True}
        ]

        self.mock_db_manager.execute_query.return_value = expected_stores

        result = self.generator._get_store_data()

        expected_dict = {1: expected_stores[0], 2: expected_stores[1]}
        self.assertEqual(result, expected_dict)

    def test_get_store_data_error(self):
        """Test store data retrieval with database error."""
        self.mock_db_manager.execute_query.side_effect = Exception(
            "Database error")

        result = self.generator._get_store_data()

        self.assertEqual(result, {})

    def test_get_material_type_data_success(self):
        """Test successful material type data retrieval."""
        expected_types = [
            {'id': 1, 'name': '成本-锅底类', 'sort_order': 1},
            {'id': 2, 'name': '成本-荤菜类', 'sort_order': 2}
        ]

        self.mock_db_manager.execute_query.return_value = expected_types

        result = self.generator._get_material_type_data()

        expected_dict = {1: expected_types[0], 2: expected_types[1]}
        self.assertEqual(result, expected_dict)

    def test_get_material_type_data_error(self):
        """Test material type data retrieval with database error."""
        self.mock_db_manager.execute_query.side_effect = Exception(
            "Database error")

        result = self.generator._get_material_type_data()

        self.assertEqual(result, {})

    def test_create_header(self):
        """Test header creation."""
        wb = Workbook()
        ws = wb.active

        self.generator._create_header(ws, "2025-06")

        # Verify header content
        self.assertEqual(ws['A1'].value, "海底捞物料使用汇总报表 - 2025-06")
        self.assertTrue(ws['A2'].value.startswith("生成时间:"))

        # Verify formatting
        self.assertEqual(ws['A1'].font.size, 16)
        self.assertTrue(ws['A1'].font.bold)

    def test_create_summary_table_single_store(self):
        """Test summary table creation with single store."""
        wb = Workbook()
        ws = wb.active

        usage_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'material_type_id': 1,
                'material_type_name': '成本-锅底类',
                'usage_amount': 25016.36
            },
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'material_type_id': 2,
                'material_type_name': '成本-荤菜类',
                'usage_amount': 65793.20
            }
        ]

        store_data = {
            1: {'id': 1, 'name': '加拿大一店', 'district_id': 1, 'is_active': True}
        }

        material_type_data = {
            1: {'id': 1, 'name': '成本-锅底类', 'sort_order': 1},
            2: {'id': 2, 'name': '成本-荤菜类', 'sort_order': 2}
        }

        self.generator._create_summary_table(
            ws, usage_data, store_data, material_type_data)

        # Verify store header is created
        self.assertEqual(ws['A4'].value, "门店: 加拿大一店")

        # Verify table headers
        self.assertEqual(ws['A5'].value, "物料分类")
        self.assertEqual(ws['B5'].value, "本月使用金额 (CAD)")

    def test_create_summary_table_multiple_stores(self):
        """Test summary table creation with multiple stores."""
        wb = Workbook()
        ws = wb.active

        usage_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'material_type_id': 1,
                'material_type_name': '成本-锅底类',
                'usage_amount': 25016.36
            },
            {
                'store_id': 2,
                'store_name': '加拿大二店',
                'material_type_id': 1,
                'material_type_name': '成本-锅底类',
                'usage_amount': 15000.00
            }
        ]

        store_data = {
            1: {'id': 1, 'name': '加拿大一店', 'district_id': 1, 'is_active': True},
            2: {'id': 2, 'name': '加拿大二店', 'district_id': 1, 'is_active': True}
        }

        material_type_data = {
            1: {'id': 1, 'name': '成本-锅底类', 'sort_order': 1}
        }

        self.generator._create_summary_table(
            ws, usage_data, store_data, material_type_data)

        # Check that both stores are included
        store_headers = []
        for row in range(1, ws.max_row + 1):
            cell_value = ws[f'A{row}'].value
            if cell_value and cell_value.startswith("门店:"):
                store_headers.append(cell_value)

        self.assertIn("门店: 加拿大一店", store_headers)
        self.assertIn("门店: 加拿大二店", store_headers)

    def test_apply_formatting(self):
        """Test worksheet formatting application."""
        wb = Workbook()
        ws = wb.active

        # Add some test data
        ws['A1'] = "Test"
        ws['B1'] = "Test"

        self.generator._apply_formatting(ws)

        # Verify column widths
        self.assertEqual(ws.column_dimensions['A'].width, 25)
        self.assertEqual(ws.column_dimensions['B'].width, 20)

        # Verify row height
        self.assertEqual(ws.row_dimensions[1].height, 20)

    def test_generate_worksheet_database_error(self):
        """Test worksheet generation with database error."""
        self.mock_db_manager.execute_query.side_effect = Exception(
            "Database connection failed")

        wb = Workbook()
        wb.remove(wb.active)

        with self.assertRaises(Exception):
            self.generator.generate_worksheet(wb, self.target_date)

    def test_date_parsing(self):
        """Test proper date parsing for year and month extraction."""
        test_dates = [
            ("2025-06-01", 2025, 6),
            ("2024-12-15", 2024, 12),
            ("2023-01-31", 2023, 1)
        ]

        for date_str, expected_year, expected_month in test_dates:
            # Mock the usage data call to verify parameters
            self.mock_db_manager.execute_query.return_value = []

            wb = Workbook()
            wb.remove(wb.active)

            self.generator.generate_worksheet(wb, date_str)

            # Verify the first call was made with correct year/month
            first_call_args = self.mock_db_manager.execute_query.call_args_list[0]
            self.assertEqual(
                first_call_args[0][1], (expected_year, expected_month))

            # Reset for next iteration
            self.mock_db_manager.reset_mock()

    def test_empty_usage_amount_handling(self):
        """Test handling of null/empty usage amounts."""
        usage_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'material_type_id': 1,
                'material_type_name': '成本-锅底类',
                'usage_amount': None  # Null amount
            },
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'material_type_id': 2,
                'material_type_name': '成本-荤菜类',
                'usage_amount': 0  # Zero amount
            }
        ]

        store_data = {1: {'id': 1, 'name': '加拿大一店',
                          'district_id': 1, 'is_active': True}}
        material_type_data = {
            1: {'id': 1, 'name': '成本-锅底类', 'sort_order': 1},
            2: {'id': 2, 'name': '成本-荤菜类', 'sort_order': 2}
        }

        wb = Workbook()
        ws = wb.active

        # Should not raise an exception
        self.generator._create_summary_table(
            ws, usage_data, store_data, material_type_data)

        # Verify that the worksheet was created (even with zero/null amounts)
        self.assertEqual(ws['A4'].value, "门店: 加拿大一店")


if __name__ == '__main__':
    unittest.main()
