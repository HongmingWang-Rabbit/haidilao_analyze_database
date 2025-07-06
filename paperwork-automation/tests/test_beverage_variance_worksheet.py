#!/usr/bin/env python3
"""
Test suite for BeverageVarianceGenerator
"""

from lib.beverage_variance_worksheet import BeverageVarianceGenerator
import unittest
import tempfile
import os
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))


class TestBeverageVarianceGenerator(unittest.TestCase):
    """Test cases for BeverageVarianceGenerator"""

    def setUp(self):
        """Set up test fixtures"""
        self.mock_data_provider = Mock()
        self.mock_db_manager = Mock()
        self.mock_data_provider.db_manager = self.mock_db_manager
        self.generator = BeverageVarianceGenerator(self.mock_data_provider)

    def test_generator_initialization(self):
        """Test generator initialization"""
        self.assertIsInstance(self.generator, BeverageVarianceGenerator)
        self.assertEqual(self.generator.data_provider, self.mock_data_provider)
        self.assertEqual(self.generator.db_manager, self.mock_db_manager)

    def test_generate_worksheet_basic(self):
        """Test basic worksheet generation with new structure"""
        workbook = Workbook()
        target_date = "2025-06-01"

        # Mock database response matching new structure
        mock_variance_data = [
            {
                'material_id': 1,
                'store_id': 1,
                'store_name': '加拿大一店',
                'material_name': '青岛啤酒',
                'material_number': 'MAT001',
                'material_unit': '瓶',
                'package_spec': '500ml',
                'theoretical_usage': 95.0,
                'combo_usage': 5.0,
                'system_record': 110.0,
                'inventory_count': 2.0,
                'variance_amount': 8.0,  # 110 - (95 + 5 + 2)
                'variance_percent': 8.0,
                'variance_status': '超量',
                'material_price': 5.0,
                'sale_price': 7.5
            }
        ]

        with patch.object(self.generator, 'get_beverage_variance_data', return_value=mock_variance_data):
            worksheet = self.generator.generate_worksheet(
                workbook, target_date)

            # Verify worksheet was created
            self.assertIsNotNone(worksheet)
            self.assertEqual(worksheet.title, "酒水差异明细表")

            # Verify title
            self.assertEqual(worksheet['A1'].value, "酒水差异明细表 - 2025年06月")

            # Find header row (should be after title and summary)
            header_row = None
            for row in range(1, 20):
                if worksheet.cell(row=row, column=1).value == "序号":
                    header_row = row
                    break

            self.assertIsNotNone(header_row, "Header row not found")

            # Verify headers match new structure
            expected_headers = [
                "序号", "门店", "物料名称", "物料号", "单位", "包装规格",
                "理论用量", "套餐用量", "系统记录", "库存盘点", "差异数量",
                "差异率(%)", "状态", "物料单价", "销售单价", "差异金额"
            ]

            for col, expected_header in enumerate(expected_headers, 1):
                actual_header = worksheet.cell(
                    row=header_row, column=col).value
                self.assertEqual(actual_header, expected_header)

            # Verify data row
            data_row = header_row + 1
            self.assertEqual(worksheet.cell(
                row=data_row, column=1).value, 1)  # 序号
            self.assertEqual(worksheet.cell(
                row=data_row, column=2).value, "加拿大一店")  # 门店
            self.assertEqual(worksheet.cell(
                row=data_row, column=3).value, "青岛啤酒")  # 物料名称
            self.assertEqual(worksheet.cell(
                row=data_row, column=4).value, "MAT001")  # 物料号
            self.assertEqual(worksheet.cell(
                row=data_row, column=14).value, 5.0)  # 物料单价
            self.assertEqual(worksheet.cell(
                row=data_row, column=15).value, 7.5)  # 销售单价

    def test_generate_worksheet_empty_data(self):
        """Test worksheet generation with empty data"""
        workbook = Workbook()
        target_date = "2025-06-01"

        with patch.object(self.generator, 'get_beverage_variance_data', return_value=[]):
            worksheet = self.generator.generate_worksheet(
                workbook, target_date)

            # Verify worksheet was created with headers only
            self.assertIsNotNone(worksheet)
            self.assertEqual(worksheet.title, "酒水差异明细表")

            # Should have title
            self.assertEqual(worksheet['A1'].value, "酒水差异明细表 - 2025年06月")

    def test_generate_worksheet_with_summary(self):
        """Test worksheet generation with summary calculation"""
        workbook = Workbook()
        target_date = "2025-06-01"

        mock_variance_data = [
            {
                'material_id': 1,
                'store_id': 1,
                'store_name': '加拿大一店',
                'material_name': '青岛啤酒',
                'material_number': 'MAT001',
                'material_unit': '瓶',
                'package_spec': '500ml',
                'theoretical_usage': 95.0,
                'combo_usage': 5.0,
                'system_record': 110.0,
                'inventory_count': 2.0,
                'variance_amount': 8.0,
                'variance_percent': 8.0,
                'variance_status': '超量',
                'material_price': 5.0,
                'sale_price': 7.5
            },
            {
                'material_id': 2,
                'store_id': 2,
                'store_name': '加拿大二店',
                'material_name': '可口可乐',
                'material_number': 'MAT002',
                'material_unit': '罐',
                'package_spec': '330ml',
                'theoretical_usage': 200.0,
                'combo_usage': 10.0,
                'system_record': 205.0,
                'inventory_count': 3.0,
                'variance_amount': -8.0,  # 205 - (200 + 10 + 3)
                'variance_percent': 3.8,
                'variance_status': '正常',
                'material_price': 3.0,
                'sale_price': 4.5
            }
        ]

        with patch.object(self.generator, 'get_beverage_variance_data', return_value=mock_variance_data):
            worksheet = self.generator.generate_worksheet(
                workbook, target_date)

            # Find summary section - should contain "酒水差异分析概览"
            summary_found = False
            for row in range(1, 20):
                if worksheet.cell(row=row, column=1).value == "酒水差异分析概览":
                    summary_found = True
                    break

            self.assertTrue(summary_found, "Summary section not found")

    def test_get_beverage_variance_data_success(self):
        """Test successful beverage variance data retrieval"""
        year, month = 2025, 6

        # Mock database connection and cursor
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_connection.cursor.return_value = mock_cursor
        mock_connection.__enter__.return_value = mock_connection
        mock_connection.__exit__.return_value = None

        self.mock_db_manager.get_connection.return_value = mock_connection

        # Mock query results matching new structure
        mock_cursor.fetchall.return_value = [
            {
                'material_id': 1,
                'store_id': 1,
                'store_name': '加拿大一店',
                'material_name': '青岛啤酒',
                'material_number': 'MAT001',
                'material_unit': '瓶',
                'package_spec': '500ml',
                'theoretical_usage': 95.0,
                'combo_usage': 5.0,
                'system_record': 110.0,
                'inventory_count': 2.0,
                'material_price': 5.0,
                'sale_price': 7.5
            }
        ]

        result = self.generator.get_beverage_variance_data(year, month)

        # Verify result
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]['store_name'], '加拿大一店')
        self.assertEqual(result[0]['material_name'], '青岛啤酒')
        self.assertEqual(result[0]['theoretical_usage'], 95.0)
        self.assertEqual(result[0]['combo_usage'], 5.0)
        self.assertEqual(result[0]['material_price'], 5.0)
        self.assertEqual(result[0]['sale_price'], 7.5)

    def test_get_beverage_variance_data_database_error(self):
        """Test beverage variance data retrieval with database error"""
        year, month = 2025, 6

        # Mock database connection to raise exception
        self.mock_db_manager.get_connection.side_effect = Exception(
            "Database connection failed")

        result = self.generator.get_beverage_variance_data(year, month)

        # Should return empty list on error
        self.assertEqual(result, [])

    def test_get_beverage_variance_data_empty_results(self):
        """Test beverage variance data retrieval with empty results"""
        year, month = 2025, 6

        # Mock database connection and cursor
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_connection.cursor.return_value = mock_cursor
        mock_connection.__enter__.return_value = mock_connection
        mock_connection.__exit__.return_value = None

        self.mock_db_manager.get_connection.return_value = mock_connection

        # Mock empty results
        mock_cursor.fetchall.return_value = []

        result = self.generator.get_beverage_variance_data(year, month)

        # Should return empty list
        self.assertEqual(result, [])

    def test_add_variance_summary_section(self):
        """Test summary section generation"""
        from openpyxl import Workbook
        ws = Workbook().active

        mock_variance_data = [
            {
                'theoretical_usage': 95.0,
                'combo_usage': 5.0,
                'system_record': 110.0,
                'inventory_count': 2.0,
                'variance_amount': 8.0,
                'variance_percent': 8.0,
                'material_price': 5.0,
                'sale_price': 7.5
            },
            {
                'theoretical_usage': 200.0,
                'combo_usage': 10.0,
                'system_record': 205.0,
                'inventory_count': 3.0,
                'variance_amount': -8.0,
                'variance_percent': 3.8,
                'material_price': 3.0,
                'sale_price': 4.5
            }
        ]

        result_row = self.generator.add_variance_summary_section(
            ws, 1, mock_variance_data)

        # Verify summary was created
        self.assertGreater(result_row, 1)
        self.assertEqual(ws['A1'].value, "酒水差异分析概览")

    def test_worksheet_styling(self):
        """Test worksheet styling and formatting"""
        workbook = Workbook()
        target_date = "2025-06-01"

        mock_variance_data = [
            {
                'material_id': 1,
                'store_id': 1,
                'store_name': '加拿大一店',
                'material_name': '青岛啤酒',
                'material_number': 'MAT001',
                'material_unit': '瓶',
                'package_spec': '500ml',
                'theoretical_usage': 95.0,
                'combo_usage': 5.0,
                'system_record': 110.0,
                'inventory_count': 2.0,
                'variance_amount': 8.0,
                'variance_percent': 8.0,  # > 5%, should be highlighted
                'variance_status': '超量',
                'material_price': 5.0,
                'sale_price': 7.5
            }
        ]

        with patch.object(self.generator, 'get_beverage_variance_data', return_value=mock_variance_data):
            worksheet = self.generator.generate_worksheet(
                workbook, target_date)

            # Verify column widths are set
            self.assertIsNotNone(worksheet.column_dimensions['A'].width)
            self.assertIsNotNone(worksheet.column_dimensions['P'].width)

            # Verify title styling
            title_cell = worksheet['A1']
            self.assertIsNotNone(title_cell.font)
            self.assertIsNotNone(title_cell.fill)

    def test_date_parsing(self):
        """Test date parsing for different formats"""
        workbook = Workbook()
        target_date = "2025-12-31"

        with patch.object(self.generator, 'get_beverage_variance_data', return_value=[]):
            worksheet = self.generator.generate_worksheet(
                workbook, target_date)

            # Verify date formatting in title
            expected_title = "酒水差异明细表 - 2025年12月"
            self.assertEqual(worksheet['A1'].value, expected_title)

    def test_variance_calculation_formulas(self):
        """Test Excel formulas for variance calculations"""
        workbook = Workbook()
        target_date = "2025-06-01"

        mock_variance_data = [
            {
                'material_id': 1,
                'store_id': 1,
                'store_name': '加拿大一店',
                'material_name': '青岛啤酒',
                'material_number': 'MAT001',
                'material_unit': '瓶',
                'package_spec': '500ml',
                'theoretical_usage': 95.0,
                'combo_usage': 5.0,
                'system_record': 110.0,
                'inventory_count': 2.0,
                'variance_amount': 8.0,
                'variance_percent': 8.0,
                'variance_status': '超量',
                'material_price': 5.0,
                'sale_price': 7.5
            }
        ]

        with patch.object(self.generator, 'get_beverage_variance_data', return_value=mock_variance_data):
            worksheet = self.generator.generate_worksheet(
                workbook, target_date)

            # Find data row
            data_row = None
            for row in range(1, 30):
                if worksheet.cell(row=row, column=1).value == 1:  # First serial number
                    data_row = row
                    break

            self.assertIsNotNone(data_row, "Data row not found")

            # Check variance amount formula (Column K)
            variance_formula = worksheet.cell(row=data_row, column=11).value
            expected_variance_formula = f"=I{data_row}-(G{data_row}+H{data_row}+J{data_row})"
            self.assertEqual(variance_formula, expected_variance_formula)

            # Check variance rate formula (Column L)
            rate_formula = worksheet.cell(row=data_row, column=12).value
            expected_rate_formula = f"=IF((G{data_row}+H{data_row})=0,IF(K{data_row}=0,0,100),ABS(K{data_row}/(G{data_row}+H{data_row}))*100)"
            self.assertEqual(rate_formula, expected_rate_formula)

            # Check net difference formula (Column P)
            net_diff_formula = worksheet.cell(row=data_row, column=16).value
            expected_net_diff_formula = f"=K{data_row}*N{data_row}"
            self.assertEqual(net_diff_formula, expected_net_diff_formula)

    def test_beverage_filtering_logic(self):
        """Test that only beverage materials are included"""
        year, month = 2025, 6

        # Mock database connection and cursor
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_connection.cursor.return_value = mock_cursor
        mock_connection.__enter__.return_value = mock_connection
        mock_connection.__exit__.return_value = None

        self.mock_db_manager.get_connection.return_value = mock_connection

        # Mock empty results (no beverage materials)
        mock_cursor.fetchall.return_value = []

        result = self.generator.get_beverage_variance_data(year, month)

        # Should return empty list when no beverage materials
        self.assertEqual(result, [])

        # Verify that query was executed
        mock_cursor.execute.assert_called_once()

        # Check that the query includes beverage filtering
        executed_query = mock_cursor.execute.call_args[0][0]
        self.assertIn("成本-酒水类", executed_query)

    def test_financial_impact_calculation(self):
        """Test financial impact calculations in summary"""
        from openpyxl import Workbook
        ws = Workbook().active

        mock_variance_data = [
            {
                'theoretical_usage': 95.0,
                'combo_usage': 5.0,
                'system_record': 110.0,
                'inventory_count': 2.0,
                'variance_amount': 8.0,
                'variance_percent': 8.0,
                'material_price': 5.0,
                'sale_price': 7.5
            }
        ]

        self.generator.add_variance_summary_section(ws, 1, mock_variance_data)

        # Check that financial impact is calculated
        # Material cost impact: 8.0 * 5.0 = 40.0
        # Sale impact: 8.0 * 7.5 = 60.0
        summary_found = False
        for row in range(1, 20):
            for col in range(1, 10):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and "材料成本影响" in str(cell_value):
                    summary_found = True
                    break

        self.assertTrue(summary_found, "Financial impact summary not found")


if __name__ == '__main__':
    unittest.main()
