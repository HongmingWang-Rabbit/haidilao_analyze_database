#!/usr/bin/env python3
"""
Test suite for BeverageSummaryGenerator
"""

from lib.beverage_summary_worksheet import BeverageSummaryGenerator
import unittest
import tempfile
import os
from unittest.mock import Mock, patch, MagicMock
from openpyxl import Workbook
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))


class TestBeverageSummaryGenerator(unittest.TestCase):
    """Test cases for BeverageSummaryGenerator"""

    def setUp(self):
        """Set up test fixtures"""
        self.mock_data_provider = Mock()
        self.mock_db_manager = Mock()
        self.mock_data_provider.db_manager = self.mock_db_manager
        self.generator = BeverageSummaryGenerator(self.mock_data_provider)

    def test_generator_initialization(self):
        """Test generator initialization"""
        self.assertIsInstance(self.generator, BeverageSummaryGenerator)
        self.assertEqual(self.generator.data_provider, self.mock_data_provider)
        self.assertEqual(self.generator.db_manager, self.mock_db_manager)

    def test_generate_worksheet_basic(self):
        """Test basic worksheet generation"""
        workbook = Workbook()
        target_date = "2025-06-01"

        # Mock database response
        mock_variance_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'store_manager': '张三',
                'regional_manager': '蒋冰洁',
                'country': '加拿大',
                'variance_quantity': '10',
                'impact_amount_local': '100',
                'impact_amount_usd': '74'
            }
        ]

        with patch.object(self.generator, 'get_beverage_variance_data', return_value=mock_variance_data):
            worksheet = self.generator.generate_worksheet(
                workbook, target_date)

            # Verify worksheet was created
            self.assertIsNotNone(worksheet)
            self.assertEqual(worksheet.title, "酒水汇总表")

            # Verify headers
            expected_headers = [
                "月份", "大区经理", "国家", "门店名称", "店经理",
                "差异数量（绝对值）", "影响金额（本币）", "影响金额（美元）"
            ]

            for col, expected_header in enumerate(expected_headers, 1):
                actual_header = worksheet.cell(row=1, column=col).value
                self.assertEqual(actual_header, expected_header)

            # Verify data row
            self.assertEqual(worksheet.cell(row=2, column=1).value, "202506")
            self.assertEqual(worksheet.cell(row=2, column=4).value, "加拿大一店")

    def test_generate_worksheet_empty_data(self):
        """Test worksheet generation with empty data"""
        workbook = Workbook()
        target_date = "2025-06-01"

        with patch.object(self.generator, 'get_beverage_variance_data', return_value=[]):
            worksheet = self.generator.generate_worksheet(
                workbook, target_date)

            # Verify worksheet was created with headers only
            self.assertIsNotNone(worksheet)
            self.assertEqual(worksheet.title, "酒水汇总表")

            # Should have headers but no data rows
            self.assertEqual(worksheet.cell(row=1, column=1).value, "月份")
            self.assertIsNone(worksheet.cell(row=2, column=1).value)

    def test_generate_worksheet_with_totals(self):
        """Test worksheet generation with totals calculation"""
        workbook = Workbook()
        target_date = "2025-06-01"

        mock_variance_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'store_manager': '张三',
                'regional_manager': '蒋冰洁',
                'country': '加拿大',
                'variance_quantity': '10',
                'impact_amount_local': '100',
                'impact_amount_usd': '74'
            },
            {
                'store_id': 2,
                'store_name': '加拿大二店',
                'store_manager': '李四',
                'regional_manager': '蒋冰洁',
                'country': '加拿大',
                'variance_quantity': '5',
                'impact_amount_local': '50',
                'impact_amount_usd': '37'
            }
        ]

        with patch.object(self.generator, 'get_beverage_variance_data', return_value=mock_variance_data):
            worksheet = self.generator.generate_worksheet(
                workbook, target_date)

            # Check totals row (should be row 4: header + 2 data rows + 1 total row)
            self.assertEqual(worksheet.cell(row=4, column=5).value, "合计")
            self.assertEqual(worksheet.cell(
                row=4, column=6).value, 15.0)  # 10 + 5
            self.assertEqual(worksheet.cell(
                row=4, column=7).value, 150.0)  # 100 + 50
            self.assertEqual(worksheet.cell(
                row=4, column=8).value, 111.0)  # 74 + 37

    def test_get_beverage_variance_data_success(self):
        """Test successful beverage variance data retrieval"""
        year, month = 2025, 6

        # Mock database connection and cursor
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_connection.cursor.return_value = mock_cursor
        mock_connection.__enter__.return_value = mock_connection
        mock_connection.__exit__.return_value = None

        self.mock_db_manager.get_connection.return_value = mock_connection

        # Mock query results
        mock_cursor.fetchall.return_value = [
            (1, '加拿大一店', '张三', '蒋冰洁', '加拿大', 10.0, 100.0, 74.0)
        ]

        result = self.generator.get_beverage_variance_data(year, month)

        # Verify result
        self.assertEqual(len(result), 1)
        self.assertEqual(result[0]['store_name'], '加拿大一店')
        self.assertEqual(result[0]['variance_quantity'], '10')
        self.assertEqual(result[0]['impact_amount_local'], '100')

    def test_get_beverage_variance_data_database_error(self):
        """Test beverage variance data retrieval with database error"""
        year, month = 2025, 6

        # Mock database connection to raise exception
        self.mock_db_manager.get_connection.side_effect = Exception(
            "Database connection failed")

        result = self.generator.get_beverage_variance_data(year, month)

        # Should return empty list on error
        self.assertEqual(result, [])

    def test_get_beverage_variance_data_empty_results(self):
        """Test beverage variance data retrieval with empty results"""
        year, month = 2025, 6

        # Mock database connection and cursor
        mock_connection = MagicMock()
        mock_cursor = Mock()
        mock_connection.cursor.return_value = mock_cursor
        mock_connection.__enter__.return_value = mock_connection
        mock_connection.__exit__.return_value = None

        self.mock_db_manager.get_connection.return_value = mock_connection

        # Mock empty query results
        mock_cursor.fetchall.return_value = []

        result = self.generator.get_beverage_variance_data(year, month)

        # Should return empty list
        self.assertEqual(result, [])

    def test_worksheet_styling(self):
        """Test worksheet styling is applied correctly"""
        workbook = Workbook()
        target_date = "2025-06-01"

        mock_variance_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'store_manager': '张三',
                'regional_manager': '蒋冰洁',
                'country': '加拿大',
                'variance_quantity': '10',
                'impact_amount_local': '100',
                'impact_amount_usd': '74'
            }
        ]

        with patch.object(self.generator, 'get_beverage_variance_data', return_value=mock_variance_data):
            worksheet = self.generator.generate_worksheet(
                workbook, target_date)

            # Check header styling
            header_cell = worksheet.cell(row=1, column=1)
            # Check font color (may have '00' prefix)
            color_value = header_cell.font.color.rgb
            self.assertTrue(color_value.endswith("FFFFFF"))
            self.assertEqual(header_cell.fill.start_color.index, "C41E3A")

            # Check column widths are set
            self.assertEqual(worksheet.column_dimensions['A'].width, 10)
            self.assertEqual(worksheet.column_dimensions['D'].width, 20)

    def test_date_parsing(self):
        """Test date parsing in worksheet generation"""
        workbook = Workbook()
        target_date = "2025-06-15"  # Different day

        with patch.object(self.generator, 'get_beverage_variance_data', return_value=[]):
            worksheet = self.generator.generate_worksheet(
                workbook, target_date)

            # The month should still be extracted correctly
            self.assertIsNotNone(worksheet)
            # Verify get_beverage_variance_data was called with correct year/month
            self.generator.get_beverage_variance_data.assert_called_with(
                2025, 6)


if __name__ == '__main__':
    unittest.main()
