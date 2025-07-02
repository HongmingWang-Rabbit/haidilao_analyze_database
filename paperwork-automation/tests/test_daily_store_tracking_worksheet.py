"""
Comprehensive tests for Daily Store Performance Tracking Worksheet Generator.

This test module covers:
- Generator initialization
- Worksheet creation and structure
- Data processing and calculations
- Formula generation and accuracy
- Formatting and styling
- Error handling and edge cases
"""

import openpyxl
from openpyxl import Workbook
from lib.daily_store_tracking_worksheet import DailyStoreTrackingGenerator
import unittest
from unittest.mock import Mock, patch, MagicMock
from datetime import datetime, timedelta
import sys
import os
from pathlib import Path

# Add project root to path for imports
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


class TestDailyStoreTrackingGenerator(unittest.TestCase):
    """Comprehensive tests for DailyStoreTrackingGenerator."""

    def setUp(self):
        """Set up test fixtures."""
        self.mock_data_provider = Mock()
        self.generator = DailyStoreTrackingGenerator(self.mock_data_provider)
        self.test_date = "2025-06-28"
        self.sample_store_data = [
            {
                'store_id': 5,
                'store_name': '加拿大五店',
                'manager_name': '陈浩',
                'seating_capacity': 55,
                'annual_avg_turnover_2024': 5.5,
                'current_turnover_rate': 5.73,
                'prev_turnover_rate': 6.71,
                'current_revenue': 4.396,
                'prev_revenue': 3.964
            },
            {
                'store_id': 6,
                'store_name': '加拿大六店',
                'manager_name': '高新菊',
                'seating_capacity': 56,
                'annual_avg_turnover_2024': 3.67,
                'current_turnover_rate': 4.05,
                'prev_turnover_rate': 4.16,
                'current_revenue': 2.851,
                'prev_revenue': 2.474
            }
        ]

    def test_generator_initialization(self):
        """Test generator initialization with data provider."""
        self.assertIsNotNone(self.generator)
        self.assertEqual(self.generator.data_provider, self.mock_data_provider)
        self.assertIsNotNone(self.generator.logger)

    def test_worksheet_creation(self):
        """Test worksheet creation with proper name."""
        workbook = Workbook()

        # Mock data provider to return sample data
        self.mock_data_provider.get_daily_store_performance.return_value = []

        self.generator.generate_worksheet(workbook, self.test_date)

        # Check worksheet was created with correct name
        self.assertIn("门店日-加拿大", workbook.sheetnames)
        ws = workbook["门店日-加拿大"]
        self.assertIsNotNone(ws)

    def test_header_generation(self):
        """Test header generation with correct structure."""
        workbook = Workbook()
        ws = workbook.active

        target_dt = datetime.strptime(self.test_date, '%Y-%m-%d')
        prev_year_dt = target_dt.replace(year=target_dt.year - 1)

        self.generator._generate_headers(ws, target_dt, prev_year_dt)

        # Check main headers (row 1)
        expected_headers = [
            "序号", "门店名称", "店经理", "餐位数", "24年全年平均翻台率",
            "日翻台率-考核", "", "", "", "",
            "日营业收入-不含税(万加元)", "", "", "", "",
            "综合得分", "综合排名"
        ]

        for col, expected in enumerate(expected_headers, 1):
            actual = ws.cell(row=1, column=col).value
            self.assertEqual(actual, expected,
                             f"Header mismatch at column {col}")

        # Check sub-headers (row 2)
        self.assertEqual(ws.cell(row=2, column=6).value, "25年6月28日")
        self.assertEqual(ws.cell(row=2, column=7).value, "24年6月28日")
        self.assertEqual(ws.cell(row=2, column=8).value, "对比")
        self.assertEqual(ws.cell(row=2, column=9).value, "基础值得分")
        self.assertEqual(ws.cell(row=2, column=10).value, "精进值得分")

        # Check formula references
        self.assertEqual(ws.cell(row=2, column=11).value, "=F2")
        self.assertEqual(ws.cell(row=2, column=12).value, "=G2")

    def test_store_data_processing(self):
        """Test store data processing and combination."""
        # Mock current year data
        current_data = [
            {'store_id': 5, 'store_name': '加拿大五店',
                'turnover_rate': 5.73, 'daily_revenue_cad_10k': 4.396},
            {'store_id': 6, 'store_name': '加拿大六店',
                'turnover_rate': 4.05, 'daily_revenue_cad_10k': 2.851}
        ]

        # Mock previous year data
        prev_data = [
            {'store_id': 5, 'turnover_rate': 6.71, 'daily_revenue_cad_10k': 3.964},
            {'store_id': 6, 'turnover_rate': 4.16, 'daily_revenue_cad_10k': 2.474}
        ]

        # Configure mock to return different data for different dates
        def mock_get_performance(date):
            if '2025' in date:
                return current_data
            else:
                return prev_data

        self.mock_data_provider.get_daily_store_performance.side_effect = mock_get_performance

        result = self.generator._get_store_performance_data(self.test_date)

        # Verify correct number of stores
        self.assertEqual(len(result), 2)

        # Verify data structure
        for store in result:
            self.assertIn('store_id', store)
            self.assertIn('current_turnover_rate', store)
            self.assertIn('prev_turnover_rate', store)
            self.assertIn('current_revenue', store)
            self.assertIn('prev_revenue', store)

    def test_mock_data_fallback(self):
        """Test mock data fallback when data provider fails."""
        # Mock data provider to raise an exception
        self.mock_data_provider.get_daily_store_performance.side_effect = Exception(
            "Database error")

        result = self.generator._get_store_performance_data(self.test_date)

        # Should return mock data
        self.assertGreater(len(result), 0)
        self.assertEqual(result[0]['store_name'], '加拿大五店')

    def test_regional_summary_generation(self):
        """Test regional summary row generation."""
        workbook = Workbook()
        ws = workbook.active

        target_dt = datetime.strptime(self.test_date, '%Y-%m-%d')
        prev_year_dt = target_dt.replace(year=target_dt.year - 1)

        self.generator._add_regional_summary(
            ws, self.sample_store_data, target_dt, prev_year_dt)

        # Check regional identifiers
        self.assertEqual(ws.cell(row=3, column=4).value, "区域")
        self.assertEqual(ws.cell(row=3, column=5).value, "蒋冰遇")

        # Check formulas
        self.assertEqual(ws.cell(row=3, column=8).value, "=F3-G3")
        self.assertEqual(ws.cell(row=3, column=13).value, "=K3-L3")

        # Check calculated averages
        expected_avg_current_turnover = (5.73 + 4.05) / 2
        expected_avg_prev_turnover = (6.71 + 4.16) / 2
        expected_avg_current_revenue = (4.396 + 2.851) / 2
        expected_avg_prev_revenue = (3.964 + 2.474) / 2

        self.assertAlmostEqual(
            ws.cell(row=3, column=6).value, expected_avg_current_turnover, places=2)
        self.assertAlmostEqual(
            ws.cell(row=3, column=7).value, expected_avg_prev_turnover, places=2)
        self.assertAlmostEqual(
            ws.cell(row=3, column=11).value, expected_avg_current_revenue, places=2)
        self.assertAlmostEqual(
            ws.cell(row=3, column=12).value, expected_avg_prev_revenue, places=2)

    def test_store_data_addition(self):
        """Test individual store data addition with formulas."""
        workbook = Workbook()
        ws = workbook.active

        target_dt = datetime.strptime(self.test_date, '%Y-%m-%d')
        prev_year_dt = target_dt.replace(year=target_dt.year - 1)

        self.generator._add_store_data(
            ws, self.sample_store_data, target_dt, prev_year_dt)

        # Check first store data (row 4)
        self.assertEqual(ws.cell(row=4, column=1).value, 1)  # Serial number
        self.assertEqual(ws.cell(row=4, column=2).value, "加拿大五店")  # Store name
        self.assertEqual(ws.cell(row=4, column=3).value, "陈浩")  # Manager name
        self.assertEqual(ws.cell(row=4, column=4).value,
                         55)  # Seating capacity

        # Check turnover data
        self.assertEqual(ws.cell(row=4, column=6).value,
                         5.73)  # Current turnover
        self.assertEqual(ws.cell(row=4, column=7).value,
                         6.71)  # Previous turnover

        # Check revenue data
        self.assertEqual(ws.cell(row=4, column=11).value,
                         4.396)  # Current revenue
        self.assertEqual(ws.cell(row=4, column=12).value,
                         3.964)  # Previous revenue

        # Check second store data (row 5)
        self.assertEqual(ws.cell(row=5, column=1).value, 2)  # Serial number
        self.assertEqual(ws.cell(row=5, column=2).value, "加拿大六店")  # Store name

    def test_formula_generation(self):
        """Test formula generation for calculations and scoring."""
        workbook = Workbook()
        ws = workbook.active

        target_dt = datetime.strptime(self.test_date, '%Y-%m-%d')
        prev_year_dt = target_dt.replace(year=target_dt.year - 1)

        self.generator._add_store_data(
            ws, self.sample_store_data, target_dt, prev_year_dt)

        # Check comparison formulas
        self.assertEqual(ws.cell(row=4, column=8).value,
                         "=F4-G4")  # Turnover comparison
        self.assertEqual(ws.cell(row=4, column=13).value,
                         "=K4-L4")  # Revenue comparison

        # Check normalized scoring formulas
        expected_basic_score = "=(F4-MIN($F$4:$F$5))/(MAX($F$4:$F$5)-MIN($F$4:$F$5))"
        expected_improvement_score = "=(H4-MIN($H$4:$H$5))/(MAX($H$4:$H$5)-MIN($H$4:$H$5))"

        self.assertEqual(ws.cell(row=4, column=9).value, expected_basic_score)
        self.assertEqual(ws.cell(row=4, column=10).value,
                         expected_improvement_score)

        # Check comprehensive scoring formula
        expected_comprehensive = "=I4*0.25+J4*0.25+N4*0.25+O4*0.25"
        self.assertEqual(ws.cell(row=4, column=16).value,
                         expected_comprehensive)

        # Check ranking formula
        expected_ranking = "=RANK(P4,$P$4:$P$5)"
        self.assertEqual(ws.cell(row=4, column=17).value, expected_ranking)

    def test_worksheet_formatting(self):
        """Test worksheet formatting and styling."""
        workbook = Workbook()
        ws = workbook.active

        # Add some test data first
        ws.cell(row=1, column=1, value="Test")
        ws.cell(row=2, column=1, value="Test")

        self.generator._apply_formatting(ws, 2)

        # Check column widths
        self.assertEqual(ws.column_dimensions['A'].width, 6)
        self.assertEqual(ws.column_dimensions['B'].width, 15)
        self.assertEqual(ws.column_dimensions['P'].width, 10)

        # Check freeze panes
        self.assertEqual(ws.freeze_panes, "A3")

        # Check header formatting
        header_cell = ws.cell(row=1, column=1)
        self.assertTrue(header_cell.font.bold)
        self.assertEqual(header_cell.font.color.rgb, "FFFFFF")

    def test_empty_store_data_handling(self):
        """Test handling of empty store data."""
        workbook = Workbook()
        ws = workbook.active

        target_dt = datetime.strptime(self.test_date, '%Y-%m-%d')
        prev_year_dt = target_dt.replace(year=target_dt.year - 1)

        # Test with empty store data
        empty_data = []
        self.generator._add_regional_summary(
            ws, empty_data, target_dt, prev_year_dt)

        # Should handle empty data gracefully
        self.assertEqual(ws.cell(row=3, column=6).value,
                         0)  # Average should be 0
        self.assertEqual(ws.cell(row=3, column=7).value, 0)
        self.assertEqual(ws.cell(row=3, column=11).value, 0)
        self.assertEqual(ws.cell(row=3, column=12).value, 0)

    def test_date_formatting_in_headers(self):
        """Test correct date formatting in headers."""
        workbook = Workbook()
        ws = workbook.active

        # Test different date
        test_date = "2025-12-31"
        target_dt = datetime.strptime(test_date, '%Y-%m-%d')
        prev_year_dt = target_dt.replace(year=target_dt.year - 1)

        self.generator._generate_headers(ws, target_dt, prev_year_dt)

        # Check date formatting
        self.assertEqual(ws.cell(row=2, column=6).value, "25年12月31日")
        self.assertEqual(ws.cell(row=2, column=7).value, "24年12月31日")

    def test_complete_worksheet_generation(self):
        """Test complete worksheet generation end-to-end."""
        workbook = Workbook()

        # Mock data provider
        self.mock_data_provider.get_daily_store_performance.return_value = []

        # Should not raise any exceptions
        self.generator.generate_worksheet(workbook, self.test_date)

        # Check worksheet exists
        self.assertIn("门店日-加拿大", workbook.sheetnames)

        ws = workbook["门店日-加拿大"]

        # Check basic structure
        self.assertEqual(ws.cell(row=1, column=1).value, "序号")
        self.assertEqual(ws.cell(row=1, column=2).value, "门店名称")
        self.assertEqual(ws.cell(row=2, column=8).value, "对比")

    def test_error_handling_in_generation(self):
        """Test error handling during worksheet generation."""
        workbook = Workbook()

        # Mock data provider to raise exception
        self.mock_data_provider.get_daily_store_performance.side_effect = Exception(
            "Test error")

        # Should handle error gracefully and still generate worksheet
        with self.assertLogs(level='ERROR') as log:
            self.generator.generate_worksheet(workbook, self.test_date)

        # Check that error was logged
        self.assertTrue(
            any("Error getting store performance data" in message for message in log.output))

    def test_invalid_date_handling(self):
        """Test handling of invalid date formats."""
        workbook = Workbook()

        with self.assertRaises(ValueError):
            self.generator.generate_worksheet(workbook, "invalid-date")

    def test_comprehensive_score_calculation_logic(self):
        """Test the logic behind comprehensive score calculation."""
        # The comprehensive score should be:
        # 25% * Basic Turnover Score + 25% * Improvement Turnover Score +
        # 25% * Basic Revenue Score + 25% * Improvement Revenue Score

        workbook = Workbook()
        ws = workbook.active

        target_dt = datetime.strptime(self.test_date, '%Y-%m-%d')
        prev_year_dt = target_dt.replace(year=target_dt.year - 1)

        # Test with known data
        test_data = [
            {
                'store_id': 1,
                'store_name': 'Test Store 1',
                'manager_name': 'Manager 1',
                'seating_capacity': 50,
                'annual_avg_turnover_2024': 5.0,
                'current_turnover_rate': 6.0,
                'prev_turnover_rate': 5.5,
                'current_revenue': 4.0,
                'prev_revenue': 3.5
            }
        ]

        self.generator._add_store_data(ws, test_data, target_dt, prev_year_dt)

        # Check the comprehensive score formula
        comprehensive_formula = ws.cell(row=4, column=16).value
        self.assertEqual(comprehensive_formula,
                         "=I4*0.25+J4*0.25+N4*0.25+O4*0.25")

        # Verify each component has equal weight (25%)
        self.assertIn("*0.25", comprehensive_formula)
        self.assertEqual(comprehensive_formula.count("*0.25"), 4)


if __name__ == '__main__':
    unittest.main()
