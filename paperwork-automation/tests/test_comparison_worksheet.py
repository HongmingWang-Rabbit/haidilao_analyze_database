#!/usr/bin/env python3
"""
Comprehensive tests for ComparisonWorksheetGenerator class.
"""

import unittest
from unittest.mock import Mock, patch, MagicMock
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent))
from lib.comparison_worksheet import ComparisonWorksheetGenerator


class TestComparisonWorksheetGenerator(unittest.TestCase):
    """Test cases for ComparisonWorksheetGenerator"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.store_names = {
            1: "加拿大一店",
            2: "加拿大二店", 
            3: "加拿大三店"
        }
        self.target_date = "2025-06-10"
        self.generator = ComparisonWorksheetGenerator(self.store_names, self.target_date)
        
        # Mock data
        self.daily_data = [
            {
                'store_id': 1,
                'revenue_tax_not_included': 25000.50,
                'tables_served_validated': 180.5,
                'customers': 520,
                'turnover_rate': 3.8,
                'takeout_tables': 12.5,
                'tables_served': 185.0
            },
            {
                'store_id': 2,
                'revenue_tax_not_included': 18000.25,
                'tables_served_validated': 145.0,
                'customers': 420,
                'turnover_rate': 3.2,
                'takeout_tables': 8.5,
                'tables_served': 148.0
            }
        ]
        
        self.monthly_data = [
            {
                'store_id': 1,
                'monthly_revenue': 280000.75,
                'monthly_tables_validated': 1850.5,
                'customers': 5200,
                'avg_turnover_rate': 3.9,
                'takeout_tables': 125.5,
                'tables_served': 1900.0
            },
            {
                'store_id': 2,
                'monthly_revenue': 195000.50,
                'monthly_tables_validated': 1450.0,
                'customers': 4200,
                'avg_turnover_rate': 3.3,
                'takeout_tables': 85.5,
                'tables_served': 1480.0
            }
        ]
        
        self.previous_month_data = [
            {
                'store_id': 1,
                'prev_monthly_revenue': 260000.00,
                'prev_monthly_tables_validated': 1750.0,
                'prev_monthly_customers': 4800,
                'prev_avg_turnover_rate': 3.7,
                'prev_takeout_tables': 115.0,
                'prev_tables_served': 1800.0
            },
            {
                'store_id': 2,
                'prev_monthly_revenue': 185000.00,
                'prev_monthly_tables_validated': 1400.0,
                'prev_monthly_customers': 4000,
                'prev_avg_turnover_rate': 3.1,
                'prev_takeout_tables': 80.0,
                'prev_tables_served': 1430.0
            }
        ]
        
        self.monthly_targets = [
            {
                'store_id': 1,
                'monthly_target': 300000.00
            },
            {
                'store_id': 2,
                'monthly_target': 200000.00
            }
        ]

    def test_generator_initialization(self):
        """Test generator initialization"""
        self.assertEqual(self.generator.store_names, self.store_names)
        self.assertEqual(self.generator.target_date, self.target_date)
        
        # Test with different data
        new_stores = {1: "Test Store"}
        new_date = "2025-01-01"
        new_generator = ComparisonWorksheetGenerator(new_stores, new_date)
        self.assertEqual(new_generator.store_names, new_stores)
        self.assertEqual(new_generator.target_date, new_date)

    def test_calculate_time_progress(self):
        """Test time progress calculation"""
        # Test June 10th (10 days into June)
        progress = self.generator.calculate_time_progress("2025-06-10")
        expected = 10 / 30  # 10 days out of 30 days in June
        self.assertAlmostEqual(progress, expected, places=4)
        
        # Test January 15th (15 days into January)
        progress = self.generator.calculate_time_progress("2025-01-15")
        expected = 15 / 31  # 15 days out of 31 days in January
        self.assertAlmostEqual(progress, expected, places=4)
        
        # Test February 28th (28 days into February)
        progress = self.generator.calculate_time_progress("2025-02-28")
        expected = 28 / 28  # 28 days out of 28 days in February (non-leap year)
        self.assertAlmostEqual(progress, expected, places=4)
        
        # Test edge case: first day of month
        progress = self.generator.calculate_time_progress("2025-03-01")
        expected = 1 / 31
        self.assertAlmostEqual(progress, expected, places=4)
        
        # Test edge case: last day of month
        progress = self.generator.calculate_time_progress("2025-04-30")
        expected = 30 / 30
        self.assertAlmostEqual(progress, expected, places=4)

    def test_get_cell_value_daily_revenue(self):
        """Test get_cell_value for daily revenue"""
        daily_dict = {row['store_id']: row for row in self.daily_data}
        monthly_dict = {row['store_id']: row for row in self.monthly_data}
        prev_month_dict = {row['store_id']: row for row in self.previous_month_data}
        targets_dict = {row['store_id']: row for row in self.monthly_targets}
        
        # Test daily revenue
        value = self.generator.get_cell_value(
            "今日营业收入(万)", "今日", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        expected = 25000.50 / 10000  # Convert to 万
        self.assertAlmostEqual(value, expected, places=2)
        
        # Test store not in data
        value = self.generator.get_cell_value(
            "今日营业收入(万)", "今日", "不存在的店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        self.assertEqual(value, 0)

    def test_get_cell_value_monthly_revenue(self):
        """Test get_cell_value for monthly revenue"""
        daily_dict = {row['store_id']: row for row in self.daily_data}
        monthly_dict = {row['store_id']: row for row in self.monthly_data}
        prev_month_dict = {row['store_id']: row for row in self.previous_month_data}
        targets_dict = {row['store_id']: row for row in self.monthly_targets}
        
        # Test monthly revenue
        value = self.generator.get_cell_value(
            "本月营业收入(万)", "本月", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        expected = 280000.75 / 10000  # Convert to 万
        self.assertAlmostEqual(value, expected, places=2)

    def test_get_cell_value_target_completion(self):
        """Test get_cell_value for target completion rate"""
        daily_dict = {row['store_id']: row for row in self.daily_data}
        monthly_dict = {row['store_id']: row for row in self.monthly_data}
        prev_month_dict = {row['store_id']: row for row in self.previous_month_data}
        targets_dict = {row['store_id']: row for row in self.monthly_targets}
        
        # Test target completion rate
        value = self.generator.get_cell_value(
            "本月截止目标完成率", "本月", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        # Monthly revenue: 280000.75, Target: 300000.00
        expected = f"{(280000.75 / 300000.00 * 100):.1f}%"
        self.assertEqual(value, expected)
        
        # Test with no target
        value = self.generator.get_cell_value(
            "本月截止目标完成率", "本月", "加拿大三店", 1,
            daily_dict, monthly_dict, prev_month_dict, {}, 0.33
        )
        self.assertEqual(value, "无目标")

    def test_get_cell_value_time_progress(self):
        """Test get_cell_value for time progress"""
        daily_dict = {row['store_id']: row for row in self.daily_data}
        monthly_dict = {row['store_id']: row for row in self.monthly_data}
        prev_month_dict = {row['store_id']: row for row in self.previous_month_data}
        targets_dict = {row['store_id']: row for row in self.monthly_targets}
        
        # Test time progress
        value = self.generator.get_cell_value(
            "标准时间进度", "本月", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        expected = f"{(0.33 * 100):.1f}%"
        self.assertEqual(value, expected)

    def test_get_cell_value_percentage_changes(self):
        """Test get_cell_value for percentage changes"""
        daily_dict = {row['store_id']: row for row in self.daily_data}
        monthly_dict = {row['store_id']: row for row in self.monthly_data}
        prev_month_dict = {row['store_id']: row for row in self.previous_month_data}
        targets_dict = {row['store_id']: row for row in self.monthly_targets}
        
        # Test revenue environment change
        value = self.generator.get_cell_value(
            "环比营业收入变化(万)", "环比", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        # Current: 280000.75, Previous: 260000.00
        expected = f"{(280000.75 - 260000.00) / 10000:+.2f}"
        self.assertEqual(value, expected)
        
        # Test percentage change
        value = self.generator.get_cell_value(
            "环比营业收入变化(%)", "环比", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        # Percentage change: (280000.75 - 260000.00) / 260000.00 * 100
        expected = f"{((280000.75 - 260000.00) / 260000.00 * 100):+.1f}%"
        self.assertEqual(value, expected)

    def test_get_cell_value_per_capita_calculations(self):
        """Test get_cell_value for per capita calculations"""
        daily_dict = {row['store_id']: row for row in self.daily_data}
        monthly_dict = {row['store_id']: row for row in self.monthly_data}
        prev_month_dict = {row['store_id']: row for row in self.previous_month_data}
        targets_dict = {row['store_id']: row for row in self.monthly_targets}
        
        # Test daily per capita
        value = self.generator.get_cell_value(
            "今日人均", "今日", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        expected = 25000.50 / 520  # Revenue / customers
        self.assertAlmostEqual(value, expected, places=2)
        
        # Test monthly per capita
        value = self.generator.get_cell_value(
            "本月人均", "本月", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        expected = 280000.75 / 5200  # Monthly revenue / customers
        self.assertAlmostEqual(value, expected, places=2)

    def test_get_cell_value_per_table_calculations(self):
        """Test get_cell_value for per table calculations"""
        daily_dict = {row['store_id']: row for row in self.daily_data}
        monthly_dict = {row['store_id']: row for row in self.monthly_data}
        prev_month_dict = {row['store_id']: row for row in self.previous_month_data}
        targets_dict = {row['store_id']: row for row in self.monthly_targets}
        
        # Test daily per table
        value = self.generator.get_cell_value(
            "今日单桌", "今日", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        expected = 25000.50 / 180.5  # Revenue / tables
        self.assertAlmostEqual(value, expected, places=2)
        
        # Test monthly per table
        value = self.generator.get_cell_value(
            "本月单桌", "本月", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        expected = 280000.75 / 1850.5  # Monthly revenue / tables
        self.assertAlmostEqual(value, expected, places=2)

    def test_get_cell_value_turnover_rate(self):
        """Test get_cell_value for turnover rate"""
        daily_dict = {row['store_id']: row for row in self.daily_data}
        monthly_dict = {row['store_id']: row for row in self.monthly_data}
        prev_month_dict = {row['store_id']: row for row in self.previous_month_data}
        targets_dict = {row['store_id']: row for row in self.monthly_targets}
        
        # Test daily turnover rate
        value = self.generator.get_cell_value(
            "今日翻台率", "今日", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        self.assertEqual(value, 3.8)
        
        # Test monthly turnover rate
        value = self.generator.get_cell_value(
            "本月翻台率", "本月", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        self.assertEqual(value, 3.9)

    def test_get_cell_value_unknown_content(self):
        """Test get_cell_value with unknown content"""
        daily_dict = {row['store_id']: row for row in self.daily_data}
        monthly_dict = {row['store_id']: row for row in self.monthly_data}
        prev_month_dict = {row['store_id']: row for row in self.previous_month_data}
        targets_dict = {row['store_id']: row for row in self.monthly_targets}
        
        # Test unknown content
        value = self.generator.get_cell_value(
            "未知内容", "今日", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        self.assertEqual(value, 0)

    def test_get_cell_value_zero_division_protection(self):
        """Test get_cell_value with zero division protection"""
        # Create data with zero values
        daily_data_zero = [{
            'store_id': 1,
            'revenue_tax_not_included': 25000.50,
            'tables_served_validated': 0,  # Zero tables
            'customers': 0,  # Zero customers
            'turnover_rate': 3.8,
            'takeout_tables': 12.5,
            'tables_served': 185.0
        }]
        
        daily_dict = {row['store_id']: row for row in daily_data_zero}
        monthly_dict = {row['store_id']: row for row in self.monthly_data}
        prev_month_dict = {row['store_id']: row for row in self.previous_month_data}
        targets_dict = {row['store_id']: row for row in self.monthly_targets}
        
        # Test per capita with zero customers
        value = self.generator.get_cell_value(
            "今日人均", "今日", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        self.assertEqual(value, 0)
        
        # Test per table with zero tables
        value = self.generator.get_cell_value(
            "今日单桌", "今日", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        self.assertEqual(value, 0)

    def test_generate_worksheet(self):
        """Test worksheet generation"""
        wb = Workbook()
        
        # Call generate_worksheet
        self.generator.generate_worksheet(
            wb, self.daily_data, self.monthly_data, self.previous_month_data, self.monthly_targets
        )
        
        # Check if worksheet was created
        self.assertIn("对比上月表", wb.sheetnames)
        ws = wb["对比上月表"]
        
        # Check if data was populated
        self.assertIsNotNone(ws['A1'].value)
        
        # Check if store names are present
        found_store_names = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    for store_name in self.store_names.values():
                        if store_name in cell:
                            found_store_names.append(store_name)
        
        self.assertTrue(len(found_store_names) > 0)

    def test_generate_worksheet_with_empty_data(self):
        """Test worksheet generation with empty data"""
        wb = Workbook()
        
        # Call generate_worksheet with empty data
        self.generator.generate_worksheet(wb, [], [], [], [])
        
        # Check if worksheet was created
        self.assertIn("对比上月表", wb.sheetnames)
        ws = wb["对比上月表"]
        
        # Check if basic structure exists
        self.assertIsNotNone(ws['A1'].value)

    def test_generate_worksheet_with_none_data(self):
        """Test worksheet generation with None data"""
        wb = Workbook()
        
        # Call generate_worksheet with None data
        self.generator.generate_worksheet(wb, None, None, None, None)
        
        # Check if worksheet was created
        self.assertIn("对比上月表", wb.sheetnames)
        ws = wb["对比上月表"]
        
        # Check if basic structure exists
        self.assertIsNotNone(ws['A1'].value)

    def test_apply_common_formatting(self):
        """Test common formatting application"""
        wb = Workbook()
        ws = wb.active
        
        # Add some test data
        ws['A1'] = "Test Header"
        ws['B1'] = "Test Data"
        ws['A2'] = "Store 1"
        ws['B2'] = 123.45
        
        # Apply formatting
        self.generator.apply_common_formatting(ws, 2)
        
        # Check if borders were applied
        self.assertIsNotNone(ws['A1'].border)
        self.assertIsNotNone(ws['B1'].border)
        
        # Check if alignment was applied
        self.assertIsNotNone(ws['A1'].alignment)
        self.assertIsNotNone(ws['B1'].alignment)

    def test_data_type_handling(self):
        """Test handling of different data types"""
        daily_dict = {row['store_id']: row for row in self.daily_data}
        monthly_dict = {row['store_id']: row for row in self.monthly_data}
        prev_month_dict = {row['store_id']: row for row in self.previous_month_data}
        targets_dict = {row['store_id']: row for row in self.monthly_targets}
        
        # Test with string values that should be converted to float
        daily_dict[1]['revenue_tax_not_included'] = "25000.50"
        
        value = self.generator.get_cell_value(
            "今日营业收入(万)", "今日", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        expected = 25000.50 / 10000
        self.assertAlmostEqual(value, expected, places=2)
        
        # Test with None values
        daily_dict[1]['revenue_tax_not_included'] = None
        
        value = self.generator.get_cell_value(
            "今日营业收入(万)", "今日", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        self.assertEqual(value, 0)

    def test_edge_cases(self):
        """Test edge cases and error handling"""
        # Test with empty store_names
        empty_generator = ComparisonWorksheetGenerator({}, self.target_date)
        wb = Workbook()
        
        # Should not crash with empty store names
        empty_generator.generate_worksheet(wb, [], [], [], [])
        self.assertIn("对比上月表", wb.sheetnames)
        
        # Test with invalid date format
        try:
            progress = self.generator.calculate_time_progress("invalid-date")
            # Should handle gracefully or raise appropriate exception
        except:
            # Expected behavior for invalid date
            pass

    def test_large_numbers_formatting(self):
        """Test handling of large numbers"""
        # Create data with large numbers
        large_daily_data = [{
            'store_id': 1,
            'revenue_tax_not_included': 999999999.99,
            'tables_served_validated': 9999.99,
            'customers': 99999,
            'turnover_rate': 99.99,
            'takeout_tables': 999.99,
            'tables_served': 9999.99
        }]
        
        daily_dict = {row['store_id']: row for row in large_daily_data}
        monthly_dict = {row['store_id']: row for row in self.monthly_data}
        prev_month_dict = {row['store_id']: row for row in self.previous_month_data}
        targets_dict = {row['store_id']: row for row in self.monthly_targets}
        
        # Test large revenue value
        value = self.generator.get_cell_value(
            "今日营业收入(万)", "今日", "加拿大一店", 1,
            daily_dict, monthly_dict, prev_month_dict, targets_dict, 0.33
        )
        expected = 999999999.99 / 10000
        self.assertAlmostEqual(value, expected, places=2)


if __name__ == '__main__':
    unittest.main() 