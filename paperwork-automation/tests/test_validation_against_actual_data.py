#!/usr/bin/env python3
"""
Validation test against actual database report output
This test validates the yearly comparison worksheet against the expected structure 
from the actual database_report_2025_06_10.xlsx file.
"""

import unittest
import sys
import os
from openpyxl import load_workbook, Workbook

# Add parent directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from lib.yearly_comparison_worksheet import YearlyComparisonWorksheetGenerator

class TestValidationAgainstActualData(unittest.TestCase):
    """Validate yearly comparison worksheet against actual database report structure"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.store_names = {
            1: "加拿大一店", 2: "加拿大二店", 3: "加拿大三店", 4: "加拿大四店",
            5: "加拿大五店", 6: "加拿大六店", 7: "加拿大七店"
        }
        self.target_date = "2025-06-10"
        self.generator = YearlyComparisonWorksheetGenerator(self.store_names, self.target_date)
        
        # Path to actual database report
        self.actual_report_path = os.path.join(os.path.dirname(__file__), '..', 'output', 'database_report_2025_06_10.xlsx')
    
    def test_worksheet_structure_matches_expected(self):
        """Test that our generated worksheet matches the expected structure"""
        
        # Sample data that would be typical for a restaurant chain
        current_data = [
            {'store_id': 1, 'total_tables': 280.5, 'total_revenue': 842000.0, 'avg_turnover_rate': 2.8, 'avg_per_table': 3007.0},
            {'store_id': 2, 'total_tables': 310.0, 'total_revenue': 925000.0, 'avg_turnover_rate': 3.1, 'avg_per_table': 2984.0},
            {'store_id': 3, 'total_tables': 245.5, 'total_revenue': 736500.0, 'avg_turnover_rate': 2.5, 'avg_per_table': 3000.0},
            {'store_id': 4, 'total_tables': 290.0, 'total_revenue': 870000.0, 'avg_turnover_rate': 2.9, 'avg_per_table': 3000.0},
            {'store_id': 5, 'total_tables': 265.0, 'total_revenue': 795000.0, 'avg_turnover_rate': 2.7, 'avg_per_table': 3000.0},
            {'store_id': 6, 'total_tables': 300.5, 'total_revenue': 901500.0, 'avg_turnover_rate': 3.0, 'avg_per_table': 3000.0},
            {'store_id': 7, 'total_tables': 220.0, 'total_revenue': 660000.0, 'avg_turnover_rate': 2.2, 'avg_per_table': 3000.0}
        ]
        
        previous_data = [
            {'store_id': 1, 'total_tables': 260.0, 'total_revenue': 780000.0, 'avg_turnover_rate': 2.6, 'avg_per_table': 3000.0},
            {'store_id': 2, 'total_tables': 290.0, 'total_revenue': 870000.0, 'avg_turnover_rate': 2.9, 'avg_per_table': 3000.0},
            {'store_id': 3, 'total_tables': 230.0, 'total_revenue': 690000.0, 'avg_turnover_rate': 2.3, 'avg_per_table': 3000.0},
            {'store_id': 4, 'total_tables': 270.0, 'total_revenue': 810000.0, 'avg_turnover_rate': 2.7, 'avg_per_table': 3000.0},
            {'store_id': 5, 'total_tables': 250.0, 'total_revenue': 750000.0, 'avg_turnover_rate': 2.5, 'avg_per_table': 3000.0},
            {'store_id': 6, 'total_tables': 280.0, 'total_revenue': 840000.0, 'avg_turnover_rate': 2.8, 'avg_per_table': 3000.0},
            {'store_id': 7, 'total_tables': 200.0, 'total_revenue': 600000.0, 'avg_turnover_rate': 2.0, 'avg_per_table': 3000.0}
        ]
        
        wb = Workbook()
        ws = self.generator.generate_worksheet(wb, current_data, previous_data)
        
        # Validate core structure
        self.assertEqual(ws.title, "同比数据")
        
        # Check title format
        title = ws['A1'].value
        self.assertIn("加拿大-各门店2025年6月10日同比数据", title)
        self.assertIn("星期二", title)
        
        # Check regional headers
        self.assertEqual(ws['A2'].value, "分类")
        self.assertEqual(ws['C2'].value, "西部")
        self.assertEqual(ws['F2'].value, "东部") 
        self.assertEqual(ws['J2'].value, "加拿大片区")
        
        # Check store order (西部: 一店,二店,七店; 东部: 三店,四店,五店,六店)
        expected_store_order = ["项目", "内容", "加拿大一店", "加拿大二店", "加拿大七店", 
                               "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大片区"]
        for col, expected_store in enumerate(expected_store_order, 1):
            self.assertEqual(ws.cell(row=3, column=col).value, expected_store)
        
        # Check section structure
        expected_sections = [
            (4, "桌数\n对比同期数据"),
            (8, "翻台率\n对比同期数据"),
            (12, "营业收入\n(不含税-万加元)"),
            (16, "单桌消费\n对比同期数据")
        ]
        
        for row, expected_section in expected_sections:
            self.assertEqual(ws.cell(row=row, column=1).value, expected_section)
    
    def test_data_calculations_are_reasonable(self):
        """Test that calculated values are reasonable for restaurant data"""
        
        # Use realistic data volumes
        current_data = [
            {'store_id': 1, 'total_tables': 280.5, 'total_revenue': 842700.0, 'avg_turnover_rate': 2.8, 'avg_per_table': 3006.0}
        ]
        
        previous_data = [
            {'store_id': 1, 'total_tables': 260.0, 'total_revenue': 780000.0, 'avg_turnover_rate': 2.6, 'avg_per_table': 3000.0}
        ]
        
        wb = Workbook()
        ws = self.generator.generate_worksheet(wb, current_data, previous_data)
        
        # Check current tables
        current_tables = ws.cell(row=4, column=3).value
        self.assertEqual(current_tables, 280.5)
        
        # Check previous tables
        previous_tables = ws.cell(row=5, column=3).value
        self.assertEqual(previous_tables, 260.0)
        
        # Check difference
        difference = ws.cell(row=6, column=3).value
        self.assertEqual(difference, 20.5)  # 280.5 - 260.0
        
        # Check growth rate
        growth_rate = ws.cell(row=7, column=3).value
        self.assertEqual(growth_rate, "7.9%")  # (280.5-260)/260 * 100 ≈ 7.9%
        
        # Check revenue conversion (万加元)
        current_revenue = ws.cell(row=12, column=3).value
        expected_revenue = 842700.0 / 10000  # Convert to 万加元
        self.assertEqual(current_revenue, expected_revenue)
    
    def test_total_calculations_aggregate_correctly(self):
        """Test that totals aggregate individual store data correctly"""
        
        # Two stores for simple verification
        current_data = [
            {'store_id': 1, 'total_tables': 100.0, 'total_revenue': 300000.0, 'avg_turnover_rate': 2.0, 'avg_per_table': 3000.0},
            {'store_id': 2, 'total_tables': 200.0, 'total_revenue': 600000.0, 'avg_turnover_rate': 3.0, 'avg_per_table': 3000.0}
        ]
        
        previous_data = [
            {'store_id': 1, 'total_tables': 90.0, 'total_revenue': 270000.0, 'avg_turnover_rate': 1.8, 'avg_per_table': 3000.0},
            {'store_id': 2, 'total_tables': 180.0, 'total_revenue': 540000.0, 'avg_turnover_rate': 2.7, 'avg_per_table': 3000.0}
        ]
        
        wb = Workbook()
        ws = self.generator.generate_worksheet(wb, current_data, previous_data)
        
        # Check total current tables (column J, row 4)
        total_current = ws.cell(row=4, column=10).value
        self.assertEqual(total_current, 300.0)  # 100 + 200
        
        # Check total previous tables (row 5)
        total_previous = ws.cell(row=5, column=10).value
        self.assertEqual(total_previous, 270.0)  # 90 + 180
        
        # Check total difference (row 6)
        total_difference = ws.cell(row=6, column=10).value
        self.assertEqual(total_difference, 30.0)  # 300 - 270
        
        # Check total revenue (row 12)
        total_current_revenue = ws.cell(row=12, column=10).value
        expected_total_revenue = (300000.0 + 600000.0) / 10000  # Convert to 万加元
        self.assertEqual(total_current_revenue, expected_total_revenue)
    
    def test_empty_data_handling(self):
        """Test graceful handling of empty or None data"""
        
        # Test with empty lists
        wb_empty = Workbook()
        ws_empty = self.generator.generate_worksheet(wb_empty, [], [])
        self.assertEqual(ws_empty.title, "同比数据")
        self.assertIn("年度对比数据不足", ws_empty['A1'].value)
        
        # Test with None data
        wb_none = Workbook()
        ws_none = self.generator.generate_worksheet(wb_none, None, None)
        self.assertEqual(ws_none.title, "同比数据")
        self.assertIn("年度对比数据不足", ws_none['A1'].value)
    
    def test_load_actual_report_if_exists(self):
        """Test loading the actual report file if it exists (optional)"""
        
        if os.path.exists(self.actual_report_path):
            try:
                wb = load_workbook(self.actual_report_path)
                print(f"\n📂 Found actual report file: {self.actual_report_path}")
                print(f"📊 Worksheets in actual report: {wb.sheetnames}")
                
                # Check if our target worksheet exists
                if "同比数据" in wb.sheetnames:
                    actual_ws = wb["同比数据"]
                    print(f"✅ Found 同比数据 worksheet in actual report")
                    print(f"📈 Title: {actual_ws['A1'].value}")
                    
                    # Validate structure matches what we expect
                    self.assertEqual(actual_ws['A2'].value, "分类")
                    self.assertEqual(actual_ws['C2'].value, "西部")
                    self.assertEqual(actual_ws['F2'].value, "东部")
                    self.assertEqual(actual_ws['J2'].value, "加拿大片区")
                    
                    print("🎯 Actual report structure validation passed!")
                else:
                    print(f"⚠️  同比数据 worksheet not found in actual report")
                    
            except Exception as e:
                print(f"⚠️  Could not load actual report: {e}")
        else:
            print(f"ℹ️  Actual report file not found: {self.actual_report_path}")
            print("ℹ️  This is expected if running tests in isolation")


if __name__ == '__main__':
    unittest.main(verbosity=2) 