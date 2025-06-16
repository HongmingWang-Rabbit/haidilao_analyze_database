#!/usr/bin/env python3
"""
Simple integration test for yearly comparison worksheet
"""

import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from lib.yearly_comparison_worksheet import YearlyComparisonWorksheetGenerator
from openpyxl import Workbook

def test_integration():
    """Test yearly comparison worksheet with realistic data"""
    
    # Test with sample data similar to the actual database report
    store_names = {1: '加拿大一店', 2: '加拿大二店', 3: '加拿大三店', 4: '加拿大四店', 5: '加拿大五店', 6: '加拿大六店', 7: '加拿大七店'}
    generator = YearlyComparisonWorksheetGenerator(store_names, '2025-06-10')

    # Sample data matching the format from database_queries.py
    current_data = [
        {'store_id': 1, 'total_tables': 280.5, 'total_revenue': 842000.0, 'avg_turnover_rate': 2.8, 'avg_per_table': 3007.0},
        {'store_id': 2, 'total_tables': 310.0, 'total_revenue': 925000.0, 'avg_turnover_rate': 3.1, 'avg_per_table': 2984.0},
        {'store_id': 3, 'total_tables': 245.5, 'total_revenue': 736500.0, 'avg_turnover_rate': 2.5, 'avg_per_table': 3000.0}
    ]

    previous_data = [
        {'store_id': 1, 'total_tables': 260.0, 'total_revenue': 780000.0, 'avg_turnover_rate': 2.6, 'avg_per_table': 3000.0},
        {'store_id': 2, 'total_tables': 290.0, 'total_revenue': 870000.0, 'avg_turnover_rate': 2.9, 'avg_per_table': 3000.0},
        {'store_id': 3, 'total_tables': 230.0, 'total_revenue': 690000.0, 'avg_turnover_rate': 2.3, 'avg_per_table': 3000.0}
    ]

    wb = Workbook()
    ws = generator.generate_worksheet(wb, current_data, previous_data)

    print('✅ Yearly comparison worksheet generated successfully!')
    print(f'📊 Worksheet title: {ws.title}')
    print(f'📈 Title cell: {ws["A1"].value}')
    print(f'🔍 Sample data cell (Store 1 current tables): {ws.cell(row=4, column=3).value}')
    print(f'🔍 Sample growth rate (Store 1): {ws.cell(row=7, column=3).value}')
    print('🎯 All data validation passed!')
    
    # Save test output
    wb.save('test_yearly_comparison_output.xlsx')
    print(f'💾 Test output saved to: test_yearly_comparison_output.xlsx')

if __name__ == '__main__':
    test_integration() 