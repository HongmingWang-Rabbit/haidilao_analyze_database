#!/usr/bin/env python3
"""
Unit tests for the time segment worksheet generator.
Tests the TimeSegmentWorksheetGenerator class functionality.
"""

import unittest
import sys
import os
from pathlib import Path
from openpyxl import Workbook
from datetime import datetime

# Add parent directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from lib.time_segment_worksheet import TimeSegmentWorksheetGenerator

class TestTimeSegmentWorksheetGenerator(unittest.TestCase):
    """Test the time segment worksheet generator functionality"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.store_names = {
            1: "加拿大一店", 2: "加拿大二店", 3: "加拿大三店", 4: "加拿大四店",
            5: "加拿大五店", 6: "加拿大六店", 7: "加拿大七店"
        }
        self.test_date = "2025-06-10"
        self.generator = TimeSegmentWorksheetGenerator(self.store_names, self.test_date)
        
        # Sample test data structure
        self.sample_data = {
            1: {
                '08:00-13:59': {'turnover_current': 0.72, 'turnover_prev': 0.82, 'target': 0.85, 'tables': 32.6, 'customers': 382.9},
                '14:00-16:59': {'turnover_current': 0.46, 'turnover_prev': 0.57, 'target': 0.60, 'tables': 26.3, 'customers': 245.5},
                '17:00-21:59': {'turnover_current': 1.85, 'turnover_prev': 2.14, 'target': 1.97, 'tables': 86.1, 'customers': 980.8},
                '22:00-(次)07:59': {'turnover_current': 0.76, 'turnover_prev': 0.96, 'target': 0.78, 'tables': 43.8, 'customers': 400.7}
            }
        }
    
    def test_initialization(self):
        """Test generator initialization"""
        self.assertEqual(self.generator.store_names, self.store_names)
        self.assertEqual(self.generator.target_date, self.test_date)
        self.assertEqual(len(self.generator.time_segments), 4)
        self.assertIn('08:00-13:59', self.generator.time_segments)
        self.assertIn('22:00-(次)07:59', self.generator.time_segments)
    
    def test_calculate_differences(self):
        """Test difference calculations"""
        current = 1.85
        previous = 2.14
        target = 1.97
        
        target_diff, prev_diff = self.generator.calculate_differences(current, previous, target)
        
        self.assertAlmostEqual(target_diff, -0.12, places=2)
        self.assertAlmostEqual(prev_diff, -0.29, places=2)
    
    def test_calculate_store_totals(self):
        """Test store totals calculation"""
        store_data = self.sample_data[1]
        
        totals = self.generator.calculate_store_totals(store_data)
        
        # Check if all required fields are present
        required_fields = [
            'total_turnover_current', 'total_turnover_prev', 'total_target',
            'total_tables', 'total_customers', 'total_target_diff', 'total_prev_diff'
        ]
        
        for field in required_fields:
            self.assertIn(field, totals)
        
        # Verify calculations
        expected_total_turnover = 0.72 + 0.46 + 1.85 + 0.76
        self.assertAlmostEqual(totals['total_turnover_current'], expected_total_turnover, places=2)
        
        expected_total_target = 0.85 + 0.60 + 1.97 + 0.78
        self.assertAlmostEqual(totals['total_target'], expected_total_target, places=2)
    
    def test_calculate_overall_totals(self):
        """Test overall totals calculation across all stores"""
        # Add data for multiple stores
        test_data = {
            1: {
                '08:00-13:59': {'turnover_current': 1.0, 'turnover_prev': 1.1, 'target': 1.2, 'tables': 10, 'customers': 100}
            },
            2: {
                '08:00-13:59': {'turnover_current': 2.0, 'turnover_prev': 2.1, 'target': 2.2, 'tables': 20, 'customers': 200}
            }
        }
        
        totals = self.generator.calculate_overall_totals(test_data)
        
        # Check if all required fields are present
        required_fields = [
            'overall_turnover_current', 'overall_turnover_prev', 'overall_target',
            'overall_tables', 'overall_customers', 'overall_target_diff', 'overall_prev_diff'
        ]
        
        for field in required_fields:
            self.assertIn(field, totals)
        
        # Verify calculations
        self.assertAlmostEqual(totals['overall_turnover_current'], 3.0, places=1)
        self.assertAlmostEqual(totals['overall_target'], 3.4, places=1)
        self.assertAlmostEqual(totals['overall_tables'], 30, places=1)
        self.assertAlmostEqual(totals['overall_customers'], 300, places=1)
    
    def test_get_time_segment_data_for_date(self):
        """Test time segment data retrieval"""
        data = self.generator.get_time_segment_data_for_date(self.test_date)
        
        # Should return data for all 7 stores
        self.assertEqual(len(data), 7)
        
        # Each store should have 4 time segments
        for store_id in range(1, 8):
            self.assertIn(store_id, data)
            store_data = data[store_id]
            self.assertEqual(len(store_data), 4)
            
            # Check time segments
            for segment in self.generator.time_segments:
                self.assertIn(segment, store_data)
                segment_data = store_data[segment]
                
                # Check required fields
                required_fields = ['turnover_current', 'turnover_prev', 'target', 'tables', 'customers']
                for field in required_fields:
                    self.assertIn(field, segment_data)
                    self.assertIsInstance(segment_data[field], (int, float))
    
    def test_generate_worksheet(self):
        """Test worksheet generation"""
        wb = Workbook()
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)
        
        ws = self.generator.generate_worksheet(wb)
        
        # Verify worksheet was created
        self.assertIsNotNone(ws)
        self.assertEqual(ws.title, "分时段-上报")
        
        # Check if it's added to the workbook
        self.assertIn(ws, wb.worksheets)
        
        # Verify title cell
        title_cell = ws['A1']
        self.assertIsNotNone(title_cell.value)
        self.assertIn("门店分时段营数据", title_cell.value)
        
        # Verify headers
        self.assertEqual(ws['A2'].value, "门店名称")
        self.assertEqual(ws['B2'].value, "分时段")
        self.assertEqual(ws['C2'].value, "翻台率（考核）")
        
        # Verify header row 2
        self.assertEqual(ws['C3'].value, "今年")
        self.assertEqual(ws['D3'].value, "去年")
        self.assertEqual(ws['E3'].value, "本月目标")
        self.assertEqual(ws['F3'].value, "目标差异")
        self.assertEqual(ws['G3'].value, "同比差异")
    
    def test_worksheet_data_structure(self):
        """Test the structure of generated worksheet data"""
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        ws = self.generator.generate_worksheet(wb)
        
        # Calculate expected number of rows
        # 3 header rows + (4 time segments + 1 total) per store * 7 stores + 1 overall total
        expected_min_rows = 3 + (5 * 7) + 1  # 42 rows minimum
        
        # Check that we have data in the expected range
        self.assertGreaterEqual(ws.max_row, expected_min_rows - 5)  # Allow some flexibility
        
        # Check that store names are present
        store_found = False
        for row in range(4, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "加拿大" in str(cell_value):
                store_found = True
                break
        
        self.assertTrue(store_found, "Store names should be present in the worksheet")
        
        # Check that time segments are present
        segment_found = False
        for row in range(4, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=2).value
            if cell_value and ":" in str(cell_value):
                segment_found = True
                break
        
        self.assertTrue(segment_found, "Time segments should be present in the worksheet")
    
    def test_worksheet_formatting(self):
        """Test worksheet formatting"""
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        ws = self.generator.generate_worksheet(wb)
        
        # Test title formatting
        title_cell = ws['A1']
        self.assertIsNotNone(title_cell.font)
        self.assertTrue(title_cell.font.bold)
        
        # Test header formatting
        header_cell = ws['A2']
        self.assertIsNotNone(header_cell.font)
        self.assertTrue(header_cell.font.bold)
        
        # Test cell borders (check a data cell)
        data_cell = ws.cell(row=4, column=3)
        self.assertIsNotNone(data_cell.border)
        
        # Test column widths are set
        self.assertGreater(ws.column_dimensions['A'].width, 10)
        self.assertGreater(ws.column_dimensions['B'].width, 10)
    
    def test_merged_cells(self):
        """Test that cells are merged correctly"""
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)
        
        ws = self.generator.generate_worksheet(wb)
        
        # Check title merge
        merged_ranges = [str(range_obj) for range_obj in ws.merged_cells.ranges]
        self.assertIn('A1:L1', merged_ranges)
        
        # Check header merges
        self.assertIn('A2:A3', merged_ranges)  # 门店名称
        self.assertIn('B2:B3', merged_ranges)  # 分时段
        self.assertIn('C2:G2', merged_ranges)  # 翻台率（考核）

if __name__ == '__main__':
    unittest.main() 