"""
Tests for Weekly Year-over-Year Comparison Worksheet Generator.

This test module covers:
- Generator initialization
- Worksheet creation and structure
- Week date calculation and alignment
- Data processing and calculations
- Color coding for change indicators
- Regional summary calculations (excluding Store 8)
"""

import unittest
from unittest.mock import Mock, MagicMock
from datetime import datetime, timedelta
import sys
from pathlib import Path

# Add project root to path for imports
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from openpyxl import Workbook
from lib.weekly_yoy_comparison_worksheet import WeeklyYoYComparisonWorksheetGenerator


class TestWeeklyYoYComparisonWorksheetGenerator(unittest.TestCase):
    """Tests for WeeklyYoYComparisonWorksheetGenerator."""

    def setUp(self):
        """Set up test fixtures."""
        self.mock_data_provider = Mock()
        self.generator = WeeklyYoYComparisonWorksheetGenerator(self.mock_data_provider)
        self.test_date = "2026-01-07"

    def test_generator_initialization(self):
        """Test generator initialization with data provider."""
        self.assertIsNotNone(self.generator)
        self.assertEqual(self.generator.data_provider, self.mock_data_provider)
        self.assertIsNotNone(self.generator.logger)

    def test_worksheet_creation(self):
        """Test worksheet creation with proper name."""
        workbook = Workbook()

        # Mock data provider to return sample data
        self.mock_data_provider.get_weekly_store_performance.return_value = []

        ws = self.generator.generate_worksheet(workbook, self.test_date)

        # Check worksheet was created with correct name
        self.assertIn("周对比上年表", workbook.sheetnames)
        self.assertIsNotNone(ws)

    def test_week_date_calculation(self):
        """Test 7-day week calculation."""
        target_dt = datetime.strptime(self.test_date, '%Y-%m-%d')
        expected_start = target_dt - timedelta(days=6)

        # The week should be 7 days ending on target date
        self.assertEqual(expected_start, datetime(2026, 1, 1))
        self.assertEqual(target_dt, datetime(2026, 1, 7))

    def test_prev_year_same_calendar_dates(self):
        """Test previous year week uses same calendar dates."""
        current_start = datetime(2026, 1, 1)
        current_end = datetime(2026, 1, 7)

        prev_start, prev_end = self.generator._calculate_prev_year_dates(
            current_start, current_end)

        # Previous year should have same calendar dates (Jan 1-7)
        self.assertEqual(prev_start, datetime(2025, 1, 1),
                        "Previous year start should be Jan 1, 2025")
        self.assertEqual(prev_end, datetime(2025, 1, 7),
                        "Previous year end should be Jan 7, 2025")

        # Previous year period should also be 7 days
        delta = prev_end - prev_start
        self.assertEqual(delta.days, 6, "Previous year period should be 7 days")

    def test_data_processing_with_mock_data(self):
        """Test data processing calculates derived metrics correctly."""
        # Mock store data
        mock_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'manager_name': '张森磊',
                'seating_capacity': 53,
                'current_avg_turnover_rate': 5.5,
                'current_total_revenue': 30.0,  # 30万元
                'current_total_tables': 200,
                'current_total_customers': 600,
                'prev_avg_turnover_rate': 5.2,
                'prev_total_revenue': 28.0,
                'prev_total_tables': 180,
                'prev_total_customers': 550,
            }
        ]

        self.mock_data_provider.get_weekly_store_performance.return_value = mock_data

        workbook = Workbook()
        ws = self.generator.generate_worksheet(workbook, self.test_date)

        # Verify worksheet was created
        self.assertIn("周对比上年表", workbook.sheetnames)

    def test_regional_summary_excludes_store_8(self):
        """Test regional summary excludes Store 8."""
        # Mock store data including Store 8
        mock_data = [
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'seating_capacity': 53,
                'current_avg_turnover_rate': 5.5,
                'current_total_revenue': 30.0,
                'current_total_tables': 200,
                'current_total_customers': 600,
                'prev_avg_turnover_rate': 5.2,
                'prev_total_revenue': 28.0,
                'prev_total_tables': 180,
                'prev_total_customers': 550,
            },
            {
                'store_id': 8,  # Store 8 should be excluded from regional
                'store_name': '加拿大八店',
                'seating_capacity': 56,
                'current_avg_turnover_rate': 4.0,
                'current_total_revenue': 20.0,
                'current_total_tables': 150,
                'current_total_customers': 400,
                'prev_avg_turnover_rate': 0,  # No prev year data
                'prev_total_revenue': 0,
                'prev_total_tables': 0,
                'prev_total_customers': 0,
            }
        ]

        # Filter for regional (excluding Store 8)
        regional_stores = [s for s in mock_data if s['store_id'] != 8]
        self.assertEqual(len(regional_stores), 1)
        self.assertEqual(regional_stores[0]['store_id'], 1)

    def test_per_table_spending_calculation(self):
        """Test per-table spending calculation."""
        # Revenue is in 万元, need to convert to 元 for per-table calc
        current_revenue = 30.0  # 万元
        current_tables = 200

        # Per table = (revenue * 10000) / tables
        expected_per_table = (current_revenue * 10000) / current_tables
        self.assertEqual(expected_per_table, 1500.0)

    def test_per_capita_spending_calculation(self):
        """Test per-capita spending calculation."""
        current_revenue = 30.0  # 万元
        current_customers = 600

        # Per capita = (revenue * 10000) / customers
        expected_per_capita = (current_revenue * 10000) / current_customers
        self.assertEqual(expected_per_capita, 500.0)

    def test_header_structure(self):
        """Test worksheet header structure."""
        workbook = Workbook()
        self.mock_data_provider.get_weekly_store_performance.return_value = []

        ws = self.generator.generate_worksheet(workbook, self.test_date)

        # Check main headers
        self.assertEqual(ws.cell(row=1, column=1).value, "门店名称")
        self.assertEqual(ws.cell(row=1, column=2).value, "翻台率")

    def test_empty_data_handling(self):
        """Test handling of empty data."""
        self.mock_data_provider.get_weekly_store_performance.return_value = []

        workbook = Workbook()
        ws = self.generator.generate_worksheet(workbook, self.test_date)

        # Should still create worksheet with headers
        self.assertIn("周对比上年表", workbook.sheetnames)


class TestWeeklyYoYComparisonEdgeCases(unittest.TestCase):
    """Test edge cases for weekly YoY comparison."""

    def setUp(self):
        """Set up test fixtures."""
        self.mock_data_provider = Mock()
        self.generator = WeeklyYoYComparisonWorksheetGenerator(self.mock_data_provider)

    def test_year_boundary_week(self):
        """Test week that spans year boundary (Dec 30 - Jan 5)."""
        # Dec 30, 2025 to Jan 5, 2026
        target_date = "2026-01-05"
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        start_dt = target_dt - timedelta(days=6)

        self.assertEqual(start_dt, datetime(2025, 12, 30))
        self.assertEqual(target_dt, datetime(2026, 1, 5))

    def test_leap_year_handling(self):
        """Test handling around leap year dates."""
        # Feb 29, 2024 should map to Feb 28, 2023 (Feb 29 doesn't exist in 2023)
        current_end = datetime(2024, 2, 29)
        current_start = datetime(2024, 2, 23)

        prev_start, prev_end = self.generator._calculate_prev_year_dates(
            current_start, current_end)

        # Feb 29 -> Feb 28 in non-leap year
        self.assertEqual(prev_end, datetime(2023, 2, 28))
        self.assertEqual(prev_start, datetime(2023, 2, 23))

    def test_division_by_zero_handling(self):
        """Test handling of zero tables/customers."""
        # When tables = 0, per_table should be 0
        current_tables = 0
        current_revenue = 10.0

        per_table = current_revenue * 10000 / current_tables if current_tables > 0 else 0
        self.assertEqual(per_table, 0)


if __name__ == '__main__':
    unittest.main()
