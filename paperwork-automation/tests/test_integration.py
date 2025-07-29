#!/usr/bin/env python3
"""
Integration tests for the modularized Haidilao paperwork automation system.
Tests end-to-end workflows using the new utility modules.
"""

import unittest
from unittest.mock import Mock, patch, MagicMock
from pathlib import Path
import tempfile
import pandas as pd
from datetime import datetime
import openpyxl

# Add project root to path
import sys
project_root = Path(__file__).parent.parent
sys.path.append(str(project_root))

from lib.excel_utils import (
    safe_read_excel, clean_dish_code, clean_material_number,
    get_material_reading_dtype, validate_required_columns
)
from lib.base_classes import BaseWorksheetGenerator, BaseExtractor
from lib.config import STORE_NAME_MAPPING, TIME_SEGMENTS, MATERIAL_TYPES
from lib.database_utils import DatabaseOperations


class TestExcelProcessingIntegration(unittest.TestCase):
    """Test integrated Excel processing workflows"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.temp_dir = tempfile.mkdtemp()
        self.addCleanup(self._cleanup_temp_dir)
        
    def _cleanup_temp_dir(self):
        """Clean up temporary directory"""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
        
    def create_realistic_material_excel(self):
        """Create realistic material Excel file with precision issues"""
        # Simulate the critical material number precision problem
        test_data = {
            '物料': [
                '000000000001500680',  # String with leading zeros
                '000000000001500681.0',  # String that looks like float
                1500682.0,  # Actual float (problematic!)
                1500683,  # Integer
                '1500684'  # Clean string
            ],
            '物料描述': [
                '锅底调料A',
                '荤菜材料B', 
                '素菜配料C',
                '酒水原料D',
                '调料包E'
            ],
            '数量': [10.5, 20, 15.75, 8, 12],
            '单价': [15.50, 25.00, 18.75, 12.00, 22.50],
            '金额': [162.75, 500.00, 296.25, 96.00, 270.00]
        }
        
        file_path = Path(self.temp_dir) / 'realistic_materials.xlsx'
        df = pd.DataFrame(test_data)
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        return file_path
        
    def create_realistic_dish_excel(self):
        """Create realistic dish Excel file with float conversion issues"""
        test_data = {
            '菜品编号': [
                90001690.0,  # Float that should be string
                90001691.0,
                '90001692',  # Already string
                90001693,    # Integer
                '90001694.0' # String with .0 suffix
            ],
            '菜品名称': [
                '麻辣锅底',
                '清汤锅底',
                '鸳鸯锅底',
                '番茄锅底',
                '菌汤锅底'
            ],
            '尺寸': ['大', '中', '小', '大', '中'],
            '数量': [5, 8, 3, 6, 4],
            '单价': [32.00, 28.00, 25.00, 30.00, 26.00]
        }
        
        file_path = Path(self.temp_dir) / 'realistic_dishes.xlsx'
        df = pd.DataFrame(test_data)
        df.to_excel(file_path, index=False, engine='openpyxl')
        
        return file_path
        
    def test_material_processing_end_to_end(self):
        """Test complete material processing workflow"""
        # Create test file
        material_file = self.create_realistic_material_excel()
        
        # Step 1: Read with proper dtype specification (critical fix)
        dtype_spec = get_material_reading_dtype()
        df = safe_read_excel(material_file, dtype_spec=dtype_spec)
        
        # Step 2: Validate required columns
        required_columns = ['物料', '物料描述', '数量', '单价']
        validate_required_columns(df, required_columns, "materials")
        
        # Step 3: Clean and process data
        df['物料_cleaned'] = df['物料'].apply(clean_material_number)
        df['金额_calculated'] = df['数量'] * df['单价']
        
        # Verify critical fixes
        expected_cleaned = ['1500680', '1500681', '1500682', '1500683', '1500684']
        actual_cleaned = df['物料_cleaned'].tolist()
        
        self.assertEqual(actual_cleaned, expected_cleaned)
        
        # Verify all are strings (not floats)
        for material_num in df['物料_cleaned']:
            self.assertIsInstance(material_num, str)
            
        # Verify calculations are correct
        expected_amounts = [162.75, 500.00, 296.25, 96.00, 270.00]
        calculated_amounts = df['金额_calculated'].tolist()
        
        for expected, calculated in zip(expected_amounts, calculated_amounts):
            self.assertAlmostEqual(calculated, expected, places=2)
            
    def test_dish_processing_end_to_end(self):
        """Test complete dish processing workflow"""
        # Create test file
        dish_file = self.create_realistic_dish_excel()
        
        # Step 1: Read Excel file
        df = safe_read_excel(dish_file)
        
        # Step 2: Validate structure
        required_columns = ['菜品编号', '菜品名称', '数量', '单价']
        validate_required_columns(df, required_columns, "dishes")
        
        # Step 3: Clean dish codes
        df['菜品编号_cleaned'] = df['菜品编号'].apply(clean_dish_code)
        
        # Verify cleaning results
        expected_codes = ['90001690', '90001691', '90001692', '90001693', '90001694']
        actual_codes = df['菜品编号_cleaned'].tolist()
        
        self.assertEqual(actual_codes, expected_codes)
        
        # Verify all are strings without .0 suffix
        for code in df['菜品编号_cleaned']:
            self.assertIsInstance(code, str)
            self.assertNotIn('.0', code)


class TestWorksheetGeneratorIntegration(unittest.TestCase):
    """Test integrated worksheet generation workflows"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.store_names = ['加拿大一店', '加拿大二店', '加拿大三店']
        self.target_date = '2025-06-15'
        
    def test_worksheet_generator_inheritance_integration(self):
        """Test worksheet generator using base class functionality"""
        
        # Create concrete implementation for testing
        class TestWorksheetGenerator(BaseWorksheetGenerator):
            def generate_worksheet(self, workbook, test_data):
                ws = workbook.create_sheet("测试工作表")
                
                # Use parent methods for common operations
                self.set_column_widths(ws, [15, 12, 12, 12])
                
                # Add title using parent method
                current_row = self.add_title_section(
                    ws, f"测试报表 - {self.target_dt.strftime('%Y年%m月%d日')}", 1, 4
                )
                
                # Add headers with parent styling
                headers = ['门店', '收入', '增长率', '排名']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    self.apply_header_style(cell)
                    
                current_row += 1
                
                # Add data with calculations
                for i, store_name in enumerate(self.store_names):
                    current_revenue = test_data['current'][i]
                    previous_revenue = test_data['previous'][i]
                    
                    # Use parent calculation methods
                    growth_rate = self.calculate_percentage_change(current_revenue, previous_revenue)
                    formatted_growth = self.format_percentage(growth_rate)
                    
                    # Add row data
                    row_data = [store_name, current_revenue, formatted_growth, i + 1]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        
                        if col == 1:  # Store name
                            self.apply_data_style(cell, align='left')
                        elif col in [2, 4]:  # Numbers
                            self.apply_data_style(cell, align='right')
                        else:  # Percentage
                            self.apply_data_style(cell, align='center')
                            
                    current_row += 1
                    
                return ws
                
        # Test the generator
        generator = TestWorksheetGenerator(self.store_names, self.target_date)
        
        # Verify base class setup
        self.assertEqual(generator.target_dt.year, 2025)
        self.assertEqual(generator.target_dt.month, 6)
        self.assertEqual(generator.target_dt.day, 15)
        
        # Test with realistic data
        test_data = {
            'current': [15000, 18000, 12000],
            'previous': [12000, 15000, 13000]
        }
        
        # Create workbook and generate worksheet
        workbook = openpyxl.Workbook()
        worksheet = generator.generate_worksheet(workbook, test_data)
        
        # Verify worksheet was created
        self.assertEqual(worksheet.title, "测试工作表")
        
        # Verify data calculations
        self.assertAlmostEqual(generator.calculate_percentage_change(15000, 12000), 0.25)
        self.assertEqual(generator.format_percentage(0.25), "25.0%")
        
        # Verify worksheet contains expected data
        self.assertEqual(worksheet.cell(row=3, column=1).value, '加拿大一店')
        self.assertEqual(worksheet.cell(row=3, column=3).value, '25.0%')


class TestExtractorIntegration(unittest.TestCase):
    """Test integrated data extraction workflows"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.temp_dir = tempfile.mkdtemp()
        self.addCleanup(self._cleanup_temp_dir)
        
    def _cleanup_temp_dir(self):
        """Clean up temporary directory"""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
        
    def test_extractor_base_functionality(self):
        """Test extractor using base class functionality"""
        
        # Create concrete implementation
        class TestExtractor(BaseExtractor):
            def extract_data(self, input_file, **kwargs):
                # Use parent validation
                if not self.validate_file_existence(input_file):
                    return {'error': 'File not found'}
                    
                # Use centralized Excel reading
                from lib.excel_utils import safe_read_excel, get_material_reading_dtype
                
                try:
                    dtype_spec = get_material_reading_dtype()
                    df = safe_read_excel(input_file, dtype_spec=dtype_spec)
                    
                    # Process data using store mapping
                    store_mapping = self.get_store_id_mapping()
                    
                    extracted_data = []
                    for _, row in df.iterrows():
                        record = {
                            'material_number': clean_material_number(row['物料']),
                            'material_name': row['物料描述'],
                            'quantity': row['数量'],
                            'price': row['单价']
                        }
                        extracted_data.append(record)
                        
                    # Log summary using parent method
                    self.log_extraction_summary({
                        'Materials': len(extracted_data)
                    })
                    
                    return {
                        'success': True,
                        'data': extracted_data,
                        'count': len(extracted_data)
                    }
                    
                except Exception as e:
                    return {'error': str(e)}
                    
        # Create test file
        material_data = {
            '物料': ['000000000001500680', '000000000001500681'],
            '物料描述': ['锅底材料', '荤菜材料'],
            '数量': [10, 20],
            '单价': [15.50, 25.00]
        }
        
        test_file = Path(self.temp_dir) / 'test_materials.xlsx'
        pd.DataFrame(material_data).to_excel(test_file, index=False)
        
        # Test extraction
        extractor = TestExtractor()
        result = extractor.extract_data(test_file)
        
        # Verify results
        self.assertTrue(result['success'])
        self.assertEqual(result['count'], 2)
        
        # Verify data processing
        extracted_materials = result['data']
        self.assertEqual(extracted_materials[0]['material_number'], '1500680')
        self.assertEqual(extracted_materials[1]['material_number'], '1500681')
        
        # Verify store mapping access
        store_mapping = extractor.get_store_id_mapping()
        self.assertEqual(len(store_mapping), 7)
        self.assertEqual(store_mapping['加拿大一店'], 1)


class TestDatabaseIntegration(unittest.TestCase):
    """Test integrated database operations"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.mock_db_manager = Mock()
        self.mock_connection = Mock()
        self.mock_cursor = Mock()
        
        # Setup mock chain
        self.mock_db_manager.get_connection.return_value.__enter__.return_value = self.mock_connection
        self.mock_db_manager.get_connection.return_value.__exit__.return_value = None
        self.mock_connection.cursor.return_value = self.mock_cursor
        
    def test_complete_material_update_workflow(self):
        """Test complete material price update workflow"""
        db_ops = DatabaseOperations(self.mock_db_manager)
        
        # Simulate material price update workflow
        material_updates = [
            {
                'material_id': 1500680,
                'store_id': 1,
                'price': 15.50,
                'effective_date': '2025-06-15'
            },
            {
                'material_id': 1500681,
                'store_id': 1,
                'price': 25.00,
                'effective_date': '2025-06-15'
            }
        ]
        
        # Step 1: Deactivate previous prices
        self.mock_cursor.rowcount = 2
        deactivated_count = 0
        
        for update in material_updates:
            count = db_ops.deactivate_previous_records(
                'material_price_history',
                {'material_id': update['material_id'], 'store_id': update['store_id']}
            )
            deactivated_count += count
            
        self.assertEqual(deactivated_count, 4)  # 2 materials × 2 old prices each
        
        # Step 2: Batch insert new prices
        processed_count = db_ops.batch_upsert(
            'material_price_history',
            material_updates,
            ['material_id', 'store_id', 'effective_date'],
            batch_size=10
        )
        
        self.assertEqual(processed_count, 2)
        
        # Verify database operations were called
        self.assertGreater(self.mock_cursor.execute.call_count, 2)
        self.assertGreater(self.mock_connection.commit.call_count, 2)
        
    def test_store_data_validation_integration(self):
        """Test integrated store data validation"""
        db_ops = DatabaseOperations(self.mock_db_manager)
        
        # Mock store data for validation
        all_stores = list(STORE_NAME_MAPPING.keys())
        all_store_ids = list(STORE_NAME_MAPPING.values())
        
        # Simulate some stores missing data
        present_stores = [1, 2, 3, 5, 7]  # Missing stores 4 and 6
        
        mock_df = pd.DataFrame({'store_id': present_stores})
        db_ops.execute_query_to_dataframe = Mock(return_value=mock_df)
        
        # Validate completeness
        is_complete, missing_stores = db_ops.validate_data_completeness(
            'daily_report', all_store_ids, '2025-06-15'
        )
        
        self.assertFalse(is_complete)
        self.assertEqual(set(missing_stores), {4, 6})
        
        # Get store summary for detailed analysis
        store_summary = {
            '加拿大一店': 1, '加拿大二店': 1, '加拿大三店': 1,
            '加拿大四店': 0, '加拿大五店': 1, '加拿大六店': 0, '加拿大七店': 1
        }
        
        db_ops.get_store_data_summary = Mock(return_value=store_summary)
        summary = db_ops.get_store_data_summary('2025-06-15')
        
        # Verify missing stores
        missing_store_names = [name for name, count in summary.items() if count == 0]
        expected_missing = ['加拿大四店', '加拿大六店']
        
        self.assertEqual(set(missing_store_names), set(expected_missing))


class TestConfigurationIntegration(unittest.TestCase):
    """Test integration between configuration modules"""
    
    def test_store_and_time_segment_integration(self):
        """Test store and time segment configuration integration"""
        # Verify store configuration consistency
        self.assertEqual(len(STORE_NAME_MAPPING), 7)
        
        # Verify all store names are in Chinese
        for store_name in STORE_NAME_MAPPING.keys():
            self.assertTrue(store_name.startswith('加拿大'))
            self.assertTrue(store_name.endswith('店'))
            
        # Verify time segments are complete
        self.assertEqual(len(TIME_SEGMENTS), 4)
        
        # Test realistic workflow: process all stores for all time segments
        total_expected_records = len(STORE_NAME_MAPPING) * len(TIME_SEGMENTS)
        self.assertEqual(total_expected_records, 28)  # 7 stores × 4 time segments
        
    def test_material_type_integration(self):
        """Test material type configuration integration"""
        # Verify material types structure
        self.assertEqual(len(MATERIAL_TYPES), 11)
        
        # Verify all material types start with "成本-"
        for material_type in MATERIAL_TYPES.values():
            self.assertTrue(material_type.startswith('成本-'))
            
        # Test material type workflow
        material_categories = set()
        for type_name in MATERIAL_TYPES.values():
            category = type_name.split('-')[1]  # Extract category after "成本-"
            material_categories.add(category)
            
        # Should have meaningful variety of categories
        self.assertGreater(len(material_categories), 5)
        self.assertIn('锅底类', material_categories)
        self.assertIn('荤菜类', material_categories)


class TestFullWorkflowIntegration(unittest.TestCase):
    """Test complete end-to-end workflow integration"""
    
    def setUp(self):
        """Set up test fixtures"""
        self.temp_dir = tempfile.mkdtemp()
        self.addCleanup(self._cleanup_temp_dir)
        
    def _cleanup_temp_dir(self):
        """Clean up temporary directory"""
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
        
    def test_complete_daily_report_workflow(self):
        """Test complete daily report processing workflow"""
        
        # Step 1: Create realistic daily report Excel file
        daily_data = {
            '门店名称': ['加拿大一店', '加拿大二店', '加拿大三店'],
            '日期': [20250615, 20250615, 20250615],
            '节假日': ['工作日', '工作日', '工作日'],
            '营业桌数': [45, 52, 38],
            '营业桌数(考核)': [40, 50, 35],
            '翻台率(考核)': [2.8, 3.2, 2.5],
            '营业收入(不含税)': [15000.50, 18500.75, 12800.25],
            '就餐人数': [180, 208, 152],
            '优惠总金额(不含税)': [750.25, 925.38, 640.15]
        }
        
        daily_file = Path(self.temp_dir) / 'daily_report.xlsx'
        pd.DataFrame(daily_data).to_excel(daily_file, index=False, sheet_name='营业基础表')
        
        # Step 2: Read and validate using centralized utilities
        df = safe_read_excel(daily_file, sheet_name='营业基础表')
        
        required_columns = [
            '门店名称', '日期', '营业桌数', '营业收入(不含税)', '翻台率(考核)'
        ]
        validate_required_columns(df, required_columns, "daily report")
        
        # Step 3: Process data using configuration
        processed_records = []
        store_mapping = STORE_NAME_MAPPING
        
        for _, row in df.iterrows():
            store_id = store_mapping.get(row['门店名称'])
            if store_id:
                record = {
                    'store_id': store_id,
                    'report_date': '2025-06-15',
                    'table_count': int(row['营业桌数']),
                    'revenue': float(row['营业收入(不含税)']),
                    'turnover_rate': float(row['翻台率(考核)']),
                    'customer_count': int(row['就餐人数']),
                    'discount_amount': float(row['优惠总金额(不含税)'])
                }
                processed_records.append(record)
                
        # Verify processing results
        self.assertEqual(len(processed_records), 3)
        self.assertEqual(processed_records[0]['store_id'], 1)  # 加拿大一店
        self.assertEqual(processed_records[1]['store_id'], 2)  # 加拿大二店
        self.assertEqual(processed_records[2]['store_id'], 3)  # 加拿大三店
        
        # Step 4: Simulate database operations
        mock_db_manager = Mock()
        mock_connection = Mock()
        mock_db_manager.get_connection.return_value.__enter__.return_value = mock_connection
        mock_db_manager.get_connection.return_value.__exit__.return_value = None
        
        db_ops = DatabaseOperations(mock_db_manager)
        
        # Simulate batch insert
        insert_count = db_ops.batch_upsert(
            'daily_report',
            processed_records,
            ['store_id', 'report_date'],
            batch_size=10
        )
        
        self.assertEqual(insert_count, 3)
        
        # Step 5: Generate worksheet using base class
        class DailyReportWorksheet(BaseWorksheetGenerator):
            def generate_worksheet(self, workbook, data):
                ws = workbook.create_sheet("日报汇总")
                
                # Add title
                current_row = self.add_title_section(
                    ws, f"日报汇总 - {self.target_dt.strftime('%Y年%m月%d日')}", 1, 5
                )
                
                # Add headers
                headers = ['门店', '桌数', '收入', '翻台率', '人数']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    self.apply_header_style(cell)
                    
                # Add data
                for i, record in enumerate(data, current_row + 1):
                    store_name = [name for name, sid in STORE_NAME_MAPPING.items() 
                                if sid == record['store_id']][0]
                                
                    row_data = [
                        store_name,
                        record['table_count'],
                        record['revenue'],
                        record['turnover_rate'],
                        record['customer_count']
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=i, column=col, value=value)
                        self.apply_data_style(cell)
                        
                return ws
                
        # Generate worksheet
        worksheet_generator = DailyReportWorksheet(
            list(store_mapping.keys())[:3], '2025-06-15'
        )
        
        workbook = openpyxl.Workbook()
        worksheet = worksheet_generator.generate_worksheet(workbook, processed_records)
        
        # Verify worksheet generation
        self.assertEqual(worksheet.title, "日报汇总")
        self.assertEqual(worksheet.cell(row=3, column=1).value, '加拿大一店')
        self.assertEqual(worksheet.cell(row=3, column=2).value, 45)  # Table count
        
        # Test calculations
        total_revenue = sum(record['revenue'] for record in processed_records)
        self.assertAlmostEqual(total_revenue, 46301.50, places=2)


if __name__ == '__main__':
    # Run all tests with detailed output
    unittest.main(verbosity=2)