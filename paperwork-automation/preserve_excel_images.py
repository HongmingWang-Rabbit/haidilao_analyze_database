#!/usr/bin/env python3
"""
Excel Image Preservation Utility
Manually preserves images and drawings when modifying Excel files with openpyxl
"""

import zipfile
import os
import tempfile
import shutil
from pathlib import Path

class ExcelImagePreserver:
    """Utility to preserve images in Excel files when using openpyxl"""
    
    def __init__(self, source_file):
        self.source_file = Path(source_file)
        self.media_files = {}
        self.drawing_files = {}
        
    def extract_media_and_drawings(self):
        """Extract media and drawing files from the source Excel file"""
        try:
            with zipfile.ZipFile(self.source_file, 'r') as zip_file:
                file_list = zip_file.namelist()
                
                # Extract media files
                for file_name in file_list:
                    if 'media/' in file_name.lower():
                        self.media_files[file_name] = zip_file.read(file_name)
                    elif 'drawings/' in file_name.lower():
                        self.drawing_files[file_name] = zip_file.read(file_name)
                        
                print(f"Extracted {len(self.media_files)} media files and {len(self.drawing_files)} drawing files")
                
        except Exception as e:
            print(f"Error extracting media/drawings: {e}")
            
    def inject_media_and_drawings(self, target_file):
        """Inject the preserved media and drawing files into the target Excel file"""
        if not self.media_files and not self.drawing_files:
            print("No media or drawings to inject")
            return False
            
        try:
            # Create a temporary file for the modified Excel
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                temp_path = temp_file.name
            
            # Copy the target file to temp
            shutil.copy2(target_file, temp_path)
            
            # Open the Excel file as a ZIP and add media/drawings
            with zipfile.ZipFile(temp_path, 'a') as zip_file:
                # Add media files
                for file_name, file_data in self.media_files.items():
                    zip_file.writestr(file_name, file_data)
                    
                # Add drawing files  
                for file_name, file_data in self.drawing_files.items():
                    zip_file.writestr(file_name, file_data)
            
            # Replace the original target file
            shutil.move(temp_path, target_file)
            
            print(f"Successfully injected {len(self.media_files)} media files and {len(self.drawing_files)} drawing files")
            return True
            
        except Exception as e:
            print(f"Error injecting media/drawings: {e}")
            # Clean up temp file if it exists
            if os.path.exists(temp_path):
                os.unlink(temp_path)
            return False

def process_excel_with_image_preservation(source_file, target_file, modification_func):
    """
    Process an Excel file while preserving images and drawings
    
    Args:
        source_file: Path to source Excel file
        target_file: Path to target Excel file
        modification_func: Function that takes a workbook and modifies it
    """
    # Step 1: Extract media and drawings from source
    preserver = ExcelImagePreserver(source_file)
    preserver.extract_media_and_drawings()
    
    # Step 2: Copy source to target
    shutil.copy2(source_file, target_file)
    
    # Step 3: Load target file with openpyxl and make modifications
    from openpyxl import load_workbook
    wb = load_workbook(target_file)
    
    # Apply the user's modifications
    modification_func(wb)
    
    # Save the modified file (this will lose images)
    wb.save(target_file)
    wb.close()
    
    # Step 4: Re-inject the preserved images and drawings
    success = preserver.inject_media_and_drawings(target_file)
    
    return success

# Test the utility
if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    
    template_path = 'Input/daily_report/bank_transactions_reports/CA全部7家店明细.xlsx'
    test_output = 'test_image_preservation_advanced.xlsx'
    
    def test_modification(workbook):
        """Test modification function"""
        ws = workbook.active
        ws['A1'] = "Test modification"
        
        # Add a new row of data to simulate bank transaction processing
        ws.append(["2025-07-29", "Test transaction", 100.00, "", "", "Test classification"])
    
    print("Testing advanced image preservation...")
    
    success = process_excel_with_image_preservation(
        template_path, 
        test_output, 
        test_modification
    )
    
    if success:
        # Verify the result
        def check_images(file_path, label):
            try:
                with zipfile.ZipFile(file_path, 'r') as zip_file:
                    media_count = len([f for f in zip_file.namelist() if 'media/' in f.lower()])
                    drawing_count = len([f for f in zip_file.namelist() if 'drawings/' in f.lower()])
                    print(f"{label}: {media_count} media, {drawing_count} drawings")
                    return media_count, drawing_count
            except:
                return 0, 0
        
        orig_media, orig_drawings = check_images(template_path, "Original")
        test_media, test_drawings = check_images(test_output, "Generated")
        
        if orig_media == test_media and orig_drawings == test_drawings:
            print("✅ SUCCESS: Advanced preservation worked!")
        else:
            print("❌ FAILED: Advanced preservation didn't work")
    
    # Clean up
    if os.path.exists(test_output):
        os.remove(test_output)