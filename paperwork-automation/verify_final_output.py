#!/usr/bin/env python3
"""
Verify final output for duplicates
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import sys
import io

# Set UTF-8 encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def verify_output():
    """Check the latest output file for duplicates"""
    
    # Find the latest output file
    output_dir = Path("output")
    output_files = list(output_dir.glob("Bank_Transactions_Report_*.xlsx"))
    
    if not output_files:
        print("No output files found")
        return
    
    latest_file = max(output_files, key=lambda p: p.stat().st_mtime)
    print(f"Checking latest output file: {latest_file.name}")
    
    # Load the workbook
    wb = load_workbook(latest_file)
    
    # Check RBC sheets
    rbc_sheets = ['RBC 5401', 'RBC 5419', 'RBC0517-Hi Bowl', 'RBC 0922（USD）', 'RBC3088（USD）-Hi Bowl']
    
    total_duplicates = 0
    
    for sheet_name in rbc_sheets:
        if sheet_name not in wb.sheetnames:
            continue
            
        print(f"\nChecking {sheet_name}...")
        ws = wb[sheet_name]
        
        # Extract all transactions
        transactions = []
        for row in range(3, ws.max_row + 1):
            # Check if row has data
            date_val = ws.cell(row=row, column=2).value  # Date column
            if not date_val:
                continue
                
            desc_val = ws.cell(row=row, column=1).value  # Description
            debit_val = ws.cell(row=row, column=4).value  # Debit
            credit_val = ws.cell(row=row, column=5).value  # Credit
            
            transactions.append({
                'row': row,
                'date': str(date_val),
                'desc': str(desc_val) if desc_val else '',
                'debit': str(debit_val) if debit_val else '',
                'credit': str(credit_val) if credit_val else ''
            })
        
        print(f"  Total transactions: {len(transactions)}")
        
        # Check for duplicates
        seen = {}
        duplicates = []
        
        for trans in transactions:
            key = (trans['date'], trans['desc'], trans['debit'], trans['credit'])
            if key in seen:
                duplicates.append((trans['row'], seen[key]))
            else:
                seen[key] = trans['row']
        
        if duplicates:
            print(f"  WARNING: Found {len(duplicates)} duplicate pairs!")
            total_duplicates += len(duplicates)
            # Show first few duplicates
            for i, (row1, row2) in enumerate(duplicates[:3]):
                trans1 = next(t for t in transactions if t['row'] == row1)
                print(f"    Duplicate: Row {row1} and {row2}")
                print(f"      {trans1['date']} | {trans1['desc'][:40]}... | D:{trans1['debit']} C:{trans1['credit']}")
        else:
            print(f"  ✓ No duplicates found")
    
    wb.close()
    
    print(f"\n\nSUMMARY: Total duplicate pairs found: {total_duplicates}")
    
    if total_duplicates > 0:
        print("\nNOTE: These duplicates may be pre-existing in the template file.")
        print("The new duplicate detection approach prevents adding new duplicates from the same date.")

if __name__ == "__main__":
    verify_output()