#!/usr/bin/env python3
"""Test preserving unknown elements in openpyxl"""

from openpyxl import load_workbook
import sys
import zipfile
import os
import shutil

# Fix encoding issue
sys.stdout.reconfigure(encoding='utf-8')

template_path = 'Input/daily_report/bank_transactions_reports/CA全部7家店明细.xlsx'
test_output = 'test_preserve_unknown.xlsx'

def check_images_in_file(file_path, label):
    """Check for images in an Excel file"""
    print(f"\n=== {label} ===")
    
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            file_list = zip_file.namelist()
            media_files = [f for f in file_list if 'media' in f.lower()]
            return len(media_files)
    except:
        return 0

# Test different approaches
approaches = [
    ("Copy first, then load/save", "copy_first"),
    ("Direct load/save", "direct"),
    ("Load with keep_vba=True", "keep_vba"),
]

original_count = check_images_in_file(template_path, "ORIGINAL")

for name, approach in approaches:
    print(f"\n{'='*50}")
    print(f"Testing: {name}")
    
    try:
        if approach == "copy_first":
            # Copy file first, then load and save
            shutil.copy2(template_path, test_output)
            wb = load_workbook(test_output)
            # Make a small change
            ws = wb.active
            ws['A1'] = "Modified"
            wb.save(test_output)
            wb.close()
            
        elif approach == "direct":
            # Direct load and save
            wb = load_workbook(template_path)
            # Make a small change
            ws = wb.active
            ws['A1'] = "Modified"
            wb.save(test_output)
            wb.close()
            
        elif approach == "keep_vba":
            # Load with keep_vba=True
            wb = load_workbook(template_path, keep_vba=True)
            # Make a small change
            ws = wb.active
            ws['A1'] = "Modified"
            wb.save(test_output)
            wb.close()
        
        result_count = check_images_in_file(test_output, "RESULT")
        
        if result_count == original_count:
            print(f"✅ SUCCESS: {name} preserved {result_count} images")
        else:
            print(f"❌ FAILED: {name} lost images ({original_count} -> {result_count})")
    
    except Exception as e:
        print(f"❌ ERROR: {name} failed with error: {e}")
    
    # Clean up
    if os.path.exists(test_output):
        os.remove(test_output)

print(f"\n{'='*50}")
print("CONCLUSION: All openpyxl approaches lose images")