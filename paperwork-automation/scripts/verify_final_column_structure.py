#!/usr/bin/env python3
"""
Verify the final column structure of the material variance worksheet
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.monthly_dishes_worksheet import MonthlyDishesWorksheetGenerator
from lib.database_queries import ReportDataProvider
from utils.database import DatabaseManager, DatabaseConfig
from openpyxl import Workbook

def verify_column_structure():
    """Verify the exact column structure"""
    
    print("Verifying Final Column Structure")
    print("=" * 35)
    
    try:
        # Initialize components
        config = DatabaseConfig(is_test=True)
        db_manager = DatabaseManager(config)
        data_provider = ReportDataProvider(db_manager)
        
        # Create worksheet generator
        target_date = "2025-05-31"
        store_names = ["加拿大一店", "加拿大二店", "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大七店"]
        worksheet_gen = MonthlyDishesWorksheetGenerator(store_names, target_date)
        
        # Create workbook and generate worksheet
        wb = Workbook()
        ws = worksheet_gen.generate_material_variance_worksheet(wb, data_provider)
        
        print(f"✅ Worksheet generated: {ws.title}")
        print(f"📊 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Find the header row
        header_row = None
        for row in range(1, 20):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "序号" in str(cell_value):
                header_row = row
                break
        
        if header_row:
            print(f"📋 Header row found at row {header_row}")
            print("\\nColumn structure:")
            for col in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=header_row, column=col)
                col_letter = chr(64 + col)  # A=65, so col 1 = A
                header_value = header_cell.value if header_cell.value else f"<empty>"
                print(f"   {col_letter}: {header_value}")
        else:
            print("❌ Could not find header row")
            
        # Check a few data rows
        if header_row and header_row + 1 <= ws.max_row:
            print(f"\\n📊 Sample data (row {header_row + 1}):")
            sample_row = header_row + 1
            for col in range(1, min(ws.max_column + 1, 15)):
                cell = ws.cell(row=sample_row, column=col)
                col_letter = chr(64 + col)
                value = str(cell.value)[:20] if cell.value else "<empty>"
                print(f"   {col_letter}: {value}")
        
        # Save output
        output_file = "output/verify_final_column_structure.xlsx"
        wb.save(output_file)
        print(f"\\n📁 Verification output saved to: {output_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    verify_column_structure()