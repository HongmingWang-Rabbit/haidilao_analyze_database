#!/usr/bin/env python3
"""
Test script to verify store-specific schema changes work correctly
"""

import sys
import os
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.monthly_dishes_worksheet import MonthlyDishesWorksheetGenerator
from lib.database_queries import ReportDataProvider
from utils.database import DatabaseManager, DatabaseConfig
from openpyxl import Workbook

def test_store_specific_schema():
    """Test that store-specific schema changes work correctly"""
    
    print("🧪 Testing Store-Specific Schema Changes")
    print("=" * 45)
    
    try:
        # Initialize components with test database
        config = DatabaseConfig(is_test=True)
        db_manager = DatabaseManager(config)
        data_provider = ReportDataProvider(db_manager)
        
        # Test parameters
        target_date = "2025-05-31"
        store_names = ["加拿大一店", "加拿大二店"]
        worksheet_gen = MonthlyDishesWorksheetGenerator(store_names, target_date)
        
        print(f"✅ Initialized components successfully")
        print(f"   Target Date: {target_date}")
        print(f"   Stores: {store_names}")
        
        # Test 1: Check if material variance data can be retrieved
        print(f"\n🔍 Test 1: Material Variance Data Retrieval")
        try:
            variance_data = worksheet_gen.get_material_variance_data(data_provider, 2025, 5)
            if variance_data:
                print(f"   ✅ Successfully retrieved {len(variance_data)} variance records")
                
                # Show first few records
                for i, record in enumerate(variance_data[:3]):
                    print(f"   📊 Record {i+1}: Store {record.get('store_name', 'Unknown')} | "
                          f"Material {record.get('material_name', 'Unknown')[:20]}...")
            else:
                print(f"   ⚠️  No variance data found (expected if no test data exists)")
                
        except Exception as e:
            print(f"   ❌ Error retrieving variance data: {e}")
            
        # Test 2: Check dish usage details functionality
        print(f"\n🔍 Test 2: Dish Usage Details (Store-Specific)")
        try:
            # Test with a common material (if exists in test data)
            test_material_id = 1
            test_store_id = 1
            
            usage_details = worksheet_gen.get_dish_usage_details(
                data_provider, test_material_id, test_store_id, 2025, 5
            )
            
            if usage_details and usage_details != "无使用记录":
                print(f"   ✅ Successfully retrieved dish usage details")
                print(f"   📝 Details preview: {usage_details[:100]}...")
            else:
                print(f"   ⚠️  No dish usage found (expected if no test data exists)")
                
        except Exception as e:
            print(f"   ❌ Error retrieving dish usage details: {e}")
            
        # Test 3: Try generating a material variance worksheet
        print(f"\n🔍 Test 3: Material Variance Worksheet Generation")
        try:
            wb = Workbook()
            ws = worksheet_gen.generate_material_variance_worksheet(wb, data_provider)
            
            print(f"   ✅ Successfully generated worksheet: {ws.title}")
            print(f"   📏 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
            
            # Check if headers are present
            header_found = False
            for row in range(1, 20):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and "序号" in str(cell_value):
                    header_found = True
                    print(f"   📋 Headers found at row {row}")
                    break
            
            if not header_found:
                print(f"   ⚠️  No data headers found (expected if no test data)")
                
        except Exception as e:
            print(f"   ❌ Error generating worksheet: {e}")
            import traceback
            traceback.print_exc()
            
        # Test 4: Database schema compatibility
        print(f"\n🔍 Test 4: Database Schema Compatibility")
        try:
            with db_manager.get_connection() as conn:
                cursor = conn.cursor()
                
                # Test if store_id columns exist in key tables
                schema_tests = [
                    ("dish", "store_id"),
                    ("material", "store_id"), 
                    ("dish_material", "store_id")
                ]
                
                for table_name, column_name in schema_tests:
                    cursor.execute(f"""
                        SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_name = %s AND column_name = %s
                    """, (table_name, column_name))
                    
                    result = cursor.fetchone()
                    if result:
                        print(f"   ✅ Table {table_name} has {column_name} column")
                    else:
                        print(f"   ❌ Table {table_name} missing {column_name} column")
                        
        except Exception as e:
            print(f"   ❌ Error checking database schema: {e}")
            
        print(f"\n🎉 Store-specific schema testing completed!")
        print(f"   Note: Some tests may show warnings if test data is not available")
        return True
        
    except Exception as e:
        print(f"❌ Critical error during testing: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_store_specific_schema()