#!/usr/bin/env python3
"""
Generate Gross Revenue Report for Dishes with Historical Comparisons
Calculates profitability for each dish by store including:
- Dish sale revenue (quantity * price)
- Material used cost (materials used * material price)  
- Net revenue (revenue - cost)
- Revenue percentage (profit margin)
- Historical price comparisons (last month and last year)
"""

import argparse
import sys
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta

# Add parent directory to path for imports - MUST come before local imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from utils.database import DatabaseManager, DatabaseConfig
from lib.config import STORE_ID_TO_NAME_MAPPING
from lib.dish_material_db import DishMaterialDatabase

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class GrossRevenueReportGenerator:
    """Generate gross revenue report for dishes by store."""
    
    def __init__(self, db_manager: DatabaseManager):
        """Initialize the report generator."""
        self.db_manager = db_manager
        self.dish_material_db = None
        
    def generate_report(self, year: int, month: int, output_path: str = None, include_history: bool = True, include_profit_summary: bool = True) -> str:
        """
        Generate the gross revenue report for all stores.
        
        Args:
            year: Target year
            month: Target month
            output_path: Optional output file path
            include_history: Whether to include historical price comparisons
            
        Returns:
            Path to generated report file
        """
        logger.info(f"Generating gross revenue report for {year}-{month:02d}")
        
        # Initialize database module with connection
        with self.db_manager.get_connection() as conn:
            self.dish_material_db = DishMaterialDatabase(conn)
        
        # Calculate comparison periods if including history
        last_month_date = None
        last_year_date = None
        if include_history:
            current_date = datetime(year, month, 1)
            last_month_date = current_date - relativedelta(months=1)
            last_year_date = current_date - relativedelta(years=1)
        
        # Create workbook
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        # Generate sheets for each store
        store_count = 0
        has_data = False
        
        for store_id, store_name in STORE_ID_TO_NAME_MAPPING.items():
            if store_id > 100:  # Skip Hi Bowl for now
                continue
                
            logger.info(f"Processing store {store_id}: {store_name}")
            
            # Get data for this store
            if include_history:
                store_data = self._get_store_dish_revenue_data_with_history(
                    store_id, year, month,
                    last_month_date.year, last_month_date.month,
                    last_year_date.year, last_year_date.month
                )
            else:
                store_data = self._get_store_dish_revenue_data(store_id, year, month)
            
            if not store_data:
                logger.warning(f"No data for store {store_id} in {year}-{month:02d}")
                continue
            
            # Create worksheet for this store
            ws = wb.create_sheet(title=f"{store_id}-{store_name}")
            if include_history:
                self._populate_store_worksheet_with_history(
                    ws, store_data, store_name, year, month, last_month_date, last_year_date
                )
            else:
                self._populate_store_worksheet(ws, store_data, store_name, year, month)
            
            # Add Theory vs Actual Usage sheet for this store
            theory_actual_ws = wb.create_sheet(title=f"{store_id}-理论vs实际")
            self._create_theory_vs_actual_sheet(theory_actual_ws, store_id, store_name, year, month)
            
            store_count += 1
            has_data = True
        
        # If no data found, create a summary sheet
        if not has_data:
            ws = wb.create_sheet(title="No Data")
            ws['A1'] = f"No data found for {year}-{month:02d}"
            ws['A3'] = "Possible reasons:"
            ws['A4'] = "1. No dish sales data has been extracted for this month"
            ws['A5'] = "2. Run extract_dishes_to_database.py first to populate dish_monthly_sale table"
            ws['A6'] = "3. Ensure material prices and dish-material mappings are set up"
            logger.warning(f"No data found for any store in {year}-{month:02d}")
        
        # Add profit summary sheet if requested and has data
        if include_profit_summary and has_data:
            logger.info("Generating profit summary sheet")
            self._create_profit_summary_sheet(wb, year, month)
        
        # Save the report
        if output_path:
            output_file = Path(output_path)
        else:
            output_dir = Path("Output/gross_revenue_reports")
            output_dir.mkdir(parents=True, exist_ok=True)
            output_file = output_dir / f"gross_revenue_report_{year}_{month:02d}.xlsx"
        
        wb.save(output_file)
        logger.info(f"Report saved to: {output_file}")
        logger.info(f"Generated {store_count} store sheets")
        
        return str(output_file)
    
    def _get_store_dish_revenue_data(self, store_id: int, year: int, month: int) -> List[Dict]:
        """
        Get dish revenue data for a specific store and month.
        
        Returns list of dicts with:
        - dish_name, dish_code
        - sale_quantity, dish_price, dish_revenue
        - material_details (list of material usage)
        - material_cost, net_revenue, profit_margin
        """
        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()
            
            # Query to get dish sales with material costs
            query = """
                WITH dish_sales AS (
                    -- Get dish sales for the month
                    SELECT 
                        d.id as dish_id,
                        d.name as dish_name,
                        d.full_code as dish_code,
                        d.size as dish_size,
                        dms.sale_amount - COALESCE(dms.return_amount, 0) as net_quantity,
                        dph.price as dish_price
                    FROM dish_monthly_sale dms
                    JOIN dish d ON d.id = dms.dish_id
                    LEFT JOIN dish_price_history dph ON 
                        dph.dish_id = d.id 
                        AND dph.store_id = dms.store_id
                        AND dph.effective_year = %s
                        AND dph.effective_month = %s
                    WHERE dms.store_id = %s 
                        AND dms.year = %s 
                        AND dms.month = %s
                        AND (dms.sale_amount - COALESCE(dms.return_amount, 0)) > 0
                ),
                material_usage AS (
                    -- Calculate material usage for each dish
                    SELECT 
                        ds.dish_id,
                        m.material_number,
                        m.name as material_name,
                        dm.standard_quantity,
                        dm.loss_rate,
                        -- Formula: net_quantity * standard_quantity / conversion_unit * loss_rate
                        -- Note: conversion_unit is stored in dm.unit_conversion_rate
                        ds.net_quantity * COALESCE(dm.standard_quantity, 0) / COALESCE(dm.unit_conversion_rate, 1.0) * COALESCE(dm.loss_rate, 1.0) as total_usage,
                        mph.price as material_price,
                        (ds.net_quantity * COALESCE(dm.standard_quantity, 0) / COALESCE(dm.unit_conversion_rate, 1.0) * COALESCE(dm.loss_rate, 1.0)) * COALESCE(mph.price, 0) as material_cost
                    FROM dish_sales ds
                    LEFT JOIN dish_material dm ON dm.dish_id = ds.dish_id AND dm.store_id = %s
                    LEFT JOIN material m ON m.id = dm.material_id
                    LEFT JOIN material_price_history mph ON 
                        mph.material_id = m.id 
                        AND mph.store_id = %s
                        AND mph.effective_year = %s
                        AND mph.effective_month = %s
                )
                SELECT 
                    ds.dish_id,
                    ds.dish_name,
                    ds.dish_code,
                    ds.dish_size,
                    ds.net_quantity,
                    COALESCE(ds.dish_price, 0) as dish_price,
                    ds.net_quantity * COALESCE(ds.dish_price, 0) as dish_revenue,
                    COALESCE(STRING_AGG(
                        CASE 
                            WHEN mu.material_number IS NOT NULL 
                            THEN mu.material_name || '-' || mu.material_number || '-' || ROUND(mu.total_usage::numeric, 4)
                            ELSE NULL
                        END, 
                        '; ' ORDER BY mu.material_name
                    ), '') as material_details,
                    COALESCE(SUM(mu.material_cost), 0) as total_material_cost
                FROM dish_sales ds
                LEFT JOIN material_usage mu ON mu.dish_id = ds.dish_id
                GROUP BY 
                    ds.dish_id, ds.dish_name, ds.dish_code, ds.dish_size,
                    ds.net_quantity, ds.dish_price
                ORDER BY ds.dish_name, ds.dish_code
            """
            
            cursor.execute(query, (
                year, month,  # for dish price
                store_id, year, month,  # for dish sales
                store_id,  # for dish_material
                store_id, year, month  # for material price
            ))
            
            results = []
            for row in cursor.fetchall():
                dish_revenue = float(row['dish_revenue'] or 0)
                material_cost = float(row['total_material_cost'] or 0)
                net_revenue = dish_revenue - material_cost
                
                # Calculate profit margin percentage
                if dish_revenue > 0:
                    profit_margin = (net_revenue / dish_revenue) * 100
                else:
                    profit_margin = 0
                
                results.append({
                    'dish_name': row['dish_name'],
                    'dish_code': row['dish_code'],
                    'dish_size': row['dish_size'],
                    'sale_quantity': float(row['net_quantity'] or 0),
                    'dish_price': float(row['dish_price'] or 0),
                    'dish_revenue': dish_revenue,
                    'material_details': row['material_details'] or 'No materials',
                    'material_cost': material_cost,
                    'net_revenue': net_revenue,
                    'profit_margin': profit_margin
                })
            
            return results
    
    def _populate_store_worksheet(self, ws, store_data: List[Dict], store_name: str, year: int, month: int):
        """Populate a worksheet with store data."""
        
        # Add title
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f"{store_name} - Gross Revenue Report ({year}-{month:02d})"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers
        headers = [
            'Dish Name',
            'Dish Code', 
            'Size',
            'Quantity Sold',
            'Unit Price',
            'Dish Revenue',
            'Materials Used',
            'Material Cost',
            'Net Revenue',
            'Profit Margin %'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Add data rows
        row_num = 4
        total_revenue = 0
        total_cost = 0
        total_net = 0
        
        for dish_data in store_data:
            ws.cell(row=row_num, column=1, value=dish_data['dish_name'])
            ws.cell(row=row_num, column=2, value=dish_data['dish_code'])
            ws.cell(row=row_num, column=3, value=dish_data['dish_size'] or '')
            ws.cell(row=row_num, column=4, value=round(dish_data['sale_quantity'], 2))
            ws.cell(row=row_num, column=5, value=round(dish_data['dish_price'], 2))
            ws.cell(row=row_num, column=6, value=round(dish_data['dish_revenue'], 2))
            ws.cell(row=row_num, column=7, value=dish_data['material_details'])
            ws.cell(row=row_num, column=8, value=round(dish_data['material_cost'], 2))
            ws.cell(row=row_num, column=9, value=round(dish_data['net_revenue'], 2))
            ws.cell(row=row_num, column=10, value=f"{dish_data['profit_margin']:.1f}%")
            
            # Apply borders
            for col in range(1, 11):
                ws.cell(row=row_num, column=col).border = border
            
            # Apply number formatting and alignment
            ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='right')
            ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='right')
            ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='right')
            ws.cell(row=row_num, column=7).alignment = Alignment(wrap_text=True)
            ws.cell(row=row_num, column=8).alignment = Alignment(horizontal='right')
            ws.cell(row=row_num, column=9).alignment = Alignment(horizontal='right')
            ws.cell(row=row_num, column=10).alignment = Alignment(horizontal='center')
            
            # Color code profit margin
            profit_cell = ws.cell(row=row_num, column=10)
            if dish_data['profit_margin'] < 30:
                profit_cell.font = Font(color="FF0000")  # Red for low margin
            elif dish_data['profit_margin'] > 70:
                profit_cell.font = Font(color="00B050")  # Green for high margin
                
            total_revenue += dish_data['dish_revenue']
            total_cost += dish_data['material_cost']
            total_net += dish_data['net_revenue']
            
            row_num += 1
        
        # Add totals row
        row_num += 1
        total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        ws.merge_cells(f'A{row_num}:C{row_num}')
        total_label = ws.cell(row=row_num, column=1, value='TOTALS')
        total_label.font = Font(bold=True)
        total_label.alignment = Alignment(horizontal='right')
        total_label.fill = total_fill
        
        ws.cell(row=row_num, column=6, value=round(total_revenue, 2)).fill = total_fill
        ws.cell(row=row_num, column=8, value=round(total_cost, 2)).fill = total_fill
        ws.cell(row=row_num, column=9, value=round(total_net, 2)).fill = total_fill
        
        if total_revenue > 0:
            overall_margin = (total_net / total_revenue) * 100
            ws.cell(row=row_num, column=10, value=f"{overall_margin:.1f}%").fill = total_fill
        
        # Apply borders to total row
        for col in range(1, 11):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            if col >= 6:
                cell.font = Font(bold=True)
        
        # Adjust column widths
        column_widths = [25, 12, 10, 12, 10, 12, 40, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Add summary statistics
        stats_row = row_num + 2
        ws.cell(row=stats_row, column=1, value='Summary Statistics:').font = Font(bold=True)
        ws.cell(row=stats_row + 1, column=1, value='Total Dishes:')
        ws.cell(row=stats_row + 1, column=2, value=len(store_data))
        ws.cell(row=stats_row + 2, column=1, value='Average Profit Margin:')
        if store_data:
            avg_margin = sum(d['profit_margin'] for d in store_data) / len(store_data)
            ws.cell(row=stats_row + 2, column=2, value=f"{avg_margin:.1f}%")
        
        # Freeze panes (keep headers visible)
        ws.freeze_panes = 'A4'

    def _get_store_dish_revenue_data_with_history(
        self, store_id: int, year: int, month: int,
        last_month_year: int, last_month_month: int,
        last_year_year: int, last_year_month: int
    ) -> List[Dict]:
        """
        Get dish revenue data with historical price comparisons.
        """
        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()
            
            # Complex query to get current and historical prices
            query = """
                WITH dish_sales AS (
                    -- Get dish sales for the current month
                    SELECT 
                        d.id as dish_id,
                        d.name as dish_name,
                        d.full_code as dish_code,
                        d.size as dish_size,
                        dms.sale_amount - COALESCE(dms.return_amount, 0) as net_quantity,
                        -- Current month price
                        dph_current.price as dish_price_current,
                        -- Last month price
                        dph_last_month.price as dish_price_last_month,
                        -- Last year price
                        dph_last_year.price as dish_price_last_year
                    FROM dish_monthly_sale dms
                    JOIN dish d ON d.id = dms.dish_id
                    -- Current month price
                    LEFT JOIN dish_price_history dph_current ON 
                        dph_current.dish_id = d.id 
                        AND dph_current.store_id = dms.store_id
                        AND dph_current.effective_year = %s
                        AND dph_current.effective_month = %s
                    -- Last month price (don't check is_active for historical prices)
                    LEFT JOIN dish_price_history dph_last_month ON 
                        dph_last_month.dish_id = d.id 
                        AND dph_last_month.store_id = dms.store_id
                        AND dph_last_month.effective_year = %s
                        AND dph_last_month.effective_month = %s
                    -- Last year price (don't check is_active for historical prices)
                    LEFT JOIN dish_price_history dph_last_year ON 
                        dph_last_year.dish_id = d.id 
                        AND dph_last_year.store_id = dms.store_id
                        AND dph_last_year.effective_year = %s
                        AND dph_last_year.effective_month = %s
                    WHERE dms.store_id = %s 
                        AND dms.year = %s 
                        AND dms.month = %s
                        AND (dms.sale_amount - COALESCE(dms.return_amount, 0)) > 0
                ),
                material_usage AS (
                    -- Calculate material usage with price history
                    SELECT 
                        ds.dish_id,
                        m.material_number,
                        m.name as material_name,
                        dm.standard_quantity,
                        dm.loss_rate,
                        dm.unit_conversion_rate,
                        -- Calculate material usage using correct formula
                        ds.net_quantity * COALESCE(dm.standard_quantity, 0) / 
                        COALESCE(dm.unit_conversion_rate, 1.0) * COALESCE(dm.loss_rate, 1.0) as total_usage,
                        -- Current prices
                        mph_current.price as material_price_current,
                        -- Last month prices
                        mph_last_month.price as material_price_last_month,
                        -- Last year prices
                        mph_last_year.price as material_price_last_year,
                        -- Calculate costs
                        (ds.net_quantity * COALESCE(dm.standard_quantity, 0) / 
                         COALESCE(dm.unit_conversion_rate, 1.0) * COALESCE(dm.loss_rate, 1.0)) * 
                         COALESCE(mph_current.price, 0) as material_cost_current
                    FROM dish_sales ds
                    LEFT JOIN dish_material dm ON dm.dish_id = ds.dish_id AND dm.store_id = %s
                    LEFT JOIN material m ON m.id = dm.material_id
                    -- Current month material price
                    LEFT JOIN material_price_history mph_current ON 
                        mph_current.material_id = m.id 
                        AND mph_current.store_id = %s
                        AND mph_current.effective_year = %s
                        AND mph_current.effective_month = %s
                    -- Last month material price (don't check is_active for historical prices)
                    LEFT JOIN material_price_history mph_last_month ON 
                        mph_last_month.material_id = m.id 
                        AND mph_last_month.store_id = %s
                        AND mph_last_month.effective_year = %s
                        AND mph_last_month.effective_month = %s
                    -- Last year material price (don't check is_active for historical prices)
                    LEFT JOIN material_price_history mph_last_year ON 
                        mph_last_year.material_id = m.id 
                        AND mph_last_year.store_id = %s
                        AND mph_last_year.effective_year = %s
                        AND mph_last_year.effective_month = %s
                )
                SELECT 
                    ds.dish_id,
                    ds.dish_name,
                    ds.dish_code,
                    ds.dish_size,
                    ds.net_quantity,
                    -- Dish prices
                    COALESCE(ds.dish_price_current, 0) as dish_price_current,
                    COALESCE(ds.dish_price_last_month, 0) as dish_price_last_month,
                    COALESCE(ds.dish_price_last_year, 0) as dish_price_last_year,
                    -- Dish revenue
                    ds.net_quantity * COALESCE(ds.dish_price_current, 0) as dish_revenue,
                    -- Material details
                    COALESCE(STRING_AGG(
                        CASE 
                            WHEN mu.material_number IS NOT NULL 
                            THEN mu.material_name || '-' || mu.material_number || 
                                 '-Qty:' || ROUND(mu.total_usage::numeric, 2) ||
                                 '-$' || ROUND(mu.material_price_current::numeric, 2) ||
                                 '(LM:$' || COALESCE(ROUND(mu.material_price_last_month::numeric, 2)::text, 'N/A') ||
                                 ',LY:$' || COALESCE(ROUND(mu.material_price_last_year::numeric, 2)::text, 'N/A') || ')'
                            ELSE NULL
                        END, 
                        '; ' ORDER BY mu.material_name
                    ), 'No materials') as material_details,
                    -- Material costs
                    COALESCE(SUM(mu.material_cost_current), 0) as total_material_cost,
                    -- Average material price changes
                    AVG(CASE 
                        WHEN mu.material_price_last_month > 0 AND mu.material_price_current > 0
                        THEN ((mu.material_price_current - mu.material_price_last_month) / mu.material_price_last_month) * 100
                        ELSE NULL
                    END) as avg_material_price_change_last_month,
                    AVG(CASE 
                        WHEN mu.material_price_last_year > 0 AND mu.material_price_current > 0
                        THEN ((mu.material_price_current - mu.material_price_last_year) / mu.material_price_last_year) * 100
                        ELSE NULL
                    END) as avg_material_price_change_last_year
                FROM dish_sales ds
                LEFT JOIN material_usage mu ON mu.dish_id = ds.dish_id
                GROUP BY 
                    ds.dish_id, ds.dish_name, ds.dish_code, ds.dish_size,
                    ds.net_quantity, ds.dish_price_current, ds.dish_price_last_month, 
                    ds.dish_price_last_year
                ORDER BY ds.dish_name, ds.dish_code
            """
            
            cursor.execute(query, (
                # Dish price parameters
                year, month,  # current
                last_month_year, last_month_month,  # last month
                last_year_year, last_year_month,  # last year
                store_id, year, month,  # for dish sales
                # Material price parameters
                store_id,  # for dish_material
                store_id, year, month,  # current material price
                store_id, last_month_year, last_month_month,  # last month material price
                store_id, last_year_year, last_year_month  # last year material price
            ))
            
            results = []
            for row in cursor.fetchall():
                dish_revenue = float(row['dish_revenue'] or 0)
                material_cost = float(row['total_material_cost'] or 0)
                net_revenue = dish_revenue - material_cost
                
                # Calculate profit margin percentage
                if dish_revenue > 0:
                    profit_margin = (net_revenue / dish_revenue) * 100
                else:
                    profit_margin = 0
                
                # Calculate dish price changes
                dish_price_current = float(row['dish_price_current'] or 0)
                dish_price_last_month = float(row['dish_price_last_month'] or 0)
                dish_price_last_year = float(row['dish_price_last_year'] or 0)
                
                dish_price_change_last_month = None
                if dish_price_last_month > 0 and dish_price_current > 0:
                    dish_price_change_last_month = ((dish_price_current - dish_price_last_month) / dish_price_last_month) * 100
                
                dish_price_change_last_year = None
                if dish_price_last_year > 0 and dish_price_current > 0:
                    dish_price_change_last_year = ((dish_price_current - dish_price_last_year) / dish_price_last_year) * 100
                
                results.append({
                    'dish_name': row['dish_name'],
                    'dish_code': row['dish_code'],
                    'dish_size': row['dish_size'],
                    'sale_quantity': float(row['net_quantity'] or 0),
                    # Dish prices
                    'dish_price': dish_price_current,  # Keep backward compatibility
                    'dish_price_current': dish_price_current,
                    'dish_price_last_month': dish_price_last_month,
                    'dish_price_last_year': dish_price_last_year,
                    'dish_price_change_last_month': dish_price_change_last_month,
                    'dish_price_change_last_year': dish_price_change_last_year,
                    # Revenue and costs
                    'dish_revenue': dish_revenue,
                    'material_details': row['material_details'] or 'No materials',
                    'material_cost': material_cost,
                    'net_revenue': net_revenue,
                    'profit_margin': profit_margin,
                    # Material price changes
                    'avg_material_price_change_last_month': row['avg_material_price_change_last_month'],
                    'avg_material_price_change_last_year': row['avg_material_price_change_last_year']
                })
            
            return results

    def _populate_store_worksheet_with_history(
        self, ws, store_data: List[Dict], store_name: str, 
        year: int, month: int, last_month_date: datetime, last_year_date: datetime
    ):
        """Populate a worksheet with store data including historical comparisons."""
        
        # Add title
        ws.merge_cells('A1:P1')
        title_cell = ws['A1']
        title_cell.value = f"{store_name} - Gross Revenue Report ({year}-{month:02d})"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add comparison period info
        ws.merge_cells('A2:P2')
        info_cell = ws['A2']
        info_cell.value = (f"Comparing with: Last Month ({last_month_date.year}-{last_month_date.month:02d}) "
                          f"and Last Year ({last_year_date.year}-{last_year_date.month:02d})")
        info_cell.font = Font(size=10, italic=True)
        info_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add headers
        headers = [
            'Dish Name',
            'Dish Code', 
            'Size',
            'Qty Sold',
            # Dish price columns
            'Current\nPrice',
            'Last Month\nPrice',
            'LM %\nChange',
            'Last Year\nPrice',
            'LY %\nChange',
            # Revenue columns
            'Dish\nRevenue',
            'Materials Used\n(with price history)',
            'Material\nCost',
            'Avg Mat\nLM % Chg',
            'Avg Mat\nLY % Chg',
            'Net\nRevenue',
            'Profit\nMargin %'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Add data rows
        row_num = 5
        total_revenue = 0
        total_cost = 0
        total_net = 0
        
        for dish_data in store_data:
            # Basic info
            ws.cell(row=row_num, column=1, value=dish_data['dish_name'])
            ws.cell(row=row_num, column=2, value=dish_data['dish_code'])
            ws.cell(row=row_num, column=3, value=dish_data['dish_size'] or '')
            ws.cell(row=row_num, column=4, value=round(dish_data['sale_quantity'], 2))
            
            # Dish prices
            ws.cell(row=row_num, column=5, value=round(dish_data['dish_price_current'], 2))
            ws.cell(row=row_num, column=6, value=round(dish_data['dish_price_last_month'], 2) if dish_data['dish_price_last_month'] else 'N/A')
            
            # Last month price change
            if dish_data.get('dish_price_change_last_month') is not None:
                lm_cell = ws.cell(row=row_num, column=7, value=f"{dish_data['dish_price_change_last_month']:.1f}%")
                if dish_data['dish_price_change_last_month'] > 0:
                    lm_cell.font = Font(color="FF0000")  # Red for increase
                elif dish_data['dish_price_change_last_month'] < 0:
                    lm_cell.font = Font(color="00B050")  # Green for decrease
            else:
                ws.cell(row=row_num, column=7, value='N/A')
            
            ws.cell(row=row_num, column=8, value=round(dish_data['dish_price_last_year'], 2) if dish_data['dish_price_last_year'] else 'N/A')
            
            # Last year price change
            if dish_data.get('dish_price_change_last_year') is not None:
                ly_cell = ws.cell(row=row_num, column=9, value=f"{dish_data['dish_price_change_last_year']:.1f}%")
                if dish_data['dish_price_change_last_year'] > 0:
                    ly_cell.font = Font(color="FF0000")  # Red for increase
                elif dish_data['dish_price_change_last_year'] < 0:
                    ly_cell.font = Font(color="00B050")  # Green for decrease
            else:
                ws.cell(row=row_num, column=9, value='N/A')
            
            # Revenue and costs
            ws.cell(row=row_num, column=10, value=round(dish_data['dish_revenue'], 2))
            ws.cell(row=row_num, column=11, value=dish_data['material_details'])
            ws.cell(row=row_num, column=12, value=round(dish_data['material_cost'], 2))
            
            # Material price changes
            if dish_data.get('avg_material_price_change_last_month') is not None:
                mat_lm_cell = ws.cell(row=row_num, column=13, value=f"{dish_data['avg_material_price_change_last_month']:.1f}%")
                if dish_data['avg_material_price_change_last_month'] > 0:
                    mat_lm_cell.font = Font(color="FF0000")
                elif dish_data['avg_material_price_change_last_month'] < 0:
                    mat_lm_cell.font = Font(color="00B050")
            else:
                ws.cell(row=row_num, column=13, value='N/A')
            
            if dish_data.get('avg_material_price_change_last_year') is not None:
                mat_ly_cell = ws.cell(row=row_num, column=14, value=f"{dish_data['avg_material_price_change_last_year']:.1f}%")
                if dish_data['avg_material_price_change_last_year'] > 0:
                    mat_ly_cell.font = Font(color="FF0000")
                elif dish_data['avg_material_price_change_last_year'] < 0:
                    mat_ly_cell.font = Font(color="00B050")
            else:
                ws.cell(row=row_num, column=14, value='N/A')
            
            ws.cell(row=row_num, column=15, value=round(dish_data['net_revenue'], 2))
            ws.cell(row=row_num, column=16, value=f"{dish_data['profit_margin']:.1f}%")
            
            # Apply borders and alignment
            for col in range(1, 17):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                if col in [4, 5, 6, 8, 10, 12, 15]:  # Number columns
                    cell.alignment = Alignment(horizontal='right')
                elif col in [7, 9, 13, 14, 16]:  # Percentage columns
                    cell.alignment = Alignment(horizontal='center')
                elif col == 11:  # Material details
                    cell.alignment = Alignment(wrap_text=True)
            
            # Color code profit margin
            profit_cell = ws.cell(row=row_num, column=16)
            if dish_data['profit_margin'] < 30:
                profit_cell.font = Font(color="FF0000")  # Red for low margin
            elif dish_data['profit_margin'] > 70:
                profit_cell.font = Font(color="00B050")  # Green for high margin
                
            total_revenue += dish_data['dish_revenue']
            total_cost += dish_data['material_cost']
            total_net += dish_data['net_revenue']
            
            row_num += 1
        
        # Add totals row
        row_num += 1
        total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        ws.merge_cells(f'A{row_num}:C{row_num}')
        total_label = ws.cell(row=row_num, column=1, value='TOTALS')
        total_label.font = Font(bold=True)
        total_label.alignment = Alignment(horizontal='right')
        total_label.fill = total_fill
        
        ws.cell(row=row_num, column=10, value=round(total_revenue, 2)).fill = total_fill
        ws.cell(row=row_num, column=12, value=round(total_cost, 2)).fill = total_fill
        ws.cell(row=row_num, column=15, value=round(total_net, 2)).fill = total_fill
        
        if total_revenue > 0:
            overall_margin = (total_net / total_revenue) * 100
            ws.cell(row=row_num, column=16, value=f"{overall_margin:.1f}%").fill = total_fill
        
        # Apply borders to total row
        for col in range(1, 17):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            if col >= 10:
                cell.font = Font(bold=True)
        
        # Adjust column widths
        column_widths = [25, 12, 8, 8, 10, 10, 8, 10, 8, 10, 50, 10, 8, 8, 10, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Add summary statistics
        stats_row = row_num + 2
        ws.cell(row=stats_row, column=1, value='Summary Statistics:').font = Font(bold=True)
        ws.cell(row=stats_row + 1, column=1, value='Total Dishes:')
        ws.cell(row=stats_row + 1, column=2, value=len(store_data))
        ws.cell(row=stats_row + 2, column=1, value='Average Profit Margin:')
        if store_data:
            avg_margin = sum(d['profit_margin'] for d in store_data) / len(store_data)
            ws.cell(row=stats_row + 2, column=2, value=f"{avg_margin:.1f}%")
        
        # Add price change summary
        ws.cell(row=stats_row + 3, column=1, value='Dishes with price increases (LM):')
        dishes_increased_lm = sum(1 for d in store_data if d.get('dish_price_change_last_month') and d['dish_price_change_last_month'] > 0)
        ws.cell(row=stats_row + 3, column=2, value=dishes_increased_lm)
        
        ws.cell(row=stats_row + 4, column=1, value='Dishes with price increases (LY):')
        dishes_increased_ly = sum(1 for d in store_data if d.get('dish_price_change_last_year') and d['dish_price_change_last_year'] > 0)
        ws.cell(row=stats_row + 4, column=2, value=dishes_increased_ly)
        
        # Freeze panes (keep headers visible)
        ws.freeze_panes = 'A5'

    def _create_profit_summary_sheet(self, wb: Workbook, year: int, month: int):
        """
        Create profit summary sheet comparing current month with previous periods.
        Shows breakdown by actual dish types from database.
        
        Args:
            wb: Workbook to add sheet to
            year: Target year
            month: Target month
        """
        from datetime import date
        from dateutil.relativedelta import relativedelta
        from decimal import Decimal
        
        # Calculate dates
        current_date = date(year, month, 1)
        last_month = current_date - relativedelta(months=1)
        last_year = current_date - relativedelta(years=1)
        
        # Get dish types from database
        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, name 
                FROM dish_type 
                WHERE id IN (3, 5, 6, 7, 8, 9)  -- Main dish types to show
                ORDER BY id
            """)
            dish_types = cursor.fetchall()
        
        # Create worksheet
        ws = wb.create_sheet(title="毛利率对比表", index=0)
        
        # Title
        total_cols = 1 + len(dish_types) * 3  # Store name + 3 cols per dish type
        ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
        ws['A1'] = f"{year}年{month}月毛利率对比表"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Build dynamic headers based on dish types
        headers = [('A2:A3', '门店名称')]
        col = 2  # Start from column B
        
        for dish_type in dish_types:
            type_name = dish_type['name']
            # Merge cells for dish type name
            end_col = col + 2
            headers.append((f'{get_column_letter(col)}2:{get_column_letter(end_col)}2', type_name))
            # Add date headers
            headers.append((f'{get_column_letter(col)}3', f'{month:02d}-01-{year}'))
            headers.append((f'{get_column_letter(col+1)}3', f'{last_month.month:02d}-01-{last_month.year}'))
            headers.append((f'{get_column_letter(col+2)}3', '环比'))
            col = end_col + 1
        
        # Apply headers
        for cell_range, text in headers:
            if ':' in cell_range:
                ws.merge_cells(cell_range)
                start_cell = cell_range.split(':')[0]
                ws[start_cell] = text
                ws[start_cell].font = Font(bold=True)
                ws[start_cell].alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws[cell_range] = text
                ws[cell_range].font = Font(bold=True)
                ws[cell_range].alignment = Alignment(horizontal='center', vertical='center')
        
        # Get data for each store
        row = 4
        with self.db_manager.get_connection() as conn:
            cursor = conn.cursor()
            
            for store_id, store_name in STORE_ID_TO_NAME_MAPPING.items():
                if store_id > 100:  # Skip Hi Bowl
                    continue
                
                # Query to get revenue by dish type and total
                dish_type_query = """
                    SELECT 
                        dt.id as dish_type_id,
                        dt.name as dish_type_name,
                        COALESCE(SUM(dms.sale_amount - COALESCE(dms.return_amount, 0)), 0) as revenue
                    FROM dish_type dt
                    LEFT JOIN dish_child_type dct ON dt.id = dct.dish_type_id
                    LEFT JOIN dish d ON d.dish_child_type_id = dct.id
                    LEFT JOIN dish_monthly_sale dms ON dms.dish_id = d.id 
                        AND dms.store_id = %s 
                        AND dms.year = %s 
                        AND dms.month = %s
                    WHERE dt.id IN (3, 5, 6, 7, 8, 9)
                    GROUP BY dt.id, dt.name
                    ORDER BY dt.id
                """
                
                # Get total revenue for percentage calculation
                total_query = """
                    SELECT COALESCE(SUM(sale_amount - COALESCE(return_amount, 0)), 0) as total_revenue
                    FROM dish_monthly_sale
                    WHERE store_id = %s AND year = %s AND month = %s
                """
                
                # Get current month total
                cursor.execute(total_query, [store_id, year, month])
                current_total = Decimal(str(cursor.fetchone()['total_revenue']) or '0')
                
                # Get last month total
                cursor.execute(total_query, [store_id, last_month.year, last_month.month])
                last_total = Decimal(str(cursor.fetchone()['total_revenue']) or '0')
                
                # Get current month data by dish type
                cursor.execute(dish_type_query, [store_id, year, month])
                current_data = {row['dish_type_id']: Decimal(str(row['revenue']) if row['revenue'] else 0) 
                               for row in cursor.fetchall()}
                
                # Get last month data by dish type
                cursor.execute(dish_type_query, [store_id, last_month.year, last_month.month])
                last_data = {row['dish_type_id']: Decimal(str(row['revenue']) if row['revenue'] else 0) 
                            for row in cursor.fetchall()}
                
                # Write store name
                ws[f'A{row}'] = store_name
                
                # Write data for each dish type
                col = 2  # Start from column B
                for dish_type in dish_types:
                    type_id = dish_type['id']
                    current_revenue = current_data.get(type_id, Decimal('0'))
                    last_revenue = last_data.get(type_id, Decimal('0'))
                    
                    # Calculate percentages of total revenue
                    current_pct = (current_revenue / current_total * 100) if current_total > 0 else 0
                    last_pct = (last_revenue / last_total * 100) if last_total > 0 else 0
                    
                    # Calculate change in percentage points
                    change = current_pct - last_pct
                    
                    # Write percentage values
                    ws.cell(row=row, column=col).value = f"{current_pct:.1f}%"
                    ws.cell(row=row, column=col+1).value = f"{last_pct:.1f}%"
                    ws.cell(row=row, column=col+2).value = f"{change:+.1f}%"  # + sign for positive
                    
                    # Apply color for change
                    change_cell = ws.cell(row=row, column=col+2)
                    if change > 0:
                        change_cell.font = Font(color="008000")  # Green for positive
                    elif change < 0:
                        change_cell.font = Font(color="FF0000")  # Red for negative
                    
                    col += 3
                
                # Apply formatting for this row
                for col_idx in range(1, total_cols + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Only set font if not already set (e.g., for colored change cells)
                    if col_idx > 1 and not hasattr(cell, '_font'):  # Numeric columns
                        if not (col_idx % 3 == 1):  # Skip change columns that already have color
                            cell.font = Font(size=10)
                
                row += 1
        
        # Apply borders to all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=total_cols):
            for cell in row_cells:
                cell.border = thin_border
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # Store name column
        for col_idx in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
    
    def _create_theory_vs_actual_sheet(self, ws: Workbook, store_id: int, store_name: str, year: int, month: int):
        """
        Create a sheet comparing theoretical usage (from dish sales) vs actual usage (from inventory).
        
        Args:
            ws: Worksheet to populate
            store_id: Store ID
            store_name: Store name
            year: Year
            month: Month
        """
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{store_name} - 理论vs实际用量对比 ({year}-{month:02d})"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Explanation
        ws.merge_cells('A2:H2')
        ws['A2'] = "理论用量 = 菜品销售量 × 标准配方 | 实际用量 = 期初 + 采购 + 调入 - 期末 - 调出"
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = [
            '物料编号',
            '物料名称',
            '理论用量',
            '实际用量',
            '差异数量',
            '差异率%',
            '单价',
            '差异金额'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Get data using the database module
        with self.db_manager.get_connection() as conn:
            self.dish_material_db = DishMaterialDatabase(conn)
            
            # Get theoretical vs actual comparison
            comparison_data = self.dish_material_db.get_theoretical_vs_actual_usage(store_id, year, month)
            
            # Get material prices
            cursor = conn.cursor()
            cursor.execute("""
                SELECT 
                    m.material_number,
                    COALESCE(mph.price, 0) as price
                FROM material m
                LEFT JOIN material_price_history mph ON 
                    mph.material_id = m.id 
                    AND mph.store_id = %s
                    AND mph.effective_year = %s
                    AND mph.effective_month = %s
                WHERE m.material_number IN %s
            """, (store_id, year, month, 
                  tuple([d['material_number'] for d in comparison_data]) if comparison_data else ('',)))
            
            prices = {row['material_number']: float(row['price']) for row in cursor.fetchall()}
        
        # Populate data
        row = 5
        total_theory = 0
        total_actual = 0
        total_variance_value = 0
        
        for item in comparison_data:
            material_num = item['material_number']
            theory = float(item['theoretical_usage'])
            actual = float(item['actual_usage'])
            variance = float(item['variance'])
            variance_pct = item['variance_percentage']
            price = prices.get(material_num, 0)
            variance_value = variance * price
            
            # Write row
            ws.cell(row=row, column=1, value=material_num)
            ws.cell(row=row, column=2, value=item.get('material_name', ''))
            ws.cell(row=row, column=3, value=f"{theory:.2f}")
            ws.cell(row=row, column=4, value=f"{actual:.2f}")
            ws.cell(row=row, column=5, value=f"{variance:.2f}")
            
            # Variance percentage with color
            variance_cell = ws.cell(row=row, column=6)
            if variance_pct is not None:
                variance_cell.value = f"{variance_pct:.1f}%"
                if variance_pct > 10:  # Over 10% more actual than theory
                    variance_cell.font = Font(color="FF0000")  # Red
                elif variance_pct < -10:  # Over 10% less actual than theory
                    variance_cell.font = Font(color="0000FF")  # Blue
            else:
                variance_cell.value = "N/A"
            
            ws.cell(row=row, column=7, value=f"${price:.2f}")
            ws.cell(row=row, column=8, value=f"${variance_value:.2f}")
            
            # Apply borders
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col >= 3:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
            
            total_theory += theory
            total_actual += actual
            total_variance_value += variance_value
            row += 1
        
        # Add summary row
        ws.cell(row=row, column=1, value="合计").font = Font(bold=True)
        ws.cell(row=row, column=2, value="").font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"{total_theory:.2f}").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"{total_actual:.2f}").font = Font(bold=True)
        ws.cell(row=row, column=5, value=f"{total_actual - total_theory:.2f}").font = Font(bold=True)
        
        if total_theory > 0:
            total_variance_pct = ((total_actual - total_theory) / total_theory * 100)
            ws.cell(row=row, column=6, value=f"{total_variance_pct:.1f}%").font = Font(bold=True)
        else:
            ws.cell(row=row, column=6, value="N/A").font = Font(bold=True)
        
        ws.cell(row=row, column=7, value="").font = Font(bold=True)
        ws.cell(row=row, column=8, value=f"${total_variance_value:.2f}").font = Font(bold=True)
        
        # Apply borders to summary row
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col >= 3:
                cell.alignment = Alignment(horizontal='right')
        
        # Set column widths
        ws.column_dimensions['A'].width = 15  # Material number
        ws.column_dimensions['B'].width = 30  # Material name
        ws.column_dimensions['C'].width = 12  # Theory
        ws.column_dimensions['D'].width = 12  # Actual
        ws.column_dimensions['E'].width = 12  # Variance
        ws.column_dimensions['F'].width = 10  # Variance %
        ws.column_dimensions['G'].width = 10  # Price
        ws.column_dimensions['H'].width = 12  # Variance value
        
        # Freeze panes
        ws.freeze_panes = 'A5'


def main():
    """Main entry point for the script."""
    parser = argparse.ArgumentParser(
        description='Generate gross revenue report for dishes by store'
    )
    parser.add_argument(
        '--year',
        type=int,
        required=True,
        help='Target year (e.g., 2025)'
    )
    parser.add_argument(
        '--month',
        type=int,
        required=True,
        help='Target month (1-12)'
    )
    parser.add_argument(
        '--output-path',
        type=str,
        help='Optional output file path'
    )
    parser.add_argument(
        '--test',
        action='store_true',
        help='Use test database'
    )
    parser.add_argument(
        '--no-history',
        action='store_true',
        help='Generate report without historical price comparisons'
    )
    
    args = parser.parse_args()
    
    # Validate month
    if not 1 <= args.month <= 12:
        logger.error("Month must be between 1 and 12")
        sys.exit(1)
    
    # Initialize database connection
    db_config = DatabaseConfig(is_test=args.test)
    db_manager = DatabaseManager(db_config)
    
    # Test database connection
    if not db_manager.test_connection():
        logger.error("Failed to connect to database")
        sys.exit(1)
    
    logger.info(f"Connected to {'test' if args.test else 'production'} database")
    
    # Generate report
    generator = GrossRevenueReportGenerator(db_manager)
    output_file = generator.generate_report(
        args.year, args.month, args.output_path, 
        include_history=not args.no_history,
        include_profit_summary=True
    )
    
    print("\n" + "="*50)
    print("GROSS REVENUE REPORT GENERATION COMPLETE")
    print("="*50)
    print(f"Report saved to: {output_file}")
    print("="*50)
    
    return 0


if __name__ == '__main__':
    sys.exit(main())