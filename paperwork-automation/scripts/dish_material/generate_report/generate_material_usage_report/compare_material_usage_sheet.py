#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate sheet to compare theoretical usage vs actual usage for materials.
"""

import sys
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import logging
from datetime import datetime
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Add parent directory to path for imports
project_root = Path(__file__).resolve().parent.parent.parent.parent.parent
sys.path.insert(0, str(project_root))

from utils.database import DatabaseManager, DatabaseConfig

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def get_material_usage_data_from_database(db_manager: DatabaseManager, year: int, month: int, store_id: int, debug: bool = False) -> Dict:
    """
    Get material usage data comparing theoretical vs actual usage.
    
    Args:
        db_manager: Database manager instance
        year: Year for the data
        month: Month for the data
        store_id: Store ID
        debug: If True, include detailed calculation info
    
    Returns:
        Dict with material_number as key containing:
        - material_name: str
        - actual_usage: float (monthly usage - ending inventory)
        - theory_usage: float (calculated from dish sales)
        - usage_details: List of dicts with dish info and usage
    """
    result = {}
    
    with db_manager.get_connection() as conn:
        cursor = conn.cursor()
        
        # Get actual material usage (directly from material_monthly_usage which now stores the pre-calculated actual usage)
        actual_usage_query = """
        SELECT 
            m.material_number,
            m.description as material_name,
            mmu.material_used as actual_usage
        FROM material_monthly_usage mmu
        JOIN material m ON m.id = mmu.material_id AND m.store_id = mmu.store_id
        WHERE mmu.year = %s
          AND mmu.month = %s
          AND mmu.store_id = %s
        """
        
        cursor.execute(actual_usage_query, (year, month, store_id))
        actual_usage_data = cursor.fetchall()
        
        for row in actual_usage_data:
            result[row['material_number']] = {
                'material_name': row['material_name'],
                'actual_usage': float(row['actual_usage']) if row['actual_usage'] else 0.0,
                'theory_usage': 0.0,
                'usage_details': []
            }
        
        # Get theoretical usage from dish sales (including combo sales)
        theory_usage_query = """
        WITH dish_sales AS (
            -- Regular dish sales
            SELECT 
                dms.dish_id,
                d.name as dish_name,
                d.short_code as dish_short_code,
                d.full_code as dish_full_code,
                d.full_code as original_full_code,
                d.size as dish_size,
                COALESCE(dms.sale_amount, 0) - COALESCE(dms.return_amount, 0) as quantity_sold,
                FALSE as is_combo
            FROM dish_monthly_sale dms
            JOIN dish d ON d.id = dms.dish_id
            WHERE dms.year = %s
              AND dms.month = %s
              AND dms.store_id = %s
            
            UNION ALL
            
            -- Combo dish sales
            SELECT 
                mcds.dish_id,
                CONCAT(d.name, '(套餐-', c.name, ')') as dish_name,
                d.short_code as dish_short_code,
                d.full_code as dish_full_code,
                d.full_code as original_full_code,
                d.size as dish_size,
                mcds.sale_amount as quantity_sold,
                TRUE as is_combo
            FROM monthly_combo_dish_sale mcds
            JOIN dish d ON d.id = mcds.dish_id
            JOIN combo c ON c.id = mcds.combo_id
            WHERE mcds.year = %s
              AND mcds.month = %s
              AND mcds.store_id = %s
        )
        SELECT 
            m.material_number,
            m.description as material_name,
            ds.dish_short_code,
            CASE 
                WHEN ds.is_combo THEN ds.dish_full_code
                ELSE CONCAT(ds.dish_full_code, '-', ds.dish_short_code)
            END as dish_full_code,
            ds.dish_name,
            ds.dish_size,
            ds.quantity_sold,
            ds.is_combo,
            dm.standard_quantity,
            dm.loss_rate,
            dm.unit_conversion_rate as unit_conversion,
            (ds.quantity_sold * dm.standard_quantity * COALESCE(dm.loss_rate, 0) / COALESCE(NULLIF(dm.unit_conversion_rate, 0), 1)) as theory_usage
        FROM dish_sales ds
        JOIN dish d2 ON d2.id = (
            -- Try to find the best matching dish with material mappings
            SELECT d3.id
            FROM dish d3
            WHERE d3.full_code = ds.original_full_code 
              AND d3.id IN (SELECT dish_id FROM dish_material WHERE store_id = %s)
            ORDER BY 
              -- Prefer exact size match
              CASE WHEN d3.size = ds.dish_size THEN 0 
                   WHEN d3.size IS NULL AND ds.dish_size IS NULL THEN 0
                   ELSE 1 
              END,
              d3.id
            LIMIT 1
        )
        JOIN dish_material dm ON dm.dish_id = d2.id AND dm.store_id = %s
        JOIN material m ON m.id = dm.material_id AND m.store_id = %s
        ORDER BY m.material_number, ds.dish_short_code, ds.is_combo
        """
        
        cursor.execute(theory_usage_query, (year, month, store_id, year, month, store_id, store_id, store_id, store_id))
        theory_usage_data = cursor.fetchall()
        
        for row in theory_usage_data:
            material_number = row['material_number']
            if material_number not in result:
                result[material_number] = {
                    'material_name': row['material_name'],
                    'actual_usage': 0.0,
                    'theory_usage': 0.0,
                    'usage_details': []
                }
            
            # Add to theory usage total
            result[material_number]['theory_usage'] += float(row['theory_usage']) if row['theory_usage'] else 0.0
            
            # Add usage detail
            # For combos, size is shown separately in the dish_code
            # For regular dishes, append size to name if present
            if row['is_combo']:
                # Combo name already has (套餐-xxx) suffix, keep as is
                dish_name = row['dish_name']
                # Add size to the dish_code for combos
                dish_code = f"{row['dish_full_code']}-{row['dish_size']}" if row['dish_size'] else row['dish_full_code']
            else:
                # Regular dish: append size to name if present
                dish_name = f"{row['dish_name']}-{row['dish_size']}" if row['dish_size'] else row['dish_name']
                dish_code = row['dish_full_code']
            
            usage_detail = {
                'dish_code': dish_code,
                'dish_name': dish_name,
                'usage': float(row['theory_usage']) if row['theory_usage'] else 0.0,
                'is_combo': row['is_combo']
            }
            
            # Add debug information if requested
            if debug:
                usage_detail['quantity_sold'] = float(row['quantity_sold']) if row['quantity_sold'] else 0.0
                usage_detail['standard_quantity'] = float(row['standard_quantity']) if row['standard_quantity'] else 0.0
                usage_detail['loss_rate'] = float(row['loss_rate']) if row['loss_rate'] else 0.0
                usage_detail['unit_conversion'] = float(row['unit_conversion']) if row['unit_conversion'] else 1.0
            
            result[material_number]['usage_details'].append(usage_detail)
    
    return result


def get_material_usage_mapping(db_manager: DatabaseManager, year: int, month: int, store_id: int, debug: bool = False) -> Dict:
    """
    Get material usage mapping in the specified format.
    
    Args:
        db_manager: Database manager instance
        year: Year for the data
        month: Month for the data
        store_id: Store ID
        debug: If True, include detailed calculation info
    
    Returns mapping like:
    {
        "1500882": {
            "material_name": "三花淡奶（354ML*48瓶/箱）",
            "usages": {
                "90000413-97684": {
                    "name": "冬阴功锅底-单锅",
                    "usage": 0.8
                },
                ...
            }
        }
    }
    """
    data = get_material_usage_data_from_database(db_manager, year, month, store_id, debug=debug)
    
    mapping = {}
    for material_number, material_data in data.items():
        usages = {}
        for detail in material_data['usage_details']:
            # Create key with combo suffix if needed
            key = detail['dish_code']
            if detail['is_combo']:
                key += "-combo"
            
            usage_info = {
                'name': detail['dish_name'],
                'usage': detail['usage']
            }
            
            # Add debug information if available
            if debug and 'quantity_sold' in detail:
                usage_info['quantity_sold'] = detail['quantity_sold']
                usage_info['standard_quantity'] = detail['standard_quantity']
                usage_info['loss_rate'] = detail['loss_rate']
                usage_info['unit_conversion'] = detail['unit_conversion']
            
            usages[key] = usage_info
        
        if usages:  # Only add if there are usages
            mapping[material_number] = {
                'material_name': material_data['material_name'],
                'usages': usages
            }
    
    return mapping


def write_material_usage_to_sheet(worksheet, db_manager: DatabaseManager, year: int, month: int, store_id: int, store_name: str, debug: bool = False):
    """
    Write material usage comparison to an Excel worksheet.
    
    Args:
        worksheet: Openpyxl worksheet object
        db_manager: Database manager instance
        year: Year for the data
        month: Month for the data
        store_id: Store ID
        store_name: Store name for display
        debug: If True, include detailed calculation info
    
    The sheet has headers: 
    物料名称, 物料号, 理论消耗来源, 总理论消耗, 实际消耗, 差异, 备注
    """
    # Add title and info at the top FIRST
    title_cell = worksheet.cell(row=1, column=1, value=f"{store_name} - 物料理论与实际消耗对比")
    title_cell.font = Font(size=14, bold=True)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    title_cell.alignment = Alignment(horizontal='center')
    
    date_cell = worksheet.cell(row=2, column=1, value=f"{year}年{month}月")
    date_cell.font = Font(size=12)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    date_cell.alignment = Alignment(horizontal='center')
    
    # Leave a blank row
    header_row = 4
    
    # Set headers
    headers = ['物料名称', '物料号', '理论消耗来源', '总理论消耗', '实际消耗', '差异', '备注']
    
    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Get data
    data = get_material_usage_data_from_database(db_manager, year, month, store_id, debug=debug)
    
    # Sort materials by material number
    sorted_materials = sorted(data.items(), key=lambda x: x[0])
    
    current_row = header_row + 1  # Start data after header row
    
    for material_number, material_data in sorted_materials:
        if not material_data['usage_details'] and material_data['actual_usage'] == 0:
            continue  # Skip materials with no usage
        
        usage_details = material_data['usage_details']
        num_details = max(len(usage_details), 1)  # At least one row even if no details
        
        # Material name and number (merged across detail rows if multiple)
        material_name_cell = worksheet.cell(row=current_row, column=1, value=material_data['material_name'])
        material_number_cell = worksheet.cell(row=current_row, column=2, value=material_number)
        
        if num_details > 1:
            # Merge cells for material name and number
            worksheet.merge_cells(start_row=current_row, start_column=1, 
                                end_row=current_row + num_details - 1, end_column=1)
            worksheet.merge_cells(start_row=current_row, start_column=2, 
                                end_row=current_row + num_details - 1, end_column=2)
        
        material_name_cell.alignment = Alignment(vertical='center', wrap_text=True)
        material_number_cell.alignment = Alignment(vertical='center')
        
        # Theory usage details
        if usage_details:
            for i, detail in enumerate(usage_details):
                if debug and 'quantity_sold' in detail:
                    # Debug format with detailed calculation info
                    theory_source = (f"{detail['dish_name']} {detail['dish_code']} "
                                   f"实收数量:{detail['quantity_sold']:.0f} "
                                   f"出品分量(kg):{detail['standard_quantity']:.2f} "
                                   f"损耗:{detail['loss_rate']:.2f} "
                                   f"物料单位:{detail['unit_conversion']:.2f} "
                                   f"计算用量:{detail['usage']:.2f}")
                else:
                    # Normal format (default)
                    theory_source = f"{detail['dish_name']}: {detail['usage']:.2f}"
                    if debug:
                        # If debug is true but no debug data, still show codes
                        theory_source = f"{detail['dish_name']} {detail['dish_code']}: {detail['usage']:.2f}"
                
                worksheet.cell(row=current_row + i, column=3, value=theory_source)
                worksheet.cell(row=current_row + i, column=3).border = thin_border
        else:
            worksheet.cell(row=current_row, column=3, value="无理论消耗")
            worksheet.cell(row=current_row, column=3).border = thin_border
        
        # Total theory usage (merged)
        total_theory_cell = worksheet.cell(row=current_row, column=4, 
                                          value=round(material_data['theory_usage'], 2))
        if num_details > 1:
            worksheet.merge_cells(start_row=current_row, start_column=4, 
                                end_row=current_row + num_details - 1, end_column=4)
        total_theory_cell.alignment = Alignment(vertical='center', horizontal='right')
        total_theory_cell.border = thin_border
        
        # Actual usage (merged)
        actual_usage_cell = worksheet.cell(row=current_row, column=5, 
                                          value=round(material_data['actual_usage'], 2))
        if num_details > 1:
            worksheet.merge_cells(start_row=current_row, start_column=5, 
                                end_row=current_row + num_details - 1, end_column=5)
        actual_usage_cell.alignment = Alignment(vertical='center', horizontal='right')
        actual_usage_cell.border = thin_border
        
        # Difference percentage (merged)
        difference = None
        if material_data['theory_usage'] != 0:
            difference = (material_data['actual_usage'] - material_data['theory_usage']) / material_data['theory_usage']
            diff_text = f"{difference:.1%}"
        else:
            diff_text = "N/A"
        
        diff_cell = worksheet.cell(row=current_row, column=6, value=diff_text)
        if num_details > 1:
            worksheet.merge_cells(start_row=current_row, start_column=6, 
                                end_row=current_row + num_details - 1, end_column=6)
        diff_cell.alignment = Alignment(vertical='center', horizontal='center')
        diff_cell.border = thin_border
        
        # Highlight large differences
        if difference is not None and abs(difference) > 0.2:  # More than 20% difference
            diff_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Notes column (merged, empty for manual input)
        notes_cell = worksheet.cell(row=current_row, column=7, value="")
        if num_details > 1:
            worksheet.merge_cells(start_row=current_row, start_column=7, 
                                end_row=current_row + num_details - 1, end_column=7)
        notes_cell.border = thin_border
        
        # Apply borders to all cells
        for row in range(current_row, current_row + num_details):
            for col in range(1, 8):
                worksheet.cell(row=row, column=col).border = thin_border
        
        current_row += num_details
    
    # Auto-adjust column widths
    column_widths = [30, 15, 50, 15, 15, 10, 20]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width
    
    logger.info(f"Completed writing material usage sheet for {store_name}")


if __name__ == "__main__":
    """Test the functions with command line arguments."""
    import argparse
    
    parser = argparse.ArgumentParser(description='Test material usage comparison functions')
    parser.add_argument('--year', type=int, default=2025, help='Year (default: 2025)')
    parser.add_argument('--month', type=int, default=1, help='Month (default: 1)')
    parser.add_argument('--store-id', type=int, default=1, help='Store ID (default: 1)')
    parser.add_argument('--debug', action='store_true', help='Enable debug output with detailed calculations')
    parser.add_argument('--test-db', action='store_true', help='Use test database')
    
    args = parser.parse_args()
    
    # Create database connection
    db_config = DatabaseConfig(is_test=args.test_db)
    db_manager = DatabaseManager(db_config)
    
    # Test get_material_usage_mapping function
    print(f"\nTesting material usage mapping for Store {args.store_id}, {args.year}/{args.month}")
    print(f"Debug mode: {args.debug}")
    print("-" * 60)
    
    mapping = get_material_usage_mapping(db_manager, args.year, args.month, args.store_id, debug=args.debug)
    
    # Print sample output (first 3 materials)
    count = 0
    for material_number, material_data in mapping.items():
        if count >= 3:
            break
        
        print(f"\nMaterial: {material_data['material_name']} ({material_number})")
        
        # Print first 2 usages for each material
        usage_count = 0
        for dish_code, usage_info in material_data['usages'].items():
            if usage_count >= 2:
                break
            
            if args.debug and 'quantity_sold' in usage_info:
                # Debug format with details
                print(f"  {usage_info['name']} {dish_code} 实收数量:{usage_info['quantity_sold']:.0f} "
                      f"出品分量(kg):{usage_info['standard_quantity']:.2f} "
                      f"损耗:{usage_info['loss_rate']:.1f} "
                      f"物料单位:{usage_info['unit_conversion']:.1f} "
                      f"计算用量:{usage_info['usage']:.2f}")
            else:
                # Normal format
                print(f"  {usage_info['name']}: {usage_info['usage']:.2f}")
            
            usage_count += 1
        
        count += 1
    
    print("\n" + "-" * 60)
    print(f"Total materials with usage: {len(mapping)}")