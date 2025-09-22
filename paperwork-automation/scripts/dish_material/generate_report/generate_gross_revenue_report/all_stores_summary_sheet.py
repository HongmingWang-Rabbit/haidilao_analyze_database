#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate summary sheet showing all stores' total revenue and actual material usage.
"""

import sys
from pathlib import Path
from typing import Dict, List
import logging
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Add parent directory to path for imports
project_root = Path(__file__).resolve().parent.parent.parent.parent.parent
sys.path.insert(0, str(project_root))

from utils.database import DatabaseManager
from lib.config import STORE_ID_TO_NAME_MAPPING

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def get_all_stores_summary_data(db_manager: DatabaseManager, year: int, month: int) -> List[Dict]:
    """
    Get summary data for all stores including total revenue and actual material usage.
    
    Args:
        db_manager: Database manager instance
        year: Year for the data
        month: Month for the data
    
    Returns:
        List of dicts containing store summary data
    """
    results = []
    
    with db_manager.get_connection() as conn:
        cursor = conn.cursor()
        
        # Query to get total revenue and actual material usage for each store
        query = """
        WITH store_revenue AS (
            -- Get total revenue from dish sales
            SELECT 
                dms.store_id,
                SUM((COALESCE(dms.sale_amount, 0) - COALESCE(dms.return_amount, 0)) * COALESCE(dph.price, 0)) as total_revenue
            FROM dish_monthly_sale dms
            LEFT JOIN dish_price_history dph ON dph.dish_id = dms.dish_id 
                AND dph.store_id = dms.store_id
                AND dph.is_active = TRUE
            WHERE dms.year = %s
              AND dms.month = %s
            GROUP BY dms.store_id
        ),
        material_usage AS (
            -- Get actual material usage from material_monthly_usage (only 成本类)
            SELECT 
                mmu.store_id,
                SUM(COALESCE(mmu.material_used, 0) * COALESCE(mph.price, 0)) as total_material_cost
            FROM material_monthly_usage mmu
            JOIN material m ON m.id = mmu.material_id AND m.store_id = mmu.store_id
            LEFT JOIN material_price_history mph ON mph.material_id = m.id 
                AND mph.store_id = m.store_id
                AND mph.is_active = TRUE
            WHERE mmu.year = %s
              AND mmu.month = %s
              AND mmu.material_use_type = '成本类'  -- Only include cost-type materials
            GROUP BY mmu.store_id
        )
        SELECT 
            sr.store_id,
            COALESCE(sr.total_revenue, 0) as total_revenue,
            COALESCE(mu.total_material_cost, 0) as total_material_cost,
            COALESCE(sr.total_revenue, 0) - COALESCE(mu.total_material_cost, 0) as gross_profit,
            CASE 
                WHEN COALESCE(sr.total_revenue, 0) > 0 THEN 
                    ((COALESCE(sr.total_revenue, 0) - COALESCE(mu.total_material_cost, 0)) / COALESCE(sr.total_revenue, 0)) * 100
                ELSE 0
            END as profit_margin
        FROM store_revenue sr
        LEFT JOIN material_usage mu ON mu.store_id = sr.store_id
        ORDER BY sr.store_id
        """
        
        cursor.execute(query, (year, month, year, month))
        raw_data = cursor.fetchall()
        
        for row in raw_data:
            store_id = row['store_id']
            
            # Skip Hi Bowl (store_id = 101) for now
            if store_id == 101:
                continue
                
            # Get store name
            store_name = STORE_ID_TO_NAME_MAPPING.get(store_id, f"Store {store_id}")
            
            results.append({
                'store_id': store_id,
                'store_name': store_name,
                'total_revenue': float(row['total_revenue']) if row['total_revenue'] else 0,
                'total_material_cost': float(row['total_material_cost']) if row['total_material_cost'] else 0,
                'gross_profit': float(row['gross_profit']) if row['gross_profit'] else 0,
                'profit_margin': float(row['profit_margin']) if row['profit_margin'] else 0
            })
    
    return results


def write_all_stores_summary_sheet(worksheet, db_manager: DatabaseManager, year: int, month: int):
    """
    Write all stores summary data to an Excel worksheet.
    
    Args:
        worksheet: Openpyxl worksheet object
        db_manager: Database manager instance
        year: Year for the data
        month: Month for the data
    """
    # Add title
    title_cell = worksheet.cell(row=1, column=1, value="全店铺实际毛利汇总")
    title_cell.font = Font(size=14, bold=True)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    title_cell.alignment = Alignment(horizontal='center')
    
    date_cell = worksheet.cell(row=2, column=1, value=f"{year}年{month}月")
    date_cell.font = Font(size=12)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    date_cell.alignment = Alignment(horizontal='center')
    
    # Leave a blank row
    header_row = 4
    
    # Set headers
    headers = ['门店', '总销售金额', '总物料使用金额', '总毛利', '总毛利百分比']
    
    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Get data
    data = get_all_stores_summary_data(db_manager, year, month)
    
    # Alternating row fills
    even_row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    current_row = header_row + 1
    
    # Write data for each store
    for i, store_data in enumerate(data):
        # Store name
        worksheet.cell(row=current_row, column=1, value=store_data['store_name'])
        
        # Total revenue
        worksheet.cell(row=current_row, column=2, value=round(store_data['total_revenue'], 2))
        
        # Total material cost (actual from inventory)
        worksheet.cell(row=current_row, column=3, value=round(store_data['total_material_cost'], 2))
        
        # Gross profit
        worksheet.cell(row=current_row, column=4, value=round(store_data['gross_profit'], 2))
        
        # Profit margin percentage
        margin_cell = worksheet.cell(row=current_row, column=5, value=f"{store_data['profit_margin']:.1f}%")
        
        # Apply color based on margin
        if store_data['profit_margin'] < 50:
            margin_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
        elif store_data['profit_margin'] > 70:
            margin_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        else:
            margin_cell.fill = even_row_fill if i % 2 == 0 else odd_row_fill
        
        # Apply borders and alignment
        for col in range(1, 6):
            cell = worksheet.cell(row=current_row, column=col)
            cell.border = thin_border
            if col > 1:  # Right align numbers
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:  # Left align store name
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Apply alternating row fill (except for margin column which has conditional formatting)
            if col != 5:
                cell.fill = even_row_fill if i % 2 == 0 else odd_row_fill
        
        current_row += 1
    
    # Add summary row
    summary_row = current_row + 1
    worksheet.cell(row=summary_row, column=1, value="合计")
    worksheet.cell(row=summary_row, column=1).font = Font(bold=True)
    
    # Calculate totals
    total_revenue = sum(d['total_revenue'] for d in data)
    total_material_cost = sum(d['total_material_cost'] for d in data)
    total_profit = total_revenue - total_material_cost
    avg_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    # Write totals
    worksheet.cell(row=summary_row, column=2, value=round(total_revenue, 2))
    worksheet.cell(row=summary_row, column=3, value=round(total_material_cost, 2))
    worksheet.cell(row=summary_row, column=4, value=round(total_profit, 2))
    worksheet.cell(row=summary_row, column=5, value=f"{avg_margin:.1f}%")
    
    # Apply formatting to summary row
    for col in range(1, 6):
        cell = worksheet.cell(row=summary_row, column=col)
        cell.font = Font(bold=True)
        cell.border = thin_border
        if col > 1:
            cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Set column widths
    column_widths = [20, 18, 20, 18, 18]
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width
    
    # Add note about data source
    note_row = summary_row + 2
    note_cell = worksheet.cell(row=note_row, column=1, 
                              value="注：物料使用金额为实际库存消耗数据（仅成本类物料），非理论计算值")
    note_cell.font = Font(italic=True, size=10, color="666666")
    worksheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    
    logger.info("Completed writing all stores summary sheet")


if __name__ == "__main__":
    """Test the functions with command line arguments."""
    import argparse
    from openpyxl import Workbook
    
    parser = argparse.ArgumentParser(description='Test all stores summary sheet generation')
    parser.add_argument('--year', type=int, default=2025, help='Year (default: 2025)')
    parser.add_argument('--month', type=int, default=6, help='Month (default: 6)')
    parser.add_argument('--test-db', action='store_true', help='Use test database')
    
    args = parser.parse_args()
    
    # Create database connection
    from utils.database import DatabaseConfig
    db_config = DatabaseConfig(is_test=args.test_db)
    db_manager = DatabaseManager(db_config)
    
    # Create a test workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "全店铺汇总"
    
    # Generate the sheet
    write_all_stores_summary_sheet(
        worksheet=worksheet,
        db_manager=db_manager,
        year=args.year,
        month=args.month
    )
    
    # Save to test file
    output_file = f"test_all_stores_summary_{args.year}_{args.month:02d}.xlsx"
    workbook.save(output_file)
    print(f"Test file saved to: {output_file}")