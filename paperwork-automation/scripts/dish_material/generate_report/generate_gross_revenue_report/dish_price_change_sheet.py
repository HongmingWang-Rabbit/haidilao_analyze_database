import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal
from typing import Optional, Dict, Any
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

logger = logging.getLogger(__name__)

class DishPriceChangeSheet:
    def __init__(self, db_manager, target_date: str):
        self.db_manager = db_manager
        self.target_date = datetime.strptime(target_date, '%Y-%m-%d')
        self.sheet_name = "菜品价格变动及损耗表"

    def _get_date_ranges(self):
        """Calculate date ranges for current, last month, and last year same month"""
        current_start = self.target_date.replace(day=1)
        current_end = self.target_date

        last_month_end = current_start - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)

        last_year_start = current_start - relativedelta(years=1)
        last_year_end = self.target_date - relativedelta(years=1)

        return {
            'current': (current_start, current_end),
            'last_month': (last_month_start, last_month_end),
            'last_year': (last_year_start, last_year_end)
        }

    def _get_dish_prices_and_sales(self, date_range: tuple, store_id: int) -> Dict[int, Dict[str, Any]]:
        """Get dish prices and sales for a specific date range and store - using dish ID for unique identification"""
        start_date, end_date = date_range
        year = start_date.year
        month = start_date.month

        query = """
            WITH dish_data AS (
                SELECT
                    d.full_code as dish_code,
                    -- Use the first name for dishes with same full_code
                    MIN(d.name) as dish_name,
                    -- Aggregate IDs for dishes with same full_code
                    ARRAY_AGG(DISTINCT d.id) as dish_ids,
                    -- Sum quantities across all sizes
                    SUM(dms.sale_amount) as total_quantity,
                    -- Weighted average price
                    SUM(dms.sale_amount * COALESCE(
                        (SELECT dph.price
                         FROM dish_price_history dph
                         WHERE dph.dish_id = d.id
                           AND dph.store_id = %s
                           AND ((dph.effective_year < %s) OR
                                (dph.effective_year = %s AND dph.effective_month <= %s))
                         ORDER BY dph.effective_year DESC, dph.effective_month DESC
                         LIMIT 1),
                        0
                    )) / NULLIF(SUM(dms.sale_amount), 0) as avg_price
                FROM dish d
                INNER JOIN dish_monthly_sale dms ON d.id = dms.dish_id
                WHERE dms.store_id = %s
                    AND dms.year = %s
                    AND dms.month = %s
                GROUP BY d.full_code
            ),
            material_usage AS (
                SELECT
                    dd.dish_code,
                    -- Calculate theoretical material cost - sum for all dishes in this dish_code group
                    SUM(
                        CASE WHEN EXISTS (
                            SELECT 1 FROM material_monthly_usage mmu
                            WHERE mmu.material_id = m.id
                              AND mmu.store_id = %s
                              AND mmu.year = %s
                              AND mmu.month = %s
                              AND mmu.material_use_type = '成本类'
                        )
                        THEN
                            -- Theoretical cost calculation
                            COALESCE(dms.sale_amount, 0) *
                            dm.standard_quantity * COALESCE(dm.loss_rate, 0) /
                            COALESCE(NULLIF(dm.unit_conversion_rate, 0), 1) * COALESCE(mph.price, 0)
                        ELSE 0
                        END
                    ) as theoretical_cost,
                    -- Calculate ACTUAL material cost using actual usage from inventory
                    SUM(
                        CASE WHEN EXISTS (
                            SELECT 1 FROM material_monthly_usage mmu
                            WHERE mmu.material_id = m.id
                              AND mmu.store_id = %s
                              AND mmu.year = %s
                              AND mmu.month = %s
                              AND mmu.material_use_type = '成本类'
                        )
                        THEN
                            -- Calculate this dish's proportion of total theoretical usage for this material
                            -- Then multiply by actual usage from inventory
                            CASE
                                WHEN (
                                    -- Total theoretical usage for this material across all dishes
                                    SELECT SUM(dms2.sale_amount * dm2.standard_quantity * COALESCE(dm2.loss_rate, 0) /
                                               COALESCE(NULLIF(dm2.unit_conversion_rate, 0), 1))
                                    FROM dish_monthly_sale dms2
                                    JOIN dish_material dm2 ON dm2.dish_id = dms2.dish_id AND dm2.store_id = dms2.store_id
                                    WHERE dm2.material_id = m.id
                                      AND dms2.store_id = %s
                                      AND dms2.year = %s
                                      AND dms2.month = %s
                                ) > 0
                                THEN
                                    -- This dish's theoretical usage
                                    (COALESCE(dms.sale_amount, 0) * dm.standard_quantity * COALESCE(dm.loss_rate, 0) /
                                     COALESCE(NULLIF(dm.unit_conversion_rate, 0), 1)) /
                                    -- Divided by total theoretical usage
                                    (SELECT SUM(dms2.sale_amount * dm2.standard_quantity * COALESCE(dm2.loss_rate, 0) /
                                                COALESCE(NULLIF(dm2.unit_conversion_rate, 0), 1))
                                     FROM dish_monthly_sale dms2
                                     JOIN dish_material dm2 ON dm2.dish_id = dms2.dish_id AND dm2.store_id = dms2.store_id
                                     WHERE dm2.material_id = m.id
                                       AND dms2.store_id = %s
                                       AND dms2.year = %s
                                       AND dms2.month = %s) *
                                    -- Multiplied by actual usage from inventory
                                    COALESCE((SELECT mmu.material_used
                                              FROM material_monthly_usage mmu
                                              WHERE mmu.material_id = m.id
                                                AND mmu.store_id = %s
                                                AND mmu.year = %s
                                                AND mmu.month = %s), 0) *
                                    -- Multiplied by material price
                                    COALESCE(mph.price, 0)
                                ELSE 0
                            END
                        ELSE 0
                        END
                    ) as actual_cost,
                    -- Concatenate unique material names (without ORDER BY due to DISTINCT)
                    STRING_AGG(
                        DISTINCT CASE WHEN EXISTS (
                            SELECT 1 FROM material_monthly_usage mmu
                            WHERE mmu.material_id = m.id
                              AND mmu.store_id = %s
                              AND mmu.year = %s
                              AND mmu.month = %s
                              AND mmu.material_use_type = '成本类'
                        )
                        THEN m.description
                        ELSE NULL
                        END, ', '
                    ) as materials_used
                FROM dish_data dd
                CROSS JOIN UNNEST(dd.dish_ids) as unnested_dish_id
                LEFT JOIN dish_monthly_sale dms ON dms.dish_id = unnested_dish_id
                    AND dms.store_id = %s
                    AND dms.year = %s
                    AND dms.month = %s
                LEFT JOIN dish_material dm ON dm.dish_id = unnested_dish_id AND dm.store_id = %s
                LEFT JOIN material m ON m.id = dm.material_id AND m.store_id = %s
                LEFT JOIN material_price_history mph ON mph.material_id = m.id
                    AND mph.store_id = %s
                    AND mph.is_active = TRUE
                GROUP BY dd.dish_code
            )
            SELECT
                dd.dish_code,
                dd.dish_name,
                dd.dish_ids,
                dd.avg_price,
                dd.total_quantity,
                COALESCE(dd.avg_price, 0) * dd.total_quantity as total_revenue,
                COALESCE(mu.theoretical_cost, 0) as theoretical_cost,
                COALESCE(mu.actual_cost, 0) as actual_cost,
                COALESCE(mu.materials_used, '') as materials_used
            FROM dish_data dd
            LEFT JOIN material_usage mu ON mu.dish_code = dd.dish_code
            WHERE dd.total_quantity > 0
        """

        try:
            with self.db_manager.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query, (
                        # dish_data CTE price subquery
                        store_id, year, year, month,
                        # dish_data CTE main query
                        store_id, year, month,
                        # material_usage CTE - first EXISTS for theoretical_cost
                        store_id, year, month,
                        # material_usage CTE - second EXISTS for actual_cost
                        store_id, year, month,
                        # material_usage CTE - actual_cost subqueries
                        store_id, year, month,  # First total theoretical subquery
                        store_id, year, month,  # Second total theoretical subquery
                        store_id, year, month,  # Actual usage subquery
                        # material_usage CTE - third EXISTS for STRING_AGG
                        store_id, year, month,
                        # material_usage CTE - JOINs
                        store_id, year, month,  # dish_monthly_sale
                        store_id, store_id, store_id  # dish_material, material, material_price_history
                    ))
                    results = cursor.fetchall()

                    dish_data = {}
                    for row in results:
                        # Use dish_code as the key for merged dishes
                        dish_data[row['dish_code']] = {
                            'dish_code': row['dish_code'],
                            'dish_name': row['dish_name'],
                            'dish_ids': row['dish_ids'],  # Array of dish IDs that were merged
                            'avg_price': float(row['avg_price']) if row['avg_price'] else 0,
                            'total_quantity': float(row['total_quantity']) if row['total_quantity'] else 0,
                            'total_revenue': float(row['total_revenue']) if row['total_revenue'] else 0,
                            'theoretical_cost': float(row['theoretical_cost']) if row['theoretical_cost'] else 0,
                            'actual_cost': float(row['actual_cost']) if row['actual_cost'] else 0,
                            'materials_used': row['materials_used'] if row['materials_used'] else ''
                        }

                    return dish_data
        except Exception as e:
            logger.error(f"Error getting dish prices and sales: {e}")
            return {}

    def _get_actual_theoretical_ratio(self, year: int, month: int, store_id: int) -> float:
        """Get the ratio of actual to theoretical material usage for the store.

        This ratio represents the actual efficiency/waste factor.
        Returns 1.0 if no data available (assumes no extra waste).
        """
        query = """
            WITH theoretical_usage AS (
                -- Calculate total theoretical material usage
                SELECT
                    m.id as material_id,
                    SUM(
                        dms.sale_amount *
                        dm.standard_quantity * COALESCE(dm.loss_rate, 0) /
                        COALESCE(NULLIF(dm.unit_conversion_rate, 0), 1)
                    ) as theoretical_qty
                FROM dish_monthly_sale dms
                JOIN dish_material dm ON dm.dish_id = dms.dish_id AND dm.store_id = dms.store_id
                JOIN material m ON m.id = dm.material_id AND m.store_id = dm.store_id
                WHERE dms.store_id = %s
                  AND dms.year = %s
                  AND dms.month = %s
                  AND EXISTS (
                      SELECT 1 FROM material_monthly_usage mmu
                      WHERE mmu.material_id = m.id
                        AND mmu.store_id = %s
                        AND mmu.year = %s
                        AND mmu.month = %s
                        AND mmu.material_use_type = '成本类'
                  )
                GROUP BY m.id
            ),
            actual_usage AS (
                -- Get actual usage from inventory data
                SELECT
                    mmu.material_id,
                    mmu.material_used as actual_qty
                FROM material_monthly_usage mmu
                WHERE mmu.year = %s
                  AND mmu.month = %s
                  AND mmu.store_id = %s
                  AND mmu.material_use_type = '成本类'
            )
            SELECT
                SUM(COALESCE(au.actual_qty, 0) * COALESCE(mph.price, 0)) as total_actual_cost,
                SUM(COALESCE(tu.theoretical_qty, 0) * COALESCE(mph.price, 0)) as total_theoretical_cost
            FROM theoretical_usage tu
            LEFT JOIN actual_usage au ON au.material_id = tu.material_id
            LEFT JOIN material_price_history mph ON mph.material_id = tu.material_id
                AND mph.store_id = %s
                AND mph.is_active = TRUE
        """

        try:
            with self.db_manager.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query, (
                        store_id, year, month,  # for theoretical_usage
                        store_id, year, month,  # for EXISTS check
                        year, month, store_id,  # for actual_usage
                        store_id  # for price history
                    ))
                    result = cursor.fetchone()

                    if result and result['total_theoretical_cost'] and result['total_theoretical_cost'] > 0:
                        # Return the ratio of actual to theoretical
                        actual = result['total_actual_cost'] or 0
                        theoretical = result['total_theoretical_cost']
                        ratio = actual / theoretical
                        logger.info(f"Store {store_id}: Actual/Theoretical ratio = {ratio:.4f}")
                        return ratio
                    else:
                        logger.warning(f"No theoretical cost data for store {store_id}")
                        return 1.0  # Default to no extra waste
        except Exception as e:
            logger.error(f"Error calculating actual/theoretical ratio: {e}")
            return 1.0

    def _get_all_dish_materials(self, store_id: int, year: int, month: int) -> Dict[str, list]:
        """Get materials for all dishes grouped by full_code"""
        query = """
            SELECT
                d.full_code as dish_code,
                m.material_number,
                m.description as material_name,
                COALESCE(mph.price, 0) as material_price,
                AVG(dm.standard_quantity * COALESCE(dm.loss_rate, 0) /
                    COALESCE(NULLIF(dm.unit_conversion_rate, 0), 1)) as usage_per_dish
            FROM dish_material dm
            JOIN dish d ON d.id = dm.dish_id
            JOIN material m ON m.id = dm.material_id AND m.store_id = dm.store_id
            LEFT JOIN material_price_history mph ON mph.material_id = m.id
                AND mph.store_id = m.store_id
                AND mph.is_active = TRUE
            WHERE dm.store_id = %s
              AND EXISTS (
                  SELECT 1 FROM material_monthly_usage mmu
                  WHERE mmu.material_id = m.id
                    AND mmu.store_id = %s
                    AND mmu.year = %s
                    AND mmu.month = %s
                    AND mmu.material_use_type = '成本类'
              )
            GROUP BY d.full_code, m.material_number, m.description, mph.price
            ORDER BY d.full_code, m.material_number
        """

        try:
            with self.db_manager.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query, (store_id, store_id, year, month))
                    results = cursor.fetchall()

                    dish_materials = {}
                    for row in results:
                        dish_code = row['dish_code']
                        if dish_code not in dish_materials:
                            dish_materials[dish_code] = []

                        # Check if this material already exists for this dish_code
                        existing = False
                        for mat in dish_materials[dish_code]:
                            if mat['material_number'] == row['material_number']:
                                existing = True
                                break

                        if not existing:
                            dish_materials[dish_code].append({
                                'material_number': row['material_number'],
                                'material_name': row['material_name'],
                                'material_price': float(row['material_price']) if row['material_price'] else 0,
                                'usage_per_dish': float(row['usage_per_dish']) if row['usage_per_dish'] else 0
                            })
                    return dish_materials
        except Exception as e:
            logger.error(f"Error getting dish materials: {e}")
            return {}

    def _get_store_data(self, store_id: int, store_name: str) -> list:
        """Get dish price change data for a specific store"""
        date_ranges = self._get_date_ranges()

        current_data = self._get_dish_prices_and_sales(date_ranges['current'], store_id)
        last_month_data = self._get_dish_prices_and_sales(date_ranges['last_month'], store_id)
        last_year_data = self._get_dish_prices_and_sales(date_ranges['last_year'], store_id)

        # Get current year/month for fetching materials
        current_year = date_ranges['current'][0].year
        current_month = date_ranges['current'][0].month

        # Get all dish materials at once for better performance
        all_dish_materials = self._get_all_dish_materials(store_id, current_year, current_month)

        # Get all dish codes that appear in any of the three periods (now grouped by full_code)
        all_dish_codes = set(current_data.keys()) | set(last_month_data.keys()) | set(last_year_data.keys())

        store_rows = []
        for dish_code in all_dish_codes:
            current = current_data.get(dish_code, {})
            last_month = last_month_data.get(dish_code, {})
            last_year = last_year_data.get(dish_code, {})

            # Skip if no current data
            if not current:
                continue

            current_price = current.get('avg_price', 0)
            last_month_price = last_month.get('avg_price', 0)
            last_year_price = last_year.get('avg_price', 0)
            current_quantity = current.get('total_quantity', 0)
            current_revenue = current.get('total_revenue', 0)

            # Calculate price changes
            # Only include actual price changes, not new dishes
            mom_price_change = current_price - last_month_price if last_month_price else 0
            yoy_price_change = current_price - last_year_price if last_year_price else 0

            # Calculate revenue impact of price changes
            mom_revenue_impact = mom_price_change * current_quantity
            yoy_revenue_impact = yoy_price_change * current_quantity

            # Get dish code and name from whichever period has data
            dish_code = current.get('dish_code', last_month.get('dish_code', last_year.get('dish_code', '')))
            dish_name = current.get('dish_name', last_month.get('dish_name', last_year.get('dish_name', '')))

            # Skip dish code 14120001
            if dish_code == '14120001':
                continue

            # Get theoretical and actual costs for this dish
            theoretical_cost = current.get('theoretical_cost', 0)
            actual_cost = current.get('actual_cost', 0)  # Now comes directly from query with material-level ratios

            # Calculate loss (actual - theoretical)
            # Positive means more usage than expected (waste)
            # Negative means less usage than expected (efficiency gain)
            loss_amount = actual_cost - theoretical_cost

            # Get materials list for this dish from pre-fetched data
            materials_list = all_dish_materials.get(dish_code, [])

            row = {
                '门店': store_name,
                '菜品编码': dish_code,
                '菜品名称': dish_name,
                '本期单价': current_price,
                '上期单价': last_month_price,
                '去年同期单价': last_year_price,
                '环比价格变动': mom_price_change,
                '同比价格变动': yoy_price_change,
                '本期销量': current_quantity,
                '本期收入': current_revenue,
                '环比影响收入': mom_revenue_impact,
                '同比影响收入': yoy_revenue_impact,
                'materials': materials_list,  # Store materials list for later processing
                '理论成本': theoretical_cost,
                '实际成本': actual_cost,
                '损耗金额': loss_amount
            }

            store_rows.append(row)

        # Sort by absolute revenue impact (descending)
        return sorted(store_rows, key=lambda x: abs(x['环比影响收入']), reverse=True)

    def generate_sheet(self, workbook):
        """Generate the dish price change sheet"""
        ws = workbook.create_sheet(self.sheet_name)

        # Define stores
        stores = [
            (1, '加拿大一店'),
            (2, '加拿大二店'),
            (3, '加拿大三店'),
            (4, '加拿大四店'),
            (5, '加拿大五店'),
            (6, '加拿大六店'),
            (7, '加拿大七店')
        ]

        # Headers
        headers = [
            '门店', '菜品编码', '菜品名称', '本期单价', '上期单价',
            '去年同期单价', '环比价格变动', '同比价格变动', '本期销量',
            '本期收入', '环比影响收入', '同比影响收入', '使用物料',
            '理论成本', '实际成本', '损耗金额'
        ]

        # Style definitions
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Collect all data
        all_data = []
        for store_id, store_name in stores:
            logger.info(f"Processing store: {store_name}")
            store_data = self._get_store_data(store_id, store_name)
            all_data.extend(store_data)

        # Write data rows
        row_num = 2
        for data in all_data:
            # Determine number of rows needed (one per material or at least 1)
            materials = data.get('materials', [])
            num_rows = max(len(materials), 1)

            # Store name (merge cells if multiple materials)
            cell = ws.cell(row=row_num, column=1, value=data['门店'])
            cell.border = border
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=1,
                             end_row=row_num + num_rows - 1, end_column=1)
                cell.alignment = Alignment(vertical='center')

            # Dish code (merge cells if multiple materials)
            cell = ws.cell(row=row_num, column=2, value=data['菜品编码'])
            cell.border = border
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=2,
                             end_row=row_num + num_rows - 1, end_column=2)
                cell.alignment = Alignment(vertical='center')

            # Dish name (merge cells if multiple materials)
            cell = ws.cell(row=row_num, column=3, value=data['菜品名称'])
            cell.border = border
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=3,
                             end_row=row_num + num_rows - 1, end_column=3)
                cell.alignment = Alignment(vertical='center', wrap_text=True)

            # Current price (merge cells if multiple materials)
            cell = ws.cell(row=row_num, column=4, value=round(data['本期单价'], 2))
            cell.border = border
            cell.number_format = '#,##0.00'
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=4,
                             end_row=row_num + num_rows - 1, end_column=4)
                cell.alignment = Alignment(vertical='center', horizontal='right')

            # Last month price (merge cells if multiple materials)
            cell = ws.cell(row=row_num, column=5, value=round(data['上期单价'], 2))
            cell.border = border
            cell.number_format = '#,##0.00'
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=5,
                             end_row=row_num + num_rows - 1, end_column=5)
                cell.alignment = Alignment(vertical='center', horizontal='right')

            # Last year price (merge cells if multiple materials)
            cell = ws.cell(row=row_num, column=6, value=round(data['去年同期单价'], 2))
            cell.border = border
            cell.number_format = '#,##0.00'
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=6,
                             end_row=row_num + num_rows - 1, end_column=6)
                cell.alignment = Alignment(vertical='center', horizontal='right')

            # MoM price change (merge cells if multiple materials)
            mom_change_cell = ws.cell(row=row_num, column=7, value=round(data['环比价格变动'], 2))
            mom_change_cell.border = border
            mom_change_cell.number_format = '#,##0.00'
            if data['环比价格变动'] > 0:
                mom_change_cell.font = Font(color="FF0000")  # Red for increase
            elif data['环比价格变动'] < 0:
                mom_change_cell.font = Font(color="008000")  # Green for decrease
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=7,
                             end_row=row_num + num_rows - 1, end_column=7)
                mom_change_cell.alignment = Alignment(vertical='center', horizontal='right')

            # YoY price change (merge cells if multiple materials)
            yoy_change_cell = ws.cell(row=row_num, column=8, value=round(data['同比价格变动'], 2))
            yoy_change_cell.border = border
            yoy_change_cell.number_format = '#,##0.00'
            if data['同比价格变动'] > 0:
                yoy_change_cell.font = Font(color="FF0000")  # Red for increase
            elif data['同比价格变动'] < 0:
                yoy_change_cell.font = Font(color="008000")  # Green for decrease
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=8,
                             end_row=row_num + num_rows - 1, end_column=8)
                yoy_change_cell.alignment = Alignment(vertical='center', horizontal='right')

            # Current quantity (merge cells if multiple materials)
            cell = ws.cell(row=row_num, column=9, value=round(data['本期销量'], 0))
            cell.border = border
            cell.number_format = '#,##0'
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=9,
                             end_row=row_num + num_rows - 1, end_column=9)
                cell.alignment = Alignment(vertical='center', horizontal='right')

            # Current revenue (merge cells if multiple materials)
            cell = ws.cell(row=row_num, column=10, value=round(data['本期收入'], 2))
            cell.border = border
            cell.number_format = '#,##0.00'
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=10,
                             end_row=row_num + num_rows - 1, end_column=10)
                cell.alignment = Alignment(vertical='center', horizontal='right')

            # MoM revenue impact (merge cells if multiple materials)
            mom_impact_cell = ws.cell(row=row_num, column=11, value=round(data['环比影响收入'], 2))
            mom_impact_cell.border = border
            mom_impact_cell.number_format = '#,##0.00'
            if data['环比影响收入'] > 0:
                mom_impact_cell.font = Font(color="FF0000")  # Red for increase
            elif data['环比影响收入'] < 0:
                mom_impact_cell.font = Font(color="008000")  # Green for decrease
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=11,
                             end_row=row_num + num_rows - 1, end_column=11)
                mom_impact_cell.alignment = Alignment(vertical='center', horizontal='right')

            # YoY revenue impact (merge cells if multiple materials)
            yoy_impact_cell = ws.cell(row=row_num, column=12, value=round(data['同比影响收入'], 2))
            yoy_impact_cell.border = border
            yoy_impact_cell.number_format = '#,##0.00'
            if data['同比影响收入'] > 0:
                yoy_impact_cell.font = Font(color="FF0000")  # Red for increase
            elif data['同比影响收入'] < 0:
                yoy_impact_cell.font = Font(color="008000")  # Green for decrease
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=12,
                             end_row=row_num + num_rows - 1, end_column=12)
                yoy_impact_cell.alignment = Alignment(vertical='center', horizontal='right')

            # Materials used (one per row)
            if materials:
                for i, material in enumerate(materials):
                    # Calculate theoretical cost for this material
                    material_theoretical_cost = material['usage_per_dish'] * data['本期销量'] * material['material_price']

                    # Calculate proportional actual cost for this material
                    if data['理论成本'] > 0:
                        material_actual_cost = (material_theoretical_cost / data['理论成本']) * data['实际成本']
                    else:
                        material_actual_cost = 0

                    # Format material info
                    material_text = (f"{material['material_name']}: "
                                   f"理论{material_theoretical_cost:.2f} "
                                   f"实际{material_actual_cost:.2f}")

                    cell = ws.cell(row=row_num + i, column=13, value=material_text)
                    cell.border = border
            else:
                cell = ws.cell(row=row_num, column=13, value="无物料")
                cell.border = border

            # Theoretical cost (merge cells if multiple materials)
            cell = ws.cell(row=row_num, column=14, value=round(data['理论成本'], 2))
            cell.border = border
            cell.number_format = '#,##0.00'
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=14,
                             end_row=row_num + num_rows - 1, end_column=14)
                cell.alignment = Alignment(vertical='center', horizontal='right')

            # Actual cost (merge cells if multiple materials)
            cell = ws.cell(row=row_num, column=15, value=round(data['实际成本'], 2))
            cell.border = border
            cell.number_format = '#,##0.00'
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=15,
                             end_row=row_num + num_rows - 1, end_column=15)
                cell.alignment = Alignment(vertical='center', horizontal='right')

            # Loss amount (merge cells if multiple materials)
            loss_cell = ws.cell(row=row_num, column=16, value=round(data['损耗金额'], 2))
            loss_cell.border = border
            loss_cell.number_format = '#,##0.00'
            if data['损耗金额'] > 0:
                loss_cell.font = Font(color="FF0000")  # Red for loss
            elif data['损耗金额'] < 0:
                loss_cell.font = Font(color="008000")  # Green for gain
            if num_rows > 1:
                ws.merge_cells(start_row=row_num, start_column=16,
                             end_row=row_num + num_rows - 1, end_column=16)
                loss_cell.alignment = Alignment(vertical='center', horizontal='right')

            # Apply borders to all cells in the range
            for row in range(row_num, row_num + num_rows):
                for col in range(1, 17):
                    ws.cell(row=row, column=col).border = border

            row_num += num_rows

        # Set column widths
        column_widths = [15, 15, 30, 12, 12, 15, 15, 15, 12, 15, 15, 15, 40, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze panes (keep header visible)
        ws.freeze_panes = 'A2'

        logger.info(f"Generated {self.sheet_name} with {row_num - 2} rows of data")
        return ws