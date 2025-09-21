#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate 12-month trend sheet showing actual gross profit margins over time.
Uses actual material usage from material_monthly_usage table (same as all stores summary).
"""

import sys
from pathlib import Path
from typing import Dict, List, Tuple
import logging
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Add parent directory to path for imports
project_root = Path(__file__).resolve().parent.parent.parent.parent.parent
sys.path.insert(0, str(project_root))

from utils.database import DatabaseManager
from lib.config import STORE_ID_TO_NAME_MAPPING

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def get_12_month_trend_data(db_manager: DatabaseManager, end_year: int, end_month: int) -> Dict:
    """
    Get 12 months of overall gross margin data for all stores using actual material usage.
    
    Args:
        db_manager: Database manager instance
        end_year: Ending year
        end_month: Ending month
    
    Returns:
        Dict with structure:
        {
            store_id: {
                (year, month): margin_percentage
            }
        }
    """
    # Calculate the 12-month period
    end_date = datetime(end_year, end_month, 1)
    start_date = end_date - relativedelta(months=11)
    
    result = {}
    
    with db_manager.get_connection() as conn:
        cursor = conn.cursor()
        
        # Query to get overall gross margin for 12 months using actual material usage
        query = """
        WITH store_revenue AS (
            -- Get total revenue from dish sales for each month
            SELECT 
                dms.store_id,
                dms.year,
                dms.month,
                SUM((COALESCE(dms.sale_amount, 0) - COALESCE(dms.return_amount, 0)) * COALESCE(dph.price, 0)) as total_revenue
            FROM dish_monthly_sale dms
            LEFT JOIN dish_price_history dph ON dph.dish_id = dms.dish_id 
                AND dph.store_id = dms.store_id
                AND dph.is_active = TRUE
            WHERE (
                (dms.year = %s AND dms.month >= %s) OR
                (dms.year > %s AND dms.year < %s) OR
                (dms.year = %s AND dms.month <= %s)
            )
            AND dms.store_id != 101  -- Exclude Hi Bowl
            GROUP BY dms.store_id, dms.year, dms.month
        ),
        material_usage AS (
            -- Get actual material usage from material_monthly_usage for each month
            SELECT 
                mmu.store_id,
                mmu.year,
                mmu.month,
                SUM(COALESCE(mmu.material_used, 0) * COALESCE(mph.price, 0)) as total_material_cost
            FROM material_monthly_usage mmu
            JOIN material m ON m.id = mmu.material_id AND m.store_id = mmu.store_id
            LEFT JOIN material_price_history mph ON mph.material_id = m.id 
                AND mph.store_id = m.store_id
                AND mph.is_active = TRUE
            WHERE (
                (mmu.year = %s AND mmu.month >= %s) OR
                (mmu.year > %s AND mmu.year < %s) OR
                (mmu.year = %s AND mmu.month <= %s)
            )
            GROUP BY mmu.store_id, mmu.year, mmu.month
        )
        SELECT 
            sr.store_id,
            sr.year,
            sr.month,
            COALESCE(sr.total_revenue, 0) as revenue,
            COALESCE(mu.total_material_cost, 0) as material_cost,
            CASE 
                WHEN COALESCE(sr.total_revenue, 0) > 0 THEN 
                    ((COALESCE(sr.total_revenue, 0) - COALESCE(mu.total_material_cost, 0)) / COALESCE(sr.total_revenue, 0)) * 100
                ELSE 0
            END as gross_margin
        FROM store_revenue sr
        LEFT JOIN material_usage mu ON mu.store_id = sr.store_id 
            AND mu.year = sr.year 
            AND mu.month = sr.month
        ORDER BY sr.store_id, sr.year, sr.month
        """
        
        cursor.execute(query, (
            # For store_revenue CTE
            start_date.year, start_date.month,  # Start year and month
            start_date.year, end_date.year,     # Years in between
            end_date.year, end_date.month,      # End year and month
            # For material_usage CTE
            start_date.year, start_date.month,  # Start year and month
            start_date.year, end_date.year,     # Years in between
            end_date.year, end_date.month       # End year and month
        ))
        
        raw_data = cursor.fetchall()
        
        # Process data into nested dictionary
        for row in raw_data:
            store_id = row['store_id']
            year = row['year']
            month = row['month']
            margin = float(row['gross_margin']) if row['gross_margin'] else 0
            
            # Initialize store dict if not exists
            if store_id not in result:
                result[store_id] = {}
            
            # Store margin for this month
            result[store_id][(year, month)] = margin
    
    return result


def write_twelve_month_trend_sheet(worksheet, db_manager: DatabaseManager, year: int, month: int):
    """
    Write 12-month overall trend data to an Excel worksheet with line chart.
    
    Args:
        worksheet: Openpyxl worksheet object
        db_manager: Database manager instance
        year: Ending year for the data
        month: Ending month for the data
    """
    # Add title
    title_cell = worksheet.cell(row=1, column=1, value="12个月实际毛利趋势分析")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    date_cell = worksheet.cell(row=2, column=1, value=f"截至{year}年{month}月")
    date_cell.font = Font(size=12)
    date_cell.alignment = Alignment(horizontal='center')
    
    # Get data
    trend_data = get_12_month_trend_data(db_manager, year, month)
    
    # Calculate months to display
    end_date = datetime(year, month, 1)
    months = []
    for i in range(11, -1, -1):
        date = end_date - relativedelta(months=i)
        months.append((date.year, date.month))
    
    # Merge title cells
    num_cols = 1 + len(months)  # Store/Category + 12 months
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    
    # Style definitions
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Colors for performance indicators
    excellent_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Dark green
    good_fill = PatternFill(start_color="E7F5E7", end_color="E7F5E7", fill_type="solid")       # Light green
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")    # Yellow
    poor_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")        # Light red
    
    # Create headers
    header_row = 4
    
    # Store header
    worksheet.cell(row=header_row, column=1, value="店铺")
    cell = worksheet.cell(row=header_row, column=1)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    
    # Month headers
    current_col = 2
    for year_month in months:
        month_str = f"{year_month[0]}-{year_month[1]:02d}"
        cell = worksheet.cell(row=header_row, column=current_col, value=month_str)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        current_col += 1
    
    # Data rows
    data_row = header_row + 1
    
    # Process each store - show overall margins only
    for store_id in sorted(trend_data.keys()):
        store_name = STORE_ID_TO_NAME_MAPPING.get(store_id, f"Store {store_id}")
        
        # Store name
        cell = worksheet.cell(row=data_row, column=1, value=store_name)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left')
        
        # Monthly data with percentage changes
        current_col = 2
        previous_margin = None
        
        for idx, year_month in enumerate(months):
            margin = trend_data[store_id].get(year_month, 0)
            
            if margin > 0:
                # Calculate percentage change
                if previous_margin is not None and previous_margin > 0:
                    change = ((margin - previous_margin) / previous_margin) * 100
                    
                    # Format based on change magnitude
                    if abs(change) < 0.5:
                        # Small change - just show margin
                        cell_value = f"{margin:.1f}%"
                        cell_fill = None
                    elif change > 0:
                        # Positive change
                        cell_value = f"{margin:.1f}% ↑{change:.1f}%"
                        cell_fill = good_fill
                    else:
                        # Negative change
                        cell_value = f"{margin:.1f}% ↓{abs(change):.1f}%"
                        cell_fill = warning_fill
                else:
                    # First month or no previous data
                    cell_value = f"{margin:.1f}%"
                    cell_fill = None
                
                cell = worksheet.cell(row=data_row, column=current_col, value=cell_value)
                
                # Apply color based on margin level if no change color
                if cell_fill:
                    cell.fill = cell_fill
                elif margin >= 75:
                    cell.fill = excellent_fill
                elif margin >= 65:
                    cell.fill = good_fill
                elif margin < 50:
                    cell.fill = poor_fill
            else:
                cell = worksheet.cell(row=data_row, column=current_col, value="-")
            
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border
            
            previous_margin = margin if margin > 0 else None
            current_col += 1
        
        data_row += 1
    
    # Add overall average row
    avg_row = data_row + 1
    cell = worksheet.cell(row=avg_row, column=1, value="全店平均")
    cell.font = Font(bold=True)
    cell.border = thin_border
    
    # Calculate and display overall averages
    current_col = 2
    for year_month in months:
        # Calculate average across all stores for this month
        margins = []
        for store_id in trend_data:
            if year_month in trend_data[store_id]:
                margin = trend_data[store_id][year_month]
                if margin > 0:
                    margins.append(margin)
        
        if margins:
            avg_margin = sum(margins) / len(margins)
            cell = worksheet.cell(row=avg_row, column=current_col, value=f"{avg_margin:.1f}%")
            
            # Color code based on average
            if avg_margin >= 70:
                cell.fill = good_fill
            elif avg_margin < 50:
                cell.fill = warning_fill
        else:
            cell = worksheet.cell(row=avg_row, column=current_col, value="-")
        
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border
        cell.font = Font(bold=True)
        current_col += 1
    
    # Auto-adjust column widths
    worksheet.column_dimensions[get_column_letter(1)].width = 18  # Store/Category
    for col in range(2, num_cols + 1):
        worksheet.column_dimensions[get_column_letter(col)].width = 10
    
    # Add legend
    legend_row = avg_row + 2
    legend_text = "说明：↑环比上升（绿色）  ↓环比下降（黄色）  深绿≥75%  红色<50%  数字为环比变化百分比"
    legend_cell = worksheet.cell(row=legend_row, column=1, value=legend_text)
    legend_cell.font = Font(italic=True, size=10)
    legend_cell.alignment = Alignment(horizontal='left')
    worksheet.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=8)
    
    logger.info("Completed writing 12-month trend sheet")


if __name__ == "__main__":
    """Test the functions with command line arguments."""
    import argparse
    from openpyxl import Workbook
    
    parser = argparse.ArgumentParser(description='Test 12-month trend sheet generation')
    parser.add_argument('--year', type=int, default=2025, help='End year (default: 2025)')
    parser.add_argument('--month', type=int, default=6, help='End month (default: 6)')
    parser.add_argument('--test-db', action='store_true', help='Use test database')
    
    args = parser.parse_args()
    
    # Create database connection
    from utils.database import DatabaseConfig
    db_config = DatabaseConfig(is_test=args.test_db)
    db_manager = DatabaseManager(db_config)
    
    # Create a test workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "12月趋势"
    
    # Generate the sheet
    write_twelve_month_trend_sheet(
        worksheet=worksheet,
        db_manager=db_manager,
        year=args.year,
        month=args.month
    )
    
    # Save to test file
    output_file = f"test_12_month_trend_{args.year}_{args.month:02d}.xlsx"
    workbook.save(output_file)
    print(f"Test file saved to: {output_file}")