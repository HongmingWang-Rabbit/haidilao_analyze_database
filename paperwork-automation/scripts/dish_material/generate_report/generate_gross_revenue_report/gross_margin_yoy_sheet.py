import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal
from typing import Optional, Dict, Any, Tuple
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

logger = logging.getLogger(__name__)

class GrossMarginYoYSheet:
    def __init__(self, db_manager, target_date: str):
        self.db_manager = db_manager
        self.target_date = datetime.strptime(target_date, '%Y-%m-%d')
        self.sheet_name = "毛利率同比"

    def _get_date_ranges(self):
        """Calculate date ranges for current and last year same month"""
        current_start = self.target_date.replace(day=1)
        current_end = self.target_date

        last_year_start = current_start - relativedelta(years=1)
        last_year_end = self.target_date - relativedelta(years=1)

        return {
            'current': (current_start, current_end),
            'last_year': (last_year_start, last_year_end)
        }

    def _get_revenue_and_cost(self, start_date: datetime, end_date: datetime, store_id: int) -> Dict[str, float]:
        """Get revenue and cost data for a specific period"""
        year = start_date.year
        month = start_date.month

        # Get revenue from dish_monthly_sale (same as summary sheet)
        revenue_query = """
            SELECT
                SUM((COALESCE(dms.sale_amount, 0) - COALESCE(dms.return_amount, 0)) * COALESCE(dph.price, 0)) as total_revenue
            FROM dish_monthly_sale dms
            LEFT JOIN dish_price_history dph ON dph.dish_id = dms.dish_id
                AND dph.store_id = dms.store_id
                AND dph.is_active = TRUE
            WHERE dms.store_id = %s
                AND dms.year = %s
                AND dms.month = %s
        """

        # Get discount from daily_report (separate query)
        discount_query = """
            SELECT
                SUM(CAST(discount_total AS DECIMAL(15,2))) as total_discount
            FROM daily_report
            WHERE store_id = %s
                AND date >= %s
                AND date <= %s
        """

        # Get material cost (only 成本类 materials, same as summary sheet)
        material_cost_query = """
            SELECT
                SUM(COALESCE(mmu.material_used, 0) * COALESCE(mph.price, 0)) as total_material_cost
            FROM material_monthly_usage mmu
            JOIN material m ON m.id = mmu.material_id AND m.store_id = mmu.store_id
            LEFT JOIN material_price_history mph ON mph.material_id = m.id
                AND mph.store_id = m.store_id
                AND mph.is_active = TRUE
            WHERE mmu.store_id = %s
                AND mmu.year = %s
                AND mmu.month = %s
                AND mmu.material_use_type = '成本类'  -- Only include cost-type materials
        """

        try:
            with self.db_manager.get_connection() as conn:
                with conn.cursor() as cursor:
                    # Get revenue data
                    cursor.execute(revenue_query, (store_id, year, month))
                    revenue_result = cursor.fetchone()
                    total_revenue = float(revenue_result['total_revenue']) if revenue_result['total_revenue'] else 0

                    # Get discount data from daily_report
                    cursor.execute(discount_query, (store_id, start_date, end_date))
                    discount_result = cursor.fetchone()
                    total_discount = float(discount_result['total_discount']) if discount_result and discount_result['total_discount'] else 0

                    # Get material cost
                    cursor.execute(material_cost_query, (store_id, year, month))
                    cost_result = cursor.fetchone()

                    total_cost = float(cost_result['total_material_cost']) if cost_result['total_material_cost'] else 0

                    return {
                        'revenue': total_revenue,
                        'discount': total_discount,
                        'cost': total_cost,
                        'gross_profit': total_revenue - total_cost,
                        'gross_margin': ((total_revenue - total_cost) / total_revenue * 100) if total_revenue > 0 else 0
                    }
        except Exception as e:
            logger.error(f"Error getting revenue and cost data: {e}")
            return {
                'revenue': 0,
                'discount': 0,
                'cost': 0,
                'gross_profit': 0,
                'gross_margin': 0
            }

    def _get_dish_price_impact(self, store_id: int, current_month: Tuple, last_year_month: Tuple) -> Dict[str, float]:
        """Calculate impact of dish price changes on gross margin (year over year)"""
        current_year, current_mon = current_month
        last_year, last_mon = last_year_month

        query = """
            WITH current_dishes AS (
                SELECT
                    d.id,
                    dms.sale_amount,
                    COALESCE(
                        (SELECT dph.price
                         FROM dish_price_history dph
                         WHERE dph.dish_id = d.id
                           AND dph.store_id = %s
                           AND ((dph.effective_year < %s) OR
                                (dph.effective_year = %s AND dph.effective_month <= %s))
                         ORDER BY dph.effective_year DESC, dph.effective_month DESC
                         LIMIT 1),
                        0
                    ) as current_price
                FROM dish d
                INNER JOIN dish_monthly_sale dms ON d.id = dms.dish_id
                WHERE dms.store_id = %s
                    AND dms.year = %s
                    AND dms.month = %s
                    AND d.full_code != '14120001'
            ),
            last_year_dishes AS (
                SELECT
                    d.id,
                    COALESCE(
                        (SELECT dph.price
                         FROM dish_price_history dph
                         WHERE dph.dish_id = d.id
                           AND dph.store_id = %s
                           AND ((dph.effective_year < %s) OR
                                (dph.effective_year = %s AND dph.effective_month <= %s))
                         ORDER BY dph.effective_year DESC, dph.effective_month DESC
                         LIMIT 1),
                        0
                    ) as last_year_price
                FROM dish d
            )
            SELECT
                SUM(CASE
                    WHEN lyd.last_year_price > 0 THEN (cd.current_price - lyd.last_year_price) * cd.sale_amount
                    ELSE 0
                END) as price_impact
            FROM current_dishes cd
            LEFT JOIN last_year_dishes lyd ON cd.id = lyd.id
            WHERE cd.sale_amount > 0
        """

        try:
            with self.db_manager.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query,
                                 (store_id, current_year, current_year, current_mon,
                                  store_id, current_year, current_mon,
                                  store_id, last_year, last_year, last_mon))
                    result = cursor.fetchone()

                    return float(result['price_impact']) if result['price_impact'] else 0
        except Exception as e:
            logger.error(f"Error calculating dish price impact: {e}")
            return 0

    def _get_material_price_impact(self, store_id: int, current_month: Tuple, last_year_month: Tuple) -> Dict[str, float]:
        """Calculate impact of material price changes on gross margin (year over year)"""
        current_year, current_mon = current_month
        last_year, last_mon = last_year_month

        query = """
            WITH current_materials AS (
                SELECT
                    mmu.material_id,
                    mmu.material_used,
                    COALESCE(
                        (SELECT mph.price
                         FROM material_price_history mph
                         WHERE mph.material_id = mmu.material_id
                           AND mph.store_id = %s
                           AND ((mph.effective_year < %s) OR
                                (mph.effective_year = %s AND mph.effective_month <= %s))
                         ORDER BY mph.effective_year DESC, mph.effective_month DESC
                         LIMIT 1),
                        0
                    ) as current_price
                FROM material_monthly_usage mmu
                WHERE mmu.store_id = %s
                    AND mmu.year = %s
                    AND mmu.month = %s
            ),
            last_year_materials AS (
                SELECT
                    mmu.material_id,
                    COALESCE(
                        (SELECT mph.price
                         FROM material_price_history mph
                         WHERE mph.material_id = mmu.material_id
                           AND mph.store_id = %s
                           AND ((mph.effective_year < %s) OR
                                (mph.effective_year = %s AND mph.effective_month <= %s))
                         ORDER BY mph.effective_year DESC, mph.effective_month DESC
                         LIMIT 1),
                        0
                    ) as last_year_price
                FROM material_monthly_usage mmu
                WHERE mmu.store_id = %s
                    AND mmu.year = %s
                    AND mmu.month = %s
            )
            SELECT
                SUM(CASE
                    WHEN lym.material_id IS NOT NULL AND lym.last_year_price > 0
                    THEN (cm.current_price - lym.last_year_price) * cm.material_used
                    ELSE 0
                END) as price_impact
            FROM current_materials cm
            LEFT JOIN last_year_materials lym ON cm.material_id = lym.material_id
            WHERE cm.material_used > 0
        """

        try:
            with self.db_manager.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query,
                                 (store_id, current_year, current_year, current_mon,
                                  store_id, current_year, current_mon,
                                  store_id, last_year, last_year, last_mon,
                                  store_id, last_year, last_mon))
                    result = cursor.fetchone()

                    return float(result['price_impact']) if result['price_impact'] else 0
        except Exception as e:
            logger.error(f"Error calculating material price impact: {e}")
            return 0

    def _calculate_dish_loss_impact(self, store_id: int, month_tuple: Tuple) -> float:
        """Calculate dish loss/waste impact (actual vs theoretical material usage)"""
        year, month = month_tuple

        # This query calculates the difference between actual material usage and theoretical usage
        # based on dishes sold and their material composition
        query = """
            WITH theoretical_usage AS (
                -- Calculate theoretical material usage based on dishes sold
                SELECT
                    dm.material_id,
                    SUM(dms.sale_amount * dm.standard_quantity * COALESCE(dm.loss_rate, 0) /
                        COALESCE(NULLIF(dm.unit_conversion_rate, 0), 1)) as theoretical_qty
                FROM dish_monthly_sale dms
                INNER JOIN dish_material dm ON dms.dish_id = dm.dish_id
                WHERE dms.store_id = %s
                    AND dms.year = %s
                    AND dms.month = %s
                    AND dm.store_id = %s
                GROUP BY dm.material_id
            ),
            actual_usage AS (
                -- Get actual material usage (only cost-type materials)
                SELECT
                    material_id,
                    material_used as actual_qty
                FROM material_monthly_usage
                WHERE store_id = %s
                    AND year = %s
                    AND month = %s
                    AND material_use_type = '成本类'  -- Only cost-type materials
            ),
            material_prices AS (
                -- Get material prices
                SELECT DISTINCT
                    m.id as material_id,
                    COALESCE(
                        (SELECT mph.price
                         FROM material_price_history mph
                         WHERE mph.material_id = m.id
                           AND mph.store_id = %s
                           AND ((mph.effective_year < %s) OR
                                (mph.effective_year = %s AND mph.effective_month <= %s))
                         ORDER BY mph.effective_year DESC, mph.effective_month DESC
                         LIMIT 1),
                        0
                    ) as price
                FROM material m
                WHERE m.store_id = %s
            )
            SELECT
                SUM(
                    (COALESCE(au.actual_qty, 0) - COALESCE(tu.theoretical_qty, 0)) * mp.price
                ) as loss_amount
            FROM material_prices mp
            LEFT JOIN theoretical_usage tu ON mp.material_id = tu.material_id
            LEFT JOIN actual_usage au ON mp.material_id = au.material_id
            WHERE tu.material_id IS NOT NULL OR au.material_id IS NOT NULL  -- Only materials with either theoretical or actual usage
        """

        try:
            with self.db_manager.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query,
                                 (store_id, year, month, store_id,
                                  store_id, year, month,
                                  store_id, year, year, month,
                                  store_id))
                    result = cursor.fetchone()

                    return float(result['loss_amount']) if result['loss_amount'] else 0
        except Exception as e:
            logger.error(f"Error calculating dish loss impact: {e}")
            return 0

    def _get_store_analysis(self, store_id: int, store_name: str) -> Dict[str, Any]:
        """Get complete gross margin year-over-year analysis for a store"""
        date_ranges = self._get_date_ranges()

        # Get basic financial data
        current_data = self._get_revenue_and_cost(
            date_ranges['current'][0], date_ranges['current'][1], store_id
        )
        last_year_data = self._get_revenue_and_cost(
            date_ranges['last_year'][0], date_ranges['last_year'][1], store_id
        )

        # Calculate impacts
        current_month = (date_ranges['current'][0].year, date_ranges['current'][0].month)
        last_year_month = (date_ranges['last_year'][0].year, date_ranges['last_year'][0].month)

        dish_price_impact = self._get_dish_price_impact(store_id, current_month, last_year_month)
        material_price_impact = self._get_material_price_impact(store_id, current_month, last_year_month)
        dish_loss_impact = self._calculate_dish_loss_impact(store_id, current_month)
        dish_loss_impact_last_year = self._calculate_dish_loss_impact(store_id, last_year_month)

        # Calculate restored margins using the Excel formulas
        # These show what the margin would have been without each specific change
        revenue = current_data['revenue']
        current_margin_decimal = current_data['gross_margin'] / 100  # Convert to decimal
        last_year_margin_decimal = last_year_data['gross_margin'] / 100

        if revenue > 0:
            # Dish price restored margin: =IFERROR(1-(1-C5)*U5/(U5-F5),0)
            # Where C5=current margin, U5=current revenue, F5=dish price impact
            if (revenue - dish_price_impact) != 0:
                dish_restored_margin = (1 - (1 - current_margin_decimal) * revenue / (revenue - dish_price_impact)) * 100
            else:
                dish_restored_margin = 0
            dish_margin_impact = current_data['gross_margin'] - dish_restored_margin

            # Material price restored margin: =IFERROR(1-((1-C5)*U5-I5)/U5,0)
            # Where I5=material price impact (negative means cost increase)
            material_restored_margin = (1 - ((1 - current_margin_decimal) * revenue - material_price_impact) / revenue) * 100
            material_margin_impact = current_data['gross_margin'] - material_restored_margin

            # Discount restored margins
            # For discount, we need to calculate the discount percentage first
            discount_pct_current = current_data['discount'] / revenue if revenue > 0 else 0
            discount_pct_last_year = last_year_data['discount'] / last_year_data['revenue'] if last_year_data['revenue'] > 0 else 0

            # Current restored: =1-(1-C5)*U5/(U5/(1-L5)) where L5=current discount %
            if discount_pct_current < 1:
                current_discount_restored_margin = (1 - (1 - current_margin_decimal) * revenue / (revenue / (1 - discount_pct_current))) * 100
            else:
                current_discount_restored_margin = 0

            # Last year restored: =1-(1-D5)*V5/(V5/(1-M5)) where M5=last year discount %
            if discount_pct_last_year < 1 and last_year_data['revenue'] > 0:
                last_year_discount_restored_margin = (1 - (1 - last_year_margin_decimal) * last_year_data['revenue'] /
                                                      (last_year_data['revenue'] / (1 - discount_pct_last_year))) * 100
            else:
                last_year_discount_restored_margin = 0

            # Discount margin impact: the difference between the two restored margins
            # This shows the impact of discount change on margin
            discount_margin_impact = current_discount_restored_margin - last_year_discount_restored_margin

            # Dish loss restored margin: =IFERROR(1-((1-C5)*U5-R5)/U5,0)
            # Where R5=dish loss impact
            loss_restored_margin = (1 - ((1 - current_margin_decimal) * revenue - dish_loss_impact) / revenue) * 100
            loss_margin_impact = current_data['gross_margin'] - loss_restored_margin
        else:
            dish_restored_margin = 0
            dish_margin_impact = 0
            material_restored_margin = 0
            material_margin_impact = 0
            current_discount_restored_margin = 0
            last_year_discount_restored_margin = 0
            discount_margin_impact = 0
            loss_restored_margin = 0
            loss_margin_impact = 0

        return {
            'store_name': store_name,
            'current_margin': current_data['gross_margin'],
            'last_year_margin': last_year_data['gross_margin'],
            'margin_change': current_data['gross_margin'] - last_year_data['gross_margin'],
            'dish_price_impact_amount': dish_price_impact,
            'dish_restored_margin': dish_restored_margin,
            'dish_margin_impact': dish_margin_impact,
            'material_price_impact_amount': material_price_impact,
            'material_restored_margin': material_restored_margin,
            'material_margin_impact': material_margin_impact,
            'current_discount': current_data['discount'],
            'last_year_discount': last_year_data['discount'],
            'discount_change': ((current_data['discount'] - last_year_data['discount']) / last_year_data['discount'] * 100) if last_year_data['discount'] > 0 else 0,
            'current_discount_restored_margin': current_discount_restored_margin,
            'last_year_discount_restored_margin': last_year_discount_restored_margin,
            'discount_margin_impact': discount_margin_impact,
            'dish_loss_amount': dish_loss_impact,
            'loss_restored_margin': loss_restored_margin,
            'loss_margin_impact': loss_margin_impact,
            'current_revenue': current_data['revenue'],
            'last_year_revenue': last_year_data['revenue']
        }

    def generate_sheet(self, workbook):
        """Generate the gross margin year-over-year analysis sheet"""
        ws = workbook.create_sheet(self.sheet_name)

        # Define stores
        stores = [
            (1, '加拿大一店'),
            (2, '加拿大二店'),
            (3, '加拿大三店'),
            (4, '加拿大四店'),
            (5, '加拿大五店'),
            (6, '加拿大六店'),
            (7, '加拿大七店')
        ]

        # Style definitions
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        subheader_font = Font(color="FFFFFF", bold=True, size=10)
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        # Create headers with merged cells
        # Row 1: Main headers
        main_headers = [
            ('门店名称', 1, 1),
            ('毛利率同比', 2, 4),
            ('菜品价格影响', 5, 7),
            ('原材料价格影响', 8, 10),
            ('优惠影响', 11, 16),
            ('菜品损耗影响', 17, 19),
            ('本月收入', 20, 20),
            ('去年同期收入', 21, 21)
        ]

        for header, start_col, end_col in main_headers:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply border to merged cells
            for col in range(start_col, end_col + 1):
                ws.cell(row=1, column=col).border = border

        # Row 2: Sub-headers
        subheaders = [
            '',  # Store name
            '本月(%)', '去年同期(%)', '同比(pp)',  # Gross margin
            '变动金额', '还原毛利率(%)', '毛利率影响(pp)',  # Dish price
            '变动金额', '还原毛利率(%)', '毛利率影响(pp)',  # Material price
            '本月', '去年同期', '同比(%)', '本月还原毛利率(%)', '去年还原毛利率(%)', '毛利率影响(pp)',  # Discount
            '变动金额', '还原毛利率(%)', '毛利率影响(pp)',  # Dish loss
            '',  # Current revenue
            ''   # Last year revenue
        ]

        for col, subheader in enumerate(subheaders, 1):
            cell = ws.cell(row=2, column=col, value=subheader)
            if subheader:
                cell.fill = subheader_fill
                cell.font = subheader_font
            else:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # Merge cells that span both rows
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)  # Store name
        ws.merge_cells(start_row=1, start_column=20, end_row=2, end_column=20)  # Current revenue
        ws.merge_cells(start_row=1, start_column=21, end_row=2, end_column=21)  # Last year revenue

        # Process each store
        row_num = 3
        for store_id, store_name in stores:
            logger.info(f"Processing YoY gross margin analysis for {store_name}")

            try:
                data = self._get_store_analysis(store_id, store_name)

                # Store name
                ws.cell(row=row_num, column=1, value=data['store_name']).border = border

                # Gross margin comparison
                ws.cell(row=row_num, column=2, value=round(data['current_margin'] / 100, 4)).border = border
                ws.cell(row=row_num, column=2).number_format = '0.00%'

                ws.cell(row=row_num, column=3, value=round(data['last_year_margin'] / 100, 4)).border = border
                ws.cell(row=row_num, column=3).number_format = '0.00%'

                ws.cell(row=row_num, column=4, value=round(data['margin_change'] / 100, 4)).border = border
                ws.cell(row=row_num, column=4).number_format = '0.00%'

                # Dish price impact
                ws.cell(row=row_num, column=5, value=round(data['dish_price_impact_amount'], 2)).border = border
                ws.cell(row=row_num, column=5).number_format = '#,##0.00'

                ws.cell(row=row_num, column=6, value=round(data['dish_restored_margin'] / 100, 4)).border = border
                ws.cell(row=row_num, column=6).number_format = '0.00%'

                ws.cell(row=row_num, column=7, value=round(data['dish_margin_impact'] / 100, 4)).border = border
                ws.cell(row=row_num, column=7).number_format = '0.00%'

                # Material price impact
                ws.cell(row=row_num, column=8, value=round(data['material_price_impact_amount'], 2)).border = border
                ws.cell(row=row_num, column=8).number_format = '#,##0.00'

                ws.cell(row=row_num, column=9, value=round(data['material_restored_margin'] / 100, 4)).border = border
                ws.cell(row=row_num, column=9).number_format = '0.00%'

                ws.cell(row=row_num, column=10, value=round(data['material_margin_impact'] / 100, 4)).border = border
                ws.cell(row=row_num, column=10).number_format = '0.00%'

                # Discount impact
                ws.cell(row=row_num, column=11, value=round(data['current_discount'], 2)).border = border
                ws.cell(row=row_num, column=11).number_format = '#,##0.00'

                ws.cell(row=row_num, column=12, value=round(data['last_year_discount'], 2)).border = border
                ws.cell(row=row_num, column=12).number_format = '#,##0.00'

                ws.cell(row=row_num, column=13, value=round(data['discount_change'] / 100, 4)).border = border
                ws.cell(row=row_num, column=13).number_format = '0.00%'

                ws.cell(row=row_num, column=14, value=round(data['current_discount_restored_margin'] / 100, 4)).border = border
                ws.cell(row=row_num, column=14).number_format = '0.00%'

                ws.cell(row=row_num, column=15, value=round(data['last_year_discount_restored_margin'] / 100, 4)).border = border
                ws.cell(row=row_num, column=15).number_format = '0.00%'

                ws.cell(row=row_num, column=16, value=round(data['discount_margin_impact'] / 100, 4)).border = border
                ws.cell(row=row_num, column=16).number_format = '0.00%'

                # Dish loss impact
                ws.cell(row=row_num, column=17, value=round(data['dish_loss_amount'], 2)).border = border
                ws.cell(row=row_num, column=17).number_format = '#,##0.00'

                ws.cell(row=row_num, column=18, value=round(data['loss_restored_margin'] / 100, 4)).border = border
                ws.cell(row=row_num, column=18).number_format = '0.00%'

                ws.cell(row=row_num, column=19, value=round(data['loss_margin_impact'] / 100, 4)).border = border
                ws.cell(row=row_num, column=19).number_format = '0.00%'

                # Revenue
                ws.cell(row=row_num, column=20, value=round(data['current_revenue'], 2)).border = border
                ws.cell(row=row_num, column=20).number_format = '#,##0.00'

                ws.cell(row=row_num, column=21, value=round(data['last_year_revenue'], 2)).border = border
                ws.cell(row=row_num, column=21).number_format = '#,##0.00'

                row_num += 1

            except Exception as e:
                logger.error(f"Error processing {store_name}: {e}")
                continue

        # Set column widths
        column_widths = [15, 10, 10, 10, 12, 15, 15, 12, 15, 15, 10, 10, 10, 15, 15, 15, 12, 15, 15, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze panes (keep headers visible)
        ws.freeze_panes = 'A3'

        logger.info(f"Generated {self.sheet_name} with {row_num - 3} stores")
        return ws