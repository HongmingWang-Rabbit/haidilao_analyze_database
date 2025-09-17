#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate gross revenue report for all stores.
Creates an Excel workbook with a sheet for each store showing revenue and profit margins.
"""

import argparse
import sys
import logging
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
import os

# Add parent directory to path for imports
project_root = Path(__file__).resolve().parent.parent.parent.parent.parent
sys.path.insert(0, str(project_root))

from utils.database import DatabaseManager, DatabaseConfig
from lib.config import STORE_ID_TO_NAME_MAPPING
from scripts.dish_material.generate_report.generate_gross_revenue_report.store_revenue_sheet import (
    write_store_revenue_sheet
)
from scripts.dish_material.generate_report.generate_gross_revenue_report.all_stores_summary_sheet import (
    write_all_stores_summary_sheet
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def generate_gross_revenue_report(year: int, month: int, output_dir: str = None, test_db: bool = False, debug: bool = False):
    """
    Generate gross revenue report for all stores.
    
    Args:
        year: Target year
        month: Target month (1-12)
        output_dir: Output directory for the Excel file
        test_db: Use test database if True
        debug: If True, include detailed columns in the output
    """
    # Set up output directory
    if output_dir is None:
        output_dir = Path(__file__).resolve().parent.parent.parent.parent.parent / "Output" / "gross_revenue_reports"
    else:
        output_dir = Path(output_dir)
    
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Create database connection
    db_config = DatabaseConfig(is_test=test_db)
    db_manager = DatabaseManager(db_config)
    
    # Create workbook
    workbook = Workbook()
    
    # Remove default sheet
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])
    
    # First, create the summary sheet for all stores
    logger.info("Creating all stores summary sheet")
    summary_sheet = workbook.create_sheet(title="全店铺实际毛利汇总", index=0)
    write_all_stores_summary_sheet(
        worksheet=summary_sheet,
        db_manager=db_manager,
        year=year,
        month=month
    )
    logger.info("Successfully created summary sheet")
    
    # Generate sheets for each store
    stores_processed = 0
    for store_id, store_name in STORE_ID_TO_NAME_MAPPING.items():
        if store_id == 101:  # Skip Hi Bowl for now
            continue
            
        logger.info(f"Processing {store_name} (Store ID: {store_id})")
        
        try:
            # Create a sheet for this store with "理论毛利" suffix
            # Shorten sheet names
            sheet_name = store_name
            if "\u52a0\u62ff\u5927" in sheet_name:
                sheet_name = sheet_name.replace("\u52a0\u62ff\u5927", "CA")
            sheet_name = f"{sheet_name}\u7406\u8bba\u6bdb\u5229"
            worksheet = workbook.create_sheet(title=sheet_name)
            
            # Write revenue data to the sheet
            write_store_revenue_sheet(
                worksheet=worksheet,
                db_manager=db_manager,
                year=year,
                month=month,
                store_id=store_id,
                store_name=store_name,
                debug=debug
            )
            
            stores_processed += 1
            logger.info(f"Successfully processed {store_name}")
            
        except Exception as e:
            logger.error(f"Error processing {store_name}: {str(e)}", exc_info=True)
            # Continue with next store even if this one fails
            continue
    
    if stores_processed == 0:
        logger.error("No stores were successfully processed")
        return None
    
    # Save the workbook
    output_filename = f"gross_revenue_report_{year}_{month:02d}.xlsx"
    output_path = output_dir / output_filename
    
    try:
        workbook.save(output_path)
        logger.info(f"Report saved to: {output_path}")
        return output_path
    except Exception as e:
        logger.error(f"Failed to save workbook: {str(e)}")
        return None


def main():
    """Main entry point for the script."""
    parser = argparse.ArgumentParser(
        description='Generate gross revenue report for all stores'
    )
    parser.add_argument(
        '--year',
        type=int,
        required=True,
        help='Target year (e.g., 2025)'
    )
    parser.add_argument(
        '--month',
        type=int,
        required=True,
        help='Target month (1-12)'
    )
    parser.add_argument(
        '--output-dir',
        type=str,
        help='Output directory for the report (optional)'
    )
    parser.add_argument(
        '--test',
        action='store_true',
        help='Use test database'
    )
    parser.add_argument(
        '--debug',
        action='store_true',
        help='Include detailed columns (dish codes, material details)'
    )
    
    args = parser.parse_args()
    
    # Validate month
    if not 1 <= args.month <= 12:
        logger.error("Month must be between 1 and 12")
        sys.exit(1)
    
    # Print header
    print("\n" + "="*60)
    print("GROSS REVENUE REPORT GENERATION")
    print("="*60)
    print(f"Year: {args.year}")
    print(f"Month: {args.month}")
    print(f"Database: {'Test' if args.test else 'Production'}")
    print(f"Debug Mode: {'Enabled' if args.debug else 'Disabled'}")
    print("="*60 + "\n")
    
    # Generate report
    output_path = generate_gross_revenue_report(
        year=args.year,
        month=args.month,
        output_dir=args.output_dir,
        test_db=args.test,
        debug=args.debug
    )
    
    # Print result
    print("\n" + "="*60)
    if output_path:
        print("[SUCCESS] Report generated successfully")
        print(f"Output file: {output_path}")
        
        # Open the file if on Windows
        if os.name == 'nt' and output_path.exists():
            try:
                os.startfile(output_path)
                print("Opening report in Excel...")
            except Exception as e:
                logger.warning(f"Could not open file automatically: {e}")
    else:
        print("[FAILED] Report generation failed")
        print("Check the logs above for error details")
    print("="*60 + "\n")
    
    sys.exit(0 if output_path else 1)


if __name__ == '__main__':
    main()