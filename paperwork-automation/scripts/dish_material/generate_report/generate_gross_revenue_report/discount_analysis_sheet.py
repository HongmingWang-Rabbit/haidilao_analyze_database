import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal
from typing import Optional, Dict, Any, Tuple
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

logger = logging.getLogger(__name__)

class DiscountAnalysisSheet:
    def __init__(self, db_manager, target_date: str):
        self.db_manager = db_manager
        self.target_date = datetime.strptime(target_date, '%Y-%m-%d')
        self.sheet_name = "打折优惠表"

    def _get_date_ranges(self):
        """Calculate date ranges for current, last month, and last year same month"""
        current_start = self.target_date.replace(day=1)
        current_end = self.target_date

        last_month_end = current_start - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)

        last_year_start = current_start - relativedelta(years=1)
        last_year_end = self.target_date - relativedelta(years=1)

        return {
            'current': (current_start, current_end),
            'last_month': (last_month_start, last_month_end),
            'last_year': (last_year_start, last_year_end)
        }

    def _get_discount_data_for_period(self, start_date: datetime, end_date: datetime, store_id: int) -> Dict[str, float]:
        """Get discount data for a specific period and store from daily_report table"""
        query = """
            SELECT
                SUM(CAST(revenue_tax_not_included AS DECIMAL(15,2))) as total_revenue,
                SUM(CAST(discount_total AS DECIMAL(15,2))) as total_discount
            FROM daily_report
            WHERE store_id = %s
                AND date >= %s
                AND date <= %s
        """

        try:
            with self.db_manager.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query, (store_id, start_date, end_date))
                    result = cursor.fetchone()

                    if result and result['total_revenue']:
                        total_revenue = float(result['total_revenue']) if result['total_revenue'] else 0
                        total_discount = float(result['total_discount']) if result['total_discount'] else 0
                        discount_percentage = (total_discount / total_revenue * 100) if total_revenue > 0 else 0

                        return {
                            'revenue': total_revenue,
                            'discount': total_discount,
                            'discount_percentage': discount_percentage
                        }
                    else:
                        return {
                            'revenue': 0,
                            'discount': 0,
                            'discount_percentage': 0
                        }
        except Exception as e:
            logger.error(f"Error getting discount data for store {store_id}: {e}")
            return {
                'revenue': 0,
                'discount': 0,
                'discount_percentage': 0
            }

    def _calculate_percentage_change(self, current: float, previous: float) -> float:
        """Calculate percentage change between two values"""
        if previous == 0:
            return 0 if current == 0 else 100
        return ((current - previous) / previous) * 100

    def _get_store_discount_analysis(self, store_id: int, store_name: str) -> Dict[str, Any]:
        """Get complete discount analysis for a store"""
        date_ranges = self._get_date_ranges()

        # Get data for all three periods
        current_data = self._get_discount_data_for_period(
            date_ranges['current'][0], date_ranges['current'][1], store_id
        )
        last_month_data = self._get_discount_data_for_period(
            date_ranges['last_month'][0], date_ranges['last_month'][1], store_id
        )
        last_year_data = self._get_discount_data_for_period(
            date_ranges['last_year'][0], date_ranges['last_year'][1], store_id
        )

        # Calculate percentage point differences (not percentage changes)
        # This is the difference in discount percentages, not the change in discount amounts
        mom_change = current_data['discount_percentage'] - last_month_data['discount_percentage']
        yoy_change = current_data['discount_percentage'] - last_year_data['discount_percentage']

        return {
            'store_name': store_name,
            'current': current_data,
            'last_month': last_month_data,
            'last_year': last_year_data,
            'mom_change': mom_change,
            'yoy_change': yoy_change
        }

    def generate_sheet(self, workbook):
        """Generate the discount analysis sheet"""
        ws = workbook.create_sheet(self.sheet_name)

        # Define stores
        stores = [
            (1, '加拿大一店'),
            (2, '加拿大二店'),
            (3, '加拿大三店'),
            (4, '加拿大四店'),
            (5, '加拿大五店'),
            (6, '加拿大六店'),
            (7, '加拿大七店')
        ]

        # Style definitions
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        subheader_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        subheader_font = Font(color="FFFFFF", bold=True, size=10)
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        # Create headers with merged cells
        # Row 1: Main headers
        main_headers = [
            ('门店名称', 1, 1),
            ('本月', 2, 4),
            ('上月', 5, 7),
            ('环比', 8, 8),
            ('去年同期', 9, 11),
            ('同比', 12, 12)
        ]

        for header, start_col, end_col in main_headers:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

            # Apply border to merged cells
            for col in range(start_col, end_col + 1):
                ws.cell(row=1, column=col).border = border

        # Row 2: Sub-headers
        subheaders = [
            '',  # Store name column
            '优惠占比(%)', '收入(本币)', '优惠总金额(本币)',  # Current month
            '优惠占比(%)', '收入(本币)', '优惠总金额(本币)',  # Last month
            '优惠变动(%)',  # MoM change
            '优惠占比(%)', '收入(本币)', '优惠总金额(本币)',  # Last year
            '优惠变动(%)'  # YoY change
        ]

        for col, subheader in enumerate(subheaders, 1):
            cell = ws.cell(row=2, column=col, value=subheader)
            if subheader:
                cell.fill = subheader_fill
                cell.font = subheader_font
            else:
                cell.fill = header_fill
                cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Merge store name cell across rows 1 and 2
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

        # Get and write data for each store
        row_num = 3
        for store_id, store_name in stores:
            logger.info(f"Processing discount data for {store_name}")

            try:
                data = self._get_store_discount_analysis(store_id, store_name)

                # Write store name
                ws.cell(row=row_num, column=1, value=data['store_name']).border = border

                # Current month data
                ws.cell(row=row_num, column=2, value=round(data['current']['discount_percentage'] / 100, 4)).border = border
                ws.cell(row=row_num, column=2).number_format = '0.00%'

                ws.cell(row=row_num, column=3, value=round(data['current']['revenue'], 2)).border = border
                ws.cell(row=row_num, column=3).number_format = '#,##0.00'

                ws.cell(row=row_num, column=4, value=round(data['current']['discount'], 2)).border = border
                ws.cell(row=row_num, column=4).number_format = '#,##0.00'

                # Last month data
                ws.cell(row=row_num, column=5, value=round(data['last_month']['discount_percentage'] / 100, 4)).border = border
                ws.cell(row=row_num, column=5).number_format = '0.00%'

                ws.cell(row=row_num, column=6, value=round(data['last_month']['revenue'], 2)).border = border
                ws.cell(row=row_num, column=6).number_format = '#,##0.00'

                ws.cell(row=row_num, column=7, value=round(data['last_month']['discount'], 2)).border = border
                ws.cell(row=row_num, column=7).number_format = '#,##0.00'

                # MoM change (percentage point difference)
                mom_cell = ws.cell(row=row_num, column=8, value=round(data['mom_change'] / 100, 4))
                mom_cell.border = border
                mom_cell.number_format = '0.00%'
                if data['mom_change'] > 0:
                    mom_cell.font = Font(color="FF0000")
                elif data['mom_change'] < 0:
                    mom_cell.font = Font(color="008000")

                # Last year data
                ws.cell(row=row_num, column=9, value=round(data['last_year']['discount_percentage'] / 100, 4)).border = border
                ws.cell(row=row_num, column=9).number_format = '0.00%'

                ws.cell(row=row_num, column=10, value=round(data['last_year']['revenue'], 2)).border = border
                ws.cell(row=row_num, column=10).number_format = '#,##0.00'

                ws.cell(row=row_num, column=11, value=round(data['last_year']['discount'], 2)).border = border
                ws.cell(row=row_num, column=11).number_format = '#,##0.00'

                # YoY change (percentage point difference)
                yoy_cell = ws.cell(row=row_num, column=12, value=round(data['yoy_change'] / 100, 4))
                yoy_cell.border = border
                yoy_cell.number_format = '0.00%'
                if data['yoy_change'] > 0:
                    yoy_cell.font = Font(color="FF0000")
                elif data['yoy_change'] < 0:
                    yoy_cell.font = Font(color="008000")

                row_num += 1

            except Exception as e:
                logger.error(f"Error processing {store_name}: {e}")
                continue

        # Add totals row
        ws.cell(row=row_num, column=1, value="总计").border = border
        ws.cell(row=row_num, column=1).font = Font(bold=True)

        # Calculate totals for each column
        for col in range(2, 13):
            if col in [2, 5, 9]:  # Percentage columns - calculate weighted average
                revenue_col = col + 1
                discount_col = col + 2
                total_revenue_formula = f"=SUM({get_column_letter(revenue_col)}3:{get_column_letter(revenue_col)}{row_num-1})"
                total_discount_formula = f"=SUM({get_column_letter(discount_col)}3:{get_column_letter(discount_col)}{row_num-1})"

                # Use formula for weighted average percentage
                cell = ws.cell(row=row_num, column=col)
                cell.value = f"={total_discount_formula}/{total_revenue_formula}"
                cell.number_format = '0.00%'
            elif col in [3, 4, 6, 7, 10, 11]:  # Revenue and discount amount columns
                formula = f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{row_num-1})"
                cell = ws.cell(row=row_num, column=col, value=formula)
                cell.number_format = '#,##0.00'
            elif col in [8, 12]:  # Change columns (percentage point differences)
                # Calculate percentage point difference for totals
                if col == 8:  # MoM (column B - column E)
                    cell = ws.cell(row=row_num, column=col)
                    cell.value = f"=B{row_num}-E{row_num}"
                else:  # YoY (column B - column I)
                    cell = ws.cell(row=row_num, column=col)
                    cell.value = f"=B{row_num}-I{row_num}"
                cell.number_format = '0.00%'

            ws.cell(row=row_num, column=col).border = border
            ws.cell(row=row_num, column=col).font = Font(bold=True)

        # Set column widths
        column_widths = [15, 12, 15, 18, 12, 15, 18, 12, 12, 15, 18, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Freeze panes (keep headers visible)
        ws.freeze_panes = 'A3'

        logger.info(f"Generated {self.sheet_name} with {row_num - 3} stores")
        return ws