import logging
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from decimal import Decimal
from typing import Optional, Dict, Any
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

logger = logging.getLogger(__name__)

class MaterialCostChangeSheet:
    def __init__(self, db_manager, target_date: str):
        self.db_manager = db_manager
        self.target_date = datetime.strptime(target_date, '%Y-%m-%d')
        self.sheet_name = "原材料成本变动表"

    def _get_date_ranges(self):
        """Calculate date ranges for current, last month, and last year same month"""
        current_start = self.target_date.replace(day=1)
        current_end = self.target_date

        last_month_end = current_start - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)

        last_year_start = current_start - relativedelta(years=1)
        last_year_end = self.target_date - relativedelta(years=1)

        return {
            'current': (current_start, current_end),
            'last_month': (last_month_start, last_month_end),
            'last_year': (last_year_start, last_year_end)
        }

    def _get_material_prices_and_usage(self, date_range: tuple, store_id: int) -> Dict[str, Dict[str, Any]]:
        """Get material prices and usage for a specific date range and store"""
        start_date, end_date = date_range
        year = start_date.year
        month = start_date.month

        query = """
            WITH material_data AS (
                SELECT
                    m.material_number,
                    m.name as material_name,
                    mmu.material_used as total_usage,
                    COALESCE(
                        (SELECT mph.price
                         FROM material_price_history mph
                         WHERE mph.material_id = m.id
                           AND mph.store_id = %s
                           AND ((mph.effective_year < %s) OR
                                (mph.effective_year = %s AND mph.effective_month <= %s))
                         ORDER BY mph.effective_year DESC, mph.effective_month DESC
                         LIMIT 1),
                        0
                    ) as avg_price
                FROM material m
                INNER JOIN material_monthly_usage mmu ON m.id = mmu.material_id
                WHERE mmu.store_id = %s
                    AND mmu.year = %s
                    AND mmu.month = %s
                    AND m.store_id = %s
            )
            SELECT
                material_number,
                material_name,
                avg_price,
                total_usage
            FROM material_data
            WHERE total_usage > 0
        """

        try:
            with self.db_manager.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query, (store_id, year, year, month, store_id, year, month, store_id))
                    results = cursor.fetchall()

                    material_data = {}
                    for row in results:
                        material_data[row['material_number']] = {
                            'material_name': row['material_name'],
                            'avg_price': float(row['avg_price']) if row['avg_price'] else 0,
                            'total_usage': float(row['total_usage']) if row['total_usage'] else 0
                        }

                    return material_data
        except Exception as e:
            logger.error(f"Error getting material prices and usage: {e}")
            return {}

    def _get_store_data(self, store_id: int, store_name: str) -> list:
        """Get material cost change data for a specific store"""
        date_ranges = self._get_date_ranges()

        current_data = self._get_material_prices_and_usage(date_ranges['current'], store_id)
        last_month_data = self._get_material_prices_and_usage(date_ranges['last_month'], store_id)
        last_year_data = self._get_material_prices_and_usage(date_ranges['last_year'], store_id)

        all_materials = set(current_data.keys()) | set(last_month_data.keys()) | set(last_year_data.keys())

        store_rows = []
        for material_code in all_materials:
            current = current_data.get(material_code, {})
            last_month = last_month_data.get(material_code, {})
            last_year = last_year_data.get(material_code, {})

            if not current:
                continue

            current_price = current.get('avg_price', 0)
            last_month_price = last_month.get('avg_price', 0)
            last_year_price = last_year.get('avg_price', 0)
            current_usage = current.get('total_usage', 0)

            mom_price_change = current_price - last_month_price if last_month_price else 0
            yoy_price_change = current_price - last_year_price if last_year_price else 0

            mom_cost_impact = mom_price_change * current_usage
            yoy_cost_impact = yoy_price_change * current_usage

            row = {
                '门店': store_name,
                '物料编码': material_code,
                '物料名称': current.get('material_name', last_month.get('material_name', last_year.get('material_name', ''))),
                '本期单价': current_price,
                '上期单价': last_month_price,
                '去年同期单价': last_year_price,
                '环比价格变动': mom_price_change,
                '同比价格变动': yoy_price_change,
                '本期用量': current_usage,
                '环比影响成本': mom_cost_impact,
                '同比影响成本': yoy_cost_impact
            }

            store_rows.append(row)

        return sorted(store_rows, key=lambda x: abs(x['环比影响成本']), reverse=True)

    def generate_sheet(self, workbook):
        """Generate the material cost change sheet"""
        ws = workbook.create_sheet(self.sheet_name)

        stores = [
            (1, '加拿大一店'),
            (2, '加拿大二店'),
            (3, '加拿大三店'),
            (4, '加拿大四店'),
            (5, '加拿大五店'),
            (6, '加拿大六店'),
            (7, '加拿大七店')
        ]

        headers = [
            '门店', '物料编码', '物料名称', '本期单价', '上期单价',
            '去年同期单价', '环比价格变动', '同比价格变动', '本期用量',
            '环比影响成本', '同比影响成本'
        ]

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        all_data = []
        for store_id, store_name in stores:
            logger.info(f"Processing store: {store_name}")
            store_data = self._get_store_data(store_id, store_name)
            all_data.extend(store_data)

        row_num = 2
        for data in all_data:
            ws.cell(row=row_num, column=1, value=data['门店']).border = border
            ws.cell(row=row_num, column=2, value=data['物料编码']).border = border
            ws.cell(row=row_num, column=3, value=data['物料名称']).border = border

            ws.cell(row=row_num, column=4, value=round(data['本期单价'], 2)).border = border
            ws.cell(row=row_num, column=4).number_format = '#,##0.00'

            ws.cell(row=row_num, column=5, value=round(data['上期单价'], 2)).border = border
            ws.cell(row=row_num, column=5).number_format = '#,##0.00'

            ws.cell(row=row_num, column=6, value=round(data['去年同期单价'], 2)).border = border
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'

            mom_change_cell = ws.cell(row=row_num, column=7, value=round(data['环比价格变动'], 2))
            mom_change_cell.border = border
            mom_change_cell.number_format = '#,##0.00'
            if data['环比价格变动'] > 0:
                mom_change_cell.font = Font(color="FF0000")
            elif data['环比价格变动'] < 0:
                mom_change_cell.font = Font(color="008000")

            yoy_change_cell = ws.cell(row=row_num, column=8, value=round(data['同比价格变动'], 2))
            yoy_change_cell.border = border
            yoy_change_cell.number_format = '#,##0.00'
            if data['同比价格变动'] > 0:
                yoy_change_cell.font = Font(color="FF0000")
            elif data['同比价格变动'] < 0:
                yoy_change_cell.font = Font(color="008000")

            ws.cell(row=row_num, column=9, value=round(data['本期用量'], 2)).border = border
            ws.cell(row=row_num, column=9).number_format = '#,##0.00'

            mom_impact_cell = ws.cell(row=row_num, column=10, value=round(data['环比影响成本'], 2))
            mom_impact_cell.border = border
            mom_impact_cell.number_format = '#,##0.00'
            if data['环比影响成本'] > 0:
                mom_impact_cell.font = Font(color="FF0000")
            elif data['环比影响成本'] < 0:
                mom_impact_cell.font = Font(color="008000")

            yoy_impact_cell = ws.cell(row=row_num, column=11, value=round(data['同比影响成本'], 2))
            yoy_impact_cell.border = border
            yoy_impact_cell.number_format = '#,##0.00'
            if data['同比影响成本'] > 0:
                yoy_impact_cell.font = Font(color="FF0000")
            elif data['同比影响成本'] < 0:
                yoy_impact_cell.font = Font(color="008000")

            row_num += 1

        column_widths = [15, 15, 30, 12, 12, 15, 15, 15, 12, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = 'A2'

        logger.info(f"Generated {self.sheet_name} with {row_num - 2} rows of data")
        return ws