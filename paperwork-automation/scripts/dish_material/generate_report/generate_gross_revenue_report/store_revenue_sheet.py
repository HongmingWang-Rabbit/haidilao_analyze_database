#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate sheet to show gross revenue and profit margins for each store.
"""

import sys
from pathlib import Path
from typing import Dict, List, Optional
import logging
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Add parent directory to path for imports
project_root = Path(__file__).resolve().parent.parent.parent.parent.parent
sys.path.insert(0, str(project_root))

from utils.database import DatabaseManager

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def get_actual_material_usage(db_manager: DatabaseManager, year: int, month: int, store_id: int) -> Dict[str, Dict]:
    """
    Get actual material usage and cost from material_monthly_usage table.
    Only includes materials with material_use_type = '成本类' (cost-type materials).
    
    Args:
        db_manager: Database manager instance
        year: Year for the data
        month: Month for the data
        store_id: Store ID
    
    Returns:
        Dict mapping material_number to dict with usage and total_cost
    """
    actual_usage = {}
    
    with db_manager.get_connection() as conn:
        cursor = conn.cursor()
        
        query = """
        SELECT 
            m.material_number,
            mmu.material_used,
            COALESCE(mph.price, 0) as material_price,
            mmu.material_used * COALESCE(mph.price, 0) as actual_cost
        FROM material_monthly_usage mmu
        JOIN material m ON m.id = mmu.material_id AND m.store_id = mmu.store_id
        LEFT JOIN material_price_history mph ON mph.material_id = m.id 
            AND mph.store_id = m.store_id
            AND mph.is_active = TRUE
        WHERE mmu.year = %s
          AND mmu.month = %s
          AND mmu.store_id = %s
          AND mmu.material_use_type = '成本类'  -- Only include cost-type materials
        """
        
        cursor.execute(query, (year, month, store_id))
        
        for row in cursor.fetchall():
            if row['material_number'] and row['material_used']:
                actual_usage[row['material_number']] = {
                    'usage': float(row['material_used']),
                    'price': float(row['material_price']) if row['material_price'] else 0,
                    'total_cost': float(row['actual_cost']) if row['actual_cost'] else 0
                }
    
    return actual_usage


def get_dish_revenue_data(db_manager: DatabaseManager, year: int, month: int, store_id: int, debug: bool = False) -> List[Dict]:
    """
    Get dish revenue and material cost data for gross profit calculation.
    
    Args:
        db_manager: Database manager instance
        year: Year for the data
        month: Month for the data
        store_id: Store ID
        debug: If True, include detailed info
    
    Returns:
        List of dicts containing dish revenue and material cost data
    """
    results = []
    
    with db_manager.get_connection() as conn:
        cursor = conn.cursor()
        
        # Get dish sales with prices and material costs
        query = """
        WITH dish_sales AS (
            -- Regular dish sales
            SELECT 
                dms.dish_id,
                d.name as dish_name,
                d.full_code as dish_full_code,
                d.short_code as dish_short_code,
                d.size as dish_size,
                COALESCE(dms.sale_amount, 0) - COALESCE(dms.return_amount, 0) as quantity_sold,
                COALESCE(dph.price, 0) as dish_price,
                FALSE as is_combo
            FROM dish_monthly_sale dms
            JOIN dish d ON d.id = dms.dish_id
            LEFT JOIN dish_price_history dph ON dph.dish_id = d.id 
                AND dph.store_id = dms.store_id
                AND dph.is_active = TRUE
            WHERE dms.year = %s
              AND dms.month = %s
              AND dms.store_id = %s
              AND (COALESCE(dms.sale_amount, 0) - COALESCE(dms.return_amount, 0)) > 0
            
            UNION ALL
            
            -- Combo dish sales
            SELECT 
                mcds.dish_id,
                CONCAT(d.name, '(套餐-', c.name, ')') as dish_name,
                d.full_code as dish_full_code,
                d.short_code as dish_short_code,
                d.size as dish_size,
                mcds.sale_amount as quantity_sold,
                0 as dish_price,  -- Combo dishes don't have individual prices
                TRUE as is_combo
            FROM monthly_combo_dish_sale mcds
            JOIN dish d ON d.id = mcds.dish_id
            JOIN combo c ON c.id = mcds.combo_id
            WHERE mcds.year = %s
              AND mcds.month = %s
              AND mcds.store_id = %s
              AND mcds.sale_amount > 0
        ),
        dish_materials AS (
            -- Get material costs for each dish
            SELECT 
                ds.dish_id,
                ds.dish_name,
                ds.dish_full_code,
                ds.dish_short_code,
                ds.dish_size,
                ds.quantity_sold,
                ds.dish_price,
                ds.is_combo,
                m.id as material_id,
                m.material_number,
                m.description as material_name,
                COALESCE(mph.price, 0) as material_price,
                dm.standard_quantity,
                dm.loss_rate,
                dm.unit_conversion_rate,
                -- Calculate material usage per dish sale
                (dm.standard_quantity * COALESCE(dm.loss_rate, 0) / COALESCE(NULLIF(dm.unit_conversion_rate, 0), 1)) as material_usage_per_dish,
                -- Calculate total material cost for this dish
                (ds.quantity_sold * dm.standard_quantity * COALESCE(dm.loss_rate, 0) / COALESCE(NULLIF(dm.unit_conversion_rate, 0), 1) * COALESCE(mph.price, 0)) as material_cost
            FROM dish_sales ds
            LEFT JOIN dish_material dm ON dm.dish_id = ds.dish_id AND dm.store_id = %s
            LEFT JOIN material m ON m.id = dm.material_id AND m.store_id = %s
            LEFT JOIN material_price_history mph ON mph.material_id = m.id 
                AND mph.store_id = %s
                AND mph.is_active = TRUE
        )
        SELECT 
            dish_id,
            dish_name,
            dish_full_code,
            dish_short_code,
            dish_size,
            quantity_sold,
            dish_price,
            is_combo,
            material_id,
            material_number,
            material_name,
            material_price,
            material_usage_per_dish,
            material_cost
        FROM dish_materials
        ORDER BY dish_full_code, dish_short_code, material_number
        """
        
        cursor.execute(query, (
            year, month, store_id,  # Regular sales
            year, month, store_id,  # Combo sales  
            store_id, store_id, store_id  # Material joins
        ))
        
        raw_data = cursor.fetchall()
        
        # Group by dish
        dish_data = {}
        for row in raw_data:
            dish_key = (row['dish_id'], row['is_combo'])
            
            if dish_key not in dish_data:
                dish_data[dish_key] = {
                    'dish_id': row['dish_id'],
                    'dish_name': row['dish_name'],
                    'dish_full_code': row['dish_full_code'],
                    'dish_short_code': row['dish_short_code'],
                    'dish_size': row['dish_size'],
                    'quantity_sold': float(row['quantity_sold']) if row['quantity_sold'] else 0,
                    'dish_price': float(row['dish_price']) if row['dish_price'] else 0,
                    'is_combo': row['is_combo'],
                    'sales_amount': float(row['quantity_sold']) * float(row['dish_price']) if row['dish_price'] and row['quantity_sold'] else 0,
                    'materials': [],
                    'total_material_cost': 0
                }
            
            # Add material info if exists
            if row['material_id']:
                material_info = {
                    'material_id': row['material_id'],
                    'material_number': row['material_number'],
                    'material_name': row['material_name'],
                    'material_price': float(row['material_price']) if row['material_price'] else 0,
                    'usage_per_dish': float(row['material_usage_per_dish']) if row['material_usage_per_dish'] else 0,
                    'total_cost': float(row['material_cost']) if row['material_cost'] else 0
                }
                dish_data[dish_key]['materials'].append(material_info)
                dish_data[dish_key]['total_material_cost'] += material_info['total_cost']
        
        # Convert to list and calculate profit margins
        for dish_info in dish_data.values():
            # Calculate gross profit
            dish_info['gross_profit'] = dish_info['sales_amount'] - dish_info['total_material_cost']
            
            # Calculate profit margin percentage
            if dish_info['sales_amount'] > 0:
                dish_info['profit_margin'] = (dish_info['gross_profit'] / dish_info['sales_amount']) * 100
            else:
                dish_info['profit_margin'] = 0
            
            results.append(dish_info)
        
        # Sort by dish full code and short code
        results.sort(key=lambda x: (x['dish_full_code'] or '', x['dish_short_code'] or ''))
    
    return results


def write_store_revenue_sheet(worksheet, db_manager: DatabaseManager, year: int, month: int, 
                             store_id: int, store_name: str, debug: bool = False):
    """
    Write store revenue and gross profit data to an Excel worksheet.
    
    Args:
        worksheet: Openpyxl worksheet object
        db_manager: Database manager instance
        year: Year for the data
        month: Month for the data
        store_id: Store ID
        store_name: Store name for display
        debug: If True, include detailed columns
    """
    # Add title
    title_cell = worksheet.cell(row=1, column=1, value=f"{store_name} - 毛利润分析")
    title_cell.font = Font(size=14, bold=True)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12 if debug else 10)
    title_cell.alignment = Alignment(horizontal='center')
    
    date_cell = worksheet.cell(row=2, column=1, value=f"{year}年{month}月")
    date_cell.font = Font(size=12)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12 if debug else 10)
    date_cell.alignment = Alignment(horizontal='center')
    
    # Leave a blank row
    header_row = 4
    
    # Set headers based on debug flag
    if debug:
        headers = ['菜品名称', '菜品编号', '菜品短编号', '规格', '菜品价格', '销售数量', '销售金额', '使用物料', '理论总成本', '实际总成本', '毛利', '毛利百分比']
    else:
        headers = ['菜品名称', '规格', '菜品价格', '销售数量', '销售金额', '使用物料', '理论总成本', '实际总成本', '毛利', '毛利百分比']
    
    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alternating row fills
    even_row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Get data
    data = get_dish_revenue_data(db_manager, year, month, store_id, debug)
    
    # Get actual material usage for the store
    actual_material_usage = get_actual_material_usage(db_manager, year, month, store_id)
    
    # Calculate total actual material cost for the store
    total_actual_material_cost = sum(mat_info['total_cost'] for mat_info in actual_material_usage.values())
    
    # Calculate total theoretical usage for each material across all dishes
    # This helps us distribute actual costs proportionally
    material_theoretical_totals = {}
    for dish_info in data:
        if dish_info['is_combo'] and dish_info['dish_price'] == 0:
            continue
        for material in dish_info['materials']:
            material_number = material['material_number']
            if material_number not in material_theoretical_totals:
                material_theoretical_totals[material_number] = 0
            # Add this dish's theoretical usage of the material
            material_theoretical_totals[material_number] += material['usage_per_dish'] * dish_info['quantity_sold']
    
    current_row = header_row + 1
    dish_index = 0
    
    for dish_info in data:
        # Skip combo dishes with no price (they're part of combo packages)
        if dish_info['is_combo'] and dish_info['dish_price'] == 0:
            continue
        
        # Calculate actual cost for each material in this dish
        dish_actual_cost = 0
        for material in dish_info['materials']:
            material_number = material['material_number']
            if material_number in actual_material_usage and material_number in material_theoretical_totals:
                # Get actual usage and cost for this material at store level
                actual_info = actual_material_usage[material_number]
                
                # Calculate this dish's proportion of the material usage
                dish_theoretical_usage = material['usage_per_dish'] * dish_info['quantity_sold']
                total_theoretical_usage = material_theoretical_totals[material_number]
                
                if total_theoretical_usage > 0:
                    # Proportion of actual cost = (dish's theoretical usage / total theoretical usage) * actual cost
                    proportion = dish_theoretical_usage / total_theoretical_usage
                    material['actual_cost'] = actual_info['total_cost'] * proportion
                    material['actual_usage'] = actual_info['usage'] * proportion
                else:
                    material['actual_cost'] = 0
                    material['actual_usage'] = 0
                
                dish_actual_cost += material['actual_cost']
            else:
                material['actual_usage'] = 0
                material['actual_cost'] = 0
        
        # Set the dish's actual total cost
        dish_info['actual_total_cost'] = dish_actual_cost
            
        num_materials = max(len(dish_info['materials']), 1)
        
        current_col = 1
        
        # Dish name
        cell = worksheet.cell(row=current_row, column=current_col, value=dish_info['dish_name'])
        if num_materials > 1:
            worksheet.merge_cells(start_row=current_row, start_column=current_col,
                                end_row=current_row + num_materials - 1, end_column=current_col)
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        current_col += 1
        
        # Debug columns: dish codes
        if debug:
            # Full code
            cell = worksheet.cell(row=current_row, column=current_col, value=dish_info['dish_full_code'])
            if num_materials > 1:
                worksheet.merge_cells(start_row=current_row, start_column=current_col,
                                    end_row=current_row + num_materials - 1, end_column=current_col)
            cell.alignment = Alignment(vertical='center')
            current_col += 1
            
            # Short code
            cell = worksheet.cell(row=current_row, column=current_col, value=dish_info['dish_short_code'])
            if num_materials > 1:
                worksheet.merge_cells(start_row=current_row, start_column=current_col,
                                    end_row=current_row + num_materials - 1, end_column=current_col)
            cell.alignment = Alignment(vertical='center')
            current_col += 1
        
        # Size
        cell = worksheet.cell(row=current_row, column=current_col, value=dish_info['dish_size'])
        if num_materials > 1:
            worksheet.merge_cells(start_row=current_row, start_column=current_col,
                                end_row=current_row + num_materials - 1, end_column=current_col)
        cell.alignment = Alignment(vertical='center')
        current_col += 1
        
        # Price
        cell = worksheet.cell(row=current_row, column=current_col, value=round(dish_info['dish_price'], 2))
        if num_materials > 1:
            worksheet.merge_cells(start_row=current_row, start_column=current_col,
                                end_row=current_row + num_materials - 1, end_column=current_col)
        cell.alignment = Alignment(vertical='center', horizontal='right')
        current_col += 1
        
        # Quantity sold
        cell = worksheet.cell(row=current_row, column=current_col, value=round(dish_info['quantity_sold'], 0))
        if num_materials > 1:
            worksheet.merge_cells(start_row=current_row, start_column=current_col,
                                end_row=current_row + num_materials - 1, end_column=current_col)
        cell.alignment = Alignment(vertical='center', horizontal='right')
        current_col += 1
        
        # Sales amount
        cell = worksheet.cell(row=current_row, column=current_col, value=round(dish_info['sales_amount'], 2))
        if num_materials > 1:
            worksheet.merge_cells(start_row=current_row, start_column=current_col,
                                end_row=current_row + num_materials - 1, end_column=current_col)
        cell.alignment = Alignment(vertical='center', horizontal='right')
        current_col += 1
        
        # Materials used
        if dish_info['materials']:
            for i, material in enumerate(dish_info['materials']):
                if debug:
                    # Detailed format with all info
                    # Calculate total material amount used (usage per dish × quantity sold)
                    total_material_used = material['usage_per_dish'] * dish_info['quantity_sold']
                    actual_usage = material.get('actual_usage', 0)
                    material_text = (f"{material['material_name']} ({material['material_number']}) "
                                   f"单价:{material['material_price']:.2f} "
                                   f"理论用量:{total_material_used:.2f} "
                                   f"实际用量:{actual_usage:.2f} "
                                   f"理论成本:{material['total_cost']:.2f} "
                                   f"实际成本:{material.get('actual_cost', 0):.2f}")
                else:
                    # Simple format - show both theoretical and proportioned actual costs
                    material_text = f"{material['material_name']}: 理论{material['total_cost']:.2f} 实际{material.get('actual_cost', 0):.2f}"
                
                worksheet.cell(row=current_row + i, column=current_col, value=material_text)
                worksheet.cell(row=current_row + i, column=current_col).border = thin_border
        else:
            worksheet.cell(row=current_row, column=current_col, value="无物料")
            worksheet.cell(row=current_row, column=current_col).border = thin_border
        current_col += 1
        
        # Theoretical total cost column
        cell = worksheet.cell(row=current_row, column=current_col, value=round(dish_info['total_material_cost'], 2))
        if num_materials > 1:
            worksheet.merge_cells(start_row=current_row, start_column=current_col,
                                end_row=current_row + num_materials - 1, end_column=current_col)
        cell.alignment = Alignment(vertical='center', horizontal='right')
        current_col += 1
        
        # Actual total cost column (now showing calculated proportional cost)
        cell = worksheet.cell(row=current_row, column=current_col, value=round(dish_info.get('actual_total_cost', 0), 2))
        if num_materials > 1:
            worksheet.merge_cells(start_row=current_row, start_column=current_col,
                                end_row=current_row + num_materials - 1, end_column=current_col)
        cell.alignment = Alignment(vertical='center', horizontal='right')
        current_col += 1
        
        # Gross profit (using actual cost now that we have it)
        actual_gross_profit = dish_info['sales_amount'] - dish_info.get('actual_total_cost', 0)
        cell = worksheet.cell(row=current_row, column=current_col, value=round(actual_gross_profit, 2))
        if num_materials > 1:
            worksheet.merge_cells(start_row=current_row, start_column=current_col,
                                end_row=current_row + num_materials - 1, end_column=current_col)
        cell.alignment = Alignment(vertical='center', horizontal='right')
        current_col += 1
        
        # Profit margin percentage (using actual cost)
        actual_profit_margin = (actual_gross_profit / dish_info['sales_amount'] * 100) if dish_info['sales_amount'] > 0 else 0
        cell = worksheet.cell(row=current_row, column=current_col, value=f"{actual_profit_margin:.1f}%")
        if num_materials > 1:
            worksheet.merge_cells(start_row=current_row, start_column=current_col,
                                end_row=current_row + num_materials - 1, end_column=current_col)
        cell.alignment = Alignment(vertical='center', horizontal='center')
        
        # Apply fill color based on actual profit margin
        if actual_profit_margin < 50:
            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
        elif actual_profit_margin > 70:
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
        
        # Apply borders and alternating fills
        fill_to_use = even_row_fill if dish_index % 2 == 0 else odd_row_fill
        for row in range(current_row, current_row + num_materials):
            for col in range(1, len(headers) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                # Don't override profit margin color
                if col != len(headers) or (actual_profit_margin >= 50 and actual_profit_margin <= 70):
                    cell.fill = fill_to_use
        
        current_row += num_materials
        dish_index += 1
    
    # Add summary row
    summary_row = current_row + 1
    worksheet.cell(row=summary_row, column=1, value="合计")
    worksheet.cell(row=summary_row, column=1).font = Font(bold=True)
    
    # Calculate totals
    total_sales = sum(d['sales_amount'] for d in data if not (d['is_combo'] and d['dish_price'] == 0))
    total_theoretical_cost = sum(d['total_material_cost'] for d in data if not (d['is_combo'] and d['dish_price'] == 0))
    # Sum up the proportioned actual costs from all dishes
    total_actual_cost = sum(d.get('actual_total_cost', 0) for d in data if not (d['is_combo'] and d['dish_price'] == 0))
    actual_total_profit = total_sales - total_actual_cost
    actual_avg_margin = (actual_total_profit / total_sales * 100) if total_sales > 0 else 0
    
    # Place totals in appropriate columns (adjusted for both theoretical and actual cost columns)
    if debug:
        worksheet.cell(row=summary_row, column=7, value=round(total_sales, 2))
        worksheet.cell(row=summary_row, column=9, value=round(total_theoretical_cost, 2))  # Theoretical total cost
        worksheet.cell(row=summary_row, column=10, value=round(total_actual_cost, 2))  # Actual total cost
        worksheet.cell(row=summary_row, column=11, value=round(actual_total_profit, 2))
        worksheet.cell(row=summary_row, column=12, value=f"{actual_avg_margin:.1f}%")
    else:
        worksheet.cell(row=summary_row, column=5, value=round(total_sales, 2))
        worksheet.cell(row=summary_row, column=7, value=round(total_theoretical_cost, 2))  # Theoretical total cost
        worksheet.cell(row=summary_row, column=8, value=round(total_actual_cost, 2))  # Actual total cost
        worksheet.cell(row=summary_row, column=9, value=round(actual_total_profit, 2))
        worksheet.cell(row=summary_row, column=10, value=f"{actual_avg_margin:.1f}%")
    
    # Apply bold font to summary row
    for col in range(1, len(headers) + 1):
        worksheet.cell(row=summary_row, column=col).font = Font(bold=True)
        worksheet.cell(row=summary_row, column=col).border = thin_border
    
    # Auto-adjust column widths (adjusted for both theoretical and actual cost columns)
    if debug:
        column_widths = [30, 15, 15, 10, 10, 10, 12, 60, 12, 12, 12, 12]
    else:
        column_widths = [30, 10, 10, 10, 12, 60, 12, 12, 12, 12]
    
    for i, width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width
    
    logger.info(f"Completed writing revenue sheet for {store_name}")


if __name__ == "__main__":
    """Test the functions with command line arguments."""
    import argparse
    from openpyxl import Workbook
    
    parser = argparse.ArgumentParser(description='Test store revenue sheet generation')
    parser.add_argument('--year', type=int, default=2025, help='Year (default: 2025)')
    parser.add_argument('--month', type=int, default=6, help='Month (default: 6)')
    parser.add_argument('--store-id', type=int, default=1, help='Store ID (default: 1)')
    parser.add_argument('--debug', action='store_true', help='Enable debug output')
    parser.add_argument('--test-db', action='store_true', help='Use test database')
    
    args = parser.parse_args()
    
    # Create database connection
    from utils.database import DatabaseConfig
    db_config = DatabaseConfig(is_test=args.test_db)
    db_manager = DatabaseManager(db_config)
    
    # Create a test workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Store {args.store_id}"
    
    # Generate the sheet
    write_store_revenue_sheet(
        worksheet=worksheet,
        db_manager=db_manager,
        year=args.year,
        month=args.month,
        store_id=args.store_id,
        store_name=f"测试店铺{args.store_id}",
        debug=args.debug
    )
    
    # Save to test file
    output_file = f"test_revenue_store_{args.store_id}_{args.year}_{args.month:02d}.xlsx"
    workbook.save(output_file)
    print(f"Test file saved to: {output_file}")