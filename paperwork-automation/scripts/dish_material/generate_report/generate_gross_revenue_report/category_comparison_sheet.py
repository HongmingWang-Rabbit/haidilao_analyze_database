#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate comparison sheet showing gross profit margins by dish broad type category across stores.
Uses the broad_type field from the dish table for categorization.
Uses actual material usage from material_monthly_usage table for cost calculations.
Includes month-over-month and year-over-year comparisons.
"""

import sys
from pathlib import Path
from typing import Dict, List, Tuple
import logging
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Add parent directory to path for imports
project_root = Path(__file__).resolve().parent.parent.parent.parent.parent
sys.path.insert(0, str(project_root))

from utils.database import DatabaseManager
from lib.config import STORE_ID_TO_NAME_MAPPING

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def get_category_gross_margin_data(db_manager: DatabaseManager, year: int, month: int) -> Dict:
    """
    Get gross margin data by dish broad type category for all stores.
    Includes current month, last month, and last year comparisons.
    Uses actual material usage distributed proportionally based on theoretical usage.
    
    Args:
        db_manager: Database manager instance
        year: Target year
        month: Target month
    
    Returns:
        Dict with structure:
        {
            store_id: {
                broad_type: {
                    'current': margin_percentage,
                    'last_month': margin_percentage,
                    'last_year': margin_percentage
                }
            }
        }
    """
    # Calculate previous month and year
    if month == 1:
        last_month = 12
        last_month_year = year - 1
    else:
        last_month = month - 1
        last_month_year = year
    
    last_year = year - 1
    
    result = {}
    
    with db_manager.get_connection() as conn:
        cursor = conn.cursor()
        
        # Query to get gross margin by broad type category using actual material usage distributed by theoretical proportions
        query = """
        WITH dish_sales AS (
            -- Get dish sales for each period
            SELECT 
                dms.store_id,
                dms.dish_id,
                d.broad_type,
                dms.year,
                dms.month,
                (COALESCE(dms.sale_amount, 0) - COALESCE(dms.return_amount, 0)) as net_quantity,
                COALESCE(dph.price, 0) as dish_price
            FROM dish_monthly_sale dms
            JOIN dish d ON d.id = dms.dish_id
            LEFT JOIN dish_price_history dph ON dph.dish_id = d.id 
                AND dph.store_id = dms.store_id
                AND dph.is_active = TRUE
            WHERE (
                -- Current month
                (dms.year = %s AND dms.month = %s) OR
                -- Last month
                (dms.year = %s AND dms.month = %s) OR
                -- Last year same month
                (dms.year = %s AND dms.month = %s)
            )
            AND dms.store_id != 101  -- Exclude Hi Bowl
        ),
        theoretical_material_usage AS (
            -- Calculate theoretical material usage by dish
            SELECT 
                ds.store_id,
                ds.year,
                ds.month,
                ds.dish_id,
                dm.material_id,
                SUM(ds.net_quantity * COALESCE(dm.standard_quantity, 0) * (1 + COALESCE(dm.loss_rate, 0))) as theoretical_usage
            FROM dish_sales ds
            LEFT JOIN dish_material dm ON dm.dish_id = ds.dish_id AND dm.store_id = ds.store_id
            WHERE dm.material_id IS NOT NULL
            GROUP BY ds.store_id, ds.year, ds.month, ds.dish_id, dm.material_id
        ),
        total_theoretical_by_material AS (
            -- Get total theoretical usage per material
            SELECT 
                store_id,
                year,
                month,
                material_id,
                SUM(theoretical_usage) as total_theoretical
            FROM theoretical_material_usage
            GROUP BY store_id, year, month, material_id
        ),
        actual_material_usage AS (
            -- Get actual material usage from material_monthly_usage (only 成本类)
            SELECT 
                mmu.store_id,
                mmu.year,
                mmu.month,
                mmu.material_id,
                COALESCE(mmu.material_used, 0) as actual_usage,
                COALESCE(mph.price, 0) as material_price
            FROM material_monthly_usage mmu
            JOIN material m ON m.id = mmu.material_id AND m.store_id = mmu.store_id
            LEFT JOIN material_price_history mph ON mph.material_id = m.id 
                AND mph.store_id = m.store_id
                AND mph.is_active = TRUE
            WHERE (
                -- Current month
                (mmu.year = %s AND mmu.month = %s) OR
                -- Last month
                (mmu.year = %s AND mmu.month = %s) OR
                -- Last year same month
                (mmu.year = %s AND mmu.month = %s)
            )
            AND mmu.material_use_type = '成本类'  -- Only include cost-type materials
        ),
        dish_actual_cost AS (
            -- Distribute actual cost to dishes based on theoretical usage proportions
            SELECT 
                tmu.store_id,
                tmu.year,
                tmu.month,
                tmu.dish_id,
                SUM(
                    CASE 
                        WHEN ttm.total_theoretical > 0 THEN 
                            (tmu.theoretical_usage / ttm.total_theoretical) * amu.actual_usage * amu.material_price
                        ELSE 0
                    END
                ) as dish_material_cost
            FROM theoretical_material_usage tmu
            JOIN total_theoretical_by_material ttm 
                ON tmu.store_id = ttm.store_id 
                AND tmu.year = ttm.year 
                AND tmu.month = ttm.month 
                AND tmu.material_id = ttm.material_id
            LEFT JOIN actual_material_usage amu 
                ON tmu.store_id = amu.store_id 
                AND tmu.year = amu.year 
                AND tmu.month = amu.month 
                AND tmu.material_id = amu.material_id
            GROUP BY tmu.store_id, tmu.year, tmu.month, tmu.dish_id
        ),
        category_aggregation AS (
            -- Aggregate by category
            SELECT 
                ds.store_id,
                COALESCE(ds.broad_type, '其他') as category,
                ds.year,
                ds.month,
                SUM(ds.net_quantity * ds.dish_price) as revenue,
                SUM(COALESCE(dac.dish_material_cost, 0)) as material_cost
            FROM dish_sales ds
            LEFT JOIN dish_actual_cost dac 
                ON ds.store_id = dac.store_id 
                AND ds.year = dac.year 
                AND ds.month = dac.month 
                AND ds.dish_id = dac.dish_id
            GROUP BY ds.store_id, COALESCE(ds.broad_type, '其他'), ds.year, ds.month
        )
        SELECT 
            store_id,
            category,
            year,
            month,
            revenue,
            material_cost,
            CASE 
                WHEN revenue > 0 THEN 
                    ((revenue - material_cost) / revenue) * 100
                ELSE 0
            END as gross_margin
        FROM category_aggregation
        ORDER BY store_id, category, year DESC, month DESC
        """
        
        cursor.execute(query, (
            year, month,  # Current month for dish_sales
            last_month_year, last_month,  # Last month for dish_sales
            last_year, month,  # Last year for dish_sales
            year, month,  # Current month for actual_material_usage
            last_month_year, last_month,  # Last month for actual_material_usage
            last_year, month  # Last year for actual_material_usage
        ))
        
        raw_data = cursor.fetchall()
        
        # Process data into nested dictionary
        for row in raw_data:
            store_id = row['store_id']
            category = row['category']
            period_year = row['year']
            period_month = row['month']
            margin = float(row['gross_margin']) if row['gross_margin'] else 0
            
            # Initialize store dict if not exists
            if store_id not in result:
                result[store_id] = {}
            
            # Initialize category dict if not exists
            if category not in result[store_id]:
                result[store_id][category] = {
                    'current': 0,
                    'last_month': 0,
                    'last_year': 0
                }
            
            # Assign margin to appropriate period
            if period_year == year and period_month == month:
                result[store_id][category]['current'] = margin
            elif period_year == last_month_year and period_month == last_month:
                result[store_id][category]['last_month'] = margin
            elif period_year == last_year and period_month == month:
                result[store_id][category]['last_year'] = margin
    
    return result


def get_all_categories(db_manager: DatabaseManager) -> List[str]:
    """
    Get all unique dish broad type categories from the database.
    
    Args:
        db_manager: Database manager instance
    
    Returns:
        List of broad type category names
    """
    with db_manager.get_connection() as conn:
        cursor = conn.cursor()
        
        query = """
        SELECT DISTINCT 
            COALESCE(broad_type, '其他') as category,
            CASE 
                WHEN COALESCE(broad_type, '其他') LIKE '%锅底%' THEN 1
                WHEN COALESCE(broad_type, '其他') LIKE '%荤%' OR COALESCE(broad_type, '其他') LIKE '%肉%' THEN 2
                WHEN COALESCE(broad_type, '其他') LIKE '%素菜%' THEN 3
                WHEN COALESCE(broad_type, '其他') LIKE '%小吃%' THEN 4
                WHEN COALESCE(broad_type, '其他') LIKE '%酒%' OR COALESCE(broad_type, '其他') LIKE '%饮%' THEN 5
                WHEN COALESCE(broad_type, '其他') LIKE '%套餐%' THEN 6
                WHEN COALESCE(broad_type, '其他') = '其他' THEN 8
                ELSE 7
            END as sort_order
        FROM dish
        WHERE is_active = TRUE
        ORDER BY sort_order, category
        """
        
        cursor.execute(query)
        categories = [row['category'] for row in cursor.fetchall() if row['category']]
    
    return categories


def write_category_comparison_sheet(worksheet, db_manager: DatabaseManager, year: int, month: int):
    """
    Write category comparison data to an Excel worksheet.
    
    Args:
        worksheet: Openpyxl worksheet object
        db_manager: Database manager instance
        year: Year for the data
        month: Month for the data
    """
    # Add title
    title_cell = worksheet.cell(row=1, column=1, value="各店铺菜品大类毛利对比分析")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    date_cell = worksheet.cell(row=2, column=1, value=f"{year}年{month}月")
    date_cell.font = Font(size=12)
    date_cell.alignment = Alignment(horizontal='center')
    
    # Get data
    margin_data = get_category_gross_margin_data(db_manager, year, month)
    categories = get_all_categories(db_manager)
    
    # Merge title cells
    num_cols = 1 + len(categories) * 5  # Store name + 5 columns per category
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    
    # Style definitions
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Colors for up/down indicators
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red
    
    # Create headers
    header_row1 = 4  # Category names
    header_row2 = 5  # Sub-headers
    
    # Store name header
    worksheet.cell(row=header_row1, column=1, value="门店")
    worksheet.merge_cells(start_row=header_row1, start_column=1, end_row=header_row2, end_column=1)
    cell = worksheet.cell(row=header_row1, column=1)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    worksheet.cell(row=header_row2, column=1).border = thin_border
    
    # Category headers
    current_col = 2
    for category in categories:
        # Main category header (merged across 5 columns)
        cell = worksheet.cell(row=header_row1, column=current_col, value=category)
        worksheet.merge_cells(start_row=header_row1, start_column=current_col, 
                            end_row=header_row1, end_column=current_col + 4)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Apply borders to merged cells
        for col in range(current_col, current_col + 5):
            worksheet.cell(row=header_row1, column=col).border = thin_border
        
        # Sub-headers
        sub_headers = ['本月', '上月', '去年', '环比', '同比']
        for i, sub_header in enumerate(sub_headers):
            cell = worksheet.cell(row=header_row2, column=current_col + i, value=sub_header)
            cell.font = Font(bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        current_col += 5
    
    # Data rows
    data_row = header_row2 + 1
    
    # Process each store
    for store_id in sorted(margin_data.keys()):
        store_name = STORE_ID_TO_NAME_MAPPING.get(store_id, f"Store {store_id}")
        
        # Store name
        worksheet.cell(row=data_row, column=1, value=store_name)
        worksheet.cell(row=data_row, column=1).border = thin_border
        
        # Category data
        current_col = 2
        for category in categories:
            category_data = margin_data[store_id].get(category, {
                'current': 0,
                'last_month': 0,
                'last_year': 0
            })
            
            # Current month
            current_val = category_data.get('current', 0)
            cell = worksheet.cell(row=data_row, column=current_col, value=f"{current_val:.1f}%")
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border
            
            # Last month
            last_month_val = category_data.get('last_month', 0)
            cell = worksheet.cell(row=data_row, column=current_col + 1, value=f"{last_month_val:.1f}%")
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border
            
            # Last year
            last_year_val = category_data.get('last_year', 0)
            cell = worksheet.cell(row=data_row, column=current_col + 2, value=f"{last_year_val:.1f}%")
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border
            
            # Month-over-month change (环比)
            if last_month_val != 0:
                mom_change = current_val - last_month_val
                cell = worksheet.cell(row=data_row, column=current_col + 3, value=f"{mom_change:+.1f}%")
                if mom_change > 0:
                    cell.fill = green_fill
                elif mom_change < 0:
                    cell.fill = red_fill
            else:
                cell = worksheet.cell(row=data_row, column=current_col + 3, value="N/A")
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border
            
            # Year-over-year change (同比)
            if last_year_val != 0:
                yoy_change = current_val - last_year_val
                cell = worksheet.cell(row=data_row, column=current_col + 4, value=f"{yoy_change:+.1f}%")
                if yoy_change > 0:
                    cell.fill = green_fill
                elif yoy_change < 0:
                    cell.fill = red_fill
            else:
                cell = worksheet.cell(row=data_row, column=current_col + 4, value="N/A")
            cell.alignment = Alignment(horizontal='right')
            cell.border = thin_border
            
            current_col += 5
        
        data_row += 1
    
    # Add average row
    avg_row = data_row + 1
    worksheet.cell(row=avg_row, column=1, value="平均")
    worksheet.cell(row=avg_row, column=1).font = Font(bold=True)
    worksheet.cell(row=avg_row, column=1).border = thin_border
    
    # Calculate and write averages for each category
    current_col = 2
    for category in categories:
        # Calculate averages
        current_vals = []
        last_month_vals = []
        last_year_vals = []
        
        for store_id in margin_data:
            if category in margin_data[store_id]:
                current_vals.append(margin_data[store_id][category].get('current', 0))
                last_month_vals.append(margin_data[store_id][category].get('last_month', 0))
                last_year_vals.append(margin_data[store_id][category].get('last_year', 0))
        
        # Current month average
        avg_current = sum(current_vals) / len(current_vals) if current_vals else 0
        cell = worksheet.cell(row=avg_row, column=current_col, value=f"{avg_current:.1f}%")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border
        
        # Last month average
        avg_last_month = sum(last_month_vals) / len(last_month_vals) if last_month_vals else 0
        cell = worksheet.cell(row=avg_row, column=current_col + 1, value=f"{avg_last_month:.1f}%")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border
        
        # Last year average
        avg_last_year = sum(last_year_vals) / len(last_year_vals) if last_year_vals else 0
        cell = worksheet.cell(row=avg_row, column=current_col + 2, value=f"{avg_last_year:.1f}%")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border
        
        # Average MoM change
        if avg_last_month != 0:
            avg_mom = avg_current - avg_last_month
            cell = worksheet.cell(row=avg_row, column=current_col + 3, value=f"{avg_mom:+.1f}%")
            if avg_mom > 0:
                cell.fill = green_fill
            elif avg_mom < 0:
                cell.fill = red_fill
        else:
            cell = worksheet.cell(row=avg_row, column=current_col + 3, value="N/A")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border
        
        # Average YoY change
        if avg_last_year != 0:
            avg_yoy = avg_current - avg_last_year
            cell = worksheet.cell(row=avg_row, column=current_col + 4, value=f"{avg_yoy:+.1f}%")
            if avg_yoy > 0:
                cell.fill = green_fill
            elif avg_yoy < 0:
                cell.fill = red_fill
        else:
            cell = worksheet.cell(row=avg_row, column=current_col + 4, value="N/A")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
        cell.border = thin_border
        
        current_col += 5
    
    # Auto-adjust column widths
    worksheet.column_dimensions[get_column_letter(1)].width = 15  # Store name
    for col in range(2, current_col):
        worksheet.column_dimensions[get_column_letter(col)].width = 10
    
    # Add legend
    legend_row = avg_row + 2
    legend_text = "说明：绿色 = 上升    红色 = 下降    环比 = 本月 vs 上月    同比 = 本月 vs 去年同月"
    legend_cell = worksheet.cell(row=legend_row, column=1, value=legend_text)
    legend_cell.font = Font(italic=True, size=10)
    legend_cell.alignment = Alignment(horizontal='left')
    
    # Merge cells for the legend (span across multiple columns for better visibility)
    worksheet.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=10)
    
    # Add color examples
    worksheet.cell(row=legend_row + 1, column=1, value="示例：")
    worksheet.cell(row=legend_row + 1, column=2, value="+5.0%")
    worksheet.cell(row=legend_row + 1, column=2).fill = green_fill
    worksheet.cell(row=legend_row + 1, column=3, value="-3.0%")
    worksheet.cell(row=legend_row + 1, column=3).fill = red_fill
    
    # Add note about actual material usage
    note_row = legend_row + 3
    note_text = "注：物料成本为实际库存消耗数据（仅成本类物料，按理论用量比例分配到各菜品后汇总），非理论计算值"
    note_cell = worksheet.cell(row=note_row, column=1, value=note_text)
    note_cell.font = Font(italic=True, size=10, color="666666")
    worksheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
    
    logger.info("Completed writing category comparison sheet")


if __name__ == "__main__":
    """Test the functions with command line arguments."""
    import argparse
    from openpyxl import Workbook
    
    parser = argparse.ArgumentParser(description='Test category comparison sheet generation')
    parser.add_argument('--year', type=int, default=2025, help='Year (default: 2025)')
    parser.add_argument('--month', type=int, default=6, help='Month (default: 6)')
    parser.add_argument('--test-db', action='store_true', help='Use test database')
    
    args = parser.parse_args()
    
    # Create database connection
    from utils.database import DatabaseConfig
    db_config = DatabaseConfig(is_test=args.test_db)
    db_manager = DatabaseManager(db_config)
    
    # Create a test workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "分类对比"
    
    # Generate the sheet
    write_category_comparison_sheet(
        worksheet=worksheet,
        db_manager=db_manager,
        year=args.year,
        month=args.month
    )
    
    # Save to test file
    output_file = f"test_category_comparison_{args.year}_{args.month:02d}.xlsx"
    workbook.save(output_file)
    print(f"Test file saved to: {output_file}")