#!/usr/bin/env python3
"""
Final verification test for the monthly material report with new cost columns
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.generate_monthly_material_report import MonthlyMaterialReportGenerator
import logging
from datetime import datetime

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def test_final_report_verification():
    """Test the monthly material report with unique filename"""
    
    print("Final Report Verification Test")
    print("=" * 35)
    
    try:
        # Generate the monthly material report with unique timestamp
        target_date = "2025-05-31"
        generator = MonthlyMaterialReportGenerator(target_date, is_test=True)
        
        # Override save method to use unique filename
        original_save = generator.save_report
        
        def save_with_timestamp(wb):
            try:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"VERIFICATION_monthly_material_report_{timestamp}.xlsx"
                output_path = generator.output_dir / filename
                wb.save(output_path)
                return output_path
            except Exception as e:
                print(f"ERROR: Error saving report: {e}")
                return None
        
        generator.save_report = save_with_timestamp
        
        print(f"Target date: {target_date}")
        print("Generating monthly material report...")
        
        # Generate the report
        output_file = generator.generate_report()
        
        if output_file:
            print(f"✅ Report generated successfully!")
            print(f"📁 Output file: {output_file}")
            
            # Check if the file exists
            if Path(output_file).exists():
                print(f"📊 File exists - Size: {Path(output_file).stat().st_size} bytes")
                
                # Quick check using openpyxl
                from openpyxl import load_workbook
                try:
                    wb = load_workbook(output_file)
                    print(f"📋 Worksheets in file: {wb.sheetnames}")
                    
                    # Check the variance worksheet specifically
                    if "物料用量差异分析" in wb.sheetnames:
                        ws = wb["物料用量差异分析"]
                        print(f"📊 物料用量差异分析 dimensions: {ws.max_row} rows × {ws.max_column} columns")
                        
                        # Find header row and verify columns
                        for row in range(1, 20):
                            cell_value = ws.cell(row=row, column=1).value
                            if cell_value and "序号" in str(cell_value):
                                header_row = row
                                print(f"📋 Header row found at row {header_row}")
                                
                                # Check for our new columns
                                col_count = ws.max_column
                                print(f"📊 Total columns: {col_count}")
                                
                                # Check last two columns
                                cost_col = ws.cell(row=header_row, column=col_count-1).value
                                variance_col = ws.cell(row=header_row, column=col_count).value
                                
                                print(f"📈 Column {col_count-1}: {cost_col}")
                                print(f"📈 Column {col_count}: {variance_col}")
                                
                                if "本月总消费金额" in str(cost_col) and "差异金额" in str(variance_col):
                                    print("✅ New cost columns confirmed in generated report!")
                                else:
                                    print("❌ Cost columns not found in expected positions")
                                break
                        
                        # Check for 包装规格 column (should be removed)
                        package_found = False
                        for col in range(1, ws.max_column + 1):
                            header_cell = ws.cell(row=header_row, column=col)
                            if header_cell.value and "包装规格" in str(header_cell.value):
                                package_found = True
                                break
                        
                        if package_found:
                            print("❌ 包装规格 column still present - not removed")
                        else:
                            print("✅ 包装规格 column successfully removed")
                            
                    wb.close()
                except Exception as e:
                    print(f"⚠️ Could not verify file contents: {e}")
                
                return True
            else:
                print(f"❌ Output file not found: {output_file}")
                return False
        else:
            print("❌ Report generation failed - no output file returned")
            return False
            
    except Exception as e:
        print(f"❌ Error during report generation: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_final_report_verification()
    if success:
        print("\n🎉 Final verification completed successfully!")
    else:
        print("\n❌ Final verification failed")
        sys.exit(1)