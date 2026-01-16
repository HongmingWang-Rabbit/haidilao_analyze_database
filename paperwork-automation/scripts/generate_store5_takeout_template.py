#!/usr/bin/env python3
"""
One-time script to generate Store 5 takeout revenue comparison template.
Fills the template with monthly takeout revenue data for 2024 vs 2025.

Template format:
| 外卖收入 | 01月 | 02月 | ... | 12月 | 求和 |
| 2025年  | ...  | ...  | ... | ...  | sum  |
| 2024年  | ...  | ...  | ... | ...  | sum  |
| 差额    | ...  | ...  | ... | ...  | diff |
"""

import sys
import codecs
from pathlib import Path

# Fix encoding issues on Windows for Chinese characters
if sys.platform.startswith('win'):
    if hasattr(sys.stdout, 'buffer'):
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
    if hasattr(sys.stderr, 'buffer'):
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from utils.database import get_database_manager
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime


def get_monthly_takeout_revenue(db_manager, store_id: int, year: int) -> dict:
    """
    Get monthly takeout revenue for a specific store and year.

    Returns:
        Dict with month number (1-12) as key and revenue amount as value
    """
    query = """
        SELECT
            EXTRACT(MONTH FROM date) as month,
            SUM(amount) as total_revenue
        FROM daily_takeout_revenue
        WHERE store_id = %s
        AND EXTRACT(YEAR FROM date) = %s
        GROUP BY EXTRACT(MONTH FROM date)
        ORDER BY month
    """

    results = db_manager.fetch_all(query, (store_id, year))

    monthly_data = {i: 0.0 for i in range(1, 13)}
    for row in results:
        month = int(row['month'])
        monthly_data[month] = float(row['total_revenue'])

    return monthly_data


def generate_takeout_template():
    """Generate the takeout revenue template for Store 5."""
    print("=" * 60)
    print("Store 5 Takeout Revenue Template Generator")
    print("=" * 60)

    # Connect to database
    db_manager = get_database_manager(is_test=False)

    if not db_manager.test_connection():
        print("ERROR: Database connection failed")
        return False

    print("Database connection successful")

    # Get data for Store 5
    store_id = 5
    store_name = "加拿大五店"

    print(f"\nFetching takeout revenue data for {store_name} (Store ID: {store_id})...")

    # Get monthly data for 2024 and 2025
    data_2025 = get_monthly_takeout_revenue(db_manager, store_id, 2025)
    data_2024 = get_monthly_takeout_revenue(db_manager, store_id, 2024)

    # Calculate differences
    differences = {month: data_2025[month] - data_2024[month] for month in range(1, 13)}

    # Calculate totals
    total_2025 = sum(data_2025.values())
    total_2024 = sum(data_2024.values())
    total_diff = total_2025 - total_2024

    # Print summary
    print(f"\n2025 Total: ${total_2025:,.2f}")
    print(f"2024 Total: ${total_2024:,.2f}")
    print(f"Difference: ${total_diff:,.2f}")

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "外卖收入对比"

    # Define styles
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header fill colors
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    year_2025_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    year_2024_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    diff_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    # Write header row
    headers = ['外卖收入'] + [f'{m:02d}月' for m in range(1, 13)] + ['求和']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    # Write 2025 row
    row_2025 = ['2025年'] + [data_2025[m] for m in range(1, 13)] + [total_2025]
    for col, value in enumerate(row_2025, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = year_2025_fill
        if col > 1:  # Number formatting for values
            cell.number_format = '#,##0.00'

    # Write 2024 row
    row_2024 = ['2024年'] + [data_2024[m] for m in range(1, 13)] + [total_2024]
    for col, value in enumerate(row_2024, 1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = year_2024_fill
        if col > 1:
            cell.number_format = '#,##0.00'

    # Write difference row
    row_diff = ['差额'] + [differences[m] for m in range(1, 13)] + [total_diff]
    for col, value in enumerate(row_diff, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = diff_fill
        if col > 1:
            cell.number_format = '#,##0.00'

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    for col_letter in 'BCDEFGHIJKLMN':
        ws.column_dimensions[col_letter].width = 12

    # Save the file
    output_dir = project_root / 'output'
    output_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = output_dir / f'store5_takeout_comparison_{timestamp}.xlsx'

    wb.save(output_file)
    print(f"\nOutput saved to: {output_file}")

    # Also print the table to console
    print("\n" + "=" * 80)
    print("Preview:")
    print("=" * 80)

    # Header
    header_line = f"{'外卖收入':<10}"
    for m in range(1, 13):
        header_line += f"{f'{m:02d}月':>10}"
    header_line += f"{'求和':>12}"
    print(header_line)

    # 2025
    line_2025 = f"{'2025年':<10}"
    for m in range(1, 13):
        line_2025 += f"{data_2025[m]:>10,.2f}"
    line_2025 += f"{total_2025:>12,.2f}"
    print(line_2025)

    # 2024
    line_2024 = f"{'2024年':<10}"
    for m in range(1, 13):
        line_2024 += f"{data_2024[m]:>10,.2f}"
    line_2024 += f"{total_2024:>12,.2f}"
    print(line_2024)

    # Difference
    line_diff = f"{'差额':<10}"
    for m in range(1, 13):
        line_diff += f"{differences[m]:>10,.2f}"
    line_diff += f"{total_diff:>12,.2f}"
    print(line_diff)

    print("=" * 80)

    return True


if __name__ == "__main__":
    success = generate_takeout_template()
    sys.exit(0 if success else 1)
