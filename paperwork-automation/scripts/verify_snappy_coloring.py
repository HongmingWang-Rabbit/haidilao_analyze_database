#!/usr/bin/env python3
"""
Verify SNAPPY transaction coloring markings in bank output
"""

from openpyxl import load_workbook
from pathlib import Path


def verify_snappy_coloring():
    """Verify SNAPPY transactions have correct coloring markings"""
    print("🎨 VERIFYING SNAPPY COLORING MARKINGS")
    print("=" * 60)

    # Check the latest output file
    output_files = list(Path("output").glob("Bank_Transactions_Report_*.xlsx"))
    if not output_files:
        print("❌ No output files found")
        return

    latest_file = max(output_files, key=lambda p: p.stat().st_mtime)
    print(f"📁 Examining: {latest_file}")

    try:
        wb = load_workbook(latest_file)

        snappy_transactions = []

        # Check all sheets for SNAPPY transactions
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Check all transaction rows
            for row in range(3, ws.max_row + 1):
                # Column H - Details
                details_cell = ws.cell(row=row, column=8).value
                transaction_type = ws.cell(
                    row=row, column=9).value  # Column I - 品名
                description = ws.cell(
                    row=row, column=10).value  # Column J - 付款详情

                # Column K: 单据号, L: 附件, M: 是否登记线下付款表, N: 是否登记支票使用表
                doc_num = ws.cell(row=row, column=11).value  # 单据号
                attachment = ws.cell(row=row, column=12).value  # 附件
                offline_payment = ws.cell(
                    row=row, column=13).value  # 是否登记线下付款表
                check_usage = ws.cell(row=row, column=14).value  # 是否登记支票使用表

                if details_cell and "SNAPPY" in str(details_cell).upper():
                    snappy_transactions.append({
                        'sheet': sheet_name,
                        'row': row,
                        'details': str(details_cell),
                        'type': str(transaction_type) if transaction_type else "",
                        'description': str(description) if description else "",
                        'doc_num': str(doc_num) if doc_num else "",
                        'attachment': str(attachment) if attachment else "",
                        'offline_payment': str(offline_payment) if offline_payment else "",
                        'check_usage': str(check_usage) if check_usage else ""
                    })

        print(f"\n📊 Found {len(snappy_transactions)} SNAPPY transactions:")

        snappydebit_correct = 0
        snappyon_correct = 0
        total_errors = 0

        for i, transaction in enumerate(snappy_transactions, 1):
            details = transaction['details']
            trans_type = transaction['type']

            print(
                f"\n🔍 Transaction {i} ({transaction['sheet']} Row {transaction['row']}):")
            print(f"   Details: '{details}'")
            print(f"   Type: '{trans_type}'")

            # Expected markings based on configuration
            if "SNAPPYDEBIT" in details.upper():
                expected_doc = "是"  # True
                expected_attach = "是"  # True
                expected_offline = "是"  # True
                expected_check = ""  # False
                transaction_category = "SNAPPYDEBIT (Platform Fee)"
            elif "SNAPPYON" in details.upper():
                expected_doc = "是"  # True
                expected_attach = "是"  # True
                expected_offline = ""  # False
                expected_check = ""  # False
                transaction_category = "SNAPPYON (Income Received)"
            else:
                continue  # Skip other SNAPPY types

            # Verify markings
            print(f"   Category: {transaction_category}")
            print(f"   Markings:")

            # Check each field
            errors_in_transaction = 0

            # 单据号
            actual_doc = transaction['doc_num']
            if actual_doc == expected_doc:
                print(
                    f"     ✅ 单据号: '{actual_doc}' (Expected: '{expected_doc}')")
            else:
                print(
                    f"     ❌ 单据号: '{actual_doc}' (Expected: '{expected_doc}')")
                errors_in_transaction += 1

            # 附件
            actual_attach = transaction['attachment']
            if actual_attach == expected_attach:
                print(
                    f"     ✅ 附件: '{actual_attach}' (Expected: '{expected_attach}')")
            else:
                print(
                    f"     ❌ 附件: '{actual_attach}' (Expected: '{expected_attach}')")
                errors_in_transaction += 1

            # 是否登记线下付款表
            actual_offline = transaction['offline_payment']
            if actual_offline == expected_offline:
                print(
                    f"     ✅ 线下付款表: '{actual_offline}' (Expected: '{expected_offline}')")
            else:
                print(
                    f"     ❌ 线下付款表: '{actual_offline}' (Expected: '{expected_offline}')")
                errors_in_transaction += 1

            # 是否登记支票使用表
            actual_check = transaction['check_usage']
            if actual_check == expected_check:
                print(
                    f"     ✅ 支票使用表: '{actual_check}' (Expected: '{expected_check}')")
            else:
                print(
                    f"     ❌ 支票使用表: '{actual_check}' (Expected: '{expected_check}')")
                errors_in_transaction += 1

            # Count correct transactions
            if errors_in_transaction == 0:
                if "SNAPPYDEBIT" in details.upper():
                    snappydebit_correct += 1
                elif "SNAPPYON" in details.upper():
                    snappyon_correct += 1
                print(f"     ✅ All markings correct!")
            else:
                total_errors += errors_in_transaction
                print(f"     ❌ {errors_in_transaction} marking errors")

        print(f"\n📋 COLORING VERIFICATION SUMMARY:")
        print(f"   ✅ SNAPPYDEBIT correctly marked: {snappydebit_correct}")
        print(f"   ✅ SNAPPYON correctly marked: {snappyon_correct}")
        print(f"   ❌ Total marking errors: {total_errors}")
        print(f"   📊 Total SNAPPY transactions: {len(snappy_transactions)}")

        if total_errors == 0:
            print(f"\n🎨 SUCCESS: All SNAPPY transactions have correct coloring markings!")
        else:
            print(
                f"\n❌ ISSUES: {total_errors} coloring markings need correction")

    except Exception as e:
        print(f"❌ Error reading output file: {e}")


if __name__ == "__main__":
    verify_snappy_coloring()
