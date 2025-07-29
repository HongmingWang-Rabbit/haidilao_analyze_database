#!/usr/bin/env python3
"""
Test the enhanced dish usage details with comprehensive information
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.monthly_dishes_worksheet import MonthlyDishesWorksheetGenerator
from lib.database_queries import ReportDataProvider
from utils.database import DatabaseManager, DatabaseConfig
from openpyxl import Workbook
from datetime import datetime

def test_enhanced_dish_usage():
    """Test the enhanced dish usage details"""
    
    print("Testing Enhanced Dish Usage Details")
    print("=" * 40)
    
    try:
        # Initialize components
        config = DatabaseConfig(is_test=True)
        db_manager = DatabaseManager(config)
        data_provider = ReportDataProvider(db_manager)
        
        # Create worksheet generator
        target_date = "2025-05-31"
        store_names = ["加拿大一店", "加拿大二店", "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大七店"]
        worksheet_gen = MonthlyDishesWorksheetGenerator(store_names, target_date)
        
        # Create workbook and generate worksheet
        wb = Workbook()
        ws = worksheet_gen.generate_material_variance_worksheet(wb, data_provider)
        
        print(f"✅ Worksheet generated: {ws.title}")
        print(f"📊 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Find the header row
        header_row = None
        for row in range(1, 20):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "序号" in str(cell_value):
                header_row = row
                break
        
        if header_row:
            print(f"📋 Header row found at row {header_row}")
            
            # Check enhanced dish usage details
            print(f"\n📊 Enhanced dish usage details samples:")
            
            sample_count = 0
            for row in range(header_row + 1, ws.max_row + 1):
                if sample_count >= 3:
                    break
                    
                usage_cell = ws.cell(row=row, column=6)  # Column F: 使用菜品详情
                material_cell = ws.cell(row=row, column=3)  # Material name
                store_cell = ws.cell(row=row, column=2)     # Store name
                
                if usage_cell.value and "materials_use-" in str(usage_cell.value):
                    material_name = material_cell.value[:25] if material_cell.value else "Unknown"
                    store_name = store_cell.value if store_cell.value else "Unknown"
                    
                    print(f"\n   📍 Row {row}: {store_name} | {material_name}...")
                    
                    # Parse and display each line
                    usage_lines = str(usage_cell.value).split('\n')
                    for i, line in enumerate(usage_lines):
                        if line.strip():
                            print(f"      Line {i+1}: {line}")
                            
                            # Verify format components
                            if "materials_use-" in line:
                                components = {
                                    'sale': 'sale-' in line,
                                    'portion': '出品分量(kg)-' in line,
                                    'loss': '损耗-' in line,
                                    'unit': '物料单位-' in line,
                                    'material_use': 'materials_use-' in line
                                }
                                
                                missing = [k for k, v in components.items() if not v]
                                if missing:
                                    print(f"         ⚠️ Missing components: {missing}")
                                else:
                                    print(f"         ✅ All components present")
                    
                    sample_count += 1
            
            # Test specific material (3000759 清油底料)
            print(f"\n🔍 Testing specific material 3000759 清油底料:")
            
            found_enhanced = False
            for row in range(header_row + 1, min(header_row + 100, ws.max_row + 1)):
                material_num_cell = ws.cell(row=row, column=4)
                if material_num_cell.value and "3000759" in str(material_num_cell.value):
                    usage_cell = ws.cell(row=row, column=6)
                    store_cell = ws.cell(row=row, column=2)
                    
                    print(f"   📍 Found at row {row}: {store_cell.value}")
                    usage_text = str(usage_cell.value)
                    
                    # Check for enhanced format
                    if "materials_use-" in usage_text and "出品分量(kg)-" in usage_text:
                        found_enhanced = True
                        print(f"   ✅ Enhanced format confirmed:")
                        
                        lines = usage_text.split('\n')
                        for line in lines[:3]:  # Show first 3 lines
                            if line.strip():
                                print(f"      {line}")
                        
                        # Check for combo usage
                        combo_lines = [line for line in lines if line.startswith('套餐')]
                        if combo_lines:
                            print(f"   📦 Combo usage: {combo_lines[0]}")
                        else:
                            print(f"   📦 No combo usage (expected if zero)")
                    else:
                        print(f"   ❌ Enhanced format not found")
                        print(f"      Raw text: {usage_text[:100]}...")
                    
                    break
            
            if not found_enhanced:
                print(f"   ❌ Material 3000759 not found or enhanced format not applied")
            
            # Check for proper combo calculations
            print(f"\n📦 Checking combo usage calculations:")
            combo_found = 0
            for row in range(header_row + 1, min(header_row + 50, ws.max_row + 1)):
                usage_cell = ws.cell(row=row, column=6)
                if usage_cell.value and "套餐 -" in str(usage_cell.value):
                    combo_line = [line for line in str(usage_cell.value).split('\n') if line.startswith('套餐')][0]
                    combo_found += 1
                    if combo_found <= 3:  # Show first 3 examples
                        material_cell = ws.cell(row=row, column=3)
                        print(f"   Row {row}: {material_cell.value[:20]}... | {combo_line}")
            
            print(f"   Total rows with combo usage: {combo_found}")
        
        # Save output
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"output/test_enhanced_dish_usage_{timestamp}.xlsx"
        wb.save(output_file)
        print(f"\n📁 Test output saved to: {output_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_enhanced_dish_usage()