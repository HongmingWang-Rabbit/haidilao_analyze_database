#!/usr/bin/env python3
"""
Generate monthly-only database report.
Works with monthly performance data (dish_monthly_sale, material_monthly_usage) 
instead of expecting daily_report data.
"""

from utils.database import DatabaseConfig, DatabaseManager
from lib.monthly_dishes_worksheet import MonthlyDishesWorksheetGenerator
from openpyxl import Workbook
from dotenv import load_dotenv
from datetime import datetime
import os
import sys
from pathlib import Path

# Add parent directory to path for imports FIRST
sys.path.insert(0, str(Path(__file__).parent.parent))

# Now import our modules

# Load environment variables
load_dotenv()


class MonthlyReportGenerator:
    """Monthly report generator for monthly-only data"""

    def __init__(self, target_date: str, is_test: bool = False):
        self.target_date = target_date
        self.is_test = is_test
        self.config = DatabaseConfig(is_test=is_test)
        self.db_manager = DatabaseManager(self.config)
        self.output_dir = Path(os.getenv('OUTPUT_DIR', './output'))
        self.output_dir.mkdir(exist_ok=True)

        # Store mapping
        self.store_names = {
            1: "加拿大一店", 2: "加拿大二店", 3: "加拿大三店", 4: "加拿大四店",
            5: "加拿大五店", 6: "加拿大六店", 7: "加拿大七店"
        }

        # Simple data provider for monthly data
        self.data_provider = self

    def get_monthly_data_summary(self):
        """Get summary of available monthly data"""
        try:
            target_dt = datetime.strptime(self.target_date, '%Y-%m-%d')
            year, month = target_dt.year, target_dt.month

            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()

                # Check dish monthly sales
                cursor.execute("""
                    SELECT COUNT(*) as count, COUNT(DISTINCT dish_id) as dishes,
                           SUM(sale_amount) as total_sales, SUM(return_amount) as total_returns
                    FROM dish_monthly_sale 
                    WHERE year = %s AND month = %s
                """, (year, month))
                dish_stats = cursor.fetchone()

                # Check material monthly usage
                cursor.execute("""
                    SELECT COUNT(*) as count, COUNT(DISTINCT material_id) as materials,
                           SUM(material_used) as total_usage
                    FROM material_monthly_usage 
                    WHERE year = %s AND month = %s
                """, (year, month))
                material_stats = cursor.fetchone()

                return {
                    'dish_sales': dish_stats,
                    'material_usage': material_stats,
                    'year': year,
                    'month': month
                }

        except Exception as e:
            print(f"ERROR: Error getting monthly data summary: {e}")
            return None

    def generate_report(self):
        """Generate monthly report with available data"""
        print(f"INFO: Generating monthly report for {self.target_date}")

        # Check what data is available
        data_summary = self.get_monthly_data_summary()
        if not data_summary:
            print("ERROR: Could not access monthly data")
            return None

        dish_stats = data_summary['dish_sales']
        material_stats = data_summary['material_usage']
        year, month = data_summary['year'], data_summary['month']

        print(f"INFO: Available data for {year}-{month:02d}:")
        print(
            f"   DISH: Dish sales: {dish_stats['count']} records, {dish_stats['dishes']} dishes")
        print(
            f"   MATERIAL: Material usage: {material_stats['count']} records, {material_stats['materials']} materials")

        if dish_stats['count'] == 0 and material_stats['count'] == 0:
            print("ERROR: No monthly data found")
            return None

        # Create Excel workbook
        wb = Workbook()
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)

        # Generate material variance analysis worksheet
        monthly_dishes_generator = MonthlyDishesWorksheetGenerator(
            self.store_names, self.target_date)
        variance_ws = monthly_dishes_generator.generate_material_variance_worksheet(
            wb, self)

        # Add a simple summary worksheet
        summary_ws = self.create_summary_worksheet(wb, data_summary)

        if not wb.worksheets:
            print("ERROR: No worksheets generated")
            return None

        # Save the report
        output_path = self.save_report(wb)
        if output_path:
            print(f"SUCCESS: Monthly report generated: {output_path}")
            return output_path
        else:
            print("ERROR: Failed to save report")
            return None

    def create_summary_worksheet(self, wb, data_summary):
        """Create a summary worksheet with monthly statistics"""
        ws = wb.create_sheet("月度统计概览", 0)  # Insert as first sheet

        from openpyxl.styles import PatternFill, Font, Alignment

        # Title
        ws['A1'] = f"海底捞月度数据统计 - {data_summary['year']}年{data_summary['month']:02d}月"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(
            start_color="C41E3A", end_color="C41E3A", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')

        # Statistics
        current_row = 3

        # Dish statistics
        ws[f'A{current_row}'] = "菜品销售统计"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 1

        dish_stats = data_summary['dish_sales']
        dish_data = [
            ("销售记录数", dish_stats['count']),
            ("销售菜品数", dish_stats['dishes']),
            ("总销售量",
             f"{dish_stats['total_sales']:.0f}" if dish_stats['total_sales'] else "0"),
            ("总退菜量",
             f"{dish_stats['total_returns']:.0f}" if dish_stats['total_returns'] else "0")
        ]

        for label, value in dish_data:
            ws[f'A{current_row}'] = label
            ws[f'B{current_row}'] = value
            current_row += 1

        current_row += 1

        # Material statistics
        ws[f'A{current_row}'] = "物料使用统计"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 1

        material_stats = data_summary['material_usage']
        material_data = [
            ("使用记录数", material_stats['count']),
            ("使用物料数", material_stats['materials']),
            ("总消耗量",
             f"{material_stats['total_usage']:.2f}" if material_stats['total_usage'] else "0")
        ]

        for label, value in material_data:
            ws[f'A{current_row}'] = label
            ws[f'B{current_row}'] = value
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

        return ws

    def save_report(self, wb):
        """Save the Excel workbook to file"""
        try:
            filename = f"monthly_report_{self.target_date.replace('-', '_')}.xlsx"
            output_path = self.output_dir / filename
            wb.save(output_path)
            return output_path
        except Exception as e:
            print(f"ERROR: Error saving report: {e}")
            return None


def main():
    """Main function"""
    import argparse

    parser = argparse.ArgumentParser(
        description="Generate monthly database report")
    parser.add_argument("--date", default="2025-06-01",
                        help="Target date (YYYY-MM-DD)")
    parser.add_argument("--test", action="store_true",
                        help="Use test database")

    args = parser.parse_args()

    try:
        generator = MonthlyReportGenerator(args.date, is_test=args.test)
        output_path = generator.generate_report()

        if output_path:
            print("SUCCESS: Monthly report generation completed successfully!")
        else:
            print("ERROR: Failed to generate monthly report")
            sys.exit(1)

    except Exception as e:
        print(f"ERROR: Failed to generate monthly report - {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
