#!/usr/bin/env python3
"""
Final verification test for status coloring in complete monthly report
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.generate_monthly_material_report import MonthlyMaterialReportGenerator
from datetime import datetime
from openpyxl import load_workbook

def test_final_coloring_verification():
    """Test coloring in the complete monthly material report"""
    
    print("Final Coloring Verification Test")
    print("=" * 35)
    
    try:
        # Generate the monthly material report with unique timestamp
        target_date = "2025-05-31"
        generator = MonthlyMaterialReportGenerator(target_date, is_test=True)
        
        # Override save method to use unique filename
        def save_with_timestamp(wb):
            try:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"COLORING_TEST_monthly_material_report_{timestamp}.xlsx"
                output_path = generator.output_dir / filename
                wb.save(output_path)
                return output_path
            except Exception as e:
                print(f"ERROR: Error saving report: {e}")
                return None
        
        generator.save_report = save_with_timestamp
        
        print(f"Target date: {target_date}")
        print("Generating monthly material report...")
        
        # Generate the report
        output_file = generator.generate_report()
        
        if output_file and Path(output_file).exists():
            print(f"✅ Report generated successfully!")
            print(f"📁 Output file: {output_file}")
            
            # Load and verify coloring
            wb = load_workbook(output_file)
            
            if "物料用量差异分析" in wb.sheetnames:
                ws = wb["物料用量差异分析"]
                print(f"📊 Analyzing 物料用量差异分析 worksheet...")
                
                # Find header row
                header_row = None
                for row in range(1, 20):
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value and "序号" in str(cell_value):
                        header_row = row
                        break
                
                if header_row:
                    print(f"📋 Header row found at row {header_row}")
                    
                    # Analyze status and colors
                    color_analysis = {"超量": {"red": 0, "green": 0, "none": 0},
                                    "不足": {"red": 0, "green": 0, "none": 0},
                                    "正常": {"red": 0, "green": 0, "none": 0}}
                    
                    for row in range(header_row + 1, min(header_row + 51, ws.max_row + 1)):
                        status_cell = ws.cell(row=row, column=12)  # Status column
                        status = status_cell.value if status_cell.value else ""
                        
                        if status in color_analysis:
                            fill_color = status_cell.fill.start_color.rgb if status_cell.fill and status_cell.fill.start_color else None
                            
                            if fill_color == "00FFE6E6":
                                color_analysis[status]["red"] += 1
                            elif fill_color == "00E8F5E8":
                                color_analysis[status]["green"] += 1
                            else:
                                color_analysis[status]["none"] += 1
                    
                    print("\n🎨 Color Verification Results:")
                    for status, colors in color_analysis.items():
                        total = sum(colors.values())
                        if total > 0:
                            print(f"   {status}: {total} rows")
                            print(f"      Red: {colors['red']}, Green: {colors['green']}, None: {colors['none']}")
                    
                    # Check if coloring is correct
                    excess_correct = color_analysis["超量"]["red"] > 0 and color_analysis["超量"]["green"] == 0
                    shortage_correct = color_analysis["不足"]["green"] > 0 and color_analysis["不足"]["red"] == 0
                    normal_correct = color_analysis["正常"]["none"] > 0 and color_analysis["正常"]["red"] == 0 and color_analysis["正常"]["green"] == 0
                    
                    print(f"\n✅ Coloring Verification:")
                    print(f"   超量 (red): {'✅' if excess_correct else '❌'}")
                    print(f"   不足 (green): {'✅' if shortage_correct else '❌'}")
                    print(f"   正常 (none): {'✅' if normal_correct else '❌'}")
                    
                    if excess_correct and shortage_correct and normal_correct:
                        print(f"\n🎉 All status coloring is correct!")
                        return True
                    else:
                        print(f"\n❌ Some status coloring is incorrect")
                        return False
                        
            wb.close()
        else:
            print("❌ Report generation failed")
            return False
            
    except Exception as e:
        print(f"❌ Error during verification: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_final_coloring_verification()
    if success:
        print("\n🎉 Final coloring verification passed!")
    else:
        print("\n❌ Final coloring verification failed")
        sys.exit(1)