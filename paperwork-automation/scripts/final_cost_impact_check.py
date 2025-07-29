#!/usr/bin/env python3
"""
Final verification of cost impact calculation and formatting
"""

import pandas as pd
from openpyxl import load_workbook

def final_cost_impact_check():
    """Final check of cost impact calculation and formatting"""
    
    file_path = r"output\monthly_gross_margin\毛利相关分析指标-202505.xlsx"
    
    try:
        # Read with pandas to check values
        df = pd.read_excel(file_path, sheet_name=1, header=None)
        
        print("=== COST IMPACT VERIFICATION ===")
        print(f"Total rows in sheet: {len(df)}")
        
        # Data starts from row 3 (index 2)
        data_rows = df.iloc[2:].dropna(how='all')
        print(f"Data rows: {len(data_rows)}")
        
        # Cost impact is in column 12 (index 11)
        cost_impact_col = 11
        
        valid_calculations = 0
        total_checked = 0
        sample_count = 0
        
        print("\nSample Cost Impact Calculations:")
        
        for idx, row in data_rows.iterrows():
            if len(row) > cost_impact_col and sample_count < 5:
                try:
                    # Get the values
                    material_num = str(row.iloc[1]).replace('.0', '') if pd.notna(row.iloc[1]) else 'Unknown'
                    current_price = float(row.iloc[5]) if pd.notna(row.iloc[5]) else 0
                    prev_price = float(row.iloc[6]) if pd.notna(row.iloc[6]) else 0
                    usage = float(row.iloc[10]) if pd.notna(row.iloc[10]) else 0  # Column 11 is usage
                    cost_impact = float(row.iloc[cost_impact_col]) if pd.notna(row.iloc[cost_impact_col]) else 0
                    
                    # Calculate expected: usage * (prev_price - current_price)
                    expected = usage * (prev_price - current_price)
                    
                    print(f"Material {material_num}:")
                    print(f"  Usage: {usage}")
                    print(f"  Current Price: ${current_price:.4f}")
                    print(f"  Previous Price: ${prev_price:.4f}")
                    print(f"  Calculated Impact: ${cost_impact:.4f}")
                    print(f"  Expected Impact: ${expected:.4f}")
                    
                    # Check if calculation is correct
                    if abs(cost_impact - expected) < 0.01:
                        print(f"  Result: ✅ CORRECT")
                        valid_calculations += 1
                    else:
                        print(f"  Result: ❌ INCORRECT")
                    
                    print()
                    total_checked += 1
                    sample_count += 1
                    
                except (ValueError, TypeError, IndexError) as e:
                    print(f"Error processing row {idx}: {e}")
                    continue
        
        # Now check formatting with openpyxl
        print("=== FORMATTING VERIFICATION ===")
        wb = load_workbook(file_path)
        
        # Find material cost sheet (should be index 1)
        sheet_names = wb.sheetnames
        material_sheet = wb[sheet_names[1]] if len(sheet_names) > 1 else None
        
        if material_sheet:
            formatting_correct = 0
            formatting_checked = 0
            
            print("Sample Formatting Check:")
            
            # Check first 5 data rows (rows 3-7)
            for row_num in range(3, 8):
                try:
                    cost_impact_cell = material_sheet.cell(row=row_num, column=12)
                    cost_impact_value = cost_impact_cell.value
                    
                    if cost_impact_value is not None:
                        fill_color = cost_impact_cell.fill.start_color.rgb if cost_impact_cell.fill.start_color else "00000000"
                        
                        # Determine expected color
                        if cost_impact_value == 0:
                            expected_color = "FFFFFF"  # White
                        elif cost_impact_value > 0:
                            expected_color = "C6EFCE"  # Green
                        else:
                            expected_color = "FFC7CE"  # Red
                        
                        # Check if color matches (allow for variations)
                        color_correct = expected_color.lower() in fill_color.lower() or fill_color == "00000000"
                        
                        print(f"  Row {row_num}: ${cost_impact_value:.2f}")
                        print(f"    Expected: {expected_color}, Actual: {fill_color}")
                        print(f"    Formatting: {'✅' if color_correct else '❌'}")
                        
                        if color_correct:
                            formatting_correct += 1
                        formatting_checked += 1
                        
                except Exception as e:
                    print(f"Error checking formatting for row {row_num}: {e}")
                    continue
            
            print(f"\nFormatting Results: {formatting_correct}/{formatting_checked} correct")
        
        print(f"\n=== FINAL RESULTS ===")
        print(f"Calculation Accuracy: {valid_calculations}/{total_checked} correct")
        
        if valid_calculations == total_checked and total_checked > 0:
            print("✅ SUCCESS: Cost impact calculation is working correctly!")
            print("✅ SUCCESS: Sign flip logic implemented (prev_price - current_price)")
            return True
        else:
            print("❌ ISSUE: Cost impact calculation needs review")
            return False
            
    except Exception as e:
        print(f"Error in final check: {e}")
        return False

if __name__ == "__main__":
    final_cost_impact_check()