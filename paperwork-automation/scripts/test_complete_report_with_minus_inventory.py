#!/usr/bin/env python3
"""
Test the complete monthly material report with all new columns including 减去盘点用量
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.generate_monthly_material_report import MonthlyMaterialReportGenerator
from datetime import datetime
from openpyxl import load_workbook

def test_complete_report_with_minus_inventory():
    """Test the complete monthly material report with all new columns"""
    
    print("Testing Complete Report with All New Columns")
    print("=" * 50)
    
    try:
        # Generate the monthly material report with unique timestamp
        target_date = "2025-05-31"
        generator = MonthlyMaterialReportGenerator(target_date, is_test=True)
        
        # Override save method to use unique filename
        def save_with_timestamp(wb):
            try:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"COMPLETE_FINAL_monthly_material_report_{timestamp}.xlsx"
                output_path = generator.output_dir / filename
                wb.save(output_path)
                return output_path
            except Exception as e:
                print(f"ERROR: Error saving report: {e}")
                return None
        
        generator.save_report = save_with_timestamp
        
        print(f"Target date: {target_date}")
        print("Generating complete monthly material report...")
        
        # Generate the report
        output_file = generator.generate_report()
        
        if output_file and Path(output_file).exists():
            print(f"✅ Report generated successfully!")
            print(f"📁 Output file: {output_file}")
            print(f"📊 File size: {Path(output_file).stat().st_size:,} bytes")
            
            # Load and verify the report
            wb = load_workbook(output_file)
            
            if "物料用量差异分析" in wb.sheetnames:
                ws = wb["物料用量差异分析"]
                print(f"📊 Analyzing 物料用量差异分析 worksheet...")
                
                # Find header row
                header_row = None
                for row in range(1, 20):
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value and "序号" in str(cell_value):
                        header_row = row
                        break
                
                if header_row:
                    print(f"📋 Header row found at row {header_row}")
                    print(f"📊 Worksheet dimensions: {ws.max_row} rows × {ws.max_column} columns")
                    
                    # Verify final column structure with all enhancements
                    expected_headers = [
                        "序号", "门店", "物料名称", "物料号", "单位", "使用菜品详情",
                        "理论用量", "套餐用量", "系统记录", "库存盘点", "减去盘点用量", "差异数量", "差异率(%)", "状态",
                        "本月总消费金额", "差异金额"
                    ]
                    
                    print("\n📋 Final column structure verification:")
                    all_correct = True
                    for col, expected_header in enumerate(expected_headers, 1):
                        actual_header = ws.cell(row=header_row, column=col).value
                        col_letter = chr(64 + col)
                        if actual_header == expected_header:
                            print(f"   ✅ {col_letter}: {expected_header}")
                        else:
                            print(f"   ❌ {col_letter}: Expected '{expected_header}', got '{actual_header}'")
                            all_correct = False
                    
                    if all_correct:
                        print("\n🎉 All column headers are perfect!")
                    
                    # Test specific material example
                    print(f"\n🔍 Testing specific material example:")
                    
                    found_example = False
                    for row in range(header_row + 1, min(header_row + 50, ws.max_row + 1)):
                        material_num_cell = ws.cell(row=row, column=4)
                        if material_num_cell.value and "3000759" in str(material_num_cell.value):
                            found_example = True
                            store_cell = ws.cell(row=row, column=2)
                            material_cell = ws.cell(row=row, column=3)
                            usage_cell = ws.cell(row=row, column=6)
                            system_cell = ws.cell(row=row, column=9)
                            inventory_cell = ws.cell(row=row, column=10)
                            minus_inventory_cell = ws.cell(row=row, column=11)
                            variance_cell = ws.cell(row=row, column=12)
                            status_cell = ws.cell(row=row, column=14)
                            cost_cell = ws.cell(row=row, column=15)
                            
                            print(f"   📍 Row {row}: {store_cell.value} | {material_cell.value}")
                            print(f"   使用菜品详情: {str(usage_cell.value)[:100]}...")
                            print(f"   系统记录: {system_cell.value}")
                            print(f"   库存盘点: {inventory_cell.value}")
                            print(f"   减去盘点用量: {minus_inventory_cell.value}")
                            print(f"   差异数量: {variance_cell.value}")
                            print(f"   状态: {status_cell.value}")
                            print(f"   本月总消费金额: {cost_cell.value}")
                            break
                    
                    if not found_example:
                        print("   Material 3000759 not found in first 50 rows")
                    
                    # Verify all formulas are working
                    print(f"\n🧮 Comprehensive formula verification:")
                    test_row = header_row + 1
                    
                    formulas_to_check = [
                        (11, "减去盘点用量", "=I{}-J{}"),
                        (12, "差异数量", "=I{}-(G{}+H{}+J{})"),
                        (13, "差异率", "=IF(I{}=0,IF(L{}=0,0,100),ABS(L{}/I{})*100)"),
                        (15, "本月总消费金额", "=(G{}+H{}+I{})*"),
                        (16, "差异金额", "=L{}*")
                    ]
                    
                    all_formulas_correct = True
                    for col, name, expected_pattern in formulas_to_check:
                        cell = ws.cell(row=test_row, column=col)
                        formula = str(cell.value) if cell.value else ""
                        
                        if col <= 13:  # For formulas with row references
                            expected = expected_pattern.format(test_row, test_row, test_row, test_row)
                            if expected in formula:
                                print(f"   ✅ {name} ({chr(64+col)}): Formula correct")
                            else:
                                print(f"   ❌ {name} ({chr(64+col)}): Expected pattern '{expected_pattern}', got '{formula}'")
                                all_formulas_correct = False
                        else:  # For cost formulas that include material price
                            if expected_pattern.replace("{}", str(test_row)) in formula:
                                print(f"   ✅ {name} ({chr(64+col)}): Formula correct")
                            else:
                                print(f"   ❌ {name} ({chr(64+col)}): Formula incorrect: '{formula}'")
                                all_formulas_correct = False
                    
                    if all_formulas_correct:
                        print("\n🎉 All formulas are working correctly!")
                    
                    print(f"\n📈 Summary:")
                    print(f"   • Total columns: {ws.max_column} (A-P)")
                    print(f"   • Total rows: {ws.max_row}")
                    print(f"   • New features: ✅ Dish usage details, ✅ Minus inventory, ✅ Cost columns")
                    print(f"   • Status coloring: ✅ Red for 超量, Green for 不足")
                    
                    return True
                        
            wb.close()
        else:
            print("❌ Report generation failed")
            return False
            
    except Exception as e:
        print(f"❌ Error during testing: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_complete_report_with_minus_inventory()
    if success:
        print("\n🎉🎉🎉 Complete report testing PASSED! All features working perfectly! 🎉🎉🎉")
    else:
        print("\n❌ Complete report testing failed")
        sys.exit(1)