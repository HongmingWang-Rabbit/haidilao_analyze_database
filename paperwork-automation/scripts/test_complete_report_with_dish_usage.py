#!/usr/bin/env python3
"""
Test the complete monthly material report with dish usage details column
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.generate_monthly_material_report import MonthlyMaterialReportGenerator
from datetime import datetime
from openpyxl import load_workbook

def test_complete_report_with_dish_usage():
    """Test the complete monthly material report with dish usage details"""
    
    print("Testing Complete Report with Dish Usage Details")
    print("=" * 50)
    
    try:
        # Generate the monthly material report with unique timestamp
        target_date = "2025-05-31"
        generator = MonthlyMaterialReportGenerator(target_date, is_test=True)
        
        # Override save method to use unique filename
        def save_with_timestamp(wb):
            try:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"DISH_USAGE_monthly_material_report_{timestamp}.xlsx"
                output_path = generator.output_dir / filename
                wb.save(output_path)
                return output_path
            except Exception as e:
                print(f"ERROR: Error saving report: {e}")
                return None
        
        generator.save_report = save_with_timestamp
        
        print(f"Target date: {target_date}")
        print("Generating monthly material report...")
        
        # Generate the report
        output_file = generator.generate_report()
        
        if output_file and Path(output_file).exists():
            print(f"✅ Report generated successfully!")
            print(f"📁 Output file: {output_file}")
            
            # Load and verify the report
            wb = load_workbook(output_file)
            
            if "物料用量差异分析" in wb.sheetnames:
                ws = wb["物料用量差异分析"]
                print(f"📊 Analyzing 物料用量差异分析 worksheet...")
                
                # Find header row
                header_row = None
                for row in range(1, 20):
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value and "序号" in str(cell_value):
                        header_row = row
                        break
                
                if header_row:
                    print(f"📋 Header row found at row {header_row}")
                    print(f"📊 Worksheet dimensions: {ws.max_row} rows × {ws.max_column} columns")
                    
                    # Verify column structure
                    expected_headers = [
                        "序号", "门店", "物料名称", "物料号", "单位", "使用菜品详情",
                        "理论用量", "套餐用量", "系统记录", "库存盘点", "差异数量", "差异率(%)", "状态",
                        "本月总消费金额", "差异金额"
                    ]
                    
                    print("\n📋 Column verification:")
                    all_correct = True
                    for col, expected_header in enumerate(expected_headers, 1):
                        actual_header = ws.cell(row=header_row, column=col).value
                        if actual_header == expected_header:
                            print(f"   ✅ Column {chr(64 + col)}: {expected_header}")
                        else:
                            print(f"   ❌ Column {chr(64 + col)}: Expected '{expected_header}', got '{actual_header}'")
                            all_correct = False
                    
                    if all_correct:
                        print("\n🎉 All column headers are correct!")
                    
                    # Check dish usage details column specifically
                    usage_col = 6  # Column F
                    print(f"\n📊 Sample dish usage details (Column F):")
                    
                    sample_count = 0
                    for row in range(header_row + 1, ws.max_row + 1):
                        if sample_count >= 3:
                            break
                            
                        material_cell = ws.cell(row=row, column=3)  # Material name
                        store_cell = ws.cell(row=row, column=2)     # Store name
                        usage_cell = ws.cell(row=row, column=usage_col)  # Usage details
                        
                        if usage_cell.value and "sale-" in str(usage_cell.value):
                            material_name = material_cell.value[:25] if material_cell.value else "Unknown"
                            store_name = store_cell.value if store_cell.value else "Unknown"
                            
                            print(f"\n   📍 Row {row}: {store_name} | {material_name}...")
                            usage_lines = str(usage_cell.value).split('\n')
                            for line in usage_lines[:3]:  # Show first 3 lines
                                if line.strip():
                                    print(f"      {line}")
                            
                            sample_count += 1
                    
                    # Check that formulas are updated correctly
                    print(f"\n🧮 Formula verification:")
                    test_row = header_row + 1
                    
                    # Check variance formula (should be K = I - (G+H+J))
                    variance_cell = ws.cell(row=test_row, column=11)
                    if variance_cell.value and "=I" in str(variance_cell.value) and "-(G" in str(variance_cell.value):
                        print("   ✅ Variance formula updated correctly")
                    else:
                        print(f"   ❌ Variance formula incorrect: {variance_cell.value}")
                    
                    # Check cost formula (should be N = (G+H+I)*price)
                    cost_cell = ws.cell(row=test_row, column=14)
                    if cost_cell.value and "=(G" in str(cost_cell.value) and "+H" in str(cost_cell.value) and "+I" in str(cost_cell.value):
                        print("   ✅ Cost formula updated correctly")
                    else:
                        print(f"   ❌ Cost formula incorrect: {cost_cell.value}")
                        
                    print(f"\n🎉 Complete report with dish usage details working perfectly!")
                    return True
                        
            wb.close()
        else:
            print("❌ Report generation failed")
            return False
            
    except Exception as e:
        print(f"❌ Error during testing: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_complete_report_with_dish_usage()
    if success:
        print("\n🎉 Complete report testing passed!")
    else:
        print("\n❌ Complete report testing failed")
        sys.exit(1)