#!/usr/bin/env python3
"""
Generate monthly gross margin report - MINIMAL VERSION
"""

import os
import sys
from pathlib import Path

# Add parent directory to path for imports FIRST
sys.path.insert(0, str(Path(__file__).parent.parent))

print("Starting minimal gross margin report generator...")

from utils.database import DatabaseConfig, DatabaseManager
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import argparse

def main():
    """Main function"""
    parser = argparse.ArgumentParser(description="Generate monthly gross margin report")
    parser.add_argument("--target-date", default="2025-07-31", help="Target date (YYYY-MM-DD)")
    args = parser.parse_args()
    
    print(f"Generating gross margin report for date: {args.target_date}")
    
    try:
        config = DatabaseConfig(is_test=False)
        db_manager = DatabaseManager(config)
        
        target_dt = datetime.strptime(args.target_date, '%Y-%m-%d')
        year, month = target_dt.year, target_dt.month
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Gross Margin Report"
        
        ws['A1'] = f"Monthly Gross Margin Report - {year}-{month:02d}"
        ws['A1'].font = Font(bold=True, size=14)
        
        output_dir = Path('./output')
        output_dir.mkdir(exist_ok=True)
        filename = f"monthly_gross_margin_report_minimal_{args.target_date.replace('-', '_')}.xlsx"
        output_path = output_dir / filename
        
        wb.save(output_path)
        print(f"SUCCESS: Gross margin report saved to {output_path}")
        sys.exit(0)  # Success
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()