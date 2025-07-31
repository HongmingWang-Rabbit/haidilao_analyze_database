#!/usr/bin/env python3
"""
Bank Transaction Processing Script
Processes all bank files and appends new transactions to existing sheets
"""

from configs.bank_desc import BankDescriptionConfig
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl import load_workbook, Workbook
import pandas as pd
import numpy as np
from pathlib import Path
import logging
import argparse
from datetime import datetime
from typing import Dict, List, Optional
import sys
import zipfile
import tempfile
import shutil
import os

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))


# Set up logging
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelImagePreserver:
    """Utility to preserve images in Excel files when using openpyxl"""
    
    def __init__(self, source_file):
        self.source_file = Path(source_file)
        self.media_files = {}
        self.drawing_files = {}
        
    def extract_media_and_drawings(self):
        """Extract media and drawing files from the source Excel file"""
        try:
            with zipfile.ZipFile(self.source_file, 'r') as zip_file:
                file_list = zip_file.namelist()
                
                # Extract various image-related files
                for file_name in file_list:
                    file_lower = file_name.lower()
                    
                    # Media files (images)
                    if 'media/' in file_lower:
                        self.media_files[file_name] = zip_file.read(file_name)
                    
                    # Drawing files and their relationships
                    elif 'drawings/' in file_lower:
                        self.drawing_files[file_name] = zip_file.read(file_name)
                    
                    # Chart files
                    elif 'charts/' in file_lower:
                        self.drawing_files[file_name] = zip_file.read(file_name)
                    
                    # Embedded objects
                    elif 'embeddings/' in file_lower:
                        self.media_files[file_name] = zip_file.read(file_name)
                    
                    # Any other image files in unusual locations
                    elif any(ext in file_lower for ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.emf', '.wmf']):
                        self.media_files[file_name] = zip_file.read(file_name)
                        
                logger.info(f"Extracted {len(self.media_files)} media files and {len(self.drawing_files)} drawing files")
                if self.media_files:
                    logger.info(f"Media files: {list(self.media_files.keys())}")
                if self.drawing_files:
                    logger.info(f"Drawing files: {list(self.drawing_files.keys())}")
                
        except Exception as e:
            logger.error(f"Error extracting media/drawings: {e}")
            
    def inject_media_and_drawings(self, target_file):
        """Inject the preserved media and drawing files into the target Excel file"""
        if not self.media_files and not self.drawing_files:
            logger.info("No media or drawing files to inject")
            return True  # Nothing to inject, but not an error
            
        try:
            # Create a temporary file for the modified Excel
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                temp_path = temp_file.name
            
            # Copy the target file to temp
            shutil.copy2(target_file, temp_path)
            logger.info(f"Created temporary file: {temp_path}")
            
            # Open the Excel file as a ZIP and add media/drawings
            with zipfile.ZipFile(temp_path, 'a', compression=zipfile.ZIP_DEFLATED) as zip_file:
                # Check existing files to avoid duplicates
                existing_files = set(zip_file.namelist())
                
                # Add media files
                injected_media = 0
                for file_name, file_data in self.media_files.items():
                    if file_name not in existing_files:
                        zip_file.writestr(file_name, file_data)
                        injected_media += 1
                        logger.debug(f"Injected media file: {file_name}")
                    else:
                        logger.debug(f"Media file already exists, skipping: {file_name}")
                    
                # Add drawing files  
                injected_drawings = 0
                for file_name, file_data in self.drawing_files.items():
                    if file_name not in existing_files:
                        zip_file.writestr(file_name, file_data)
                        injected_drawings += 1
                        logger.debug(f"Injected drawing file: {file_name}")
                    else:
                        logger.debug(f"Drawing file already exists, skipping: {file_name}")
            
            # Replace the original target file
            shutil.move(temp_path, target_file)
            
            logger.info(f"Successfully injected {injected_media} media files and {injected_drawings} drawing files")
            return True
            
        except Exception as e:
            logger.error(f"Error injecting media/drawings: {e}")
            import traceback
            logger.error(f"Full traceback: {traceback.format_exc()}")
            # Clean up temp file if it exists
            if 'temp_path' in locals() and os.path.exists(temp_path):
                try:
                    os.unlink(temp_path)
                except:
                    pass
            return False


class BankTransactionProcessor:
    """Process bank transactions from multiple sources and append to existing worksheets"""

    def __init__(self, target_year: int, target_month: int):
        self.target_year = target_year
        self.target_month = target_month
        self.input_dir = Path("Input/daily_report/bank_transactions_reports")
        self.template_file = self.input_dir / "CA全部7家店明细.xlsx"

        # Always save to output folder with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        self.output_file = Path("output") / \
            f"Bank_Transactions_Report_{timestamp}.xlsx"

        # Account mapping: input account identifier -> output sheet name
        self.account_mapping = {
            # BMO accounts (from ReconciliationReport)
            "3817": "CA1D-3817",
            "6027": "CA2D-6027",
            "1680": "CA3D-1680 ",
            "1699": "CA4D-1699",
            "6333": "CA5D-6333",
            "6317": "CA6D-6317",
            "0798": "BMO美金-0798",

            # CIBC account (from TransactionDetail.csv)
            "0401": "CA7D-CIBC 0401",

            # RBC accounts (from individual files)
            "5401": "RBC 5401",
            "5419": "RBC 5419",
            "0517": "RBC0517-Hi Bowl",
            "0922": "RBC 0922（USD）",
            "3088": "RBC3088（USD）-Hi Bowl"
        }

    def process_all_transactions(self) -> None:
        """Main processing function"""
        logger.info(
            f"Processing bank transactions for {self.target_year}-{self.target_month:02d}")

        # Get last existing dates from template to avoid duplicates
        last_existing_dates = self.get_last_existing_dates_from_template()

        # Collect all transactions by sheet name
        all_transactions = {}

        # Process BMO reconciliation report (multiple accounts)
        bmo_transactions = self.process_bmo_reconciliation(last_existing_dates)
        for sheet_name, transactions in bmo_transactions.items():
            if transactions:
                all_transactions[sheet_name] = transactions
                logger.info(
                    f"Extracted {len(transactions)} transactions for {sheet_name}")

        # Process CIBC transaction detail
        cibc_transactions = self.process_cibc_file(last_existing_dates)
        if cibc_transactions:
            all_transactions["CA7D-CIBC 0401"] = cibc_transactions
            logger.info(
                f"Extracted {len(cibc_transactions)} transactions for CA7D-CIBC 0401")

        # Process RBC files (individual accounts)
        rbc_transactions = self.process_rbc_files(last_existing_dates)
        for sheet_name, transactions in rbc_transactions.items():
            if transactions:
                all_transactions[sheet_name] = transactions
                logger.info(
                    f"Extracted {len(transactions)} transactions for {sheet_name}")

        if not all_transactions:
            logger.warning("No transactions found for processing")
            return

        # Update existing workbook
        self.append_to_existing_workbook(all_transactions)

    def process_bmo_reconciliation(self, last_existing_dates: Dict[str, str]) -> Dict[str, List[Dict]]:
        """Process BMO reconciliation report containing multiple accounts"""
        file_path = None
        for f in self.input_dir.glob("ReconciliationReport*.xls"):
            file_path = f
            break

        if not file_path or not file_path.exists():
            logger.warning("BMO ReconciliationReport not found")
            return {}

        logger.info(f"Processing BMO reconciliation: {file_path.name}")

        try:
            # Read the Excel file
            df = pd.read_excel(file_path, engine='xlrd')

            # The file structure has account headers and transaction rows
            transactions_by_account = {}
            current_account = None

            for idx, row in df.iterrows():
                # Look for account header lines
                first_col = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""

                # Check if this is an account header
                for account_num in ["3817", "6027", "1680", "1699", "6333", "6317", "0798"]:
                    if account_num in first_col:
                        current_account = account_num
                        sheet_name = self.account_mapping[account_num]
                        if sheet_name not in transactions_by_account:
                            transactions_by_account[sheet_name] = []
                        break

                # If we have a current account and this looks like transaction data
                if current_account and self.is_transaction_row(row):
                    transaction = self.parse_bmo_transaction_row(
                        row, current_account, last_existing_dates)
                    if transaction:  # Date filtering now done in parse method
                        sheet_name = self.account_mapping[current_account]
                        transactions_by_account[sheet_name].append(transaction)

            return transactions_by_account

        except Exception as e:
            logger.error(f"Error processing BMO reconciliation: {e}")
            return {}

    def is_transaction_row(self, row) -> bool:
        """Check if a row contains transaction data"""
        # Look for date pattern in first column and amount in other columns
        first_col = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""

        # Check if first column looks like a date
        if any(month in first_col for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                                                "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]):
            return True
        return False

    def parse_bmo_transaction_row(self, row, account_num: str, last_existing_dates: Dict[str, str]) -> Optional[Dict]:
        """Parse a BMO transaction row"""
        try:
            # Map columns based on BMO format
            date_str = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""

            # Parse date
            try:
                date_obj = pd.to_datetime(date_str)
                # Get sheet name for this account
                sheet_name = self.account_mapping[account_num]
                
                # Check if we should process this transaction (target month and after last existing date)
                if not self.should_process_transaction(date_obj.strftime('%Y-%m-%d'), sheet_name, last_existing_dates):
                    return None
            except:
                return None

            # BMO Column Structure:
            # Column 0: Date
            # Column 1: Flag (* or NaN)
            # Column 2: Transaction Description
            # Column 3: Customer Reference
            # Column 4: Bank Reference
            # Column 5: Debit
            # Column 6: Credit
            # Column 7: Balance
            # Column 8: Details

            description = str(row.iloc[2]) if len(
                row) > 2 and pd.notna(row.iloc[2]) else ""
            customer_ref = str(row.iloc[3]) if len(
                row) > 3 and pd.notna(row.iloc[3]) else ""
            bank_ref = str(row.iloc[4]) if len(
                row) > 4 and pd.notna(row.iloc[4]) else ""
            details = str(row.iloc[8]) if len(
                row) > 8 and pd.notna(row.iloc[8]) else ""

            # Extract debit/credit from specific columns
            debit = ""
            credit = ""

            if len(row) > 5 and pd.notna(row.iloc[5]):
                debit_val = row.iloc[5]
                if isinstance(debit_val, (int, float)) and debit_val != 0:
                    debit = abs(debit_val)

            if len(row) > 6 and pd.notna(row.iloc[6]):
                credit_val = row.iloc[6]
                if isinstance(credit_val, (int, float)) and credit_val != 0:
                    credit = abs(credit_val)

            # Get classification using concatenated description and details
            # Some transactions have info in Transaction Description, others in Details
            combined_text = f"{description} {details}".strip()
            
            # Calculate transaction amount and determine transaction type
            transaction_amount = None
            transaction_type = None
            if credit and isinstance(credit, (int, float)) and credit > 0:
                transaction_amount = float(credit)
                transaction_type = 'credit'
            elif debit and isinstance(debit, (int, float)) and debit > 0:
                transaction_amount = float(debit)  # Store as positive amount
                transaction_type = 'debit'
                
            classification = BankDescriptionConfig.get_transaction_info(
                combined_text, transaction_amount, transaction_type)

            return {
                'Date': date_obj.strftime('%b %d, %Y'),
                'Transaction Description': description,
                'Customer Reference': customer_ref,
                'Bank Reference': bank_ref,
                'Debit': debit,
                'Credit': credit,
                'Details': details,
                '品名': classification['品名'],
                '付款详情': classification['付款详情'],
                '单据号': "",
                '附件': "",
                '是否登记线下付款表': "",
                '是否登记支票使用表': "",
                '_account': account_num
            }

        except Exception as e:
            logger.error(f"Error parsing BMO transaction row: {e}")
            return None

    def process_cibc_file(self, last_existing_dates: Dict[str, str]) -> List[Dict]:
        """Process CIBC TransactionDetail.xlsx with special format"""
        file_path = self.input_dir / "TransactionDetail.xlsx"
        if not file_path.exists():
            logger.warning("CIBC TransactionDetail.xlsx not found")
            return []

        logger.info(f"Processing CIBC: {file_path.name}")

        try:
            df = pd.read_excel(file_path, engine='openpyxl', header=None)
            transactions = []
            current_section = None  # 'debit' or 'credit'
            
            for idx, row in df.iterrows():
                # Convert row to list for easier processing
                row_data = row.tolist()
                first_col = str(row_data[0]) if pd.notna(row_data[0]) else ""
                
                # Check for section headers
                if 'Debit transactions' in first_col:
                    current_section = 'debit'
                    continue
                elif 'Credit transactions' in first_col:
                    current_section = 'credit'
                    continue
                elif 'Total debits' in first_col or 'Total credits' in first_col:
                    current_section = None
                    continue
                
                # Skip header rows and empty rows
                if current_section is None or 'Description' in first_col or first_col == '' or pd.isna(row_data[0]):
                    continue
                
                # Process transaction rows (should have description in column 0, date in column 6, amount in column 7)
                if len(row_data) >= 8 and pd.notna(row_data[6]) and pd.notna(row_data[7]):
                    try:
                        # Extract transaction data
                        description = str(row_data[0]).strip()
                        date_value = pd.to_datetime(row_data[6])
                        amount = float(row_data[7])
                        bank_ref = str(row_data[10]) if len(row_data) > 10 and pd.notna(row_data[10]) else ""
                        client_ref = str(row_data[11]) if len(row_data) > 11 and pd.notna(row_data[11]) else ""
                        
                        date_str = date_value.strftime('%Y-%m-%d')
                        
                        # Check if we should process this transaction
                        if not self.should_process_transaction(date_str, "CA7D-CIBC 0401", last_existing_dates):
                            continue
                        
                        # Determine debit/credit based on section
                        if current_section == 'debit':
                            debit = abs(amount)
                            credit = ''
                        else:  # credit section
                            debit = ''
                            credit = abs(amount)
                        
                        # Get classification with amount and transaction type information
                        transaction_amount = abs(amount)  # Store as positive amount
                        transaction_type = current_section  # 'credit' or 'debit'
                            
                        classification = BankDescriptionConfig.get_transaction_info(description, transaction_amount, transaction_type)
                        
                        transaction = {
                            'Date': date_value.strftime('%b %d, %Y'),
                            'Transaction Description': description,
                            'Customer Reference': client_ref,
                            'Bank Reference': bank_ref,
                            'Debit': debit,
                            'Credit': credit,
                            'Details': description,
                            '品名': classification['品名'],
                            '付款详情': classification['付款详情'],
                            '单据号': "",
                            '附件': "",
                            '是否登记线下付款表': "",
                            '是否登记支票使用表': "",
                            '_account': '0401'
                        }
                        transactions.append(transaction)
                        
                    except Exception as e:
                        logger.warning(f"Error parsing CIBC transaction row {idx}: {e}")
                        continue

            logger.info(f"Successfully processed {len(transactions)} CIBC transactions")
            return transactions

        except Exception as e:
            logger.error(f"Error processing CIBC file: {e}")
            return []

    def process_rbc_files(self, last_existing_dates: Dict[str, str]) -> Dict[str, List[Dict]]:
        """Process individual RBC files"""
        transactions_by_account = {}

        rbc_files = {
            "5401": "RBC Business Bank Account (5401)*.xlsx",
            "5419": "RBC Business Bank Account (5419)*.xlsx",
            "0517": "RBC Business Bank Account (0517)*.xlsx",
            "0922": "RBC Business Bank Account (0922)*.xlsx",
            "3088": "RBC Business Bank Account (3088)*.xlsx"
        }

        for account_num, pattern in rbc_files.items():
            files = list(self.input_dir.glob(pattern))
            if not files:
                logger.warning(f"RBC file for account {account_num} not found")
                continue

            file_path = files[0]  # Take first match
            logger.info(f"Processing RBC {account_num}: {file_path.name}")

            try:
                df = pd.read_excel(file_path, engine='openpyxl')
                transactions = []

                for _, row in df.iterrows():
                    date_value = pd.to_datetime(row['Date'])
                    date_str = date_value.strftime('%Y-%m-%d')
                    
                    # Check if we should process this transaction (target month and after last existing date)
                    sheet_name = self.account_mapping[account_num]
                    if not self.should_process_transaction(date_str, sheet_name, last_existing_dates):
                        continue

                    # Combine description fields with dashes
                    desc_fields = ['Description 1', 'Description 2',
                                   'Description 3', 'Description 4', 'Description 5']
                    description = " - ".join([str(row.get(
                        field, '')) for field in desc_fields if pd.notna(row.get(field, '')) and str(row.get(field, '')).strip()]).strip()

                    withdrawals = row.get('Withdrawals', 0)
                    deposits = row.get('Deposits', 0)

                    debit = withdrawals if pd.notna(
                        withdrawals) and withdrawals > 0 else ''
                    credit = deposits if pd.notna(
                        deposits) and deposits > 0 else ''

                    # Calculate transaction amount for classification
                    transaction_amount = None
                    transaction_type = None
                    if credit and isinstance(credit, (int, float)) and credit > 0:
                        transaction_amount = float(credit)
                        transaction_type = 'credit'
                    elif debit and isinstance(debit, (int, float)) and debit > 0:
                        transaction_amount = float(debit)  # Store as positive amount
                        transaction_type = 'debit'

                    classification = BankDescriptionConfig.get_transaction_info(
                        description, transaction_amount, transaction_type)
                    transaction = {
                        'Date': date_value.strftime('%b %d, %Y'),
                        'Transaction Description': description,
                        'Customer Reference': "",
                        'Bank Reference': "",
                        'Debit': debit,
                        'Credit': credit,
                        'Details': description,  # Use description as details for RBC
                        '品名': classification['品名'],
                        '付款详情': classification['付款详情'],
                        '单据号': "",
                        '附件': "",
                        '是否登记线下付款表': "",
                        '是否登记支票使用表': "",
                        '_account': account_num
                    }
                    transactions.append(transaction)

                if transactions:
                    sheet_name = self.account_mapping[account_num]
                    transactions_by_account[sheet_name] = transactions

            except Exception as e:
                logger.error(f"Error processing RBC file {account_num}: {e}")

        return transactions_by_account

    def is_target_month(self, date_str: str) -> bool:
        """Check if date is in target month/year"""
        try:
            date_obj = pd.to_datetime(date_str)
            return date_obj.year == self.target_year and date_obj.month == self.target_month
        except:
            return False

    def should_process_transaction(self, date_str: str, sheet_name: str, last_existing_dates: Dict[str, str]) -> bool:
        """Check if transaction should be processed (in target month and after last existing date)"""
        # First check if it's in target month
        if not self.is_target_month(date_str):
            return False
        
        # If no last existing date for this sheet, process all target month transactions
        if sheet_name not in last_existing_dates:
            return True
        
        # Only process transactions after the last existing date
        try:
            transaction_date = pd.to_datetime(date_str)
            last_existing_date = pd.to_datetime(last_existing_dates[sheet_name])
            return transaction_date > last_existing_date
        except:
            # If date parsing fails, fall back to target month check
            return True

    def append_to_existing_workbook(self, all_transactions: Dict[str, List[Dict]]) -> None:
        """Append transactions to existing workbook sheets with image preservation"""
        try:
            if not self.template_file.exists():
                logger.error(
                    f"Template file does not exist: {self.template_file}")
                return
            
            # Create output directory first
            self.output_file.parent.mkdir(exist_ok=True)
            
            # Step 1: Extract images and drawings from template for preservation
            logger.info(f"Extracting images and drawings from template: {self.template_file}")
            image_preserver = ExcelImagePreserver(self.template_file)
            image_preserver.extract_media_and_drawings()
            
            # Step 2: Copy the template file to output location
            logger.info(f"Copying template file: {self.template_file}")
            shutil.copy2(self.template_file, self.output_file)
            
            # Step 3: Load the copied file and modify it
            logger.info(f"Loading copied workbook: {self.output_file}")
            wb = load_workbook(self.output_file)

            total_added = 0
            total_skipped = 0

            for sheet_name, transactions in all_transactions.items():
                if sheet_name not in wb.sheetnames:
                    logger.warning(
                        f"Sheet '{sheet_name}' not found in workbook")
                    continue

                ws = wb[sheet_name]
                added, skipped = self.append_to_worksheet(ws, transactions)
                total_added += added
                total_skipped += skipped
                logger.info(
                    f"Sheet '{sheet_name}': Added {added} new transactions (skipped {skipped} duplicates)")

            # Step 4: Save the modified file (this will lose images temporarily)
            try:
                wb.save(self.output_file)
                wb.close()
                logger.info(f"Saved modified workbook: {self.output_file}")
                
                # Give the file system a moment to finalize the file
                import time
                time.sleep(0.5)
                
                # Step 5: Re-inject the preserved images and drawings
                logger.info(f"Re-injecting preserved images and drawings to: {self.output_file}")
                success = image_preserver.inject_media_and_drawings(self.output_file)
                if success:
                    logger.info(f"Successfully preserved images in output file")
                else:
                    logger.warning(f"Failed to preserve some images in output file")
                    # Continue anyway - the file is still usable
                
                print(f"SUCCESS: Bank report saved to output folder: {self.output_file}")
                print(f"SUMMARY: Added {total_added} new transactions, skipped {total_skipped} duplicates")
                
            except Exception as e:
                logger.error(f"Failed to save output file: {e}")
                print(f"ERROR: Could not save output file: {e}")
                # Clean up failed file
                if self.output_file.exists():
                    self.output_file.unlink()

        except Exception as e:
            logger.error(f"Error updating workbook: {e}")
            raise

    def get_header_positions_for_sheet(self, sheet_name: str) -> Dict[str, int]:
        """Get bank-specific header positions based on sheet name"""
        if 'CIBC' in sheet_name:
            # CIBC format: 11 columns
            return {
                'Date': 1,
                'Transaction Description': 2,  # "Transaction details"
                'Customer Reference': 2,       # Combined in Transaction details
                'Bank Reference': 2,           # Combined in Transaction details
                'Debit': 3,
                'Credit': 4,
                'Details': 2,                  # Use Transaction details
                '品名': 6,
                '付款详情': 7,
                '单据号': 8,
                '附件': 9,
                '是否登记线下付款表': 10,
                '是否登记支票使用表': 11
            }
        elif sheet_name.startswith('RBC'):
            # RBC format: 12 columns
            return {
                'Date': 2,                     # "Effective Date"
                'Transaction Description': 1,  # "Description"
                'Customer Reference': 3,       # "Serial Number"
                'Bank Reference': 3,           # "Serial Number"
                'Debit': 4,                    # "Debits"
                'Credit': 5,                   # "Credits"
                'Details': 1,                  # Use Description
                '品名': 7,
                '付款详情': 8,
                '单据号': 9,
                '附件': 10,
                '是否登记线下付款表': 11,
                '是否登记支票使用表': 12
            }
        else:
            # BMO format: 14 columns (default)
            return {
                'Date': 1,
                'Transaction Description': 3,
                'Customer Reference': 4,
                'Bank Reference': 5,
                'Debit': 6,
                'Credit': 7,
                'Details': 8,
                '品名': 9,
                '付款详情': 10,
                '单据号': 11,
                '附件': 12,
                '是否登记线下付款表': 13,
                '是否登记支票使用表': 14
            }

    def append_to_worksheet(self, ws, transactions: List[Dict]) -> tuple:
        """Append transactions to existing worksheet"""
        # Find the last row with data (starting from row 3)
        last_row = 2  # Headers are in row 2
        for row in range(3, ws.max_row + 1):
            if any(ws.cell(row=row, column=col).value for col in range(1, 12)):
                last_row = row
            else:
                break

        # Get bank-specific header positions
        header_positions = self.get_header_positions_for_sheet(ws.title)

        added_count = 0
        skipped_count = 0

        for transaction in transactions:
            # Check for duplicates against current worksheet state (including newly added rows)
            current_last_row = last_row + added_count
            is_duplicate = self.is_duplicate_transaction(
                ws, transaction, current_last_row)
            if is_duplicate:
                skipped_count += 1
                continue

            # Add new transaction
            new_row = last_row + 1 + added_count

            for field, col_idx in header_positions.items():
                if field in transaction:
                    cell = ws.cell(row=new_row, column=col_idx,
                                   value=transaction[field])

                    # Apply green background color for coloring fields based on configuration
                    if field in ['单据号', '附件', '是否登记线下付款表', '是否登记支票使用表']:
                        # Get classification for this transaction's details
                        details = transaction.get('Details', '')
                        
                        # Calculate transaction amount for classification
                        transaction_amount = None
                        transaction_type = None
                        if transaction.get('Credit') and isinstance(transaction.get('Credit'), (int, float)):
                            transaction_amount = float(transaction.get('Credit'))
                            transaction_type = 'credit'
                        elif transaction.get('Debit') and isinstance(transaction.get('Debit'), (int, float)):
                            transaction_amount = float(transaction.get('Debit'))  # Store as positive amount
                            transaction_type = 'debit'
                            
                        classification = BankDescriptionConfig.get_transaction_info(
                            details, transaction_amount, transaction_type)
                        if classification.get(field, False):
                            cell.fill = PatternFill(
                                start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green

                    # Apply red background color for uncategorized transactions (需要手动分类)
                    elif field == '品名' and transaction.get(field) == '未分类交易':
                        cell.fill = PatternFill(
                            start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')  # Light red

            added_count += 1

        return added_count, skipped_count

    def is_duplicate_transaction(self, ws, transaction: Dict, last_row: int) -> bool:
        """Check if transaction already exists in worksheet"""
        # Get bank-specific header positions for this sheet
        header_positions = self.get_header_positions_for_sheet(ws.title)
        
        # Extract column positions for key fields
        date_col = header_positions.get('Date', 1)
        desc_col = header_positions.get('Transaction Description', 3)
        debit_col = header_positions.get('Debit', 6)
        credit_col = header_positions.get('Credit', 7)
        
        # Check all existing rows for duplicates (more thorough but slower)
        # Start from row 3 to skip headers
        start_row = 3

        for row in range(start_row, last_row + 1):
            existing_date = ws.cell(row=row, column=date_col).value
            existing_desc = ws.cell(row=row, column=desc_col).value
            existing_debit = ws.cell(row=row, column=debit_col).value
            existing_credit = ws.cell(row=row, column=credit_col).value

            # Normalize dates for comparison (handle different date formats)
            existing_date_normalized = self.normalize_date(existing_date)
            transaction_date_normalized = self.normalize_date(transaction.get('Date'))

            # Convert other fields to strings for comparison, handling None values properly
            existing_desc_str = str(existing_desc) if existing_desc else ""
            existing_debit_str = str(existing_debit) if existing_debit and existing_debit != "" else ""
            existing_credit_str = str(existing_credit) if existing_credit and existing_credit != "" else ""

            transaction_desc_str = str(transaction.get('Transaction Description', ''))
            transaction_debit_str = str(transaction.get('Debit', '')) if transaction.get('Debit') and transaction.get('Debit') != "" else ""
            transaction_credit_str = str(transaction.get('Credit', '')) if transaction.get('Credit') and transaction.get('Credit') != "" else ""

            # Check if key fields match (using normalized dates)
            if (existing_date_normalized == transaction_date_normalized and
                existing_desc_str == transaction_desc_str and
                existing_debit_str == transaction_debit_str and
                    existing_credit_str == transaction_credit_str):
                return True

        return False

    def normalize_date(self, date_value) -> str:
        """Normalize date to a standard format for comparison"""
        if not date_value:
            return ""
        
        try:
            # Convert to pandas datetime (handles multiple formats)
            date_obj = pd.to_datetime(date_value)
            # Return in a standard format for comparison
            return date_obj.strftime('%Y-%m-%d')
        except:
            # If parsing fails, return original string
            return str(date_value)

    def get_last_existing_date(self, ws) -> str:
        """Find the last existing date in the worksheet to avoid processing duplicates"""
        # Get bank-specific header positions to find the correct date column
        header_positions = self.get_header_positions_for_sheet(ws.title)
        date_col = header_positions.get('Date', 1)
        
        last_date = None
        
        # Check all data rows for the most recent date (don't limit to last 50 rows)
        for row in range(3, ws.max_row + 1):
            date_val = ws.cell(row=row, column=date_col).value
            if date_val and str(date_val) not in ['Date', '']:
                try:
                    parsed_date = pd.to_datetime(date_val)
                    if last_date is None or parsed_date > last_date:
                        last_date = parsed_date
                except:
                    continue
        
        # Return as string in YYYY-MM-DD format, or empty if no date found
        return last_date.strftime('%Y-%m-%d') if last_date else ""

    def get_last_existing_dates_from_template(self) -> Dict[str, str]:
        """Get the last existing date from each sheet in the template"""
        last_dates = {}
        
        try:
            if self.template_file.exists():
                wb = load_workbook(self.template_file)
                
                # Check all sheets that might contain bank data
                for sheet_name in wb.sheetnames:
                    if any(bank in sheet_name for bank in ['CA', 'RBC', 'BMO', 'CIBC']):
                        ws = wb[sheet_name]
                        last_date = self.get_last_existing_date(ws)
                        if last_date:
                            last_dates[sheet_name] = last_date
                            logger.info(f"Sheet '{sheet_name}': Last existing date is {last_date}")
                        else:
                            logger.info(f"Sheet '{sheet_name}': No existing dates found")
                
                wb.close()
        except Exception as e:
            logger.error(f"Error reading template file: {e}")
        
        return last_dates


def main():
    parser = argparse.ArgumentParser(description='Process bank transactions')
    parser.add_argument('--target-date', type=str, required=True,
                        help='Target date in YYYY-MM-DD format')
    args = parser.parse_args()

    try:
        target_date = datetime.strptime(args.target_date, '%Y-%m-%d')
        processor = BankTransactionProcessor(
            target_date.year, target_date.month)
        processor.process_all_transactions()
        logger.info("Bank transaction processing completed successfully")
        logger.info(f"Output saved to: {processor.output_file}")
        print("Bank transaction processing completed successfully!")

    except Exception as e:
        logger.error(f"Error processing bank transactions: {e}")
        print("ERROR: Bank transaction processing failed!")
        sys.exit(1)


if __name__ == "__main__":
    main()
