#!/usr/bin/env python3
"""
Test the status coloring changes - 不足 should be green, 超量 should be red
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.monthly_dishes_worksheet import MonthlyDishesWorksheetGenerator
from lib.database_queries import ReportDataProvider
from utils.database import DatabaseManager, DatabaseConfig
from openpyxl import Workbook
from datetime import datetime

def test_status_coloring():
    """Test the status coloring changes"""
    
    print("Testing Status Coloring Changes")
    print("=" * 35)
    
    try:
        # Initialize components
        config = DatabaseConfig(is_test=True)
        db_manager = DatabaseManager(config)
        data_provider = ReportDataProvider(db_manager)
        
        # Create worksheet generator
        target_date = "2025-05-31"
        store_names = ["加拿大一店", "加拿大二店", "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大七店"]
        worksheet_gen = MonthlyDishesWorksheetGenerator(store_names, target_date)
        
        # Create workbook and generate worksheet
        wb = Workbook()
        ws = worksheet_gen.generate_material_variance_worksheet(wb, data_provider)
        
        print(f"✅ Worksheet generated: {ws.title}")
        print(f"📊 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Find the header row
        header_row = None
        for row in range(1, 20):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "序号" in str(cell_value):
                header_row = row
                break
        
        if header_row:
            print(f"📋 Header row found at row {header_row}")
            
            # Check a few data rows for status and color
            color_examples = {"红色": [], "绿色": [], "正常": []}
            
            for row in range(header_row + 1, min(header_row + 21, ws.max_row + 1)):
                status_cell = ws.cell(row=row, column=12)  # Status column
                variance_cell = ws.cell(row=row, column=10)  # Variance amount column
                
                status = status_cell.value if status_cell.value else ""
                fill_color = status_cell.fill.start_color.rgb if status_cell.fill and status_cell.fill.start_color else "None"
                
                if status == "超量":
                    color_examples["红色"].append((row, status, fill_color))
                elif status == "不足":
                    color_examples["绿色"].append((row, status, fill_color))
                elif status == "正常":
                    color_examples["正常"].append((row, status, fill_color))
            
            print("\n🎨 Color Analysis:")
            print(f"   红色 (超量): {len(color_examples['红色'])} rows")
            for row, status, color in color_examples["红色"][:3]:
                print(f"      Row {row}: {status} - Color: {color}")
                
            print(f"   绿色 (不足): {len(color_examples['绿色'])} rows")
            for row, status, color in color_examples["绿色"][:3]:
                print(f"      Row {row}: {status} - Color: {color}")
                
            print(f"   正常: {len(color_examples['正常'])} rows")
            for row, status, color in color_examples["正常"][:3]:
                print(f"      Row {row}: {status} - Color: {color}")
        
        # Save output
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"output/test_status_coloring_{timestamp}.xlsx"
        wb.save(output_file)
        print(f"\n📁 Test output saved to: {output_file}")
        
        # Check specific color codes
        print("\n🎨 Expected Color Codes:")
        print("   FFE6E6 = Light Red (for 超量)")
        print("   E8F5E8 = Light Green (for 不足)")
        print("   No fill = White (for 正常)")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_status_coloring()