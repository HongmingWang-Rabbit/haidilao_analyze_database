#!/usr/bin/env python3
"""
Generate year-over-year comparison report (同比数据) Excel file from database data.
This script compares current year data with the same period from the previous year.
"""

import os
import sys
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent))
from utils.database import DatabaseConfig, DatabaseManager

# Load environment variables
load_dotenv()

class YearlyComparisonReportGenerator:
    """Generate year-over-year comparison report from database data"""
    
    def __init__(self, target_date: str, is_test: bool = False):
        self.target_date = target_date
        self.is_test = is_test
        self.config = DatabaseConfig(is_test=is_test)
        self.db_manager = DatabaseManager(self.config)
        self.output_dir = Path(os.getenv('OUTPUT_DIR', './output'))
        self.output_dir.mkdir(exist_ok=True)
        
        # Store mapping
        self.store_names = {
            1: "加拿大一店", 2: "加拿大二店", 3: "加拿大三店", 4: "加拿大四店",
            5: "加拿大五店", 6: "加拿大六店", 7: "加拿大七店"
        }
        
        # Calculate previous year date
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        self.previous_year_date = target_dt.replace(year=target_dt.year - 1).strftime('%Y-%m-%d')
        self.current_year = target_dt.year
        self.previous_year = target_dt.year - 1
        self.month = target_dt.month
        self.day = target_dt.day
    
    def get_monthly_data_up_to_date(self, year: int, month: int, day: int):
        """Get monthly data up to a specific date"""
        sql = """
        SELECT 
            s.id as store_id,
            s.name as store_name,
            SUM(dr.tables_served_validated) as total_tables,
            SUM(dr.revenue_tax_included) as total_revenue,
            AVG(dr.turnover_rate) as avg_turnover_rate,
            AVG(dr.revenue_tax_included / NULLIF(dr.tables_served_validated, 0)) as avg_per_table
        FROM daily_report dr
        JOIN store s ON dr.store_id = s.id
        WHERE EXTRACT(YEAR FROM dr.date) = %s 
        AND EXTRACT(MONTH FROM dr.date) = %s
        AND EXTRACT(DAY FROM dr.date) <= %s
        GROUP BY s.id, s.name
        ORDER BY s.id
        """
        
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(sql, (year, month, day))
                columns = [desc[0] for desc in cursor.description] if cursor.description else []
                rows = cursor.fetchall()
                return [dict(zip(columns, row)) for row in rows]
        except Exception as e:
            print(f"❌ Error fetching monthly data for {year}-{month:02d}: {e}")
            return []
    
    def get_daily_data(self, date: str):
        """Get daily report data for a specific date"""
        sql = """
        SELECT 
            s.id as store_id,
            s.name as store_name,
            dr.tables_served_validated as tables,
            dr.revenue_tax_included as revenue,
            dr.turnover_rate,
            (dr.revenue_tax_included / NULLIF(dr.tables_served_validated, 0)) as per_table_consumption
        FROM daily_report dr
        JOIN store s ON dr.store_id = s.id
        WHERE dr.date = %s
        ORDER BY s.id
        """
        
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(sql, (date,))
                columns = [desc[0] for desc in cursor.description] if cursor.description else []
                rows = cursor.fetchall()
                return [dict(zip(columns, row)) for row in rows]
        except Exception as e:
            print(f"❌ Error fetching daily data for {date}: {e}")
            return []
    
    def calculate_percentage_change(self, current, previous):
        """Calculate percentage change between current and previous values"""
        if previous == 0 or previous is None:
            return 0.0
        return ((current - previous) / previous) * 100
    
    def format_percentage_change(self, change):
        """Format percentage change with appropriate color coding"""
        if change > 0:
            return f"{change:.1f}%"
        elif change < 0:
            return f"{change:.1f}%"
        else:
            return "0.0%"
    
    def generate_comparison_data(self):
        """Generate year-over-year comparison data"""
        print(f"📊 Generating year-over-year comparison for {self.target_date}")
        
        # Get current year data (month-to-date)
        current_mtd_data = self.get_monthly_data_up_to_date(self.current_year, self.month, self.day)
        
        # Get previous year data (same period)
        previous_mtd_data = self.get_monthly_data_up_to_date(self.previous_year, self.month, self.day)
        
        if not current_mtd_data:
            print(f"❌ No current year data found for {self.current_year}-{self.month:02d}")
            return None
        
        if not previous_mtd_data:
            print(f"❌ No previous year data found for {self.previous_year}-{self.month:02d}")
            return None
        
        print(f"✅ Found data for {len(current_mtd_data)} stores (current) and {len(previous_mtd_data)} stores (previous)")
        
        # Convert to dictionaries for easier lookup
        current_dict = {row['store_id']: row for row in current_mtd_data}
        previous_dict = {row['store_id']: row for row in previous_mtd_data}
        
        # Build comparison data structure
        comparison_data = {}
        
        # Calculate data for each store
        for store_id in current_dict.keys():
            store_name = self.store_names.get(store_id, f"Store {store_id}")
            current = current_dict.get(store_id, {})
            previous = previous_dict.get(store_id, {})
            
            # Current year values
            current_tables = float(current.get('total_tables', 0))
            current_revenue = float(current.get('total_revenue', 0))
            current_turnover = float(current.get('avg_turnover_rate', 0))
            current_per_table = float(current.get('avg_per_table', 0))
            
            # Previous year values
            previous_tables = float(previous.get('total_tables', 0)) if previous else 0
            previous_revenue = float(previous.get('total_revenue', 0)) if previous else 0
            previous_turnover = float(previous.get('avg_turnover_rate', 0)) if previous else 0
            previous_per_table = float(previous.get('avg_per_table', 0)) if previous else 0
            
            # Calculate changes
            tables_change = current_tables - previous_tables
            revenue_change = current_revenue - previous_revenue
            turnover_change = current_turnover - previous_turnover
            per_table_change = current_per_table - previous_per_table
            
            # Calculate percentage changes
            tables_pct = self.calculate_percentage_change(current_tables, previous_tables)
            revenue_pct = self.calculate_percentage_change(current_revenue, previous_revenue)
            turnover_pct = self.calculate_percentage_change(current_turnover, previous_turnover)
            per_table_pct = self.calculate_percentage_change(current_per_table, previous_per_table)
            
            comparison_data[store_name] = {
                # 桌数对比同期数据
                '本月截止目前': round(current_tables, 2),
                '去年截止同期': round(previous_tables, 2),
                '对比去年同期': round(tables_change, 2),
                '桌数增长率': self.format_percentage_change(tables_pct),
                
                # 翻台率对比同期数据
                '本月截止目前翻台率': round(current_turnover, 2),
                '去年截止同期翻台率': round(previous_turnover, 2),
                '对比去年同期翻台率': round(turnover_change, 2),
                '翻台率增长率': self.format_percentage_change(turnover_pct),
                
                # 营业收入(不含税-万加元)
                '本月截止目前收入': round(current_revenue / 10000, 2),
                '去年截止同期收入': round(previous_revenue / 10000, 2),
                '对比去年同期收入': round(revenue_change / 10000, 2),
                '收入增长率': self.format_percentage_change(revenue_pct),
                
                # 单桌消费对比同期数据
                '本月截止目前单桌': round(current_per_table, 2),
                '去年截止同期单桌': round(previous_per_table, 2),
                '对比去年同期单桌': round(per_table_change, 2),
                '单桌消费增长率': self.format_percentage_change(per_table_pct)
            }
        
        # Calculate totals for 加拿大片区
        total_current_tables = sum(float(row['total_tables']) for row in current_mtd_data)
        total_previous_tables = sum(float(row['total_tables']) for row in previous_mtd_data)
        total_current_revenue = sum(float(row['total_revenue']) for row in current_mtd_data)
        total_previous_revenue = sum(float(row['total_revenue']) for row in previous_mtd_data)
        total_current_turnover = sum(float(row['avg_turnover_rate']) for row in current_mtd_data) / len(current_mtd_data)
        total_previous_turnover = sum(float(row['avg_turnover_rate']) for row in previous_mtd_data) / len(previous_mtd_data) if previous_mtd_data else 0
        total_current_per_table = sum(float(row['avg_per_table']) for row in current_mtd_data) / len(current_mtd_data)
        total_previous_per_table = sum(float(row['avg_per_table']) for row in previous_mtd_data) / len(previous_mtd_data) if previous_mtd_data else 0
        
        # Calculate total changes
        total_tables_change = total_current_tables - total_previous_tables
        total_revenue_change = total_current_revenue - total_previous_revenue
        total_turnover_change = total_current_turnover - total_previous_turnover
        total_per_table_change = total_current_per_table - total_previous_per_table
        
        # Calculate total percentage changes
        total_tables_pct = self.calculate_percentage_change(total_current_tables, total_previous_tables)
        total_revenue_pct = self.calculate_percentage_change(total_current_revenue, total_previous_revenue)
        total_turnover_pct = self.calculate_percentage_change(total_current_turnover, total_previous_turnover)
        total_per_table_pct = self.calculate_percentage_change(total_current_per_table, total_previous_per_table)
        
        comparison_data['加拿大片区'] = {
            '本月截止目前': round(total_current_tables, 2),
            '去年截止同期': round(total_previous_tables, 2),
            '对比去年同期': round(total_tables_change, 2),
            '桌数增长率': self.format_percentage_change(total_tables_pct),
            '本月截止目前翻台率': round(total_current_turnover, 2),
            '去年截止同期翻台率': round(total_previous_turnover, 2),
            '对比去年同期翻台率': round(total_turnover_change, 2),
            '翻台率增长率': self.format_percentage_change(total_turnover_pct),
            '本月截止目前收入': round(total_current_revenue / 10000, 2),
            '去年截止同期收入': round(total_previous_revenue / 10000, 2),
            '对比去年同期收入': round(total_revenue_change / 10000, 2),
            '收入增长率': self.format_percentage_change(total_revenue_pct),
            '本月截止目前单桌': round(total_current_per_table, 2),
            '去年截止同期单桌': round(total_previous_per_table, 2),
            '对比去年同期单桌': round(total_per_table_change, 2),
            '单桌消费增长率': self.format_percentage_change(total_per_table_pct)
        }
        
        return comparison_data
    
    def create_excel_report(self, comparison_data):
        """Create formatted Excel report matching the uploaded image format"""
        if not comparison_data:
            print("❌ No comparison data to create report")
            return None
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("同比数据")
        else:
            ws.title = "同比数据"
        
        # Get weekday in Chinese
        target_dt = datetime.strptime(self.target_date, '%Y-%m-%d')
        weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
        weekday = weekdays[target_dt.weekday()]
        
        # Title
        title = f"加拿大-各门店{self.current_year}年{self.month}月{self.day}日同比数据-{weekday}"
        ws.merge_cells('A1:J1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        # Headers - split into regions like the image
        # Row 2: Main headers
        ws.merge_cells('A2:B2')
        ws['A2'] = "分类"
        ws.merge_cells('C2:E2')
        ws['C2'] = "西部"
        ws.merge_cells('F2:H2')
        ws['F2'] = "东部"
        ws['I2'] = "加拿大片区"
        
        # Row 3: Store names
        headers_row3 = ["项目", "内容", "加拿大一店", "加拿大二店", "加拿大七店", "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大片区"]
        for col, header in enumerate(headers_row3, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply header formatting
        for row in [2, 3]:
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows matching the image structure
        data_rows = [
            # 桌数对比同期数据 section
            ("桌数\n对比同期数据", "本月截止目前", "FFFF99"),
            ("", "去年截止同期", "FFFF99"),
            ("", "对比去年同期", "FFFF99"),
            ("", "桌数增长率", "FFFF00"),  # Highlighted
            
            # 翻台率对比同期数据 section
            ("翻台率\n对比同期数据", "本月截止目前", "E6F3FF"),
            ("", "去年截止同期", "E6F3FF"),
            ("", "对比去年同期", "E6F3FF"),
            ("", "翻台率增长率", "FFFF00"),  # Highlighted
            
            # 营业收入 section
            ("营业收入\n(不含税-万加元)", "本月截止目前", "FFFF99"),
            ("", "去年截止同期", "FFFF99"),
            ("", "对比去年同期", "FFFF99"),
            ("", "收入增长率", "FFFF00"),  # Highlighted
            
            # 单桌消费对比同期数据 section
            ("单桌消费\n对比同期数据", "本月截止目前", "E6F3FF"),
            ("", "去年截止同期", "E6F3FF"),
            ("", "对比去年同期", "E6F3FF"),
            ("", "单桌消费增长率", "FFFF00"),  # Highlighted
        ]
        
        # Store order matching the image (西部: 一店,二店,七店; 东部: 三店,四店,五店,六店)
        store_order = ["加拿大一店", "加拿大二店", "加拿大七店", "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大片区"]
        
        # Add data to worksheet
        current_row = 4
        
        for category, content, color in data_rows:
            # Add category (column A)
            if category:
                ws.cell(row=current_row, column=1, value=category)
            
            # Add content (column B)
            ws.cell(row=current_row, column=2, value=content)
            
            # Add data for each store in the specified order
            for col, store_name in enumerate(store_order, 3):
                if store_name in comparison_data:
                    # Map content to the correct data key
                    data_key = content
                    if content == "本月截止目前" and "翻台率" in category:
                        data_key = "本月截止目前翻台率"
                    elif content == "去年截止同期" and "翻台率" in category:
                        data_key = "去年截止同期翻台率"
                    elif content == "对比去年同期" and "翻台率" in category:
                        data_key = "对比去年同期翻台率"
                    elif content == "本月截止目前" and "收入" in category:
                        data_key = "本月截止目前收入"
                    elif content == "去年截止同期" and "收入" in category:
                        data_key = "去年截止同期收入"
                    elif content == "对比去年同期" and "收入" in category:
                        data_key = "对比去年同期收入"
                    elif content == "本月截止目前" and "单桌" in category:
                        data_key = "本月截止目前单桌"
                    elif content == "去年截止同期" and "单桌" in category:
                        data_key = "去年截止同期单桌"
                    elif content == "对比去年同期" and "单桌" in category:
                        data_key = "对比去年同期单桌"
                    
                    value = comparison_data[store_name].get(data_key, "")
                    cell = ws.cell(row=current_row, column=col, value=value)
                    
                    # Apply background color
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    
                    # Apply red color for negative percentage changes
                    if isinstance(value, str) and value.endswith('%') and value.startswith('-'):
                        cell.font = Font(color="FF0000")  # Red color for negative percentages
            
            # Apply background color to category and content cells
            ws.cell(row=current_row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=current_row, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            current_row += 1
        
        # Merge category cells
        merge_ranges = [
            (4, 7),   # 桌数对比同期数据 - 4 rows
            (8, 11),  # 翻台率对比同期数据 - 4 rows
            (12, 15), # 营业收入 - 4 rows
            (16, 19)  # 单桌消费对比同期数据 - 4 rows
        ]
        
        for start_row, end_row in merge_ranges:
            ws.merge_cells(f'A{start_row}:A{end_row}')
            cell = ws[f'A{start_row}']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
        
        # Apply borders
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for row in range(1, current_row):
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = thin_border
        
        # Set column widths
        column_widths = [15, 20, 12, 12, 12, 12, 12, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Set row heights
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 20
        
        return wb
    
    def generate_report(self):
        """Main method to generate the complete year-over-year comparison report"""
        print(f"📈 Generating Year-over-Year Comparison Report for {self.target_date}")
        print("=" * 70)
        
        # Generate comparison data from database
        comparison_data = self.generate_comparison_data()
        if not comparison_data:
            return None
        
        # Create Excel report
        wb = self.create_excel_report(comparison_data)
        if not wb:
            return None
        
        # Save the file
        filename = f"yearly_comparison_report_{self.target_date.replace('-', '_')}.xlsx"
        output_path = self.output_dir / filename
        
        wb.save(output_path)
        
        print(f"✅ Year-over-year comparison report generated successfully!")
        print(f"📁 Saved to: {output_path}")
        print(f"📊 Report compares {self.current_year} vs {self.previous_year} data")
        print(f"📈 Contains data for {len(comparison_data)-1} stores + total")
        
        return output_path

def main():
    """Main function"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate year-over-year comparison report from database")
    parser.add_argument("--date", default="2025-06-10", help="Target date (YYYY-MM-DD)")
    parser.add_argument("--test", action="store_true", help="Use test database")
    
    args = parser.parse_args()
    
    try:
        generator = YearlyComparisonReportGenerator(args.date, is_test=args.test)
        output_path = generator.generate_report()
        
        if output_path:
            print(f"\n🎯 Success! Year-over-year comparison report created: {output_path}")
            print(f"📁 Located in OUTPUT_DIR: {os.getenv('OUTPUT_DIR', './output')}")
            print(f"📊 Compares {args.date} with same period from previous year")
        else:
            print("\n❌ Failed to generate report")
            sys.exit(1)
            
    except Exception as e:
        print(f"❌ Error generating report: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main() 