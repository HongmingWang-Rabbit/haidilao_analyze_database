#!/usr/bin/env python3
"""Verify the final corrected output file"""

import openpyxl
import sys

# Set UTF-8 encoding for Windows
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Check the final corrected file
output_file = "output/hi-bowl/final_corrected.xlsx"

wb = openpyxl.load_workbook(output_file, data_only=True)
ws = wb['海外新业态管报收入数据-本位币']

print("Final corrected Hi-Bowl report values:")
print("-" * 50)

# Check all the filled values
checks = [
    ("K8", "优惠总金额（含税）", 2627.30),
    ("L8", "优惠总金额（不含税）", 2284.61),
    ("M8", "小费", 0.02),
    ("N8", "实际营业天数", 3),
    ("O8", "应营业天数", 3),
    ("P8", "工作日天数", 3),
    ("Q8", "节假日天数", 0),
    ("W8", "全月就餐人数", 1615),
    ("X8", "工作日就餐人数", 1615),
    ("Y8", "节假日就餐人数", 0),
    ("AA8", "全月订单数量", 555),
    ("AB8", "工作日订单数", 555),
    ("AC8", "节假日订单数", 0),
    ("AS8", "08:00-13:59销售收入", 10703.59),
    ("AT8", "14:00-16:59销售收入", 7873.93),
    ("AU8", "17:00-21:59销售收入", 36682.54),
    ("AV8", "22:00-07:59销售收入", 12000.36),
    ("BC8", "08:00-13:59订单数", 96),
    ("BD8", "14:00-16:59订单数", 74),
    ("BE8", "17:00-21:59订单数", 283),
    ("BF8", "22:00-(次)07:59订单数", 102),
]

all_match = True
for cell_ref, desc, expected in checks:
    actual = ws[cell_ref].value or 0
    match = abs(actual - expected) < 0.01 if isinstance(expected, float) else actual == expected
    status = "✓" if match else "✗"
    print(f"{status} {cell_ref} ({desc}): {actual} {'=' if match else '≠'} {expected}")
    if not match:
        all_match = False

# Calculate time segment total
segment_total = ws["AS8"].value + ws["AT8"].value + ws["AU8"].value + ws["AV8"].value
print(f"\nTime segment revenue total: ${segment_total:,.2f}")
print(f"Expected total: $67,260.42")
print(f"Match: {'YES ✅' if abs(segment_total - 67260.42) < 0.01 else 'NO ❌'}")

print(f"\nAll values correct: {'YES ✅' if all_match else 'NO ❌'}")

wb.close()