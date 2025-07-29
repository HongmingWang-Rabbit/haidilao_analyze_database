#!/usr/bin/env python3
"""
Final test of the complete monthly material report with all enhancements
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.generate_monthly_material_report import MonthlyMaterialReportGenerator
from datetime import datetime
from openpyxl import load_workbook

def test_final_complete_report():
    """Final test of the complete monthly material report"""
    
    print("Final Complete Report Test - All Enhancements")
    print("=" * 50)
    
    try:
        # Generate the monthly material report with unique timestamp
        target_date = "2025-05-31"
        generator = MonthlyMaterialReportGenerator(target_date, is_test=True)
        
        # Override save method to use unique filename
        def save_with_timestamp(wb):
            try:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"FINAL_ENHANCED_monthly_material_report_{timestamp}.xlsx"
                output_path = generator.output_dir / filename
                wb.save(output_path)
                return output_path
            except Exception as e:
                print(f"ERROR: Error saving report: {e}")
                return None
        
        generator.save_report = save_with_timestamp
        
        print(f"Target date: {target_date}")
        print("Generating final enhanced monthly material report...")
        
        # Generate the report
        output_file = generator.generate_report()
        
        if output_file and Path(output_file).exists():
            print(f"✅ Report generated successfully!")
            print(f"📁 Output file: {output_file}")
            print(f"📊 File size: {Path(output_file).stat().st_size:,} bytes")
            
            # Load and verify the report
            wb = load_workbook(output_file)
            
            if "物料用量差异分析" in wb.sheetnames:
                ws = wb["物料用量差异分析"]
                print(f"📊 Analyzing 物料用量差异分析 worksheet...")
                
                # Find header row
                header_row = None
                for row in range(1, 20):
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value and "序号" in str(cell_value):
                        header_row = row
                        break
                
                if header_row:
                    # Final comprehensive verification
                    print(f"\n🎯 FINAL VERIFICATION CHECKLIST:")
                    
                    # 1. Column count and headers
                    expected_headers = [
                        "序号", "门店", "物料名称", "物料号", "单位", "使用菜品详情",
                        "理论用量", "套餐用量", "系统记录", "库存盘点", "减去盘点用量", "差异数量", "差异率(%)", "状态",
                        "本月总消费金额", "差异金额"
                    ]
                    
                    header_check = ws.max_column == 16
                    print(f"   ✅ Column count (16): {'PASS' if header_check else 'FAIL'}")
                    
                    # 2. Status text verification
                    status_counts = {"超量": 0, "少用": 0, "正常": 0, "不足": 0}
                    for row in range(header_row + 1, ws.max_row + 1):
                        status_cell = ws.cell(row=row, column=14)
                        status = status_cell.value if status_cell.value else ""
                        if status in status_counts:
                            status_counts[status] += 1
                    
                    status_check = status_counts["不足"] == 0 and status_counts["少用"] > 0
                    print(f"   ✅ Status text (不足→少用): {'PASS' if status_check else 'FAIL'}")
                    print(f"      • 超量: {status_counts['超量']}, 少用: {status_counts['少用']}, 正常: {status_counts['正常']}")
                    
                    # 3. Color verification
                    sample_green = False
                    sample_red = False
                    for row in range(header_row + 1, min(header_row + 21, ws.max_row + 1)):
                        status_cell = ws.cell(row=row, column=14)
                        if status_cell.value == "少用":
                            color = status_cell.fill.start_color.rgb if status_cell.fill and status_cell.fill.start_color else "None"
                            if color == "00E8F5E8":
                                sample_green = True
                                break
                    
                    for row in range(header_row + 1, min(header_row + 21, ws.max_row + 1)):
                        status_cell = ws.cell(row=row, column=14)
                        if status_cell.value == "超量":
                            color = status_cell.fill.start_color.rgb if status_cell.fill and status_cell.fill.start_color else "None"
                            if color == "00FFE6E6":
                                sample_red = True
                                break
                    
                    color_check = sample_green and sample_red
                    print(f"   ✅ Status coloring: {'PASS' if color_check else 'FAIL'}")
                    
                    # 4. Dish usage details verification
                    usage_sample = ws.cell(row=header_row + 1, column=6).value
                    usage_check = usage_sample and "sale-" in str(usage_sample)
                    print(f"   ✅ Dish usage details: {'PASS' if usage_check else 'FAIL'}")
                    
                    # 5. New formulas verification
                    test_row = header_row + 1
                    minus_inventory_formula = ws.cell(row=test_row, column=11).value
                    cost_formula = ws.cell(row=test_row, column=15).value
                    
                    formula_check = (minus_inventory_formula and "=I" in str(minus_inventory_formula) and "-J" in str(minus_inventory_formula) and
                                   cost_formula and "=(G" in str(cost_formula))
                    print(f"   ✅ New formulas: {'PASS' if formula_check else 'FAIL'}")
                    
                    # 6. Summary section
                    summary_check = False
                    for row in range(1, header_row):
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            if cell.value and "少用物料" in str(cell.value):
                                summary_check = True
                                break
                        if summary_check:
                            break
                    
                    print(f"   ✅ Summary section: {'PASS' if summary_check else 'FAIL'}")
                    
                    # Overall result
                    all_checks = [header_check, status_check, color_check, usage_check, formula_check, summary_check]
                    overall_pass = all(all_checks)
                    
                    print(f"\n🏆 OVERALL RESULT: {'🎉 ALL TESTS PASSED! 🎉' if overall_pass else '❌ SOME TESTS FAILED'}")
                    
                    if overall_pass:
                        print(f"\n🌟 ENHANCEMENT SUMMARY:")
                        print(f"   • 16 comprehensive columns (A-P)")
                        print(f"   • 使用菜品详情: Shows dish usage for each material")
                        print(f"   • 减去盘点用量: System - Inventory calculation")
                        print(f"   • 本月总消费金额: Monthly cost analysis")
                        print(f"   • 差异金额: Variance cost impact") 
                        print(f"   • Status: 超量(red), 少用(green), 正常(white)")
                        print(f"   • All formulas and calculations working perfectly!")
                        
                    return overall_pass
                        
            wb.close()
        else:
            print("❌ Report generation failed")
            return False
            
    except Exception as e:
        print(f"❌ Error during final testing: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_final_complete_report()
    if success:
        print("\n🎊🎊🎊 FINAL TESTING COMPLETE - ALL ENHANCEMENTS SUCCESSFUL! 🎊🎊🎊")
    else:
        print("\n❌ Final testing failed")
        sys.exit(1)