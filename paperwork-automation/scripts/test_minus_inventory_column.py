#!/usr/bin/env python3
"""
Test the new 减去盘点用量 column (系统记录 - 库存盘点)
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.monthly_dishes_worksheet import MonthlyDishesWorksheetGenerator
from lib.database_queries import ReportDataProvider
from utils.database import DatabaseManager, DatabaseConfig
from openpyxl import Workbook
from datetime import datetime

def test_minus_inventory_column():
    """Test the new 减去盘点用量 column"""
    
    print("Testing 减去盘点用量 Column")
    print("=" * 35)
    
    try:
        # Initialize components
        config = DatabaseConfig(is_test=True)
        db_manager = DatabaseManager(config)
        data_provider = ReportDataProvider(db_manager)
        
        # Create worksheet generator
        target_date = "2025-05-31"
        store_names = ["加拿大一店", "加拿大二店", "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大七店"]
        worksheet_gen = MonthlyDishesWorksheetGenerator(store_names, target_date)
        
        # Create workbook and generate worksheet
        wb = Workbook()
        ws = worksheet_gen.generate_material_variance_worksheet(wb, data_provider)
        
        print(f"✅ Worksheet generated: {ws.title}")
        print(f"📊 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Find the header row
        header_row = None
        for row in range(1, 20):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "序号" in str(cell_value):
                header_row = row
                break
        
        if header_row:
            print(f"📋 Header row found at row {header_row}")
            print("\nColumn structure:")
            for col in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=header_row, column=col)
                col_letter = chr(64 + col)  # A=65, so col 1 = A
                header_value = header_cell.value if header_cell.value else f"<empty>"
                print(f"   {col_letter}: {header_value}")
            
            # Check for the 减去盘点用量 column
            minus_inventory_col_found = False
            minus_inventory_col_num = None
            for col in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=header_row, column=col)
                if header_cell.value and "减去盘点用量" in str(header_cell.value):
                    minus_inventory_col_found = True
                    minus_inventory_col_num = col
                    break
            
            if minus_inventory_col_found:
                print(f"\n✅ 减去盘点用量 column found at column {chr(64 + minus_inventory_col_num)}")
                
                # Verify the formula and sample calculations
                print("\n📊 Sample calculations (系统记录 - 库存盘点):")
                for row in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
                    material_cell = ws.cell(row=row, column=3)  # Material name
                    store_cell = ws.cell(row=row, column=2)     # Store name
                    system_cell = ws.cell(row=row, column=9)    # 系统记录 (I)
                    inventory_cell = ws.cell(row=row, column=10) # 库存盘点 (J)
                    minus_inventory_cell = ws.cell(row=row, column=11)  # 减去盘点用量 (K)
                    
                    material_name = material_cell.value[:30] if material_cell.value else "Unknown"
                    store_name = store_cell.value if store_cell.value else "Unknown"
                    system_value = system_cell.value
                    inventory_value = inventory_cell.value
                    formula = minus_inventory_cell.value
                    
                    print(f"   Row {row}: {store_name} | {material_name}")
                    print(f"      系统记录: {system_value}")
                    print(f"      库存盘点: {inventory_value}")
                    print(f"      Formula: {formula}")
                    print()
                    
            else:
                print("❌ 减去盘点用量 column not found")
                
            # Verify expected column count (should be 16 now)
            expected_cols = 16
            if ws.max_column == expected_cols:
                print(f"✅ Column count correct: {ws.max_column} columns")
            else:
                print(f"❌ Column count incorrect: expected {expected_cols}, got {ws.max_column}")
                
            # Check that other formulas were updated correctly
            print(f"\n🧮 Formula verification:")
            test_row = header_row + 1
            
            # Check variance formula (should now be L = I - (G+H+J))
            variance_cell = ws.cell(row=test_row, column=12)  # L column
            variance_formula = variance_cell.value
            print(f"   差异数量 formula (L): {variance_formula}")
            if variance_formula and "=I" in str(variance_formula) and "-(G" in str(variance_formula):
                print("   ✅ Variance formula correct")
            else:
                print("   ❌ Variance formula incorrect")
            
            # Check cost formula (should now be O = (G+H+I)*price)
            cost_cell = ws.cell(row=test_row, column=15)  # O column
            cost_formula = cost_cell.value
            print(f"   本月总消费金额 formula (O): {cost_formula}")
            if cost_formula and "=(G" in str(cost_formula) and "+H" in str(cost_formula) and "+I" in str(cost_formula):
                print("   ✅ Cost formula correct")
            else:
                print("   ❌ Cost formula incorrect")
        
        # Save output
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"output/test_minus_inventory_column_{timestamp}.xlsx"
        wb.save(output_file)
        print(f"\n📁 Test output saved to: {output_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_minus_inventory_column()