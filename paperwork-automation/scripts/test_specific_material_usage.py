#!/usr/bin/env python3
"""
Test dish usage details for specific material 3000759 清油底料（300G*40包/件）
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.monthly_dishes_worksheet import MonthlyDishesWorksheetGenerator
from lib.database_queries import ReportDataProvider
from utils.database import DatabaseManager, DatabaseConfig
from openpyxl import Workbook

def test_specific_material_usage():
    """Test dish usage details for specific material"""
    
    print("Testing Specific Material Usage (3000759)")
    print("=" * 45)
    
    try:
        # Initialize components
        config = DatabaseConfig(is_test=True)
        db_manager = DatabaseManager(config)
        data_provider = ReportDataProvider(db_manager)
        
        # Create worksheet generator
        target_date = "2025-05-31"
        store_names = ["加拿大一店", "加拿大二店", "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大七店"]
        worksheet_gen = MonthlyDishesWorksheetGenerator(store_names, target_date)
        
        # Create workbook and generate worksheet
        wb = Workbook()
        ws = worksheet_gen.generate_material_variance_worksheet(wb, data_provider)
        
        print(f"✅ Worksheet generated: {ws.title}")
        print(f"📊 Searching for material 3000759...")
        
        # Find the header row
        header_row = None
        for row in range(1, 20):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "序号" in str(cell_value):
                header_row = row
                break
        
        if header_row:
            # Search for material 3000759
            found_materials = []
            
            for row in range(header_row + 1, ws.max_row + 1):
                material_num_cell = ws.cell(row=row, column=4)  # Material number column
                material_name_cell = ws.cell(row=row, column=3)  # Material name column
                store_cell = ws.cell(row=row, column=2)          # Store column
                usage_cell = ws.cell(row=row, column=6)          # Usage details column
                
                material_num = str(material_num_cell.value) if material_num_cell.value else ""
                material_name = material_name_cell.value if material_name_cell.value else ""
                store_name = store_cell.value if store_cell.value else ""
                usage_details = usage_cell.value if usage_cell.value else ""
                
                # Check for the specific material number or name containing 清油底料
                if "3000759" in material_num or "清油底料" in material_name:
                    found_materials.append({
                        'row': row,
                        'store': store_name,
                        'material_name': material_name,
                        'material_number': material_num,
                        'usage_details': usage_details
                    })
            
            if found_materials:
                print(f"✅ Found {len(found_materials)} records for 清油底料:")
                for material in found_materials:
                    print(f"\n📍 Row {material['row']}: {material['store']}")
                    print(f"   Material: {material['material_name']}")
                    print(f"   Number: {material['material_number']}")
                    print(f"   Usage Details:")
                    for line in material['usage_details'].split('\n'):
                        if line.strip():
                            print(f"      {line}")
            else:
                print("❌ Material 3000759 清油底料 not found in the report")
                
                # Search for any materials containing "清油" or "底料"
                print("\n🔍 Searching for materials containing '清油' or '底料'...")
                similar_materials = []
                
                for row in range(header_row + 1, min(header_row + 21, ws.max_row + 1)):
                    material_name_cell = ws.cell(row=row, column=3)
                    material_name = material_name_cell.value if material_name_cell.value else ""
                    
                    if "清油" in material_name or "底料" in material_name:
                        material_num_cell = ws.cell(row=row, column=4)
                        store_cell = ws.cell(row=row, column=2)
                        
                        similar_materials.append({
                            'row': row,
                            'store': store_cell.value,
                            'material_name': material_name,
                            'material_number': material_num_cell.value
                        })
                
                if similar_materials:
                    print(f"Found {len(similar_materials)} similar materials:")
                    for material in similar_materials[:5]:  # Show first 5
                        print(f"   Row {material['row']}: {material['store']} | {material['material_name']} | {material['material_number']}")
                else:
                    print("No similar materials found")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_specific_material_usage()