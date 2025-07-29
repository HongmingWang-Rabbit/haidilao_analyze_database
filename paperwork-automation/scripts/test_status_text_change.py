#!/usr/bin/env python3
"""
Test the status text change from 不足 to 少用
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.monthly_dishes_worksheet import MonthlyDishesWorksheetGenerator
from lib.database_queries import ReportDataProvider
from utils.database import DatabaseManager, DatabaseConfig
from openpyxl import Workbook
from datetime import datetime

def test_status_text_change():
    """Test the status text change from 不足 to 少用"""
    
    print("Testing Status Text Change (不足 → 少用)")
    print("=" * 40)
    
    try:
        # Initialize components
        config = DatabaseConfig(is_test=True)
        db_manager = DatabaseManager(config)
        data_provider = ReportDataProvider(db_manager)
        
        # Create worksheet generator
        target_date = "2025-05-31"
        store_names = ["加拿大一店", "加拿大二店", "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大七店"]
        worksheet_gen = MonthlyDishesWorksheetGenerator(store_names, target_date)
        
        # Create workbook and generate worksheet
        wb = Workbook()
        ws = worksheet_gen.generate_material_variance_worksheet(wb, data_provider)
        
        print(f"✅ Worksheet generated: {ws.title}")
        print(f"📊 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Find the header row
        header_row = None
        for row in range(1, 20):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "序号" in str(cell_value):
                header_row = row
                break
        
        if header_row:
            print(f"📋 Header row found at row {header_row}")
            
            # Count status occurrences
            status_counts = {"超量": 0, "少用": 0, "正常": 0, "不足": 0}
            status_examples = {"超量": [], "少用": [], "正常": []}
            
            for row in range(header_row + 1, ws.max_row + 1):
                status_cell = ws.cell(row=row, column=14)  # Status column (N)
                status = status_cell.value if status_cell.value else ""
                
                if status in status_counts:
                    status_counts[status] += 1
                    
                    # Collect examples with color info
                    if status in status_examples and len(status_examples[status]) < 3:
                        fill_color = status_cell.fill.start_color.rgb if status_cell.fill and status_cell.fill.start_color else "None"
                        material_cell = ws.cell(row=row, column=3)
                        store_cell = ws.cell(row=row, column=2)
                        
                        status_examples[status].append({
                            'row': row,
                            'store': store_cell.value,
                            'material': material_cell.value[:30] if material_cell.value else "Unknown",
                            'color': fill_color
                        })
            
            print(f"\n📊 Status Distribution:")
            for status, count in status_counts.items():
                print(f"   {status}: {count} occurrences")
            
            # Verify no old "不足" status exists
            if status_counts["不足"] > 0:
                print(f"\n❌ ERROR: Found {status_counts['不足']} occurrences of old '不足' status!")
            else:
                print(f"\n✅ SUCCESS: No old '不足' status found")
            
            # Verify new "少用" status exists and has correct color
            if status_counts["少用"] > 0:
                print(f"\n✅ SUCCESS: Found {status_counts['少用']} occurrences of new '少用' status")
                
                print(f"\n🎨 Status Examples with Colors:")
                for status, examples in status_examples.items():
                    if examples:
                        print(f"   {status}:")
                        for example in examples:
                            print(f"      Row {example['row']}: {example['store']} | {example['material']} | Color: {example['color']}")
                
                # Verify color coding
                if status_examples["少用"]:
                    first_example = status_examples["少用"][0]
                    if first_example['color'] == "00E8F5E8":
                        print(f"\n✅ Color verification: 少用 status has correct green color")
                    else:
                        print(f"\n❌ Color verification: 少用 status has incorrect color: {first_example['color']}")
                        
            else:
                print(f"\n⚠️ No '少用' status found in current data (might be normal if no under-usage)")
            
            # Check summary section for updated text
            print(f"\n📋 Checking summary section...")
            summary_found = False
            for row in range(1, header_row):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and "少用物料" in str(cell.value):
                        summary_found = True
                        print(f"   ✅ Found updated summary text: '{cell.value}' at row {row}")
                        break
                if summary_found:
                    break
            
            if not summary_found:
                print(f"   ⚠️ Updated summary text not found (checking nearby cells)")
                # Check for old text
                old_text_found = False
                for row in range(1, header_row):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value and "使用不足物料" in str(cell.value):
                            old_text_found = True
                            print(f"   ❌ Found old summary text: '{cell.value}' at row {row}")
                            break
                    if old_text_found:
                        break
                
                if not old_text_found:
                    print(f"   ✅ No old summary text found")
        
        # Save output
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"output/test_status_text_change_{timestamp}.xlsx"
        wb.save(output_file)
        print(f"\n📁 Test output saved to: {output_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_status_text_change()