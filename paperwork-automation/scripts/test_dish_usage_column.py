#!/usr/bin/env python3
"""
Test the new dish usage details column in the material variance worksheet
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.monthly_dishes_worksheet import MonthlyDishesWorksheetGenerator
from lib.database_queries import ReportDataProvider
from utils.database import DatabaseManager, DatabaseConfig
from openpyxl import Workbook
from datetime import datetime

def test_dish_usage_column():
    """Test the dish usage details column"""
    
    print("Testing Dish Usage Details Column")
    print("=" * 35)
    
    try:
        # Initialize components
        config = DatabaseConfig(is_test=True)
        db_manager = DatabaseManager(config)
        data_provider = ReportDataProvider(db_manager)
        
        # Create worksheet generator
        target_date = "2025-05-31"
        store_names = ["加拿大一店", "加拿大二店", "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大七店"]
        worksheet_gen = MonthlyDishesWorksheetGenerator(store_names, target_date)
        
        # Create workbook and generate worksheet
        wb = Workbook()
        ws = worksheet_gen.generate_material_variance_worksheet(wb, data_provider)
        
        print(f"✅ Worksheet generated: {ws.title}")
        print(f"📊 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Find the header row
        header_row = None
        for row in range(1, 20):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "序号" in str(cell_value):
                header_row = row
                break
        
        if header_row:
            print(f"📋 Header row found at row {header_row}")
            print("\nColumn structure:")
            for col in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=header_row, column=col)
                col_letter = chr(64 + col)  # A=65, so col 1 = A
                header_value = header_cell.value if header_cell.value else f"<empty>"
                print(f"   {col_letter}: {header_value}")
            
            # Check for the dish usage details column
            usage_col_found = False
            usage_col_num = None
            for col in range(1, ws.max_column + 1):
                header_cell = ws.cell(row=header_row, column=col)
                if header_cell.value and "使用菜品详情" in str(header_cell.value):
                    usage_col_found = True
                    usage_col_num = col
                    break
            
            if usage_col_found:
                print(f"\n✅ 使用菜品详情 column found at column {chr(64 + usage_col_num)}")
                
                # Check a few sample dish usage details
                print("\n📊 Sample dish usage details:")
                for row in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
                    material_cell = ws.cell(row=row, column=3)  # Material name
                    store_cell = ws.cell(row=row, column=2)     # Store name
                    usage_cell = ws.cell(row=row, column=usage_col_num)  # Usage details
                    
                    material_name = material_cell.value[:30] if material_cell.value else "Unknown"
                    store_name = store_cell.value if store_cell.value else "Unknown"
                    usage_details = usage_cell.value[:100] if usage_cell.value else "No details"
                    
                    print(f"   Row {row}: {store_name} | {material_name}")
                    print(f"      Usage: {usage_details}")
                    print()
                    
            else:
                print("❌ 使用菜品详情 column not found")
                
            # Verify expected column count (should be 15 now)
            expected_cols = 15
            if ws.max_column == expected_cols:
                print(f"✅ Column count correct: {ws.max_column} columns")
            else:
                print(f"❌ Column count incorrect: expected {expected_cols}, got {ws.max_column}")
        
        # Save output
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"output/test_dish_usage_column_{timestamp}.xlsx"
        wb.save(output_file)
        print(f"\n📁 Test output saved to: {output_file}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_dish_usage_column()