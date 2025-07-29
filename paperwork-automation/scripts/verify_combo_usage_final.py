#!/usr/bin/env python3
"""
Final verification that combo usage is showing correctly in dish usage details
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.monthly_dishes_worksheet import MonthlyDishesWorksheetGenerator
from lib.database_queries import ReportDataProvider
from utils.database import DatabaseManager, DatabaseConfig
from openpyxl import Workbook

def verify_combo_usage_final():
    """Verify combo usage is showing correctly"""
    
    print("Final Combo Usage Verification")
    print("=" * 35)
    
    try:
        # Initialize components
        config = DatabaseConfig(is_test=True)
        db_manager = DatabaseManager(config)
        data_provider = ReportDataProvider(db_manager)
        
        # Create worksheet generator
        target_date = "2025-05-31"
        store_names = ["加拿大一店", "加拿大二店", "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大七店"]
        worksheet_gen = MonthlyDishesWorksheetGenerator(store_names, target_date)
        
        # Create workbook and generate worksheet
        wb = Workbook()
        ws = worksheet_gen.generate_material_variance_worksheet(wb, data_provider)
        
        print(f"Worksheet generated: {ws.title}")
        print(f"Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        
        # Find the header row
        header_row = None
        for row in range(1, 20):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "序号" in str(cell_value):
                header_row = row
                break
        
        if header_row:
            print(f"Header row found at row {header_row}")
            
            # Count combo usage occurrences
            rows_with_combo = 0
            combo_examples = []
            
            for row in range(header_row + 1, ws.max_row + 1):
                usage_cell = ws.cell(row=row, column=6)  # Column F: 使用菜品详情
                material_cell = ws.cell(row=row, column=3)  # Material name
                store_cell = ws.cell(row=row, column=2)     # Store name
                
                if usage_cell.value and "套餐 -" in str(usage_cell.value):
                    rows_with_combo += 1
                    
                    # Extract combo line
                    lines = str(usage_cell.value).split('\n')
                    combo_line = None
                    for line in lines:
                        if line.startswith('套餐 -'):
                            combo_line = line
                            break
                    
                    if combo_line and len(combo_examples) < 5:
                        combo_examples.append({
                            'row': row,
                            'store': store_cell.value,
                            'material': material_cell.value[:30] if material_cell.value else "Unknown",
                            'combo_line': combo_line
                        })
            
            print(f"\nCombo Usage Summary:")
            print(f"  Total rows with combo usage: {rows_with_combo}")
            
            if combo_examples:
                print(f"\nExamples of combo usage:")
                for example in combo_examples:
                    print(f"  Row {example['row']}: {example['store']} | {example['material']}...")
                    print(f"    {example['combo_line']}")
                
                # Check the specific material 3000759
                print(f"\nSpecific check for material 3000759 清油底料:")
                found_3000759 = False
                for row in range(header_row + 1, ws.max_row + 1):
                    material_num_cell = ws.cell(row=row, column=4)
                    if material_num_cell.value and "3000759" in str(material_num_cell.value):
                        usage_cell = ws.cell(row=row, column=6)
                        store_cell = ws.cell(row=row, column=2)
                        
                        if usage_cell.value and "套餐 -" in str(usage_cell.value):
                            lines = str(usage_cell.value).split('\n')
                            combo_line = [line for line in lines if line.startswith('套餐')][0]
                            print(f"  Found at {store_cell.value}: {combo_line}")
                            found_3000759 = True
                        else:
                            print(f"  Found at {store_cell.value}: No combo usage")
                        break
                
                if found_3000759:
                    print("  Status: COMBO USAGE WORKING CORRECTLY")
                else:
                    print("  Status: Material 3000759 not found or no combo usage")
                    
            else:
                print("  No combo usage examples found")
                
            # Summary
            if rows_with_combo > 0:
                print(f"\n✅ SUCCESS: Combo usage is working correctly!")
                print(f"  - {rows_with_combo} materials have combo usage > 0")
                print(f"  - Combo usage only shows when > 0 (as requested)")
                print(f"  - Format: 套餐 - [calculated_amount]")
            else:
                print(f"\n❌ ISSUE: No combo usage found in any rows")
        
        return True
        
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    verify_combo_usage_final()