#!/usr/bin/env python3
"""
One-time script to generate 2025 vs 2024 time segment comparison report.
Compares time segment performance (turnover rate, table counts) across all stores
for the full year 2025 vs 2024.
"""

import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.database import get_database_manager
from lib.config import TIME_SEGMENTS, STORE_ID_TO_NAME_MAPPING


class YearlyTimeSegmentComparison:
    """Generate yearly time segment comparison report (2025 vs 2024)"""

    def __init__(self, db_manager):
        self.db_manager = db_manager
        self.time_segments = TIME_SEGMENTS
        self.store_names = STORE_ID_TO_NAME_MAPPING

    def get_yearly_time_segment_data(self, year: int):
        """Get aggregated time segment data for a full year"""
        sql = """
        SELECT
            str.store_id,
            ts.label as time_segment,
            AVG(str.turnover_rate) as avg_turnover_rate,
            SUM(str.tables_served_validated) as total_tables,
            COUNT(DISTINCT str.date) as days_count,
            SUM(str.tables_served_validated) / NULLIF(COUNT(DISTINCT str.date), 0) as avg_daily_tables
        FROM store_time_report str
        JOIN time_segment ts ON str.time_segment_id = ts.id
        WHERE EXTRACT(YEAR FROM str.date) = %s
            AND str.store_id BETWEEN 1 AND 8
        GROUP BY str.store_id, ts.label, ts.id
        ORDER BY str.store_id, ts.id
        """

        results = self.db_manager.fetch_all(sql, (year,))

        # Organize by store_id and time_segment
        data = {}
        for row in results:
            store_id = row['store_id']
            time_segment = row['time_segment']

            if store_id not in data:
                data[store_id] = {}

            data[store_id][time_segment] = {
                'avg_turnover_rate': float(row['avg_turnover_rate'] or 0),
                'total_tables': int(row['total_tables'] or 0),
                'days_count': int(row['days_count'] or 0),
                'avg_daily_tables': float(row['avg_daily_tables'] or 0)
            }

        return data

    def get_monthly_breakdown(self, year: int):
        """Get monthly breakdown of time segment data for a year"""
        sql = """
        SELECT
            str.store_id,
            EXTRACT(MONTH FROM str.date) as month,
            ts.label as time_segment,
            AVG(str.turnover_rate) as avg_turnover_rate,
            SUM(str.tables_served_validated) as total_tables,
            COUNT(DISTINCT str.date) as days_count
        FROM store_time_report str
        JOIN time_segment ts ON str.time_segment_id = ts.id
        WHERE EXTRACT(YEAR FROM str.date) = %s
            AND str.store_id BETWEEN 1 AND 8
        GROUP BY str.store_id, EXTRACT(MONTH FROM str.date), ts.label, ts.id
        ORDER BY str.store_id, EXTRACT(MONTH FROM str.date), ts.id
        """

        results = self.db_manager.fetch_all(sql, (year,))

        # Organize by store_id, month, time_segment
        data = {}
        for row in results:
            store_id = row['store_id']
            month = int(row['month'])
            time_segment = row['time_segment']

            if store_id not in data:
                data[store_id] = {}
            if month not in data[store_id]:
                data[store_id][month] = {}

            data[store_id][month][time_segment] = {
                'avg_turnover_rate': float(row['avg_turnover_rate'] or 0),
                'total_tables': int(row['total_tables'] or 0),
                'days_count': int(row['days_count'] or 0)
            }

        return data

    def generate_report(self, output_path: str = None):
        """Generate the comparison Excel report"""
        print("正在获取2025年分时段数据...")
        data_2025 = self.get_yearly_time_segment_data(2025)

        print("正在获取2024年分时段数据...")
        data_2024 = self.get_yearly_time_segment_data(2024)

        print("正在获取月度明细数据...")
        monthly_2025 = self.get_monthly_breakdown(2025)
        monthly_2024 = self.get_monthly_breakdown(2024)

        # Create workbook
        wb = Workbook()

        # Generate worksheets
        self._generate_summary_worksheet(wb, data_2025, data_2024)
        self._generate_store_detail_worksheets(wb, data_2025, data_2024)
        self._generate_monthly_comparison_worksheet(wb, monthly_2025, monthly_2024)
        self._generate_time_segment_trend_worksheet(wb, monthly_2025, monthly_2024)

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # Save workbook
        if output_path is None:
            output_dir = Path(__file__).parent.parent / "output"
            output_dir.mkdir(exist_ok=True)
            output_path = str(output_dir / f"time_segment_comparison_2025_vs_2024_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

        wb.save(output_path)
        print(f"\n报告已保存至: {output_path}")
        return output_path

    def _generate_summary_worksheet(self, wb, data_2025, data_2024):
        """Generate summary comparison worksheet"""
        ws = wb.create_sheet("汇总")

        # Styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        positive_font = Font(color="008000")
        negative_font = Font(color="FF0000")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:I1')
        ws['A1'] = "2025年 vs 2024年 分时段数据对比 - 汇总"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Headers
        headers = [
            "门店", "分时段",
            "2025年翻台率", "2024年翻台率", "翻台率差异",
            "2025年桌数", "2024年桌数", "桌数差异", "桌数同比%"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        row = 4
        store_colors = ["FFE6E6", "E6F3FF", "E6FFE6", "FFFFD0", "FFE6CC", "F0E6FF", "E6FFFF", "F5F5F5"]

        for store_id in sorted(set(list(data_2025.keys()) + list(data_2024.keys()))):
            store_name = self.store_names.get(store_id, f"门店{store_id}")
            store_data_2025 = data_2025.get(store_id, {})
            store_data_2024 = data_2024.get(store_id, {})

            color = store_colors[(store_id - 1) % len(store_colors)]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            start_row = row
            for time_segment in self.time_segments:
                ts_2025 = store_data_2025.get(time_segment, {})
                ts_2024 = store_data_2024.get(time_segment, {})

                turnover_2025 = ts_2025.get('avg_turnover_rate', 0)
                turnover_2024 = ts_2024.get('avg_turnover_rate', 0)
                turnover_diff = turnover_2025 - turnover_2024

                tables_2025 = ts_2025.get('total_tables', 0)
                tables_2024 = ts_2024.get('total_tables', 0)
                tables_diff = tables_2025 - tables_2024
                tables_pct = ((tables_2025 / tables_2024) - 1) * 100 if tables_2024 > 0 else 0

                ws.cell(row=row, column=1, value=store_name if row == start_row else "")
                ws.cell(row=row, column=2, value=time_segment)
                ws.cell(row=row, column=3, value=round(turnover_2025, 2))
                ws.cell(row=row, column=4, value=round(turnover_2024, 2))

                diff_cell = ws.cell(row=row, column=5, value=round(turnover_diff, 2))
                diff_cell.font = positive_font if turnover_diff > 0 else (negative_font if turnover_diff < 0 else Font())

                ws.cell(row=row, column=6, value=tables_2025)
                ws.cell(row=row, column=7, value=tables_2024)

                tables_diff_cell = ws.cell(row=row, column=8, value=tables_diff)
                tables_diff_cell.font = positive_font if tables_diff > 0 else (negative_font if tables_diff < 0 else Font())

                pct_cell = ws.cell(row=row, column=9, value=f"{tables_pct:.1f}%")
                pct_cell.font = positive_font if tables_pct > 0 else (negative_font if tables_pct < 0 else Font())

                # Apply styling
                for col in range(1, 10):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill
                    cell.border = thin_border
                    if col > 2:
                        cell.alignment = Alignment(horizontal='right')

                row += 1

            # Merge store name cells
            if row > start_row + 1:
                ws.merge_cells(f'A{start_row}:A{row-1}')
                ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')

            # Add store totals row
            store_total_turnover_2025 = sum(ts.get('avg_turnover_rate', 0) for ts in store_data_2025.values())
            store_total_turnover_2024 = sum(ts.get('avg_turnover_rate', 0) for ts in store_data_2024.values())
            store_total_tables_2025 = sum(ts.get('total_tables', 0) for ts in store_data_2025.values())
            store_total_tables_2024 = sum(ts.get('total_tables', 0) for ts in store_data_2024.values())

            ws.cell(row=row, column=1, value=f"{store_name}汇总")
            ws.cell(row=row, column=3, value=round(store_total_turnover_2025, 2))
            ws.cell(row=row, column=4, value=round(store_total_turnover_2024, 2))
            ws.cell(row=row, column=5, value=round(store_total_turnover_2025 - store_total_turnover_2024, 2))
            ws.cell(row=row, column=6, value=store_total_tables_2025)
            ws.cell(row=row, column=7, value=store_total_tables_2024)
            ws.cell(row=row, column=8, value=store_total_tables_2025 - store_total_tables_2024)

            total_pct = ((store_total_tables_2025 / store_total_tables_2024) - 1) * 100 if store_total_tables_2024 > 0 else 0
            ws.cell(row=row, column=9, value=f"{total_pct:.1f}%")

            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
                cell.border = thin_border

            row += 1

        # Add 加拿大片区 total section
        row += 1  # Empty row for separation

        # Regional header
        ws.merge_cells(f'A{row}:I{row}')
        ws[f'A{row}'] = "加拿大片区汇总"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1

        # Calculate regional totals by time segment
        regional_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

        for time_segment in self.time_segments:
            # Sum across all stores for this time segment
            turnover_2025_sum = 0
            turnover_2024_sum = 0
            tables_2025_sum = 0
            tables_2024_sum = 0
            store_count_2025 = 0
            store_count_2024 = 0

            for store_id in set(list(data_2025.keys()) + list(data_2024.keys())):
                ts_2025 = data_2025.get(store_id, {}).get(time_segment, {})
                ts_2024 = data_2024.get(store_id, {}).get(time_segment, {})

                if ts_2025.get('avg_turnover_rate', 0) > 0:
                    turnover_2025_sum += ts_2025.get('avg_turnover_rate', 0)
                    store_count_2025 += 1
                if ts_2024.get('avg_turnover_rate', 0) > 0:
                    turnover_2024_sum += ts_2024.get('avg_turnover_rate', 0)
                    store_count_2024 += 1

                tables_2025_sum += ts_2025.get('total_tables', 0)
                tables_2024_sum += ts_2024.get('total_tables', 0)

            # Average turnover (weighted by store count)
            avg_turnover_2025 = turnover_2025_sum / store_count_2025 if store_count_2025 > 0 else 0
            avg_turnover_2024 = turnover_2024_sum / store_count_2024 if store_count_2024 > 0 else 0
            turnover_diff = avg_turnover_2025 - avg_turnover_2024

            tables_diff = tables_2025_sum - tables_2024_sum
            tables_pct = ((tables_2025_sum / tables_2024_sum) - 1) * 100 if tables_2024_sum > 0 else 0

            ws.cell(row=row, column=1, value="加拿大片区")
            ws.cell(row=row, column=2, value=time_segment)
            ws.cell(row=row, column=3, value=round(avg_turnover_2025, 2))
            ws.cell(row=row, column=4, value=round(avg_turnover_2024, 2))

            diff_cell = ws.cell(row=row, column=5, value=round(turnover_diff, 2))
            diff_cell.font = positive_font if turnover_diff > 0 else (negative_font if turnover_diff < 0 else Font())

            ws.cell(row=row, column=6, value=tables_2025_sum)
            ws.cell(row=row, column=7, value=tables_2024_sum)

            tables_diff_cell = ws.cell(row=row, column=8, value=tables_diff)
            tables_diff_cell.font = positive_font if tables_diff > 0 else (negative_font if tables_diff < 0 else Font())

            pct_cell = ws.cell(row=row, column=9, value=f"{tables_pct:.1f}%")
            pct_cell.font = positive_font if tables_pct > 0 else (negative_font if tables_pct < 0 else Font())

            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.fill = regional_fill
                cell.border = thin_border
                if col > 2:
                    cell.alignment = Alignment(horizontal='right')

            row += 1

        # Merge 加拿大片区 cells
        start_regional_row = row - len(self.time_segments)
        if len(self.time_segments) > 1:
            ws.merge_cells(f'A{start_regional_row}:A{row-1}')
            ws[f'A{start_regional_row}'].alignment = Alignment(horizontal='center', vertical='center')

        # Regional grand total row
        total_turnover_2025 = 0
        total_turnover_2024 = 0
        total_tables_2025 = 0
        total_tables_2024 = 0

        for store_id in set(list(data_2025.keys()) + list(data_2024.keys())):
            store_data_2025 = data_2025.get(store_id, {})
            store_data_2024 = data_2024.get(store_id, {})
            total_turnover_2025 += sum(ts.get('avg_turnover_rate', 0) for ts in store_data_2025.values())
            total_turnover_2024 += sum(ts.get('avg_turnover_rate', 0) for ts in store_data_2024.values())
            total_tables_2025 += sum(ts.get('total_tables', 0) for ts in store_data_2025.values())
            total_tables_2024 += sum(ts.get('total_tables', 0) for ts in store_data_2024.values())

        total_turnover_diff = total_turnover_2025 - total_turnover_2024
        total_tables_diff = total_tables_2025 - total_tables_2024
        total_tables_pct = ((total_tables_2025 / total_tables_2024) - 1) * 100 if total_tables_2024 > 0 else 0

        ws.cell(row=row, column=1, value="加拿大片区总计")
        ws.cell(row=row, column=3, value=round(total_turnover_2025, 2))
        ws.cell(row=row, column=4, value=round(total_turnover_2024, 2))
        ws.cell(row=row, column=5, value=round(total_turnover_diff, 2))
        ws.cell(row=row, column=6, value=total_tables_2025)
        ws.cell(row=row, column=7, value=total_tables_2024)
        ws.cell(row=row, column=8, value=total_tables_diff)
        ws.cell(row=row, column=9, value=f"{total_tables_pct:.1f}%")

        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
            cell.border = thin_border

        # Set column widths
        column_widths = [15, 18, 14, 14, 12, 14, 14, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _generate_store_detail_worksheets(self, wb, data_2025, data_2024):
        """Generate detailed worksheets for each store"""
        # Already covered in summary - skip individual store sheets for simplicity
        pass

    def _generate_monthly_comparison_worksheet(self, wb, monthly_2025, monthly_2024):
        """Generate monthly breakdown comparison"""
        ws = wb.create_sheet("月度明细")

        # Styles
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        positive_font = Font(color="008000")
        negative_font = Font(color="FF0000")

        # Chinese month names
        chinese_months = {
            1: "一月", 2: "二月", 3: "三月", 4: "四月",
            5: "五月", 6: "六月", 7: "七月", 8: "八月",
            9: "九月", 10: "十月", 11: "十一月", 12: "十二月"
        }

        # Title
        ws.merge_cells('A1:O1')
        ws['A1'] = "月度分时段对比 - 2025年 vs 2024年"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        row = 3
        months_available = set()
        for store_data in monthly_2025.values():
            months_available.update(store_data.keys())
        for store_data in monthly_2024.values():
            months_available.update(store_data.keys())

        for month in sorted(months_available):
            # Month header
            month_name = chinese_months.get(month, f"{month}月")
            ws.merge_cells(f'A{row}:O{row}')
            ws[f'A{row}'] = f"{month_name}对比"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            row += 1

            # Column headers for this month
            headers = ["门店"]
            for ts in self.time_segments:
                short_ts = ts.replace("(次)", "").replace(":", "")[:8]
                headers.extend([f"{short_ts} 25年", f"{short_ts} 24年", f"{short_ts} 差异"])

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            row += 1

            # Data for each store
            for store_id in sorted(set(list(monthly_2025.keys()) + list(monthly_2024.keys()))):
                store_name = self.store_names.get(store_id, f"门店{store_id}")
                store_2025 = monthly_2025.get(store_id, {}).get(month, {})
                store_2024 = monthly_2024.get(store_id, {}).get(month, {})

                ws.cell(row=row, column=1, value=store_name).border = thin_border

                col = 2
                for ts in self.time_segments:
                    ts_2025 = store_2025.get(ts, {})
                    ts_2024 = store_2024.get(ts, {})

                    turnover_2025 = ts_2025.get('avg_turnover_rate', 0)
                    turnover_2024 = ts_2024.get('avg_turnover_rate', 0)
                    diff = turnover_2025 - turnover_2024

                    ws.cell(row=row, column=col, value=round(turnover_2025, 2)).border = thin_border
                    ws.cell(row=row, column=col+1, value=round(turnover_2024, 2)).border = thin_border

                    diff_cell = ws.cell(row=row, column=col+2, value=round(diff, 2))
                    diff_cell.border = thin_border
                    diff_cell.font = positive_font if diff > 0 else (negative_font if diff < 0 else Font())

                    col += 3

                row += 1

            row += 1  # Space between months

        # Set column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, 15):
            ws.column_dimensions[get_column_letter(col)].width = 10

    def _generate_time_segment_trend_worksheet(self, wb, monthly_2025, monthly_2024):
        """Generate time segment trend analysis across all stores"""
        ws = wb.create_sheet("区域分析")

        # Styles
        header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        positive_font = Font(color="008000")
        negative_font = Font(color="FF0000")

        # Chinese month names
        chinese_months = {
            1: "一月", 2: "二月", 3: "三月", 4: "四月",
            5: "五月", 6: "六月", 7: "七月", 8: "八月",
            9: "九月", 10: "十月", 11: "十一月", 12: "十二月"
        }

        # Title
        ws.merge_cells('A1:N1')
        ws['A1'] = "分时段表现分析 - 全区域汇总"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        row = 3

        # Headers
        headers = ["月份"]
        for ts in self.time_segments:
            short_ts = ts.replace("(次)", "").replace(":", "")[:8]
            headers.extend([f"{short_ts} 25年", f"{short_ts} 24年", f"{short_ts} 差异"])

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        row += 1

        # Aggregate data across all stores by month
        months_available = set()
        for store_data in monthly_2025.values():
            months_available.update(store_data.keys())
        for store_data in monthly_2024.values():
            months_available.update(store_data.keys())

        for month in sorted(months_available):
            month_name = chinese_months.get(month, f"{month}月")
            ws.cell(row=row, column=1, value=month_name).border = thin_border

            col = 2
            for ts in self.time_segments:
                # Aggregate across all stores
                turnover_2025_sum = 0
                turnover_2024_sum = 0
                count_2025 = 0
                count_2024 = 0

                for store_id in set(list(monthly_2025.keys()) + list(monthly_2024.keys())):
                    ts_2025 = monthly_2025.get(store_id, {}).get(month, {}).get(ts, {})
                    ts_2024 = monthly_2024.get(store_id, {}).get(month, {}).get(ts, {})

                    if ts_2025.get('avg_turnover_rate', 0) > 0:
                        turnover_2025_sum += ts_2025.get('avg_turnover_rate', 0)
                        count_2025 += 1
                    if ts_2024.get('avg_turnover_rate', 0) > 0:
                        turnover_2024_sum += ts_2024.get('avg_turnover_rate', 0)
                        count_2024 += 1

                avg_2025 = turnover_2025_sum / count_2025 if count_2025 > 0 else 0
                avg_2024 = turnover_2024_sum / count_2024 if count_2024 > 0 else 0
                diff = avg_2025 - avg_2024

                ws.cell(row=row, column=col, value=round(avg_2025, 2)).border = thin_border
                ws.cell(row=row, column=col+1, value=round(avg_2024, 2)).border = thin_border

                diff_cell = ws.cell(row=row, column=col+2, value=round(diff, 2))
                diff_cell.border = thin_border
                diff_cell.font = positive_font if diff > 0 else (negative_font if diff < 0 else Font())

                col += 3

            row += 1

        # Add yearly totals
        row += 1
        ws.cell(row=row, column=1, value="年度平均").font = Font(bold=True)
        ws.cell(row=row, column=1).border = thin_border

        col = 2
        for ts in self.time_segments:
            turnover_2025_total = 0
            turnover_2024_total = 0
            count_2025 = 0
            count_2024 = 0

            for store_id in set(list(monthly_2025.keys()) + list(monthly_2024.keys())):
                for month in months_available:
                    ts_2025 = monthly_2025.get(store_id, {}).get(month, {}).get(ts, {})
                    ts_2024 = monthly_2024.get(store_id, {}).get(month, {}).get(ts, {})

                    if ts_2025.get('avg_turnover_rate', 0) > 0:
                        turnover_2025_total += ts_2025.get('avg_turnover_rate', 0)
                        count_2025 += 1
                    if ts_2024.get('avg_turnover_rate', 0) > 0:
                        turnover_2024_total += ts_2024.get('avg_turnover_rate', 0)
                        count_2024 += 1

            avg_2025 = turnover_2025_total / count_2025 if count_2025 > 0 else 0
            avg_2024 = turnover_2024_total / count_2024 if count_2024 > 0 else 0
            diff = avg_2025 - avg_2024

            cell_2025 = ws.cell(row=row, column=col, value=round(avg_2025, 2))
            cell_2025.border = thin_border
            cell_2025.font = Font(bold=True)

            cell_2024 = ws.cell(row=row, column=col+1, value=round(avg_2024, 2))
            cell_2024.border = thin_border
            cell_2024.font = Font(bold=True)

            diff_cell = ws.cell(row=row, column=col+2, value=round(diff, 2))
            diff_cell.border = thin_border
            diff_cell.font = Font(bold=True, color="008000" if diff > 0 else ("FF0000" if diff < 0 else "000000"))

            col += 3

        # Set column widths
        ws.column_dimensions['A'].width = 12
        for c in range(2, 14):
            ws.column_dimensions[get_column_letter(c)].width = 10


def main():
    """Main entry point"""
    print("=" * 60)
    print("2025年 vs 2024年 分时段数据对比报告生成器")
    print("=" * 60)

    try:
        # Get database connection
        db_manager = get_database_manager()

        if not db_manager.test_connection():
            print("数据库连接失败")
            return 1

        print("数据库连接成功")

        # Generate report
        generator = YearlyTimeSegmentComparison(db_manager)
        output_path = generator.generate_report()

        print("\n报告生成完成!")
        return 0

    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
