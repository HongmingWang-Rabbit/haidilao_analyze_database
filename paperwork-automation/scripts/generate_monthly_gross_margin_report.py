#!/usr/bin/env python3
"""
Generate Monthly Gross Margin Report (月度毛利相关分析指标)
Generates comprehensive monthly gross margin analysis reports matching the manual file structure:
- 菜品价格变动及菜品损耗表 (Dish Price Changes and Loss Analysis)
- 原材料成本变动表 (Material Cost Changes)
- 打折优惠表 (Discount Analysis)
- 各店毛利率分析 (Store Gross Margin Analysis)
- 月度毛利汇总 (Monthly Gross Margin Summary)
- 同比环比分析 (YoY and MoM Analysis)
"""

import argparse
import sys
import logging
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from utils.database import DatabaseManager, DatabaseConfig
from lib.database_queries import ReportDataProvider
from lib.store_gross_profit_worksheet import StoreGrossProfitWorksheetGenerator
from lib.gross_margin_worksheet import GrossMarginWorksheetGenerator


def setup_logging():
    """Set up logging configuration"""
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )


class MonthlyGrossMarginReportGenerator:
    """Generate comprehensive monthly gross margin report with all analysis sheets"""
    
    def __init__(self, target_date: str):
        """Initialize with target date (should be last day of month)"""
        self.target_date = target_date
        self.target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        self.logger = logging.getLogger(__name__)
        
        # Calculate date ranges
        self.month_start = self.target_dt.replace(day=1)
        # Fix: month_end should be the last day of the month, not the target_dt
        self.month_end = self.month_start + timedelta(days=calendar.monthrange(self.month_start.year, self.month_start.month)[1] - 1)
        
        # Previous month
        if self.month_start.month == 1:
            self.prev_month_start = self.month_start.replace(year=self.month_start.year - 1, month=12)
        else:
            self.prev_month_start = self.month_start.replace(month=self.month_start.month - 1)
        self.prev_month_end = self.prev_month_start + timedelta(days=calendar.monthrange(self.prev_month_start.year, self.prev_month_start.month)[1] - 1)
        
        # Last year same month
        self.last_year_month_start = self.month_start.replace(year=self.month_start.year - 1)
        self.last_year_month_end = self.last_year_month_start + timedelta(days=calendar.monthrange(self.last_year_month_start.year, self.last_year_month_start.month)[1] - 1)
        
    def generate_report(self, db_manager: DatabaseManager, output_path: str = None):
        """Generate the complete monthly gross margin report"""
        
        self.logger.info(f"🍲 Starting monthly gross margin report generation for {self.target_dt.strftime('%Y年%m月')}")
        
        try:
            # Initialize data provider
            data_provider = ReportDataProvider(db_manager)
            
            # Create workbook
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # 1. Generate Dish Price Changes and Loss Analysis (菜品价格变动及菜品损耗表)
            self._generate_dish_price_loss_sheet(wb, data_provider)
            
            # 2. Generate Material Cost Changes (原材料成本变动表)
            self._generate_material_cost_changes_sheet(wb, data_provider)
            
            # 3. Generate Discount Analysis (打折优惠表)
            self._generate_discount_analysis_sheet(wb, data_provider)
            
            # 4. Generate Store Gross Margin Analysis (各店毛利率分析)
            self._generate_store_gross_margin_sheet(wb, data_provider)
            
            # 5. Generate Monthly Summary (月度毛利汇总)
            self._generate_monthly_summary_sheet(wb, data_provider)
            
            # 6. Generate YoY and MoM Analysis (同比环比分析)
            self._generate_yoy_mom_analysis_sheet(wb, data_provider)
            
            # Determine output filename
            if output_path:
                output_file = Path(output_path)
            else:
                output_dir = Path("output") / "monthly_gross_margin"
                output_dir.mkdir(parents=True, exist_ok=True)
                year_month = self.target_dt.strftime('%Y%m')
                output_file = output_dir / f"毛利相关分析指标-{year_month}.xlsx"
            
            # Save workbook
            wb.save(output_file)
            self.logger.info(f"✅ Monthly gross margin report saved to: {output_file}")
            
            # Report summary
            self.logger.info("📋 Report generation completed successfully!")
            self.logger.info(f"📊 Generated worksheets:")
            self.logger.info(f"   - 菜品价格变动及菜品损耗表")
            self.logger.info(f"   - 原材料成本变动表")
            self.logger.info(f"   - 打折优惠表")
            self.logger.info(f"   - 各店毛利率分析")
            self.logger.info(f"   - 月度毛利汇总")
            self.logger.info(f"   - 同比环比分析")
            
            return True
            
        except Exception as e:
            self.logger.error(f"❌ Error generating monthly gross margin report: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _generate_dish_price_loss_sheet(self, wb, data_provider):
        """Generate dish price changes and loss analysis sheet"""
        self.logger.info("📊 Generating dish price changes and loss analysis sheet...")
        
        try:
            # Use existing GrossMarginWorksheetGenerator for this sheet
            worksheet_generator = GrossMarginWorksheetGenerator(self.target_date)
            
            # Get data for the entire month
            dish_price_data = data_provider.get_gross_margin_dish_price_data(self.target_date)
            self.logger.info(f"✅ Retrieved {len(dish_price_data)} dish price records")
            
            # Generate worksheet
            worksheet_generator.generate_detailed_revenue_worksheet(wb, dish_price_data)
            self.logger.info("✅ Dish price changes and loss analysis sheet generated successfully")
            
        except Exception as e:
            self.logger.error(f"❌ Error generating dish price loss sheet: {e}")
            # Create placeholder sheet
            ws = wb.create_sheet("菜品价格变动及菜品损耗表")
            ws.cell(row=1, column=1, value="数据生成错误")
    
    def _generate_material_cost_changes_sheet(self, wb, data_provider):
        """Generate material cost changes analysis sheet"""
        self.logger.info("📊 Generating material cost changes sheet...")
        
        ws = wb.create_sheet("原材料成本变动表")
        
        try:
            # Title
            ws.merge_cells('A1:M1')
            ws['A1'] = f"原材料成本变动表 - {self.target_dt.strftime('%Y年%m月')}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Headers
            headers = [
                "门店名称", "物料编码", "物料名称", "物料类型", "单位",
                "本月均价", "上月均价", "去年同期均价",
                "环比变动率%", "同比变动率%", 
                "本月用量", "环比成本影响金额", "同比成本影响金额"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            
            # Get material cost data
            material_data = self._get_material_cost_analysis_data(data_provider)
            
            # Add data rows
            row = 3
            for material in material_data:
                ws.cell(row=row, column=1, value=material['store_name'])
                ws.cell(row=row, column=2, value=material['material_number'])
                ws.cell(row=row, column=3, value=material['material_name'])
                ws.cell(row=row, column=4, value=material['material_type_id'])
                ws.cell(row=row, column=5, value=material['unit'])
                ws.cell(row=row, column=6, value=material['current_avg_price'])
                ws.cell(row=row, column=7, value=material['prev_avg_price'])
                ws.cell(row=row, column=8, value=material['last_year_avg_price'])
                
                # Calculate changes
                if material['prev_avg_price'] > 0:
                    mom_change = ((material['current_avg_price'] - material['prev_avg_price']) / material['prev_avg_price']) * 100
                else:
                    mom_change = 0
                    
                if material['last_year_avg_price'] > 0:
                    yoy_change = ((material['current_avg_price'] - material['last_year_avg_price']) / material['last_year_avg_price']) * 100
                else:
                    yoy_change = 0
                
                ws.cell(row=row, column=9, value=f"{mom_change:.2f}%")
                ws.cell(row=row, column=10, value=f"{yoy_change:.2f}%")
                ws.cell(row=row, column=11, value=material['current_usage'])
                
                # 环比成本影响金额 (Month-over-Month Cost Impact)
                mom_cost_impact = material['current_usage'] * (material['prev_avg_price'] - material['current_avg_price'])
                mom_cost_impact_cell = ws.cell(row=row, column=12, value=mom_cost_impact)
                
                # Apply conditional formatting for 环比成本影响金额
                if mom_cost_impact == 0:
                    # White background for zero
                    mom_cost_impact_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                elif mom_cost_impact > 0:
                    # Green background for positive (cost savings)
                    mom_cost_impact_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    mom_cost_impact_cell.font = Font(color="006100")
                else:
                    # Red background for negative (cost increases)
                    mom_cost_impact_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    mom_cost_impact_cell.font = Font(color="9C0006")
                
                # 同比成本影响金额 (Year-over-Year Cost Impact)
                yoy_cost_impact = material['current_usage'] * (material['last_year_avg_price'] - material['current_avg_price'])
                yoy_cost_impact_cell = ws.cell(row=row, column=13, value=yoy_cost_impact)
                
                # Apply conditional formatting for 同比成本影响金额
                if yoy_cost_impact == 0:
                    # White background for zero
                    yoy_cost_impact_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                elif yoy_cost_impact > 0:
                    # Green background for positive (cost savings)
                    yoy_cost_impact_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    yoy_cost_impact_cell.font = Font(color="006100")
                else:
                    # Red background for negative (cost increases)
                    yoy_cost_impact_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    yoy_cost_impact_cell.font = Font(color="9C0006")
                
                row += 1
            
            # Apply formatting
            self._apply_basic_formatting(ws, row - 1, 13)
            
            self.logger.info("✅ Material cost changes sheet generated successfully")
            
        except Exception as e:
            self.logger.error(f"❌ Error generating material cost sheet: {e}")
            ws.cell(row=1, column=1, value="数据生成错误")
    
    def _generate_discount_analysis_sheet(self, wb, data_provider):
        """Generate discount analysis sheet"""
        self.logger.info("📊 Generating discount analysis sheet...")
        
        ws = wb.create_sheet("打折优惠表")
        
        try:
            # Title
            ws.merge_cells('A1:I1')
            ws['A1'] = f"打折优惠分析表 - {self.target_dt.strftime('%Y年%m月')}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Headers
            headers = [
                "门店名称", "折扣类型", "本月折扣金额", "上月折扣金额", 
                "去年同期折扣金额", "环比变动", "同比变动",
                "占营业额比例%", "折扣次数"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
            
            # Get discount data
            discount_data = self._get_discount_analysis_data(data_provider)
            
            # Add data rows
            row = 3
            for discount in discount_data:
                ws.cell(row=row, column=1, value=discount['store_name'])
                ws.cell(row=row, column=2, value=discount['discount_type'])
                ws.cell(row=row, column=3, value=discount['current_amount'])
                ws.cell(row=row, column=4, value=discount['prev_amount'])
                ws.cell(row=row, column=5, value=discount['last_year_amount'])
                ws.cell(row=row, column=6, value=discount['current_amount'] - discount['prev_amount'])
                ws.cell(row=row, column=7, value=discount['current_amount'] - discount['last_year_amount'])
                ws.cell(row=row, column=8, value=f"{discount['revenue_ratio']:.2f}%")
                ws.cell(row=row, column=9, value=discount['discount_count'])
                
                row += 1
            
            # Apply formatting
            self._apply_basic_formatting(ws, row - 1, 9)
            
            self.logger.info("✅ Discount analysis sheet generated successfully")
            
        except Exception as e:
            self.logger.error(f"❌ Error generating discount sheet: {e}")
            ws.cell(row=1, column=1, value="数据生成错误")
    
    def _generate_store_gross_margin_sheet(self, wb, data_provider):
        """Generate store gross margin analysis sheet"""
        self.logger.info("📊 Generating store gross margin analysis sheet...")
        
        try:
            # Use existing StoreGrossProfitWorksheetGenerator
            generator = StoreGrossProfitWorksheetGenerator(data_provider)
            generator.generate_worksheet(wb, self.target_date)
            self.logger.info("✅ Store gross margin sheet generated successfully")
            
        except Exception as e:
            self.logger.error(f"❌ Error generating store gross margin sheet: {e}")
            ws = wb.create_sheet("各店毛利率分析")
            ws.cell(row=1, column=1, value="数据生成错误")
    
    def _generate_monthly_summary_sheet(self, wb, data_provider):
        """Generate monthly gross margin summary sheet"""
        self.logger.info("📊 Generating monthly summary sheet...")
        
        ws = wb.create_sheet("月度毛利汇总")
        
        try:
            # Title
            ws.merge_cells('A1:H1')
            ws['A1'] = f"月度毛利汇总表 - {self.target_dt.strftime('%Y年%m月')}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Get summary data for all stores
            summary_data = self._get_monthly_summary_data(data_provider)
            
            # Headers
            headers = [
                "门店名称", "营业收入", "食材成本", "毛利额", "毛利率%",
                "上月毛利率%", "去年同期毛利率%", "环比变动"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            
            # Add data rows
            row = 3
            total_revenue = 0
            total_cost = 0
            
            for store in summary_data:
                ws.cell(row=row, column=1, value=store['store_name'])
                ws.cell(row=row, column=2, value=store['revenue'])
                ws.cell(row=row, column=3, value=store['material_cost'])
                ws.cell(row=row, column=4, value=store['gross_profit'])
                ws.cell(row=row, column=5, value=f"{store['gross_margin']:.2f}%")
                ws.cell(row=row, column=6, value=f"{store['prev_gross_margin']:.2f}%")
                ws.cell(row=row, column=7, value=f"{store['last_year_gross_margin']:.2f}%")
                ws.cell(row=row, column=8, value=f"{store['gross_margin'] - store['prev_gross_margin']:.2f}%")
                
                total_revenue += store['revenue']
                total_cost += store['material_cost']
                
                row += 1
            
            # Add total row
            ws.cell(row=row, column=1, value="合计")
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=total_revenue)
            ws.cell(row=row, column=3, value=total_cost)
            ws.cell(row=row, column=4, value=total_revenue - total_cost)
            if total_revenue > 0:
                ws.cell(row=row, column=5, value=f"{((total_revenue - total_cost) / total_revenue * 100):.2f}%")
            
            # Apply formatting
            self._apply_basic_formatting(ws, row, 8)
            
            # Number formatting for currency columns
            for r in range(3, row + 1):
                for c in [2, 3, 4]:  # Revenue, cost, gross profit columns
                    ws.cell(row=r, column=c).number_format = '#,##0.00'
            
            self.logger.info("✅ Monthly summary sheet generated successfully")
            
        except Exception as e:
            self.logger.error(f"❌ Error generating monthly summary sheet: {e}")
            ws.cell(row=1, column=1, value="数据生成错误")
    
    def _generate_yoy_mom_analysis_sheet(self, wb, data_provider):
        """Generate year-over-year and month-over-month analysis sheet"""
        self.logger.info("📊 Generating YoY/MoM analysis sheet...")
        
        ws = wb.create_sheet("同比环比分析")
        
        try:
            # Title
            ws.merge_cells('A1:K1')
            ws['A1'] = f"同比环比分析表 - {self.target_dt.strftime('%Y年%m月')}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Section 1: Revenue Analysis
            ws['A3'] = "一、营业收入分析"
            ws['A3'].font = Font(bold=True, size=12)
            
            headers = [
                "门店名称", "本月收入", "上月收入", "环比增长额", "环比增长率%",
                "去年同期收入", "同比增长额", "同比增长率%",
                "日均收入", "客单价", "交易笔数"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            
            # Get comparative data
            comparative_data = self._get_comparative_analysis_data(data_provider)
            
            # Add revenue data
            row = 5
            for store in comparative_data:
                ws.cell(row=row, column=1, value=store['store_name'])
                ws.cell(row=row, column=2, value=store['current_revenue'])
                ws.cell(row=row, column=3, value=store['prev_revenue'])
                ws.cell(row=row, column=4, value=store['current_revenue'] - store['prev_revenue'])
                
                if store['prev_revenue'] > 0:
                    mom_rate = ((store['current_revenue'] - store['prev_revenue']) / store['prev_revenue']) * 100
                else:
                    mom_rate = 0
                ws.cell(row=row, column=5, value=f"{mom_rate:.2f}%")
                
                ws.cell(row=row, column=6, value=store['last_year_revenue'])
                ws.cell(row=row, column=7, value=store['current_revenue'] - store['last_year_revenue'])
                
                if store['last_year_revenue'] > 0:
                    yoy_rate = ((store['current_revenue'] - store['last_year_revenue']) / store['last_year_revenue']) * 100
                else:
                    yoy_rate = 0
                ws.cell(row=row, column=8, value=f"{yoy_rate:.2f}%")
                
                # Calculate daily average
                days_in_month = calendar.monthrange(self.target_dt.year, self.target_dt.month)[1]
                daily_avg = store['current_revenue'] / days_in_month
                ws.cell(row=row, column=9, value=daily_avg)
                
                ws.cell(row=row, column=10, value=store.get('avg_ticket', 0))
                ws.cell(row=row, column=11, value=store.get('transaction_count', 0))
                
                row += 1
            
            # Apply formatting to revenue section
            self._apply_basic_formatting(ws, row - 1, 11, start_row=4)
            
            # Section 2: Cost Analysis (starting 3 rows below revenue section)
            row += 2
            ws.cell(row=row, column=1, value="二、成本分析")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            
            row += 1
            cost_headers = [
                "门店名称", "本月食材成本", "上月食材成本", "环比变动",
                "去年同期成本", "同比变动", "成本率%", "主要成本类别", "占比%"
            ]
            
            for col, header in enumerate(cost_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
            
            # Add cost data
            row += 1
            start_row = row
            for store in comparative_data:
                ws.cell(row=row, column=1, value=store['store_name'])
                ws.cell(row=row, column=2, value=store['current_cost'])
                ws.cell(row=row, column=3, value=store['prev_cost'])
                ws.cell(row=row, column=4, value=store['current_cost'] - store['prev_cost'])
                ws.cell(row=row, column=5, value=store['last_year_cost'])
                ws.cell(row=row, column=6, value=store['current_cost'] - store['last_year_cost'])
                
                if store['current_revenue'] > 0:
                    cost_rate = (store['current_cost'] / store['current_revenue']) * 100
                else:
                    cost_rate = 0
                ws.cell(row=row, column=7, value=f"{cost_rate:.2f}%")
                
                ws.cell(row=row, column=8, value=store.get('main_cost_category', ''))
                ws.cell(row=row, column=9, value=f"{store.get('main_cost_ratio', 0):.2f}%")
                
                row += 1
            
            # Apply formatting to cost section
            self._apply_basic_formatting(ws, row - 1, 9, start_row=start_row - 1)
            
            self.logger.info("✅ YoY/MoM analysis sheet generated successfully")
            
        except Exception as e:
            self.logger.error(f"❌ Error generating YoY/MoM sheet: {e}")
            ws.cell(row=1, column=1, value="数据生成错误")
    
    def _apply_basic_formatting(self, ws, max_row, max_col, start_row=2):
        """Apply basic formatting to worksheet"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders
        for row in range(start_row, max_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).border = thin_border
        
        # Auto-adjust column widths
        for col in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Freeze panes
        ws.freeze_panes = f'A{start_row + 1}'
    
    def _get_material_cost_analysis_data(self, data_provider):
        """Get material cost analysis data"""
        # This would query the database for material cost data
        # For now, returning sample structure
        query = """
        WITH current_month AS (
            SELECT 
                s.name as store_name,
                mmu.material_id,
                m.material_number,
                m.name as material_name,
                m.material_type_id,
                m.unit,
                mmu.material_used as total_usage,
                COALESCE(mph.price, 0) as avg_price
            FROM material_monthly_usage mmu
            JOIN material m ON mmu.material_id = m.id
            JOIN store s ON mmu.store_id = s.id
            LEFT JOIN LATERAL (
                SELECT price 
                FROM material_price_history 
                WHERE material_id = mmu.material_id 
                    AND store_id = mmu.store_id 
                    AND is_active = true
                    AND effective_date <= (date_trunc('month', make_date(%s, %s, 1)) + interval '1 month - 1 day')::date
                ORDER BY effective_date DESC
                LIMIT 1
            ) mph ON true
            WHERE mmu.year = %s AND mmu.month = %s
        ),
        prev_month AS (
            SELECT 
                s.name as store_name,
                mmu.material_id,
                COALESCE(mph.price, 0) as avg_price
            FROM material_monthly_usage mmu
            JOIN store s ON mmu.store_id = s.id
            LEFT JOIN LATERAL (
                SELECT price 
                FROM material_price_history 
                WHERE material_id = mmu.material_id 
                    AND store_id = mmu.store_id 
                    AND effective_date = %s
                ORDER BY effective_date DESC
                LIMIT 1
            ) mph ON true
            WHERE mmu.year = %s AND mmu.month = %s
        ),
        last_year AS (
            SELECT 
                s.name as store_name,
                mmu.material_id,
                COALESCE(mph.price, 0) as avg_price
            FROM material_monthly_usage mmu
            JOIN store s ON mmu.store_id = s.id
            LEFT JOIN LATERAL (
                SELECT price 
                FROM material_price_history 
                WHERE material_id = mmu.material_id 
                    AND store_id = mmu.store_id 
                    AND effective_date = %s
                ORDER BY effective_date DESC
                LIMIT 1
            ) mph ON true
            WHERE mmu.year = %s AND mmu.month = %s
        )
        SELECT 
            cm.store_name,
            cm.material_number,
            cm.material_name,
            cm.material_type_id,
            cm.unit,
            cm.avg_price as current_avg_price,
            COALESCE(pm.avg_price, 0) as prev_avg_price,
            COALESCE(ly.avg_price, 0) as last_year_avg_price,
            cm.total_usage as current_usage
        FROM current_month cm
        LEFT JOIN prev_month pm ON cm.store_name = pm.store_name 
            AND cm.material_id = pm.material_id
        LEFT JOIN last_year ly ON cm.store_name = ly.store_name 
            AND cm.material_id = ly.material_id
        ORDER BY cm.store_name, cm.material_type_id, cm.material_number
        """
        
        try:
            result = data_provider.db_manager.fetch_all(query, (
                # Current month: price date, usage year, usage month
                self.target_dt.year, self.target_dt.month,
                self.target_dt.year, self.target_dt.month,
                # Previous month: price date, usage year, usage month (current)
                self.prev_month_end.strftime('%Y-%m-%d'),
                self.target_dt.year, self.target_dt.month,
                # Last year: price date, usage year, usage month (current)
                self.last_year_month_end.strftime('%Y-%m-%d'),
                self.target_dt.year, self.target_dt.month
            ))
            
            return result
        except Exception as e:
            self.logger.error(f"Error getting material cost data: {e}")
            return []
    
    def _get_discount_analysis_data(self, data_provider):
        """Get discount analysis data"""
        year = self.target_dt.year
        month = self.target_dt.month
        
        # Previous month
        prev_month = month - 1
        prev_year = year
        if prev_month < 1:
            prev_month = 12
            prev_year = year - 1
        
        # Last year same month
        last_year = year - 1
        
        query = """
        WITH discount_data AS (
            -- Current month data (aggregate from daily_discount_detail)
            SELECT 
                s.id as store_id,
                s.name as store_name,
                dt.name as discount_type,
                COALESCE(cm.total_discount_amount, 0) as current_amount,
                COALESCE(cm.total_discount_count, 0) as current_count,
                COALESCE(pm.total_discount_amount, 0) as prev_amount,
                COALESCE(pm.total_discount_count, 0) as prev_count,
                COALESCE(ly.total_discount_amount, 0) as last_year_amount,
                COALESCE(ly.total_discount_count, 0) as last_year_count,
                COALESCE(dr.revenue_tax_not_included, 0) as current_revenue
            FROM store s
            CROSS JOIN discount_type dt
            -- Current month (aggregate from daily data)
            LEFT JOIN (
                SELECT 
                    store_id,
                    discount_type_id,
                    SUM(discount_amount) as total_discount_amount,
                    SUM(discount_count) as total_discount_count
                FROM daily_discount_detail
                WHERE EXTRACT(YEAR FROM date) = %s 
                    AND EXTRACT(MONTH FROM date) = %s
                GROUP BY store_id, discount_type_id
            ) cm ON s.id = cm.store_id AND cm.discount_type_id = dt.id
            -- Previous month (aggregate from daily data)
            LEFT JOIN (
                SELECT 
                    store_id,
                    discount_type_id,
                    SUM(discount_amount) as total_discount_amount,
                    SUM(discount_count) as total_discount_count
                FROM daily_discount_detail
                WHERE EXTRACT(YEAR FROM date) = %s 
                    AND EXTRACT(MONTH FROM date) = %s
                GROUP BY store_id, discount_type_id
            ) pm ON s.id = pm.store_id AND pm.discount_type_id = dt.id
            -- Last year same month (aggregate from daily data)
            LEFT JOIN (
                SELECT 
                    store_id,
                    discount_type_id,
                    SUM(discount_amount) as total_discount_amount,
                    SUM(discount_count) as total_discount_count
                FROM daily_discount_detail
                WHERE EXTRACT(YEAR FROM date) = %s 
                    AND EXTRACT(MONTH FROM date) = %s
                GROUP BY store_id, discount_type_id
            ) ly ON s.id = ly.store_id AND ly.discount_type_id = dt.id
            -- Current month revenue
            LEFT JOIN (
                SELECT 
                    store_id,
                    SUM(revenue_tax_not_included) as revenue_tax_not_included
                FROM daily_report
                WHERE EXTRACT(YEAR FROM date) = %s 
                    AND EXTRACT(MONTH FROM date) = %s
                GROUP BY store_id
            ) dr ON s.id = dr.store_id
            WHERE s.id BETWEEN 1 AND 7
                AND (cm.total_discount_amount > 0 
                    OR pm.total_discount_amount > 0 
                    OR ly.total_discount_amount > 0)
        )
        SELECT 
            store_name,
            discount_type,
            current_amount,
            prev_amount,
            last_year_amount,
            current_count as discount_count,
            CASE WHEN current_revenue > 0 
                THEN ROUND((current_amount / current_revenue * 100), 2) 
                ELSE 0 
            END as revenue_ratio
        FROM discount_data
        ORDER BY store_name, discount_type
        """
        
        try:
            results = data_provider.db_manager.fetch_all(query, (
                year, month,  # current
                prev_year, prev_month,  # previous
                last_year, month,  # last year
                year, month  # for revenue
            ))
            
            return [
                {
                    'store_name': row['store_name'],
                    'discount_type': row['discount_type'],
                    'current_amount': row['current_amount'],
                    'prev_amount': row['prev_amount'],
                    'last_year_amount': row['last_year_amount'],
                    'discount_count': row['discount_count'],
                    'revenue_ratio': row['revenue_ratio']
                }
                for row in results
            ]
        except Exception as e:
            self.logger.error(f"Error getting discount data: {e}")
            return []
    
    def _get_monthly_summary_data(self, data_provider):
        """Get monthly summary data for all stores"""
        # Use the same approach as store gross profit but for monthly
        query = """
        WITH current_month AS (
            SELECT 
                s.id as store_id,
                s.name as store_name,
                COALESCE(SUM(dms.sale_amount), 0) as revenue,
                0 as material_cost  -- Will calculate separately
            FROM store s
            LEFT JOIN dish_monthly_sale dms ON s.id = dms.store_id
                AND dms.year = %s AND dms.month = %s
            WHERE s.id BETWEEN 1 AND 7
            GROUP BY s.id, s.name
        ),
        current_costs AS (
            SELECT 
                mmu.store_id,
                SUM(mmu.material_used * COALESCE(mph.price, 0)) as material_cost
            FROM material_monthly_usage mmu
            LEFT JOIN LATERAL (
                SELECT price 
                FROM material_price_history 
                WHERE material_id = mmu.material_id 
                    AND store_id = mmu.store_id 
                    AND is_active = true
                    AND effective_date <= (date_trunc('month', make_date(%s, %s, 1)) + interval '1 month - 1 day')::date
                ORDER BY effective_date DESC
                LIMIT 1
            ) mph ON true
            WHERE mmu.year = %s AND mmu.month = %s
            GROUP BY mmu.store_id
        ),
        prev_month AS (
            SELECT 
                dms.store_id,
                COALESCE(SUM(dms.sale_amount), 0) as revenue
            FROM dish_monthly_sale dms
            WHERE dms.year = %s AND dms.month = %s
            GROUP BY dms.store_id
        ),
        prev_costs AS (
            SELECT 
                mmu.store_id,
                SUM(mmu.material_used * COALESCE(mph.price, 0)) as material_cost
            FROM material_monthly_usage mmu
            LEFT JOIN LATERAL (
                SELECT price 
                FROM material_price_history 
                WHERE material_id = mmu.material_id 
                    AND store_id = mmu.store_id 
                    AND is_active = true
                    AND effective_date <= (date_trunc('month', make_date(%s, %s, 1)) + interval '1 month - 1 day')::date
                ORDER BY effective_date DESC
                LIMIT 1
            ) mph ON true
            WHERE mmu.year = %s AND mmu.month = %s
            GROUP BY mmu.store_id
        ),
        last_year AS (
            SELECT 
                dms.store_id,
                COALESCE(SUM(dms.sale_amount), 0) as revenue
            FROM dish_monthly_sale dms
            WHERE dms.year = %s AND dms.month = %s
            GROUP BY dms.store_id
        ),
        last_year_costs AS (
            SELECT 
                mmu.store_id,
                SUM(mmu.material_used * COALESCE(mph.price, 0)) as material_cost
            FROM material_monthly_usage mmu
            LEFT JOIN LATERAL (
                SELECT price 
                FROM material_price_history 
                WHERE material_id = mmu.material_id 
                    AND store_id = mmu.store_id 
                    AND is_active = true
                    AND effective_date <= (date_trunc('month', make_date(%s, %s, 1)) + interval '1 month - 1 day')::date
                ORDER BY effective_date DESC
                LIMIT 1
            ) mph ON true
            WHERE mmu.year = %s AND mmu.month = %s
            GROUP BY mmu.store_id
        )
        SELECT 
            cm.store_name,
            cm.revenue,
            COALESCE(cc.material_cost, 0) as material_cost,
            cm.revenue - COALESCE(cc.material_cost, 0) as gross_profit,
            CASE WHEN cm.revenue > 0 
                THEN ((cm.revenue - COALESCE(cc.material_cost, 0)) / cm.revenue * 100) 
                ELSE 0 END as gross_margin,
            CASE WHEN COALESCE(pm.revenue, 0) > 0 
                THEN ((COALESCE(pm.revenue, 0) - COALESCE(pc.material_cost, 0)) / COALESCE(pm.revenue, 0) * 100) 
                ELSE 0 END as prev_gross_margin,
            CASE WHEN COALESCE(ly.revenue, 0) > 0 
                THEN ((COALESCE(ly.revenue, 0) - COALESCE(lyc.material_cost, 0)) / COALESCE(ly.revenue, 0) * 100) 
                ELSE 0 END as last_year_gross_margin
        FROM current_month cm
        LEFT JOIN current_costs cc ON cm.store_id = cc.store_id
        LEFT JOIN prev_month pm ON cm.store_id = pm.store_id
        LEFT JOIN prev_costs pc ON cm.store_id = pc.store_id
        LEFT JOIN last_year ly ON cm.store_id = ly.store_id
        LEFT JOIN last_year_costs lyc ON cm.store_id = lyc.store_id
        ORDER BY cm.store_name
        """
        
        try:
            result = data_provider.db_manager.fetch_all(query, (
                # Current month revenue
                self.target_dt.year, self.target_dt.month,
                # Current costs
                self.target_dt.year, self.target_dt.month,  # price date
                self.target_dt.year, self.target_dt.month,  # where clause
                # Previous month revenue
                self.prev_month_start.year, self.prev_month_start.month,
                # Previous costs
                self.prev_month_start.year, self.prev_month_start.month,  # price date
                self.prev_month_start.year, self.prev_month_start.month,  # where clause
                # Last year revenue
                self.last_year_month_start.year, self.last_year_month_start.month,
                # Last year costs
                self.last_year_month_start.year, self.last_year_month_start.month,  # price date
                self.last_year_month_start.year, self.last_year_month_start.month   # where clause
            ))
            
            return result
        except Exception as e:
            self.logger.error(f"Error getting monthly summary data: {e}")
            return []
    
    def _get_comparative_analysis_data(self, data_provider):
        """Get comparative analysis data for YoY/MoM sheet"""
        query = """
        WITH current_month AS (
            SELECT 
                s.id as store_id,
                s.name as store_name,
                COALESCE(SUM(dr.revenue_tax_not_included), 0) as revenue,
                COUNT(DISTINCT dr.date) as days_open,
                COALESCE(SUM(dr.customers), 0) as total_customers,
                COALESCE(AVG(dr.revenue_tax_not_included / NULLIF(dr.customers, 0)), 0) as avg_ticket
            FROM store s
            LEFT JOIN daily_report dr ON s.id = dr.store_id 
                AND dr.date >= %s AND dr.date <= %s
            GROUP BY s.id, s.name
        ),
        prev_month AS (
            SELECT 
                s.id as store_id,
                COALESCE(SUM(dr.revenue_tax_not_included), 0) as revenue
            FROM store s
            LEFT JOIN daily_report dr ON s.id = dr.store_id 
                AND dr.date >= %s AND dr.date <= %s
            GROUP BY s.id
        ),
        last_year AS (
            SELECT 
                s.id as store_id,
                COALESCE(SUM(dr.revenue_tax_not_included), 0) as revenue
            FROM store s
            LEFT JOIN daily_report dr ON s.id = dr.store_id 
                AND dr.date >= %s AND dr.date <= %s
            GROUP BY s.id
        ),
        current_costs AS (
            SELECT 
                mmu.store_id,
                SUM(mmu.material_used * COALESCE(mph.price, 0)) as material_cost
            FROM material_monthly_usage mmu
            LEFT JOIN LATERAL (
                SELECT price 
                FROM material_price_history 
                WHERE material_id = mmu.material_id 
                    AND store_id = mmu.store_id 
                    AND is_active = true
                    AND effective_date <= %s
                ORDER BY effective_date DESC
                LIMIT 1
            ) mph ON true
            WHERE mmu.year = %s AND mmu.month = %s
            GROUP BY mmu.store_id
        ),
        prev_costs AS (
            SELECT 
                mmu.store_id,
                SUM(mmu.material_used * COALESCE(mph.price, 0)) as material_cost
            FROM material_monthly_usage mmu
            LEFT JOIN LATERAL (
                SELECT price 
                FROM material_price_history 
                WHERE material_id = mmu.material_id 
                    AND store_id = mmu.store_id 
                    AND is_active = true
                    AND effective_date <= %s
                ORDER BY effective_date DESC
                LIMIT 1
            ) mph ON true
            WHERE mmu.year = %s AND mmu.month = %s
            GROUP BY mmu.store_id
        ),
        last_year_costs AS (
            SELECT 
                mmu.store_id,
                SUM(mmu.material_used * COALESCE(mph.price, 0)) as material_cost
            FROM material_monthly_usage mmu
            LEFT JOIN LATERAL (
                SELECT price 
                FROM material_price_history 
                WHERE material_id = mmu.material_id 
                    AND store_id = mmu.store_id 
                    AND is_active = true
                    AND effective_date <= %s
                ORDER BY effective_date DESC
                LIMIT 1
            ) mph ON true
            WHERE mmu.year = %s AND mmu.month = %s
            GROUP BY mmu.store_id
        )
        SELECT 
            cm.store_name,
            cm.revenue as current_revenue,
            COALESCE(pm.revenue, 0) as prev_revenue,
            COALESCE(ly.revenue, 0) as last_year_revenue,
            COALESCE(cc.material_cost, 0) as current_cost,
            COALESCE(pc.material_cost, 0) as prev_cost,
            COALESCE(lyc.material_cost, 0) as last_year_cost,
            cm.days_open,
            cm.total_customers,
            cm.avg_ticket
        FROM current_month cm
        LEFT JOIN prev_month pm ON cm.store_id = pm.store_id
        LEFT JOIN last_year ly ON cm.store_id = ly.store_id
        LEFT JOIN current_costs cc ON cm.store_id = cc.store_id
        LEFT JOIN prev_costs pc ON cm.store_id = pc.store_id
        LEFT JOIN last_year_costs lyc ON cm.store_id = lyc.store_id
        ORDER BY cm.store_name
        """
        
        try:
            result = data_provider.db_manager.fetch_all(query, (
                self.month_start, self.month_end,  # current month revenue
                self.prev_month_start, self.prev_month_end,  # prev month revenue
                self.last_year_month_start, self.last_year_month_end,  # last year revenue
                self.target_dt,  # current costs price date
                self.target_dt.year, self.target_dt.month,  # current costs where clause
                self.prev_month_end,  # prev costs price date
                self.prev_month_start.year, self.prev_month_start.month,  # prev costs where clause
                self.last_year_month_end,  # last year costs price date
                self.last_year_month_start.year, self.last_year_month_start.month  # last year costs where clause
            ))
            
            return result
        except Exception as e:
            self.logger.error(f"Error getting comparative data: {e}")
            return []


def main():
    """Main function to handle command line arguments and generate report"""
    parser = argparse.ArgumentParser(
        description='Generate Monthly Gross Margin Report (月度毛利相关分析指标)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python generate_monthly_gross_margin_report.py --target-date 2025-06-30
  python generate_monthly_gross_margin_report.py --target-date 2025-06-30 --output custom_report.xlsx
        '''
    )
    
    parser.add_argument(
        '--target-date',
        type=str,
        required=True,
        help='Target date (last day of month) in YYYY-MM-DD format'
    )
    
    parser.add_argument(
        '--output',
        type=str,
        help='Output file path (optional)'
    )
    
    parser.add_argument(
        '--verbose',
        action='store_true',
        help='Enable verbose logging'
    )
    
    args = parser.parse_args()
    
    # Setup logging
    setup_logging()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Validate target date format
    try:
        target_dt = datetime.strptime(args.target_date, '%Y-%m-%d')
        
        # Check if it's the last day of month
        last_day = calendar.monthrange(target_dt.year, target_dt.month)[1]
        if target_dt.day != last_day:
            print(f"⚠️  Warning: {args.target_date} is not the last day of the month.")
            print(f"   The last day of {target_dt.strftime('%Y-%m')} is {target_dt.year}-{target_dt.month:02d}-{last_day}")
            
    except ValueError:
        print("❌ Error: Invalid date format. Please use YYYY-MM-DD")
        sys.exit(1)
    
    # Initialize database connection
    db_config = DatabaseConfig()
    db_manager = DatabaseManager(db_config)
    
    # Generate report
    generator = MonthlyGrossMarginReportGenerator(args.target_date)
    success = generator.generate_report(db_manager, args.output)
    
    if success:
        print(f"✅ Monthly gross margin report generated successfully!")
        sys.exit(0)
    else:
        print(f"❌ Failed to generate monthly gross margin report")
        sys.exit(1)


if __name__ == "__main__":
    main()