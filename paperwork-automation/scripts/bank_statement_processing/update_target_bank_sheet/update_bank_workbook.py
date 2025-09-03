import sys
import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Set, Tuple
import logging
from pathlib import Path
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import calendar

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../../..')))

from configs.bank_statement.banks import BankBrands
from configs.bank_statement.processing_sheet import BankWorkSheet, BanWorkSheetToFormattedName
from type.bank_processing import BankRecord
from scripts.bank_statement_processing.extract_bank_statements.extract_bank_statements import extract_bank_statements
from scripts.bank_statement_processing.read_target_bank_workbook.read_target_file import read_all_worksheet, get_file_by_datetime
from scripts.bank_statement_processing.update_target_bank_sheet.BMO import append_bmo_records_to_worksheet
from scripts.bank_statement_processing.update_target_bank_sheet.RBC import append_rbc_records_to_worksheet
from scripts.bank_statement_processing.update_target_bank_sheet.CIBC import append_cibc_records_to_worksheet

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def get_month_date_range(current_date: datetime) -> Tuple[datetime, datetime]:
    """
    Get the start and end date for the current month.
    
    Args:
        current_date: The current date
        
    Returns:
        Tuple of (month_start, month_end) datetime objects
    """
    year = current_date.year
    month = current_date.month
    
    # First day of the month
    month_start = datetime(year, month, 1)
    
    # Last day of the month
    last_day = calendar.monthrange(year, month)[1]
    month_end = datetime(year, month, last_day, 23, 59, 59)
    
    return month_start, month_end

def create_record_key(record: BankRecord) -> str:
    """
    Create a unique key for a bank record for comparison.
    
    Args:
        record: BankRecord object
        
    Returns:
        String key for the record
    """
    # Use date, amount, and description to create a unique key
    date_str = record.date.strftime('%Y-%m-%d') if record.date else ''
    desc = record.short_desctiption or record.full_desctiption or ''
    # Include both debit and credit to handle the sign differences
    return f"{date_str}|{desc}|{abs(record.debit):.2f}|{abs(record.credit):.2f}"

def filter_records_by_month(records: List[BankRecord], month_start: datetime, month_end: datetime) -> List[BankRecord]:
    """
    Filter records to only include those within the specified month.
    
    Args:
        records: List of BankRecord objects
        month_start: Start of the month
        month_end: End of the month
        
    Returns:
        Filtered list of BankRecord objects
    """
    filtered = []
    for record in records:
        if record.date and month_start <= record.date <= month_end:
            filtered.append(record)
    return filtered

def find_new_records(
    extracted_records: List[BankRecord], 
    existing_records: List[BankRecord]
) -> List[BankRecord]:
    """
    Find records that exist in extracted but not in existing.
    
    Args:
        extracted_records: Records extracted from bank statements
        existing_records: Records already in the workbook
        
    Returns:
        List of new records to be added
    """
    # Create sets of record keys for comparison
    existing_keys = set(create_record_key(record) for record in existing_records)
    
    new_records = []
    for record in extracted_records:
        record_key = create_record_key(record)
        if record_key not in existing_keys:
            new_records.append(record)
            logger.debug(f"New record found: {record_key}")
    
    return new_records

def get_bank_brand_from_account(account_id: str) -> BankBrands:
    """
    Extract bank brand from account ID.
    
    Args:
        account_id: Account identifier (e.g., BMO3817, RBC5419)
        
    Returns:
        BankBrands enum value
    """
    if account_id.startswith('BMO'):
        return BankBrands.BMO
    elif account_id.startswith('RBC'):
        return BankBrands.RBC
    elif account_id.startswith('CIBC'):
        return BankBrands.CIBC
    else:
        raise ValueError(f"Unknown bank brand for account: {account_id}")

# Bank-specific append functions are now imported from separate modules

def append_records_to_worksheet(wb, sheet_name: str, new_records: List[BankRecord]):
    """
    Append new records to a worksheet using the appropriate bank-specific function.
    
    Args:
        wb: Openpyxl workbook object
        sheet_name: Name of the worksheet
        new_records: List of BankRecord objects to append
    """
    if sheet_name not in BankWorkSheet:
        logger.warning(f"Sheet {sheet_name} not in BankWorkSheet mapping")
        return
    
    bank_brand = BankWorkSheet[sheet_name]
    
    if bank_brand == BankBrands.BMO:
        append_bmo_records_to_worksheet(wb, sheet_name, new_records)
    elif bank_brand == BankBrands.RBC:
        append_rbc_records_to_worksheet(wb, sheet_name, new_records)
    elif bank_brand == BankBrands.CIBC:
        append_cibc_records_to_worksheet(wb, sheet_name, new_records)
    else:
        logger.warning(f"Unsupported bank brand: {bank_brand}")

def update_bank_workbook(current_date: datetime, output_folder: str = None):
    """
    Main function to update bank workbook with new records.
    
    Args:
        current_date: The current date for processing
        output_folder: Optional output folder path. If not provided, uses default.
    """
    # Get month range
    month_start, month_end = get_month_date_range(current_date)
    logger.info(f"Processing records for month: {month_start.strftime('%Y-%m')}")
    
    # Get the source workbook file
    source_file = get_file_by_datetime(current_date)
    if not source_file:
        logger.error(f"No workbook file found for {current_date.strftime('%Y-%m')}")
        return
    
    # Determine output folder
    if not output_folder:
        # Default output folder
        base_path = Path(__file__).parent.parent.parent.parent
        output_folder = base_path / "output" / "bank_statements" / current_date.strftime("%Y-%m")
    else:
        output_folder = Path(output_folder)
    
    # Create output folder if it doesn't exist
    output_folder.mkdir(parents=True, exist_ok=True)
    
    # Copy workbook to output folder
    source_path = Path(source_file)
    output_file = output_folder / f"Updated_{source_path.name}"
    
    logger.info(f"Copying workbook to: {output_file}")
    shutil.copy2(source_file, output_file)
    
    # Read existing records from workbook
    logger.info("Reading existing records from workbook...")
    existing_records_by_account = read_all_worksheet(current_date)
    
    # Extract bank statements for the current month
    logger.info("Extracting bank statements...")
    extracted_records_dict = extract_bank_statements(current_date)
    
    # Convert extracted records to account-based format
    extracted_by_account = {}
    for bank_brand, records in extracted_records_dict.items():
        # Group by account ID
        for record in records:
            if record.serial_number:
                account_id = record.serial_number.rsplit('_', 1)[0]
                if account_id not in extracted_by_account:
                    extracted_by_account[account_id] = []
                extracted_by_account[account_id].append(record)
    
    # Load workbook for updating
    logger.info(f"Loading workbook for updates: {output_file}")
    wb = load_workbook(output_file)
    
    # Process each account
    updates_made = False
    for sheet_name, bank_brand in BankWorkSheet.items():
        # Use the predefined mapping to get the account ID for this sheet
        expected_account_id = BanWorkSheetToFormattedName.get(sheet_name)
        
        if not expected_account_id:
            logger.debug(f"No account mapping found for sheet {sheet_name}")
            continue
        
        # Get existing and extracted records for this account
        existing = existing_records_by_account.get(expected_account_id, [])
        extracted = extracted_by_account.get(expected_account_id, [])
        
        # Skip if no extracted records for this account
        if not extracted:
            logger.debug(f"No extracted records for account {expected_account_id}")
            continue
        
        # Filter records to current month only
        existing_month = filter_records_by_month(existing, month_start, month_end)
        extracted_month = filter_records_by_month(extracted, month_start, month_end)
        
        logger.info(f"Sheet {sheet_name} / Account {expected_account_id}:")
        logger.info(f"  Existing records in month: {len(existing_month)}")
        logger.info(f"  Extracted records in month: {len(extracted_month)}")
        
        # Find new records
        new_records = find_new_records(extracted_month, existing_month)
        
        if new_records:
            logger.info(f"  Found {len(new_records)} new records to append")
            append_records_to_worksheet(wb, sheet_name, new_records)
            updates_made = True
        else:
            logger.info(f"  No new records to append")
    
    # Save the updated workbook
    if updates_made:
        logger.info(f"Saving updated workbook: {output_file}")
        wb.save(output_file)
        logger.info("Update complete!")
    else:
        logger.info("No updates needed - workbook is already up to date")
        wb.close()

def test_update():
    """Test function to update workbook for August 2025."""
    test_date = datetime(2025, 8, 28)  # Mid-August to ensure we're in the right month
    
    print(f"\nTesting update_bank_workbook for {test_date.strftime('%Y-%m')}")
    print("="*60)
    
    update_bank_workbook(test_date)

if __name__ == "__main__":
    test_update()