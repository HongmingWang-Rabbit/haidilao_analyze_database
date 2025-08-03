#!/usr/bin/env python3
"""
Generate monthly beverage report - MINIMAL VERSION
"""

import os
import sys
from pathlib import Path

# Add parent directory to path for imports FIRST
sys.path.insert(0, str(Path(__file__).parent.parent))

print("Starting minimal beverage report generator...")

from utils.database import DatabaseConfig, DatabaseManager
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import argparse

def main():
    """Main function"""
    parser = argparse.ArgumentParser(description="Generate monthly beverage report")
    parser.add_argument("--date", default="2025-07-31", help="Target date (YYYY-MM-DD)")
    args = parser.parse_args()
    
    print(f"Generating beverage report for date: {args.date}")
    
    try:
        config = DatabaseConfig(is_test=False)
        db_manager = DatabaseManager(config)
        
        target_dt = datetime.strptime(args.date, '%Y-%m-%d')
        year, month = target_dt.year, target_dt.month
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Beverage Report"
        
        ws['A1'] = f"Monthly Beverage Report - {year}-{month:02d}"
        ws['A1'].font = Font(bold=True, size=14)
        
        output_dir = Path('./output')
        output_dir.mkdir(exist_ok=True)
        filename = f"monthly_beverage_report_minimal_{args.date.replace('-', '_')}.xlsx"
        output_path = output_dir / filename
        
        wb.save(output_path)
        print(f"SUCCESS: Beverage report saved to {output_path}")
        sys.exit(0)  # Success
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()