#!/usr/bin/env python3
"""
Final test of the complete monthly material report with enhanced dish usage details
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from scripts.generate_monthly_material_report import MonthlyMaterialReportGenerator
from datetime import datetime
from openpyxl import load_workbook

def test_final_enhanced_report():
    """Final test of the complete monthly material report with all enhancements"""
    
    print("Final Enhanced Report Test - All Features")
    print("=" * 50)
    
    try:
        # Generate the monthly material report with unique timestamp
        target_date = "2025-05-31"
        generator = MonthlyMaterialReportGenerator(target_date, is_test=True)
        
        # Override save method to use unique filename
        def save_with_timestamp(wb):
            try:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"FINAL_FULLY_ENHANCED_monthly_material_report_{timestamp}.xlsx"
                output_path = generator.output_dir / filename
                wb.save(output_path)
                return output_path
            except Exception as e:
                print(f"ERROR: Error saving report: {e}")
                return None
        
        generator.save_report = save_with_timestamp
        
        print(f"Target date: {target_date}")
        print("Generating final fully enhanced monthly material report...")
        
        # Generate the report
        output_file = generator.generate_report()
        
        if output_file and Path(output_file).exists():
            print(f"✅ Report generated successfully!")
            print(f"📁 Output file: {output_file}")
            print(f"📊 File size: {Path(output_file).stat().st_size:,} bytes")
            
            # Load and verify the report
            wb = load_workbook(output_file)
            
            if "物料用量差异分析" in wb.sheetnames:
                ws = wb["物料用量差异分析"]
                print(f"📊 Analyzing 物料用量差异分析 worksheet...")
                
                # Find header row
                header_row = None
                for row in range(1, 20):
                    cell_value = ws.cell(row=row, column=1).value
                    if cell_value and "序号" in str(cell_value):
                        header_row = row
                        break
                
                if header_row:
                    print(f"\n🎯 FINAL COMPREHENSIVE VERIFICATION:")
                    
                    # 1. Enhanced dish usage details verification
                    print(f"\n1️⃣ Enhanced Dish Usage Details:")
                    
                    # Check for specific material 3000759
                    found_enhanced_example = False
                    for row in range(header_row + 1, min(header_row + 100, ws.max_row + 1)):
                        material_num_cell = ws.cell(row=row, column=4)
                        if material_num_cell.value and "3000759" in str(material_num_cell.value):
                            usage_cell = ws.cell(row=row, column=6)
                            store_cell = ws.cell(row=row, column=2)
                            
                            if usage_cell.value and "materials_use-" in str(usage_cell.value):
                                found_enhanced_example = True
                                print(f"   ✅ Enhanced format confirmed for 清油底料 at {store_cell.value}")
                                
                                # Parse first line
                                first_line = str(usage_cell.value).split('\n')[0]
                                print(f"   📄 Example: {first_line}")
                                
                                # Check components
                                components = ['sale-', '出品分量(kg)-', '损耗-', '物料单位-', 'materials_use-']
                                all_present = all(comp in first_line for comp in components)
                                print(f"   📋 All components present: {'✅' if all_present else '❌'}")
                                break
                    
                    if not found_enhanced_example:
                        print(f"   ⚠️ Enhanced format example not found")
                    
                    # 2. Column structure verification
                    expected_headers = [
                        "序号", "门店", "物料名称", "物料号", "单位", "使用菜品详情",
                        "理论用量", "套餐用量", "系统记录", "库存盘点", "减去盘点用量", "差异数量", "差异率(%)", "状态",
                        "本月总消费金额", "差异金额"
                    ]
                    
                    print(f"\n2️⃣ Column Structure (16 columns A-P):")
                    header_check = ws.max_column == 16
                    print(f"   ✅ Column count: {'PASS' if header_check else 'FAIL'} ({ws.max_column}/16)")
                    
                    # 3. Status verification (不足 → 少用)
                    print(f"\n3️⃣ Status Text (不足 → 少用):")
                    status_counts = {"超量": 0, "少用": 0, "正常": 0, "不足": 0}
                    for row in range(header_row + 1, ws.max_row + 1):
                        status_cell = ws.cell(row=row, column=14)
                        status = status_cell.value if status_cell.value else ""
                        if status in status_counts:
                            status_counts[status] += 1
                    
                    status_check = status_counts["不足"] == 0 and status_counts["少用"] > 0
                    print(f"   ✅ Status text: {'PASS' if status_check else 'FAIL'}")
                    print(f"   📊 Distribution: 超量({status_counts['超量']}), 少用({status_counts['少用']}), 正常({status_counts['正常']})")
                    
                    # 4. Formula verification
                    print(f"\n4️⃣ Enhanced Formulas:")
                    test_row = header_row + 1
                    
                    # Check minus inventory formula (K = I - J)
                    minus_inventory_cell = ws.cell(row=test_row, column=11)
                    minus_inventory_formula = str(minus_inventory_cell.value)
                    minus_inventory_check = "=I" in minus_inventory_formula and "-J" in minus_inventory_formula
                    print(f"   ✅ 减去盘点用量 (K): {'PASS' if minus_inventory_check else 'FAIL'}")
                    
                    # Check variance formula (L = I - (G+H+J))
                    variance_cell = ws.cell(row=test_row, column=12)
                    variance_formula = str(variance_cell.value)
                    variance_check = "=I" in variance_formula and "-(G" in variance_formula
                    print(f"   ✅ 差异数量 (L): {'PASS' if variance_check else 'FAIL'}")
                    
                    # Check cost formula (O = (G+H+I)*price)
                    cost_cell = ws.cell(row=test_row, column=15)
                    cost_formula = str(cost_cell.value)
                    cost_check = "=(G" in cost_formula and "+H" in cost_formula and "+I" in cost_formula
                    print(f"   ✅ 本月总消费金额 (O): {'PASS' if cost_check else 'FAIL'}")
                    
                    # 5. Color verification
                    print(f"\n5️⃣ Status Colors:")
                    sample_colors = {"超量": None, "少用": None, "正常": None}
                    for row in range(header_row + 1, min(header_row + 21, ws.max_row + 1)):
                        status_cell = ws.cell(row=row, column=14)
                        status = status_cell.value
                        if status in sample_colors and sample_colors[status] is None:
                            color = status_cell.fill.start_color.rgb if status_cell.fill and status_cell.fill.start_color else "None"
                            sample_colors[status] = color
                    
                    color_checks = {
                        "超量": sample_colors["超量"] == "00FFE6E6",
                        "少用": sample_colors["少用"] == "00E8F5E8", 
                        "正常": sample_colors["正常"] == "00000000" or sample_colors["正常"] == "None"
                    }
                    
                    for status, is_correct in color_checks.items():
                        print(f"   ✅ {status} color: {'PASS' if is_correct else 'FAIL'}")
                    
                    # Overall assessment
                    all_checks = [
                        found_enhanced_example,
                        header_check,
                        status_check,
                        minus_inventory_check,
                        variance_check, 
                        cost_check,
                        all(color_checks.values())
                    ]
                    
                    overall_pass = all(all_checks)
                    
                    print(f"\n🏆 FINAL ASSESSMENT: {'🎉 ALL TESTS PASSED! 🎉' if overall_pass else '❌ SOME TESTS FAILED'}")
                    
                    if overall_pass:
                        print(f"\n🌟 COMPLETE FEATURE SET:")
                        print(f"   • 📊 16 comprehensive columns (A-P)")
                        print(f"   • 📋 Enhanced dish usage details with:")
                        print(f"     - Sale quantities")
                        print(f"     - 出品分量 (portion size)")  
                        print(f"     - 损耗 (loss rate)")
                        print(f"     - 物料单位 (material unit)")
                        print(f"     - materials_use (calculated usage)")
                        print(f"   • ➖ 减去盘点用量: System - Inventory calculation")
                        print(f"   • 💰 Cost analysis: Monthly total & variance cost")
                        print(f"   • 🎨 Smart status coloring: 超量(red), 少用(green), 正常(white)")
                        print(f"   • 📝 Updated terminology: 不足 → 少用")
                        print(f"   • ✅ All formulas and calculations working perfectly!")
                        
                    return overall_pass
                        
            wb.close()
        else:
            print("❌ Report generation failed")
            return False
            
    except Exception as e:
        print(f"❌ Error during final testing: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_final_enhanced_report()
    if success:
        print("\n🎊🎊🎊 FINAL ENHANCEMENT COMPLETE - ALL FEATURES WORKING PERFECTLY! 🎊🎊🎊")
    else:
        print("\n❌ Final enhancement testing failed")
        sys.exit(1)