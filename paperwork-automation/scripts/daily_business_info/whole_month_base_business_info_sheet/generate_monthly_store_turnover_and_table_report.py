#!/usr/bin/env python3
"""
Generate monthly business information report for all stores.

Each store gets its own sheet with:
- 日期 (Date)
- 星期 (Day of week, e.g., 星期一)
- 翻台率 (Turnover rate)
- 当日桌数 (Validated table count today)
"""

import argparse
import sys
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import calendar

# Add project root to path
project_root = Path(__file__).parent.parent.parent.parent
sys.path.insert(0, str(project_root))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from utils.database import DatabaseManager, DatabaseConfig
from lib.config import STORE_ID_TO_NAME_MAPPING, STORE_NAME_MAPPING


class MonthlyStoreReportGenerator:
    """Generate monthly business information report for all stores."""

    # Day of week mapping
    WEEKDAY_NAMES = {
        0: '星期一',
        1: '星期二',
        2: '星期三',
        3: '星期四',
        4: '星期五',
        5: '星期六',
        6: '星期日'
    }

    def __init__(self, db_manager: DatabaseManager, year: int, month: int):
        """
        Initialize the generator.

        Args:
            db_manager: Database manager instance
            year: Target year (e.g., 2025)
            month: Target month (1-12)
        """
        self.db_manager = db_manager
        self.year = year
        self.month = month

        # Calculate month start and end dates
        self.month_start = datetime(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        self.month_end = datetime(year, month, last_day)

        # Setup styles
        self.setup_styles()

    def setup_styles(self):
        """Setup common openpyxl styles."""
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.header_font = Font(bold=True, size=12)
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def apply_header_style(self, cell):
        """Apply standard header styling to a cell."""
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = self.header_alignment
        cell.border = self.thin_border

    def set_column_widths(self, ws, widths: List[int]):
        """Set column widths for worksheet."""
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def fetch_store_data(self, store_id: int) -> List[Dict]:
        """
        Fetch daily business data for a store for the entire month.

        Args:
            store_id: Store ID

        Returns:
            List of daily data records
        """
        query = """
            SELECT
                date,
                turnover_rate,
                tables_served_validated
            FROM daily_report
            WHERE store_id = %s
                AND date >= %s
                AND date <= %s
            ORDER BY date
        """

        try:
            results = self.db_manager.fetch_all(
                query,
                (store_id, self.month_start.date(), self.month_end.date())
            )
            return results if results else []
        except Exception as e:
            print(f"Error fetching data for store {store_id}: {e}")
            return []

    def generate_store_worksheet(self, workbook: Workbook, store_id: int, store_name: str):
        """
        Generate worksheet for a single store.

        Args:
            workbook: Excel workbook object
            store_id: Store ID
            store_name: Store name
        """
        # Create worksheet with store name
        ws = workbook.create_sheet(store_name)

        # Set column widths
        self.set_column_widths(ws, [15, 12, 15, 20])

        # Write header
        headers = ['日期', '星期', '翻台率', '当日桌数']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            self.apply_header_style(cell)

        # Fetch data
        data_records = self.fetch_store_data(store_id)
        data_by_date = {record['date']: record for record in data_records}

        # Write data for each day in the month
        current_date = self.month_start
        row_idx = 2

        while current_date <= self.month_end:
            date_obj = current_date.date()

            # Column A: Date
            ws.cell(row=row_idx, column=1, value=date_obj.strftime('%Y-%m-%d'))

            # Column B: Day of week
            weekday_name = self.WEEKDAY_NAMES[date_obj.weekday()]
            ws.cell(row=row_idx, column=2, value=weekday_name)

            # Fetch data for this date
            if date_obj in data_by_date:
                record = data_by_date[date_obj]

                # Column C: Turnover rate
                turnover_rate = record.get('turnover_rate')
                if turnover_rate is not None:
                    ws.cell(row=row_idx, column=3, value=float(turnover_rate))
                    ws.cell(row=row_idx, column=3).number_format = '0.00'
                else:
                    ws.cell(row=row_idx, column=3, value='N/A')

                # Column D: Validated table count
                table_count = record.get('tables_served_validated')
                if table_count is not None:
                    ws.cell(row=row_idx, column=4, value=int(table_count))
                else:
                    ws.cell(row=row_idx, column=4, value='N/A')
            else:
                # No data for this date
                ws.cell(row=row_idx, column=3, value='N/A')
                ws.cell(row=row_idx, column=4, value='N/A')

            # Apply alignment
            for col_idx in range(1, 5):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    horizontal='center',
                    vertical='center'
                )

            current_date += timedelta(days=1)
            row_idx += 1

    def generate_report(self, output_path: str) -> bool:
        """
        Generate the complete monthly report for all stores.

        Args:
            output_path: Path to save the Excel file

        Returns:
            True if successful, False otherwise
        """
        try:
            workbook = Workbook()
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])

            # Generate worksheet for each store (IDs 1-7)
            for store_id in range(1, 8):
                store_name = STORE_ID_TO_NAME_MAPPING.get(store_id, f"Store_{store_id}")
                print(f"Generating worksheet for {store_name}...")
                self.generate_store_worksheet(workbook, store_id, store_name)

            # Save workbook
            workbook.save(output_path)
            print(f"\nReport generated successfully: {output_path}")
            return True

        except Exception as e:
            print(f"Error generating report: {e}")
            import traceback
            traceback.print_exc()
            return False


def main():
    """Main execution function."""
    parser = argparse.ArgumentParser(
        description='Generate monthly business information report for all stores'
    )
    parser.add_argument(
        '--year',
        type=int,
        required=True,
        help='Target year (e.g., 2025)'
    )
    parser.add_argument(
        '--month',
        type=int,
        required=True,
        help='Target month (1-12)'
    )
    parser.add_argument(
        '--output',
        help='Output file path (optional, defaults to auto-generated name)'
    )

    args = parser.parse_args()

    # Validate month
    if not 1 <= args.month <= 12:
        print("Error: Month must be between 1 and 12")
        sys.exit(1)

    # Generate default output path if not provided
    if args.output:
        output_path = args.output
    else:
        year_month = f"{args.year}-{args.month:02d}"
        output_dir = project_root / 'output'
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / f'monthly_store_report_{year_month}.xlsx'

    print(f"Generating monthly report for {args.year}年{args.month:02d}月...")
    print(f"Output: {output_path}\n")

    # Initialize database manager
    db_config = DatabaseConfig()
    db_manager = DatabaseManager(db_config)

    try:
        # Generate report
        generator = MonthlyStoreReportGenerator(db_manager, args.year, args.month)
        success = generator.generate_report(str(output_path))

        sys.exit(0 if success else 1)

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
