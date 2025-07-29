#!/usr/bin/env python3
"""
Verify the monthly gross margin report generation and compare with manual file
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

def verify_report_generation():
    """Verify the report can be generated and check its contents"""
    import subprocess
    import os
    from datetime import datetime
    
    print("=== MONTHLY GROSS MARGIN REPORT VERIFICATION ===\n")
    
    # 1. First, let's generate the report for May 2025
    print("1. Generating monthly gross margin report for 2025-05-31...")
    
    cmd = [
        sys.executable,
        "scripts/generate_monthly_gross_margin_report.py",
        "--target-date", "2025-05-31"
    ]
    
    # Set up environment
    env = os.environ.copy()
    env['PYTHONPATH'] = str(Path.cwd())
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, env=env)
        
        if result.returncode == 0:
            print("✅ Report generated successfully!")
            print("\nOutput:")
            print(result.stdout)
        else:
            print("❌ Report generation failed!")
            print("\nError output:")
            print(result.stderr)
            return False
            
    except Exception as e:
        print(f"❌ Error running report generation: {e}")
        return False
    
    # 2. Check if the output file exists
    output_file = Path("output/monthly_gross_margin/毛利相关分析指标-202505.xlsx")
    if output_file.exists():
        print(f"\n✅ Output file created: {output_file}")
        print(f"   File size: {output_file.stat().st_size:,} bytes")
        print(f"   Modified: {datetime.fromtimestamp(output_file.stat().st_mtime)}")
    else:
        print(f"\n❌ Output file not found: {output_file}")
        return False
    
    # 3. Try to read and analyze the generated file
    print("\n2. Analyzing generated report structure...")
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(output_file, read_only=True, data_only=True)
        
        print(f"\nSheets in generated report ({len(wb.sheetnames)}):")
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"  {i}. {sheet_name}")
            
        # Check each sheet for data
        print("\nChecking sheet contents:")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Count non-empty rows (excluding headers)
            data_rows = 0
            for row_idx in range(4, ws.max_row + 1):  # Start from row 4 (after headers)
                has_data = False
                for col_idx in range(1, min(5, ws.max_column + 1)):
                    if ws.cell(row=row_idx, column=col_idx).value:
                        has_data = True
                        break
                if has_data:
                    data_rows += 1
            
            print(f"\n  {sheet_name}:")
            print(f"    - Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            print(f"    - Data rows: {data_rows}")
            
            # Show sample data from first few rows
            if data_rows > 0:
                print("    - Sample data (first data row):")
                for col_idx in range(1, min(6, ws.max_column + 1)):
                    header = ws.cell(row=3, column=col_idx).value or ws.cell(row=2, column=col_idx).value or f"Col{col_idx}"
                    value = ws.cell(row=4, column=col_idx).value
                    if value:
                        print(f"      {header}: {value}")
        
        wb.close()
        
    except ImportError:
        print("⚠️  openpyxl not available, skipping detailed analysis")
    except Exception as e:
        print(f"⚠️  Error analyzing report: {e}")
    
    # 4. Check the manual file for comparison
    print("\n3. Checking manual file for comparison...")
    manual_file = Path("data/dishes_related/附件3-毛利相关分析指标-2505.xlsx")
    
    if manual_file.exists():
        print(f"✅ Manual file found: {manual_file}")
        print(f"   File size: {manual_file.stat().st_size:,} bytes")
        
        try:
            from openpyxl import load_workbook
            
            wb_manual = load_workbook(manual_file, read_only=True, data_only=True)
            
            print(f"\nSheets in manual file ({len(wb_manual.sheetnames)}):")
            for i, sheet_name in enumerate(wb_manual.sheetnames, 1):
                print(f"  {i}. {sheet_name}")
                
            wb_manual.close()
            
        except:
            pass
    else:
        print(f"❌ Manual file not found: {manual_file}")
    
    # 5. Summary
    print("\n" + "="*50)
    print("VERIFICATION SUMMARY:")
    print("="*50)
    
    if output_file.exists():
        print("✅ Report generation: SUCCESS")
        print("✅ Output file created")
        print("\n📋 Next steps:")
        print("1. Open both files in Excel to compare:")
        print(f"   - Generated: {output_file}")
        print(f"   - Manual: {manual_file}")
        print("2. Check if data values match between files")
        print("3. Verify calculations are correct")
    else:
        print("❌ Report generation: FAILED")
        print("Please check the error messages above")
    
    return True

def check_database_data():
    """Quick check of database data for May 2025"""
    print("\n\n=== DATABASE DATA CHECK ===\n")
    
    # Create a simple SQL script to check data
    sql_script = """
-- Check May 2025 data availability
SELECT 'dish_monthly_sale' as table_name, COUNT(*) as records, SUM(sale_amount) as total_amount
FROM dish_monthly_sale WHERE year = 2025 AND month = 5
UNION ALL
SELECT 'material_monthly_usage', COUNT(*), SUM(material_used)
FROM material_monthly_usage WHERE year = 2025 AND month = 5
UNION ALL
SELECT 'material_price_history May', COUNT(*), AVG(price)
FROM material_price_history WHERE effective_date BETWEEN '2025-05-01' AND '2025-05-31'
UNION ALL
SELECT 'dish_price_history May', COUNT(*), AVG(price)
FROM dish_price_history WHERE effective_date BETWEEN '2025-05-01' AND '2025-05-31';
"""
    
    # Save SQL script
    sql_file = Path("check_may_data.sql")
    sql_file.write_text(sql_script)
    print(f"SQL script saved to: {sql_file}")
    print("\nTo check database data, run:")
    print(f"psql -U hongming -d haidilao-paperwork -f {sql_file}")
    
    # Clean up
    sql_file.unlink()

if __name__ == "__main__":
    print("Starting Monthly Gross Margin Report Verification")
    print("="*60)
    
    # Run verification
    verify_report_generation()
    
    # Check database
    check_database_data()
    
    print("\n✅ Verification complete!")