#!/usr/bin/env python3
"""
Generate monthly material report - MINIMAL VERSION for debugging
"""

import os
import sys
from pathlib import Path

# Add parent directory to path for imports FIRST
sys.path.insert(0, str(Path(__file__).parent.parent))

print("Starting minimal material report generator...")

# Now import our modules after path is set
from utils.database import DatabaseConfig, DatabaseManager
from openpyxl import Workbook
from datetime import datetime
import argparse

def main():
    """Main function"""
    parser = argparse.ArgumentParser(description="Generate monthly material report")
    parser.add_argument("--date", default="2025-07-31", help="Target date (YYYY-MM-DD)")
    args = parser.parse_args()
    
    print(f"Generating report for date: {args.date}")
    
    try:
        # Database connection
        config = DatabaseConfig(is_test=False)
        db_manager = DatabaseManager(config)
        
        # Parse date
        target_dt = datetime.strptime(args.date, '%Y-%m-%d')
        year, month = target_dt.year, target_dt.month
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Material Report"
        
        # Add header
        ws['A1'] = f"Monthly Material Report - {year}-{month:02d}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Query material usage data
        with db_manager.get_connection() as conn:
            cursor = conn.cursor()
            
            # Simple query for material usage summary
            cursor.execute("""
                SELECT 
                    s.name as store_name,
                    COUNT(DISTINCT mmu.material_id) as material_count,
                    SUM(mmu.material_used) as total_usage
                FROM material_monthly_usage mmu
                JOIN store s ON mmu.store_id = s.id
                WHERE mmu.year = %s AND mmu.month = %s
                GROUP BY s.id, s.name
                ORDER BY s.id
            """, (year, month))
            
            # Add headers
            ws['A3'] = "Store"
            ws['B3'] = "Material Count"
            ws['C3'] = "Total Usage"
            
            # Add data
            row = 4
            for record in cursor.fetchall():
                store_name = record['store_name'] if isinstance(record, dict) else record[0]
                material_count = record['material_count'] if isinstance(record, dict) else record[1]
                total_usage = record['total_usage'] if isinstance(record, dict) else record[2]
                
                ws[f'A{row}'] = store_name
                ws[f'B{row}'] = material_count
                ws[f'C{row}'] = float(total_usage) if total_usage else 0
                row += 1
        
        # Save file
        output_dir = Path('./output')
        output_dir.mkdir(exist_ok=True)
        filename = f"monthly_material_report_minimal_{args.date.replace('-', '_')}.xlsx"
        output_path = output_dir / filename
        
        wb.save(output_path)
        print(f"SUCCESS: Report saved to {output_path}")
        sys.exit(0)  # Success
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    from openpyxl.styles import Font
    main()