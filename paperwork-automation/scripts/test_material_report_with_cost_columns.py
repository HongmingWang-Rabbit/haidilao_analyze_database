#!/usr/bin/env python3
"""
Test the material report with the new cost columns
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

from lib.monthly_dishes_worksheet import MonthlyDishesWorksheetGenerator
from lib.database_queries import ReportDataProvider
from utils.database import DatabaseManager, DatabaseConfig
from openpyxl import Workbook
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def test_material_report_with_cost_columns():
    """Test the material report with the new cost columns"""
    
    print("Testing Material Report with New Cost Columns")
    print("=" * 55)
    
    try:
        # Initialize components with test database
        config = DatabaseConfig(is_test=True)
        db_manager = DatabaseManager(config)
        data_provider = ReportDataProvider(db_manager)
        
        # Create worksheet generator
        target_date = "2025-05-31"  # Use end of May
        store_names = ["加拿大一店", "加拿大二店", "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大七店"]
        worksheet_gen = MonthlyDishesWorksheetGenerator(store_names, target_date)
        
        print(f"Target date: {target_date}")
        print("Generating material variance worksheet with new cost columns...")
        
        # Create workbook and generate worksheet
        wb = Workbook()
        ws = worksheet_gen.generate_material_variance_worksheet(wb, data_provider)
        
        print(f"✅ Worksheet generated successfully!")
        print(f"   Worksheet name: {ws.title}")
        
        # Save test output
        output_file = "output/test_material_report_with_cost_columns.xlsx"
        wb.save(output_file)
        
        print(f"📁 Test output saved to: {output_file}")
        
        # Show some basic info about the worksheet
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"📊 Worksheet dimensions: {max_row} rows × {max_col} columns")
        
        # Check headers
        print("\\n📋 Column headers:")
        for col in range(1, min(max_col + 1, 16)):  # Show up to column O
            header_cell = None
            for row in range(1, 10):  # Search first 10 rows for headers
                cell = ws.cell(row=row, column=col)
                if cell.value and "序号" in str(cell.value):
                    # Found header row, get this column's header
                    header_cell = ws.cell(row=row, column=col)
                    break
                elif row > 1 and cell.value and isinstance(cell.value, str):
                    if any(keyword in cell.value for keyword in ["序号", "门店", "物料", "用量", "金额"]):
                        header_cell = cell
                        break
            
            if header_cell and header_cell.value:
                col_letter = chr(64 + col)  # A=65, so col 1 = A
                print(f"   {col_letter}: {header_cell.value}")
        
        print("\\n✅ Test completed successfully!")
        
        return True
        
    except Exception as e:
        print(f"❌ Error during test: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_material_report_with_cost_columns()
    if not success:
        sys.exit(1)