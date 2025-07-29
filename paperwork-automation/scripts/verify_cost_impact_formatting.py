#!/usr/bin/env python3
"""
Verify that cost impact calculation and conditional formatting are working correctly
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def verify_cost_impact_formatting():
    """Check if cost impact calculation and formatting are correct"""
    
    file_path = "output/monthly_gross_margin/毛利相关分析指标-202505.xlsx"
    
    try:
        # Load workbook to check formatting
        wb = load_workbook(file_path)
        ws_names = wb.sheetnames
        print(f"Available sheets: {ws_names}")
        
        # Find the material cost changes sheet (should be sheet 2)
        material_sheet = None
        for i, name in enumerate(ws_names):
            if "原材料成本变动" in name or i == 1:  # Sheet index 1 (2nd sheet)
                material_sheet = wb[name]
                break
        
        if not material_sheet:
            print("Could not find material cost changes sheet")
            return False
        
        print(f"Checking sheet: {material_sheet.title}")
        
        # Check a few rows for cost impact values and formatting
        cost_impact_samples = []
        formatting_samples = []
        
        # Start from row 3 (skip headers), check first 10 data rows
        for row_num in range(3, 13):
            try:
                # Cost impact should be in column L (12)
                cost_impact_cell = material_sheet.cell(row=row_num, column=12)
                cost_impact_value = cost_impact_cell.value
                
                if cost_impact_value is not None and cost_impact_value != 0:
                    cost_impact_samples.append({
                        'row': row_num,
                        'value': cost_impact_value,
                        'fill_color': cost_impact_cell.fill.start_color.rgb if cost_impact_cell.fill else None
                    })
                    
                    # Check formatting based on value
                    expected_color = None
                    if cost_impact_value == 0:
                        expected_color = "FFFFFF"  # White
                    elif cost_impact_value > 0:
                        expected_color = "C6EFCE"  # Green
                    else:
                        expected_color = "FFC7CE"  # Red
                    
                    actual_color = cost_impact_cell.fill.start_color.rgb if cost_impact_cell.fill else "FFFFFF"
                    
                    formatting_samples.append({
                        'row': row_num,
                        'value': cost_impact_value,
                        'expected_color': expected_color,
                        'actual_color': actual_color,
                        'correct': actual_color == expected_color
                    })
                    
            except Exception as e:
                print(f"Error checking row {row_num}: {e}")
                continue
        
        print(f"\nCost Impact Sample Values:")
        for sample in cost_impact_samples[:5]:  # Show first 5
            print(f"  Row {sample['row']}: ${sample['value']:.4f}")
        
        print(f"\nFormatting Verification:")
        correct_formatting = 0
        for sample in formatting_samples[:5]:  # Show first 5
            status = "✅" if sample['correct'] else "❌"
            print(f"  Row {sample['row']}: ${sample['value']:.4f} - {status}")
            print(f"    Expected: {sample['expected_color']}, Actual: {sample['actual_color']}")
            if sample['correct']:
                correct_formatting += 1
        
        if len(formatting_samples) > 0:
            accuracy = correct_formatting / len(formatting_samples[:5]) * 100
            print(f"\nFormatting Accuracy: {accuracy:.1f}% ({correct_formatting}/{len(formatting_samples[:5])})")
            
            if accuracy >= 80:
                print("SUCCESS: Cost impact formatting is working correctly!")
                return True
            else:
                print("ISSUE: Cost impact formatting needs adjustment")
                return False
        else:
            print("No cost impact data found to verify")
            return False
            
    except Exception as e:
        print(f"Error verifying cost impact: {e}")
        return False

if __name__ == "__main__":
    verify_cost_impact_formatting()