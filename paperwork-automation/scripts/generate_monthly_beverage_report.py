#!/usr/bin/env python3
"""
Generate monthly beverage report.
Focuses on beverage sales vs inventory variance analysis.
Compares system sales data with actual inventory count for beverages.
"""

import os
import sys
from pathlib import Path

# Add parent directory to path for imports FIRST
sys.path.insert(0, str(Path(__file__).parent.parent))

# Now import our modules after path is set
from utils.database import DatabaseConfig, DatabaseManager
from lib.beverage_summary_worksheet import BeverageSummaryGenerator
from lib.beverage_variance_worksheet import BeverageVarianceGenerator
from openpyxl import Workbook
from dotenv import load_dotenv
from datetime import datetime

# Load environment variables
load_dotenv()


class MonthlyBeverageReportGenerator:
    """Monthly beverage report generator focusing on beverage sales vs inventory variance"""

    def __init__(self, target_date: str, is_test: bool = False):
        self.target_date = target_date
        self.is_test = is_test
        self.config = DatabaseConfig(is_test=is_test)
        self.db_manager = DatabaseManager(self.config)
        self.output_dir = Path(os.getenv('OUTPUT_DIR', './output'))
        self.output_dir.mkdir(exist_ok=True)

        # Store mapping
        self.store_names = {
            1: "加拿大一店", 2: "加拿大二店", 3: "加拿大三店", 4: "加拿大四店",
            5: "加拿大五店", 6: "加拿大六店", 7: "加拿大七店"
        }

        # Data provider for beverage data
        self.data_provider = self

    def get_beverage_data_summary(self):
        """Get summary of available beverage data"""
        try:
            target_dt = datetime.strptime(self.target_date, '%Y-%m-%d')
            year, month = target_dt.year, target_dt.month

            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()

                # Check beverage sales data - simplified approach
                beverage_sales = {'count': 0, 'beverages': 0,
                                  'total_sales': 0, 'total_returns': 0}

                # Check inventory count data for beverages - with proper material filtering
                cursor.execute("""
                    SELECT 
                        COUNT(*) as count, 
                        COUNT(CASE WHEN mt.name = '成本-酒水类' THEN 1 END) as materials,
                        SUM(ic.counted_quantity) as total_counted, 
                        0 as total_system
                    FROM inventory_count ic
                    LEFT JOIN material m ON ic.material_id = m.id
                    LEFT JOIN material_type mt ON m.material_type_id = mt.id
                    WHERE EXTRACT(year FROM ic.count_date) = %s AND EXTRACT(month FROM ic.count_date) = %s
                """, (year, month))
                inventory_counts = cursor.fetchone()
                return {
                    'beverage_sales': beverage_sales,
                    'inventory_counts': inventory_counts,
                    'year': year,
                    'month': month
                }

        except Exception as e:
            print(f"ERROR: Error getting beverage data summary: {e}")
            return None

    def generate_report(self):
        """Generate monthly beverage report with variance analysis"""
        print(
            f"INFO: Generating monthly beverage report for {self.target_date}")

        # Check what data is available
        data_summary = self.get_beverage_data_summary()
        if not data_summary:
            print("ERROR: Could not access beverage data")
            return None

        beverage_sales = data_summary['beverage_sales']
        inventory_counts = data_summary['inventory_counts']
        year, month = data_summary['year'], data_summary['month']

        print(f"INFO: Available beverage data for {year}-{month:02d}:")
        print(
            f"   SALES: Beverage sales: {beverage_sales['count']} records, {beverage_sales['beverages']} beverages")
        print(
            f"   INVENTORY: Inventory counts: {inventory_counts['count']} records, {inventory_counts['materials']} materials")

        if beverage_sales['count'] == 0 and inventory_counts['count'] == 0:
            print("ERROR: No beverage data found")
            return None

        # Create Excel workbook
        wb = Workbook()
        # Remove default sheet
        if wb.active:
            wb.remove(wb.active)

        # Generate beverage summary worksheet (汇总表)
        if beverage_sales['count'] > 0 or inventory_counts['count'] > 0:
            print("INFO: Generating Beverage Summary worksheet...")
            summary_generator = BeverageSummaryGenerator(self)
            summary_generator.generate_worksheet(wb, self.target_date)

        # Generate beverage variance worksheet (差异明细表)
        if inventory_counts['count'] > 0:
            print("INFO: Generating Beverage Variance worksheet...")
            variance_generator = BeverageVarianceGenerator(self)
            variance_generator.generate_worksheet(wb, self.target_date)

        # Add overview worksheet
        overview_ws = self.create_overview_worksheet(wb, data_summary)

        if not wb.worksheets:
            print("ERROR: No worksheets generated")
            return None

        # Save the report
        output_path = self.save_report(wb)
        if output_path:
            print(f"SUCCESS: Monthly beverage report generated: {output_path}")
            return output_path
        else:
            print("ERROR: Failed to save beverage report")
            return None

    def create_overview_worksheet(self, wb, data_summary):
        """Create overview worksheet with beverage statistics"""
        ws = wb.create_sheet("酒水统计概览", 0)  # Insert as first sheet

        from openpyxl.styles import PatternFill, Font, Alignment

        # Title
        ws['A1'] = f"海底捞酒水数据统计 - {data_summary['year']}年{data_summary['month']:02d}月"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = PatternFill(
            start_color="C41E3A", end_color="C41E3A", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')

        # Statistics
        current_row = 3

        # Beverage sales statistics
        ws[f'A{current_row}'] = "酒水销售统计"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 1

        sales_stats = data_summary['beverage_sales']
        sales_data = [
            ("销售记录数", sales_stats['count']),
            ("销售酒水品种", sales_stats['beverages']),
            ("总销售量",
             f"{sales_stats['total_sales']:.0f}" if sales_stats['total_sales'] else "0"),
            ("总退货量",
             f"{sales_stats['total_returns']:.0f}" if sales_stats['total_returns'] else "0")
        ]

        for label, value in sales_data:
            ws[f'A{current_row}'] = label
            ws[f'B{current_row}'] = value
            current_row += 1

        current_row += 1

        # Inventory statistics
        ws[f'A{current_row}'] = "酒水库存统计"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        current_row += 1

        inventory_stats = data_summary['inventory_counts']
        inventory_data = [
            ("盘点记录数", inventory_stats['count']),
            ("盘点酒水品种", inventory_stats['materials']),
            ("总盘点数量",
             f"{inventory_stats['total_counted']:.2f}" if inventory_stats['total_counted'] else "0"),
            ("总系统数量",
             f"{inventory_stats['total_system']:.2f}" if inventory_stats['total_system'] else "0")
        ]

        for label, value in inventory_data:
            ws[f'A{current_row}'] = label
            ws[f'B{current_row}'] = value
            current_row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15

        return ws

    def save_report(self, wb):
        """Save the Excel workbook to file"""
        try:
            filename = f"monthly_beverage_report_{self.target_date.replace('-', '_')}.xlsx"
            output_path = self.output_dir / filename
            wb.save(output_path)
            return output_path
        except Exception as e:
            print(f"ERROR: Error saving beverage report: {e}")
            return None


def main():
    """Main function"""
    import argparse

    parser = argparse.ArgumentParser(
        description="Generate monthly beverage report")
    parser.add_argument("--date", default="2025-06-01",
                        help="Target date (YYYY-MM-DD)")
    parser.add_argument("--test", action="store_true",
                        help="Use test database")

    args = parser.parse_args()

    try:
        generator = MonthlyBeverageReportGenerator(
            args.date, is_test=args.test)
        output_path = generator.generate_report()

        if output_path:
            print("SUCCESS: Monthly beverage report generation completed successfully!")
        else:
            print("ERROR: Failed to generate monthly beverage report")
            sys.exit(1)

    except Exception as e:
        print(f"ERROR: Failed to generate monthly beverage report - {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
