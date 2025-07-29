#!/usr/bin/env python3
"""Analyze daily report files to find discount/promotion data"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from openpyxl import load_workbook
    import re
    
    # Find a daily report file
    daily_report_dir = Path("history_files/daily_report_inputs/20240401-20250724/daily_store_report")
    excel_files = list(daily_report_dir.glob("*.xlsx"))
    
    if not excel_files:
        print("No Excel files found!")
        sys.exit(1)
    
    # Use the first file
    excel_file = excel_files[0]
    print(f"Analyzing file: {excel_file.name}")
    print("=" * 60)
    
    # Load workbook
    wb = load_workbook(excel_file, read_only=True, data_only=True)
    
    print(f"Sheets in file: {wb.sheetnames}")
    print()
    
    # Check each sheet for discount-related columns
    discount_keywords = ['折扣', '优惠', '减免', '打折', 'discount', '赠送', '免单', '折让']
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print("-" * 40)
        
        # Get headers (usually in first few rows)
        headers_found = []
        for row_idx in range(1, min(10, ws.max_row + 1)):
            row_headers = []
            for col_idx in range(1, min(30, ws.max_column + 1)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value)
                    # Check if this might be a header row
                    if any(keyword in cell_str for keyword in discount_keywords):
                        row_headers.append(f"Col{col_idx}: {cell_str}")
            
            if row_headers:
                print(f"  Row {row_idx} discount-related headers:")
                for header in row_headers:
                    print(f"    {header}")
                headers_found.extend(row_headers)
        
        # If no discount headers found, show all headers from likely header row
        if not headers_found and ws.max_row > 0:
            # Find the row with most non-empty cells (likely header)
            header_row = 1
            max_cells = 0
            for row_idx in range(1, min(10, ws.max_row + 1)):
                non_empty = sum(1 for col in range(1, ws.max_column + 1) 
                              if ws.cell(row=row_idx, column=col).value)
                if non_empty > max_cells:
                    max_cells = non_empty
                    header_row = row_idx
            
            print(f"  All headers in row {header_row}:")
            for col_idx in range(1, min(20, ws.max_column + 1)):
                value = ws.cell(row=header_row, column=col_idx).value
                if value:
                    print(f"    Col{col_idx}: {value}")
    
    wb.close()
    
except ImportError:
    print("openpyxl not available. Trying alternative approach...")
    
    # Alternative: Use pandas if available
    try:
        import pandas as pd
        
        daily_report_dir = Path("history_files/daily_report_inputs/20240401-20250724/daily_store_report")
        excel_files = list(daily_report_dir.glob("*.xlsx"))
        
        if excel_files:
            excel_file = excel_files[0]
            print(f"\nAnalyzing with pandas: {excel_file.name}")
            
            # Read all sheets
            excel_data = pd.ExcelFile(excel_file)
            
            for sheet_name in excel_data.sheet_names:
                print(f"\nSheet: {sheet_name}")
                df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=20)
                
                # Look for discount columns
                discount_cols = [col for col in df.columns 
                               if any(keyword in str(col) for keyword in 
                                     ['折扣', '优惠', '减免', '打折', 'discount', '赠送', '免单', '折让'])]
                
                if discount_cols:
                    print(f"  Discount columns found: {discount_cols}")
                else:
                    print(f"  All columns: {list(df.columns)[:10]}...")  # Show first 10
                    
    except ImportError:
        print("Neither openpyxl nor pandas available")
        print("\nManual inspection approach:")
        print("1. Open a daily report file in Excel")
        print("2. Look for columns with keywords: 折扣, 优惠, 减免, 打折, discount, 赠送, 免单, 折让")
        print("3. Common discount data includes:")
        print("   - 折扣金额 (Discount amount)")
        print("   - 优惠金额 (Promotion amount)")
        print("   - 赠送金额 (Complimentary amount)")
        print("   - 折扣率 (Discount rate)")