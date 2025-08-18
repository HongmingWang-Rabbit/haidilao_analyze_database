#!/usr/bin/env python3
"""
Verify RBC transactions in output file have no duplicates
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

def verify_output():
    """Check the latest output file for duplicates"""
    
    # Find the latest output file
    output_dir = Path("output")
    output_files = list(output_dir.glob("Bank_Transactions_Report_*.xlsx"))
    
    if not output_files:
        print("No output files found")
        return
    
    latest_file = max(output_files, key=lambda p: p.stat().st_mtime)
    print(f"Checking latest output file: {latest_file.name}")
    
    # Load the workbook
    wb = load_workbook(latest_file)
    
    # Check all RBC sheets
    rbc_sheets = [sheet for sheet in wb.sheetnames if 'RBC' in sheet]
    
    for sheet_name in rbc_sheets:
        print(f"\nChecking {sheet_name}...")
        ws = wb[sheet_name]
        
        # Extract transactions (skip headers)
        transactions = []
        for row in range(3, ws.max_row + 1):
            # Check if row has data
            if not any(ws.cell(row=row, column=col).value for col in range(1, 7)):
                break
                
            # Get key fields based on RBC format
            date_val = ws.cell(row=row, column=2).value  # Effective Date
            desc_val = ws.cell(row=row, column=1).value  # Description
            debit_val = ws.cell(row=row, column=4).value  # Debits
            credit_val = ws.cell(row=row, column=5).value  # Credits
            
            if date_val:
                transactions.append({
                    'row': row,
                    'date': str(date_val),
                    'desc': str(desc_val) if desc_val else '',
                    'debit': str(debit_val) if debit_val else '',
                    'credit': str(credit_val) if credit_val else ''
                })
        
        print(f"  Total transactions: {len(transactions)}")
        
        # Check for duplicates
        seen = {}
        duplicates = []
        
        for trans in transactions:
            key = (trans['date'], trans['desc'], trans['debit'], trans['credit'])
            if key in seen:
                duplicates.append((trans['row'], seen[key]))
            else:
                seen[key] = trans['row']
        
        if duplicates:
            print(f"  WARNING: Found {len(duplicates)} duplicate pairs!")
            for dup in duplicates[:5]:  # Show first 5
                print(f"    Rows {dup[0]} and {dup[1]} are duplicates")
        else:
            print(f"  ✓ No duplicates found")
    
    wb.close()

if __name__ == "__main__":
    verify_output()