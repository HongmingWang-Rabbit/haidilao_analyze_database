#!/usr/bin/env python3
"""More detailed check for images and objects in Excel file"""

from openpyxl import load_workbook
import sys
import zipfile
import os

# Fix encoding issue
sys.stdout.reconfigure(encoding='utf-8')

def check_excel_internals(file_path):
    """Check Excel file internals for embedded objects"""
    print("=== Checking Excel file internals ===")
    
    try:
        # Excel files are ZIP archives - check contents
        with zipfile.ZipFile(file_path, 'r') as zip_file:
            file_list = zip_file.namelist()
            
            # Look for media files (images)
            media_files = [f for f in file_list if 'media' in f.lower()]
            if media_files:
                print(f"Media files found: {len(media_files)}")
                for media in media_files:
                    print(f"  {media}")
            else:
                print("No media files found in Excel archive")
            
            # Look for drawing files
            drawing_files = [f for f in file_list if 'drawing' in f.lower()]
            if drawing_files:
                print(f"Drawing files found: {len(drawing_files)}")
                for drawing in drawing_files:
                    print(f"  {drawing}")
            else:
                print("No drawing files found")
            
            # Look for chart files
            chart_files = [f for f in file_list if 'chart' in f.lower()]
            if chart_files:
                print(f"Chart files found: {len(chart_files)}")
                for chart in chart_files:
                    print(f"  {chart}")
            else:
                print("No chart files found")
                
            # Show all files for reference
            print(f"\nAll files in Excel archive ({len(file_list)} total):")
            [print(f"  {f}") for f in sorted(file_list)[:20]]  # Show first 20
            if len(file_list) > 20:
                print(f"  ... and {len(file_list) - 20} more files")
                
    except Exception as e:
        print(f"Error checking Excel internals: {e}")

def check_openpyxl_objects(file_path):
    """Check using openpyxl object detection"""
    print("\n=== Checking with openpyxl ===")
    
    try:
        wb = load_workbook(file_path)
        
        for sheet_name in wb.sheetnames[:3]:  # Check first 3 sheets
            ws = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ---")
            
            # Check all possible object containers
            attrs_to_check = ['_images', '_charts', '_drawing', '_shapes', 'drawing']
            
            for attr in attrs_to_check:
                if hasattr(ws, attr):
                    obj = getattr(ws, attr)
                    if obj:
                        if hasattr(obj, '__len__'):
                            print(f"  {attr}: {len(obj)} items")
                        else:
                            print(f"  {attr}: exists (not countable)")
                    else:
                        print(f"  {attr}: empty/None")
                else:
                    print(f"  {attr}: attribute not found")
        
        wb.close()
        
    except Exception as e:
        print(f"Error with openpyxl check: {e}")

# Check the template file
template_path = 'Input/daily_report/bank_transactions_reports/CA全部7家店明细.xlsx'

if os.path.exists(template_path):
    print(f"Checking template file: {template_path}")
    check_excel_internals(template_path)
    check_openpyxl_objects(template_path)
else:
    print(f"Template file not found: {template_path}")