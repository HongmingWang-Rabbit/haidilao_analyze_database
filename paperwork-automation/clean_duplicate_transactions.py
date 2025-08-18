#!/usr/bin/env python3
"""
Clean duplicate transactions from the output file
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import sys
import io
from datetime import datetime

# Set UTF-8 encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def clean_duplicates():
    """Remove duplicate transactions and sort by date"""
    
    # Find the latest output file
    output_dir = Path("output")
    output_files = list(output_dir.glob("Bank_Transactions_Report_*.xlsx"))
    
    if not output_files:
        print("No output files found")
        return
    
    latest_file = max(output_files, key=lambda p: p.stat().st_mtime)
    print(f"Processing file: {latest_file.name}")
    
    # Create a cleaned output file
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    cleaned_file = output_dir / f"Bank_Transactions_Report_CLEANED_{timestamp}.xlsx"
    
    # Load the workbook
    wb = load_workbook(latest_file)
    
    sheets_processed = 0
    total_duplicates_removed = 0
    
    for sheet_name in wb.sheetnames:
        # Skip summary or non-transaction sheets
        if 'Summary' in sheet_name or 'Cover' in sheet_name:
            continue
            
        ws = wb[sheet_name]
        
        # Determine column positions based on sheet type
        if 'RBC' in sheet_name:
            date_col = 2
            desc_col = 1
            debit_col = 4
            credit_col = 5
        elif 'CIBC' in sheet_name:
            date_col = 1
            desc_col = 2
            debit_col = 3
            credit_col = 4
        else:  # BMO format
            date_col = 1
            desc_col = 3
            debit_col = 6
            credit_col = 7
        
        # Extract all transactions
        transactions = []
        headers = []
        
        # Save headers (rows 1-2)
        for row in range(1, 3):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append({
                    'value': cell.value,
                    'has_fill': cell.fill.patternType is not None if cell.fill else False,
                    'fill_color': cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
                })
            headers.append(row_data)
        
        # Extract transaction data
        for row in range(3, ws.max_row + 1):
            # Check if row has data
            date_val = ws.cell(row=row, column=date_col).value
            if not date_val or str(date_val).lower() in ['date', '']:
                continue
            
            # Extract all cell data and formatting
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append({
                    'value': cell.value,
                    'has_fill': cell.fill.patternType is not None if cell.fill else False,
                    'fill_color': cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
                })
            
            # Create transaction record
            transactions.append({
                'date': date_val,
                'desc': ws.cell(row=row, column=desc_col).value,
                'debit': ws.cell(row=row, column=debit_col).value,
                'credit': ws.cell(row=row, column=credit_col).value,
                'row_data': row_data,
                'original_row': row
            })
        
        if not transactions:
            continue
        
        print(f"\nProcessing {sheet_name}:")
        print(f"  Original transactions: {len(transactions)}")
        
        # Remove duplicates - keep first occurrence
        seen = set()
        unique_transactions = []
        duplicates_removed = 0
        
        for trans in transactions:
            # Create a key for duplicate detection
            key = (
                str(trans['date']).strip(),
                str(trans['desc']).strip() if trans['desc'] else '',
                str(trans['debit']).strip() if trans['debit'] else '',
                str(trans['credit']).strip() if trans['credit'] else ''
            )
            
            if key not in seen:
                seen.add(key)
                unique_transactions.append(trans)
            else:
                duplicates_removed += 1
        
        print(f"  Duplicates removed: {duplicates_removed}")
        
        # Sort by date
        try:
            # Try to parse dates for proper sorting
            for trans in unique_transactions:
                trans['parsed_date'] = pd.to_datetime(trans['date'])
            
            unique_transactions.sort(key=lambda x: x['parsed_date'])
        except:
            # Fall back to string sorting if date parsing fails
            unique_transactions.sort(key=lambda x: str(x['date']))
        
        print(f"  Final transactions: {len(unique_transactions)}")
        
        # Clear the worksheet and rewrite
        # Clear all rows after headers
        for row in range(3, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
        
        # Write back unique, sorted transactions
        for idx, trans in enumerate(unique_transactions):
            new_row = idx + 3  # Start from row 3
            
            # Write all cell data with formatting
            for col_idx, cell_data in enumerate(trans['row_data'], 1):
                cell = ws.cell(row=new_row, column=col_idx)
                cell.value = cell_data['value']
                
                # Apply fill color if it existed
                if cell_data.get('has_fill') and cell_data.get('fill_color'):
                    cell.fill = PatternFill(start_color=cell_data['fill_color'], 
                                          end_color=cell_data['fill_color'], 
                                          fill_type='solid')
        
        sheets_processed += 1
        total_duplicates_removed += duplicates_removed
    
    # Save the cleaned workbook
    wb.save(cleaned_file)
    wb.close()
    
    print(f"\n{'='*60}")
    print(f"SUMMARY:")
    print(f"  Sheets processed: {sheets_processed}")
    print(f"  Total duplicates removed: {total_duplicates_removed}")
    print(f"  Cleaned file saved as: {cleaned_file.name}")

if __name__ == "__main__":
    clean_duplicates()