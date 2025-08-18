#!/usr/bin/env python3
"""
Check output file for duplicates with more detail
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import sys
import io
from collections import defaultdict

# Set UTF-8 encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def check_all_sheets():
    """Check all sheets in the output file for duplicates"""
    
    # Find the latest output file
    output_dir = Path("output")
    output_files = list(output_dir.glob("Bank_Transactions_Report_*.xlsx"))
    
    if not output_files:
        print("No output files found")
        return
    
    latest_file = max(output_files, key=lambda p: p.stat().st_mtime)
    print(f"Checking latest output file: {latest_file.name}\n")
    
    # Load the workbook
    wb = load_workbook(latest_file)
    
    # Check all sheets
    total_duplicates_found = 0
    
    for sheet_name in wb.sheetnames:
        # Skip summary or non-transaction sheets
        if 'Summary' in sheet_name or 'Cover' in sheet_name:
            continue
            
        ws = wb[sheet_name]
        
        # Determine column positions based on sheet type
        if 'RBC' in sheet_name:
            date_col = 2
            desc_col = 1
            debit_col = 4
            credit_col = 5
        elif 'CIBC' in sheet_name:
            date_col = 1
            desc_col = 2
            debit_col = 3
            credit_col = 4
        else:  # BMO format
            date_col = 1
            desc_col = 3
            debit_col = 6
            credit_col = 7
        
        # Extract all transactions
        transactions = []
        for row in range(3, ws.max_row + 1):
            # Check if row has data
            date_val = ws.cell(row=row, column=date_col).value
            if not date_val or str(date_val).lower() in ['date', '']:
                continue
                
            desc_val = ws.cell(row=row, column=desc_col).value
            debit_val = ws.cell(row=row, column=debit_col).value
            credit_val = ws.cell(row=row, column=credit_col).value
            
            # Normalize values for comparison
            date_str = str(date_val).strip()
            desc_str = str(desc_val).strip() if desc_val else ''
            debit_str = str(debit_val).strip() if debit_val and str(debit_val).strip() != '' else ''
            credit_str = str(credit_val).strip() if credit_val and str(credit_val).strip() != '' else ''
            
            transactions.append({
                'row': row,
                'date': date_str,
                'desc': desc_str,
                'debit': debit_str,
                'credit': credit_str
            })
        
        if not transactions:
            continue
            
        print(f"Sheet: {sheet_name}")
        print(f"  Total transactions: {len(transactions)}")
        
        # Group transactions by key to find duplicates
        transaction_groups = defaultdict(list)
        for trans in transactions:
            key = (trans['date'], trans['desc'], trans['debit'], trans['credit'])
            transaction_groups[key].append(trans['row'])
        
        # Find duplicates
        duplicates = {k: v for k, v in transaction_groups.items() if len(v) > 1}
        
        if duplicates:
            print(f"  WARNING: Found {len(duplicates)} groups of duplicate transactions!")
            total_duplicates_found += sum(len(rows) - 1 for rows in duplicates.values())
            
            # Show details of duplicates
            for i, (key, rows) in enumerate(list(duplicates.items())[:5]):  # Show first 5
                date, desc, debit, credit = key
                print(f"\n  Duplicate group {i+1} (rows: {rows}):")
                print(f"    Date: {date}")
                print(f"    Description: {desc[:60]}...")
                print(f"    Debit: {debit}, Credit: {credit}")
                print(f"    Appears in rows: {', '.join(map(str, rows))}")
        else:
            print(f"  ✓ No duplicates found")
        
        print()
    
    wb.close()
    
    print(f"\nTOTAL DUPLICATE TRANSACTIONS FOUND: {total_duplicates_found}")
    
    # Also check the template file to understand pre-existing duplicates
    print("\n" + "="*80)
    print("Checking template file for comparison...")
    
    template_file = Path("Input/daily_report/bank_transactions_reports/CA全部7家店明细.xlsx")
    if template_file.exists():
        wb_template = load_workbook(template_file)
        
        print("\nTemplate file duplicate check:")
        for sheet_name in ['RBC 5401', 'RBC 5419']:
            if sheet_name in wb_template.sheetnames:
                ws = wb_template[sheet_name]
                
                # Count rows with data
                row_count = 0
                for row in range(3, ws.max_row + 1):
                    if ws.cell(row=row, column=2).value:  # Date column
                        row_count += 1
                    else:
                        break
                
                print(f"  {sheet_name}: {row_count} transactions in template")
        
        wb_template.close()

if __name__ == "__main__":
    check_all_sheets()