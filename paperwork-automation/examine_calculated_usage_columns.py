#!/usr/bin/env python3
"""
Examine Calculated Usage Columns

This script examines the actual column names in the calculated dish material usage file
to identify the correct column name for unit conversion rate.
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
import logging
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def examine_calculated_usage_columns():
    """Examine the calculated usage file columns"""
    
    # Look for the calculated usage file
    usage_file = Path("Input/monthly_report/calculated_dish_material_usage/inventory_calculation_data_manual_extracted.xlsx")
    
    if not usage_file.exists():
        logger.error(f"File not found: {usage_file}")
        return False
    
    try:
        logger.info(f"Examining file: {usage_file}")
        
        # Load workbook to see sheet structure
        workbook = load_workbook(usage_file, read_only=True)
        logger.info(f"Sheets: {workbook.sheetnames}")
        
        # Read the first sheet
        sheet_name = workbook.sheetnames[0]
        df = pd.read_excel(usage_file, sheet_name=sheet_name, nrows=5)
        
        logger.info(f"Column names in '{sheet_name}' sheet:")
        for i, col in enumerate(df.columns):
            logger.info(f"  Column {i+1}: '{col}'")
        
        # Check for conversion rate related columns
        conversion_columns = []
        for col in df.columns:
            col_str = str(col).lower()
            if '转换' in col_str or '转换率' in col_str or '单位' in col_str:
                conversion_columns.append(col)
        
        logger.info(f"\nPotential conversion rate columns:")
        for col in conversion_columns:
            logger.info(f"  - '{col}'")
            
            # Show sample values
            sample_values = df[col].dropna().head(3).tolist()
            logger.info(f"    Sample values: {sample_values}")
        
        # Show a sample row to understand the data structure  
        logger.info(f"\nSample data (first row):")
        if len(df) > 0:
            sample_row = df.iloc[0]
            for col, value in sample_row.items():
                logger.info(f"  {col}: {value}")
        
        workbook.close()
        
    except Exception as e:
        logger.error(f"Error examining file: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return True


if __name__ == "__main__":
    success = examine_calculated_usage_columns()
    if success:
        print("Column examination completed!")
    else:
        print("Column examination failed!")