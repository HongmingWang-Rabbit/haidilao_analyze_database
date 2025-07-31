#!/usr/bin/env python3
"""
Examine Monthly Report Calculations

This script examines the monthly material report to see where theoretical usage
calculations are performed and whether they're using division or multiplication.
"""

import sys
import os
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

import pandas as pd
import logging
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def examine_monthly_report_calculations():
    """Examine the monthly report file for usage calculations"""
    
    # Check the monthly material report
    report_file = Path("output/monthly_material_report_2025_05_31.xlsx")
    
    if not report_file.exists():
        logger.error(f"Report file not found: {report_file}")
        return False
    
    try:
        logger.info(f"Examining file: {report_file}")
        
        # Load workbook to see sheet structure
        workbook = load_workbook(report_file, read_only=True)
        logger.info(f"Sheets: {workbook.sheetnames}")
        
        # Look for the material variance analysis sheet
        target_sheets = []
        for sheet_name in workbook.sheetnames:
            if '物料用量差异分析' in sheet_name or '差异分析' in sheet_name:
                target_sheets.append(sheet_name)
        
        if not target_sheets:
            logger.warning("No variance analysis sheets found")
            target_sheets = workbook.sheetnames[:3]  # Check first few sheets
        
        logger.info(f"Examining sheets: {target_sheets}")
        
        for sheet_name in target_sheets:
            logger.info(f"\n--- Sheet: {sheet_name} ---")
            
            try:
                df = pd.read_excel(report_file, sheet_name=sheet_name, nrows=20)
                
                logger.info(f"Columns in '{sheet_name}':")
                for i, col in enumerate(df.columns):
                    logger.info(f"  Column {i+1}: '{col}'")
                
                # Look for theoretical usage or materials_use columns
                usage_columns = []
                for col in df.columns:
                    col_str = str(col).lower()
                    if '理论' in col_str or 'theoretical' in col_str or 'materials_use' in col_str:
                        usage_columns.append(col)
                
                if usage_columns:
                    logger.info(f"Found usage calculation columns:")
                    for col in usage_columns:
                        logger.info(f"  - '{col}'")
                        
                        # Show sample values
                        sample_values = df[col].dropna().head(5).tolist()
                        logger.info(f"    Sample values: {sample_values}")
                
                # Look for beef brisket material specifically
                beef_rows = df[df.astype(str).apply(lambda x: x.str.contains('牛腩|1500677', na=False)).any(axis=1)]
                if len(beef_rows) > 0:
                    logger.info(f"Found {len(beef_rows)} beef brisket rows:")
                    for i, (idx, row) in enumerate(beef_rows.head(3).iterrows()):
                        logger.info(f"  Row {i+1}: {dict(row)}")
                
            except Exception as e:
                logger.warning(f"Could not read sheet '{sheet_name}': {e}")
        
        workbook.close()
        
    except Exception as e:
        logger.error(f"Error examining report: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    return True


if __name__ == "__main__":
    success = examine_monthly_report_calculations()
    if success:
        print("Monthly report examination completed!")
    else:
        print("Monthly report examination failed!")