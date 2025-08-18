#!/usr/bin/env python3
"""
Analyze why duplicates are appearing in append operation
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import sys
import io

# Set UTF-8 encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_issue():
    """Analyze the append operation issue"""
    
    # Check template file
    template_file = Path("Input/daily_report/bank_transactions_reports/CA全部7家店明细.xlsx")
    print(f"Analyzing template file: {template_file.name}\n")
    
    wb = load_workbook(template_file)
    
    # Check a specific sheet
    sheet_name = "CA1D-3817"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        print(f"Sheet: {sheet_name}")
        
        # Find last few transactions
        transactions = []
        for row in range(3, min(ws.max_row + 1, 100)):  # Check up to row 100
            date_val = ws.cell(row=row, column=1).value  # BMO date is in column 1
            if date_val and str(date_val).lower() != 'date':
                desc_val = ws.cell(row=row, column=3).value
                transactions.append({
                    'row': row,
                    'date': str(date_val),
                    'desc': str(desc_val)[:50] if desc_val else ''
                })
        
        print(f"Total transactions found: {len(transactions)}")
        
        if transactions:
            print("\nFirst 5 transactions:")
            for t in transactions[:5]:
                print(f"  Row {t['row']}: {t['date']} | {t['desc']}...")
                
            print("\nLast 5 transactions:")
            for t in transactions[-5:]:
                print(f"  Row {t['row']}: {t['date']} | {t['desc']}...")
    
    wb.close()
    
    # Check output file
    print("\n" + "="*80)
    output_files = list(Path("output").glob("Bank_Transactions_Report_*.xlsx"))
    if output_files:
        latest_output = max(output_files, key=lambda p: p.stat().st_mtime)
        print(f"\nAnalyzing output file: {latest_output.name}\n")
        
        wb_out = load_workbook(latest_output)
        
        if sheet_name in wb_out.sheetnames:
            ws_out = wb_out[sheet_name]
            
            # Check where transactions start appearing twice
            print(f"Checking for duplicate pattern in {sheet_name}...")
            
            # Look for where August transactions appear
            for row in range(3, min(ws_out.max_row + 1, 100)):
                date_val = ws_out.cell(row=row, column=1).value
                desc_val = ws_out.cell(row=row, column=3).value
                
                if date_val and 'Aug' in str(date_val):
                    print(f"\nFirst August transaction at row {row}: {date_val}")
                    
                    # Check next few rows
                    print("Next few rows:")
                    for r in range(row, min(row + 10, ws_out.max_row + 1)):
                        d = ws_out.cell(row=r, column=1).value
                        desc = ws_out.cell(row=r, column=3).value
                        print(f"  Row {r}: {d} | {str(desc)[:40]}...")
                    break
        
        wb_out.close()

if __name__ == "__main__":
    analyze_issue()