#!/usr/bin/env python3
"""
Yearly comparison worksheet generator (同比数据).
Only handles worksheet creation - receives data from main report generator.
"""

from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class YearlyComparisonWorksheetGenerator:
    """Generate yearly comparison worksheet (同比数据) from provided data"""

    def __init__(self, store_names, target_date):
        self.store_names = store_names
        self.target_date = target_date
        # Calculate previous year date
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        self.current_year = target_dt.year
        self.previous_year = target_dt.year - 1
        self.month = target_dt.month
        self.day = target_dt.day

    def calculate_percentage_change(self, current, previous):
        """Calculate percentage change between current and previous values"""
        if previous == 0 or previous is None:
            return 0.0
        return ((current - previous) / previous) * 100

    def format_percentage_change(self, change):
        """Format percentage change with appropriate color coding"""
        if change > 0:
            return f"{change:.1f}%"
        elif change < 0:
            return f"{change:.1f}%"
        else:
            return "0.0%"

    def generate_worksheet(self, wb, yearly_current, yearly_previous):
        """Generate the yearly comparison worksheet using provided data"""
        ws = wb.create_sheet("同比数据")

        # Check if we have sufficient data
        if not yearly_current or not yearly_previous:
            # Still generate worksheet but with limited data
            pass

        # Convert to dictionaries for easier lookup
        current_dict = {row['store_id']: row for row in yearly_current}
        previous_dict = {row['store_id']: row for row in yearly_previous}

        # Build comparison data structure
        comparison_data = {}

        # Calculate data for each store
        for store_id in current_dict.keys():
            store_name = self.store_names.get(store_id, f"Store {store_id}")
            current = current_dict.get(store_id, {})
            previous = previous_dict.get(store_id, {})

            # Safe conversion function to handle None and Decimal types
            def safe_float(value):
                if value is None:
                    return 0.0
                return float(value)

            # Current year values - use explicit field names
            current_tables = safe_float(current.get(
                'total_tables_validated', 0))  # For table count display
            current_revenue = safe_float(current.get('total_revenue', 0))
            current_turnover = safe_float(current.get('avg_turnover_rate', 0))
            current_per_table = safe_float(current.get(
                'avg_per_table', 0))  # This uses served tables

            # Previous year values - use explicit field names
            previous_tables = safe_float(previous.get(
                'total_tables_validated', 0)) if previous else 0.0
            previous_revenue = safe_float(previous.get(
                'total_revenue', 0)) if previous else 0.0
            previous_turnover = safe_float(previous.get(
                'avg_turnover_rate', 0)) if previous else 0.0
            previous_per_table = safe_float(previous.get(
                'avg_per_table', 0)) if previous else 0.0

            # Calculate changes
            tables_change = current_tables - previous_tables
            revenue_change = current_revenue - previous_revenue
            turnover_change = current_turnover - previous_turnover
            per_table_change = current_per_table - previous_per_table

            # Calculate percentage changes
            tables_pct = self.calculate_percentage_change(
                current_tables, previous_tables)
            revenue_pct = self.calculate_percentage_change(
                current_revenue, previous_revenue)
            turnover_pct = self.calculate_percentage_change(
                current_turnover, previous_turnover)
            per_table_pct = self.calculate_percentage_change(
                current_per_table, previous_per_table)

            comparison_data[store_name] = {
                # 桌数对比同期数据
                '本月截止目前': round(current_tables, 2),
                '去年截止同期': round(previous_tables, 2),
                '对比去年同期': round(tables_change, 2),
                '桌数增长率': self.format_percentage_change(tables_pct),

                # 翻台率对比同期数据
                '本月截止目前翻台率': round(current_turnover, 2),
                '去年截止同期翻台率': round(previous_turnover, 2),
                '对比去年同期翻台率': round(turnover_change, 2),
                '翻台率增长率': self.format_percentage_change(turnover_pct),

                # 营业收入(不含税-万加元)
                '本月截止目前收入': round(current_revenue / 10000, 2),
                '去年截止同期收入': round(previous_revenue / 10000, 2),
                '对比去年同期收入': round(revenue_change / 10000, 2),
                '收入增长率': self.format_percentage_change(revenue_pct),

                # 单桌消费对比同期数据
                '本月截止目前单桌': round(current_per_table, 2),
                '去年截止同期单桌': round(previous_per_table, 2),
                '对比去年同期单桌': round(per_table_change, 2),
                '单桌消费增长率': self.format_percentage_change(per_table_pct)
            }

        # Safe conversion function for totals (same as above)
        def safe_float_total(value):
            if value is None:
                return 0.0
            return float(value)

        # Calculate totals for 加拿大片区
        # For table count display: use validated tables explicitly
        total_current_tables = sum(
            safe_float_total(row['total_tables_validated']) for row in yearly_current)
        total_previous_tables = sum(
            safe_float_total(row['total_tables_validated']) for row in yearly_previous)

        # For consumption calculation: use direct calculation (revenue / non-validated tables)
        total_current_revenue = sum(
            safe_float_total(row['total_revenue']) for row in yearly_current)
        total_previous_revenue = sum(
            safe_float_total(row['total_revenue']) for row in yearly_previous)

        # Calculate total non-validated tables for proper per-table calculations
        total_current_tables_served = sum(
            safe_float_total(row['total_tables']) for row in yearly_current)  # non-validated
        total_previous_tables_served = sum(
            safe_float_total(row['total_tables']) for row in yearly_previous)  # non-validated

        # Calculate per-table consumption using direct division (not average of averages)
        total_current_per_table = total_current_revenue / \
            total_current_tables_served if total_current_tables_served > 0 else 0.0
        total_previous_per_table = total_previous_revenue / \
            total_previous_tables_served if total_previous_tables_served > 0 else 0.0
        total_current_turnover = sum(
            safe_float_total(row['avg_turnover_rate']) for row in yearly_current) / max(1, len(yearly_current))
        total_previous_turnover = sum(safe_float_total(
            row['avg_turnover_rate']) for row in yearly_previous) / max(1, len(yearly_previous)) if yearly_previous else 0.0

        # Calculate total changes
        total_tables_change = total_current_tables - total_previous_tables
        total_revenue_change = total_current_revenue - total_previous_revenue
        total_turnover_change = total_current_turnover - total_previous_turnover
        total_per_table_change = total_current_per_table - total_previous_per_table

        # Calculate total percentage changes
        total_tables_pct = self.calculate_percentage_change(
            total_current_tables, total_previous_tables)
        total_revenue_pct = self.calculate_percentage_change(
            total_current_revenue, total_previous_revenue)
        total_turnover_pct = self.calculate_percentage_change(
            total_current_turnover, total_previous_turnover)
        total_per_table_pct = self.calculate_percentage_change(
            total_current_per_table, total_previous_per_table)

        comparison_data['加拿大片区'] = {
            '本月截止目前': round(total_current_tables, 2),
            '去年截止同期': round(total_previous_tables, 2),
            '对比去年同期': round(total_tables_change, 2),
            '桌数增长率': self.format_percentage_change(total_tables_pct),
            '本月截止目前翻台率': round(total_current_turnover, 2),
            '去年截止同期翻台率': round(total_previous_turnover, 2),
            '对比去年同期翻台率': round(total_turnover_change, 2),
            '翻台率增长率': self.format_percentage_change(total_turnover_pct),
            '本月截止目前收入': round(total_current_revenue / 10000, 2),
            '去年截止同期收入': round(total_previous_revenue / 10000, 2),
            '对比去年同期收入': round(total_revenue_change / 10000, 2),
            '收入增长率': self.format_percentage_change(total_revenue_pct),
            '本月截止目前单桌': round(total_current_per_table, 2),
            '去年截止同期单桌': round(total_previous_per_table, 2),
            '对比去年同期单桌': round(total_per_table_change, 2),
            '单桌消费增长率': self.format_percentage_change(total_per_table_pct)
        }

        # Get weekday in Chinese
        target_dt = datetime.strptime(self.target_date, '%Y-%m-%d')
        weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
        weekday = weekdays[target_dt.weekday()]

        # Title
        title = f"加拿大-各门店{self.current_year}年{self.month}月{self.day}日同比数据-{weekday}"
        ws.merge_cells('A1:J1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid")

        # Headers - split into regions like the image
        # Row 2: Main headers
        ws.merge_cells('A2:B2')
        ws['A2'] = "分类"
        ws.merge_cells('C2:E2')
        ws['C2'] = "西部"
        ws.merge_cells('F2:I2')
        ws['F2'] = "东部"
        ws['J2'] = "加拿大片区"

        # Row 3: Store names
        headers_row3 = ["项目", "内容", "加拿大一店", "加拿大二店", "加拿大七店",
                        "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大片区"]
        for col, header in enumerate(headers_row3, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Apply header formatting
        for row in [2, 3]:
            for col in range(1, 11):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="FFD700", end_color="FFD700", fill_type="solid")
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')

        # Data rows matching the image structure
        data_rows = [
            # 桌数对比同期数据 section
            ("桌数\n对比同期数据", "本月截止目前", "FFFF99"),
            ("", "去年截止同期", "FFFF99"),
            ("", "对比去年同期", "FFFF99"),
            ("", "桌数增长率", "FFFF00"),  # Highlighted

            # 翻台率对比同期数据 section
            ("翻台率\n对比同期数据", "本月截止目前", "E6F3FF"),
            ("", "去年截止同期", "E6F3FF"),
            ("", "对比去年同期", "E6F3FF"),
            ("", "翻台率增长率", "FFFF00"),  # Highlighted

            # 营业收入 section
            ("营业收入\n(不含税-万加元)", "本月截止目前", "FFFF99"),
            ("", "去年截止同期", "FFFF99"),
            ("", "对比去年同期", "FFFF99"),
            ("", "收入增长率", "FFFF00"),  # Highlighted

            # 单桌消费对比同期数据 section
            ("单桌消费\n对比同期数据", "本月截止目前", "E6F3FF"),
            ("", "去年截止同期", "E6F3FF"),
            ("", "对比去年同期", "E6F3FF"),
            ("", "单桌消费增长率", "FFFF00"),  # Highlighted
        ]

        # Store order matching the image (西部: 一店,二店,七店; 东部: 三店,四店,五店,六店)
        store_order = ["加拿大一店", "加拿大二店", "加拿大七店",
                       "加拿大三店", "加拿大四店", "加拿大五店", "加拿大六店", "加拿大片区"]

        # Add data to worksheet
        current_row = 4
        current_section = ""  # Track the current section for proper data mapping

        for category, content, color in data_rows:
            # Update current section when we encounter a new category
            if category:
                current_section = category
                ws.cell(row=current_row, column=1, value=category)

            # Add content (column B)
            ws.cell(row=current_row, column=2, value=content)

            # Add data for each store in the specified order
            for col, store_name in enumerate(store_order, 3):
                if store_name in comparison_data:
                    # Map content to the correct data key using current_section context
                    data_key = content
                    if content == "本月截止目前" and "翻台率" in current_section:
                        data_key = "本月截止目前翻台率"
                    elif content == "去年截止同期" and "翻台率" in current_section:
                        data_key = "去年截止同期翻台率"
                    elif content == "对比去年同期" and "翻台率" in current_section:
                        data_key = "对比去年同期翻台率"
                    elif content == "翻台率增长率":
                        data_key = "翻台率增长率"
                    elif content == "本月截止目前" and "收入" in current_section:
                        data_key = "本月截止目前收入"
                    elif content == "去年截止同期" and "收入" in current_section:
                        data_key = "去年截止同期收入"
                    elif content == "对比去年同期" and "收入" in current_section:
                        data_key = "对比去年同期收入"
                    elif content == "收入增长率":
                        data_key = "收入增长率"
                    elif content == "本月截止目前" and "单桌" in current_section:
                        data_key = "本月截止目前单桌"
                    elif content == "去年截止同期" and "单桌" in current_section:
                        data_key = "去年截止同期单桌"
                    elif content == "对比去年同期" and "单桌" in current_section:
                        data_key = "对比去年同期单桌"
                    elif content == "单桌消费增长率":
                        data_key = "单桌消费增长率"
                    elif content == "桌数增长率":
                        data_key = "桌数增长率"

                    value = comparison_data[store_name].get(data_key, "")
                    cell = ws.cell(row=current_row, column=col, value=value)

                    # Apply background color
                    cell.fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid")

                    # Apply red color for negative percentage changes
                    if isinstance(value, str) and value.endswith('%') and value.startswith('-'):
                        # Red color for negative percentages
                        cell.font = Font(color="FF0000")

            # Apply background color to category and content cells
            ws.cell(row=current_row, column=1).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=current_row, column=2).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid")

            current_row += 1

        # Merge category cells
        merge_ranges = [
            (4, 7),   # 桌数对比同期数据 - 4 rows
            (8, 11),  # 翻台率对比同期数据 - 4 rows
            (12, 15),  # 营业收入 - 4 rows
            (16, 19)  # 单桌消费对比同期数据 - 4 rows
        ]

        for start_row, end_row in merge_ranges:
            ws.merge_cells(f'A{start_row}:A{end_row}')
            cell = ws[f'A{start_row}']
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)

        # Apply borders
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for row in range(1, current_row):
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = thin_border

        # Set column widths
        column_widths = [15, 20, 12, 12, 12, 12, 12, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Set row heights
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 20

        return ws
