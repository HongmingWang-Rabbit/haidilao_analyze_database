#!/usr/bin/env python3
"""
Comparison worksheet generator (对比上月表).
Only handles worksheet creation - receives data from main report generator.
"""

from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ComparisonWorksheetGenerator:
    """Generate comparison worksheet (对比上月表) from provided data"""

    def __init__(self, store_names, target_date):
        self.store_names = store_names
        self.target_date = target_date

    def calculate_time_progress(self, target_date: str):
        """Calculate actual time progress through the month"""
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')

        # Get last day of the month
        if target_dt.month == 12:
            next_month = target_dt.replace(
                year=target_dt.year + 1, month=1, day=1)
        else:
            next_month = target_dt.replace(month=target_dt.month + 1, day=1)

        last_day = (next_month - timedelta(days=1)).day
        current_day = target_dt.day

        progress = (current_day / last_day) * 100
        return round(progress, 2)

    def get_cell_value(self, content, category, store_name, col, daily_dict, monthly_dict,
                       prev_month_dict, targets_dict, current_mtd_dict, prev_mtd_dict,
                       daily_ranking, monthly_ranking, daily_ranking_values, monthly_ranking_values,
                       time_progress, target_dt):
        """Calculate cell value from processed data"""

        # Handle 加拿大片区 (totals) vs individual stores
        if store_name == "加拿大片区":
            # Calculate totals from all stores
            tables_served = sum(float(row['tables_served'])
                                for row in daily_dict.values())
            tables_served_validated = sum(
                float(row['tables_served_validated']) for row in daily_dict.values())
            takeout_tables = sum(float(row['takeout_tables'])
                                 for row in daily_dict.values())
            tables_validated = sum(
                float(row['tables_served_validated']) for row in daily_dict.values())
            revenue = sum(float(row['revenue_tax_not_included'])
                          for row in daily_dict.values())
            customers = sum(float(row['customers'])
                            for row in daily_dict.values())
            discount = sum(float(row['discount_total'])
                           for row in daily_dict.values())
            turnover_rate = sum(float(row['turnover_rate'])
                                for row in daily_dict.values()) / len(daily_dict)

            monthly_discount = sum(
                float(row['monthly_discount_total']) for row in monthly_dict.values())
            monthly_tables = sum(float(row['monthly_tables'])
                                 for row in monthly_dict.values())
            monthly_tables_validated = sum(
                float(row['monthly_tables_validated']) for row in monthly_dict.values())
            monthly_revenue = sum(float(row['monthly_revenue'])
                                  for row in monthly_dict.values())

            # Calculate average turnover rate using time segment method for consistency
            # This should match the time segment worksheet calculation (3.67)
            from lib.time_segment_worksheet import TimeSegmentWorksheetGenerator
            time_segment_generator = TimeSegmentWorksheetGenerator(
                self.store_names, self.target_date)
            time_segment_data = time_segment_generator.get_time_segment_data_for_date(
                self.target_date)

            store_turnover_totals = []
            for store_id in self.store_names.keys():
                if store_id in time_segment_data:
                    store_data = time_segment_data[store_id]
                    store_totals = time_segment_generator.calculate_store_totals(
                        store_data)
                    store_turnover_totals.append(
                        store_totals['total_turnover_current'])

            avg_monthly_turnover = sum(
                store_turnover_totals) / len(store_turnover_totals) if store_turnover_totals else 0

            # Previous month totals
            prev_month_tables = sum(float(row['prev_monthly_tables'])
                                    for row in prev_month_dict.values()) if prev_month_dict else 0
            prev_month_tables_validated = sum(float(
                row['prev_monthly_tables_validated']) for row in prev_month_dict.values()) if prev_month_dict else 0
            prev_month_revenue = sum(float(row['prev_monthly_revenue'])
                                     for row in prev_month_dict.values()) if prev_month_dict else 0

            # Month-to-date totals
            current_mtd_tables = sum(float(row['mtd_tables']) for row in current_mtd_dict.values(
            )) if current_mtd_dict else monthly_tables
            current_mtd_revenue = sum(float(row['mtd_revenue']) for row in current_mtd_dict.values(
            )) if current_mtd_dict else monthly_revenue
            current_mtd_tables_validated = sum(float(row['mtd_tables']) for row in current_mtd_dict.values(
            )) if current_mtd_dict else monthly_tables_validated
            current_mtd_discount = sum(float(row['mtd_discount_total']) for row in current_mtd_dict.values(
            )) if current_mtd_dict else monthly_discount

            # Previous month MTD totals (for proper comparison)
            prev_mtd_tables = sum(float(
                row['prev_mtd_tables']) for row in prev_mtd_dict.values()) if prev_mtd_dict else 0
            prev_mtd_tables_validated = sum(float(
                row['prev_monthly_tables_validated']) for row in prev_mtd_dict.values()) if prev_mtd_dict else 0
            prev_mtd_revenue = sum(float(
                row['prev_mtd_revenue']) for row in prev_mtd_dict.values()) if prev_mtd_dict else 0

            # Target totals - handle None values
            target_revenue = sum(float(row['target_revenue']) if row['target_revenue'] is not None else 0
                                 for row in targets_dict.values())

        else:
            # Get store_id from store_name
            store_id = None
            for sid, sname in self.store_names.items():
                if sname == store_name:
                    store_id = sid
                    break

            if not store_id or store_id not in daily_dict:
                return ""

            # Get data for this individual store
            daily_row = daily_dict[store_id]
            monthly_row = monthly_dict.get(store_id, {})
            prev_month_row = prev_month_dict.get(store_id, {})
            target_row = targets_dict.get(store_id, {})
            current_mtd_row = current_mtd_dict.get(store_id, {})
            prev_mtd_row = prev_mtd_dict.get(store_id, {})

            # Extract individual store values
            tables_served = float(daily_row.get('tables_served', 0))
            tables_served_validated = float(
                daily_row.get('tables_served_validated', 0))
            takeout_tables = float(daily_row.get('takeout_tables', 0))
            tables_validated = float(
                daily_row.get('tables_served_validated', 0))
            # Safe conversion function to handle Decimal types consistently

            def safe_float(value):
                return float(value) if value is not None else 0.0

            # Convert all values to float for consistent arithmetic
            revenue = safe_float(daily_row.get('revenue_tax_not_included', 0))
            customers = safe_float(daily_row.get('customers', 0))
            discount = safe_float(daily_row.get('discount_total', 0))
            turnover_rate = safe_float(daily_row.get('turnover_rate', 0))

            # Monthly data
            monthly_discount = safe_float(monthly_row.get(
                'monthly_discount_total', 0)) if monthly_row else 0.0
            monthly_tables = safe_float(monthly_row.get(
                'monthly_tables', 0)) if monthly_row else 0.0
            monthly_tables_validated = safe_float(monthly_row.get(
                'monthly_tables_validated', 0)) if monthly_row else 0.0
            monthly_revenue = safe_float(monthly_row.get(
                'monthly_revenue', 0)) if monthly_row else 0.0
            avg_monthly_turnover = safe_float(monthly_row.get(
                'avg_turnover_rate', 0)) if monthly_row else 0.0

            # Previous month data
            prev_month_tables = float(prev_month_row.get(
                'prev_monthly_tables', 0)) if prev_month_row else 0
            prev_month_tables_validated = float(prev_month_row.get(
                'prev_monthly_tables_validated', 0)) if prev_month_row else 0
            prev_month_revenue = float(prev_month_row.get(
                'prev_monthly_revenue', 0)) if prev_month_row else 0

            # Month-to-date data
            current_mtd_tables = safe_float(current_mtd_row.get(
                'mtd_tables', 0)) if current_mtd_row else monthly_tables
            current_mtd_tables_validated = safe_float(current_mtd_row.get(
                'mtd_tables_validated', 0)) if current_mtd_row else monthly_tables_validated
            current_mtd_discount = safe_float(current_mtd_row.get(
                'mtd_discount_total', 0)) if current_mtd_row else monthly_discount
            current_mtd_revenue = safe_float(current_mtd_row.get(
                'mtd_revenue', 0)) if current_mtd_row else monthly_revenue

            # Previous month MTD data (for proper comparison)
            prev_mtd_revenue = safe_float(prev_mtd_row.get(
                'prev_mtd_revenue', 0)) if prev_mtd_row else 0.0
            prev_mtd_tables = safe_float(prev_mtd_row.get(
                'prev_mtd_tables', 0)) if prev_mtd_row else 0.0
            prev_mtd_tables_validated = safe_float(prev_mtd_row.get(
                'prev_monthly_tables_validated', 0)) if prev_mtd_row else 0.0

            # Target data
            target_revenue = safe_float(target_row.get(
                'target_revenue', 100000)) if target_row else 100000.0

        # Calculate derived values (same logic for both totals and individual stores)
        excluded_customers = tables_served - tables_validated
        avg_per_table = revenue / tables_served if tables_served > 0 else 0
        per_capita = revenue / customers if customers > 0 else 0
        discount_pct = discount / revenue * 100 if revenue > 0 else 0
        mtd_discount_pct = current_mtd_discount / \
            current_mtd_revenue * 100 if current_mtd_revenue > 0 else 0
        completion_rate = monthly_revenue / \
            target_revenue * 100 if target_revenue > 0 else 0

        # Fix issue 1: 上月同期总桌数 should be MTD, not full month
        table_change = monthly_tables_validated - prev_mtd_tables_validated

        # Fix issue 2: 上月截止目前营业收入(万) should be MTD
        revenue_change = current_mtd_revenue - prev_mtd_revenue
        current_mtd_avg_table = current_mtd_revenue / \
            current_mtd_tables if current_mtd_tables > 0 else 0

        # Fix issue 3: 上月单桌消费 should be average of whole month using total tables served
        prev_month_avg_table = prev_month_revenue / \
            prev_month_tables if prev_month_tables > 0 else 0
        avg_table_change = current_mtd_avg_table - prev_month_avg_table

        # Return value based on content (unified logic for both totals and individual stores)
        if content == "今日总桌数":
            return round(tables_validated, 2)
        elif content == "今日外卖桌数":
            return round(takeout_tables, 2)
        elif content == "今日未计入考核桌数":
            return round(excluded_customers, 2)
        elif content == f"{target_dt.month}月总桌数":
            return round(monthly_tables_validated, 2)
        elif content == "上月同期总桌数":
            return round(prev_mtd_tables_validated, 2)
        elif content == "对比上月同期总桌数":
            return f"{'上升' if table_change >= 0 else '下降'}{abs(table_change):.2f}桌"
        elif content == "今日营业收入(万)":
            return round(revenue/10000, 2)
        elif content == "本月截止目前营业收入(万)":
            return round(current_mtd_revenue/10000, 2)
        elif content == "上月截止目前营业收入(万)":
            return round(prev_mtd_revenue/10000, 2)
        elif content == "环比营业收入变化(万)":
            return round(revenue_change/10000, 2)
        elif content == "本月营业收入目标(万)":
            return round(target_revenue/10000, 2)
        elif content == "本月截止目标完成率":
            return f"{completion_rate:.2f}%"
        elif content == "标准时间进度":
            return f"{time_progress:.2f}%"
        elif content == "当月累计优惠总金额(万)":
            return round(current_mtd_discount/10000, 2)
        elif content == "当月累计优惠占比":
            return f"{mtd_discount_pct:.2f}%"
        elif content == "今日人均消费":
            return round(per_capita, 2)
        elif content == "今日消费客数":
            return round(customers, 2)
        elif content == "今日单桌消费":
            return round(avg_per_table, 2)
        elif content == "截止今日单桌消费":
            return round(current_mtd_avg_table, 2)
        elif content == "上月单桌消费":
            return round(prev_month_avg_table, 2)
        elif content == "环比上月变化":
            return round(avg_table_change, 2)
        elif content == "名次":
            if store_name == "加拿大片区":
                return "当月累计平均翻台率"
            else:
                col_index = col - 3
                if col_index < 8:
                    return f"第{col_index + 1}名"
        elif content == f"{target_dt.month}月{target_dt.day}日翻台率排名店铺":
            if store_name == "加拿大片区":
                result = round(avg_monthly_turnover, 2)
                return result
            else:
                col_index = col - 3
                if col_index < len(daily_ranking):
                    return daily_ranking[col_index]
                return ""
        elif content == f"{target_dt.month}月{target_dt.day}日翻台率排名":
            if store_name == "加拿大片区":
                result = round(avg_monthly_turnover, 2)
                return result
            else:
                col_index = col - 3
                if col_index < len(daily_ranking_values):
                    return round(daily_ranking_values[col_index], 2)
                return ""
        elif content == f"{target_dt.month}月平均翻台率排名店铺":
            if store_name == "加拿大片区":
                result = round(avg_monthly_turnover, 2)
                return result
            else:
                col_index = col - 3
                if col_index < len(monthly_ranking):
                    return monthly_ranking[col_index]
                return ""
        elif content == f"{target_dt.month}月平均翻台率排名":
            if store_name == "加拿大片区":
                # FIX: Calculate accurate weighted average to avoid precision loss
                # Use the same calculation as in get_cell_value method
                if monthly_dict:
                    # Store seating capacity mapping (consistent with database_queries.py)
                    store_seating_capacity = {
                        1: 53, 2: 36, 3: 48, 4: 70, 5: 55, 6: 56, 7: 57}

                    # Calculate weighted average turnover rate using total tables served and seating capacity
                    total_tables_validated = sum(
                        float(row['monthly_tables_validated']) for row in monthly_dict.values())
                    total_seat_days = sum(store_seating_capacity.get(store_id, 53) * row['days_count']
                                          for store_id, row in monthly_dict.items())

                    if total_seat_days > 0:
                        accurate_avg = total_tables_validated / total_seat_days
                        return round(accurate_avg, 2)
                    else:
                        # Fallback to arithmetic mean if seating data is incomplete
                        result = sum(float(row['avg_turnover_rate'])
                                     for row in monthly_dict.values()) / len(monthly_dict)
                        return round(result, 2)
                else:
                    # Legacy fallback: Average of monthly ranking values (less precise)
                    if monthly_ranking_values and len(monthly_ranking_values) > 0:
                        correct_avg = sum(monthly_ranking_values) / \
                            len(monthly_ranking_values)
                        return round(correct_avg, 2)
                    else:
                        # Fallback to the average calculated from time segment data
                        from lib.time_segment_worksheet import TimeSegmentWorksheetGenerator
                        time_segment_generator = TimeSegmentWorksheetGenerator(
                            self.store_names, self.target_date)
                        time_segment_data = time_segment_generator.get_time_segment_data_for_date(
                            self.target_date)

                        store_turnover_totals = []
                        for store_id in self.store_names.keys():
                            if store_id in time_segment_data:
                                store_data = time_segment_data[store_id]
                                store_totals = time_segment_generator.calculate_store_totals(
                                    store_data)
                                store_turnover_totals.append(
                                    store_totals['total_turnover_current'])

                        fallback_avg = sum(store_turnover_totals) / \
                            len(store_turnover_totals) if store_turnover_totals else 0
                        return round(fallback_avg, 2)
            else:
                col_index = col - 3
                if col_index < len(monthly_ranking_values):
                    return round(monthly_ranking_values[col_index], 2)
                return ""

        return ""

    def generate_worksheet(self, wb, daily_data, monthly_data, previous_month_data,
                           monthly_targets, current_mtd, prev_mtd,
                           daily_ranking, monthly_ranking, daily_ranking_values, monthly_ranking_values):
        """Generate the comparison worksheet from provided data"""

        ws = wb.create_sheet("对比上月表")

        # Parse target date for calculations
        target_dt = datetime.strptime(self.target_date, '%Y-%m-%d')
        year, month, day = target_dt.year, target_dt.month, target_dt.day
        time_progress = self.calculate_time_progress(self.target_date)

        # Convert to dictionaries for easier access
        daily_dict = {row['store_id']: row for row in daily_data}
        monthly_dict = {row['store_id']: row for row in monthly_data}
        prev_month_dict = {row['store_id']: row for row in previous_month_data}
        targets_dict = {row['store_id']: row for row in monthly_targets}
        current_mtd_dict = {row['store_id']: row for row in current_mtd}
        prev_mtd_dict = {row['store_id']: row for row in prev_mtd}

        # Get weekday in Chinese
        weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
        weekday = weekdays[target_dt.weekday()]

        # Title
        title = f"加拿大-各门店{self.target_date.replace('-', '年', 1).replace('-', '月', 1)}日环比数据-{weekday}"
        ws.merge_cells('A1:K1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid")

        # Headers
        headers = ["项目", "内容"] + list(self.store_names.values()) + ["加拿大片区"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="FFD700", end_color="FFD700", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data rows with exact structure from screenshots
        data_rows = [
            # 桌数(考核) section
            ("桌数\n(考核)", "今日总桌数", "FFFF99"),
            ("", "今日外卖桌数", "FFFF99"),
            ("", "今日未计入考核桌数", "FFFF99"),
            ("", f"{target_dt.month}月总桌数", "FFFF99"),
            ("", "上月同期总桌数", "FFFF99"),
            ("", "对比上月同期总桌数", "FFFF00"),  # Highlighted

            # 收入 section
            ("收入\n(不含税-万加元)", "今日营业收入(万)", "E6F3FF"),
            ("", "本月截止目前营业收入(万)", "E6F3FF"),
            ("", "上月截止目前营业收入(万)", "E6F3FF"),
            ("", "环比营业收入变化(万)", "E6F3FF"),
            ("", "本月营业收入目标(万)", "E6F3FF"),
            ("", "本月截止目标完成率", "FFFF00"),  # Highlighted
            ("", "标准时间进度", "E6F3FF"),
            ("", "当月累计优惠总金额(万)", "E6F3FF"),
            ("", "当月累计优惠占比", "E6F3FF"),

            # 单桌消费 section
            ("单桌消费\n(不含税)", "今日人均消费", "FFFF99"),
            ("", "今日消费客数", "FFFF99"),
            ("", "今日单桌消费", "FFFF99"),
            ("", "截止今日单桌消费", "FFFF99"),  # Highlighted
            ("", "上月单桌消费", "FFFF99"),
            ("", "环比上月变化", "FFFF99"),

            # 翻台率 section
            ("翻台率", "名次", "FFFF00"),  # Highlighted
            ("", f"{target_dt.month}月{target_dt.day}日翻台率排名店铺", "E6F3FF"),
            ("", f"{target_dt.month}月{target_dt.day}日翻台率排名", "FFFF00"),
            ("", f"{target_dt.month}月平均翻台率排名店铺", "E6F3FF"),
            ("", f"{target_dt.month}月平均翻台率排名", "FFFF00"),
        ]

        # Add data to worksheet
        current_row = 3

        for category, content, color in data_rows:
            # Add category (column A)
            if category:
                ws.cell(row=current_row, column=1, value=category)

            # Add content (column B)
            ws.cell(row=current_row, column=2, value=content)

            # Add data for each store (columns C-I) and total (column J)
            for col, store_name in enumerate(list(self.store_names.values()) + ["加拿大片区"], 3):
                value = self.get_cell_value(content, category, store_name, col,
                                            daily_dict, monthly_dict, prev_month_dict, targets_dict,
                                            current_mtd_dict, prev_mtd_dict, daily_ranking, monthly_ranking,
                                            daily_ranking_values, monthly_ranking_values,
                                            time_progress, target_dt)

                cell = ws.cell(row=current_row, column=col, value=value)
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid")

            # Apply background color to category and content cells
            ws.cell(row=current_row, column=1).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid")
            ws.cell(row=current_row, column=2).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid")

            current_row += 1

        # Merge category cells
        merge_ranges = [
            (3, 8),   # 桌数(考核) - 6 rows
            (9, 17),  # 收入 - 11 rows
            (18, 23),  # 单桌消费 - 4 rows
            (24, 28),  # 翻台率 - 2 rows
        ]

        for start_row, end_row in merge_ranges:
            if start_row < end_row:
                ws.merge_cells(f'A{start_row}:A{end_row}')
                cell = ws[f'A{start_row}']
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.font = Font(bold=True)

        # Apply common formatting
        self.apply_common_formatting(ws, current_row)

        # manual modify some displays
        # #当月累计平均翻台率 - Calculate the correct average from monthly ranking values
        ws.merge_cells('K25:K28')
        mtd_avg_turnover = ws['K25']

        # FIX: Calculate accurate weighted average to avoid precision loss
        # Use the same calculation as in get_cell_value method
        if monthly_dict:
            # Store seating capacity mapping (consistent with database_queries.py)
            store_seating_capacity = {1: 53, 2: 36,
                                      3: 48, 4: 70, 5: 55, 6: 56, 7: 57}

            # Calculate weighted average turnover rate using total tables served and seating capacity
            total_tables_validated = sum(
                float(row['monthly_tables_validated']) for row in monthly_dict.values())
            total_seat_days = sum(store_seating_capacity.get(store_id, 53) * row['days_count']
                                  for store_id, row in monthly_dict.items())

            if total_seat_days > 0:
                accurate_avg = total_tables_validated / total_seat_days
                mtd_avg_turnover.value = round(accurate_avg, 2)
            else:
                # Fallback to arithmetic mean if seating data is incomplete
                result = sum(float(row['avg_turnover_rate'])
                             for row in monthly_dict.values()) / len(monthly_dict)
                mtd_avg_turnover.value = round(result, 2)
        else:
            # Legacy fallback: Average of monthly ranking values (less precise)
            if monthly_ranking_values and len(monthly_ranking_values) > 0:
                correct_avg = sum(monthly_ranking_values) / \
                    len(monthly_ranking_values)
                mtd_avg_turnover.value = round(correct_avg, 2)
            else:
                # Fallback to the average calculated from time segment data
                from lib.time_segment_worksheet import TimeSegmentWorksheetGenerator
                time_segment_generator = TimeSegmentWorksheetGenerator(
                    self.store_names, self.target_date)
                time_segment_data = time_segment_generator.get_time_segment_data_for_date(
                    self.target_date)

                store_turnover_totals = []
                for store_id in self.store_names.keys():
                    if store_id in time_segment_data:
                        store_data = time_segment_data[store_id]
                        store_totals = time_segment_generator.calculate_store_totals(
                            store_data)
                        store_turnover_totals.append(
                            store_totals['total_turnover_current'])

                fallback_avg = sum(store_turnover_totals) / \
                    len(store_turnover_totals) if store_turnover_totals else 0
                mtd_avg_turnover.value = round(fallback_avg, 2)

        mtd_avg_turnover.font = Font(bold=True, size=20, color="FF0000")
        mtd_avg_turnover.alignment = Alignment(
            horizontal='center', vertical='center')
        mtd_avg_turnover.fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid")

        return ws

    def apply_common_formatting(self, ws, current_row):
        """Apply common formatting to worksheet"""
        # Apply borders
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for row in range(1, current_row):
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

                # Apply number formatting for numeric values
                if isinstance(cell.value, (int, float)) and col > 2:
                    cell.number_format = '0.00'
                elif isinstance(cell.value, str) and col > 2:
                    # Check if it's a percentage
                    if cell.value.endswith('%'):
                        cell.number_format = '0.00%'

        # Set column widths
        column_widths = [20, 20, 12, 12, 12, 12, 12, 12, 12, 12, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Set row height for title
        ws.row_dimensions[1].height = 25
