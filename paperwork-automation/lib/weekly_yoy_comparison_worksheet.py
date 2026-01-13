#!/usr/bin/env python3
"""
MTD (Month-to-Date) Year-over-Year Comparison Worksheet Generator

This module generates a MTD YoY comparison worksheet (周对比上年表) with a
store-centric layout where all challenges for each store are grouped together.

Structure per store (7 rows):
1. 翻台率挑战 - Turnover rate challenge
2. 桌数挑战 - Tables challenge
3-6. 时段挑战 - Time segment challenges (4 segments)
7. 外卖挑战 - Takeout revenue challenge

All configuration values are sourced from configs/challenge_targets/.

Normalization Formula:
For MTD comparisons, previous year data is normalized using:
    prev_year_normalized = prev_year_month_total / prev_year_month_days * current_days
This ensures fair comparison between periods with different number of days.
"""

import calendar
import logging
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Import centralized configurations
from configs.store_config import STORE_SEATING_CAPACITY, STORE_MANAGERS, REGIONAL_MANAGER
from configs.challenge_targets import (
    get_store_turnover_target,
    is_store_excluded_from_regional,
    STORE_TARGET_CONFIG,
    DEFAULT_TURNOVER_IMPROVEMENT,
    DEFAULT_SLOW_TIME_TARGET,
    AFTERNOON_SLOW_TARGETS,
    LATE_NIGHT_TARGETS,
    TIME_SEGMENT_LABELS,
    TIME_SEGMENT_CONFIG,
    get_takeout_daily_improvement_cad,
    get_absolute_time_segment_target,
    get_takeout_target,
    is_using_absolute_targets,
    JANUARY_2026_TURNOVER_TARGETS,
)


class WeeklyYoYComparisonWorksheetGenerator:
    """Generator for MTD year-over-year comparison worksheet."""

    # Excel styling constants
    HEADER_COLOR = "DC2626"
    GREEN_COLOR = "E6FFE6"
    RED_COLOR = "FFE6E6"
    SUMMARY_COLOR = "FFF3CD"
    EXCLUDED_COLOR = "E0E0E0"
    SEPARATOR_COLOR = "333333"

    # Different colors for each store (8 stores)
    STORE_COLORS = [
        "4472C4",  # Store 1 - Blue
        "ED7D31",  # Store 2 - Orange
        "A5A5A5",  # Store 3 - Gray
        "FFC000",  # Store 4 - Gold
        "5B9BD5",  # Store 5 - Light Blue
        "70AD47",  # Store 6 - Green
        "9E480E",  # Store 7 - Brown
        "7030A0",  # Store 8 - Purple
    ]

    SECTION_COLORS = {
        '翻台率': "FDE9D9",
        '桌数': "DAEEF3",
        '时段': "E4DFEC",
        '外卖': "D8E4BC",
    }

    # Number of rows per store
    ROWS_PER_STORE = 7  # 1 turnover + 1 tables + 4 time segments + 1 takeout

    def __init__(self, data_provider):
        """Initialize the generator with data provider."""
        self.data_provider = data_provider
        self.logger = logging.getLogger(__name__)

        # Standard styling
        self.header_fill = PatternFill(
            start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.wrap_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Color fills
        self.green_fill = PatternFill(start_color=self.GREEN_COLOR, end_color=self.GREEN_COLOR, fill_type="solid")
        self.red_fill = PatternFill(start_color=self.RED_COLOR, end_color=self.RED_COLOR, fill_type="solid")
        self.summary_fill = PatternFill(start_color=self.SUMMARY_COLOR, end_color=self.SUMMARY_COLOR, fill_type="solid")
        self.excluded_fill = PatternFill(start_color=self.EXCLUDED_COLOR, end_color=self.EXCLUDED_COLOR, fill_type="solid")

    @staticmethod
    def _format_date_period(start_dt: datetime, end_dt: datetime) -> str:
        """Format date range as Chinese date period string."""
        year = end_dt.year % 100
        return f"{year}年{start_dt.month}月{start_dt.day}日-{end_dt.month}月{end_dt.day}日"

    @staticmethod
    def _normalize_to_mtd(full_month_total: float, full_month_days: int, current_days: int) -> float:
        """
        Normalize a full month total to MTD (Month-to-Date) period.

        Formula: full_month_total / full_month_days * current_days

        Args:
            full_month_total: Total value for the full month
            full_month_days: Number of days in the full month
            current_days: Number of days in current MTD period

        Returns:
            Normalized value for the MTD period
        """
        if full_month_days <= 0:
            return 0
        return full_month_total / full_month_days * current_days

    @staticmethod
    def _prorate_monthly_target(monthly_target: float, target_year: int, target_month: int, current_days: int) -> float:
        """
        Prorate a monthly target to MTD (Month-to-Date) period.

        Args:
            monthly_target: Full monthly target value
            target_year: Year of the target month
            target_month: Month number (1-12)
            current_days: Number of days in current MTD period

        Returns:
            Prorated target for the MTD period
        """
        days_in_month = calendar.monthrange(target_year, target_month)[1]
        return monthly_target / days_in_month * current_days

    def generate_worksheet(self, workbook: Workbook, target_date: str) -> Worksheet:
        """
        Generate MTD YoY comparison worksheet with store-centric layout.

        Args:
            workbook: Excel workbook to add worksheet to
            target_date: Target date in YYYY-MM-DD format

        Returns:
            The generated worksheet
        """
        try:
            self.logger.info(f"Generating MTD YoY comparison worksheet for {target_date}")

            ws = workbook.create_sheet("周对比上年表")

            # Parse dates
            target_dt = datetime.strptime(target_date, '%Y-%m-%d')
            start_dt = target_dt.replace(day=1)
            num_days = target_dt.day
            prev_year = target_dt.year - 1

            # Calculate previous year dates
            prev_start_dt = start_dt.replace(year=prev_year)
            prev_end_dt = target_dt.replace(year=prev_year)

            self.logger.info(f"Current MTD period: {start_dt.strftime('%Y-%m-%d')} to {target_dt.strftime('%Y-%m-%d')} ({num_days} days)")

            # Get all data
            store_data = self._get_mtd_data(start_dt, target_dt)
            time_segment_data = self.data_provider.get_time_segment_mtd_data(target_date)
            takeout_data = self.data_provider.get_takeout_mtd_data(target_date)

            if not store_data:
                self.logger.warning("No store data found")
                return ws

            # Format date periods for headers
            current_period = self._format_date_period(start_dt, target_dt)
            prev_period = self._format_date_period(prev_start_dt, prev_end_dt)

            # Generate headers
            self._generate_headers(ws, current_period, prev_period, prev_year, target_dt)

            # Add store data (store-centric layout)
            current_row = 3
            for store_index, store in enumerate(store_data):
                current_row = self._add_store_section(
                    ws, store, current_row, num_days,
                    time_segment_data, takeout_data,
                    target_dt, prev_period, current_period,
                    store_index
                )
                # Add separator row between stores
                current_row = self._add_separator_row(ws, current_row)

            # Add Canada total summary section
            current_row = self._add_canada_summary(
                ws, store_data, current_row, num_days,
                time_segment_data, takeout_data, target_dt
            )

            # Apply column widths
            self._apply_column_widths(ws)

            # Freeze panes
            ws.freeze_panes = "C3"

            self.logger.info(f"MTD YoY comparison worksheet generated with {len(store_data)} stores")
            self.logger.info("Formatting applied to weekly YoY comparison worksheet")

            return ws

        except Exception as e:
            self.logger.error(f"Error generating worksheet: {str(e)}")
            raise

    def _generate_headers(self, ws: Worksheet, current_period: str, prev_period: str,
                          prev_year: int, target_dt: datetime) -> None:
        """Generate worksheet headers."""
        # Row 1 - Main header
        headers_row1 = [
            "门店",           # A - Store name (merged)
            "挑战类型",       # B - Challenge type
            "去年数据",       # C - Previous year
            "目标",           # D - Target
            "今年数据",       # E - Current year
            "差距",           # F - Gap
            "备注"            # G - Notes
        ]

        for col, header in enumerate(headers_row1, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

        # Row 2 - Sub headers with date info
        sub_headers = [
            "",
            "",
            prev_period,
            f"去年+精进",
            current_period,
            "今年-目标",
            ""
        ]

        for col, header in enumerate(sub_headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

    def _add_store_section(self, ws: Worksheet, store: Dict, start_row: int,
                           num_days: int, time_segment_data: Dict,
                           takeout_data: Dict, target_dt: datetime,
                           prev_period: str, current_period: str,
                           store_index: int = 0) -> int:
        """
        Add all challenge data for a single store.

        Returns:
            Next available row after this store's section
        """
        store_id = store.get('store_id', 0)
        store_name = store.get('store_name', '')
        seating_capacity = STORE_SEATING_CAPACITY.get(store_id, 50)
        target_date_str = target_dt.strftime('%Y-%m-%d')

        # Check if using absolute targets (January 2026)
        use_absolute = is_using_absolute_targets(target_date_str)

        # Get store-specific color
        store_color = self.STORE_COLORS[store_index % len(self.STORE_COLORS)]
        store_fill = PatternFill(start_color=store_color, end_color=store_color, fill_type="solid")

        current_row = start_row
        section_start_row = start_row

        # === Row 1: 翻台率挑战 ===
        prev_turnover = float(store.get('prev_avg_turnover_rate', 0))
        current_turnover = float(store.get('current_avg_turnover_rate', 0))
        target_turnover = get_store_turnover_target(store_id, prev_turnover, target_date_str)
        turnover_gap = current_turnover - target_turnover

        self._write_data_row(ws, current_row, "翻台率",
                             prev_turnover, target_turnover, current_turnover, turnover_gap,
                             "翻台率", number_format='0.00')
        current_row += 1

        # === Row 2: 桌数挑战 ===
        prev_tables = int(prev_turnover * seating_capacity * num_days)
        target_tables = int(target_turnover * seating_capacity * num_days)
        current_tables = int(current_turnover * seating_capacity * num_days)

        if not is_store_excluded_from_regional(store_id):
            tables_gap = current_tables - target_tables
            self._write_data_row(ws, current_row, "桌数",
                                 prev_tables, target_tables, current_tables, tables_gap,
                                 "桌数", number_format='0')
        else:
            self._write_data_row(ws, current_row, "桌数",
                                 prev_tables, "N/A", current_tables, "N/A",
                                 "桌数", notes="不参与区域考核")
        current_row += 1

        # === Rows 3-6: 时段挑战 (4 segments) ===
        store_ts_data = time_segment_data.get(store_id, {}).get('time_segments', {}) if time_segment_data else {}

        for segment_config in TIME_SEGMENT_CONFIG:
            segment_label = segment_config['label']
            segment_key = segment_config['key']
            is_slow = segment_config['type'] == 'slow'

            segment_data = store_ts_data.get(segment_label, {})
            prev_daily_tables = segment_data.get('prev_total_tables', 0) / num_days if num_days > 0 else 0
            current_daily_tables = segment_data.get('current_total_tables', 0) / num_days if num_days > 0 else 0

            # For January 2026, use absolute targets for slow times
            if use_absolute and is_slow:
                absolute_target = get_absolute_time_segment_target(store_id, segment_key, target_date_str)
                if absolute_target is not None:
                    target_daily = absolute_target
                    gap = current_daily_tables - target_daily
                    self._write_data_row(ws, current_row, segment_label,
                                         prev_daily_tables, target_daily, current_daily_tables, gap,
                                         "时段", number_format='0.00',
                                         notes="日均桌数 (固定目标)")
                    current_row += 1
                    continue

            # Fall back to improvement-based calculation
            hardcoded_targets = segment_config['targets']
            if is_slow and hardcoded_targets:
                daily_target = hardcoded_targets.get(store_id, DEFAULT_SLOW_TIME_TARGET)
            else:
                daily_target = self._calculate_busy_time_target(
                    store_id, store, time_segment_data, segment_key, num_days, target_date_str
                )

            # Target is improvement over last year
            target_daily = prev_daily_tables + daily_target
            gap = current_daily_tables - target_daily

            self._write_data_row(ws, current_row, segment_label,
                                 prev_daily_tables, target_daily, current_daily_tables, gap,
                                 "时段", number_format='0.00',
                                 notes=f"日均桌数 (目标+{daily_target:.1f})")
            current_row += 1

        # === Row 7: 外卖挑战 ===
        store_takeout = takeout_data.get(store_id, {}) if takeout_data else {}
        current_mtd_total = store_takeout.get('current_mtd_total', 0)
        prev_year_month_total = store_takeout.get('prev_year_month_total', 0)
        prev_year_month_days = store_takeout.get('prev_year_month_days', 0)

        # Normalize previous year data to current MTD period
        prev_year_normalized = self._normalize_to_mtd(prev_year_month_total, prev_year_month_days, num_days)

        # Check for fixed takeout target
        fixed_takeout_target = get_takeout_target(store_id, target_date_str)
        if fixed_takeout_target is not None:
            # Fixed target is in 万加币 (10k CAD), convert to CAD and prorate to MTD
            target_monthly_cad = fixed_takeout_target * 10000
            target_mtd = self._prorate_monthly_target(target_monthly_cad, target_dt.year, target_dt.month, num_days)
            monthly_gap = current_mtd_total - target_mtd

            self._write_data_row(ws, current_row, "外卖收入",
                                 prev_year_normalized, target_mtd, current_mtd_total, monthly_gap,
                                 "外卖", number_format='"$"#,##0',
                                 notes=f"月目标 {fixed_takeout_target:.2f}万 (MTD)")
        else:
            # Legacy: daily improvement-based calculation
            prev_month_days = store_takeout.get('prev_year_month_days', 30) or 30
            current_days = store_takeout.get('current_days', num_days) or num_days

            prev_daily_avg = prev_year_month_total / prev_month_days if prev_month_days > 0 else 0
            daily_improvement_cad = get_takeout_daily_improvement_cad(target_dt.year)
            daily_target = prev_daily_avg + daily_improvement_cad
            current_daily_avg = current_mtd_total / current_days if current_days > 0 else 0
            daily_gap = current_daily_avg - daily_target

            self._write_data_row(ws, current_row, "外卖收入",
                                 prev_daily_avg, daily_target, current_daily_avg, daily_gap,
                                 "外卖", number_format='"$"#,##0.00',
                                 notes="日均 (目标+$200 USD)")
        current_row += 1

        # Merge store name cell across all rows
        ws.merge_cells(start_row=section_start_row, start_column=1,
                       end_row=current_row - 1, end_column=1)
        store_cell = ws.cell(row=section_start_row, column=1, value=store_name)
        store_cell.font = Font(bold=True, size=11, color="FFFFFF")
        store_cell.fill = store_fill
        store_cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation=0)
        store_cell.border = self.thin_border

        # Apply borders to merged area
        for r in range(section_start_row, current_row):
            ws.cell(row=r, column=1).border = self.thin_border

        return current_row

    def _write_data_row(self, ws: Worksheet, row: int, challenge_type: str,
                        prev_value: Any, target_value: Any, current_value: Any,
                        gap_value: Any, section: str, number_format: str = '0.00',
                        notes: str = "") -> None:
        """Write a single data row with formatting."""
        # Get section color
        section_fill = PatternFill(
            start_color=self.SECTION_COLORS.get(section, "FFFFFF"),
            end_color=self.SECTION_COLORS.get(section, "FFFFFF"),
            fill_type="solid"
        )

        # Column B - Challenge type
        cell_b = ws.cell(row=row, column=2, value=challenge_type)
        cell_b.fill = section_fill
        cell_b.alignment = self.center_alignment
        cell_b.border = self.thin_border

        # Column C - Previous year
        cell_c = ws.cell(row=row, column=3, value=prev_value)
        cell_c.alignment = self.center_alignment
        cell_c.border = self.thin_border
        if isinstance(prev_value, (int, float)):
            cell_c.number_format = number_format

        # Column D - Target
        cell_d = ws.cell(row=row, column=4, value=target_value)
        cell_d.alignment = self.center_alignment
        cell_d.border = self.thin_border
        if isinstance(target_value, (int, float)):
            cell_d.number_format = number_format

        # Column E - Current year
        cell_e = ws.cell(row=row, column=5, value=current_value)
        cell_e.alignment = self.center_alignment
        cell_e.border = self.thin_border
        if isinstance(current_value, (int, float)):
            cell_e.number_format = number_format

        # Column F - Gap (with color)
        cell_f = ws.cell(row=row, column=6, value=gap_value)
        cell_f.alignment = self.center_alignment
        cell_f.border = self.thin_border
        if isinstance(gap_value, (int, float)):
            cell_f.number_format = number_format
            cell_f.fill = self.green_fill if gap_value >= 0 else self.red_fill
        elif gap_value == "N/A":
            cell_f.fill = self.excluded_fill

        # Column G - Notes
        cell_g = ws.cell(row=row, column=7, value=notes)
        cell_g.alignment = self.center_alignment
        cell_g.border = self.thin_border
        cell_g.font = Font(size=9, color="666666")

    def _apply_column_widths(self, ws: Worksheet) -> None:
        """Apply column widths."""
        widths = {
            'A': 12,   # Store name
            'B': 18,   # Challenge type
            'C': 14,   # Previous year
            'D': 14,   # Target
            'E': 14,   # Current year
            'F': 12,   # Gap
            'G': 25    # Notes
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def _add_separator_row(self, ws: Worksheet, row: int) -> int:
        """Add a dark separator row between stores."""
        separator_fill = PatternFill(
            start_color=self.SEPARATOR_COLOR,
            end_color=self.SEPARATOR_COLOR,
            fill_type="solid"
        )
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col, value="")
            cell.fill = separator_fill
            cell.border = self.thin_border

        # Set row height to be thin
        ws.row_dimensions[row].height = 6

        return row + 1

    def _add_canada_summary(self, ws: Worksheet, store_data: List[Dict],
                            start_row: int, num_days: int,
                            time_segment_data: Dict, takeout_data: Dict,
                            target_dt: datetime) -> int:
        """
        Add Canada total summary section at the end.

        Returns:
            Next available row after summary section
        """
        current_row = start_row
        section_start_row = start_row
        target_date_str = target_dt.strftime('%Y-%m-%d')
        use_absolute = is_using_absolute_targets(target_date_str)

        # Use all stores for Canada summary (matches individual store rows)
        all_stores = store_data
        total_seats = sum(STORE_SEATING_CAPACITY.get(s['store_id'], 50) for s in all_stores)

        # === Row 1: 翻台率 (weighted average by seating capacity) ===
        prev_weighted_turnover = sum(
            float(s.get('prev_avg_turnover_rate', 0)) * STORE_SEATING_CAPACITY.get(s['store_id'], 50)
            for s in all_stores
        ) / total_seats if total_seats > 0 else 0

        current_weighted_turnover = sum(
            float(s.get('current_avg_turnover_rate', 0)) * STORE_SEATING_CAPACITY.get(s['store_id'], 50)
            for s in all_stores
        ) / total_seats if total_seats > 0 else 0

        # Use weighted average of fixed targets when available
        if use_absolute:
            target_weighted_turnover = sum(
                get_store_turnover_target(s['store_id'], float(s.get('prev_avg_turnover_rate', 0)), target_date_str) *
                STORE_SEATING_CAPACITY.get(s['store_id'], 50)
                for s in all_stores
            ) / total_seats if total_seats > 0 else 0
        else:
            target_weighted_turnover = prev_weighted_turnover + DEFAULT_TURNOVER_IMPROVEMENT

        turnover_gap = current_weighted_turnover - target_weighted_turnover

        self._write_summary_row(ws, current_row, "翻台率 (加权平均)",
                                prev_weighted_turnover, target_weighted_turnover,
                                current_weighted_turnover, turnover_gap,
                                number_format='0.00', notes="")
        current_row += 1

        # === Row 2: 桌数 (sum of all stores) ===
        prev_tables_total = sum(
            int(float(s.get('prev_avg_turnover_rate', 0)) * STORE_SEATING_CAPACITY.get(s['store_id'], 50) * num_days)
            for s in all_stores
        )
        target_tables_total = sum(
            int(get_store_turnover_target(s['store_id'], float(s.get('prev_avg_turnover_rate', 0)), target_date_str) *
                STORE_SEATING_CAPACITY.get(s['store_id'], 50) * num_days)
            for s in all_stores
        )
        current_tables_total = sum(
            int(float(s.get('current_avg_turnover_rate', 0)) * STORE_SEATING_CAPACITY.get(s['store_id'], 50) * num_days)
            for s in all_stores
        )
        tables_gap = current_tables_total - target_tables_total

        self._write_summary_row(ws, current_row, "桌数 (合计)",
                                prev_tables_total, target_tables_total,
                                current_tables_total, tables_gap,
                                number_format='0', notes="")
        current_row += 1

        # === Rows 3-6: 时段 (sum of all stores) ===
        for segment_config in TIME_SEGMENT_CONFIG:
            segment_label = segment_config['label']
            segment_key = segment_config['key']
            is_slow = segment_config['type'] == 'slow'

            prev_total = 0
            current_total = 0
            target_total = 0

            for store in all_stores:
                store_id = store.get('store_id', 0)
                store_ts_data = time_segment_data.get(store_id, {}).get('time_segments', {}) if time_segment_data else {}
                segment_data = store_ts_data.get(segment_label, {})

                prev_daily = segment_data.get('prev_total_tables', 0) / num_days if num_days > 0 else 0
                current_daily = segment_data.get('current_total_tables', 0) / num_days if num_days > 0 else 0

                prev_total += prev_daily
                current_total += current_daily

                # For January 2026, use absolute targets for slow times
                if use_absolute and is_slow:
                    absolute_target = get_absolute_time_segment_target(store_id, segment_key, target_date_str)
                    if absolute_target is not None:
                        target_total += absolute_target
                        continue

                # Fall back to improvement-based calculation
                if is_slow and segment_config['targets']:
                    daily_target_improvement = segment_config['targets'].get(store_id, DEFAULT_SLOW_TIME_TARGET)
                    target_total += prev_daily + daily_target_improvement
                else:
                    daily_target_improvement = self._calculate_busy_time_target(
                        store_id, store, time_segment_data, segment_key, num_days, target_date_str
                    )
                    target_total += prev_daily + daily_target_improvement

            gap = current_total - target_total

            self._write_summary_row(ws, current_row, segment_label,
                                    prev_total, target_total, current_total, gap,
                                    number_format='0.00', notes="日均合计")
            current_row += 1

        # === Row 7: 外卖 (sum of all stores) ===
        current_takeout_total = 0
        target_takeout_total = 0
        prev_year_takeout_normalized_total = 0

        # Check if using fixed takeout targets
        if use_absolute:
            for store in all_stores:
                store_id = store.get('store_id', 0)
                store_takeout = takeout_data.get(store_id, {}) if takeout_data else {}
                current_mtd_total = store_takeout.get('current_mtd_total', 0)
                prev_year_month_total = store_takeout.get('prev_year_month_total', 0)
                prev_year_month_days = store_takeout.get('prev_year_month_days', 0)

                # Normalize previous year to current MTD period
                prev_year_normalized = self._normalize_to_mtd(
                    prev_year_month_total, prev_year_month_days, num_days
                )

                fixed_target = get_takeout_target(store_id, target_date_str)
                if fixed_target is not None:
                    # Prorate monthly target to MTD
                    target_takeout_total += self._prorate_monthly_target(
                        fixed_target * 10000, target_dt.year, target_dt.month, num_days
                    )
                current_takeout_total += current_mtd_total
                prev_year_takeout_normalized_total += prev_year_normalized

            takeout_gap = current_takeout_total - target_takeout_total

            self._write_summary_row(ws, current_row, "外卖收入 (合计)",
                                    prev_year_takeout_normalized_total, target_takeout_total, current_takeout_total, takeout_gap,
                                    number_format='"$"#,##0', notes="MTD目标合计")
        else:
            # Legacy: daily improvement-based calculation
            prev_takeout_total = 0
            daily_improvement_cad = get_takeout_daily_improvement_cad(target_dt.year)

            for store in all_stores:
                store_id = store.get('store_id', 0)
                store_takeout = takeout_data.get(store_id, {}) if takeout_data else {}

                prev_month_total = store_takeout.get('prev_year_month_total', 0)
                prev_month_days = store_takeout.get('prev_year_month_days', 30) or 30
                current_mtd_total = store_takeout.get('current_mtd_total', 0)
                current_days = store_takeout.get('current_days', num_days) or num_days

                prev_daily = prev_month_total / prev_month_days if prev_month_days > 0 else 0
                current_daily = current_mtd_total / current_days if current_days > 0 else 0

                prev_takeout_total += prev_daily
                current_takeout_total += current_daily
                target_takeout_total += prev_daily + daily_improvement_cad

            takeout_gap = current_takeout_total - target_takeout_total

            self._write_summary_row(ws, current_row, "外卖收入 (合计)",
                                    prev_takeout_total, target_takeout_total,
                                    current_takeout_total, takeout_gap,
                                    number_format='"$"#,##0.00', notes="日均合计")
        current_row += 1

        # Merge "加拿大合计" cell across all rows
        ws.merge_cells(start_row=section_start_row, start_column=1,
                       end_row=current_row - 1, end_column=1)
        summary_cell = ws.cell(row=section_start_row, column=1, value="加拿大合计")
        summary_cell.font = Font(bold=True, size=12, color="FFFFFF")
        summary_cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        summary_cell.alignment = Alignment(horizontal="center", vertical="center")
        summary_cell.border = self.thin_border

        # Apply borders to merged area
        for r in range(section_start_row, current_row):
            ws.cell(row=r, column=1).border = self.thin_border

        return current_row

    def _write_summary_row(self, ws: Worksheet, row: int, challenge_type: str,
                           prev_value: Any, target_value: Any, current_value: Any,
                           gap_value: Any, number_format: str = '0.00',
                           notes: str = "") -> None:
        """Write a summary row with special formatting."""
        # Column B - Challenge type
        cell_b = ws.cell(row=row, column=2, value=challenge_type)
        cell_b.fill = self.summary_fill
        cell_b.alignment = self.center_alignment
        cell_b.border = self.thin_border
        cell_b.font = Font(bold=True)

        # Column C - Previous year
        cell_c = ws.cell(row=row, column=3, value=prev_value)
        cell_c.fill = self.summary_fill
        cell_c.alignment = self.center_alignment
        cell_c.border = self.thin_border
        cell_c.font = Font(bold=True)
        if isinstance(prev_value, (int, float)):
            cell_c.number_format = number_format

        # Column D - Target
        cell_d = ws.cell(row=row, column=4, value=target_value)
        cell_d.fill = self.summary_fill
        cell_d.alignment = self.center_alignment
        cell_d.border = self.thin_border
        cell_d.font = Font(bold=True)
        if isinstance(target_value, (int, float)):
            cell_d.number_format = number_format

        # Column E - Current year
        cell_e = ws.cell(row=row, column=5, value=current_value)
        cell_e.fill = self.summary_fill
        cell_e.alignment = self.center_alignment
        cell_e.border = self.thin_border
        cell_e.font = Font(bold=True)
        if isinstance(current_value, (int, float)):
            cell_e.number_format = number_format

        # Column F - Gap (with color)
        cell_f = ws.cell(row=row, column=6, value=gap_value)
        cell_f.alignment = self.center_alignment
        cell_f.border = self.thin_border
        cell_f.font = Font(bold=True)
        if isinstance(gap_value, (int, float)):
            cell_f.number_format = number_format
            cell_f.fill = self.green_fill if gap_value >= 0 else self.red_fill

        # Column G - Notes
        cell_g = ws.cell(row=row, column=7, value=notes)
        cell_g.fill = self.summary_fill
        cell_g.alignment = self.center_alignment
        cell_g.border = self.thin_border
        cell_g.font = Font(size=9, color="666666", bold=True)

    def _get_mtd_data(self, start_dt: datetime, end_dt: datetime) -> List[Dict[str, Any]]:
        """Get MTD store performance data from database."""
        try:
            store_data = self.data_provider.get_weekly_store_performance(
                start_dt.strftime('%Y-%m-%d'), end_dt.strftime('%Y-%m-%d'))

            if not store_data:
                self.logger.warning("No MTD store performance data found")
                return []

            self.logger.info(f"Retrieved MTD data for {len(store_data)} stores")
            return store_data

        except Exception as e:
            self.logger.error(f"Error getting MTD data: {str(e)}")
            return []

    def _calculate_busy_time_target(self, store_id: int, store: Dict,
                                     time_segment_data: Dict, segment_key: str,
                                     num_days: int, target_date: str = None) -> float:
        """
        Calculate busy time target from leftover after slow times.

        For January 2026 with absolute targets:
        - Total daily tables = target_turnover × seating_capacity
        - Leftover = total_daily_tables - afternoon_absolute - late_night_absolute
        - Busy time target = leftover × (this_segment_turnover / total_busy_turnover)

        For other periods (improvement-based):
        - Leftover = total_daily_improvement - afternoon_improvement - late_night_improvement
        - Busy time target = leftover × (this_segment_turnover / total_busy_turnover)
        """
        seating_capacity = STORE_SEATING_CAPACITY.get(store_id, 50)

        prev_turnover = float(store.get('prev_avg_turnover_rate', 0))
        target_turnover = get_store_turnover_target(store_id, prev_turnover, target_date)

        # Check if using absolute targets (January 2026)
        use_absolute = is_using_absolute_targets(target_date) if target_date else False

        if use_absolute:
            # For absolute targets, calculate leftover from total daily tables
            target_daily_tables = target_turnover * seating_capacity
            afternoon_absolute = get_absolute_time_segment_target(store_id, 'afternoon', target_date) or 0
            late_night_absolute = get_absolute_time_segment_target(store_id, 'late_night', target_date) or 0
            leftover = target_daily_tables - afternoon_absolute - late_night_absolute
        else:
            # For improvement-based, calculate leftover from improvement
            prev_daily_tables = prev_turnover * seating_capacity
            target_daily_tables = target_turnover * seating_capacity
            total_daily_improvement = target_daily_tables - prev_daily_tables

            afternoon_target = AFTERNOON_SLOW_TARGETS.get(store_id, DEFAULT_SLOW_TIME_TARGET)
            late_night_target = LATE_NIGHT_TARGETS.get(store_id, DEFAULT_SLOW_TIME_TARGET)
            leftover = total_daily_improvement - afternoon_target - late_night_target

        if leftover <= 0:
            return 0

        store_ts_data = time_segment_data.get(store_id, {}).get('time_segments', {}) if time_segment_data else {}
        morning_label = TIME_SEGMENT_LABELS['morning']
        evening_label = TIME_SEGMENT_LABELS['evening']

        morning_data = store_ts_data.get(morning_label, {})
        evening_data = store_ts_data.get(evening_label, {})

        morning_turnover = morning_data.get('prev_avg_turnover', 0)
        evening_turnover = evening_data.get('prev_avg_turnover', 0)
        total_busy_turnover = morning_turnover + evening_turnover

        if total_busy_turnover <= 0:
            return leftover / 2

        if segment_key == 'morning':
            return leftover * morning_turnover / total_busy_turnover
        else:
            return leftover * evening_turnover / total_busy_turnover

    def generate_detail_worksheet(self, workbook: Workbook, target_date: str) -> Worksheet:
        """
        Generate detailed daily data worksheet for verification.

        Shows daily turnover rate and tables for both current and previous year.
        """
        ws = workbook.create_sheet("每日详细数据")

        # Parse dates
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        start_dt = target_dt.replace(day=1)
        num_days = target_dt.day
        prev_year = target_dt.year - 1

        # Get daily data from database
        daily_data = self._get_daily_detail_data(start_dt, target_dt)

        if not daily_data:
            ws.cell(row=1, column=1, value="无数据")
            return ws

        # Generate headers
        current_row = 1

        # Title
        title_cell = ws.cell(row=current_row, column=1, value=f"每日详细数据 ({start_dt.strftime('%Y-%m-%d')} 至 {target_dt.strftime('%Y-%m-%d')})")
        title_cell.font = Font(bold=True, size=14)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        current_row += 2

        # Column headers
        headers = ["门店", "日期", f"{prev_year}翻台率", f"{target_dt.year}翻台率", "翻台率差距",
                   f"{prev_year}桌数", f"{target_dt.year}桌数", "桌数差距"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

        current_row += 1

        # Store name mapping
        store_names = {
            1: '加拿大一店', 2: '加拿大二店', 3: '加拿大三店', 4: '加拿大四店',
            5: '加拿大五店', 6: '加拿大六店', 7: '加拿大七店', 8: '加拿大八店'
        }

        # Write data rows
        current_store = None
        store_start_row = current_row

        for row_data in daily_data:
            store_id = row_data['store_id']
            store_name = store_names.get(store_id, f'店{store_id}')

            # Add separator between stores
            if current_store is not None and store_id != current_store:
                # Merge store name cells for previous store
                if current_row - 1 > store_start_row:
                    ws.merge_cells(start_row=store_start_row, start_column=1,
                                   end_row=current_row - 1, end_column=1)

                # Add separator row
                for col in range(1, 9):
                    cell = ws.cell(row=current_row, column=col, value="")
                    cell.fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
                ws.row_dimensions[current_row].height = 4
                current_row += 1
                store_start_row = current_row

            current_store = store_id

            # Store name (will be merged later)
            ws.cell(row=current_row, column=1, value=store_name).border = self.thin_border

            # Date
            date_str = f"{row_data['month']}月{row_data['day']}日"
            ws.cell(row=current_row, column=2, value=date_str).border = self.thin_border

            # Previous year turnover
            prev_turnover = row_data.get('prev_turnover', 0) or 0
            cell = ws.cell(row=current_row, column=3, value=prev_turnover)
            cell.number_format = '0.00'
            cell.border = self.thin_border

            # Current year turnover
            curr_turnover = row_data.get('curr_turnover', 0) or 0
            cell = ws.cell(row=current_row, column=4, value=curr_turnover)
            cell.number_format = '0.00'
            cell.border = self.thin_border

            # Turnover gap
            turnover_gap = curr_turnover - prev_turnover
            cell = ws.cell(row=current_row, column=5, value=turnover_gap)
            cell.number_format = '+0.00;-0.00;0.00'
            cell.border = self.thin_border
            cell.fill = self.green_fill if turnover_gap >= 0 else self.red_fill

            # Previous year tables
            prev_tables = row_data.get('prev_tables', 0) or 0
            cell = ws.cell(row=current_row, column=6, value=prev_tables)
            cell.number_format = '0'
            cell.border = self.thin_border

            # Current year tables
            curr_tables = row_data.get('curr_tables', 0) or 0
            cell = ws.cell(row=current_row, column=7, value=curr_tables)
            cell.number_format = '0'
            cell.border = self.thin_border

            # Tables gap
            tables_gap = curr_tables - prev_tables
            cell = ws.cell(row=current_row, column=8, value=tables_gap)
            cell.number_format = '+0;-0;0'
            cell.border = self.thin_border
            cell.fill = self.green_fill if tables_gap >= 0 else self.red_fill

            current_row += 1

        # Merge last store's name cells
        if current_row - 1 > store_start_row:
            ws.merge_cells(start_row=store_start_row, start_column=1,
                           end_row=current_row - 1, end_column=1)

        # Apply column widths
        widths = {'A': 12, 'B': 10, 'C': 12, 'D': 12, 'E': 12, 'F': 10, 'G': 10, 'H': 10}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        # Freeze panes
        ws.freeze_panes = "C4"

        return ws

    def _get_daily_detail_data(self, start_dt: datetime, end_dt: datetime) -> list:
        """Get daily detail data for both current and previous year."""
        prev_year = end_dt.year - 1
        curr_year = end_dt.year
        month = end_dt.month
        max_day = end_dt.day

        sql = """
        WITH current_year AS (
            SELECT store_id, EXTRACT(DAY FROM date) as day,
                   turnover_rate, tables_served_validated
            FROM daily_report
            WHERE EXTRACT(YEAR FROM date) = %s
              AND EXTRACT(MONTH FROM date) = %s
              AND EXTRACT(DAY FROM date) <= %s
        ),
        prev_year AS (
            SELECT store_id, EXTRACT(DAY FROM date) as day,
                   turnover_rate, tables_served_validated
            FROM daily_report
            WHERE EXTRACT(YEAR FROM date) = %s
              AND EXTRACT(MONTH FROM date) = %s
              AND EXTRACT(DAY FROM date) <= %s
        )
        SELECT
            COALESCE(cy.store_id, py.store_id) as store_id,
            COALESCE(cy.day, py.day) as day,
            py.turnover_rate as prev_turnover,
            cy.turnover_rate as curr_turnover,
            py.tables_served_validated as prev_tables,
            cy.tables_served_validated as curr_tables
        FROM current_year cy
        FULL OUTER JOIN prev_year py
            ON cy.store_id = py.store_id AND cy.day = py.day
        WHERE COALESCE(cy.store_id, py.store_id) BETWEEN 1 AND 8
        ORDER BY COALESCE(cy.store_id, py.store_id), COALESCE(cy.day, py.day)
        """

        try:
            results = self.data_provider.db_manager.fetch_all(sql, (
                curr_year, month, max_day,
                prev_year, month, max_day
            ))

            # Add month to results
            return [
                {
                    'store_id': r['store_id'],
                    'day': int(r['day']),
                    'month': month,
                    'prev_turnover': r['prev_turnover'],
                    'curr_turnover': r['curr_turnover'],
                    'prev_tables': r['prev_tables'],
                    'curr_tables': r['curr_tables']
                }
                for r in results
            ]

        except Exception as e:
            self.logger.error(f"Error getting daily detail data: {e}")
            return []
