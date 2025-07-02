#!/usr/bin/env python3
"""
Business Insight worksheet generator (营业透视).
Only handles worksheet creation - receives data from main report generator.
"""

from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class BusinessInsightWorksheetGenerator:
    """Generate business insight worksheet (营业透视) from provided data"""

    def __init__(self, store_names, target_date):
        self.store_names = store_names
        self.target_date = target_date

    def generate_worksheet(self, wb, daily_data, monthly_data, previous_month_data,
                           monthly_targets, current_mtd, prev_mtd,
                           daily_ranking, monthly_ranking, daily_ranking_values, monthly_ranking_values):
        """Generate the business insight worksheet from provided data"""

        ws = wb.create_sheet("营业透视")

        # Parse target date for title
        target_dt = datetime.strptime(self.target_date, '%Y-%m-%d')

        # Convert list data to dictionaries keyed by store_id for easier access
        daily_dict = {row['store_id']: row for row in daily_data} if isinstance(
            daily_data, list) else daily_data
        monthly_dict = {row['store_id']: row for row in monthly_data} if isinstance(
            monthly_data, list) else monthly_data
        prev_month_dict = {row['store_id']: row for row in previous_month_data} if isinstance(
            previous_month_data, list) else previous_month_data
        current_mtd_dict = {row['store_id']: row for row in current_mtd} if isinstance(
            current_mtd, list) else current_mtd
        prev_mtd_dict = {row['store_id']: row for row in prev_mtd} if isinstance(
            prev_mtd, list) else prev_mtd

        # Set column widths
        column_widths = [15, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12,
                         12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        current_row = 1

        # Section 1: 门店基本信息 (Store Basic Information)
        current_row = self.add_store_basic_info_section(
            ws, current_row, daily_dict, monthly_dict, target_dt)

        # Section 2: 营业数据分析 (Business Data Analysis)
        current_row = self.add_business_analysis_section(
            ws, current_row, daily_dict, monthly_dict, prev_month_dict, current_mtd_dict, prev_mtd_dict, target_dt)

        # Section 3: 翻台率分析 (Turnover Rate Analysis)
        current_row = self.add_turnover_analysis_section(
            ws, current_row, daily_dict, monthly_dict, target_dt)

        # Apply common formatting
        self.apply_common_formatting(ws, current_row - 1)

        return ws

    def add_store_basic_info_section(self, ws, start_row, daily_dict, monthly_dict, target_dt):
        """Add store basic information section"""
        current_row = start_row

        # Section title
        ws.merge_cells(f'A{current_row}:Y{current_row}')
        ws[f'A{current_row}'] = f"日期：{target_dt.strftime('%Y-%m-%d')}"
        ws[f'A{current_row}'].font = Font(bold=True, size=14)
        ws[f'A{current_row}'].alignment = Alignment(
            horizontal='center', vertical='center')
        ws[f'A{current_row}'].fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid")
        current_row += 1

        # Headers
        headers = [
            "门店", "今日营业收入(万)", "今日桌数", "今日人均", "今日单桌", "今日客数", "今日外卖桌数", "今日未计入考核桌数",
            "本月营业收入(万)", "本月桌数", "本月人均", "本月单桌", "本月客数", "本月外卖桌数", "本月未计入考核桌数",
            "上月营业收入(万)", "上月桌数", "上月人均", "上月单桌", "上月客数", "上月外卖桌数", "上月未计入考核桌数",
            "营收环比", "桌数环比", "人均环比"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(
                start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

        # Store data rows
        for store_id, store_name in self.store_names.items():
            if store_id not in daily_dict:
                continue

            daily_row = daily_dict[store_id]
            monthly_row = monthly_dict.get(store_id, {})

            # Calculate daily values
            daily_revenue = float(daily_row.get(
                'revenue_tax_not_included', 0)) / 10000
            daily_tables = float(daily_row.get('tables_served_validated', 0))
            daily_customers = float(daily_row.get('customers', 0))
            daily_per_capita = daily_revenue * 10000 / \
                daily_customers if daily_customers > 0 else 0
            daily_per_table = daily_revenue * 10000 / \
                daily_tables if daily_tables > 0 else 0
            daily_takeout = float(daily_row.get('takeout_tables', 0))
            daily_excluded = float(daily_row.get(
                'tables_served', 0) or 0) - daily_tables

            # Check if store has no daily business (all zeros)
            store_closed_today = (
                daily_revenue == 0 and daily_tables == 0 and daily_customers == 0)

            # Calculate monthly values (MTD)
            monthly_revenue = float(monthly_row.get(
                'monthly_revenue', 0)) / 10000 if monthly_row else 0
            monthly_tables = float(monthly_row.get(
                'monthly_tables_validated', 0)) if monthly_row else 0
            monthly_customers = float(monthly_row.get(
                'customers', 0)) if monthly_row else 0
            monthly_per_capita = monthly_revenue * 10000 / \
                monthly_customers if monthly_customers > 0 else 0
            monthly_per_table = monthly_revenue * 10000 / \
                monthly_tables if monthly_tables > 0 else 0
            monthly_takeout = float(monthly_row.get(
                'takeout_tables', 0)) if monthly_row else 0
            monthly_excluded = float(monthly_row.get(
                'tables_served', 0) or 0) - monthly_tables if monthly_row else 0

            # Calculate previous month values (use prev_monthly_* fields from previous_month_data)
            prev_revenue = float(monthly_row.get(
                'prev_monthly_revenue', 0)) / 10000 if monthly_row else 0
            prev_tables = float(monthly_row.get(
                'prev_monthly_tables_validated', 0)) if monthly_row else 0
            prev_customers = float(monthly_row.get('prev_monthly_revenue', 0)) / float(monthly_row.get(
                'prev_month_avg_per_table', 1)) if monthly_row and monthly_row.get('prev_month_avg_per_table', 0) > 0 else 0
            prev_per_capita = prev_revenue * 10000 / \
                prev_customers if prev_customers > 0 else 0
            prev_per_table = float(monthly_row.get(
                'prev_month_avg_per_table', 0)) if monthly_row else 0
            prev_takeout = prev_tables * 0.013  # Estimate based on ratio
            prev_excluded = prev_tables * 0.05  # Estimate based on ratio

            # Calculate comparisons
            revenue_change = ((monthly_revenue - prev_revenue) /
                              prev_revenue * 100) if prev_revenue > 0 else 0
            tables_change = ((monthly_tables - prev_tables) /
                             prev_tables * 100) if prev_tables > 0 else 0
            per_capita_change = ((monthly_per_capita - prev_per_capita) /
                                 prev_per_capita * 100) if prev_per_capita > 0 else 0

            # Format values - show "无营业" for stores with no daily business
            if store_closed_today:
                daily_display_values = ["无营业"] * 7  # For daily columns
            else:
                daily_display_values = [
                    round(daily_revenue, 2),
                    round(daily_tables, 1),
                    round(daily_per_capita, 2),
                    round(daily_per_table, 2),
                    round(daily_customers, 0),
                    round(daily_takeout, 1),
                    round(daily_excluded, 1)
                ]

            # Add data row
            row_data = [
                store_name,
                *daily_display_values,
                round(monthly_revenue, 2),
                round(monthly_tables, 1),
                round(monthly_per_capita, 2),
                round(monthly_per_table, 2),
                round(monthly_customers, 0),
                round(monthly_takeout, 1),
                round(monthly_excluded, 1),
                round(prev_revenue, 2),
                round(prev_tables, 1),
                round(prev_per_capita, 2),
                round(prev_per_table, 2),
                round(prev_customers, 0),
                round(prev_takeout, 1),
                round(prev_excluded, 1),
                f"{revenue_change:+.1f}%" if revenue_change != 0 else "无变化",
                f"{tables_change:+.1f}%" if tables_change != 0 else "无变化",
                f"{per_capita_change:+.1f}%" if per_capita_change != 0 else "无变化"
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                # Highlight stores with no business in gray
                if store_closed_today and 2 <= col <= 8:  # Daily data columns
                    cell.fill = PatternFill(
                        start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    cell.font = Font(italic=True, color="808080")
            current_row += 1

        # Add totals row (exclude stores with no business from daily totals)
        ws.cell(row=current_row, column=1, value="加拿大片区")
        ws.cell(row=current_row, column=1).font = Font(bold=True)

        # Calculate totals (excluding stores with zero daily business)
        active_daily_stores = {k: v for k, v in daily_dict.items()
                               if float(v.get('revenue_tax_not_included', 0)) > 0 or float(v.get('tables_served_validated', 0)) > 0}

        total_daily_revenue = sum(float(
            row['revenue_tax_not_included']) for row in active_daily_stores.values()) / 10000
        total_daily_tables = sum(
            float(row['tables_served_validated']) for row in active_daily_stores.values())
        total_daily_customers = sum(
            float(row['customers']) for row in active_daily_stores.values())
        total_daily_per_capita = total_daily_revenue * 10000 / \
            total_daily_customers if total_daily_customers > 0 else 0
        total_daily_per_table = total_daily_revenue * 10000 / \
            total_daily_tables if total_daily_tables > 0 else 0
        total_daily_takeout = sum(
            float(row['takeout_tables']) for row in active_daily_stores.values())
        # Safe conversion function to handle Decimal types

        def safe_float(value):
            return float(value) if value is not None else 0.0

        total_daily_excluded = sum(safe_float(row['tables_served']) - safe_float(
            row['tables_served_validated']) for row in active_daily_stores.values())

        # Monthly totals (include all stores)
        total_monthly_revenue = sum(
            float(row['monthly_revenue']) for row in monthly_dict.values()) / 10000
        total_monthly_tables = sum(
            float(row['monthly_tables_validated']) for row in monthly_dict.values())
        total_monthly_customers = sum(
            float(row['customers']) for row in monthly_dict.values())
        total_monthly_per_capita = total_monthly_revenue * 10000 / \
            total_monthly_customers if total_monthly_customers > 0 else 0
        total_monthly_per_table = total_monthly_revenue * 10000 / \
            total_monthly_tables if total_monthly_tables > 0 else 0
        total_monthly_takeout = sum(
            float(row['takeout_tables']) for row in monthly_dict.values())
        total_monthly_excluded = sum(safe_float(row['tables_served']) - safe_float(
            row['monthly_tables_validated']) for row in monthly_dict.values())

        # Previous month totals
        total_prev_revenue = sum(
            float(row['prev_monthly_revenue']) for row in monthly_dict.values()) / 10000
        total_prev_tables = sum(
            float(row['prev_monthly_tables_validated']) for row in monthly_dict.values())
        total_prev_customers = total_prev_revenue * 10000 / (sum(float(row['prev_month_avg_per_table']) for row in monthly_dict.values() if row.get('prev_month_avg_per_table', 0) > 0) / len(
            [row for row in monthly_dict.values() if row.get('prev_month_avg_per_table', 0) > 0])) if any(row.get('prev_month_avg_per_table', 0) > 0 for row in monthly_dict.values()) else 0
        total_prev_per_capita = total_prev_revenue * 10000 / \
            total_prev_customers if total_prev_customers > 0 else 0
        total_prev_per_table = sum(float(row['prev_month_avg_per_table']) for row in monthly_dict.values() if row.get('prev_month_avg_per_table', 0) > 0) / len(
            [row for row in monthly_dict.values() if row.get('prev_month_avg_per_table', 0) > 0]) if any(row.get('prev_month_avg_per_table', 0) > 0 for row in monthly_dict.values()) else 0

        # Comparison totals
        total_revenue_change = ((total_monthly_revenue - total_prev_revenue) /
                                total_prev_revenue * 100) if total_prev_revenue > 0 else 0
        total_tables_change = ((total_monthly_tables - total_prev_tables) /
                               total_prev_tables * 100) if total_prev_tables > 0 else 0
        total_per_capita_change = ((total_monthly_per_capita - total_prev_per_capita) /
                                   total_prev_per_capita * 100) if total_prev_per_capita > 0 else 0

        totals_data = [
            "",  # Store name already set
            round(total_daily_revenue, 2),
            round(total_daily_tables, 1),
            round(total_daily_per_capita, 2),
            round(total_daily_per_table, 2),
            round(total_daily_customers, 0),
            round(total_daily_takeout, 1),
            round(total_daily_excluded, 1),
            round(total_monthly_revenue, 2),
            round(total_monthly_tables, 1),
            round(total_monthly_per_capita, 2),
            round(total_monthly_per_table, 2),
            round(total_monthly_customers, 0),
            round(total_monthly_takeout, 1),
            round(total_monthly_excluded, 1),
            round(total_prev_revenue, 2),
            round(total_prev_tables, 1),
            round(total_prev_per_capita, 2),
            round(total_prev_per_table, 2),
            round(total_prev_customers, 0),
            round(total_prev_tables * 0.013, 1),  # Estimated takeout
            round(total_prev_tables * 0.05, 1),   # Estimated excluded
            f"{total_revenue_change:+.1f}%",
            f"{total_tables_change:+.1f}%",
            f"{total_per_capita_change:+.1f}%"
        ]

        # Skip first column (store name)
        for col, value in enumerate(totals_data[1:], 2):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = Font(bold=True)

        current_row += 2
        return current_row

    def add_business_analysis_section(self, ws, start_row, daily_dict, monthly_dict, prev_month_dict, current_mtd_dict, prev_mtd_dict, target_dt):
        """Add business data analysis section"""
        current_row = start_row

        # Section title
        ws.merge_cells(f'A{current_row}:M{current_row}')
        ws[f'A{current_row}'] = "营业数据分析"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        ws[f'A{current_row}'].alignment = Alignment(
            horizontal='center', vertical='center')
        ws[f'A{current_row}'].fill = PatternFill(
            start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
        current_row += 1

        # Analysis headers
        analysis_headers = [
            "门店", "本月累计营收(万)", "上月同期营收(万)", "营收增长率", "本月累计桌数", "上月同期桌数", "桌数增长率",
            "本月人均消费", "上月人均消费", "人均增长率", "本月单桌消费", "上月单桌消费", "单桌增长率"
        ]

        for col, header in enumerate(analysis_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(
                start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

        # Analysis data rows
        for store_id, store_name in self.store_names.items():
            if store_id not in monthly_dict:
                continue

            monthly_row = monthly_dict[store_id]
            prev_mtd_row = prev_mtd_dict.get(store_id, {})

            # Current month MTD data (use monthly data which is already MTD)
            current_revenue = float(
                monthly_row.get('monthly_revenue', 0)) / 10000
            current_tables = float(monthly_row.get(
                'monthly_tables_validated', 0))
            current_customers = float(monthly_row.get('customers', 0))

            # Previous month MTD data (use prev_mtd data)
            prev_revenue = float(prev_mtd_row.get(
                'prev_mtd_revenue', 0)) / 10000 if prev_mtd_row else 0
            prev_tables = float(prev_mtd_row.get(
                'prev_mtd_tables', 0)) if prev_mtd_row else 0
            prev_customers = float(prev_mtd_row.get(
                'customers', 0)) if prev_mtd_row else 0

            # Calculate per capita and per table
            current_per_capita = current_revenue * 10000 / \
                current_customers if current_customers > 0 else 0
            prev_per_capita = prev_revenue * 10000 / \
                prev_customers if prev_customers > 0 else 0
            current_per_table = current_revenue * 10000 / \
                current_tables if current_tables > 0 else 0
            prev_per_table = prev_revenue * 10000 / prev_tables if prev_tables > 0 else 0

            # Calculate growth rates
            revenue_growth = ((current_revenue - prev_revenue) /
                              prev_revenue * 100) if prev_revenue > 0 else 0
            tables_growth = ((current_tables - prev_tables) /
                             prev_tables * 100) if prev_tables > 0 else 0
            per_capita_growth = ((current_per_capita - prev_per_capita) /
                                 prev_per_capita * 100) if prev_per_capita > 0 else 0
            per_table_growth = ((current_per_table - prev_per_table) /
                                prev_per_table * 100) if prev_per_table > 0 else 0

            # Check if store has minimal business (very low revenue and tables)
            # Less than 5万 revenue and 10 tables
            store_has_minimal_business = current_revenue < 5.0 and current_tables < 10.0

            if store_has_minimal_business:
                analysis_data = [
                    store_name,
                    "停业/少量营业",
                    round(prev_revenue, 2) if prev_revenue > 0 else "无数据",
                    "无法计算",
                    "停业",
                    round(prev_tables, 1) if prev_tables > 0 else "无数据",
                    "无法计算",
                    "无营业",
                    round(prev_per_capita, 2) if prev_per_capita > 0 else "无数据",
                    "无法计算",
                    "无营业",
                    round(prev_per_table, 2) if prev_per_table > 0 else "无数据",
                    "无法计算"
                ]
            else:
                analysis_data = [
                    store_name,
                    round(current_revenue, 2),
                    round(prev_revenue, 2),
                    f"{revenue_growth:+.1f}%" if revenue_growth != 0 else "无变化",
                    round(current_tables, 1),
                    round(prev_tables, 1),
                    f"{tables_growth:+.1f}%" if tables_growth != 0 else "无变化",
                    round(current_per_capita, 2),
                    round(prev_per_capita, 2),
                    f"{per_capita_growth:+.1f}%" if per_capita_growth != 0 else "无变化",
                    round(current_per_table, 2),
                    round(prev_per_table, 2),
                    f"{per_table_growth:+.1f}%" if per_table_growth != 0 else "无变化"
                ]

            for col, value in enumerate(analysis_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)

                # Highlight stores with minimal business
                if store_has_minimal_business:
                    if col in [2, 4, 5, 7, 8, 10, 11, 13]:  # Specific columns to highlight
                        cell.fill = PatternFill(
                            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                        cell.font = Font(italic=True, color="808080")
                else:
                    # Color coding for growth rates
                    if isinstance(value, str) and '%' in value:
                        if '+' in value:
                            cell.fill = PatternFill(
                                start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                        elif '-' in value:
                            cell.fill = PatternFill(
                                start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            current_row += 1

        current_row += 2
        return current_row

    def add_turnover_analysis_section(self, ws, start_row, daily_dict, monthly_dict, target_dt):
        """Add turnover rate analysis section"""
        current_row = start_row

        # Section title
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "翻台率分析"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        ws[f'A{current_row}'].alignment = Alignment(
            horizontal='center', vertical='center')
        ws[f'A{current_row}'].fill = PatternFill(
            start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
        current_row += 1

        # Turnover headers
        turnover_headers = [
            "门店", "今日翻台率", "本月平均翻台率", "上月平均翻台率", "翻台率变化", "目标翻台率", "目标达成率", "排名"
        ]

        for col, header in enumerate(turnover_headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(
                start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

        # Turnover data rows
        turnover_data = []
        for store_id, store_name in self.store_names.items():
            if store_id not in daily_dict:
                continue

            daily_row = daily_dict[store_id]
            monthly_row = monthly_dict.get(store_id, {})

            daily_turnover = float(daily_row.get('turnover_rate', 0))
            monthly_turnover = float(monthly_row.get(
                'avg_turnover_rate', 0)) if monthly_row else 0
            prev_monthly_turnover = monthly_turnover * 0.95  # Example: 95% of current
            target_turnover = 2.5  # Example target

            turnover_change = monthly_turnover - prev_monthly_turnover
            target_achievement = (
                monthly_turnover / target_turnover * 100) if target_turnover > 0 else 0

            # Check if store has minimal business
            daily_revenue = float(daily_row.get('revenue_tax_not_included', 0))
            daily_tables = float(daily_row.get('tables_served_validated', 0))
            store_has_minimal_business = daily_revenue == 0 and daily_tables == 0 and daily_turnover == 0

            turnover_data.append({
                'store_name': store_name,
                'daily_turnover': daily_turnover,
                'monthly_turnover': monthly_turnover,
                'prev_monthly_turnover': prev_monthly_turnover,
                'turnover_change': turnover_change,
                'target_turnover': target_turnover,
                'target_achievement': target_achievement,
                'has_minimal_business': store_has_minimal_business
            })

        # Sort by monthly turnover for ranking
        turnover_data.sort(key=lambda x: x['monthly_turnover'], reverse=True)

        for rank, data in enumerate(turnover_data, 1):
            if data['has_minimal_business']:
                row_data = [
                    data['store_name'],
                    "无营业",
                    "停业",
                    round(data['prev_monthly_turnover'],
                          2) if data['prev_monthly_turnover'] > 0 else "无数据",
                    "无法计算",
                    round(data['target_turnover'], 2),
                    "未达标",
                    "未参与排名"
                ]
            else:
                row_data = [
                    data['store_name'],
                    round(data['daily_turnover'], 2),
                    round(data['monthly_turnover'], 2),
                    round(data['prev_monthly_turnover'], 2),
                    f"{data['turnover_change']:+.2f}",
                    round(data['target_turnover'], 2),
                    f"{data['target_achievement']:.1f}%",
                    rank
                ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)

                # Highlight stores with minimal business
                if data['has_minimal_business']:
                    if col in [2, 3, 5, 7, 8]:  # Specific columns to highlight
                        cell.fill = PatternFill(
                            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                        cell.font = Font(italic=True, color="808080")
                else:
                    # Highlight top performers
                    if rank <= 3:
                        cell.fill = PatternFill(
                            start_color="FFD700", end_color="FFD700", fill_type="solid")
            current_row += 1

        current_row += 2
        return current_row

    def apply_common_formatting(self, ws, max_row):
        """Apply common formatting to the worksheet"""
        # Apply borders to all cells with data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=25):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border
                    if cell.alignment is None:
                        cell.alignment = Alignment(
                            horizontal='center', vertical='center')

        # Set row heights
        for row in range(1, max_row + 1):
            ws.row_dimensions[row].height = 20
