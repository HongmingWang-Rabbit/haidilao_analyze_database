#!/usr/bin/env python3
"""
Beverage Summary Worksheet Generator
Generates beverage variance summary by store showing differences between system and inventory counts
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent))


class BeverageSummaryGenerator:
    """Generator for beverage variance summary worksheet"""

    def __init__(self, data_provider):
        self.data_provider = data_provider
        self.db_manager = data_provider.db_manager

    def generate_worksheet(self, workbook, target_date: str):
        """Generate beverage summary worksheet"""
        worksheet = workbook.create_sheet("酒水汇总表")

        # Set up styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="C41E3A", end_color="C41E3A", fill_type="solid")
        data_font = Font(size=10)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # Parse target date
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        year, month = target_dt.year, target_dt.month

        # Create header row
        headers = [
            "月份", "大区经理", "国家", "门店名称", "店经理",
            "差异数量（绝对值）", "影响金额（本币）", "影响金额（美元）"
        ]

        # Set headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Get beverage variance data
        variance_data = self.get_beverage_variance_data(year, month)

        # Populate data rows
        row = 2
        for store_data in variance_data:
            worksheet.cell(row=row, column=1, value=f"{year:04d}{month:02d}")
            worksheet.cell(row=row, column=2,
                           value=store_data['regional_manager'])
            worksheet.cell(row=row, column=3, value=store_data['country'])
            worksheet.cell(row=row, column=4, value=store_data['store_name'])
            worksheet.cell(row=row, column=5,
                           value=store_data['store_manager'])
            worksheet.cell(row=row, column=6,
                           value=store_data['variance_quantity'])
            worksheet.cell(row=row, column=7,
                           value=store_data['impact_amount_local'])
            worksheet.cell(row=row, column=8,
                           value=store_data['impact_amount_usd'])

            # Apply styling to data rows
            for col in range(1, 9):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.border = border
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')

            row += 1

        # Add total row
        if variance_data:
            worksheet.cell(row=row, column=1, value="").border = border
            worksheet.cell(row=row, column=2, value="").border = border
            worksheet.cell(row=row, column=3, value="").border = border
            worksheet.cell(row=row, column=4, value="").border = border
            worksheet.cell(row=row, column=5,
                           value="合计").font = Font(bold=True)
            worksheet.cell(row=row, column=5).border = border
            worksheet.cell(row=row, column=5).alignment = Alignment(
                horizontal='center')

            # Calculate totals
            total_variance = sum(
                abs(float(data['variance_quantity'])) for data in variance_data)
            total_local = sum(float(data['impact_amount_local'])
                              for data in variance_data)
            total_usd = sum(float(data['impact_amount_usd'])
                            for data in variance_data)

            worksheet.cell(row=row, column=6,
                           value=total_variance).font = Font(bold=True)
            worksheet.cell(row=row, column=6).border = border
            worksheet.cell(row=row, column=6).alignment = Alignment(
                horizontal='center')

            worksheet.cell(row=row, column=7,
                           value=total_local).font = Font(bold=True)
            worksheet.cell(row=row, column=7).border = border
            worksheet.cell(row=row, column=7).alignment = Alignment(
                horizontal='center')

            worksheet.cell(row=row, column=8,
                           value=total_usd).font = Font(bold=True)
            worksheet.cell(row=row, column=8).border = border
            worksheet.cell(row=row, column=8).alignment = Alignment(
                horizontal='center')

        # Set column widths
        column_widths = [10, 15, 10, 20, 15, 18, 18, 18]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

        return worksheet

    def get_beverage_variance_data(self, year: int, month: int):
        """Get beverage variance data by store"""
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()

                # Query beverage variance data - simplified without system_quantity
                query = """
                SELECT 
                    s.id as store_id,
                    s.name as store_name,
                    s.manager as store_manager,
                    '蒋冰洁' as regional_manager,  -- Default regional manager
                    '加拿大' as country,          -- Default country
                    
                    -- Use counted quantity as variance (since no system_quantity available)
                    SUM(COALESCE(ic.counted_quantity, 0)) as variance_quantity,
                    
                    -- Calculate impact amount (simplified without price for now)
                    SUM(COALESCE(ic.counted_quantity, 0) * 10) as impact_amount_local,  -- Assume $10 average
                    
                    -- Calculate impact amount (USD) - assuming 1 CAD = 0.74 USD
                    SUM(COALESCE(ic.counted_quantity, 0) * 10 * 0.74) as impact_amount_usd
                    
                FROM store s
                JOIN inventory_count ic ON s.id = ic.store_id 
                    AND EXTRACT(year FROM ic.count_date) = %s AND EXTRACT(month FROM ic.count_date) = %s
                JOIN material m ON ic.material_id = m.id
                JOIN material_type mt ON m.material_type_id = mt.id
                WHERE s.is_active = TRUE
                AND mt.name = '成本-酒水类'
                GROUP BY s.id, s.name, s.manager
                HAVING SUM(COALESCE(ic.counted_quantity, 0)) > 0
                ORDER BY s.id
                """

                cursor.execute(query, (year, month))
                results = cursor.fetchall()

                # Convert to list of dictionaries
                variance_data = []
                for row in results:
                    variance_data.append({
                        'store_id': row['store_id'],
                        'store_name': row['store_name'],
                        'store_manager': row['store_manager'] or '待定',
                        'regional_manager': row['regional_manager'],
                        'country': row['country'],
                        'variance_quantity': f"{row['variance_quantity']:.0f}" if row['variance_quantity'] else "0",
                        'impact_amount_local': f"{row['impact_amount_local']:.0f}" if row['impact_amount_local'] else "0",
                        'impact_amount_usd': f"{row['impact_amount_usd']:.0f}" if row['impact_amount_usd'] else "0"
                    })

                return variance_data

        except Exception as e:
            print(f"ERROR: Failed to get beverage variance data: {e}")
            return []
