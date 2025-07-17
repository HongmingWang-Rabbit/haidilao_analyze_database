#!/usr/bin/env python3
"""
Monthly Dishes Report worksheet generator (菜品用料月报).
Displays comprehensive dish-material relationship data from the database.
"""

from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict


class MonthlyDishesWorksheetGenerator:
    """Generate monthly dishes report worksheet (菜品用料月报) from database data"""

    def __init__(self, store_names, target_date):
        self.store_names = store_names
        self.target_date = target_date

    def generate_worksheet(self, wb, data_provider):
        """Generate the monthly dishes worksheet from database data"""

        ws = wb.create_sheet("菜品用料月报")

        # Parse target date for title
        target_dt = datetime.strptime(self.target_date, '%Y-%m-%d')

        # Get dish-material data from database
        dish_data = self.get_dish_material_data(data_provider)

        if not dish_data:
            # Create empty worksheet with message
            ws['A1'] = "无可用数据 - No Data Available"
            ws['A1'].font = Font(bold=True, size=14)
            return ws

        # Set column widths for readability
        column_widths = [8, 12, 15, 12, 12, 15, 12, 25,
                         25, 12, 15, 12, 12, 15, 15, 15, 15, 15, 15, 10]
        for i, width in enumerate(column_widths, 1):
            if i <= len(column_widths):
                ws.column_dimensions[get_column_letter(i)].width = width

        current_row = 1

        # Title section
        ws.merge_cells(f'A{current_row}:T{current_row}')
        ws[f'A{current_row}'] = f"海底捞菜品用料月报 - {target_dt.strftime('%Y年%m月')}"
        ws[f'A{current_row}'].font = Font(bold=True, size=16, color="FFFFFF")
        ws[f'A{current_row}'].alignment = Alignment(
            horizontal='center', vertical='center')
        ws[f'A{current_row}'].fill = PatternFill(
            start_color="C41E3A", end_color="C41E3A", fill_type="solid")
        current_row += 2

        # Summary section
        current_row = self.add_summary_section(ws, current_row, dish_data)

        # Main data section
        current_row = self.add_dish_data_section(ws, current_row, dish_data)

        # Apply common formatting
        self.apply_common_formatting(ws, current_row - 1)

        return ws

    def generate_material_variance_worksheet(self, wb, data_provider):
        """Generate material variance analysis worksheet comparing calculated vs actual usage"""

        ws = wb.create_sheet("物料用量差异分析")

        # Parse target date for title
        target_dt = datetime.strptime(self.target_date, '%Y-%m-%d')
        year, month = target_dt.year, target_dt.month

        # Get material variance data
        variance_data = self.get_material_variance_data(
            data_provider, year, month)

        if not variance_data:
            # Create empty worksheet with message
            ws['A1'] = "无可用数据 - No Variance Data Available"
            ws['A1'].font = Font(bold=True, size=14)
            return ws

        # Set column widths for readability (added combo usage column)
        column_widths = [8, 12, 15, 12, 15, 15, 15, 15, 15, 15, 12, 20, 12]
        for i, width in enumerate(column_widths, 1):
            if i <= len(column_widths):
                ws.column_dimensions[get_column_letter(i)].width = width

        current_row = 1

        # Title section (now extends to M column for 13 columns)
        ws.merge_cells(f'A{current_row}:M{current_row}')
        ws[f'A{current_row}'] = f"物料用量差异分析 - {target_dt.strftime('%Y年%m月')}"
        ws[f'A{current_row}'].font = Font(bold=True, size=16, color="FFFFFF")
        ws[f'A{current_row}'].alignment = Alignment(
            horizontal='center', vertical='center')
        ws[f'A{current_row}'].fill = PatternFill(
            start_color="C41E3A", end_color="C41E3A", fill_type="solid")
        current_row += 2

        # Summary section
        current_row = self.add_variance_summary_section(
            ws, current_row, variance_data)

        # Main variance data section
        current_row = self.add_variance_data_section(
            ws, current_row, variance_data)

        # Apply variance formatting
        self.apply_variance_formatting(ws, current_row - 1)

        return ws

    def get_dish_material_data(self, data_provider):
        """Query database for comprehensive dish-material relationship data"""
        try:
            with data_provider.db_manager.get_connection() as conn:
                cursor = conn.cursor()

                # Get target date info for monthly sales join
                target_dt = datetime.strptime(self.target_date, '%Y-%m-%d')
                year, month = target_dt.year, target_dt.month

                # Comprehensive query to get dish-material data with all related information
                # Include both regular dish sales and combo dish sales
                sql = """
                WITH combined_dish_sales AS (
                    -- Regular dish sales
                    SELECT 
                        dish_id, store_id,
                        sale_amount, return_amount, free_meal_amount, gift_amount
                    FROM dish_monthly_sale 
                    WHERE year = %s AND month = %s
                    
                    UNION ALL
                    
                    -- Combo dish sales (no return/free/gift amounts for combos)
                    SELECT 
                        dish_id, store_id,
                        sale_amount, 0 as return_amount, 0 as free_meal_amount, 0 as gift_amount
                    FROM monthly_combo_dish_sale 
                    WHERE year = %s AND month = %s
                ),
                aggregated_sales AS (
                    SELECT 
                        dish_id, store_id,
                        SUM(sale_amount) as total_sale_amount,
                        SUM(return_amount) as total_return_amount,
                        SUM(free_meal_amount) as total_free_meal_amount,
                        SUM(gift_amount) as total_gift_amount
                    FROM combined_dish_sales
                    GROUP BY dish_id, store_id
                )
                SELECT 
                    dt.name as dish_type_name,
                    dct.name as dish_child_type_name,
                    d.id as dish_id,
                    d.name as dish_name,
                    d.system_name as dish_system_name,
                    d.full_code as dish_code,
                    d.short_code as dish_short_code,
                    d.size as dish_size,
                    d.unit as dish_unit,
                    d.serving_size_kg,
                    m.id as material_id,
                    m.name as material_name,
                    m.material_number,
                    m.unit as material_unit,
                    m.package_spec,
                    dm.standard_quantity,
                    -- Get current price for first store (if available)
                    dph.price as current_price,
                    dph.currency,
                    dph.effective_date as price_date,
                    -- Monthly sales data (aggregated from both regular and combo sales)
                    COALESCE(ads.total_sale_amount, 0) as sale_amount,
                    COALESCE(ads.total_return_amount, 0) as return_amount,
                    COALESCE(ads.total_free_meal_amount, 0) as free_meal_amount,
                    COALESCE(ads.total_gift_amount, 0) as gift_amount,
                    -- Material usage data
                    COALESCE(mmu.material_used, 0) as material_used,
                    -- Count total dish-material relationships for this dish
                    COUNT(*) OVER (PARTITION BY d.id) as material_count_for_dish
                FROM dish d
                JOIN dish_child_type dct ON d.dish_child_type_id = dct.id
                JOIN dish_type dt ON dct.dish_type_id = dt.id
                LEFT JOIN dish_material dm ON d.id = dm.dish_id
                LEFT JOIN material m ON dm.material_id = m.id
                LEFT JOIN dish_price_history dph ON d.id = dph.dish_id 
                    AND dph.is_active = TRUE 
                    AND dph.store_id = 1  -- Use first store for reference price
                LEFT JOIN aggregated_sales ads ON d.id = ads.dish_id
                LEFT JOIN material_monthly_usage mmu ON m.id = mmu.material_id 
                    AND mmu.year = %s AND mmu.month = %s
                    AND mmu.store_id = ads.store_id  -- Match store for material usage
                WHERE d.is_active = TRUE
                ORDER BY ads.store_id, dt.name, dct.name, d.name, m.name
                """

                cursor.execute(sql, (year, month, year, month, year, month))
                return cursor.fetchall()

        except Exception as e:
            print(f"❌ Error fetching dish-material data: {e}")
            return []

    def add_summary_section(self, ws, start_row, dish_data):
        """Add summary statistics section"""
        current_row = start_row

        # Calculate summary statistics
        total_dishes = len(set(row['dish_id'] for row in dish_data))
        total_materials = len(set(row['material_id']
                              for row in dish_data if row['material_id']))
        total_relationships = len(
            [row for row in dish_data if row['material_id']])
        total_dish_types = len(set(row['dish_type_name'] for row in dish_data))

        # Dishes with/without materials
        dishes_with_materials = len(
            set(row['dish_id'] for row in dish_data if row['material_id']))
        dishes_without_materials = total_dishes - dishes_with_materials

        # Monthly performance totals
        total_sales = sum(row['sale_amount'] or 0 for row in dish_data)
        total_returns = sum(row['return_amount'] or 0 for row in dish_data)
        total_free_meals = sum(
            row['free_meal_amount'] or 0 for row in dish_data)
        total_gifts = sum(row['gift_amount'] or 0 for row in dish_data)

        # Create summary section
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "数据统计概览"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        ws[f'A{current_row}'].fill = PatternFill(
            start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        current_row += 1

        summary_data = [
            ("菜品总数", total_dishes),
            ("物料总数", total_materials),
            ("配方关系总数", total_relationships),
            ("菜品大类数", total_dish_types),
            ("有配方菜品", dishes_with_materials),
            ("无配方菜品", dishes_without_materials),
            ("总销量", f"{total_sales:.0f}"),
            ("总退菜", f"{total_returns:.0f}"),
            ("总免费餐", f"{total_free_meals:.0f}"),
            ("总赠送", f"{total_gifts:.0f}")
        ]

        for i, (label, value) in enumerate(summary_data):
            row = current_row + (i // 2)
            col = 1 if i % 2 == 0 else 3

            ws.cell(row=row, column=col, value=label).font = Font(bold=True)
            ws.cell(row=row, column=col+1, value=value)

            # Highlight performance metrics in a different color
            if i >= 6:  # Performance metrics start from index 6
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                ws.cell(row=row, column=col+1).fill = PatternFill(
                    start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")

        current_row += 3
        return current_row + 1

    def add_dish_data_section(self, ws, start_row, dish_data):
        """Add main dish data section"""
        current_row = start_row

        # Headers - Updated to include monthly performance data
        headers = [
            "序号", "菜品大类", "菜品子类", "菜品编码", "菜品短编码", "菜品名称",
            "系统名称", "规格", "单位", "出品分量(kg)", "物料号", "物料名称",
            "包装规格", "标准用量", "物料单位", "当前售价", "货币", "价格日期",
            "月销量", "退菜量", "免费餐", "赠送量", "物料消耗",
            "配方数量", "备注"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

        # Group data by dish to handle dishes with multiple materials
        dishes_processed = set()
        row_number = 1

        for row_data in dish_data:
            dish_id = row_data['dish_id']

            # For dishes with materials, show each material relationship
            if row_data['material_id']:
                self.add_dish_row(ws, current_row, row_number,
                                  row_data, has_material=True)
                row_number += 1
                current_row += 1
            # For dishes without materials, show once (if not already shown)
            elif dish_id not in dishes_processed:
                self.add_dish_row(ws, current_row, row_number,
                                  row_data, has_material=False)
                dishes_processed.add(dish_id)
                row_number += 1
                current_row += 1

            dishes_processed.add(dish_id)

        return current_row

    def add_dish_row(self, ws, row, row_number, data, has_material=True):
        """Add a single dish data row"""

        # Determine remarks based on material availability
        if has_material:
            total_materials = data['material_count_for_dish'] if data['material_count_for_dish'] else 0
            remarks = f"共{total_materials}种物料" if total_materials > 1 else "单一物料"
        else:
            remarks = "暂无配方"

        row_data = [
            row_number,  # 序号
            data['dish_type_name'],  # 菜品大类
            data['dish_child_type_name'],  # 菜品子类
            data['dish_code'],  # 菜品编码
            data['dish_short_code'] or '',  # 菜品短编码
            data['dish_name'],  # 菜品名称
            data['dish_system_name'] or data['dish_name'],  # 系统名称
            data['dish_size'] or '',  # 规格
            data['dish_unit'] or '份',  # 单位
            data['serving_size_kg'] or 0.0,  # 出品分量(kg)
            data['material_number'] or '',  # 物料号
            data['material_name'] or '',  # 物料名称
            data['package_spec'] or '',  # 包装规格
            data['standard_quantity'] or '',  # 标准用量
            data['material_unit'] or '',  # 物料单位
            data['current_price'] or '',  # 当前售价
            data['currency'] or '',  # 货币
            data['price_date'].strftime(
                '%Y-%m-%d') if data['price_date'] else '',  # 价格日期
            # Monthly performance data
            data['sale_amount'] or 0,  # 月销量
            data['return_amount'] or 0,  # 退菜量
            data['free_meal_amount'] or 0,  # 免费餐
            data['gift_amount'] or 0,  # 赠送量
            data['material_used'] or 0,  # 物料消耗
            data['material_count_for_dish'] or 0,  # 配方数量
            remarks  # 备注
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)

            # Highlight dishes without materials
            if not has_material:
                cell.fill = PatternFill(
                    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            # Format numeric columns
            # Serving size, standard quantity, price, monthly performance data
            if col in [10, 14, 16, 19, 20, 21, 22, 23, 24]:
                if isinstance(value, (int, float)) and value != 0:
                    if col == 16:  # Price
                        cell.number_format = '#,##0.00'
                    # Sales performance data (whole numbers)
                    elif col in [19, 20, 21, 22, 23]:
                        cell.number_format = '#,##0'
                    else:  # Material quantities (decimal)
                        cell.number_format = '#,##0.0000'

            # Center align certain columns
            # Serial, short code, size, unit, material unit, currency, count, performance metrics
            if col in [1, 4, 8, 9, 15, 17, 19, 20, 21, 22, 23, 24, 25]:
                cell.alignment = Alignment(horizontal='center')

    def apply_common_formatting(self, ws, max_row):
        """Apply common formatting to the worksheet"""

        # Add borders to data section
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Find where data section starts (after summary)
        data_start_row = 1
        for row in range(1, max_row + 1):
            if ws.cell(row=row, column=1).value == "序号":
                data_start_row = row
                break

        # Apply borders to data section
        for row in range(data_start_row, max_row + 1):
            for col in range(1, 26):  # Updated to 25 columns (removed 2 stock columns)
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

                # Wrap text for long content columns
                if col in [6, 7, 12, 13, 25]:  # Name columns and remarks
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Freeze panes at data header
        if data_start_row > 1:
            ws.freeze_panes = f'A{data_start_row + 1}'

        # Auto-fit row heights for wrapped text
        for row in range(data_start_row, max_row + 1):
            ws.row_dimensions[row].height = None  # Auto height

    def get_material_variance_data(self, data_provider, year: int, month: int):
        """Calculate material usage variance between theoretical and system record"""
        try:
            with data_provider.db_manager.get_connection() as conn:
                cursor = conn.cursor()

                # Get only materials that have dish_material records with standard_quantity
                cursor.execute("""
                    SELECT DISTINCT m.id as material_id
                    FROM material m
                    WHERE m.is_active = TRUE
                    AND EXISTS (SELECT 1 FROM dish_material dm WHERE dm.material_id = m.id)
                """, ())

                all_materials = cursor.fetchall()
                material_ids = [row['material_id'] for row in all_materials]

                if not material_ids:
                    print("⚠️  No materials found with dish-material relationships")
                    return [{
                        'material_id': 0,
                        'store_id': 1,
                        'store_name': '系统说明',
                        'material_name': '暂无配方物料数据',
                        'material_number': '',
                        'material_unit': '',
                        'package_spec': '',
                        'theoretical_usage': 0,
                        'combo_usage': 0,
                        'system_record': 0,
                        'inventory_count': 0,
                        'variance_amount': 0,
                        'variance_percent': 0,
                        'variance_status': '正常'
                    }]

                # Get regular theoretical usage (dish sales * standard_quantity * loss_rate)
                regular_theoretical_sql = """
                WITH aggregated_dish_sales AS (
                    SELECT 
                        dish_id,
                        store_id,
                        SUM(COALESCE(sale_amount, 0)) as total_sale_amount,
                        SUM(COALESCE(return_amount, 0)) as total_return_amount
                    FROM dish_monthly_sale
                    WHERE year = %s AND month = %s
                    GROUP BY dish_id, store_id
                ),
                dish_sales AS (
                    SELECT 
                        d.id as dish_id,
                        d.name as dish_name,
                        ads.store_id,
                        s.name as store_name,
                        COALESCE(ads.total_sale_amount, 0) - COALESCE(ads.total_return_amount, 0) as net_sales
                    FROM dish d
                    LEFT JOIN aggregated_dish_sales ads ON d.id = ads.dish_id
                    LEFT JOIN store s ON ads.store_id = s.id
                    WHERE d.is_active = TRUE AND ads.store_id IS NOT NULL
                ),
                regular_theoretical_usage AS (
                    SELECT 
                        m.id as material_id,
                        m.name as material_name,
                        m.material_number,
                        m.unit as material_unit,
                        m.package_spec,
                        ds.store_id,
                        ds.store_name,
                        SUM(ds.net_sales * COALESCE(dm.standard_quantity, 0) * COALESCE(dm.loss_rate, 1.0)) as theoretical_total
                    FROM material m
                    LEFT JOIN dish_material dm ON m.id = dm.material_id
                    LEFT JOIN dish_sales ds ON dm.dish_id = ds.dish_id
                    WHERE m.is_active = TRUE AND ds.store_id IS NOT NULL
                        AND m.id = ANY(%s)
                    GROUP BY m.id, m.name, m.material_number, m.unit, m.package_spec, ds.store_id, ds.store_name
                )
                SELECT * FROM regular_theoretical_usage
                ORDER BY store_name, material_name
                """

                cursor.execute(regular_theoretical_sql,
                               (year, month, material_ids))
                regular_theoretical_data = cursor.fetchall()

                # Get combo usage (combo dish sales * standard_quantity * loss_rate)
                combo_usage_sql = """
                WITH combo_dish_sales AS (
                    SELECT 
                        d.id as dish_id,
                        d.name as dish_name,
                        mcds.store_id,
                        s.name as store_name,
                        COALESCE(mcds.sale_amount, 0) as net_sales
                    FROM dish d
                    LEFT JOIN monthly_combo_dish_sale mcds ON d.id = mcds.dish_id 
                        AND mcds.year = %s AND mcds.month = %s
                    LEFT JOIN store s ON mcds.store_id = s.id
                    WHERE d.is_active = TRUE AND mcds.store_id IS NOT NULL
                ),
                combo_theoretical_usage AS (
                    SELECT 
                        m.id as material_id,
                        m.name as material_name,
                        m.material_number,
                        m.unit as material_unit,
                        m.package_spec,
                        cds.store_id,
                        cds.store_name,
                        SUM(cds.net_sales * COALESCE(dm.standard_quantity, 0) * COALESCE(dm.loss_rate, 1.0)) as combo_total
                    FROM material m
                    LEFT JOIN dish_material dm ON m.id = dm.material_id
                    LEFT JOIN combo_dish_sales cds ON dm.dish_id = cds.dish_id
                    WHERE m.is_active = TRUE AND cds.store_id IS NOT NULL
                        AND m.id = ANY(%s)
                    GROUP BY m.id, m.name, m.material_number, m.unit, m.package_spec, cds.store_id, cds.store_name
                )
                SELECT * FROM combo_theoretical_usage
                ORDER BY store_name, material_name
                """

                cursor.execute(combo_usage_sql, (year, month, material_ids))
                combo_usage_data = cursor.fetchall()

                # Get system record (material_monthly_usage.material_used)
                system_sql = """
                SELECT 
                    m.id as material_id,
                    m.name as material_name,
                    m.material_number,
                    m.unit as material_unit,
                    m.package_spec,
                    mmu.store_id,
                    s.name as store_name,
                    COALESCE(mmu.material_used, 0) as system_record
                FROM material m
                LEFT JOIN material_monthly_usage mmu ON m.id = mmu.material_id 
                    AND mmu.year = %s AND mmu.month = %s
                LEFT JOIN store s ON mmu.store_id = s.id
                WHERE m.is_active = TRUE AND mmu.store_id IS NOT NULL
                    AND m.id = ANY(%s)
                ORDER BY s.name, m.name
                """

                cursor.execute(system_sql, (year, month, material_ids))
                system_data = cursor.fetchall()

                # Get inventory count (inventory_count.counted_quantity)
                inventory_sql = """
                SELECT 
                    m.id as material_id,
                    m.name as material_name,
                    m.material_number,
                    m.unit as material_unit,
                    m.package_spec,
                    ic.store_id,
                    s.name as store_name,
                    COALESCE(ic.counted_quantity, 0) as inventory_count
                FROM material m
                LEFT JOIN inventory_count ic ON m.id = ic.material_id
                LEFT JOIN store s ON ic.store_id = s.id
                WHERE m.is_active = TRUE AND ic.store_id IS NOT NULL
                    AND m.id = ANY(%s)
                    AND EXTRACT(year FROM ic.count_date) = %s 
                    AND EXTRACT(month FROM ic.count_date) = %s
                ORDER BY s.name, m.id
                """

                cursor.execute(inventory_sql, (material_ids, year, month))
                inventory_data = cursor.fetchall()

                # Combine all data
                regular_theoretical_dict = {
                    (row['material_id'], row['store_id']): row for row in regular_theoretical_data}
                combo_usage_dict = {
                    (row['material_id'], row['store_id']): row for row in combo_usage_data}
                system_dict = {
                    (row['material_id'], row['store_id']): row for row in system_data}
                inventory_dict = {
                    (row['material_id'], row['store_id']): row for row in inventory_data}

                # Create combined variance data
                variance_data = []

                # Get all unique material-store combinations
                all_combinations = set()
                for row in regular_theoretical_data + combo_usage_data + system_data + inventory_data:
                    all_combinations.add((row['material_id'], row['store_id']))

                for material_id, store_id in all_combinations:
                    regular_theoretical_row = regular_theoretical_dict.get(
                        (material_id, store_id), {})
                    combo_usage_row = combo_usage_dict.get(
                        (material_id, store_id), {})
                    system_row = system_dict.get((material_id, store_id), {})
                    inventory_row = inventory_dict.get(
                        (material_id, store_id), {})

                    # Get material info from any available row
                    info_row = regular_theoretical_row or combo_usage_row or system_row or inventory_row
                    if not info_row:
                        continue

                    # Ensure we have all required fields with defaults
                    material_name = info_row.get(
                        'material_name', f'Material_{material_id}')
                    material_number = info_row.get('material_number', '')
                    material_unit = info_row.get('material_unit', '')
                    package_spec = info_row.get('package_spec', '')
                    store_name = info_row.get(
                        'store_name', f'Store_{store_id}')

                    regular_theoretical_usage = regular_theoretical_row.get(
                        'theoretical_total', 0) or 0
                    combo_usage = combo_usage_row.get('combo_total', 0) or 0
                    theoretical_usage = regular_theoretical_usage + \
                        combo_usage  # Combined for variance calculation
                    system_record = system_row.get('system_record', 0) or 0
                    inventory_count = inventory_row.get(
                        'inventory_count', 0) or 0

                    # Calculate variance to match Excel formula: H - (G + I)
                    # This matches the Excel formula: =I{row}-(G{row}+H{row}) (will be updated with new column positions)
                    variance_amount = system_record - \
                        (theoretical_usage + inventory_count)

                    # Calculate variance percentage to match Excel formula logic
                    if theoretical_usage > 0:
                        variance_percent = abs(
                            variance_amount) / theoretical_usage * 100
                    elif variance_amount != 0:
                        variance_percent = 100.0
                    else:
                        variance_percent = 0

                    # Determine variance status based on calculated percentage
                    variance_status = "正常"
                    if variance_percent > 5:
                        if variance_amount > 0:
                            variance_status = "超量"
                        else:
                            variance_status = "不足"

                    variance_data.append({
                        'material_id': material_id,
                        'store_id': store_id,
                        'store_name': store_name,
                        'material_name': material_name,
                        'material_number': material_number,
                        'material_unit': material_unit,
                        'package_spec': package_spec,
                        'theoretical_usage': regular_theoretical_usage,
                        'combo_usage': combo_usage,
                        'system_record': system_record,
                        'inventory_count': inventory_count,
                        'variance_amount': variance_amount,
                        'variance_percent': variance_percent,
                        'variance_status': variance_status
                    })

                return sorted(variance_data, key=lambda x: (x['store_name'], x['material_name']))

        except Exception as e:
            print(f"❌ Error calculating material variance data: {e}")
            return []

    def add_variance_summary_section(self, ws, start_row, variance_data):
        """Add variance summary statistics section"""
        current_row = start_row

        # Calculate summary statistics
        total_materials = len(variance_data)
        materials_over_threshold = len(
            [row for row in variance_data if abs(row['variance_percent']) > 5])
        materials_normal = total_materials - materials_over_threshold

        over_usage = len(
            [row for row in variance_data if row['variance_percent'] > 5])
        under_usage = len(
            [row for row in variance_data if row['variance_percent'] < -5])

        total_theoretical = sum(row['theoretical_usage']
                                for row in variance_data)
        total_combo_usage = sum(row['combo_usage']
                                for row in variance_data)
        total_combined_theoretical = total_theoretical + total_combo_usage
        total_system = sum(row['system_record'] for row in variance_data)
        total_inventory = sum(row['inventory_count'] for row in variance_data)
        total_variance = total_system - total_combined_theoretical
        overall_variance_percent = (
            total_variance / total_combined_theoretical * 100) if total_combined_theoretical > 0 else 0

        # Create summary section
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "差异分析概览"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        ws[f'A{current_row}'].fill = PatternFill(
            start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        current_row += 1

        summary_data = [
            ("物料总数", total_materials),
            ("差异超5%物料", materials_over_threshold),
            ("正常范围物料", materials_normal),
            ("超量使用物料", over_usage),
            ("使用不足物料", under_usage),
            ("总理论用量", f"{total_theoretical:.2f}"),
            ("总套餐用量", f"{total_combo_usage:.2f}"),
            ("总系统记录", f"{total_system:.2f}"),
            ("总库存盘点", f"{total_inventory:.2f}"),
            ("总差异", f"{total_variance:.2f}"),
            ("总差异率", f"{overall_variance_percent:.1f}%")
        ]

        for i, (label, value) in enumerate(summary_data):
            row = current_row + (i // 2)
            col = 1 + (i % 2) * 3

            cell_label = ws.cell(row=row, column=col, value=label)
            cell_value = ws.cell(row=row, column=col+1, value=value)
            cell_label.font = Font(bold=True)

            # Color code summary metrics
            if i >= 5:  # Usage metrics
                fill_color = "E8F5E8"
                # Overall variance (now index 10)
                if i == 10 and abs(overall_variance_percent) > 5:
                    fill_color = "FFE6E6" if overall_variance_percent > 0 else "FFF0E6"
                cell_label.fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell_value.fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid")

        current_row += 3
        return current_row + 1

    def add_variance_data_section(self, ws, start_row, variance_data):
        """Add main variance data section"""
        current_row = start_row

        # Headers (added combo usage column)
        headers = [
            "序号", "门店", "物料名称", "物料号", "单位", "包装规格",
            "理论用量", "套餐用量", "系统记录", "库存盘点", "差异数量", "差异率(%)", "状态"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1

        # Data rows
        for row_number, data in enumerate(variance_data, 1):
            # Basic data (columns A-J)
            basic_data = [
                row_number,                      # A: 序号
                data['store_name'],              # B: 门店
                data['material_name'],           # C: 物料名称
                data['material_number'] or '',   # D: 物料号
                data['material_unit'] or '',     # E: 单位
                data['package_spec'] or '',      # F: 包装规格
                data['theoretical_usage'],       # G: 理论用量 (regular only)
                data['combo_usage'],             # H: 套餐用量 (NEW COLUMN)
                data['system_record'],           # I: 系统记录
                data['inventory_count']          # J: 库存盘点
            ]

            # Insert basic data (A-J)
            for col, value in enumerate(basic_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)

                # Format numeric columns
                if col in [7, 8, 9, 10]:  # Usage amounts, combo usage, system record, inventory count
                    cell.number_format = '#,##0.0000'

                # Center align certain columns
                if col in [1, 2, 5]:  # Serial, store, unit
                    cell.alignment = Alignment(horizontal='center')

            # K: 差异数量 (Variance Amount) - Excel formula: = I - (G+H+J)
            variance_formula = f"=I{current_row}-(G{current_row}+H{current_row}+J{current_row})"
            variance_cell = ws.cell(
                row=current_row, column=11, value=variance_formula)
            variance_cell.number_format = '#,##0.0000'

            # L: 差异率 (Variance Rate) - Excel formula: = ABS(K/(G+H))*100 with error handling
            # Use combined theoretical usage (G+H) for percentage calculation
            rate_formula = f"=IF((G{current_row}+H{current_row})=0,IF(K{current_row}=0,0,100),ABS(K{current_row}/(G{current_row}+H{current_row}))*100)"
            rate_cell = ws.cell(row=current_row, column=12, value=rate_formula)
            # Fixed percentage format with exactly 2 decimal places
            rate_cell.number_format = '0.00"%"'
            rate_cell.alignment = Alignment(horizontal='center')

            # M: 状态 (Status)
            status_cell = ws.cell(row=current_row, column=13,
                                  value=data['variance_status'])
            status_cell.alignment = Alignment(horizontal='center')

            # Apply row-level formatting based on variance percentage
            if abs(data['variance_percent']) > 5:
                fill_color = "FFE6E6" if data['variance_percent'] > 5 else "FFF0E6"
                row_fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid")

                # Apply fill to all cells in this row (now 13 columns)
                for col in range(1, 14):
                    ws.cell(row=current_row, column=col).fill = row_fill

            current_row += 1

        return current_row

    def apply_variance_formatting(self, ws, max_row):
        """Apply formatting to variance worksheet"""
        # Add borders to data section
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Find data section start
        data_start_row = 1
        for row in range(1, max_row + 1):
            if ws.cell(row=row, column=1).value == "序号":
                data_start_row = row
                break

        # Apply borders to data section
        for row in range(data_start_row, max_row + 1):
            # 13 columns for variance data (added combo usage column)
            for col in range(1, 14):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

        # Freeze panes at data header
        if data_start_row > 1:
            ws.freeze_panes = f'A{data_start_row + 1}'
