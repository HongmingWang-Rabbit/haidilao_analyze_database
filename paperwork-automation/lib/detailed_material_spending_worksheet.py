#!/usr/bin/env python3
"""
Detailed Material Spending Worksheet Generator

This generator creates detailed material spending worksheets for each store,
showing individual material usage with amounts and classifications.
Similar to the screenshot provided - shows granular material spending data.
"""

import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class DetailedMaterialSpendingGenerator:
    """Generator for detailed material spending worksheets by store."""

    def __init__(self, data_provider):
        """Initialize the generator with data provider."""
        self.data_provider = data_provider
        self.logger = logging.getLogger(__name__)

    def generate_worksheets(self, workbook, target_date: str) -> None:
        """Generate detailed material spending worksheets for each store."""
        try:
            # Parse target date
            target_dt = datetime.strptime(target_date, '%Y-%m-%d')
            year, month = target_dt.year, target_dt.month

            self.logger.info(
                f"Generating detailed material spending worksheets for {year}-{month:02d}")

            # Get store data
            store_data = self._get_store_data()
            if not store_data:
                self.logger.warning("No store data found")
                return

            # Generate worksheet for each store
            for store_id, store_info in store_data.items():
                # Get detailed material spending for this store
                spending_data = self._get_detailed_material_spending(
                    store_id, year, month)

                if not spending_data:
                    self.logger.warning(
                        f"No material spending data found for store {store_info['name']}")
                    continue

                # Create worksheet for this store
                worksheet_name = f"物料明细-{store_info['name']}"
                # Ensure worksheet name is not too long for Excel
                if len(worksheet_name) > 31:
                    worksheet_name = f"物料明细-{store_id}"

                ws = workbook.create_sheet(title=worksheet_name)

                # Generate the worksheet content
                self._create_detailed_spending_worksheet(
                    ws, store_info, spending_data, f"{year}-{month:02d}")

            self.logger.info(
                "Detailed material spending worksheets generated successfully")

        except Exception as e:
            self.logger.error(
                f"Error generating detailed material spending worksheets: {e}")
            raise

    def _get_store_data(self) -> Dict[int, Dict[str, Any]]:
        """Get store information."""
        try:
            query = """
            SELECT id, name, district_id, is_active
            FROM store 
            WHERE is_active = TRUE
            ORDER BY id
            """

            stores = self.data_provider.db_manager.fetch_all(query)
            return {store['id']: store for store in stores}

        except Exception as e:
            self.logger.error(f"Error fetching store data: {e}")
            return {}

    def _get_detailed_material_spending(self, store_id: int, year: int, month: int) -> List[Dict[str, Any]]:
        """Get detailed material spending data for a specific store."""
        try:
            query = """
            SELECT 
                m.material_number,
                m.name as material_name,
                mt.name as material_type_name,
                mct.name as material_child_type_name,
                mmu.material_used,
                COALESCE(mph.price, 0) as unit_price,
                (mmu.material_used * COALESCE(mph.price, 0)) as total_amount,
                m.unit,
                m.package_spec,
                mt.sort_order as type_sort_order,
                COALESCE(mct.sort_order, 999) as child_type_sort_order
            FROM material_monthly_usage mmu
            JOIN material m ON mmu.material_id = m.id
            LEFT JOIN material_type mt ON m.material_type_id = mt.id
            LEFT JOIN material_child_type mct ON m.material_child_type_id = mct.id
            LEFT JOIN material_price_history mph ON m.id = mph.material_id 
                AND mph.is_active = TRUE
                AND mph.store_id = %s
            WHERE mmu.store_id = %s 
                AND mmu.year = %s 
                AND mmu.month = %s
                AND mmu.material_used > 0
            ORDER BY 
                COALESCE(mt.sort_order, 999), 
                mt.name,
                COALESCE(mct.sort_order, 999),
                mct.name,
                m.name
            """

            return self.data_provider.db_manager.fetch_all(query, (store_id, store_id, year, month))

        except Exception as e:
            self.logger.error(
                f"Error fetching detailed material spending for store {store_id}: {e}")
            return []

    def _create_detailed_spending_worksheet(self, ws, store_info: Dict, spending_data: List[Dict], period: str) -> None:
        """Create the detailed spending worksheet for a store."""
        current_row = 1

        # Title
        ws[f'A{current_row}'] = f"海底捞物料消耗明细报表 - {store_info['name']}"
        ws[f'A{current_row}'].font = Font(name='SimSun', size=16, bold=True)
        ws[f'A{current_row}'].alignment = Alignment(
            horizontal='center', vertical='center')
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 1

        # Period and generation time
        ws[f'A{current_row}'] = f"报表期间: {period} | 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{current_row}'].font = Font(name='SimSun', size=10)
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2

        # Headers
        headers = [
            "物料编号 - 物料名称",
            "本月使用金额",
            "分类",
            "使用数量",
            "单价",
            "单位"
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(name='SimSun', size=11, bold=True)
            cell.fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(
                horizontal='center', vertical='center')

        current_row += 1
        data_start_row = current_row

        # Group data by material type
        grouped_data = {}
        for row in spending_data:
            material_type = row['material_type_name'] or "未分类"
            if material_type not in grouped_data:
                grouped_data[material_type] = {
                    'items': [],
                    'total': 0,
                    'sort_order': row['type_sort_order'] or 999
                }
            grouped_data[material_type]['items'].append(row)
            grouped_data[material_type]['total'] += float(
                row['total_amount'] or 0)

        # Sort material types by sort_order
        sorted_types = sorted(grouped_data.items(),
                              key=lambda x: x[1]['sort_order'])

        grand_total = 0

        # Generate data rows
        for material_type, type_data in sorted_types:
            # Material type header
            ws[f'A{current_row}'] = f"● {material_type}"
            ws[f'A{current_row}'].font = Font(
                name='SimSun', size=11, bold=True, color="0066CC")
            ws[f'A{current_row}'].fill = PatternFill(
                start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            ws.merge_cells(f'A{current_row}:F{current_row}')
            current_row += 1

            type_total = 0

            # Material items
            for item in type_data['items']:
                material_name = f"{item['material_number']} {item['material_name']}"
                total_amount = float(item['total_amount'] or 0)
                used_quantity = float(item['material_used'] or 0)
                unit_price = float(item['unit_price'] or 0)
                unit = item['unit'] or ""

                # Material row
                ws[f'A{current_row}'] = material_name
                ws[f'B{current_row}'] = round(total_amount, 2)
                ws[f'C{current_row}'] = material_type
                ws[f'D{current_row}'] = round(used_quantity, 4)
                ws[f'E{current_row}'] = round(unit_price, 4)
                ws[f'F{current_row}'] = unit

                # Format cells
                ws[f'A{current_row}'].font = Font(name='SimSun', size=10)
                ws[f'B{current_row}'].font = Font(name='SimSun', size=10)
                ws[f'B{current_row}'].number_format = '#,##0.00'
                ws[f'B{current_row}'].alignment = Alignment(horizontal='right')
                ws[f'C{current_row}'].font = Font(name='SimSun', size=10)
                ws[f'D{current_row}'].font = Font(name='SimSun', size=10)
                ws[f'D{current_row}'].number_format = '#,##0.0000'
                ws[f'D{current_row}'].alignment = Alignment(horizontal='right')
                ws[f'E{current_row}'].font = Font(name='SimSun', size=10)
                ws[f'E{current_row}'].number_format = '#,##0.0000'
                ws[f'E{current_row}'].alignment = Alignment(horizontal='right')
                ws[f'F{current_row}'].font = Font(name='SimSun', size=10)
                ws[f'F{current_row}'].alignment = Alignment(
                    horizontal='center')

                type_total += total_amount
                current_row += 1

            # Type subtotal
            ws[f'A{current_row}'] = f"小计 - {material_type}"
            ws[f'B{current_row}'] = round(type_total, 2)

            ws[f'A{current_row}'].font = Font(
                name='SimSun', size=11, bold=True)
            ws[f'B{current_row}'].font = Font(
                name='SimSun', size=11, bold=True)
            ws[f'A{current_row}'].fill = PatternFill(
                start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            ws[f'B{current_row}'].fill = PatternFill(
                start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            ws[f'B{current_row}'].number_format = '#,##0.00'
            ws[f'B{current_row}'].alignment = Alignment(horizontal='right')

            grand_total += type_total
            current_row += 2  # Space between categories

        # Grand total
        ws[f'A{current_row}'] = "总计"
        ws[f'B{current_row}'] = round(grand_total, 2)

        ws[f'A{current_row}'].font = Font(
            name='SimSun', size=12, bold=True, color="FF0000")
        ws[f'B{current_row}'].font = Font(
            name='SimSun', size=12, bold=True, color="FF0000")
        ws[f'A{current_row}'].fill = PatternFill(
            start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        ws[f'B{current_row}'].fill = PatternFill(
            start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        ws[f'B{current_row}'].number_format = '#,##0.00'
        ws[f'B{current_row}'].alignment = Alignment(horizontal='right')

        # Apply formatting
        self._apply_formatting(ws, data_start_row, current_row)

    def _apply_formatting(self, ws, data_start_row: int, data_end_row: int) -> None:
        """Apply formatting to the worksheet."""
        # Set column widths
        column_widths = {
            'A': 35,  # Material number and name
            'B': 15,  # Amount
            'C': 20,  # Category
            'D': 12,  # Quantity
            'E': 12,  # Unit price
            'F': 10   # Unit
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Add borders to data area
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(max(1, data_start_row - 1), data_end_row + 1):
            for col in range(1, 7):  # A to F
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
