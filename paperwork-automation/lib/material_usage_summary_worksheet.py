#!/usr/bin/env python3
"""
Material Usage Summary Worksheet Generator
Generates a summary of material usage by store and material type for monthly reports.
"""

import logging
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any
from datetime import datetime

logger = logging.getLogger(__name__)


class MaterialUsageSummaryGenerator:
    """Generator for Material Usage Summary worksheet."""

    def __init__(self, data_provider):
        """
        Initialize the generator.

        Args:
            data_provider: Database data provider instance
        """
        self.data_provider = data_provider
        self.logger = logging.getLogger(__name__)

    def generate_worksheet(self, workbook, target_date: str) -> None:
        """
        Generate Material Usage Summary worksheet.

        Args:
            workbook: Excel workbook to add worksheet to
            target_date: Target date in YYYY-MM-DD format
        """
        try:
            self.logger.info("Generating Material Usage Summary worksheet")

            # Parse target date to get year and month
            target_dt = datetime.strptime(target_date, '%Y-%m-%d')
            year, month = target_dt.year, target_dt.month

            # Create worksheet
            ws = workbook.create_sheet("物料使用汇总")

            # Get data
            usage_data = self._get_material_usage_data(year, month)
            store_data = self._get_store_data()
            material_type_data = self._get_material_type_data()

            if not usage_data:
                self.logger.warning("No material usage data found")
                # Create empty worksheet with header
                self._create_header(ws, f"{year}-{month:02d}")
                return

            # Generate the worksheet content
            self._create_header(ws, f"{year}-{month:02d}")
            self._create_summary_table(
                ws, usage_data, store_data, material_type_data)
            self._apply_formatting(ws)

            self.logger.info(
                "Material Usage Summary worksheet generated successfully")

        except Exception as e:
            self.logger.error(
                f"Error generating Material Usage Summary worksheet: {e}")
            raise

    def _get_material_usage_data(self, year: int, month: int) -> List[Dict[str, Any]]:
        """Get material usage data by store and material type."""
        try:
            query = """
            SELECT 
                s.id as store_id,
                s.name as store_name,
                mt.id as material_type_id,
                mt.name as material_type_name,
                SUM(mmu.material_used * mph.price) as usage_amount
            FROM material_monthly_usage mmu
            JOIN material m ON mmu.material_id = m.id
            JOIN material_type mt ON m.material_type_id = mt.id
            JOIN store s ON mmu.store_id = s.id
            LEFT JOIN material_price_history mph ON m.id = mph.material_id 
                AND mph.is_active = TRUE
                AND mph.district_id = s.district_id
            WHERE mmu.year = %s AND mmu.month = %s
            GROUP BY s.id, s.name, mt.id, mt.name
            ORDER BY s.id, mt.sort_order, mt.name
            """

            return self.data_provider.db_manager.fetch_all(query, (year, month))

        except Exception as e:
            self.logger.error(f"Error fetching material usage data: {e}")
            return []

    def _get_store_data(self) -> Dict[int, Dict[str, Any]]:
        """Get store information."""
        try:
            query = """
            SELECT id, name, district_id, is_active
            FROM store 
            WHERE is_active = TRUE
            ORDER BY id
            """

            stores = self.data_provider.db_manager.fetch_all(query)
            return {store['id']: store for store in stores}

        except Exception as e:
            self.logger.error(f"Error fetching store data: {e}")
            return {}

    def _get_material_type_data(self) -> Dict[int, Dict[str, Any]]:
        """Get material type information."""
        try:
            query = """
            SELECT id, name, sort_order
            FROM material_type 
            WHERE is_active = TRUE
            ORDER BY sort_order, name
            """

            types = self.data_provider.db_manager.fetch_all(query)
            return {mt['id']: mt for mt in types}

        except Exception as e:
            self.logger.error(f"Error fetching material type data: {e}")
            return {}

    def _create_header(self, ws, period: str) -> None:
        """Create worksheet header."""
        # Title
        ws['A1'] = f"海底捞物料使用汇总报表 - {period}"
        ws['A1'].font = Font(name='SimSun', size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Merge title cells
        ws.merge_cells('A1:F1')

        # Generation time
        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(name='SimSun', size=10)
        ws.merge_cells('A2:F2')

    def _create_summary_table(self, ws, usage_data: List[Dict], store_data: Dict, material_type_data: Dict) -> None:
        """Create the main summary table."""
        start_row = 4

        # Group data by store
        store_usage = {}
        for row in usage_data:
            store_id = row['store_id']
            if store_id not in store_usage:
                store_usage[store_id] = {
                    'store_name': row['store_name'],
                    'material_types': {},
                    'total': 0
                }

            material_type_id = row['material_type_id']
            usage_amount = float(row['usage_amount']
                                 ) if row['usage_amount'] else 0

            store_usage[store_id]['material_types'][material_type_id] = {
                'name': row['material_type_name'],
                'amount': usage_amount
            }
            store_usage[store_id]['total'] += usage_amount

        # Get all material types for consistent columns
        all_material_types = list(material_type_data.keys())

        current_row = start_row

        for store_id in sorted(store_usage.keys()):
            store_info = store_usage[store_id]

            # Store header
            ws[f'A{current_row}'] = f"门店: {store_info['store_name']}"
            ws[f'A{current_row}'].font = Font(
                name='SimSun', size=12, bold=True)
            ws[f'A{current_row}'].fill = PatternFill(
                start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            ws.merge_cells(f'A{current_row}:F{current_row}')
            current_row += 1

            # Table headers
            ws[f'A{current_row}'] = "物料分类"
            ws[f'B{current_row}'] = "本月使用金额 (CAD)"

            for col in ['A', 'B']:
                cell = ws[f'{col}{current_row}']
                cell.font = Font(name='SimSun', size=11, bold=True)
                cell.fill = PatternFill(
                    start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')

            current_row += 1
            data_start_row = current_row

            # Material type data
            for material_type_id in all_material_types:
                material_type_info = material_type_data[material_type_id]
                material_usage = store_info['material_types'].get(
                    material_type_id, {'name': material_type_info['name'], 'amount': 0})

                # Only show rows with usage > 0
                if material_usage['amount'] > 0:
                    ws[f'A{current_row}'] = material_usage['name']
                    ws[f'B{current_row}'] = round(material_usage['amount'], 2)

                    # Format the cells
                    ws[f'A{current_row}'].font = Font(name='SimSun', size=10)
                    ws[f'B{current_row}'].font = Font(name='SimSun', size=10)
                    ws[f'B{current_row}'].number_format = '#,##0.00'
                    ws[f'B{current_row}'].alignment = Alignment(
                        horizontal='right')

                    current_row += 1

            # Store total
            ws[f'A{current_row}'] = "总计"
            ws[f'B{current_row}'] = round(store_info['total'], 2)

            for col in ['A', 'B']:
                cell = ws[f'{col}{current_row}']
                cell.font = Font(name='SimSun', size=11, bold=True)
                cell.fill = PatternFill(
                    start_color="FFFF99", end_color="FFFF99", fill_type="solid")

            ws[f'B{current_row}'].number_format = '#,##0.00'
            ws[f'B{current_row}'].alignment = Alignment(horizontal='right')

            # Add borders to the store section
            self._add_table_borders(
                ws, data_start_row - 1, current_row, 'A', 'B')

            current_row += 2  # Space between stores

        # Grand total across all stores
        if store_usage:
            grand_total = sum(store_info['total']
                              for store_info in store_usage.values())

            ws[f'A{current_row}'] = "所有门店总计"
            ws[f'B{current_row}'] = round(grand_total, 2)

            for col in ['A', 'B']:
                cell = ws[f'{col}{current_row}']
                cell.font = Font(name='SimSun', size=12, bold=True)
                cell.fill = PatternFill(
                    start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

            ws[f'B{current_row}'].number_format = '#,##0.00'
            ws[f'B{current_row}'].alignment = Alignment(horizontal='right')

            # Add border to grand total
            self._add_table_borders(ws, current_row, current_row, 'A', 'B')

    def _add_table_borders(self, ws, start_row: int, end_row: int, start_col: str, end_col: str) -> None:
        """Add borders to a table range."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        start_col_idx = ord(start_col) - ord('A') + 1
        end_col_idx = ord(end_col) - ord('A') + 1

        for row in range(start_row, end_row + 1):
            for col_idx in range(start_col_idx, end_col_idx + 1):
                col_letter = get_column_letter(col_idx)
                ws[f'{col_letter}{row}'].border = thin_border

    def _apply_formatting(self, ws) -> None:
        """Apply general formatting to the worksheet."""
        # Set column widths
        ws.column_dimensions['A'].width = 25  # Material type names
        ws.column_dimensions['B'].width = 20  # Usage amounts

        # Set row height for better readability
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
