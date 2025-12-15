"""Monthly Store Revenue Compare Worksheet Generator

Generates monthly revenue and turnover rate comparison reports for all Canadian stores.
"""

import os
import shutil
import logging
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from decimal import Decimal

from utils.database import DatabaseManager, DatabaseConfig

logger = logging.getLogger(__name__)


class MonthlyStoreRevenueCompareWorksheet:
    """Generate monthly store revenue and turnover rate comparison worksheet"""

    def __init__(self, database_manager: DatabaseManager = None):
        """Initialize the worksheet generator
        
        Args:
            database_manager: Database manager instance (optional, will create if not provided)
        """
        if database_manager is None:
            config = DatabaseConfig()
            database_manager = DatabaseManager(config)
        self.db_manager = database_manager

    def generate_report(self, year: int, month: int, output_file: str) -> bool:
        """Generate the monthly store revenue comparison report
        
        Args:
            year: Target year (e.g., 2025)
            month: Target month (1-12)
            output_file: Path to save the output file
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Get template file path
            template_file = self._get_template_path()
            if not os.path.exists(template_file):
                logger.error(f"Template file not found: {template_file}")
                return False
            
            # Create output directory if needed
            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
            
            # Copy template to output location
            shutil.copy2(template_file, output_file)
            logger.info(f"Copied template to: {output_file}")
            
            # Load workbook
            wb = openpyxl.load_workbook(output_file)
            ws = wb.active
            
            # Fetch data from database
            current_data = self._fetch_monthly_data(year, month)
            prev_year_data = self._fetch_monthly_data(year - 1, month)  # 同比 - same month last year
            
            # Calculate previous month for 环比
            if month == 1:
                prev_month_year = year - 1
                prev_month_num = 12
            else:
                prev_month_year = year
                prev_month_num = month - 1
            prev_month_data = self._fetch_monthly_data(prev_month_year, prev_month_num)  # 环比 - last month
            
            if not current_data:
                logger.warning(f"No data found for {year}-{month:02d}")
            
            # Fill the worksheet
            self._fill_worksheet(ws, current_data, prev_year_data, prev_month_data, year, month)
            
            # Save workbook
            wb.save(output_file)
            wb.close()
            
            logger.info(f"Successfully generated monthly store revenue comparison report: {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Error generating monthly store revenue comparison report: {str(e)}")
            return False

    def _get_template_path(self) -> str:
        """Get the template file path"""
        current_file = os.path.abspath(__file__)
        project_root = os.path.dirname(os.path.dirname(current_file))
        
        template_path = os.path.join(
            project_root, 'sheet-templates', 
            'monthly-store-revenue-compare-template.xlsx'
        )
        
        return template_path

    def _fetch_monthly_data(self, year: int, month: int, fetch_usd_rate: bool = True) -> Dict[str, Dict]:
        """Fetch monthly data for all stores from database
        
        Args:
            year: Target year
            month: Target month
            
        Returns:
            Dict mapping store name to data
        """
        try:
            # Build date range for the month
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(days=1)
            
            query = """
                SELECT
                    s.name as store_name,
                    s.seats_total as seating_capacity,
                    COALESCE(SUM(d.revenue_tax_not_included), 0) as total_revenue,
                    COALESCE(SUM(d.tables_served_validated), 0) as total_tables_served,
                    COALESCE(AVG(d.tables_served), 0) as avg_tables_served,
                    COALESCE(AVG(d.tables_served_validated), 0) as avg_tables_validated,
                    COUNT(DISTINCT d.date) as operating_days
                FROM store s
                LEFT JOIN daily_report d ON s.id = d.store_id
                    AND d.date >= %s
                    AND d.date <= %s
                WHERE s.id <= 8  -- Only Canadian stores
                GROUP BY s.id, s.name, s.seats_total
                ORDER BY s.id
            """
            
            with self.db_manager.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query, (start_date, end_date))
                    results = cursor.fetchall()
            
            # Convert to dictionary
            data = {}
            for row in results:
                # Handle both tuple and dict cursor results
                if isinstance(row, dict):
                    store_name = row['store_name']
                    seating_capacity = float(row['seating_capacity']) if row['seating_capacity'] else 50
                    total_tables_served = float(row['total_tables_served']) if row['total_tables_served'] else 0
                    operating_days = row['operating_days'] if row['operating_days'] else 0

                    # Calculate correct monthly average turnover rate
                    # Formula: total_tables_served / (operating_days × seating_capacity)
                    if operating_days > 0 and seating_capacity > 0:
                        turnover_rate = total_tables_served / (operating_days * seating_capacity)
                    else:
                        turnover_rate = 0

                    data[store_name] = {
                        'revenue': float(row['total_revenue']) if row['total_revenue'] else 0,
                        'turnover_rate': turnover_rate,
                        'tables_served': float(row['avg_tables_served']) if row['avg_tables_served'] else 0,
                        'tables_validated': float(row['avg_tables_validated']) if row['avg_tables_validated'] else 0,
                        'operating_days': operating_days,
                        'seating_capacity': seating_capacity,
                        'total_tables_served': total_tables_served
                    }
                else:
                    # Tuple format handling
                    store_name = row[0]
                    seating_capacity = float(row[1]) if row[1] else 50
                    total_revenue = float(row[2]) if row[2] else 0
                    total_tables_served = float(row[3]) if row[3] else 0
                    operating_days = row[6] if row[6] else 0

                    # Calculate correct monthly average turnover rate
                    if operating_days > 0 and seating_capacity > 0:
                        turnover_rate = total_tables_served / (operating_days * seating_capacity)
                    else:
                        turnover_rate = 0

                    data[store_name] = {
                        'revenue': total_revenue,
                        'turnover_rate': turnover_rate,
                        'operating_days': operating_days,
                        'seating_capacity': seating_capacity,
                        'total_tables_served': total_tables_served
                    }
            
            logger.info(f"Fetched data for {len(data)} stores for {year}-{month:02d}")
            # Fetch USD rate if requested
            if fetch_usd_rate:
                usd_rate = self._fetch_usd_rate(year, month)
                for store_data in data.values():
                    store_data['usd_rate'] = usd_rate
            
            return data
            
        except Exception as e:
            logger.error(f"Error fetching monthly data: {str(e)}", exc_info=True)
            return {}
    
    def _get_store_seats(self, store_name: str) -> int:
        """Get seats_total for a store by name
        
        Args:
            store_name: Store name in Chinese
            
        Returns:
            Number of seats for the store
        """
        # Hardcoded based on database values to avoid extra queries
        seats_mapping = {
            '加拿大一店': 53,
            '加拿大二店': 36,
            '加拿大三店': 48,
            '加拿大四店': 70,
            '加拿大五店': 55,
            '加拿大六店': 56,
            '加拿大七店': 57,
            "加拿大八店": 56
        }
        return seats_mapping.get(store_name, 50)  # Default to 50 if not found
    
    def _fetch_usd_rate(self, year: int, month: int) -> float:
        """Fetch CAD to USD conversion rate for the given month
        
        Args:
            year: Target year
            month: Target month
            
        Returns:
            CAD to USD conversion rate
        """
        try:
            query = """
                SELECT cad_usd_rate 
                FROM month_static_data 
                WHERE month = %s
            """
            
            month_date = datetime(year, month, 1)
            
            with self.db_manager.get_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(query, (month_date,))
                    result = cursor.fetchone()
                    
                    if result and result['cad_usd_rate']:
                        return float(result['cad_usd_rate'])
                    else:
                        logger.warning(f"No USD rate found for {year}-{month:02d}, using default 0.75")
                        return 0.75
                        
        except Exception as e:
            logger.error(f"Error fetching USD rate: {str(e)}")
            return 0.75

    def _write_cell_value(self, ws, cell_ref: str, value):
        """Write value to a cell, handling merged cells properly"""
        cell = ws[cell_ref]
        # Check if cell is part of a merged range
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Write to the top-left cell of the merged range
                min_col, min_row = merged_range.min_col, merged_range.min_row
                ws.cell(row=min_row, column=min_col, value=value)
                return
        # Not merged, write directly
        cell.value = value
    
    def _fill_worksheet(self, ws, current_data: Dict, prev_year_data: Dict, prev_month_data: Dict, year: int, month: int):
        """Fill the worksheet with data
        
        Args:
            ws: Worksheet object
            current_data: Current year/month data
            prev_data: Previous year same month data  
            year: Target year
            month: Target month
        """
        # Update headers with year-month
        current_period = f"{year}-{month:01d}"  # e.g., "2025-7"
        prev_period = f"{year-1}-{month:01d}"    # e.g., "2024-7"
        
        # Update period headers in row 2 and row 16 for both sections
        # Row 2 for revenue section
        self._write_cell_value(ws, 'C2', current_period)  # CAD current
        self._write_cell_value(ws, 'H2', current_period)  # USD current
        
        # Row 16 for turnover section
        self._write_cell_value(ws, 'C16', current_period)  # CAD current
        self._write_cell_value(ws, 'H16', current_period)  # USD current
        
        # Row 15 has merged cells C15:G15, so skip updating them
        # The template already has the proper structure
        
        # Store name mapping (matching template order)
        store_mapping = {
            '加拿大一店': ('张森磊', 3),
            '加拿大二店': ('潘多治', 4),
            '加拿大三店': ('Bao Xiaoyun', 5),
            '加拿大四店': ('李俊娟', 6),
            '加拿大五店': ('陈浩', 7),
            '加拿大六店': ('高新荣', 8),
            '加拿大七店': ('海参威', 9),
            '加拿大八店': ('', 10)
        }
        
        # Fill revenue data (top section, rows 3-9)
        # LEFT section (CAD): C=current, D=同比(previous year), F=环比(previous month - we'll use prev year for now)
        #   - Shows revenue in 万元 (divide by 10000)
        # RIGHT section (labeled "USD" but actually shows turnover rates): H=current, I=同比, L=环比
        #   - Shows turnover rates (tables_served / seats_total)
        # NOTE: Despite being labeled "USD", the right section shows turnover rates, not USD revenue
        
        # Get USD rates
        current_usd_rate = list(current_data.values())[0].get('usd_rate', 0.75) if current_data else 0.75
        prev_usd_rate = list(prev_year_data.values())[0].get('usd_rate', 0.75) if prev_year_data else 0.75
        
        for store_name, (manager, row) in store_mapping.items():
            # Current year revenue (convert to 万元)
            current_revenue = current_data.get(store_name, {}).get('revenue', 0)
            current_revenue_wan = current_revenue / 10000 if current_revenue > 0 else 0
            
            # Previous year revenue (convert to 万元) - for 同比
            prev_year_revenue = prev_year_data.get(store_name, {}).get('revenue', 0)
            prev_year_revenue_wan = prev_year_revenue / 10000 if prev_year_revenue > 0 else 0
            
            # Previous month revenue (convert to 万元) - for 环比
            prev_month_revenue = prev_month_data.get(store_name, {}).get('revenue', 0)
            prev_month_revenue_wan = prev_month_revenue / 10000 if prev_month_revenue > 0 else 0
            
            # LEFT section (CAD values in C, D, F columns)
            if current_revenue_wan > 0:
                self._write_cell_value(ws, f'C{row}', round(current_revenue_wan, 2))
            
            if prev_year_revenue_wan > 0:
                self._write_cell_value(ws, f'D{row}', round(prev_year_revenue_wan, 2))  # 同比 - last year
            
            if prev_month_revenue_wan > 0:
                self._write_cell_value(ws, f'F{row}', round(prev_month_revenue_wan, 2))  # 环比 - last month
            
            # RIGHT section (H, I, L columns) - Turnover rates
            # Despite column headers saying "USD", this section shows turnover rates
            # Use the turnover_rate directly from database (already validated)
            current_turnover = current_data.get(store_name, {}).get('turnover_rate', 0)
            prev_year_turnover = prev_year_data.get(store_name, {}).get('turnover_rate', 0)  # 同比
            prev_month_turnover = prev_month_data.get(store_name, {}).get('turnover_rate', 0)  # 环比
            
            if current_turnover > 0:
                self._write_cell_value(ws, f'H{row}', round(current_turnover, 2))
            
            if prev_year_turnover > 0:
                self._write_cell_value(ws, f'I{row}', round(prev_year_turnover, 2))  # 同比 - last year
            
            if prev_month_turnover > 0:
                self._write_cell_value(ws, f'L{row}', round(prev_month_turnover, 2))  # 环比 - last month
        
        # Row 10 - Override the USD section (H, I, L columns) with weighted average for turnover rates
        # The left section (C, D, F) has SUM formulas for revenue which are correct
        # The right section (H, I, L) should use weighted average for turnover rates

        # Calculate weighted average for current period turnover (same calculation as row 24)
        total_weighted_turnover_10 = 0
        total_capacity_10 = 0

        for store_name in store_mapping.keys():
            capacity = self._get_store_seats(store_name)
            turnover = current_data.get(store_name, {}).get('turnover_rate', 0)
            if turnover > 0:
                total_weighted_turnover_10 += turnover * capacity
                total_capacity_10 += capacity

        current_weighted_avg_10 = total_weighted_turnover_10 / total_capacity_10 if total_capacity_10 > 0 else 0

        # Calculate weighted average for previous year
        total_weighted_prev_year_10 = 0
        total_capacity_prev_year_10 = 0

        for store_name in store_mapping.keys():
            capacity = self._get_store_seats(store_name)
            turnover = prev_year_data.get(store_name, {}).get('turnover_rate', 0)
            if turnover > 0:
                total_weighted_prev_year_10 += turnover * capacity
                total_capacity_prev_year_10 += capacity

        prev_year_weighted_avg_10 = total_weighted_prev_year_10 / total_capacity_prev_year_10 if total_capacity_prev_year_10 > 0 else 0

        # Calculate weighted average for previous month
        total_weighted_prev_month_10 = 0
        total_capacity_prev_month_10 = 0

        for store_name in store_mapping.keys():
            capacity = self._get_store_seats(store_name)
            turnover = prev_month_data.get(store_name, {}).get('turnover_rate', 0)
            if turnover > 0:
                total_weighted_prev_month_10 += turnover * capacity
                total_capacity_prev_month_10 += capacity

        prev_month_weighted_avg_10 = total_weighted_prev_month_10 / total_capacity_prev_month_10 if total_capacity_prev_month_10 > 0 else 0

        # Write weighted averages to row 10 (revenue section, USD/turnover columns)
        if current_weighted_avg_10 > 0:
            self._write_cell_value(ws, 'H10', round(current_weighted_avg_10, 2))
        if prev_year_weighted_avg_10 > 0:
            self._write_cell_value(ws, 'I10', round(prev_year_weighted_avg_10, 2))
        if prev_month_weighted_avg_10 > 0:
            self._write_cell_value(ws, 'L10', round(prev_month_weighted_avg_10, 2))

        # Fill turnover rate data (bottom section, rows 17-23)
        # Note: Turnover section starts at row 17, not 16
        # IMPORTANT: Based on Sheet1 example analysis:
        # CAD section: Shows CAD revenue converted to USD (CAD revenue × USD rate)
        # USD section: Shows turnover rates (tables/seats) - same as revenue section USD values
        
        for store_name, (manager, orig_row) in store_mapping.items():
            revenue_row = orig_row  # Row 3-9 for revenue
            turnover_row = orig_row + 14  # Row 17-23 for turnover (17 = 3 + 14)
            
            # Get revenue data (same as used in revenue section)
            current_revenue = current_data.get(store_name, {}).get('revenue', 0)
            prev_year_revenue = prev_year_data.get(store_name, {}).get('revenue', 0)
            prev_month_revenue = prev_month_data.get(store_name, {}).get('revenue', 0)
            
            # Convert to 万元
            current_revenue_wan = current_revenue / 10000 if current_revenue > 0 else 0
            prev_year_revenue_wan = prev_year_revenue / 10000 if prev_year_revenue > 0 else 0
            prev_month_revenue_wan = prev_month_revenue / 10000 if prev_month_revenue > 0 else 0
            
            # Get USD rates for each period
            prev_month_usd_rate = list(prev_month_data.values())[0].get('usd_rate', 0.75) if prev_month_data else 0.75
            
            # CAD section (C, D, F columns) - CAD revenue converted to USD
            if current_revenue_wan > 0:
                current_revenue_usd = current_revenue_wan * current_usd_rate
                self._write_cell_value(ws, f'C{turnover_row}', round(current_revenue_usd, 2))
            
            if prev_year_revenue_wan > 0:
                prev_year_revenue_usd = prev_year_revenue_wan * prev_usd_rate
                self._write_cell_value(ws, f'D{turnover_row}', round(prev_year_revenue_usd, 2))  # 同比
            
            if prev_month_revenue_wan > 0:
                prev_month_revenue_usd = prev_month_revenue_wan * prev_month_usd_rate
                self._write_cell_value(ws, f'F{turnover_row}', round(prev_month_revenue_usd, 2))  # 环比
            
            # USD section (H, I, L columns) - Shows turnover rates
            # Same as revenue section USD values (both show turnover_rate from database)
            current_turnover = current_data.get(store_name, {}).get('turnover_rate', 0)
            prev_year_turnover = prev_year_data.get(store_name, {}).get('turnover_rate', 0)
            prev_month_turnover = prev_month_data.get(store_name, {}).get('turnover_rate', 0)
            
            if current_turnover > 0:
                self._write_cell_value(ws, f'H{turnover_row}', round(current_turnover, 2))
            
            if prev_year_turnover > 0:
                self._write_cell_value(ws, f'I{turnover_row}', round(prev_year_turnover, 2))  # 同比
            
            if prev_month_turnover > 0:
                self._write_cell_value(ws, f'L{turnover_row}', round(prev_month_turnover, 2))  # 环比
        
        # Row 24 - Calculate weighted averages for turnover rates (not simple AVERAGE)
        # Use weighted average by seating capacity for accuracy (same as daily tracking and time segment sheets)

        # Calculate weighted average for current period turnover (column H, row 24)
        total_weighted_turnover = 0
        total_capacity = 0

        for store_name in store_mapping.keys():
            capacity = self._get_store_seats(store_name)
            turnover = current_data.get(store_name, {}).get('turnover_rate', 0)
            if turnover > 0:  # Only include stores with data
                total_weighted_turnover += turnover * capacity
                total_capacity += capacity

        current_weighted_avg = total_weighted_turnover / total_capacity if total_capacity > 0 else 0

        # Calculate weighted average for previous year (同比)
        total_weighted_prev_year = 0
        total_capacity_prev_year = 0

        for store_name in store_mapping.keys():
            capacity = self._get_store_seats(store_name)
            turnover = prev_year_data.get(store_name, {}).get('turnover_rate', 0)
            if turnover > 0:
                total_weighted_prev_year += turnover * capacity
                total_capacity_prev_year += capacity

        prev_year_weighted_avg = total_weighted_prev_year / total_capacity_prev_year if total_capacity_prev_year > 0 else 0

        # Calculate weighted average for previous month (环比)
        total_weighted_prev_month = 0
        total_capacity_prev_month = 0

        for store_name in store_mapping.keys():
            capacity = self._get_store_seats(store_name)
            turnover = prev_month_data.get(store_name, {}).get('turnover_rate', 0)
            if turnover > 0:
                total_weighted_prev_month += turnover * capacity
                total_capacity_prev_month += capacity

        prev_month_weighted_avg = total_weighted_prev_month / total_capacity_prev_month if total_capacity_prev_month > 0 else 0

        # Write weighted averages to row 24 (turnover section)
        # Row 24 is the average row for the turnover section
        if current_weighted_avg > 0:
            self._write_cell_value(ws, 'H24', round(current_weighted_avg, 2))
        if prev_year_weighted_avg > 0:
            self._write_cell_value(ws, 'I24', round(prev_year_weighted_avg, 2))
        if prev_month_weighted_avg > 0:
            self._write_cell_value(ws, 'L24', round(prev_month_weighted_avg, 2))

        # The template already has formulas for revenue totals (rows 10) and calculations
        # We override the turnover average (row 24) with weighted average for accuracy

        logger.info(f"Filled worksheet with data for {year}-{month:02d}")