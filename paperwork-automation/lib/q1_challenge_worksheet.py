#!/usr/bin/env python3
"""
Q1 2026 Challenge Tracking Worksheet Generator

This module generates a Q1 2026 challenge tracking worksheet (2026 Q1 挑战)
that compares weekly performance against Q1 2026 challenge targets.

Target Rules (from Canada region Q1 2026 challenge document):
- Store 6: Fixed turnover target of 3.65 (road construction exemption)
- Store 8: Fixed turnover target of 4.0 (new store, excluded from regional YoY)
- Stores 1-5, 7: Turnover target = last year same week + 0.16
- All stores except 8: Tables target = last year same week + 56 (8 tables/day × 7 days)

Visual indicators:
- Green (E6FFE6): Target met (current >= target)
- Red (FFE6E6): Target not met (current < target)
"""

import logging
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Import centralized configurations
from configs.store_config import STORE_SEATING_CAPACITY, STORE_MANAGERS, REGIONAL_MANAGER
from configs.challenge_targets import (
    get_store_turnover_target,
    get_store_tables_target,
    is_store_excluded_from_regional,
    is_q1_2026_active,
    STORE_TARGET_CONFIG,
    DEFAULT_TURNOVER_IMPROVEMENT,
    WEEKLY_TABLES_IMPROVEMENT
)


class Q1ChallengeWorksheetGenerator:
    """Generator for Q1 2026 challenge tracking worksheet."""

    # Standard color fills
    GREEN_FILL = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    SUMMARY_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    EXCLUDED_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    def __init__(self, data_provider):
        """
        Initialize the generator with data provider.

        Args:
            data_provider: ReportDataProvider instance for database access
        """
        self.data_provider = data_provider
        self.logger = logging.getLogger(__name__)

        # Standard styling
        self.header_font = Font(bold=True, color="FFFFFF")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")

    def generate_worksheet(self, workbook: Workbook, target_date: str) -> Optional[Worksheet]:
        """
        Generate Q1 2026 challenge tracking worksheet.

        Args:
            workbook: Excel workbook to add worksheet to
            target_date: Target date in YYYY-MM-DD format (end of 7-day period)

        Returns:
            The generated worksheet, or None if not in Q1 2026
        """
        # Check if we're in Q1 2026
        if not is_q1_2026_active(target_date):
            self.logger.info(f"Target date {target_date} is not in Q1 2026, skipping challenge worksheet")
            return None

        try:
            self.logger.info(
                f"Generating Q1 2026 challenge worksheet for week ending {target_date}")

            # Create worksheet
            ws = workbook.create_sheet("2026 Q1 挑战")

            # Parse target date and calculate date ranges
            target_dt = datetime.strptime(target_date, '%Y-%m-%d')
            start_dt = target_dt - timedelta(days=6)

            # Calculate previous year dates with same weekday alignment
            prev_year_start, prev_year_end = self._calculate_prev_year_dates(
                start_dt, target_dt)

            # Get weekly store performance data
            store_data = self._get_weekly_data(start_dt, target_dt)

            if not store_data:
                self.logger.warning("No data available for Q1 challenge worksheet")
                return ws

            # Generate headers
            self._generate_headers(ws, start_dt, target_dt, prev_year_start, prev_year_end)

            # Add store data rows with targets
            self._add_store_challenge_data(ws, store_data, start_row=3)

            # Add regional summary (excluding Store 8)
            self._add_regional_summary(ws, store_data, row=3 + len(store_data))

            # Apply formatting
            self._apply_formatting(ws, len(store_data))

            self.logger.info(
                f"Q1 challenge worksheet generated with {len(store_data)} stores")

            return ws

        except Exception as e:
            self.logger.error(
                f"Error generating Q1 challenge worksheet: {str(e)}")
            raise

    def _calculate_prev_year_dates(self, current_start: datetime,
                                    current_end: datetime) -> Tuple[datetime, datetime]:
        """Calculate previous year week with same calendar dates."""
        prev_year = current_end.year - 1

        # Use same calendar dates in previous year
        # Handle leap year dates (Feb 29 -> Feb 28 in non-leap year)
        try:
            prev_year_end = current_end.replace(year=prev_year)
        except ValueError:
            prev_year_end = current_end.replace(year=prev_year, day=28)

        try:
            prev_year_start = current_start.replace(year=prev_year)
        except ValueError:
            prev_year_start = current_start.replace(year=prev_year, day=28)

        return prev_year_start, prev_year_end

    def _get_weekly_data(self, start_dt: datetime, end_dt: datetime) -> List[Dict[str, Any]]:
        """Get weekly store performance data from database."""
        try:
            store_data = self.data_provider.get_weekly_store_performance(
                start_dt.strftime('%Y-%m-%d'), end_dt.strftime('%Y-%m-%d'))

            if not store_data:
                self.logger.warning("No weekly store performance data found")
                return []

            self.logger.info(f"Retrieved weekly data for {len(store_data)} stores")
            return store_data

        except Exception as e:
            self.logger.error(f"Error getting weekly data: {str(e)}")
            return []

    def _generate_headers(self, ws: Worksheet, start_dt: datetime, end_dt: datetime,
                          prev_start_dt: datetime, prev_end_dt: datetime) -> None:
        """Generate worksheet headers."""
        # Format date strings
        current_year = end_dt.year % 100
        prev_year = prev_end_dt.year % 100

        current_period = f"{current_year}年{start_dt.month}月{start_dt.day}日-{end_dt.month}月{end_dt.day}日"
        prev_period = f"{prev_year}年{prev_start_dt.month}月{prev_start_dt.day}日-{prev_end_dt.month}月{prev_end_dt.day}日"

        # Row 1 - Main headers
        headers_row1 = [
            "门店名称",      # A
            "翻台率挑战",    # B - Turnover challenge section
            "", "", "", "",  # C-F
            "桌数挑战",      # G - Tables challenge section
            "", "", "", "",  # H-K
            "备注"          # L - Notes
        ]

        for col, header in enumerate(headers_row1, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.HEADER_FILL
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

        # Merge header cells
        ws.merge_cells('B1:F1')  # 翻台率挑战
        ws.merge_cells('G1:K1')  # 桌数挑战

        # Row 2 - Sub headers
        sub_headers = [
            "",              # A - Store name
            prev_period,     # B - Previous year turnover
            "目标",          # C - Target turnover
            current_period,  # D - Current turnover
            "差距",          # E - Gap
            "达成",          # F - Achievement
            prev_period,     # G - Previous year tables
            "目标",          # H - Target tables
            current_period,  # I - Current tables
            "差距",          # J - Gap
            "达成",          # K - Achievement
            ""               # L - Notes
        ]

        for col, header in enumerate(sub_headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.HEADER_FILL
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

    def _add_store_challenge_data(self, ws: Worksheet, store_data: List[Dict[str, Any]],
                                   start_row: int) -> None:
        """Add store challenge data rows with targets and color coding."""
        for i, store in enumerate(store_data):
            row = start_row + i
            store_id = store.get('store_id', 0)

            # Get target configuration
            config = STORE_TARGET_CONFIG.get(store_id, {})
            is_excluded = is_store_excluded_from_regional(store_id)

            # Store name
            ws.cell(row=row, column=1, value=store['store_name'])

            # === Turnover Rate Section ===
            prev_turnover = float(store.get('prev_avg_turnover_rate', 0))
            current_turnover = float(store.get('current_avg_turnover_rate', 0))
            target_turnover = get_store_turnover_target(store_id, prev_turnover)
            turnover_gap = current_turnover - target_turnover
            turnover_achieved = current_turnover >= target_turnover

            # Previous year turnover
            ws.cell(row=row, column=2, value=prev_turnover)
            # Target turnover
            ws.cell(row=row, column=3, value=target_turnover)
            # Current turnover
            ws.cell(row=row, column=4, value=current_turnover)
            # Gap
            gap_cell = ws.cell(row=row, column=5, value=turnover_gap)
            # Achievement
            achievement_cell = ws.cell(row=row, column=6, value="达成" if turnover_achieved else "未达成")
            self._apply_achievement_color(achievement_cell, turnover_achieved)

            # === Tables Section ===
            prev_tables = float(store.get('prev_total_tables', 0))
            current_tables = float(store.get('current_total_tables', 0))
            target_tables = get_store_tables_target(store_id, int(prev_tables))

            # Previous year tables
            ws.cell(row=row, column=7, value=int(prev_tables))

            if target_tables is not None:
                tables_gap = current_tables - target_tables
                tables_achieved = current_tables >= target_tables

                # Target tables
                ws.cell(row=row, column=8, value=target_tables)
                # Current tables
                ws.cell(row=row, column=9, value=int(current_tables))
                # Gap
                ws.cell(row=row, column=10, value=int(tables_gap))
                # Achievement
                achievement_cell = ws.cell(row=row, column=11, value="达成" if tables_achieved else "未达成")
                self._apply_achievement_color(achievement_cell, tables_achieved)
            else:
                # Store 8 - no tables target
                ws.cell(row=row, column=8, value="N/A")
                ws.cell(row=row, column=9, value=int(current_tables))
                ws.cell(row=row, column=10, value="N/A")
                na_cell = ws.cell(row=row, column=11, value="不考核")
                na_cell.fill = self.EXCLUDED_FILL

            # Notes
            notes = config.get('notes', '')
            exemption = config.get('exemption_reason', '')
            if exemption:
                notes = f"{notes} ({exemption})" if notes else exemption
            ws.cell(row=row, column=12, value=notes)

    def _add_regional_summary(self, ws: Worksheet, store_data: List[Dict[str, Any]],
                               row: int) -> None:
        """Add regional summary row (excluding Store 8)."""
        # Filter out Store 8
        regional_stores = [s for s in store_data if s.get('store_id') != 8]

        if not regional_stores:
            return

        ws.cell(row=row, column=1, value="区域汇总(不含8店)")

        # Calculate weighted average turnover
        total_seats = sum(
            STORE_SEATING_CAPACITY.get(s['store_id'], 50) for s in regional_stores)

        # Previous year weighted turnover
        prev_weighted_turnover = sum(
            float(s.get('prev_avg_turnover_rate', 0)) *
            STORE_SEATING_CAPACITY.get(s['store_id'], 50)
            for s in regional_stores
        ) / total_seats if total_seats > 0 else 0

        # Current weighted turnover
        current_weighted_turnover = sum(
            float(s.get('current_avg_turnover_rate', 0)) *
            STORE_SEATING_CAPACITY.get(s['store_id'], 50)
            for s in regional_stores
        ) / total_seats if total_seats > 0 else 0

        # Regional target = prev + 0.16 (the standard improvement)
        target_weighted_turnover = prev_weighted_turnover + DEFAULT_TURNOVER_IMPROVEMENT

        ws.cell(row=row, column=2, value=prev_weighted_turnover)
        ws.cell(row=row, column=3, value=target_weighted_turnover)
        ws.cell(row=row, column=4, value=current_weighted_turnover)

        turnover_gap = current_weighted_turnover - target_weighted_turnover
        turnover_achieved = current_weighted_turnover >= target_weighted_turnover

        ws.cell(row=row, column=5, value=turnover_gap)
        achievement_cell = ws.cell(row=row, column=6, value="达成" if turnover_achieved else "未达成")
        self._apply_achievement_color(achievement_cell, turnover_achieved)

        # Sum tables for regional
        prev_tables = sum(float(s.get('prev_total_tables', 0)) for s in regional_stores)
        current_tables = sum(float(s.get('current_total_tables', 0)) for s in regional_stores)
        # Regional target = prev + (56 per store × 7 stores)
        target_tables = prev_tables + (WEEKLY_TABLES_IMPROVEMENT * len(regional_stores))

        ws.cell(row=row, column=7, value=int(prev_tables))
        ws.cell(row=row, column=8, value=int(target_tables))
        ws.cell(row=row, column=9, value=int(current_tables))

        tables_gap = current_tables - target_tables
        tables_achieved = current_tables >= target_tables

        ws.cell(row=row, column=10, value=int(tables_gap))
        achievement_cell = ws.cell(row=row, column=11, value="达成" if tables_achieved else "未达成")
        self._apply_achievement_color(achievement_cell, tables_achieved)

        ws.cell(row=row, column=12, value="区域整体目标(翻台率+0.16,日均桌数+8)")

        # Style the summary row
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.SUMMARY_FILL
            cell.font = Font(bold=True)

    def _apply_achievement_color(self, cell, achieved: bool) -> None:
        """Apply green/red color based on achievement."""
        cell.fill = self.GREEN_FILL if achieved else self.RED_FILL
        cell.font = Font(bold=True, color="006400" if achieved else "8B0000")

    def _apply_formatting(self, ws: Worksheet, num_stores: int) -> None:
        """Apply professional formatting to the worksheet."""
        # Set column widths
        column_widths = {
            'A': 18,  # Store name
            'B': 12, 'C': 10, 'D': 12, 'E': 10, 'F': 10,  # Turnover
            'G': 12, 'H': 10, 'I': 12, 'J': 10, 'K': 10,  # Tables
            'L': 30   # Notes
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Format data rows
        data_rows = range(3, 3 + num_stores + 1)  # Store data + summary
        for row in data_rows:
            for col in range(1, 13):
                cell = ws.cell(row=row, column=col)
                cell.border = self.thin_border
                cell.alignment = self.center_alignment

                # Number formatting
                if col in [2, 3, 4, 5]:  # Turnover rates
                    if cell.value != "N/A":
                        cell.number_format = '0.00'
                elif col in [7, 8, 9, 10]:  # Tables (integers)
                    if cell.value != "N/A":
                        cell.number_format = '0'

        # Freeze panes
        ws.freeze_panes = "B3"

        self.logger.info("Formatting applied to Q1 challenge worksheet")
