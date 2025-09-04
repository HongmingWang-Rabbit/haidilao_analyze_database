"""
Daily Store Performance Tracking Worksheet Generator

This module generates a daily store performance tracking worksheet (跟踪表-加拿大)
that compares current year vs previous year performance with normalized scoring and ranking.

The worksheet includes:
- Table turnover rate comparison and scoring
- Daily revenue comparison and scoring  
- Comprehensive weighted scoring (25% each metric)
- Automatic ranking system
"""

import logging
from typing import Dict, List, Tuple, Any
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class DailyStoreTrackingGenerator:
    """Generator for daily store performance tracking worksheet."""

    def __init__(self, data_provider):
        """Initialize the generator with data provider."""
        self.data_provider = data_provider
        self.logger = logging.getLogger(__name__)

    def generate_worksheet(self, workbook: Workbook, target_date: str) -> Worksheet:
        """
        Generate daily store performance tracking worksheet.

        Args:
            workbook: Excel workbook to add worksheet to
            target_date: Target date in YYYY-MM-DD format

        Returns:
            The generated worksheet
        """
        try:
            self.logger.info(
                f"Generating daily store tracking worksheet for {target_date}")

            # Create worksheet
            ws = workbook.create_sheet("门店日-加拿大")

            # Parse target date
            target_dt = datetime.strptime(target_date, '%Y-%m-%d')
            prev_year_dt = target_dt.replace(year=target_dt.year - 1)

            # Generate headers
            self._generate_headers(ws, target_dt, prev_year_dt)

            # Get store data
            store_data = self._get_store_performance_data(target_date)

            # Add regional summary (row 3)
            self._add_regional_summary(ws, store_data, target_dt, prev_year_dt)

            # Add store data (rows 4+)
            self._add_store_data(ws, store_data, target_dt, prev_year_dt)

            # Apply formatting
            self._apply_formatting(ws, len(store_data))

            self.logger.info(
                f"Daily store tracking worksheet generated with {len(store_data)} stores")

            return ws

        except Exception as e:
            self.logger.error(
                f"Error generating daily store tracking worksheet: {str(e)}")
            raise

    def _generate_headers(self, ws: Worksheet, target_dt: datetime, prev_year_dt: datetime) -> None:
        """Generate worksheet headers."""
        # Row 1 - Main headers
        headers_row1 = [
            "序号",           # A1 - Serial Number
            "门店名称",       # B1 - Store Name
            "店经理",         # C1 - Store Manager
            "餐位数",         # D1 - Seating Capacity
            "24年全年平均翻台率",  # E1 - 2024 Annual Average Table Turnover Rate
            "日翻台率-考核",   # F1 - Daily Table Turnover Rate Assessment
            "",              # G1 - Empty
            "",              # H1 - Empty
            "",              # I1 - Empty
            "日营业收入-不含税(万加元)",  # J1 - Daily Revenue Excluding Tax (10k CAD)
            "",              # K1 - Empty
            "",              # L1 - Empty
            "",              # M1 - Empty
            "综合得分",       # N1 - Comprehensive Score
            "综合排名"        # O1 - Comprehensive Ranking
        ]

        for col, header in enumerate(headers_row1, 1):
            ws.cell(row=1, column=col, value=header)

        # Merge header cells for better visual grouping
        ws.merge_cells('F1:I1')  # Merge table turnover rate section (日翻台率-考核)
        ws.merge_cells('J1:M1')  # Merge revenue section (日营业收入-不含税)

        # Center align the merged headers
        ws['F1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['J1'].alignment = Alignment(horizontal='center', vertical='center')

        # Row 2 - Sub headers with dates
        target_date_str = target_dt.strftime('%y年%m月%d日')
        prev_date_str = prev_year_dt.strftime('%y年%m月%d日')

        # Date headers for table turnover rate section
        ws.cell(row=2, column=6,
                value=f"25年{target_dt.month}月{target_dt.day}日")  # F2
        ws.cell(row=2, column=7,
                value=f"24年{prev_year_dt.month}月{prev_year_dt.day}日")  # G2
        ws.cell(row=2, column=8, value="对比")  # H2 - Comparison
        ws.cell(row=2, column=9, value="精进值得分")  # I2 - Improvement Score

        # Date headers for revenue section (formulas referencing F2, G2)
        ws.cell(row=2, column=10, value="=F2")  # J2
        ws.cell(row=2, column=11, value="=G2")  # K2
        ws.cell(row=2, column=12, value="对比")  # L2 - Comparison
        ws.cell(row=2, column=13, value="精进值得分")  # M2 - Improvement Score

    def _get_store_performance_data(self, target_date: str) -> List[Dict[str, Any]]:
        """
        Get store performance data for the target date.

        Args:
            target_date: Target date in YYYY-MM-DD format

        Returns:
            List of store performance data dictionaries
        """
        try:
            # Get store performance data (now includes both current and previous year data)
            store_data = self.data_provider.get_daily_store_performance(
                target_date)

            if not store_data:
                self.logger.warning(
                    "⚠️ No store performance data found, using mock data")
                return self._get_mock_store_data()

            self.logger.info(
                f"✅ Retrieved {len(store_data)} stores with real database data")

            # Validate that we have the required fields
            for store in store_data:
                required_fields = ['store_id', 'store_name', 'manager_name', 'seating_capacity',
                                   'current_turnover_rate', 'current_revenue',
                                   'prev_turnover_rate', 'prev_revenue']

                for field in required_fields:
                    if field not in store:
                        self.logger.warning(
                            f"⚠️ Missing field {field} for store {store.get('store_id', 'unknown')}")

            return store_data

        except Exception as e:
            self.logger.error(
                f"❌ Error getting store performance data: {str(e)}")
            self.logger.warning("⚠️ Falling back to mock data")
            return self._get_mock_store_data()

    def _get_mock_store_data(self) -> List[Dict[str, Any]]:
        """Get mock store data for development/testing."""
        # Mock data in the CORRECT order from reference file: 5,6,3,4,1,7,2
        return [
            {
                'store_id': 5,
                'store_name': '加拿大五店',
                'manager_name': '陈浩',
                'seating_capacity': 55,
                'annual_avg_turnover_2024': 5.5,
                'current_turnover_rate': 5.73,
                'prev_turnover_rate': 6.71,
                'current_revenue': 4.40,
                'prev_revenue': 3.96
            },
            {
                'store_id': 6,
                'store_name': '加拿大六店',
                'manager_name': '高新菊',
                'seating_capacity': 56,
                'annual_avg_turnover_2024': 3.67,
                'current_turnover_rate': 4.05,
                'prev_turnover_rate': 4.16,
                'current_revenue': 3.14,
                'prev_revenue': 2.57
            },
            {
                'store_id': 3,
                'store_name': '加拿大三店',
                'manager_name': 'Bao Xiaoyun',
                'seating_capacity': 48,
                'annual_avg_turnover_2024': 5.72,
                'current_turnover_rate': 5.55,
                'prev_turnover_rate': 7.83,
                'current_revenue': 3.79,
                'prev_revenue': 3.79
            },
            {
                'store_id': 4,
                'store_name': '加拿大四店',
                'manager_name': '李俊娟',
                'seating_capacity': 70,
                'annual_avg_turnover_2024': 4.49,
                'current_turnover_rate': 4.61,
                'prev_turnover_rate': 5.14,
                'current_revenue': 4.54,
                'prev_revenue': 3.84
            },
            {
                'store_id': 1,
                'store_name': '加拿大一店',
                'manager_name': '张森磊',
                'seating_capacity': 53,
                'annual_avg_turnover_2024': 4.57,
                'current_turnover_rate': 4.83,
                'prev_turnover_rate': 5.28,
                'current_revenue': 4.26,
                'prev_revenue': 4.14
            },
            {
                'store_id': 7,
                'store_name': '加拿大七店',
                'manager_name': '潘幸远',
                'seating_capacity': 57,
                'annual_avg_turnover_2024': 3.32,
                'current_turnover_rate': 4.86,
                'prev_turnover_rate': 3.86,
                'current_revenue': 4.31,
                'prev_revenue': 3.64
            },
            {
                'store_id': 2,
                'store_name': '加拿大二店',
                'manager_name': '潘幸远',
                'seating_capacity': 36,
                'annual_avg_turnover_2024': 4.02,
                'current_turnover_rate': 2.71,
                'prev_turnover_rate': 4.06,
                'current_revenue': 1.59,
                'prev_revenue': 1.87
            }
        ]

    def _add_regional_summary(self, ws: Worksheet, store_data: List[Dict[str, Any]],
                              target_dt: datetime, prev_year_dt: datetime) -> None:
        """Add regional summary row (row 3)."""
        row = 3

        # Regional identifier
        ws.cell(row=row, column=4, value="区域")  # D3
        ws.cell(row=row, column=5, value="蒋冰遇")  # E3 - Regional manager

        # Calculate regional statistics with high precision
        if store_data:
            # Turnover rates: use WEIGHTED AVERAGE (total tables served / total seats)
            total_current_tables = sum(
                (s['current_turnover_rate'] or 0) * s['seating_capacity'] for s in store_data)
            total_prev_tables = sum(
                (s['prev_turnover_rate'] or 0) * s['seating_capacity'] for s in store_data)
            total_seats = sum(s['seating_capacity'] for s in store_data)

            # Calculate weighted averages (maintain 5 decimal precision)
            avg_current_turnover = total_current_tables / \
                total_seats if total_seats > 0 else 0
            avg_prev_turnover = total_prev_tables / total_seats if total_seats > 0 else 0

            # Revenue: use SUM for regional totals
            sum_current_revenue = sum((s['current_revenue'] or 0) for s in store_data)
            sum_prev_revenue = sum((s['prev_revenue'] or 0) for s in store_data)
        else:
            avg_current_turnover = avg_prev_turnover = 0
            sum_current_revenue = sum_prev_revenue = 0

        # Add turnover data (averages for regional performance)
        ws.cell(row=row, column=6, value=avg_current_turnover)  # F3
        ws.cell(row=row, column=7, value=avg_prev_turnover)     # G3
        # H3 - Comparison formula
        ws.cell(row=row, column=8, value="=F3-G3")

        # Add revenue data (sums for regional totals)
        ws.cell(row=row, column=10, value=sum_current_revenue)  # J3
        ws.cell(row=row, column=11, value=sum_prev_revenue)     # K3
        # L3 - Comparison formula
        ws.cell(row=row, column=12, value="=J3-K3")

    def _calculate_comprehensive_scores(self, store_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Calculate comprehensive scores for all stores and sort by ranking."""
        if not store_data:
            return store_data

        # Extract data for normalization (convert Decimal to float, handle None)
        current_turnovers = [float(s['current_turnover_rate'] or 0)
                             for s in store_data]
        prev_turnovers = [float(s['prev_turnover_rate'] or 0) for s in store_data]
        current_revenues = [float(s['current_revenue'] or 0) for s in store_data]
        prev_revenues = [float(s['prev_revenue'] or 0) for s in store_data]

        # Calculate comparison values
        turnover_diffs = [curr - prev for curr,
                          prev in zip(current_turnovers, prev_turnovers)]
        revenue_diffs = [curr - prev for curr,
                         prev in zip(current_revenues, prev_revenues)]

        # Calculate min/max for normalization
        min_turnover, max_turnover = min(
            current_turnovers), max(current_turnovers)
        min_turnover_diff, max_turnover_diff = min(
            turnover_diffs), max(turnover_diffs)
        min_revenue, max_revenue = min(current_revenues), max(current_revenues)
        min_revenue_diff, max_revenue_diff = min(
            revenue_diffs), max(revenue_diffs)

        # Calculate normalized scores for each store
        for i, store in enumerate(store_data):
            # Normalized current turnover score (I column equivalent)
            turnover_basic_score = 0
            if max_turnover != min_turnover:
                turnover_basic_score = (
                    current_turnovers[i] - min_turnover) / (max_turnover - min_turnover)

            # Normalized turnover improvement score (J column equivalent)
            turnover_improvement_score = 0
            if max_turnover_diff != min_turnover_diff:
                turnover_improvement_score = (
                    turnover_diffs[i] - min_turnover_diff) / (max_turnover_diff - min_turnover_diff)

            # Normalized current revenue score (N column equivalent)
            revenue_basic_score = 0
            if max_revenue != min_revenue:
                revenue_basic_score = (
                    current_revenues[i] - min_revenue) / (max_revenue - min_revenue)

            # Normalized revenue improvement score (O column equivalent)
            revenue_improvement_score = 0
            if max_revenue_diff != min_revenue_diff:
                revenue_improvement_score = (
                    revenue_diffs[i] - min_revenue_diff) / (max_revenue_diff - min_revenue_diff)

            # Comprehensive score (P column equivalent) - 25% weight each
            comprehensive_score = (turnover_basic_score * 0.25 +
                                   turnover_improvement_score * 0.25 +
                                   revenue_basic_score * 0.25 +
                                   revenue_improvement_score * 0.25)

            # Add calculated values to store data
            store['turnover_basic_score'] = turnover_basic_score
            store['turnover_improvement_score'] = turnover_improvement_score
            store['revenue_basic_score'] = revenue_basic_score
            store['revenue_improvement_score'] = revenue_improvement_score
            store['comprehensive_score'] = comprehensive_score
            store['turnover_diff'] = turnover_diffs[i]
            store['revenue_diff'] = revenue_diffs[i]

        # Sort by comprehensive score (descending - highest score = rank 1)
        sorted_stores = sorted(
            store_data, key=lambda x: x['comprehensive_score'], reverse=True)

        # Add ranking to sorted data
        for rank, store in enumerate(sorted_stores, 1):
            store['ranking'] = rank

        return sorted_stores

    def _add_store_data(self, ws: Worksheet, store_data: List[Dict[str, Any]],
                        target_dt: datetime, prev_year_dt: datetime) -> None:
        """Add individual store data rows with Excel formulas for scoring."""
        # Calculate comprehensive scores and sort by ranking for initial ordering
        sorted_store_data = self._calculate_comprehensive_scores(store_data)

        start_row = 4
        end_row = start_row + len(sorted_store_data) - 1

        for i, store in enumerate(sorted_store_data):
            row = start_row + i

            # Basic store information
            # A - Serial number (now matches ranking)
            ws.cell(row=row, column=1, value=store['ranking'])
            # B - Store name
            ws.cell(row=row, column=2, value=store['store_name'])
            # C - Manager name
            ws.cell(row=row, column=3, value=store['manager_name'])
            # D - Seating capacity
            ws.cell(row=row, column=4, value=store['seating_capacity'])
            # E - 2024 avg turnover
            ws.cell(row=row, column=5, value=store['annual_avg_turnover_2024'])

            # Table turnover rate data
            # F - Current turnover
            ws.cell(row=row, column=6, value=store['current_turnover_rate'])
            # G - Previous turnover
            ws.cell(row=row, column=7, value=store['prev_turnover_rate'])
            # H - Comparison (Excel formula)
            ws.cell(row=row, column=8, value=f"=F{row}-G{row}")

            # Normalized scoring - EXACT formulas from reference sheet
            # I - Improvement score (Turnover difference normalization)
            ws.cell(row=row, column=9,
                    value=f"=(H{row}-MIN($H$4:$H$10))/(MAX($H$4:$H$10)-MIN($H$4:$H$10))")

            # Revenue data
            # J - Current revenue
            ws.cell(row=row, column=10, value=store['current_revenue'])
            # K - Previous revenue
            ws.cell(row=row, column=11, value=store['prev_revenue'])
            # L - Revenue comparison (Excel formula)
            ws.cell(row=row, column=12, value=f"=J{row}-K{row}")

            # Normalized scoring - EXACT formulas from reference sheet
            # M - Revenue improvement score (Revenue difference normalization)
            ws.cell(row=row, column=13,
                    value=f"=(L{row}-MIN($L$4:$L$10))/(MAX($L$4:$L$10)-MIN($L$4:$L$10))")

            # Comprehensive scoring - Modified formula without basic scores
            # N - Comprehensive score (50% weight each improvement score)
            ws.cell(row=row, column=14,
                    value=f"=I{row}*0.5+M{row}*0.5")

            # Ranking - EXACT formula from reference sheet
            # O - Ranking (RANK function)
            ws.cell(row=row, column=15, value=f"=RANK(N{row},$N$4:$N$10)")

    def _apply_formatting(self, ws: Worksheet, num_stores: int) -> None:
        """Apply professional formatting to the worksheet."""
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="DC2626", end_color="DC2626", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Set column widths
        column_widths = {
            'A': 6,   # Serial number
            'B': 15,  # Store name
            'C': 10,  # Manager name
            'D': 8,   # Seating capacity
            'E': 12,  # Annual avg turnover
            'F': 12,  # Current turnover
            'G': 12,  # Previous turnover
            'H': 8,   # Comparison
            'I': 10,  # Improvement score
            'J': 15,  # Current revenue
            'K': 15,  # Previous revenue
            'L': 8,   # Revenue comparison
            'M': 10,  # Revenue improvement score
            'N': 10,  # Comprehensive score
            'O': 8    # Ranking
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Format headers (rows 1-2)
        for row in [1, 2]:
            for col in range(1, 16):  # A to O
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

        # Format data rows
        data_rows = range(3, 4 + num_stores)  # Regional summary + store data
        for row in data_rows:
            for col in range(1, 16):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = center_alignment

                # Number formatting for specific columns (maintain 5 decimal precision)
                if col in [5, 6, 7]:  # Turnover rates - 5 decimal places
                    cell.number_format = '0.00000'
                elif col in [10, 11]:  # Revenue - 5 decimal places
                    cell.number_format = '0.00000'
                elif col in [8, 12]:  # Comparisons - 5 decimal places
                    cell.number_format = '0.00000'
                elif col in [9, 13, 14]:  # Scores - 5 decimal places
                    cell.number_format = '0.00000'

        # Freeze panes at row 3 (after headers)
        ws.freeze_panes = "A3"

        self.logger.info(
            "Formatting applied to daily store tracking worksheet")
