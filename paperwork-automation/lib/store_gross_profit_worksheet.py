import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from lib.database_queries import ReportDataProvider


class StoreGrossProfitWorksheetGenerator:
    """Generator for all store gross profit comparison worksheet."""

    def __init__(self, data_provider: ReportDataProvider):
        self.data_provider = data_provider
        self.logger = logging.getLogger(__name__)

    def generate_worksheet(self, workbook: Workbook, target_date: str) -> None:
        """
        Generate all store gross profit worksheet.

        Args:
            workbook: Excel workbook object
            target_date: Target date in YYYY-MM-DD format
        """
        try:
            self.logger.info(
                f"🔄 Generating store gross profit worksheet for {target_date}")

            # Get store gross profit data
            store_gross_profit_data = self.data_provider.get_store_gross_profit_data(
                target_date)

            if not store_gross_profit_data:
                self.logger.warning("⚠️ No store gross profit data found")
                return

            # Create the worksheet
            ws = workbook.create_sheet(title="各店毛利率分析")

            # Generate the content
            self._generate_store_gross_profit_analysis(
                ws, store_gross_profit_data, target_date)

            # Apply formatting
            self._apply_formatting(ws)

            self.logger.info(
                f"✅ Store gross profit worksheet generated successfully with {len(store_gross_profit_data)} records")

        except Exception as e:
            self.logger.error(
                f"❌ Error generating store gross profit worksheet: {e}")
            raise

    def _generate_store_gross_profit_analysis(self, ws, data: List[Dict[str, Any]], target_date: str) -> None:
        """Generate store gross profit analysis content."""
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')

        # Title
        ws.merge_cells('A1:J1')
        ws['A1'] = f"各店毛利率分析表 - {target_dt.strftime('%Y年%m月')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = [
            "门店名称", "本月销售额", "本月成本", "本月毛利", "本月毛利率",
            "上月销售额", "上月成本", "上月毛利", "上月毛利率", "环比变动"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col).value = header
            ws.cell(row=3, column=col).font = Font(bold=True)
            ws.cell(row=3, column=col).alignment = Alignment(
                horizontal='center', vertical='center')

        # Data rows
        current_row = 4

        # Group data by store
        store_data = {}
        for record in data:
            store_name = record.get('store_name', '')
            if store_name not in store_data:
                store_data[store_name] = {
                    'current_revenue': 0,
                    'current_cost': 0,
                    'previous_revenue': 0,
                    'previous_cost': 0
                }

            store_data[store_name]['current_revenue'] += self._safe_float(
                record.get('current_revenue', 0))
            store_data[store_name]['current_cost'] += self._safe_float(
                record.get('current_cost', 0))
            store_data[store_name]['previous_revenue'] += self._safe_float(
                record.get('previous_revenue', 0))
            store_data[store_name]['previous_cost'] += self._safe_float(
                record.get('previous_cost', 0))

        # Calculate totals
        total_current_revenue = 0
        total_current_cost = 0
        total_previous_revenue = 0
        total_previous_cost = 0

        for store_name in sorted(store_data.keys()):
            store_info = store_data[store_name]

            # Current month calculations
            current_revenue = store_info['current_revenue']
            current_cost = store_info['current_cost']
            current_profit = current_revenue - current_cost
            current_margin = (current_profit / current_revenue *
                              100) if current_revenue > 0 else 0

            # Previous month calculations
            previous_revenue = store_info['previous_revenue']
            previous_cost = store_info['previous_cost']
            previous_profit = previous_revenue - previous_cost
            previous_margin = (
                previous_profit / previous_revenue * 100) if previous_revenue > 0 else 0

            # Month-over-month change
            margin_change = current_margin - previous_margin

            # Fill row data
            ws.cell(row=current_row, column=1).value = store_name
            ws.cell(row=current_row, column=2).value = current_revenue
            ws.cell(row=current_row, column=3).value = current_cost
            ws.cell(row=current_row, column=4).value = current_profit
            ws.cell(row=current_row, column=5).value = current_margin / \
                100  # Convert to percentage
            ws.cell(row=current_row, column=6).value = previous_revenue
            ws.cell(row=current_row, column=7).value = previous_cost
            ws.cell(row=current_row, column=8).value = previous_profit
            ws.cell(row=current_row, column=9).value = previous_margin / \
                100  # Convert to percentage
            ws.cell(row=current_row, column=10).value = margin_change / \
                100  # Convert to percentage

            # Add to totals
            total_current_revenue += current_revenue
            total_current_cost += current_cost
            total_previous_revenue += previous_revenue
            total_previous_cost += previous_cost

            current_row += 1

        # Add total row
        if store_data:
            total_current_profit = total_current_revenue - total_current_cost
            total_current_margin = (
                total_current_profit / total_current_revenue * 100) if total_current_revenue > 0 else 0

            total_previous_profit = total_previous_revenue - total_previous_cost
            total_previous_margin = (
                total_previous_profit / total_previous_revenue * 100) if total_previous_revenue > 0 else 0

            total_margin_change = total_current_margin - total_previous_margin

            ws.cell(row=current_row, column=1).value = "总计"
            ws.cell(row=current_row, column=2).value = total_current_revenue
            ws.cell(row=current_row, column=3).value = total_current_cost
            ws.cell(row=current_row, column=4).value = total_current_profit
            ws.cell(row=current_row, column=5).value = total_current_margin / 100
            ws.cell(row=current_row, column=6).value = total_previous_revenue
            ws.cell(row=current_row, column=7).value = total_previous_cost
            ws.cell(row=current_row, column=8).value = total_previous_profit
            ws.cell(row=current_row, column=9).value = total_previous_margin / 100
            ws.cell(row=current_row, column=10).value = total_margin_change / 100

            # Bold the total row
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).font = Font(bold=True)

    def _apply_formatting(self, ws) -> None:
        """Apply formatting to the worksheet."""
        # Column widths
        column_widths = {
            1: 15,  # 门店名称
            2: 15,  # 本月销售额
            3: 15,  # 本月成本
            4: 15,  # 本月毛利
            5: 12,  # 本月毛利率
            6: 15,  # 上月销售额
            7: 15,  # 上月成本
            8: 15,  # 上月毛利
            9: 12,  # 上月毛利率
            10: 12  # 环比变动
        }

        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Number formats
        # Currency format for revenue, cost, profit columns
        currency_format = '#,##0.00'
        percentage_format = '0.00%'

        # Apply number formatting to all data rows
        for row in range(4, ws.max_row + 1):
            # Currency columns (2,3,4,6,7,8)
            for col in [2, 3, 4, 6, 7, 8]:
                ws.cell(row=row, column=col).number_format = currency_format

            # Percentage columns (5,9,10)
            for col in [5, 9, 10]:
                ws.cell(row=row, column=col).number_format = percentage_format

        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply borders to all cells with data
        for row in range(1, ws.max_row + 1):
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = thin_border

        # Header styling
        header_fill = PatternFill(
            start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        for col in range(1, 11):
            ws.cell(row=3, column=col).fill = header_fill

        # Title styling
        title_fill = PatternFill(start_color="B0C4DE",
                                 end_color="B0C4DE", fill_type="solid")
        ws['A1'].fill = title_fill

    def _safe_float(self, value: Any) -> float:
        """Safely convert value to float."""
        if value is None:
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
