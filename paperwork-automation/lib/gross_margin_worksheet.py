#!/usr/bin/env python3
"""
Gross Margin Report worksheet generator (毛利报表).
Generates detailed revenue data and other gross margin analysis worksheets.
"""

from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging


class GrossMarginWorksheetGenerator:
    """Generate gross margin analysis worksheets from provided data"""

    def __init__(self, target_date: str):
        self.target_date = target_date
        self.logger = logging.getLogger(__name__)

    def generate_detailed_revenue_worksheet(self, workbook, dish_price_data):
        """Generate detailed revenue data worksheet (菜品价格变动及菜品损耗表)"""
        ws = workbook.create_sheet("菜品价格变动及菜品损耗表")

        # Parse target date for display
        target_dt = datetime.strptime(self.target_date, '%Y-%m-%d')

        # Set column widths based on the Excel template structure
        column_widths = [
            8,   # 区域
            12,  # 门店名称
            20,  # 唯一编码
            20,  # 套餐唯一编码
            12,  # 菜品编码
            12,  # 菜品短编码
            20,  # 菜品名称
            10,  # 菜品份量KG
            8,   # 菜品单位
            8,   # 规格
            10,  # 本期销量
            10,  # 理论耗用量
            10,  # 本期单价
            10,  # 上期单价
            10,  # 去年同期单价
            12,  # 环比价格变动
            12,  # 同比价格变动
            15,  # 本期收入
            20,  # 对应物料名称
            10,  # 本期耗用量
            10,  # 套餐、拼盘用量
            10,  # 套餐销量
            12,  # 环比影响收入
            12,  # 同比影响收入
            12,  # 调价日期
            10,  # 实际毛利率
            10,  # 理论耗用量
            10,  # 实际耗用量
            12,  # 本月损耗影响成本金额
            12,  # 上月损耗影响成本金额
            12,  # 损耗环比变动金额
            12,  # 可比期间损耗影响成本金额
            12,  # 损耗同比变动金额
        ]

        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        current_row = 1

        # Add title section
        current_row = self._add_title_section(ws, current_row, target_dt)

        # Add main headers
        current_row = self._add_main_headers(ws, current_row)

        # Add sub headers
        current_row = self._add_sub_headers(ws, current_row)

        # Add data rows
        current_row = self._add_data_rows(ws, current_row, dish_price_data)

        # Apply formatting
        self._apply_formatting(ws, current_row - 1)

        return ws

    def _add_title_section(self, ws, start_row, target_dt):
        """Add title section with date and main categories"""
        current_row = start_row

        # Main title row
        ws.merge_cells(f'A{current_row}:AC{current_row}')
        ws[f'A{current_row}'] = f"菜品销售报表（本币） - {target_dt.strftime('%Y年%m月')}"
        ws[f'A{current_row}'].font = Font(bold=True, size=14)
        ws[f'A{current_row}'].alignment = Alignment(
            horizontal='center', vertical='center')
        ws[f'A{current_row}'].fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        current_row += 1

        # Category headers row
        category_headers = [
            ("A", "N", "菜品销售报表（本币）"),
            ("P", "R", "原材料耗用（本币）"),
            ("S", "T", "菜品价格调整影响"),
            ("U", "V", "理论与实际耗用量"),
            ("W", "AC", "菜品损耗影响金额（环比及同比）")
        ]

        for start_col, end_col, title in category_headers:
            ws.merge_cells(f'{start_col}{current_row}:{end_col}{current_row}')
            ws[f'{start_col}{current_row}'] = title
            ws[f'{start_col}{current_row}'].font = Font(bold=True, size=11)
            ws[f'{start_col}{current_row}'].alignment = Alignment(
                horizontal='center', vertical='center')
            ws[f'{start_col}{current_row}'].fill = PatternFill(
                start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        current_row += 1
        return current_row

    def _add_main_headers(self, ws, start_row):
        """Add main column headers"""
        current_row = start_row

        # Main headers (based on Excel template structure) - removed unnecessary columns, added revenue
        headers = [
            "区域", "门店名称", "菜品编码", "菜品短编码", "菜品名称",
            "规格", "本期销量", "理论耗用量", "本期单价", "上期单价",
            "去年同期单价", "环比价格变动", "同比价格变动", "本期收入", "对应物料名称（对应zfi0156每个门店V）",
            "本期耗用量(对应zfi0156每个门店V）", "套餐、拼盘用量",
            "套餐销量", "环比影响收入", "同比影响收入", "调价日期", "实际毛利率",
            "理论耗用量", "实际耗用量", "本月损耗影响成本金额", "上月损耗影响成本金额",
            "损耗环比变动金额", "可比期间损耗影响成本金额", "损耗同比变动金额"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
            cell.fill = PatternFill(
                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        current_row += 1
        return current_row

    def _add_sub_headers(self, ws, start_row):
        """Add sub-headers if needed"""
        # This method can be used for additional sub-categorization if needed
        return start_row

    def _safe_float(self, value):
        """Safely convert value to float, handling None and invalid strings"""
        if value is None:
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0

    def _add_data_rows(self, ws, start_row, dish_price_data):
        """Add data rows with dish price and loss information"""
        current_row = start_row

        if not dish_price_data:
            # Add placeholder row if no data
            ws.cell(row=current_row, column=1, value="暂无数据")
            ws.cell(row=current_row, column=1).font = Font(italic=True)
            ws.cell(row=current_row, column=1).alignment = Alignment(
                horizontal='center')
            current_row += 1
            return current_row

        # Group data by store for better organization
        data_by_store = {}
        for row in dish_price_data:
            store_id = row.get('store_id')
            if store_id not in data_by_store:
                data_by_store[store_id] = []
            data_by_store[store_id].append(row)

        # Add data rows for each store
        for store_id, store_data in data_by_store.items():
            for data_row in store_data:
                # Extract data values
                region = data_row.get('region', '加拿大')
                store_name = data_row.get('store_name', '')
                dish_code = data_row.get('dish_code', '')
                dish_short_code = data_row.get('dish_short_code', '')
                dish_name = data_row.get('dish_name', '')
                specification = data_row.get('specification', '')
                current_sales = self._safe_float(
                    data_row.get('current_period_sales', 0))
                theoretical_usage = self._safe_float(
                    data_row.get('theoretical_usage', 0))
                current_price = self._safe_float(
                    data_row.get('current_price', 0))
                previous_price = self._safe_float(
                    data_row.get('previous_price', 0))
                last_year_price = self._safe_float(
                    data_row.get('last_year_price', 0))

                # Calculate price changes
                price_change_mom = ((current_price - previous_price) /
                                    previous_price * 100) if previous_price > 0 else 0
                price_change_yoy = ((current_price - last_year_price) /
                                    last_year_price * 100) if last_year_price > 0 else 0

                # Calculate revenue (sales amount * price)
                current_revenue = current_sales * current_price

                # Material information - using aggregated data
                # All materials formatted as "material_number - name - quantity"
                material_list_with_usage = data_row.get(
                    'material_list_with_usage', '')
                total_material_cost = self._safe_float(
                    data_row.get('total_material_cost', 0))  # Total spend for all materials
                combo_usage = self._safe_float(data_row.get('combo_usage', 0))
                combo_sales = self._safe_float(data_row.get('combo_sales', 0))

                # Revenue impact
                revenue_impact_mom = price_change_mom * \
                    current_sales / 100 if current_sales > 0 else 0
                revenue_impact_yoy = price_change_yoy * \
                    current_sales / 100 if current_sales > 0 else 0

                # Additional fields
                price_adjustment_date = data_row.get(
                    'price_adjustment_date', '')

                # Calculate actual gross margin: (revenue - cost) / revenue * 100%
                if current_revenue > 0:
                    actual_gross_margin = (
                        (current_revenue - total_material_cost) / current_revenue) * 100
                else:
                    actual_gross_margin = 0

                # Loss analysis
                theoretical_usage_loss = self._safe_float(
                    data_row.get('theoretical_usage_loss', 0))
                actual_usage_loss = self._safe_float(
                    data_row.get('actual_usage_loss', 0))
                current_month_loss_cost = self._safe_float(
                    data_row.get('current_month_loss_cost', 0))
                previous_month_loss_cost = self._safe_float(
                    data_row.get('previous_month_loss_cost', 0))
                loss_change_mom = current_month_loss_cost - previous_month_loss_cost
                comparable_period_loss_cost = self._safe_float(
                    data_row.get('comparable_period_loss_cost', 0))
                loss_change_yoy = current_month_loss_cost - comparable_period_loss_cost

                # Create row data (removed unnecessary columns: click_rate, dish_category)
                row_data = [
                    region,
                    store_name,
                    dish_code,
                    dish_short_code,
                    dish_name,
                    specification,
                    current_sales,
                    theoretical_usage,
                    current_price,
                    previous_price,
                    last_year_price,
                    f"{price_change_mom:+.2f}%" if price_change_mom != 0 else "0%",
                    f"{price_change_yoy:+.2f}%" if price_change_yoy != 0 else "0%",
                    current_revenue,           # New revenue column
                    # All materials with usage amounts (now includes material number)
                    material_list_with_usage,
                    total_material_cost,       # Total cost for all materials
                    combo_usage,
                    combo_sales,
                    revenue_impact_mom,
                    revenue_impact_yoy,
                    price_adjustment_date,
                    f"{actual_gross_margin:.2f}%" if actual_gross_margin != 0 else "0%",
                    theoretical_usage_loss,
                    actual_usage_loss,
                    current_month_loss_cost,
                    previous_month_loss_cost,
                    loss_change_mom,
                    comparable_period_loss_cost,
                    loss_change_yoy
                ]

                # Add row to worksheet
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)

                    # Apply text wrapping for material list column (column 15, shifted due to new revenue column)
                    if col == 15:  # material_list_with_usage column
                        cell.alignment = Alignment(
                            wrap_text=True, vertical='top')

                    # Apply number formatting for numeric columns (updated column numbers with new revenue column)
                    if col in [7, 8, 9, 10, 11, 14, 16, 17, 18, 19, 20, 21, 22, 25, 26, 27, 28, 29, 30, 31]:
                        if isinstance(value, (int, float)) and value != 0:
                            cell.number_format = '#,##0.00'

                current_row += 1

        return current_row

    def _apply_formatting(self, ws, max_row):
        """Apply common formatting to the worksheet"""
        # Add borders to all cells with data
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in range(1, max_row + 1):
            for col in range(1, 32):  # 31 columns + 1 (added revenue column)
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border

                # Center align headers
                if row <= 3:
                    cell.alignment = Alignment(
                        horizontal='center', vertical='center')

        # Set column widths (updated for new revenue column)
        column_widths = {
            1: 8,   # 区域
            2: 12,  # 门店名称
            3: 12,  # 菜品编码
            4: 10,  # 菜品短编码
            5: 20,  # 菜品名称
            6: 10,  # 规格
            7: 10,  # 本期销量
            8: 10,  # 理论耗用量
            9: 10,  # 本期单价
            10: 10,  # 上期单价
            11: 12,  # 去年同期单价
            12: 12,  # 环比价格变动
            13: 12,  # 同比价格变动
            14: 15,  # 本期收入
            15: 30,  # 对应物料名称（多行显示，包含材料编号）
            16: 15,  # 本期耗用量（总材料成本）
            17: 12,  # 套餐、拼盘用量
            18: 10,  # 套餐销量
            19: 12,  # 环比影响收入
            20: 12,  # 同比影响收入
            21: 12,  # 调价日期
            22: 12,  # 实际毛利率
            23: 12,  # 理论耗用量
            24: 12,  # 实际耗用量
            25: 15,  # 本月损耗影响成本金额
            26: 15,  # 上月损耗影响成本金额
            27: 15,  # 损耗环比变动金额
            28: 18,  # 可比期间损耗影响成本金额
            29: 15   # 损耗同比变动金额
        }

        from openpyxl.utils import get_column_letter
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze panes at row 4 (after headers)
        ws.freeze_panes = 'A4'

        # Set row height for header rows
        for row in range(1, 4):
            ws.row_dimensions[row].height = 20

        # Set row height for data rows (larger to accommodate multi-line material lists)
        for row in range(4, max_row + 1):
            # Increased height for multi-line content
            ws.row_dimensions[row].height = 50

    def generate_material_cost_worksheet(self, workbook, material_cost_data):
        """Generate material cost change worksheet (原材料成本变动表)"""
        ws = workbook.create_sheet("原材料成本变动表")

        # This will be implemented in the next iteration
        ws.cell(row=1, column=1, value="原材料成本变动表 - 开发中")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        return ws

    def generate_discount_analysis_worksheet(self, workbook, discount_data):
        """Generate discount analysis worksheet (打折优惠表)"""
        ws = workbook.create_sheet("打折优惠表")

        # This will be implemented in the next iteration
        ws.cell(row=1, column=1, value="打折优惠表 - 开发中")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        return ws
