#!/usr/bin/env python3
"""
Beverage Variance Worksheet Generator
Generates detailed beverage variance analysis matching the structure of material usage variance analysis
with additional pricing information for financial impact analysis.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
from pathlib import Path

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent))


class BeverageVarianceGenerator:
    """Generator for beverage variance detail worksheet"""

    def __init__(self, data_provider):
        self.data_provider = data_provider
        self.db_manager = data_provider.db_manager

    def generate_worksheet(self, workbook, target_date: str):
        """Generate beverage variance detail worksheet matching material variance analysis structure"""
        worksheet = workbook.create_sheet("酒水差异明细表")

        # Set up styles
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid")
        data_font = Font(size=10)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # Parse target date
        target_dt = datetime.strptime(target_date, '%Y-%m-%d')
        year, month = target_dt.year, target_dt.month

        # Create title section
        current_row = 1
        worksheet.merge_cells(f'A{current_row}:P{current_row}')
        worksheet[f'A{current_row}'] = f"酒水差异明细表 - {target_dt.strftime('%Y年%m月')}"
        worksheet[f'A{current_row}'].font = Font(
            bold=True, size=16, color="FFFFFF")
        worksheet[f'A{current_row}'].alignment = Alignment(
            horizontal='center', vertical='center')
        worksheet[f'A{current_row}'].fill = PatternFill(
            start_color="C41E3A", end_color="C41E3A", fill_type="solid")
        current_row += 2

        # Get detailed variance data
        variance_data = self.get_beverage_variance_data(year, month)

        # Add summary section
        current_row = self.add_variance_summary_section(
            worksheet, current_row, variance_data)

        # Add headers matching material variance structure plus pricing columns
        headers = [
            "序号",          # A: Serial
            "门店",          # B: Store
            "物料名称",      # C: Material Name
            "物料号",        # D: Material Number
            "单位",          # E: Unit
            "包装规格",      # F: Package Spec
            "理论用量",      # G: Theoretical Usage
            "套餐用量",      # H: Combo Usage
            "系统记录",      # I: System Record
            "库存盘点",      # J: Inventory Count
            "差异数量",      # K: Variance Amount
            "差异率(%)",     # L: Variance Rate
            "状态",          # M: Status
            "物料单价",      # N: Material Price (NEW)
            "销售单价",      # O: Sale Price (NEW)
            "差异金额"       # P: Net Difference Amount (NEW)
        ]

        # Set headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        current_row += 1
        data_start_row = current_row

        # Populate data rows
        for row_number, data in enumerate(variance_data, 1):
            # Basic data (columns A-J) - matching material variance structure
            basic_data = [
                row_number,                      # A: 序号
                data['store_name'],              # B: 门店
                data['material_name'],           # C: 物料名称
                data['material_number'] or '',   # D: 物料号
                data['material_unit'] or '',     # E: 单位
                data['package_spec'] or '',      # F: 包装规格
                data['theoretical_usage'],       # G: 理论用量
                data['combo_usage'],             # H: 套餐用量
                data['system_record'],           # I: 系统记录
                data['inventory_count']          # J: 库存盘点
            ]

            # Insert basic data (A-J)
            for col, value in enumerate(basic_data, 1):
                cell = worksheet.cell(row=current_row, column=col, value=value)

                # Format numeric columns
                if col in [7, 8, 9, 10]:  # Usage amounts, combo usage, system record, inventory count
                    cell.number_format = '#,##0.0000'

                # Center align certain columns
                if col in [1, 2, 5]:  # Serial, store, unit
                    cell.alignment = Alignment(horizontal='center')

                cell.border = border

            # K: 差异数量 (Variance Amount) - Excel formula: = I - (G+H+J)
            variance_formula = f"=I{current_row}-(G{current_row}+H{current_row}+J{current_row})"
            variance_cell = worksheet.cell(
                row=current_row, column=11, value=variance_formula)
            variance_cell.number_format = '#,##0.0000'
            variance_cell.border = border

            # L: 差异率 (Variance Rate) - Excel formula: = ABS(K/(G+H))*100 with error handling
            rate_formula = f"=IF((G{current_row}+H{current_row})=0,IF(K{current_row}=0,0,100),ABS(K{current_row}/(G{current_row}+H{current_row}))*100)"
            rate_cell = worksheet.cell(
                row=current_row, column=12, value=rate_formula)
            rate_cell.number_format = '0.00"%"'
            rate_cell.alignment = Alignment(horizontal='center')
            rate_cell.border = border

            # M: 状态 (Status)
            status_cell = worksheet.cell(
                row=current_row, column=13, value=data['variance_status'])
            status_cell.alignment = Alignment(horizontal='center')
            status_cell.border = border

            # N: 物料单价 (Material Price) - NEW
            material_price_cell = worksheet.cell(
                row=current_row, column=14, value=data['material_price'])
            material_price_cell.number_format = '#,##0.0000'
            material_price_cell.alignment = Alignment(horizontal='right')
            material_price_cell.border = border

            # O: 销售单价 (Sale Price) - NEW
            sale_price_cell = worksheet.cell(
                row=current_row, column=15, value=data['sale_price'])
            sale_price_cell.number_format = '#,##0.0000'
            sale_price_cell.alignment = Alignment(horizontal='right')
            sale_price_cell.border = border

            # P: 差异金额 (Net Difference Amount) - NEW: Excel formula: = K * N
            net_diff_formula = f"=K{current_row}*N{current_row}"
            net_diff_cell = worksheet.cell(
                row=current_row, column=16, value=net_diff_formula)
            net_diff_cell.number_format = '#,##0.00'
            net_diff_cell.alignment = Alignment(horizontal='right')
            net_diff_cell.border = border

            # Apply row-level formatting based on variance percentage
            if abs(data['variance_percent']) > 5:
                fill_color = "FFE6E6" if data['variance_percent'] > 5 else "FFF0E6"
                row_fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid")

                # Apply fill to all cells in this row (now 16 columns)
                for col in range(1, 17):
                    worksheet.cell(row=current_row, column=col).fill = row_fill

            current_row += 1

        # Set column widths for better readability
        column_widths = [
            6,   # A: 序号
            12,  # B: 门店
            25,  # C: 物料名称
            12,  # D: 物料号
            8,   # E: 单位
            15,  # F: 包装规格
            12,  # G: 理论用量
            12,  # H: 套餐用量
            12,  # I: 系统记录
            12,  # J: 库存盘点
            12,  # K: 差异数量
            10,  # L: 差异率
            8,   # M: 状态
            12,  # N: 物料单价
            12,  # O: 销售单价
            12   # P: 差异金额
        ]

        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

        # Freeze panes at data header
        if data_start_row > 1:
            worksheet.freeze_panes = f'A{data_start_row}'

        return worksheet

    def add_variance_summary_section(self, ws, start_row, variance_data):
        """Add variance summary statistics section"""
        current_row = start_row

        # Calculate summary statistics
        total_materials = len(variance_data)
        materials_over_threshold = len(
            [row for row in variance_data if abs(row['variance_percent']) > 5])
        materials_normal = total_materials - materials_over_threshold

        over_usage = len(
            [row for row in variance_data if row['variance_percent'] > 5])
        under_usage = len(
            [row for row in variance_data if row['variance_percent'] < -5])

        total_theoretical = sum(row['theoretical_usage']
                                for row in variance_data)
        total_combo_usage = sum(row['combo_usage'] for row in variance_data)
        total_combined_theoretical = total_theoretical + total_combo_usage
        total_system = sum(row['system_record'] for row in variance_data)
        total_inventory = sum(row['inventory_count'] for row in variance_data)
        total_variance = total_system - total_combined_theoretical
        overall_variance_percent = (
            total_variance / total_combined_theoretical * 100) if total_combined_theoretical > 0 else 0

        # Financial impact calculations
        total_material_cost = sum(
            row['material_price'] * row['variance_amount'] for row in variance_data)
        total_sale_impact = sum(
            row['sale_price'] * row['variance_amount'] for row in variance_data)

        # Create summary section
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'] = "酒水差异分析概览"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        ws[f'A{current_row}'].fill = PatternFill(
            start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        current_row += 1

        summary_data = [
            ("酒水物料总数", total_materials),
            ("差异超5%物料", materials_over_threshold),
            ("正常范围物料", materials_normal),
            ("超量使用物料", over_usage),
            ("使用不足物料", under_usage),
            ("总理论用量", f"{total_theoretical:.2f}"),
            ("总套餐用量", f"{total_combo_usage:.2f}"),
            ("总系统记录", f"{total_system:.2f}"),
            ("总库存盘点", f"{total_inventory:.2f}"),
            ("总差异", f"{total_variance:.2f}"),
            ("总差异率", f"{overall_variance_percent:.1f}%"),
            ("材料成本影响", f"{total_material_cost:.2f} CAD"),
            ("销售影响", f"{total_sale_impact:.2f} CAD")
        ]

        for i, (label, value) in enumerate(summary_data):
            row = current_row + (i // 2)
            col = 1 + (i % 2) * 3

            cell_label = ws.cell(row=row, column=col, value=label)
            cell_value = ws.cell(row=row, column=col+1, value=value)
            cell_label.font = Font(bold=True)

            # Color code summary metrics
            if i >= 5:  # Usage metrics
                fill_color = "E8F5E8"
                # Overall variance and financial impact
                if i == 10 and abs(overall_variance_percent) > 5:
                    fill_color = "FFE6E6" if overall_variance_percent > 0 else "FFF0E6"
                elif i >= 11:  # Financial impact
                    fill_color = "FFF8E1"

                cell_label.fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell_value.fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid")

        current_row += 4
        return current_row + 1

    def get_beverage_variance_data(self, year: int, month: int):
        """Get detailed beverage variance data matching material variance analysis structure"""
        try:
            with self.db_manager.get_connection() as conn:
                cursor = conn.cursor()

                # Working query from the direct test
                query = """
                WITH beverage_materials AS (
                    SELECT DISTINCT m.id as material_id
                    FROM material m
                    JOIN material_type mt ON m.material_type_id = mt.id
                    WHERE m.is_active = TRUE
                    AND mt.name = '成本-酒水类'
                ),
                
                -- Aggregate dish sales across all sales modes first
                aggregated_dish_sales AS (
                    SELECT 
                        dish_id,
                        store_id,
                        SUM(COALESCE(sale_amount, 0)) as total_sale_amount,
                        SUM(COALESCE(return_amount, 0)) as total_return_amount
                    FROM dish_monthly_sale
                    WHERE year = %s AND month = %s
                    GROUP BY dish_id, store_id
                ),
                -- Get theoretical usage from dish sales (only for materials with dish relationships)
                theoretical_usage AS (
                    SELECT 
                        m.id as material_id,
                        ads.store_id,
                        SUM((COALESCE(ads.total_sale_amount, 0) - COALESCE(ads.total_return_amount, 0)) * 
                            COALESCE(dm.standard_quantity, 0) * COALESCE(dm.loss_rate, 1.0)) as theoretical_total
                    FROM material m
                    JOIN beverage_materials bm ON m.id = bm.material_id
                    JOIN dish_material dm ON m.id = dm.material_id
                    JOIN aggregated_dish_sales ads ON dm.dish_id = ads.dish_id
                    GROUP BY m.id, ads.store_id
                ),
                
                -- Get combo usage from combo dish sales (only for materials with dish relationships)
                combo_usage AS (
                    SELECT 
                        m.id as material_id,
                        mcds.store_id,
                        SUM(COALESCE(mcds.sale_amount, 0) * 
                            COALESCE(dm.standard_quantity, 0) * COALESCE(dm.loss_rate, 1.0)) as combo_total
                    FROM material m
                    JOIN beverage_materials bm ON m.id = bm.material_id
                    JOIN dish_material dm ON m.id = dm.material_id
                    JOIN monthly_combo_dish_sale mcds ON dm.dish_id = mcds.dish_id 
                        AND mcds.year = %s AND mcds.month = %s
                    GROUP BY m.id, mcds.store_id
                ),
                
                -- Get all store-material combinations with any data
                base_data AS (
                    SELECT 
                        m.id as material_id,
                        m.name as material_name,
                        m.material_number,
                        m.unit as material_unit,
                        m.package_spec,
                        s.id as store_id,
                        s.name as store_name,
                        -- System record from material usage
                        COALESCE(mmu.material_used, 0) as system_record,
                        -- Inventory count
                        COALESCE(ic.counted_quantity, 0) as inventory_count,
                        -- Material price
                        COALESCE(mph.price, 10.0) as material_price
                    FROM material m
                    JOIN beverage_materials bm ON m.id = bm.material_id
                    CROSS JOIN store s
                    LEFT JOIN material_monthly_usage mmu ON m.id = mmu.material_id AND s.id = mmu.store_id
                        AND mmu.year = %s AND mmu.month = %s
                    LEFT JOIN inventory_count ic ON m.id = ic.material_id AND s.id = ic.store_id
                        AND EXTRACT(year FROM ic.count_date) = %s AND EXTRACT(month FROM ic.count_date) = %s
                    LEFT JOIN material_price_history mph ON m.id = mph.material_id AND s.id = mph.store_id
                        AND mph.is_active = TRUE
                    WHERE s.is_active = TRUE
                    -- Only include records where we have some data
                    AND (mmu.material_used IS NOT NULL OR ic.counted_quantity IS NOT NULL)
                )
                
                -- Final result combining all data
                SELECT 
                    bd.material_id,
                    bd.store_id,
                    bd.store_name,
                    bd.material_name,
                    bd.material_number,
                    bd.material_unit,
                    bd.package_spec,
                    COALESCE(tu.theoretical_total, 0) as theoretical_usage,
                    COALESCE(cu.combo_total, 0) as combo_usage,
                    bd.system_record,
                    bd.inventory_count,
                    bd.material_price,
                    bd.material_price * 1.5 as sale_price
                FROM base_data bd
                LEFT JOIN theoretical_usage tu ON bd.material_id = tu.material_id AND bd.store_id = tu.store_id
                LEFT JOIN combo_usage cu ON bd.material_id = cu.material_id AND bd.store_id = cu.store_id
                ORDER BY bd.store_name, bd.material_name
                """

                cursor.execute(query, (year, month, year, month, year,
                               month, year, month, year, month))
                results = cursor.fetchall()

                # Convert to list of dictionaries with variance calculations
                variance_data = []
                for row in results:
                    # Access dictionary-style (RealDictRow objects)
                    theoretical_usage = float(
                        row['theoretical_usage']) if row['theoretical_usage'] else 0
                    combo_usage = float(
                        row['combo_usage']) if row['combo_usage'] else 0
                    system_record = float(
                        row['system_record']) if row['system_record'] else 0
                    inventory_count = float(
                        row['inventory_count']) if row['inventory_count'] else 0
                    material_price = float(
                        row['material_price']) if row['material_price'] else 10.0
                    sale_price = float(
                        row['sale_price']) if row['sale_price'] else 15.0

                    # Calculate variance amount: System Record - (Theoretical + Combo + Inventory)
                    variance_amount = system_record - \
                        (theoretical_usage + combo_usage + inventory_count)

                    # Calculate variance percentage
                    combined_theoretical = theoretical_usage + combo_usage
                    if combined_theoretical > 0:
                        variance_percent = abs(
                            variance_amount) / combined_theoretical * 100
                    elif variance_amount != 0:
                        variance_percent = 100.0
                    else:
                        variance_percent = 0

                    # Determine variance status
                    variance_status = "正常"
                    if variance_percent > 5:
                        variance_status = "超量" if variance_amount > 0 else "不足"

                    variance_data.append({
                        'material_id': row['material_id'],
                        'store_id': row['store_id'],
                        'store_name': row['store_name'] or f"Store_{row['store_id']}",
                        'material_name': row['material_name'] or f"Material_{row['material_id']}",
                        'material_number': row['material_number'] or '',
                        'material_unit': row['material_unit'] or '个',
                        'package_spec': row['package_spec'] or '',
                        'theoretical_usage': theoretical_usage,
                        'combo_usage': combo_usage,
                        'system_record': system_record,
                        'inventory_count': inventory_count,
                        'variance_amount': variance_amount,
                        'variance_percent': variance_percent,
                        'variance_status': variance_status,
                        'material_price': material_price,
                        'sale_price': sale_price
                    })

                return variance_data

        except Exception as e:
            print(f"ERROR: Failed to get beverage variance data: {e}")
            import traceback
            traceback.print_exc()
            return []
