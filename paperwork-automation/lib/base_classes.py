#!/usr/bin/env python3
"""
Base classes for common patterns in Haidilao paperwork automation.
Consolidates duplicate code from 13+ worksheet generators and 50+ extraction scripts.
"""

import logging
import sys
from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Union
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import our centralized utilities
from .excel_utils import suppress_excel_warnings, safe_read_excel
from .config import STORE_NAME_MAPPING, DEFAULT_DATE_FORMAT


class BaseWorksheetGenerator(ABC):
    """
    Base class for all worksheet generators.
    Consolidates common patterns from 13+ worksheet generator classes.
    """
    
    def __init__(self, store_names: List[str], target_date: str):
        """
        Initialize worksheet generator with common setup.
        
        Args:
            store_names: List of store names to process
            target_date: Target date in YYYY-MM-DD format
        """
        self.store_names = store_names
        self.target_date = target_date
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # Parse target date for convenience
        try:
            self.target_dt = datetime.strptime(target_date, DEFAULT_DATE_FORMAT)
        except ValueError as e:
            self.logger.error(f"Invalid target date format: {target_date}")
            raise ValueError(f"Target date must be in YYYY-MM-DD format: {e}")
        
        # Setup common styles
        self.setup_common_styles()
    
    def setup_common_styles(self):
        """Setup common openpyxl styles used across all worksheets."""
        # Header styles
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.header_font = Font(bold=True, size=12)
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Data styles
        self.data_font = Font(size=11)
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Border styles
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Number formatting
        self.percentage_format = '0.0%'
        self.currency_format = '#,##0.00'
        self.integer_format = '#,##0'
    
    def set_column_widths(self, ws, widths: List[int]):
        """
        Set column widths for worksheet.
        Standardizes column width setting across all generators.
        
        Args:
            ws: Worksheet object
            widths: List of column widths
        """
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def apply_header_style(self, cell):
        """Apply standard header styling to a cell."""
        cell.fill = self.header_fill
        cell.font = self.header_font
        cell.alignment = self.header_alignment
        cell.border = self.thin_border
    
    def apply_data_style(self, cell, align='left'):
        """Apply standard data styling to a cell."""
        cell.font = self.data_font
        cell.border = self.thin_border
        
        if align == 'center':
            cell.alignment = self.center_alignment
        elif align == 'right':
            cell.alignment = self.right_alignment
    
    def format_percentage(self, value: float, decimals: int = 1) -> str:
        """
        Standard percentage formatting.
        
        Args:
            value: Percentage value (as decimal, e.g., 0.15 for 15%)
            decimals: Number of decimal places
            
        Returns:
            Formatted percentage string
        """
        if pd.isna(value) or value is None:
            return "0.0%"
        return f"{value * 100:.{decimals}f}%"
    
    def calculate_percentage_change(self, current: float, previous: float) -> float:
        """
        Standard percentage change calculation.
        Used across multiple worksheet generators.
        
        Args:
            current: Current period value
            previous: Previous period value
            
        Returns:
            Percentage change as decimal (e.g., 0.15 for 15% increase)
        """
        if previous == 0 or previous is None or pd.isna(previous):
            return 0.0
        if current is None or pd.isna(current):
            return -1.0
        
        return (current - previous) / previous
    
    def safe_divide(self, numerator: float, denominator: float, default: float = 0.0) -> float:
        """
        Safe division with default value for zero denominator.
        
        Args:
            numerator: Numerator value
            denominator: Denominator value
            default: Default value if denominator is zero
            
        Returns:
            Division result or default value
        """
        if denominator == 0 or denominator is None or pd.isna(denominator):
            return default
        if numerator is None or pd.isna(numerator):
            return default
            
        return numerator / denominator
    
    def add_title_section(self, ws, title: str, row: int, span_columns: int) -> int:
        """
        Add a title section with standard formatting.
        
        Args:
            ws: Worksheet object
            title: Title text
            row: Starting row
            span_columns: Number of columns to span
            
        Returns:
            Next available row
        """
        # Merge cells for title
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_columns)
        title_cell = ws.cell(row=row, column=1, value=title)
        
        # Apply title styling
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        return row + 1
    
    @abstractmethod
    def generate_worksheet(self, workbook, *args, **kwargs):
        """
        Abstract method for generating the specific worksheet type.
        Must be implemented by each concrete worksheet generator.
        """
        pass


class BaseExtractor(ABC):
    """
    Base class for all data extraction scripts.
    Consolidates common patterns from 50+ extraction scripts.
    """
    
    def __init__(self, database_manager=None):
        """
        Initialize extractor with common setup.
        
        Args:
            database_manager: Optional database manager instance
        """
        self.db_manager = database_manager
        self.logger = logging.getLogger(self.__class__.__name__)
        self.setup_project_path()
        suppress_excel_warnings()
    
    def setup_project_path(self):
        """
        Standard project path setup.
        Consolidates duplicate path manipulation from all scripts.
        """
        # Get project root (parent of parent of current file)
        project_root = Path(__file__).parent.parent
        if str(project_root) not in sys.path:
            sys.path.append(str(project_root))
        
        self.project_root = project_root
        self.logger.info(f"Project root set to: {project_root}")
    
    def setup_database_connection(self):
        """
        Standard database connection setup with error handling.
        """
        if not self.db_manager:
            try:
                from utils.database import get_database_manager
                self.db_manager = get_database_manager()
                self.logger.info("Database connection established")
            except ImportError:
                self.logger.warning("Database module not available - SQL file generation only")
                return False
        return True
    
    def safe_database_operation(self, operation_func, *args, **kwargs):
        """
        Execute database operation with proper error handling and rollback.
        
        Args:
            operation_func: Function to execute
            *args, **kwargs: Arguments for the function
            
        Returns:
            Operation result or None if failed
        """
        if not self.db_manager:
            self.logger.error("No database manager available")
            return None
        
        try:
            with self.db_manager.get_connection() as conn:
                return operation_func(conn, *args, **kwargs)
        except Exception as e:
            self.logger.error(f"Database operation failed: {e}")
            return None
    
    def batch_insert_with_conflict_handling(self, table_name: str, data_list: List[Dict], 
                                          conflict_columns: List[str], batch_size: int = 1000):
        """
        Standard batch insert with conflict resolution.
        Consolidates duplicate batch insert patterns.
        
        Args:
            table_name: Target table name
            data_list: List of dictionaries with data to insert
            conflict_columns: Columns to check for conflicts (for upsert)
            batch_size: Number of records per batch
            
        Returns:
            Number of records processed
        """
        if not data_list:
            self.logger.info(f"No data to insert into {table_name}")
            return 0
        
        total_processed = 0
        
        for i in range(0, len(data_list), batch_size):
            batch = data_list[i:i + batch_size]
            
            def insert_batch(conn, batch_data):
                # This would be implemented with specific SQL for each table
                # For now, this is a placeholder for the pattern
                self.logger.info(f"Processing batch of {len(batch_data)} records for {table_name}")
                return len(batch_data)
            
            processed = self.safe_database_operation(insert_batch, batch)
            if processed:
                total_processed += processed
        
        self.logger.info(f"Successfully processed {total_processed} records for {table_name}")
        return total_processed
    
    def get_store_id_mapping(self) -> Dict[str, int]:
        """
        Get standardized store name to ID mapping.
        
        Returns:
            Dictionary mapping store names to IDs
        """
        return STORE_NAME_MAPPING.copy()
    
    def validate_file_existence(self, file_path: Union[str, Path]) -> bool:
        """
        Validate that input file exists and is readable.
        
        Args:
            file_path: Path to file to validate
            
        Returns:
            True if file exists and is readable
        """
        file_path = Path(file_path)
        if not file_path.exists():
            self.logger.error(f"Input file does not exist: {file_path}")
            return False
        
        if not file_path.is_file():
            self.logger.error(f"Path is not a file: {file_path}")
            return False
        
        self.logger.info(f"Input file validated: {file_path}")
        return True
    
    def log_extraction_summary(self, extracted_counts: Dict[str, int]):
        """
        Log standardized extraction summary.
        
        Args:
            extracted_counts: Dictionary of entity type to count
        """
        self.logger.info("=" * 50)
        self.logger.info("EXTRACTION SUMMARY")
        self.logger.info("=" * 50)
        
        total = 0
        for entity_type, count in extracted_counts.items():
            self.logger.info(f"{entity_type:.<30} {count:>5}")
            total += count
        
        self.logger.info("-" * 50)
        self.logger.info(f"{'TOTAL RECORDS':.<30} {total:>5}")
        self.logger.info("=" * 50)
    
    @abstractmethod
    def extract_data(self, input_file: Union[str, Path], **kwargs) -> Dict[str, Any]:
        """
        Abstract method for data extraction.
        Must be implemented by each concrete extractor.
        
        Args:
            input_file: Path to input file
            **kwargs: Additional extraction parameters
            
        Returns:
            Dictionary with extraction results
        """
        pass


class BaseReportGenerator:
    """
    Base class for report generators that combine multiple worksheets.
    """
    
    def __init__(self, target_date: str):
        """
        Initialize report generator.
        
        Args:
            target_date: Target date in YYYY-MM-DD format
        """
        self.target_date = target_date
        self.logger = logging.getLogger(self.__class__.__name__)
        
        # Parse target date
        try:
            self.target_dt = datetime.strptime(target_date, DEFAULT_DATE_FORMAT)
        except ValueError as e:
            raise ValueError(f"Invalid target date format: {e}")
    
    def generate_output_filename(self, base_name: str, extension: str = '.xlsx') -> str:
        """
        Generate standardized output filename with timestamp.
        
        Args:
            base_name: Base name for the file
            extension: File extension
            
        Returns:
            Formatted filename with date
        """
        date_str = self.target_dt.strftime('%Y_%m_%d')
        return f"{base_name}_{date_str}{extension}"
    
    def setup_output_directory(self, output_dir: Union[str, Path] = "output") -> Path:
        """
        Setup and create output directory if it doesn't exist.
        
        Args:
            output_dir: Output directory path
            
        Returns:
            Path object for output directory
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        self.logger.info(f"Output directory: {output_path.absolute()}")
        return output_path


# Export base classes
__all__ = [
    'BaseWorksheetGenerator',
    'BaseExtractor', 
    'BaseReportGenerator'
]