"""Hi-Bowl Daily Report Processor

Processes Hi-Bowl daily transaction data and generates formatted reports
for the overseas business reporting template.
"""

import os
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path
import shutil
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import Dict, List, Tuple, Optional
import logging
import re

from lib.excel_utils import safe_read_excel, suppress_excel_warnings
from lib.config import EXCEL_FORMATS

# Set up logging
logger = logging.getLogger(__name__)


class HiBowlDailyProcessor:
    """Process Hi-Bowl daily reports and generate output"""

    def __init__(self):
        """Initialize the processor"""
        suppress_excel_warnings()
        # Use a more visible blue color with full opacity
        self.blue_fill = PatternFill(
            start_color="FF5B9BD5", end_color="FF5B9BD5", fill_type="solid")

    def process_daily_file(self, input_file: str, output_file: str, target_month: str = None) -> bool:
        """
        Process a Hi-Bowl daily data file and generate the output report.

        Args:
            input_file: Path to the input Excel file with daily data
            output_file: Path where the output file should be saved
            target_month: Target month in YYYYMM format (optional, will extract from data if not provided)

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            logger.info(f"Processing Hi-Bowl daily file: {input_file}")

            # Copy template to output location
            template_file = self._get_template_file()
            if not os.path.exists(template_file):
                logger.error(f"Template file not found: {template_file}")
                return False

            # Create output directory if needed
            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)

            # Copy template
            shutil.copy2(template_file, output_file)
            logger.info(f"Copied template to: {output_file}")

            # Process daily data
            result = self._read_daily_data(input_file)
            if result is None:
                return False

            daily_data, sheet_count = result

            # Calculate summary statistics
            summary = self._calculate_summary(
                daily_data, target_month, sheet_count)

            # Write to output file
            success = self._write_to_template(output_file, summary)

            if success:
                logger.info(
                    f"Successfully generated Hi-Bowl report: {output_file}")

            return success

        except Exception as e:
            logger.error(f"Error processing Hi-Bowl daily file: {str(e)}")
            return False

    def _get_template_file(self) -> str:
        """Get the template file path"""
        # Get project root
        current_file = os.path.abspath(__file__)
        project_root = os.path.dirname(os.path.dirname(current_file))

        template_path = os.path.join(
            project_root, 'Input', 'daily_report', 'hi-bowl-report',
            'output-template', '海外新业态管报数据-本位币-新模板.xlsx'
        )

        return template_path

    def _read_daily_data(self, input_file: str) -> Optional[Tuple[pd.DataFrame, int]]:
        """Read and consolidate daily data from all sheets

        Returns:
            Tuple of (DataFrame, sheet_count) or None if error
        """
        try:
            # Read all sheets
            xl_file = pd.ExcelFile(input_file)
            all_data = []
            valid_sheet_count = 0

            for sheet_name in xl_file.sheet_names:
                logger.info(f"Reading sheet: {sheet_name}")

                # Read raw data to find header
                df_raw = safe_read_excel(
                    input_file, sheet_name=sheet_name, header=None)

                # Find header row (contains '开台时间')
                header_row = None
                for idx in range(min(5, len(df_raw))):
                    row = df_raw.iloc[idx]
                    # Look for the actual headers like '订单编号' etc
                    if any('订单编号' in str(val) or '订单税前实收金额' in str(val) for val in row if pd.notna(val)):
                        header_row = idx
                        break

                if header_row is None:
                    logger.warning(
                        f"Could not find header in sheet {sheet_name}")
                    continue

                # Read with proper header
                df = safe_read_excel(
                    input_file, sheet_name=sheet_name, skiprows=header_row)

                # Handle duplicate column names by making them unique
                cols = df.columns.tolist()
                new_cols = []
                col_counts = {}
                for col in cols:
                    if col in col_counts:
                        col_counts[col] += 1
                        new_cols.append(f"{col}_{col_counts[col]}")
                    else:
                        col_counts[col] = 0
                        new_cols.append(col)
                df.columns = new_cols

                # Extract date from sheet name
                try:
                    date = pd.to_datetime(sheet_name)
                    df['date'] = date
                except:
                    logger.warning(
                        f"Could not parse date from sheet name: {sheet_name}")
                    continue

                # Don't clean column names yet - we'll do it after combining
                all_data.append(df)
                valid_sheet_count += 1

            if not all_data:
                logger.error("No valid data found in any sheet")
                return None

            # Combine all data
            combined_df = pd.concat(all_data, ignore_index=True)

            # Parse monetary values
            combined_df = self._parse_monetary_columns(combined_df)

            logger.info(
                f"Loaded {len(combined_df)} total transactions from {valid_sheet_count} sheets")
            return combined_df, valid_sheet_count

        except Exception as e:
            logger.error(f"Error reading daily data: {str(e)}")
            return None

    def _clean_column_name(self, col_name: str) -> str:
        """Clean column names for easier access"""
        mappings = {
            '开台时间': 'open_time',
            '完结时间': 'close_time',
            '订单编号': 'order_id',
            '桌号': 'table_no',
            '就餐人数': 'guest_count',
            '订单税前实收金额Totals': 'revenue_before_tax',
            '订单税金Tax': 'tax',
            'GST': 'gst',
            'QST': 'qst',
            '折扣金额（合计）': 'discount_amount',
            '赠菜金额（合计）': 'comp_amount',
            '免单金额（合计）': 'void_amount',
            '订单类型': 'order_type',
            '订单税前应收金额（不含税）': 'gross_revenue',
            '小费': 'tips',
            '舍入': 'rounding',
            '桌数': 'table_count',
            '退款订单': 'refund_order',
            '原始单号': 'original_order',
            '订单折扣率': 'discount_rate'
        }

        # Clean unnamed columns
        if 'Unnamed' in str(col_name):
            return str(col_name)

        # Apply mapping if exists
        for chinese, english in mappings.items():
            if chinese in str(col_name):
                return english

        return str(col_name)

    def _parse_monetary_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """Parse monetary columns from string format (e.g., '$123.45') to float"""
        # Map original column names to our clean names
        column_mappings = {
            '开台时间': '开台时间',  # Keep original for time parsing
            '完结时间': '完结时间',  # Keep original for time parsing
            '订单税前实收金额Totals': 'revenue_before_tax',
            '订单税金Tax': 'tax',
            'GST': 'gst',
            'QST': 'qst',
            '折扣金额（合计）': 'discount_amount',
            '赠菜金额（合计）': 'comp_amount',
            '免单金额（合计）': 'void_amount',
            '订单税前应收金额（不含税）': 'gross_revenue',
            '小费': 'tips',
            '舍入': 'rounding',
            '就餐人数': 'guest_count',
            '桌数': 'table_count',
            '退款订单': 'refund_order'
        }

        # Rename columns based on content
        # Keep track of renamed columns to avoid duplicate renaming
        renamed_cols = set()
        for idx, col in enumerate(df.columns):
            if col not in renamed_cols:
                for chinese, english in column_mappings.items():
                    if chinese in str(col):
                        # Check if the target name already exists
                        if english not in df.columns or english in renamed_cols:
                            df.rename(columns={col: english}, inplace=True)
                            renamed_cols.add(english)
                        break

        # Parse monetary columns
        monetary_columns = ['revenue_before_tax', 'tax', 'gst', 'qst',
                            'discount_amount', 'comp_amount', 'void_amount',
                            'gross_revenue', 'tips', 'rounding']

        for col in monetary_columns:
            if col in df.columns:
                # Remove $ sign and convert to float
                df[col] = df[col].apply(self._parse_money)

        # Ensure numeric columns
        numeric_columns = ['guest_count', 'table_count']
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        return df

    def _parse_money(self, value) -> float:
        """Parse money string to float"""
        if pd.isna(value):
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        # Remove $ and convert
        try:
            return float(str(value).replace('$', '').replace(',', ''))
        except:
            return 0.0

    def _calculate_summary(self, df: pd.DataFrame, target_month: str = None, sheet_count: int = 0) -> Dict:
        """Calculate summary statistics from daily data"""
        summary = {}

        # Store sheet count as scheduled operating days
        summary['scheduled_operating_days'] = sheet_count

        # Filter by target month if specified
        if target_month:
            target_year = int(target_month[:4])
            target_mon = int(target_month[4:6])
            df = df[df['date'].dt.year == target_year]
            df = df[df['date'].dt.month == target_mon]

        if len(df) == 0:
            logger.warning("No data for target month")
            return summary

        # Convert time strings to datetime for time segment analysis
        # Parse open time and close time
        df['open_datetime'] = pd.to_datetime(df['开台时间'], errors='coerce')
        df['close_datetime'] = pd.to_datetime(df['完结时间'], errors='coerce')

        # Basic statistics - removed revenue calculations as they're not needed

        # Calculate discount amounts (they come as negative values in the data)
        # Only use 折扣金额（合计） for the total discount amount
        discount_total = abs(df['discount_amount'].sum())
        summary['total_discount_with_tax'] = discount_total * 1.13
        # Assuming the discount amount includes 13% tax (GST + QST in Quebec)
        summary['total_discount_without_tax'] = discount_total 

        summary['total_tips'] = df['tips'].sum()

        # Operating days
        summary['actual_operating_days'] = df['date'].nunique()

        # Weekday/Weekend split (Monday-Friday as weekday, Saturday-Sunday as weekend/holiday)
        df['weekday'] = df['date'].dt.weekday
        weekday_mask = df['weekday'] < 5  # 0-4 are Monday-Friday

        summary['weekday_days'] = df[weekday_mask]['date'].nunique()
        summary['weekend_days'] = df[~weekday_mask]['date'].nunique()  # 节假日天数

        # Guest and order counts
        # Check if refund_order column exists
        if 'refund_order' in df.columns:
            # Create a mask for non-refunded orders (refund_order is null or empty)
            non_refund_mask = df['refund_order'].isna() | (df['refund_order'] == '')
            
            # Count refunded orders for logging
            refunded_count = len(df[~non_refund_mask])
            if refunded_count > 0:
                logger.info(f"Excluding {refunded_count} refunded orders from order counts")
            
            # Filter out refunded orders
            df_non_refund = df[non_refund_mask]
            df_non_refund_weekday = df[non_refund_mask & weekday_mask]
            df_non_refund_weekend = df[non_refund_mask & ~weekday_mask]
        else:
            # If no refund column, use all data
            logger.warning("No refund_order column found, including all orders in counts")
            df_non_refund = df
            df_non_refund_weekday = df[weekday_mask]
            df_non_refund_weekend = df[~weekday_mask]
            
        # Calculate guest counts (include all orders, even refunded ones for guest tracking)
        summary['total_guests'] = int(df['guest_count'].sum())
        summary['weekday_guests'] = int(df[weekday_mask]['guest_count'].sum())
        summary['weekend_guests'] = int(df[~weekday_mask]['guest_count'].sum())

        # Calculate order counts (exclude refunded orders)
        summary['total_orders'] = len(df_non_refund)
        summary['weekday_orders'] = len(df_non_refund_weekday)
        summary['weekend_orders'] = len(df_non_refund_weekend)

        # Time segment analysis (for Singapore/Malaysia)
        # Define time segments
        time_segments = {
            '08:00-13:59': (8, 14),
            '14:00-16:59': (14, 17),
            '17:00-21:59': (17, 22),
            '22:00-07:59': (22, 8)  # Overnight segment
        }

        # Get hour from close time (do this ONCE outside the loop)
        df['close_hour'] = df['close_datetime'].dt.hour

        # Log revenue calculation method
        logger.info("Calculating revenue by deducting 订单税金Tax from 订单税前实收金额Totals")
        
        # Log overall totals for verification (including refunded orders as they have negative amounts)
        total_revenue_with_tax = df['revenue_before_tax'].sum()
        total_tax = df['tax'].sum() if 'tax' in df.columns else 0
        total_revenue_without_tax = total_revenue_with_tax - total_tax
        logger.info(f"Total revenue with tax (including refunds): ${total_revenue_with_tax:.2f}, Total tax: ${total_tax:.2f}, Total revenue without tax: ${total_revenue_without_tax:.2f}")

        # Initialize segment totals for verification
        segment_revenue_total = 0
        segment_tax_total = 0
        
        # Calculate revenue and orders by time segment
        for segment_name, (start_hour, end_hour) in time_segments.items():

            if start_hour < end_hour:
                # Normal case (same day)
                segment_mask = (df['close_hour'] >= start_hour) & (
                    df['close_hour'] < end_hour)
            else:
                # Overnight case (22:00-07:59)
                segment_mask = (df['close_hour'] >= start_hour) | (
                    df['close_hour'] < 8)

            # Calculate revenue for this segment (before tax only)
            # Include refunded orders as they already have negative amounts
            revenue_mask = segment_mask
                
            # Always calculate by deducting actual tax from revenue_before_tax
            segment_revenue_with_tax = df[revenue_mask]['revenue_before_tax'].sum()
            segment_tax = df[revenue_mask]['tax'].sum() if 'tax' in df.columns else 0
            segment_revenue = segment_revenue_with_tax - segment_tax
            
            # Add to totals for verification
            segment_revenue_total += segment_revenue
            segment_tax_total += segment_tax
            
            # Log for debugging - changed to INFO level for visibility
            if segment_revenue_with_tax > 0:
                logger.info(f"Segment {segment_name}: Revenue with tax: ${segment_revenue_with_tax:.2f}, Tax: ${segment_tax:.2f}, Revenue without tax: ${segment_revenue:.2f}")
            
            # Calculate orders excluding refunded ones
            if 'refund_order' in df.columns:
                non_refund_segment_mask = segment_mask & (df['refund_order'].isna() | (df['refund_order'] == ''))
                segment_orders = len(df[non_refund_segment_mask])
            else:
                segment_orders = len(df[segment_mask])

            # Store in summary
            summary[f'revenue_{segment_name}'] = segment_revenue
            summary[f'orders_{segment_name}'] = segment_orders
            
        # Log segment totals vs overall totals
        logger.info(f"Sum of all segments - Revenue without tax: ${segment_revenue_total:.2f}")
        logger.info(f"Expected total (from overall calculation): ${total_revenue_without_tax:.2f}")
        if abs(segment_revenue_total - total_revenue_without_tax) > 0.01:
            logger.warning(f"Discrepancy detected: Segment sum differs from total by ${abs(segment_revenue_total - total_revenue_without_tax):.2f}")

        # Period (YYMM format)
        if target_month:
            summary['period'] = target_month[2:6]  # e.g., '2507' for July 2025
        else:
            # Get from first date
            first_date = df['date'].min()
            summary['period'] = first_date.strftime('%y%m')

        logger.info(f"Calculated summary for period {summary['period']}")
        logger.info(
            f"Total orders (excluding refunds): {summary['total_orders']}, Total guests: {summary['total_guests']}")
        logger.info(
            f"Weekday days: {summary['weekday_days']}, Weekend days: {summary['weekend_days']}")

        return summary

    def _write_to_template(self, output_file: str, summary: Dict) -> bool:
        """Write summary data to the template file"""
        try:
            # Open workbook
            wb = openpyxl.load_workbook(output_file)
            ws = wb['海外新业态管报收入数据-本位币']

            # Find the data row for '加拿大hi1' (row 8 based on analysis)
            data_row = 8

            # ONLY fill the requested fields with CORRECT column mappings

            # Column K (11): 优惠总金额（含税）
            ws.cell(row=data_row, column=11).value = summary.get(
                'total_discount_with_tax', 0)

            # Column L (12): 优惠总金额（不含税）
            ws.cell(row=data_row, column=12).value = summary.get(
                'total_discount_without_tax', 0)

            # Column M (13): 小费-仅美国面馆/美国快餐填
            ws.cell(row=data_row, column=13).value = summary.get(
                'total_tips', 0)

            # Column N (14): 实际营业天数
            ws.cell(row=data_row, column=14).value = summary.get(
                'actual_operating_days', 0)

            # Column O (15): 应营业天数
            ws.cell(row=data_row, column=15).value = summary.get(
                'scheduled_operating_days', 0)

            # Column P (16): 工作日天数
            ws.cell(row=data_row, column=16).value = summary.get(
                'weekday_days', 0)

            # Column Q (17): 节假日天数
            ws.cell(row=data_row, column=17).value = summary.get(
                'weekend_days', 0)

            # Column W (23): 全月就餐人数
            ws.cell(row=data_row, column=23).value = summary.get(
                'total_guests', 0)

            # Column X (24): 工作日就餐人数
            ws.cell(row=data_row, column=24).value = summary.get(
                'weekday_guests', 0)

            # Column Y (25): 节假日就餐人数
            ws.cell(row=data_row, column=25).value = summary.get(
                'weekend_guests', 0)

            # Column AA (27): 全月订单数量-美国面馆/马来清真火锅/美国快餐
            ws.cell(row=data_row, column=27).value = summary.get(
                'total_orders', 0)

            # Column AB (28): 工作日订单数-美国面馆/马来清真火锅/美国快餐
            ws.cell(row=data_row, column=28).value = summary.get(
                'weekday_orders', 0)

            # Column AC (29): 节假日订单数-美国面馆/马来清真火锅/美国快餐
            ws.cell(row=data_row, column=29).value = summary.get(
                'weekend_orders', 0)

            # Time segment revenue (columns AS-AV, 45-48)
            # Note: All revenue values are tax-excluded (sales revenue without tax)
            # Column AS (45): 08:00-13:59销售收入-新加坡面馆&马来清真火锅
            ws.cell(row=data_row, column=45).value = summary.get(
                'revenue_08:00-13:59', 0)

            # Column AT (46): 14:00-16:59销售收入-新加坡面馆&马来清真火锅
            ws.cell(row=data_row, column=46).value = summary.get(
                'revenue_14:00-16:59', 0)

            # Column AU (47): 17:00-21:59销售收入-新加坡面馆&马来清真火锅
            ws.cell(row=data_row, column=47).value = summary.get(
                'revenue_17:00-21:59', 0)

            # Column AV (48): 22:00-(次)07:59销售收入-新加坡面馆&马来清真火锅
            ws.cell(row=data_row, column=48).value = summary.get(
                'revenue_22:00-07:59', 0)

            # Time segment orders (columns BC-BF, 55-58)
            # Column BC (55): 08:00-13:59订单数-美国面馆&马来清真火锅
            ws.cell(row=data_row, column=55).value = summary.get(
                'orders_08:00-13:59', 0)

            # Column BD (56): 14:00-16:59订单数-美国面馆&马来清真火锅
            ws.cell(row=data_row, column=56).value = summary.get(
                'orders_14:00-16:59', 0)

            # Column BE (57): 17:00-21:59订单数-美国面馆&马来清真火锅
            ws.cell(row=data_row, column=57).value = summary.get(
                'orders_17:00-21:59', 0)

            # Column BF (58): 22:00-(次)07:59订单数-美国面馆&马来清真火锅
            ws.cell(row=data_row, column=58).value = summary.get(
                'orders_22:00-07:59', 0)

            # Apply blue fill to all cells that were filled
            cells_to_fill = [11, 12, 13, 14, 15, 16, 17, 23, 24,
                             25, 27, 28, 29, 45, 46, 47, 48, 55, 56, 57, 58]
            for col in cells_to_fill:
                cell = ws.cell(row=data_row, column=col)
                cell.fill = self.blue_fill
                # Also format numbers nicely
                if col in [11, 12, 13, 45, 46, 47, 48]:  # Revenue/money columns
                    cell.number_format = '#,##0.00'
                elif col in [14, 15, 16, 17, 23, 24, 25, 27, 28, 29, 55, 56, 57, 58]:  # Count columns
                    cell.number_format = '#,##0'

            # Save workbook
            wb.save(output_file)
            wb.close()

            logger.info("Successfully wrote data to template")
            return True

        except Exception as e:
            logger.error(f"Error writing to template: {str(e)}")
            return False
